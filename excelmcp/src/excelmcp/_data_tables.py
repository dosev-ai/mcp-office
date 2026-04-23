"""Tables and named ranges: list, read, write.

Internal sub-module. Import via excelmcp._data (public router).
"""
from __future__ import annotations

from openpyxl.workbook.defined_name import DefinedName as _DefinedName

from ._core import (
    _NAMED_RANGE_RE,
    _NAMED_RANGE_RESERVED,
    _RANGE_ADDRESS_RE,
    ValidationError,
    _audit_log,
    _check_confirm,
    _check_path,
    _check_range_size,
    _check_write,
    _flush_if_dirty,
    _load_wb,
    _parse_a1_range,
    _parse_named_range_ref,
    _read_range_cells,
    _save_and_evict,
)

__all__ = [
    "list_tables",
    "read_table",
    "get_named_ranges",
    "read_named_range",
    "write_named_range",
]


def list_tables(path: str, sheet: str | None = None) -> list[dict]:
    """List all Excel tables in *sheet*, or across all sheets when *sheet* is omitted.

    Each entry: ``{"sheet", "name", "display_name", "ref", "header_row",
    "totals_row"}``.  The ``"sheet"`` key is always present for workbook-wide
    discovery.
    """
    resolved = _check_path(path)
    wb = _load_wb(resolved)
    sheet_names: list[str]
    if sheet is not None:
        if sheet not in wb.sheetnames:
            raise ValidationError(f"Sheet not found: {sheet!r}")
        sheet_names = [sheet]
    else:
        sheet_names = list(wb.sheetnames)
    results: list[dict] = []
    for sn in sheet_names:
        ws = wb[sn]
        try:
            table_iter = ws.tables.values()
        except AttributeError:
            table_iter = iter(ws.tables) if ws.tables else iter([])
        for t in table_iter:
            header_row = True
            try:
                if t.headerRowCount is not None:
                    header_row = t.headerRowCount != 0
            except AttributeError:
                pass
            totals_row = False
            try:
                totals_row = bool(t.totalsRowCount)
            except AttributeError:
                pass
            results.append(
                {
                    "sheet": sn,
                    "name": t.name,
                    "display_name": t.displayName,
                    "ref": t.ref,
                    "header_row": header_row,
                    "totals_row": totals_row,
                }
            )
    return results


def read_table(path: str, sheet: str, table_name: str) -> dict:
    """Read all rows from table *table_name* in *sheet*.

    Returns ``{"table_name", "ref", "headers": [...], "rows": [[...], ...]}``.
    """
    resolved = _check_path(path)
    wb = _load_wb(resolved)
    if sheet not in wb.sheetnames:
        raise ValidationError(f"Sheet not found: {sheet!r}")
    ws = wb[sheet]

    target_table = None
    try:
        target_table = ws.tables[table_name]
    except (KeyError, TypeError):
        pass
    if target_table is None:
        try:
            for t in ws.tables.values():
                if t.name == table_name:
                    target_table = t
                    break
        except AttributeError:
            pass
    if target_table is None:
        raise ValidationError(f"Table not found: {table_name!r}")

    min_col, min_row, max_col, max_row = _parse_a1_range(target_table.ref)
    all_rows = list(
        ws.iter_rows(
            min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col
        )
    )
    if not all_rows:
        return {
            "table_name": table_name,
            "ref": target_table.ref,
            "headers": [],
            "rows": [],
        }
    headers = [str(c.value) for c in all_rows[0]]
    rows = [[c.value for c in row] for row in all_rows[1:]]
    return {
        "table_name": table_name,
        "ref": target_table.ref,
        "headers": headers,
        "rows": rows,
    }


def get_named_ranges(path: str) -> list[dict]:
    """Return all defined (named) ranges in the workbook.

    Each entry: ``{"name": str, "refers_to": str, "scope": None}``.
    Returns ``[]`` if none found.
    """
    resolved = _check_path(path)
    wb = _load_wb(resolved)
    results: list[dict] = []
    try:
        items = list(wb.defined_names.items())
    except AttributeError:
        try:
            items = [(defn.name, defn) for defn in wb.defined_names.definedName]
        except AttributeError:
            return []
    for name, defn in items:
        try:
            refers_to = (
                getattr(defn, "attr_text", None) or getattr(defn, "value", None) or ""
            )
            if not refers_to:
                continue
            results.append({"name": name, "refers_to": refers_to, "scope": None})
        except Exception:
            continue
    return results


def read_named_range(path: str, name: str) -> dict:
    """Read the cells referenced by a named range.

    Returns ``{"name": name, "refers_to": str, "data": list[list[dict]]}``.
    """
    resolved = _check_path(path)
    wb = _load_wb(resolved)
    defn = None
    try:
        defn = wb.defined_names[name]
    except KeyError:
        pass
    if defn is None:
        raise ValidationError(f"Named range not found: {name!r}")
    refers_to = getattr(defn, "attr_text", None) or getattr(defn, "value", None) or ""
    if not refers_to:
        raise ValidationError(f"Named range {name!r} has no reference")
    sheet_name, range_address = _parse_named_range_ref(refers_to)
    if sheet_name not in wb.sheetnames:
        raise ValidationError(f"Sheet not found: {sheet_name!r}")
    ws = wb[sheet_name]
    min_col, min_row, max_col, max_row = _parse_a1_range(range_address)
    n_cells = (max_row - min_row + 1) * (max_col - min_col + 1)
    _check_range_size(n_cells)
    data_list = _read_range_cells(ws, min_row, min_col, max_row, max_col)
    return {"name": name, "refers_to": refers_to, "data": data_list}


def write_named_range(
    path: str,
    name: str,
    sheet: str,
    range_address: str,
    confirm: bool = False,
) -> dict:
    """Create or update a named range definition in the workbook.

    Returns ``{"ok": True, "name": name, "refers_to": full_ref, "sheet": sheet}``.
    Requires EXCEL_ENABLE_WRITE=true and confirm=True.
    """
    _check_write()
    _check_confirm(confirm)
    resolved = _check_path(path)
    if not _NAMED_RANGE_RE.match(name):
        raise ValidationError(
            f"Invalid named range name {name!r}: must start with a letter or underscore "
            "and contain only letters, digits, underscores, dots, question marks, or backslashes."
        )
    if _NAMED_RANGE_RESERVED.match(name):
        raise ValidationError(
            f"Named range name {name!r} uses a reserved prefix (__xlnm.* or _xlfn.*) "
            "and cannot be written."
        )
    # Validate range_address format
    if not _RANGE_ADDRESS_RE.match(range_address):
        raise ValidationError(
            f"Invalid range address {range_address!r}: expected a cell reference (A1), "
            "a range (A1:D4), or a sheet-qualified range ('Sheet1'!A1:D4)."
        )
    _flush_if_dirty(str(resolved))
    wb = _load_wb(resolved, for_write=True)
    if sheet not in wb.sheetnames:
        raise ValidationError(f"Sheet not found: {sheet!r}")
    if "!" not in range_address:
        safe_sheet = sheet.replace("'", "''")
        full_ref = f"'{safe_sheet}'!{range_address}"
    else:
        full_ref = range_address
    wb.defined_names[name] = _DefinedName(name, attr_text=full_ref)
    _save_and_evict(wb, resolved)
    _audit_log("write_named_range", resolved, sheet, range_address, None)
    return {"ok": True, "name": name, "refers_to": full_ref, "sheet": sheet}
