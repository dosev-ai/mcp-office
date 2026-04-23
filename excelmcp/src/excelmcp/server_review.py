"""MCP wiring for Excel render review and evidence primitives.

Tools registered on the shared mcp instance at import time (side-effect import):
    review_workbook_render         — Openpyxl structural quality gate (T-RR1)
    produce_export_evidence_bundle — Aggregate PDF export + review findings (T-RR2)
    export_changed_sheets_only     — COM PDF export of changed sheets (W4)

Also registers:
    Resource: excelmcp://prompts/excel_render_check_iterate_v1
    Prompt:   excel_render_check_iterate_v1

Import side-effect: importing this module registers all tools/prompts/resources
on the shared mcp instance from excelmcp._server_instance.
Do not instantiate FastMCP here.

The module is mostly openpyxl (read-only), but export_changed_sheets_only
performs a deferred COM import for PDF export (write side-effect).
"""
from __future__ import annotations

import hashlib
import logging
import os
from datetime import datetime, timezone
from pathlib import Path

import openpyxl
from fastmcp.exceptions import ToolError

from excelmcp._core import ValidationError, _check_confirm, _check_path, _check_write
from excelmcp._server_instance import mcp
from excelmcp.review_excel import review_workbook_render as _review_workbook_render

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Canonical Render-Check-Iterate workflow prompt (Excel adaptation of PPT v1)
# ---------------------------------------------------------------------------

_RENDER_CHECK_ITERATE_V1_PROMPT = """\
# Excel Render-Check-Iterate Workflow v1

## Required Environment Gates
Before starting, verify:
- EXCEL_ENABLE_WRITE=true   (recalculate_workbook step)
- EXCEL_ENABLE_COM=true     (recalculate_workbook + export_as_pdf steps)
- EXCEL_ALLOWLIST_ROOTS=<workbook directory>
- EXCEL_EXPORT_ROOTS=<PDF output directory>
- confirm=True on every write/export call

---

## STEP 1 — Recalculate
Flush stale formula caches BEFORE reviewing or exporting.

```
recalculate_workbook(
    path="<absolute workbook path>",
    full_calc=True,     # rebuilds entire dependency tree
    confirm=True        # required; modifies the workbook in-place
)
```

If `EXCEL_ENABLE_COM` is not available, document the gap and proceed — cached
values may be stale. Note this in the evidence bundle.

---

## STEP 2 — Review
Run the render quality gate on the recalculated workbook.

```
review_workbook_render(
    path="<absolute workbook path>",
    sheet_names=None,   # None = all non-chart sheets; or list specific sheets
    recalculated=True,  # suppresses FORMULA_STALE finding after Step 1
)
```

Inspect the returned `findings` list. Look for any WARN entries with
`severity="medium"` or `severity="high"` — these block the export.

---

## STEP 3 — Fix
If `passed=False`, address every medium/high severity finding before proceeding.

Common fixes by criterion:
- **PRINT_AREA_UNSET**:
  Call `set_print_area(path, sheet, address="A1:<last data cell>", confirm=True)`
- **SHEET_EMPTY**:
  Confirm the sheet should be exported; if empty by design, exclude it via `sheet_names`.
- **FORMULA_STALE**:
  Re-run Step 1 (recalculate_workbook). Only fires if recalculated=False.
- **PRINT_AREA_EXCEEDS_PAGE** (low — does not block):
  Optionally narrow the print area or use landscape orientation.

After fixes, loop back to Step 2 to re-review until `passed=True`.
Maximum recommended iterations: 3.

---

## STEP 4 — Export
Export the reviewed workbook to PDF.

### Full workbook export (all visible sheets)
```
export_as_pdf(
    scope="workbook",
    path="<absolute workbook path>",
    output_path="<absolute output path>/<name>.pdf",
    quality=0,
    confirm=True
)
```

### Per-sheet export
```
export_as_pdf(
    scope="sheet",
    path="<absolute workbook path>",
    output_path="<absolute output path>/<SheetName>.pdf",
    sheet="<sheet name>",
    quality=0,
    confirm=True
)
```

Each call returns `sha256`, `size_bytes`, `exported_at`, and `elapsed_ms` as evidence.

---

## STEP 5 — Bundle
Aggregate export results and review findings into a single evidence record.

```
produce_export_evidence_bundle(
    path="<absolute workbook path>",
    export_results=[<list of export_as_pdf return dicts>],
    review_results=[<list of review_workbook_render return dicts>],
    exports_dir="<optional directory containing PDF files>",
)
```

Returns `overall_passed`, `total_exports`, `total_review_findings`,
`review_summary`, and `generated_at`.

---

## STEP 6 — Iterate or Complete
- If `overall_passed=True`: the render loop is complete. Attach the bundle to
    the issue tracker record for this task as evidence.
- If `overall_passed=False`: return to Step 3, fix remaining findings, and
  repeat the loop (Steps 3 → 4 → 5).

Evidence bundle format for review record attachment:
```json
{
  "bundle_version": "1.0",
  "path": "<workbook path>",
  "overall_passed": true,
  "total_exports": 1,
  "total_review_findings": 0,
  "exports": [{"sha256": "...", "exported_at": "..."}],
  "review_summary": [],
  "generated_at": "2026-03-11T12:00:00Z"
}
```
"""


# ---------------------------------------------------------------------------
# Tool T-RR1: review_workbook_render
# ---------------------------------------------------------------------------

@mcp.tool()
def review_workbook_render(
    path: str,
    sheet_names: list[str] | None = None,
    recalculated: bool = False,
) -> dict:
    """Review an Excel workbook's render quality before/after PDF export.

    Runs structural checks (PRINT_AREA_UNSET, SHEET_EMPTY, FORMULA_STALE,
    PRINT_AREA_EXCEEDS_PAGE) on the specified sheets and returns findings.

    Read-only — does NOT modify the workbook.

    Args:
        path: Absolute path to the .xlsx workbook.
        sheet_names: Sheet names to review. None → all non-chart sheets.
        recalculated: Set True if recalculate_workbook() was called before export.
                      Suppresses FORMULA_STALE finding.

    Returns a dict with keys: path, sheets_reviewed, passed, findings,
    checks_run, findings_count.
    """
    try:
        _check_path(path)
        return _review_workbook_render(
            path=path,
            sheet_names=sheet_names,
            recalculated=recalculated,
        )
    except ToolError:
        raise
    except Exception as exc:
        logger.error("[excelmcp] review_workbook_render error: %s", exc, exc_info=True)
        raise ToolError(str(exc)) from exc


# ---------------------------------------------------------------------------
# Tool T-RR2: produce_export_evidence_bundle
# ---------------------------------------------------------------------------

@mcp.tool()
def produce_export_evidence_bundle(
    path: str,
    export_results: list[dict],
    review_results: list[dict],
    exports_dir: str | None = None,
) -> dict:
    """Produce a structured evidence bundle aggregating PDF export + review findings.

    Call after review_workbook_render and export_as_pdf to close the
    render-check-iterate loop.

    Args:
        path: Absolute path to the workbook.
        export_results: List of export result dicts from export_as_pdf calls.
                        Must be a list — raises ToolError if not.
                        Each entry must contain at least 'sha256' and 'exported_at'.
        review_results: List of review result dicts from review_workbook_render.
                        Must be a list — raises ToolError if not.
        exports_dir: Optional directory containing the exported PDF(s).

    Returns a bundle dict with keys: bundle_version, path, overall_passed,
    total_exports, total_review_findings, exports, review_summary, generated_at.
    """
    try:
        if not isinstance(export_results, list):
            raise ToolError("export_results must be a list")
        if not isinstance(review_results, list):
            raise ToolError("review_results must be a list")

        overall_passed = (
            len(export_results) > 0
            and all(r.get("passed", False) for r in review_results)
        )

        total_review_findings = sum(
            r.get("findings_count", 0) for r in review_results
        )

        # Flatten all findings from all review results
        review_summary: list[dict] = []
        for r in review_results:
            review_summary.extend(r.get("findings", []))

        generated_at = datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace(
            "+00:00", "Z"
        )

        bundle: dict = {
            "bundle_version": "1.0",
            "path": path,
            "overall_passed": overall_passed,
            "total_exports": len(export_results),
            "total_review_findings": total_review_findings,
            "exports": export_results,
            "review_summary": review_summary,
            "generated_at": generated_at,
        }
        if exports_dir is not None:
            bundle["exports_dir"] = exports_dir

        return bundle
    except ToolError:
        raise
    except Exception as exc:
        logger.error(
            "[excelmcp] produce_export_evidence_bundle error: %s", exc, exc_info=True
        )
        raise ToolError(str(exc)) from exc


# ---------------------------------------------------------------------------
# Resource + Prompt: excel_render_check_iterate_v1
# ---------------------------------------------------------------------------

# ---------------------------------------------------------------------------
# Private helper: _hash_sheet_content
# ---------------------------------------------------------------------------

def _hash_sheet_content(ws) -> str:
    """Produce a sha256 fingerprint of a worksheet's cell values."""
    h = hashlib.sha256()
    for row in ws.iter_rows(values_only=True):
        for cell in row:
            h.update(str(cell).encode())
    return h.hexdigest()


# ---------------------------------------------------------------------------
# Tool W4: export_changed_sheets_only
# ---------------------------------------------------------------------------

@mcp.tool()
def export_changed_sheets_only(
    path: str,
    base_dir: str,
    previous_hashes: dict[str, str],
    confirm: bool = False,
) -> dict:
    """
    Re-export only sheets whose content has changed since the last export run.

    Uses openpyxl cell-value hashing to detect mutations. Exports changed
    sheets via export_as_pdf. Returns hashes for ALL sheets so the caller
    can store them for the next iteration.

    Requires EXCEL_ENABLE_WRITE=true. If EXCEL_ENABLE_COM is not true, falls back
    to hash-only mode (no PDF export). confirm=True is always required.

    Args:
        path: Absolute path to the .xlsx workbook.
        base_dir: Directory where per-run export subdirectories are created.
        previous_hashes: Dict mapping sheet_name → sha256 hash string from
                         the previous run. Pass {} for the first run (all sheets
                         are treated as changed).
        confirm: Must be True to proceed (mutation guard).

    Returns:
        On success:
            {
              ok: True,
              run_dir: str,                 # Full path to the created run subdirectory
              exported_sheets: list[str],   # Sheet names that were re-exported
              unchanged_sheets: list[str],  # Sheet names whose hash matched
              new_hashes: dict[str, str],   # Updated hash map for all sheets
              export_details: list[dict],   # Per-exported-sheet export_as_pdf result
            }
        When EXCEL_ENABLE_COM is unavailable (graceful degradation):
            {
              ok: False,
              error: str,
              new_hashes: dict[str, str],
              exported_sheets: [],
              unchanged_sheets: list[str],
              run_dir: None,
              export_details: [],
            }
    """
    _check_write()
    _check_confirm(confirm)
    _check_path(path)

    try:
        # Load workbook for hashing (data_only=True — cached values; no COM needed)
        wb = openpyxl.load_workbook(path, data_only=True)

        # Compute hashes for all sheets
        new_hashes: dict[str, str] = {}
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            new_hashes[sheet_name] = _hash_sheet_content(ws)

        # Find changed vs unchanged sheets
        changed_sheets = [
            name for name in wb.sheetnames
            if new_hashes[name] != previous_hashes.get(name)
        ]
        unchanged_sheets = [
            name for name in wb.sheetnames
            if new_hashes[name] == previous_hashes.get(name)
        ]

        # Graceful degradation: COM not available
        com_available = os.environ.get("EXCEL_ENABLE_COM", "").lower() == "true"
        if not com_available:
            return {
                "ok": False,
                "error": "COM not available — set EXCEL_ENABLE_COM=true in environment",
                "new_hashes": new_hashes,
                "exported_sheets": [],
                "unchanged_sheets": list(wb.sheetnames),
                "run_dir": None,
                "export_details": [],
            }

        # COM available — validate base_dir (directory, not a file)
        if "\x00" in base_dir:
            raise ValidationError("Path contains illegal null byte")
        resolved_base = Path(base_dir).resolve()
        if str(resolved_base).startswith("\\\\") or str(resolved_base).startswith("//"):
            raise ToolError("UNC paths are not allowed")
        roots_env = os.getenv("EXCEL_EXPORT_ROOTS", "").strip()
        if not roots_env:
            raise ValidationError("No allowlist configured")
        roots = [r.strip() for r in roots_env.split(",") if r.strip()]
        if not any(
            resolved_base.is_relative_to(Path(r).resolve()) for r in roots
        ):
            raise ValidationError("Path not in allowlist")
        run_dir = os.path.join(
            base_dir,
            f"run_{datetime.now(timezone.utc).strftime('%Y%m%dT%H%M%SZ')}",
        )
        os.makedirs(run_dir, exist_ok=True)

        export_details: list[dict] = []
        try:
            from excelmcp.server_com import export_as_pdf  # COM import deferred
        except ImportError as exc:
            raise ToolError(
                "COM export requires excelmcp.server_com — use export_as_pdf directly"
            ) from exc

        for sheet_name in changed_sheets:
            out_pdf = os.path.join(run_dir, f"{sheet_name}.pdf")
            try:
                result = export_as_pdf(
                    scope="sheet",
                    path=path,
                    output_path=out_pdf,
                    sheet=sheet_name,
                    quality=0,
                    confirm=True,
                )
                export_details.append(result)
            except ToolError:
                raise
            except Exception as exc:
                raise ToolError(
                    f"COM export failed for sheet {sheet_name!r}: {exc}"
                ) from exc

        return {
            "ok": True,
            "run_dir": run_dir,
            "exported_sheets": changed_sheets,
            "unchanged_sheets": unchanged_sheets,
            "new_hashes": new_hashes,
            "export_details": export_details,
        }
    except ToolError:
        raise
    except Exception as exc:
        raise ToolError(str(exc)) from exc


@mcp.resource("excelmcp://prompts/excel_render_check_iterate_v1")
def excel_render_check_iterate_v1_resource() -> str:
    """Excel Render-Check-Iterate v1 workflow prompt resource."""
    return _RENDER_CHECK_ITERATE_V1_PROMPT


@mcp.prompt()
def excel_render_check_iterate_v1() -> str:
    """6-step canonical loop: recalculate → review → export → bundle → iterate."""
    return _RENDER_CHECK_ITERATE_V1_PROMPT
