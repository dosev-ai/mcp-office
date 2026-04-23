"""Workbook-level metadata storage on a hidden worksheet."""
from __future__ import annotations

import json
import os

from openpyxl.worksheet.worksheet import Worksheet

from ._core import (
    ValidationError,
    _audit_log,
    _check_confirm,
    _check_path,
    _check_write,
    _load_wb,
    _save_and_evict,
    _wb_cache,
)

_METADATA_SHEET_NAME = "_MCP_META"
_METADATA_MAGIC = "EXCELMCP_WORKBOOK_METADATA"
_METADATA_VERSION = "v1"
_METADATA_CELL_MAGIC = "A1"
_METADATA_CELL_VERSION = "A2"
_METADATA_CELL_PAYLOAD = "A3"
_METADATA_CELL_TEXT_LIMIT = 32767


def _metadata_limit_bytes() -> int:
    raw_limit = os.getenv("EXCEL_MAX_METADATA_BYTES", "32768")
    try:
        limit = int(raw_limit)
    except (TypeError, ValueError) as exc:
        raise ValidationError(
            "Invalid EXCEL_MAX_METADATA_BYTES: expected positive integer, "
            f"got {raw_limit!r}"
        ) from exc
    if limit <= 0:
        raise ValidationError(
            "Invalid EXCEL_MAX_METADATA_BYTES: expected positive integer, "
            f"got {raw_limit!r}"
        )
    return limit


def _check_payload_bytes_limit(payload_bytes: int) -> None:
    limit = _metadata_limit_bytes()
    if payload_bytes > limit:
        raise ValidationError(
            f"Workbook metadata payload exceeds limit {limit} bytes"
        )


def _check_payload_cell_limit(payload_text: str) -> None:
    if len(payload_text) > _METADATA_CELL_TEXT_LIMIT:
        raise ValidationError(
            "Workbook metadata payload exceeds Excel single-cell limit "
            f"{_METADATA_CELL_TEXT_LIMIT} chars"
        )


def _encode_payload(payload: dict) -> tuple[str, int]:
    if not isinstance(payload, dict):
        raise ValidationError("Workbook metadata payload must be a top-level JSON object")

    try:
        encoded_text = json.dumps(
            payload,
            separators=(",", ":"),
            ensure_ascii=False,
            allow_nan=False,
        )
    except ValueError as exc:
        raise ValidationError(f"Invalid workbook metadata payload: {exc}") from exc
    encoded_bytes = encoded_text.encode("utf-8")
    _check_payload_bytes_limit(len(encoded_bytes))
    _check_payload_cell_limit(encoded_text)
    return encoded_text, len(encoded_bytes)


def _parse_payload(raw_payload: object) -> tuple[dict, int]:
    if not isinstance(raw_payload, str) or raw_payload == "":
        raise ValidationError("Workbook metadata payload missing or invalid")
    payload_bytes = len(raw_payload.encode("utf-8"))
    _check_payload_bytes_limit(payload_bytes)
    try:
        metadata = json.loads(raw_payload)
    except json.JSONDecodeError as exc:
        raise ValidationError(f"Workbook metadata payload is not valid JSON: {exc}") from exc
    if not isinstance(metadata, dict):
        raise ValidationError("Workbook metadata payload must be a top-level JSON object")
    return metadata, payload_bytes


def _get_metadata_sheet(wb, *, require_markers: bool) -> Worksheet | None:
    if _METADATA_SHEET_NAME not in wb.sheetnames:
        return None

    ws = wb[_METADATA_SHEET_NAME]
    if (
        ws[_METADATA_CELL_MAGIC].value != _METADATA_MAGIC
        or ws[_METADATA_CELL_VERSION].value != _METADATA_VERSION
    ):
        raise ValidationError(
            "Metadata sheet collision: _MCP_META exists without exact ExcelMCP markers"
        )
    if require_markers:
        return ws
    return ws


def _validate_read_contract(result: dict) -> None:
    required = {"present", "sheet_name", "visibility", "version", "metadata", "bytes"}
    if set(result) != required:
        raise ValidationError("Workbook metadata read contract validation failed")


def _validate_write_contract(result: dict) -> None:
    required = {
        "ok",
        "path",
        "sheet_name",
        "created",
        "replaced",
        "visibility",
        "version",
        "bytes_written",
    }
    if set(result) != required:
        raise ValidationError("Workbook metadata write contract validation failed")


def read_workbook_metadata(path: str) -> dict:
    """Read workbook metadata from the hidden _MCP_META worksheet."""
    resolved = _check_path(path)
    wb = _load_wb(resolved)
    ws = _get_metadata_sheet(wb, require_markers=True)
    if ws is None:
        result = {
            "present": False,
            "sheet_name": _METADATA_SHEET_NAME,
            "visibility": None,
            "version": None,
            "metadata": None,
            "bytes": 0,
        }
        _validate_read_contract(result)
        return result

    metadata, payload_bytes = _parse_payload(ws[_METADATA_CELL_PAYLOAD].value)
    result = {
        "present": True,
        "sheet_name": ws.title,
        "visibility": ws.sheet_state,
        "version": ws[_METADATA_CELL_VERSION].value,
        "metadata": metadata,
        "bytes": payload_bytes,
    }
    _validate_read_contract(result)
    return result


def write_workbook_metadata(
    path: str,
    payload: dict,
    confirm: bool = False,
) -> dict:
    """Write workbook metadata to the hidden _MCP_META worksheet."""
    _check_write()
    _check_confirm(confirm)
    payload_text, payload_bytes = _encode_payload(payload)
    resolved = _check_path(path)
    wb = _load_wb(resolved, for_write=True)
    try:
        ws = _get_metadata_sheet(wb, require_markers=False)
        created = ws is None
        replaced = ws is not None
        if ws is None:
            ws = wb.create_sheet(_METADATA_SHEET_NAME)

        ws.sheet_state = "hidden"
        ws[_METADATA_CELL_MAGIC] = _METADATA_MAGIC
        ws[_METADATA_CELL_VERSION] = _METADATA_VERSION
        ws[_METADATA_CELL_PAYLOAD] = payload_text

        result = {
            "ok": True,
            "path": str(resolved),
            "sheet_name": ws.title,
            "created": created,
            "replaced": replaced,
            "visibility": ws.sheet_state,
            "version": _METADATA_VERSION,
            "bytes_written": payload_bytes,
        }
        _validate_write_contract(result)
        _save_and_evict(wb, resolved)
        _audit_log("write_workbook_metadata", resolved, ws.title, "A1:A3", 1)
        return result
    except Exception:
        _wb_cache.pop((str(resolved), True), None)
        raise


def resolve_workbook_metadata_resource(path: str) -> dict:
    """Resolve workbook metadata resource content."""
    return read_workbook_metadata(path)
