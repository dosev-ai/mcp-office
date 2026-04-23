"""
Shared constants, error classes, parsing helpers, sheet-utility functions,
and type helpers for the Excel MCP layer.

Extracted from _core.py to keep that module focused on path validation,
workbook I/O, and governance guards.

No intra-package imports — only stdlib + openpyxl.
No MCP / FastMCP imports.
"""
from __future__ import annotations

import re

import openpyxl
import openpyxl.utils
import openpyxl.utils.cell

# ---------------------------------------------------------------------------
# Error taxonomy
# ---------------------------------------------------------------------------

class ExcelMCPError(Exception):
    """Base class for all Excel MCP errors."""


class ValidationError(ExcelMCPError):
    """Input did not pass allowlist / type / format validation."""


class NotAllowedError(ExcelMCPError):
    """Operation blocked by safety gate (EXCEL_ENABLE_WRITE or confirm=False)."""


class OfficeCOMError(ExcelMCPError):
    """Wraps lower-level Office / file errors with a safe, user-visible message."""


# ---------------------------------------------------------------------------
# Module-level constants (shared by _format_cell, _data_tables, _advanced, …)
# ---------------------------------------------------------------------------

_HEX_COLOR_RE = re.compile(r"^[0-9A-Fa-f]{6}$")
_VALID_BORDER_STYLES = frozenset(
    {"thin", "medium", "thick", "dashed", "dotted", "double"}
)
_NAMED_RANGE_RE = re.compile(r'^[A-Za-z_][A-Za-z0-9_.?\\]*$')
_NAMED_RANGE_RESERVED = re.compile(r'^(__xlnm\.|_xlfn\.)', re.IGNORECASE)
# Valid range address: a cell ref (A1), a range (A1:D4), or a cross-sheet ref ('Sheet1'!A1:D4)
# Allow: letters, digits, $, :, !, ', spaces (for sheet names); - is NOT valid
_RANGE_ADDRESS_RE = re.compile(
    r"^(?:'[^']+'|[A-Za-z0-9_]+)?!?\$?[A-Z]{1,3}\$?\d+(?::\$?[A-Z]{1,3}\$?\d+)?$"
)


# ---------------------------------------------------------------------------
# Parsing helpers
# ---------------------------------------------------------------------------

def _parse_a1_range(a1_range: str) -> tuple[int, int, int, int]:
    """Return (min_col, min_row, max_col, max_row) from an A1:B2-style string."""
    try:
        min_col, min_row, max_col, max_row = openpyxl.utils.cell.range_boundaries(
            a1_range.upper()
        )
    except Exception as exc:
        raise ValidationError(f"Invalid range reference: {a1_range!r}") from exc
    return min_col, min_row, max_col, max_row


def _parse_named_range_ref(refers_to: str) -> tuple[str, str]:
    """Parse ``'Sheet1'!$A$1:$B$3`` → ``("Sheet1", "A1:B3")``."""
    if "!" not in refers_to:
        raise ValidationError(f"Cannot parse named range reference: {refers_to!r}")
    raw_sheet, raw_range = refers_to.rsplit("!", 1)
    sheet_name = raw_sheet.strip("'\"")
    range_address = raw_range.replace("$", "")
    return sheet_name, range_address


# ---------------------------------------------------------------------------
# Type / formula helpers
# ---------------------------------------------------------------------------

def _validate_formula_safe(formula: str) -> None:
    """Validate that a formula string is safe to write to a cell.

    Raises :class:`ValidationError` if:
    - the formula does not start with ``=`` (after stripping leading/trailing
      whitespace), or
    - the formula exceeds 8 192 characters (Excel's cell-content limit).

    No semantic parsing is performed — only structural guards.
    """
    stripped = formula.strip() if formula else ""
    if not stripped.startswith("="):
        raise ValidationError("Formula must start with '='")
    if len(formula) > 8192:
        raise ValidationError("Formula too long")


# ---------------------------------------------------------------------------
# Sheet-level utilities (used by _io, _sheet, _data — no MCP logic)
# ---------------------------------------------------------------------------

def _get_used_range(ws: openpyxl.worksheet.worksheet.Worksheet) -> dict:
    """Return used-range dimensions for an already-loaded Worksheet.

    Returns a dict with min_row, max_row, min_col, max_col, address.
    Returns empty=True when the sheet has no data.
    """
    min_row = ws.min_row
    max_row = ws.max_row
    min_col = ws.min_column
    max_col = ws.max_column
    if min_col is None or min_row is None:
        return {
            "min_row": None,
            "max_row": None,
            "min_col": None,
            "max_col": None,
            "address": None,
            "empty": True,
        }
    min_col_letter = openpyxl.utils.get_column_letter(min_col)
    max_col_letter = openpyxl.utils.get_column_letter(max_col)
    address = f"{min_col_letter}{min_row}:{max_col_letter}{max_row}"
    return {
        "min_row": min_row,
        "max_row": max_row,
        "min_col": min_col,
        "max_col": max_col,
        "address": address,
    }


def _resolve_table_last_row(ws: "openpyxl.worksheet.worksheet.Worksheet") -> int:
    """Return the last data row of the first valid Excel Table on *ws*.

    Falls back to ``(ws.max_row or 0)`` when no valid table exists,
    preserving backward-compatible behaviour for plain (non-table) sheets.

    Handles malformed workbooks where ``tbl.ref`` is ``None`` or an empty
    string by silently skipping that table (BLOCKER 1 guard).
    """
    for tbl in ws.tables.values():
        tbl_ref = tbl.ref or ""          # BLOCKER 1: guard None / ""
        if not tbl_ref:
            continue
        try:
            _min_col, _min_row, _max_col, max_row = (
                openpyxl.utils.cell.range_boundaries(tbl_ref.upper())
            )
        except Exception:               # guard: malformed ref string
            continue
        return max_row                  # first valid table wins
    return ws.max_row or 0             # fallback: no valid table found


def _sync_autofilter_to_table(
    ws: "openpyxl.worksheet.worksheet.Worksheet",
) -> None:
    """Align ``ws.auto_filter.ref`` to the first matching Excel Table ref.

    Matches are determined by top-left cell equality.  When the autofilter
    ref and a table's ref share the same top-left cell, the autofilter is
    updated to the table's current (possibly expanded/contracted) ref.

    If ``ws.auto_filter.ref`` is ``None`` (no autofilter set) this function
    is a no-op — it does NOT create an autofilter (BLOCKER 2-A guard).
    Tables with ``tbl.ref`` equal to ``None`` or ``""`` are silently
    skipped (BLOCKER 2-B guard).
    """
    af_ref = ws.auto_filter.ref
    if not af_ref:                      # BLOCKER 2-A: None / "" → no-op
        return
    try:
        af_min_col, af_min_row, _af_max_col, _af_max_row = (
            openpyxl.utils.cell.range_boundaries(af_ref.upper())
        )
    except Exception:                   # malformed autofilter ref → no-op
        return
    af_top_left = (af_min_col, af_min_row)

    for tbl in ws.tables.values():
        tbl_ref = tbl.ref or ""         # BLOCKER 2-B: guard None / ""
        if not tbl_ref:
            continue
        try:
            tbl_min_col, tbl_min_row, _tbl_max_col, _tbl_max_row = (
                openpyxl.utils.cell.range_boundaries(tbl_ref.upper())
            )
        except Exception:               # malformed tbl ref → skip table
            continue
        if (tbl_min_col, tbl_min_row) == af_top_left:
            ws.auto_filter.ref = tbl_ref
            break                       # first match wins; stop iterating


def _adjust_table_refs_after_delete(
    ws: "openpyxl.worksheet.worksheet.Worksheet",
    start_row: int,
    amount: int,
) -> None:
    """Update ``tbl.ref`` on all tables after ``ws.delete_rows()`` is called.

    openpyxl does NOT auto-update table refs on row deletion, so callers must
    invoke this helper immediately after ``ws.delete_rows(start_row, amount)``.

    Algorithm:
      - Table entirely above deletion (t_max_row < start_row): no change.
      - Table entirely below deletion (t_min_row > end_row): shift up by amount.
      - Deletion overlaps the end of the table (t_min_row < start_row and
        t_max_row >= start_row): contract the table's last row.
      - Other overlap configurations are skipped (tbl.ref left unchanged).

    Tables with ``tbl.ref`` equal to ``None`` or ``""`` are silently skipped.
    """
    end_row = start_row + amount - 1
    for tbl in ws.tables.values():
        tbl_ref = tbl.ref or ""
        if not tbl_ref:
            continue
        try:
            t_min_col, t_min_row, t_max_col, t_max_row = (
                openpyxl.utils.cell.range_boundaries(tbl_ref.upper())
            )
        except Exception:
            continue

        if t_max_row < start_row:
            # Table entirely above deletion -- no change
            continue

        if t_min_row > end_row:
            # Table entirely below deletion -- shift the whole table up
            new_min_row = t_min_row - amount
            new_max_row = t_max_row - amount
        elif t_min_row < start_row and t_max_row >= start_row:
            # Deletion overlaps end of table -- contract bottom
            rows_removed = min(t_max_row, end_row) - start_row + 1
            new_min_row = t_min_row
            new_max_row = t_max_row - rows_removed
        else:
            # Other configurations (e.g. deletion covers table start) -- skip
            continue

        min_col_letter = openpyxl.utils.get_column_letter(t_min_col)
        max_col_letter = openpyxl.utils.get_column_letter(t_max_col)
        tbl.ref = f"{min_col_letter}{new_min_row}:{max_col_letter}{new_max_row}"


def _adjust_table_refs_after_insert(
    ws: "openpyxl.worksheet.worksheet.Worksheet",
    row: int,
    count: int,
) -> None:
    """Update ``tbl.ref`` on all tables after ``ws.insert_rows()`` is called.

    openpyxl does NOT auto-update table refs on row insertion, so callers must
    invoke this helper immediately after ``ws.insert_rows(row, amount=count)``.

    Algorithm:
      - Table entirely above insertion (t_max_row < row): no change.
      - Insertion at or before table start (row <= t_min_row): shift table down.
      - Insertion inside table (t_min_row < row <= t_max_row): expand table.

    Tables with ``tbl.ref`` equal to ``None`` or ``""`` are silently skipped.
    """
    for tbl in ws.tables.values():
        tbl_ref = tbl.ref or ""
        if not tbl_ref:
            continue
        try:
            t_min_col, t_min_row, t_max_col, t_max_row = (
                openpyxl.utils.cell.range_boundaries(tbl_ref.upper())
            )
        except Exception:
            continue

        if t_max_row < row:
            # Table entirely above insertion -- no change
            continue

        if t_min_row >= row:
            # Insertion at or before table start -- shift entire table down
            new_min_row = t_min_row + count
            new_max_row = t_max_row + count
        else:
            # t_min_row < row <= t_max_row: insertion inside table -- expand
            new_min_row = t_min_row
            new_max_row = t_max_row + count

        min_col_letter = openpyxl.utils.get_column_letter(t_min_col)
        max_col_letter = openpyxl.utils.get_column_letter(t_max_col)
        tbl.ref = f"{min_col_letter}{new_min_row}:{max_col_letter}{new_max_row}"


# ---------------------------------------------------------------------------
# Range cell reader
# ---------------------------------------------------------------------------

def _read_range_cells(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    min_row: int,
    min_col: int,
    max_row: int,
    max_col: int,
) -> list:
    """Return a 2-D list of cell dicts for a bounded region of *ws*.

    Each cell dict: {address, value, data_type}.
    Caller is responsible for range-size validation before calling this.
    """
    result: list = []
    for row in ws.iter_rows(
        min_row=min_row,
        max_row=max_row,
        min_col=min_col,
        max_col=max_col,
    ):
        result.append(
            [
                {
                    "address": cell.coordinate,
                    "value": cell.value,
                    "data_type": cell.data_type,
                }
                for cell in row
            ]
        )
    return result
