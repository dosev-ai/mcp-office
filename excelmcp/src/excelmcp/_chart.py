"""
Chart operations: create, list, and delete charts in worksheets.
"""
from __future__ import annotations

import re as _re

from ._core import (
    ExcelMCPError,
    ValidationError,
    _audit_log,
    _check_confirm,
    _check_path,
    _check_write,
    _flush_if_dirty,
    _load_wb,
    _save_and_evict,
    openpyxl,
)


def create_chart(
    path: str,
    sheet: str,
    chart_type: str,
    data_range: str,
    title: str = "",
    target_cell: str = "D1",
    anchor: str | None = None,
    width_px: int | None = None,
    height_px: int | None = None,
    confirm: bool = False,
) -> dict:
    """Create a chart (bar/line/scatter/pie) in the sheet.

    anchor overrides target_cell when provided (backward-compatible alias).
    width_px / height_px set chart size in pixels (1–5000); default is openpyxl default (~14x7.5 cm).
    Requires EXCEL_ENABLE_WRITE=true and confirm=True.
    """
    from openpyxl.chart import BarChart, LineChart, PieChart, Reference, ScatterChart

    _check_write()
    _check_confirm(confirm)
    resolved = _check_path(path)

    _valid_types = {"bar", "line", "scatter", "pie"}
    if chart_type not in _valid_types:
        raise ExcelMCPError(
            f"Invalid chart_type: {chart_type!r}. Valid values: {sorted(_valid_types)}"
        )

    if ":" not in data_range:
        raise ExcelMCPError(
            f"data_range must be a range like 'A1:B5', got {data_range!r}"
        )

    try:
        min_col, min_row, max_col, max_row = openpyxl.utils.cell.range_boundaries(
            data_range.upper()
        )
    except Exception as exc:
        raise ExcelMCPError(f"Invalid data_range {data_range!r}: {exc}") from exc

    # Resolve anchor vs target_cell
    effective_cell = anchor if anchor is not None else target_cell

    # Validate effective_cell is a bare cell reference (e.g. 'D1')
    if not _re.fullmatch(r"[A-Za-z]{1,3}\d+", effective_cell):
        raise ExcelMCPError(
            f"target_cell/anchor must be a bare cell reference (e.g. 'D1'), "
            f"got {effective_cell!r}"
        )

    # Validate px dimensions
    if width_px is not None and not (0 < width_px <= 5000):
        raise ExcelMCPError(f"width_px must be 1–5000, got {width_px}")
    if height_px is not None and not (0 < height_px <= 5000):
        raise ExcelMCPError(f"height_px must be 1–5000, got {height_px}")
    _flush_if_dirty(str(resolved))
    wb = _load_wb(resolved, for_write=True)
    if sheet not in wb.sheetnames:
        raise ValidationError(f"Sheet not found: {sheet!r}")
    ws = wb[sheet]

    chart_map = {
        "bar": BarChart,
        "line": LineChart,
        "scatter": ScatterChart,
        "pie": PieChart,
    }
    chart = chart_map[chart_type]()
    if title:
        chart.title = title

    data_ref = Reference(
        ws,
        min_col=min_col,
        min_row=min_row,
        max_col=max_col,
        max_row=max_row,
    )
    chart.add_data(data_ref, titles_from_data=True)

    if width_px is not None:
        chart.width = round(width_px * 2.54 / 96, 4)
    if height_px is not None:
        chart.height = round(height_px * 2.54 / 96, 4)

    ws.add_chart(chart, effective_cell)

    _save_and_evict(wb, resolved)
    _audit_log("create_chart", resolved, sheet, data_range, None)
    return {
        "ok": True,
        "sheet": sheet,
        "chart_type": chart_type,
        "title": title,
        "target_cell": effective_cell,
        "width_cm": round(chart.width, 4),
        "height_cm": round(chart.height, 4),
        "message": (
            f"Created {chart_type} chart '{title}' at {effective_cell} in sheet {sheet!r}"
        ),
    }


def update_chart_data(
    path: str,
    sheet: str,
    chart_index: int,
    data_range: str,
    confirm: bool = False,
) -> dict:
    """Replace the data range of an existing chart (0-based chart_index).

    chart_index is 0-based (see list_charts to discover existing charts).
    data_range must be a range containing ':', e.g. 'A1:B5'.
    Requires EXCEL_ENABLE_WRITE=true and confirm=True.
    """
    from openpyxl.chart import Reference

    # Security gate sequence (required order per codebase convention)
    _check_write()
    _check_confirm(confirm)
    resolved = _check_path(path)
    _flush_if_dirty(str(resolved))

    # Validate data_range
    if ":" not in data_range:
        raise ExcelMCPError(
            f"data_range must be a range (e.g. A1:B10), got {data_range!r}"
        )
    try:
        min_col, min_row, max_col, max_row = openpyxl.utils.cell.range_boundaries(
            data_range.upper()
        )
    except Exception as exc:
        raise ExcelMCPError(f"Invalid data_range {data_range!r}: {exc}") from exc

    wb = _load_wb(resolved, for_write=True)
    if sheet not in wb.sheetnames:
        raise ValidationError(f"Sheet not found: {sheet!r}")
    ws = wb[sheet]

    # Bounds check chart_index
    charts = ws._charts
    if chart_index < 0 or chart_index >= len(charts):
        raise ExcelMCPError(
            f"chart_index {chart_index} out of range "
            f"(sheet {sheet!r} has {len(charts)} chart(s))"
        )

    chart = charts[chart_index]
    chart_type_name = type(chart).__name__

    # Replace series
    chart.series.clear()
    data_ref = Reference(
        ws,
        min_col=min_col,
        min_row=min_row,
        max_col=max_col,
        max_row=max_row,
    )
    chart.add_data(data_ref, titles_from_data=True)

    _save_and_evict(wb, resolved)
    _audit_log("update_chart_data", resolved, sheet, data_range, None)
    return {
        "ok": True,
        "chart_index": chart_index,
        "chart_type": chart_type_name,
        "data_range": data_range,
        "message": (
            f"Updated chart {chart_index} ({chart_type_name}) data to "
            f"{data_range} in sheet {sheet!r}"
        ),
    }


def list_charts(path: str, sheet: str) -> dict:
    """List all charts in a worksheet.

    Returns {"charts": [...], "count": int} where each entry has
    title, type, and anchor.
    """
    resolved = _check_path(path)
    wb = _load_wb(resolved)
    if sheet not in wb.sheetnames:
        raise ValidationError(f"Sheet not found: {sheet!r}")
    ws = wb[sheet]

    charts = []
    for idx, chart in enumerate(ws._charts):
        # chart.title may be a Title object
        # (openpyxl wraps string in Title on assignment)
        raw_title = chart.title
        if raw_title is None:
            title_str = ""
        elif isinstance(raw_title, str):
            title_str = raw_title
        else:
            # openpyxl Title → tx.rich.p[0].r[0].t
            try:
                parts = []
                for para in raw_title.tx.rich.p:
                    for run in para.r:
                        if run.t:
                            parts.append(run.t)
                title_str = " ".join(parts)
            except Exception:
                title_str = ""
        # Simplify anchor to "col,row" string
        anc = chart.anchor
        if anc is None:
            anchor_str = ""
        else:
            try:
                anchor_str = f"{anc._from.col},{anc._from.row}"
            except Exception:
                anchor_str = str(anc)
        charts.append(
            {
                "index": idx,
                "title": title_str,
                "type": type(chart).__name__,
                "anchor": anchor_str,
            }
        )
    return {"charts": charts, "count": len(charts)}


def delete_chart(
    path: str,
    sheet: str,
    chart_index: int,
    confirm: bool = False,
) -> dict:
    """Delete a chart by 0-based index from a worksheet.

    Requires EXCEL_ENABLE_WRITE=true and confirm=True.
    """
    _check_write()
    _check_confirm(confirm)
    resolved = _check_path(path)
    _flush_if_dirty(str(resolved))
    wb = _load_wb(resolved, for_write=True)
    if sheet not in wb.sheetnames:
        raise ValidationError(f"Sheet not found: {sheet!r}")
    ws = wb[sheet]

    charts = ws._charts
    if chart_index < 0 or chart_index >= len(charts):
        raise ExcelMCPError(
            f"chart_index {chart_index} out of range "
            f"(sheet has {len(charts)} chart(s))"
        )

    del charts[chart_index]
    _save_and_evict(wb, resolved)
    _audit_log("delete_chart", resolved, sheet, None, None)
    return {
        "ok": True,
        "sheet": sheet,
        "chart_index": chart_index,
        "message": f"Deleted chart at index {chart_index} from sheet {sheet!r}",
    }
