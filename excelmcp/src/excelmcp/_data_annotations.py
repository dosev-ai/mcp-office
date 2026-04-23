"""Cell annotations: comments, hyperlinks.

Internal sub-module. Import via excelmcp._data (public router).

SECURITY: _validate_hyperlink_url is private (not in __all__).
EXCEL_ALLOW_FILE_HYPERLINKS is read inside function body via os.getenv().
"""
from __future__ import annotations

from ._core import (
    ValidationError,
    _audit_log,
    _check_confirm,
    _check_path,
    _check_write,
    _flush_if_dirty,
    _load_wb,
    _save_and_evict,
    openpyxl,
)

__all__ = [
    "get_hyperlinks",
    "add_hyperlink",
    "remove_hyperlink",
    "bulk_add_hyperlinks",
    "add_cell_comment",
    "delete_cell_comment",
    "get_cell_comment",
    "bulk_add_comments",
]


def _validate_hyperlink_url(url: str) -> None:
    """Validate a hyperlink URL for safe schemes (OWASP A03/A05 injection prevention)."""
    import os as _os
    import urllib.parse

    parsed = urllib.parse.urlparse(url.strip())
    scheme = parsed.scheme.lower()
    _DEFAULT_ALLOWED = {"https", "http", "mailto"}
    allow_file = _os.getenv("EXCEL_ALLOW_FILE_HYPERLINKS", "").lower() == "true"
    if scheme == "file":
        if not allow_file:
            raise ValidationError(
                "file:// URLs are disabled by default (NTLM credential relay risk on "
                "Windows). Set EXCEL_ALLOW_FILE_HYPERLINKS=true to enable."
            )
        return
    _BLOCKED = {"javascript", "data", "vbscript"}
    if scheme in _BLOCKED:
        raise ValidationError(
            f"URL scheme {scheme!r} is not allowed. Permitted: {sorted(_DEFAULT_ALLOWED)}."
        )
    if scheme not in _DEFAULT_ALLOWED:
        raise ValidationError(
            f"URL scheme {scheme!r} is not allowed. Permitted: {sorted(_DEFAULT_ALLOWED)}. "
            "Pass file:// URLs only when EXCEL_ALLOW_FILE_HYPERLINKS=true."
        )
    if len(url) > 2048:
        raise ValidationError(f"URL exceeds max length (2048 chars): {len(url)}")


def get_hyperlinks(path: str, sheet: str) -> dict:
    """Return all hyperlinks in a worksheet.

    Read-only — no write gate required.
    Returns ``{"sheet": sheet, "count": int, "hyperlinks": list}``.
    Each entry has: address, url, display_text, tooltip.
    """
    resolved = _check_path(path)
    wb_obj = _load_wb(resolved)
    if sheet not in wb_obj.sheetnames:
        raise ValidationError(f"Sheet not found: {sheet!r}")
    ws = wb_obj[sheet]
    links = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.hyperlink is not None:
                hl = cell.hyperlink
                links.append(
                    {
                        "address": cell.coordinate,
                        "url": getattr(hl, "target", str(hl)),
                        "display_text": (
                            cell.value if isinstance(cell.value, str) else None
                        ),
                        "tooltip": getattr(hl, "tooltip", None),
                    }
                )
    return {"sheet": sheet, "count": len(links), "hyperlinks": links}


def add_hyperlink(
    path: str,
    sheet: str,
    address: str,
    url: str,
    display_text: str | None = None,
    tooltip: str | None = None,
    confirm: bool = False,
) -> dict:
    """Add or update a hyperlink on a cell.

    Returns ``{"ok": True, "sheet": sheet, "address": address, "url": url, ...}``.
    Requires EXCEL_ENABLE_WRITE=true and confirm=True.
    """
    _check_write()
    _check_confirm(confirm)
    _validate_hyperlink_url(url)
    from openpyxl.worksheet.hyperlink import Hyperlink  # noqa: PLC0415

    resolved = _check_path(path)
    _flush_if_dirty(str(resolved))
    wb_obj = _load_wb(resolved, for_write=True)
    if sheet not in wb_obj.sheetnames:
        raise ValidationError(f"Sheet not found: {sheet!r}")
    ws = wb_obj[sheet]
    addr = address.upper()
    # Validate display_text BEFORE writing anything to the workbook (R-003).
    if display_text is not None and isinstance(display_text, str):
        _stripped_dt = display_text.lstrip()
        try:
            float(_stripped_dt)
        except (ValueError, TypeError):
            if _stripped_dt[:1] in ("=", "+", "-", "@"):
                raise ValidationError(
                    "Formula injection blocked in display_text: strings starting with "
                    "'=', '+', '-', or '@' are not allowed."
                )
    hl = Hyperlink(ref=addr, target=url)
    if tooltip:
        hl.tooltip = tooltip
    ws[addr].hyperlink = hl
    if display_text is not None:
        ws[addr].value = display_text  # already validated above
    _save_and_evict(wb_obj, resolved)
    _audit_log("add_hyperlink", resolved, sheet, addr, None)
    return {
        "ok": True,
        "sheet": sheet,
        "address": addr,
        "url": url,
        "display_text": display_text,
        "message": f"Hyperlink added to {addr}",
    }


def remove_hyperlink(
    path: str,
    sheet: str,
    address: str,
    confirm: bool = False,
) -> dict:
    """Remove a hyperlink from a cell.

    Returns ``{"ok": True, "sheet": sheet, "address": address, "message": ...}``.
    Requires EXCEL_ENABLE_WRITE=true and confirm=True.
    """
    _check_write()
    _check_confirm(confirm)
    resolved = _check_path(path)
    _flush_if_dirty(str(resolved))
    wb_obj = _load_wb(resolved, for_write=True)
    if sheet not in wb_obj.sheetnames:
        raise ValidationError(f"Sheet not found: {sheet!r}")
    ws = wb_obj[sheet]
    addr = address.upper()
    ws[addr].hyperlink = None
    _save_and_evict(wb_obj, resolved)
    _audit_log("remove_hyperlink", resolved, sheet, addr, None)
    return {
        "ok": True,
        "sheet": sheet,
        "address": addr,
        "message": f"Hyperlink removed from {addr}",
    }


def bulk_add_hyperlinks(
    path: str,
    sheet: str,
    links: list[dict],
    apply_style: bool = True,
    confirm: bool = False,
) -> dict:
    """Add multiple hyperlinks in a single atomic call.

    Each item in ``links`` is a dict with keys:
    - ``address`` (str, required): cell address in A1 notation.
    - ``url`` (str, required): target URL (validated for safe schemes).
    - ``display_text`` (str | None): text to write into the cell.
    - ``tooltip`` (str | None): hover tooltip — accepted but not stored by
      openpyxl's cell model (limitation of the library; no error is raised).

    All items are validated before any write occurs (fail-fast atomicity).

    If ``apply_style=True`` (default), a blue-underline hyperlink font
    (``Font(color="0563C1", underline="single")``) is applied to each cell.

    Returns ``{"ok": True, "sheet": sheet, "links_added": int}``.
    Requires EXCEL_ENABLE_WRITE=true and confirm=True.
    """
    from openpyxl.styles import Font as _Font  # noqa: PLC0415
    from openpyxl.worksheet.hyperlink import Hyperlink  # noqa: PLC0415

    _check_write()
    _check_confirm(confirm)
    if not links:
        raise ValidationError("links must be non-empty")
    resolved = _check_path(path)
    if len(links) > 500:
        raise ValidationError(
            f"too many links — max 500 per call (got {len(links)})"
        )
    # Validate ALL items before writing (fail-fast atomicity)
    for idx, item in enumerate(links):
        address = item.get("address")
        url = item.get("url")
        if not address:
            raise ValidationError(f"links[{idx}] missing 'address'")
        if not url:
            raise ValidationError(f"links[{idx}] missing 'url'")
        try:
            openpyxl.utils.cell.coordinate_to_tuple(address.upper())
        except Exception as exc:
            raise ValidationError(
                f"links[{idx}] invalid address {address!r}: {exc}"
            ) from exc
        _validate_hyperlink_url(url)
        display_text = item.get("display_text")
        if display_text is not None and isinstance(display_text, str):
            _stripped_dt = display_text.lstrip()
            try:
                float(_stripped_dt)
            except (ValueError, TypeError):
                if _stripped_dt[:1] in ("=", "+", "-", "@"):
                    raise ValidationError(
                        f"Formula injection blocked in display_text at links[{idx}]: "
                        "strings starting with '=', '+', '-', or '@' are not allowed."
                    )
    _flush_if_dirty(str(resolved))
    wb_obj = _load_wb(resolved, for_write=True)
    if sheet not in wb_obj.sheetnames:
        raise ValidationError(f"Sheet not found: {sheet!r}")
    ws = wb_obj[sheet]
    _hl_font = _Font(color="0563C1", underline="single") if apply_style else None
    for item in links:
        addr = item["address"].upper()
        url = item["url"]
        hl = Hyperlink(ref=addr, target=url)
        ws[addr].hyperlink = hl
        if item.get("display_text") is not None:
            ws[addr].value = item["display_text"]  # already validated above
        if _hl_font is not None:
            ws[addr].font = _hl_font
    _save_and_evict(wb_obj, resolved)
    _audit_log("bulk_add_hyperlinks", resolved, sheet, None, len(links))
    return {"ok": True, "sheet": sheet, "links_added": len(links)}


def add_cell_comment(
    path: str,
    sheet: str,
    address: str,
    text: str,
    author: str = "ExcelMCP",
    confirm: bool = False,
) -> dict:
    """Add a comment to a cell.

    Returns ``{"ok": True, "sheet": sheet, "address": address, "author": author, "message": ...}``.
    Requires EXCEL_ENABLE_WRITE=true and confirm=True.
    """
    from openpyxl.comments import Comment as _Comment  # noqa: PLC0415

    _check_write()
    _check_confirm(confirm)
    resolved = _check_path(path)
    _flush_if_dirty(str(resolved))
    wb = _load_wb(resolved, for_write=True)
    if sheet not in wb.sheetnames:
        raise ValidationError(f"Sheet not found: {sheet!r}")
    ws = wb[sheet]
    addr = address.upper()
    ws[addr].comment = _Comment(text, author)
    _save_and_evict(wb, resolved)
    _audit_log("add_cell_comment", resolved, sheet, addr, None)
    return {
        "ok": True,
        "sheet": sheet,
        "address": addr,
        "author": author,
        "message": f"Added comment to {addr}",
    }


def delete_cell_comment(
    path: str,
    sheet: str,
    address: str,
    confirm: bool = False,
) -> dict:
    """Delete the comment from a cell.

    Returns ``{"ok": True, "sheet": sheet, "address": address, "message": ...}``.
    Requires EXCEL_ENABLE_WRITE=true and confirm=True.
    """
    _check_write()
    _check_confirm(confirm)
    resolved = _check_path(path)
    _flush_if_dirty(str(resolved))
    wb = _load_wb(resolved, for_write=True)
    if sheet not in wb.sheetnames:
        raise ValidationError(f"Sheet not found: {sheet!r}")
    ws = wb[sheet]
    addr = address.upper()
    ws[addr].comment = None
    _save_and_evict(wb, resolved)
    _audit_log("delete_cell_comment", resolved, sheet, addr, None)
    return {
        "ok": True,
        "sheet": sheet,
        "address": addr,
        "message": f"Deleted comment from {addr}",
    }


def get_cell_comment(path: str, sheet: str, address: str) -> dict:
    """Return the comment on a cell, if any.

    Returns a dict with keys:
    - ``address``: normalised cell address
    - ``has_comment``: True when a comment is present
    - ``text``: comment text or None
    - ``author``: comment author or None
    """
    resolved = _check_path(path)
    wb = _load_wb(resolved)
    if sheet not in wb.sheetnames:
        raise ValidationError(f"Sheet not found: {sheet!r}")
    ws = wb[sheet]
    addr = address.upper()
    try:
        cell = ws[addr]
    except (KeyError, ValueError) as exc:
        raise ValidationError(f"Invalid cell address {address!r}: {exc}") from exc
    if cell.comment is not None:
        try:
            text = cell.comment.text
            author = cell.comment.author
        except AttributeError:
            text = None
            author = None
        return {"address": addr, "has_comment": True, "text": text, "author": author}
    return {"address": addr, "has_comment": False, "text": None, "author": None}


def bulk_add_comments(
    path: str,
    sheet: str,
    comments: list[dict],
    overwrite: bool = True,
    confirm: bool = False,
) -> dict:
    """Add multiple cell comments in one call.

    Each dict: ``{"address": str, "text": str, "author": str | None}``.
    ``overwrite=True`` (default) replaces existing comments.
    ``overwrite=False`` skips cells that already have a comment.

    Returns ``{"ok": True, "comments_added": int, "skipped": int, "cells": list[str]}``.
    Requires EXCEL_ENABLE_WRITE=true and confirm=True.
    """
    from openpyxl.comments import Comment as _Comment  # noqa: PLC0415

    _check_write()
    _check_confirm(confirm)

    if not comments:
        raise ValidationError("comments must be non-empty")
    if len(comments) > 500:
        raise ValidationError("too many comments — max 500 per call")

    # Validate ALL items BEFORE writing any (fail-fast)
    validated: list[tuple[str, str, str]] = []
    for i, item in enumerate(comments):
        addr = (item.get("address") or "").strip()
        text = item.get("text", "")
        author = item.get("author") or "ExcelMCP"

        try:
            openpyxl.utils.cell.coordinate_from_string(addr.upper())
        except Exception as exc:
            raise ValidationError(
                f"Invalid cell address at index {i}: {addr!r}"
            ) from exc

        if not isinstance(text, str) or not text.strip():
            raise ValidationError(f"text must be non-empty at index {i}")
        if len(text) > 32767:
            raise ValidationError(
                f"text too long at index {i}: max 32767 chars, got {len(text)}"
            )

        validated.append((addr.upper(), text, str(author)))

    # Deduplicate: last item in list wins for duplicate addresses
    deduped: dict[str, tuple[str, str]] = {}
    for addr, text, author in validated:
        deduped[addr] = (text, author)

    resolved = _check_path(path)
    _flush_if_dirty(str(resolved))
    wb_obj = _load_wb(resolved, for_write=True)
    if sheet not in wb_obj.sheetnames:
        raise ValidationError(f"Sheet not found: {sheet!r}")
    ws = wb_obj[sheet]

    comments_added = 0
    skipped = 0
    cells: list[str] = []

    for addr, (text, author) in deduped.items():
        cell = ws[addr]
        if not overwrite and cell.comment is not None:
            skipped += 1
            continue
        cell.comment = _Comment(text, author)
        comments_added += 1
        cells.append(addr)

    _save_and_evict(wb_obj, resolved)
    _audit_log("bulk_add_comments", resolved, sheet, None, comments_added)
    return {
        "ok": True,
        "sheet": sheet,
        "comments_added": comments_added,
        "skipped": skipped,
        "cells": cells,
    }
