"""
Sheet and workbook property operations: sheet_properties, workbook protection,
calculation mode, and table creation.
"""
from __future__ import annotations

from ._active_wb import _check_file_not_open_in_excel
from ._core import (
    ValidationError,
    _audit_log,
    _check_confirm,
    _check_path,
    _check_write,
    _flush_if_dirty,
    _load_wb,
    _save_and_evict,
)


def get_sheet_properties(path: str, sheet: str) -> dict:
    """Return tab_color and visibility state for a sheet.

    Returns ``{"sheet": sheet, "tab_color": str | None, "state": str}``.
    Read-only; no confirm or write gate required.
    """
    resolved = _check_path(path)
    wb = _load_wb(resolved)
    if sheet not in wb.sheetnames:
        raise ValidationError(f"Sheet not found: {sheet!r}")
    ws = wb[sheet]
    color = ws.sheet_properties.tabColor
    # strip ARGB alpha prefix if present (e.g. "FFFF0000" -> "FF0000")
    tab_color = color.rgb[-6:] if (color and color.rgb) else None
    state = ws.sheet_state
    return {"sheet": sheet, "tab_color": tab_color, "state": state}


def set_sheet_properties(
    path: str,
    sheet: str,
    tab_color: str | None = None,
    state: str | None = None,
    show_gridlines: bool | None = None,
    show_row_col_headers: bool | None = None,
    zoom_scale: int | None = None,
    confirm: bool = False,
) -> dict:
    """Set tab colour, visibility state, and/or view options for a sheet.

    tab_color: 6 hex digits without '#', e.g. 'FF0000'. Pass None to leave unchanged.
    state: one of 'visible', 'hidden', 'veryHidden'. Pass None to leave unchanged.
    show_gridlines: True/False to show/hide gridlines. None = no change.
    show_row_col_headers: True/False to show/hide row & column headers. None = no change.

    Returns ``{"ok": True, "sheet": sheet, "tab_color": tab_color, "state": state,
    "show_gridlines": ..., "show_row_col_headers": ..., "message": ...}``.
    Requires EXCEL_ENABLE_WRITE=true and confirm=True.
    """
    import re as _re

    from openpyxl.styles import Color as _Color

    _check_write()
    _check_confirm(confirm)
    if tab_color is not None and not _re.match(r"^[0-9A-Fa-f]{6}$", tab_color):
        raise ValidationError(
            f"Invalid tab_color {tab_color!r}: must be exactly 6 hex digits (e.g. 'FF0000')"
        )
    _VALID_STATES = {"visible", "hidden", "veryHidden"}
    if state is not None and state not in _VALID_STATES:
        raise ValidationError(
            f"Invalid state {state!r}: must be one of {sorted(_VALID_STATES)}"
        )
    resolved = _check_path(path)
    _check_file_not_open_in_excel(resolved)   # LC-2: block if file is open in Excel
    _flush_if_dirty(str(resolved))
    wb = _load_wb(resolved, for_write=True)
    if sheet not in wb.sheetnames:
        raise ValidationError(f"Sheet not found: {sheet!r}")
    ws = wb[sheet]
    if tab_color is not None:
        ws.sheet_properties.tabColor = _Color(rgb="FF" + tab_color)
    if state is not None:
        ws.sheet_state = state
    if show_gridlines is not None:
        ws.sheet_view.showGridLines = show_gridlines
    if show_row_col_headers is not None:
        ws.sheet_view.showRowColHeaders = show_row_col_headers
    if zoom_scale is not None:
        if not (10 <= zoom_scale <= 400):
            raise ValidationError("zoom_scale must be between 10 and 400")
        ws.sheet_view.zoomScale = zoom_scale
    _save_and_evict(wb, resolved)
    _audit_log("set_sheet_properties", str(resolved), sheet, None, None)
    return {
        "ok": True,
        "sheet": sheet,
        "tab_color": tab_color,
        "state": state,
        "show_gridlines": show_gridlines,
        "show_row_col_headers": show_row_col_headers,
        "zoom_scale": zoom_scale,
        "message": "Sheet properties updated",
    }


_VALID_CALC_MODES: frozenset[str] = frozenset({"auto", "manual", "autoNoTable"})


def set_workbook_calc_mode(
    path: str,
    calc_mode: str | None = None,
    full_calc_on_load: bool | None = None,
    confirm: bool = False,
) -> dict:
    """Set workbook calculation mode and/or fullCalcOnLoad flag.

    Args:
        path: Path to the workbook.
        calc_mode: One of 'auto', 'manual', 'autoNoTable'. None = no change.
        full_calc_on_load: If True, Excel will recalculate all formulas when the
            workbook is next opened. None = no change.
        confirm: Must be True to apply changes.
    """
    _check_write()
    _check_confirm(confirm)
    if calc_mode is not None and calc_mode not in _VALID_CALC_MODES:
        raise ValidationError(
            f"calc_mode must be one of {sorted(_VALID_CALC_MODES)}, got {calc_mode!r}"
        )
    resolved = _check_path(path)
    _check_file_not_open_in_excel(resolved)   # LC-2: block if file is open in Excel
    _flush_if_dirty(str(resolved))
    wb = _load_wb(resolved, for_write=True)
    if calc_mode is not None:
        wb.calculation.calcMode = calc_mode
    if full_calc_on_load is not None:
        wb.calculation.fullCalcOnLoad = full_calc_on_load
    _save_and_evict(wb, resolved)
    _audit_log("set_workbook_calc_mode", str(resolved), None, None, None)
    return {
        "calc_mode": wb.calculation.calcMode,
        "full_calc_on_load": wb.calculation.fullCalcOnLoad,
    }


def protect_workbook(
    path: str,
    password: str | None = None,
    lock_structure: bool = True,
    lock_windows: bool = False,
    confirm: bool = False,
) -> dict:
    """Enable workbook-level structure/windows protection.

    password: optional password string (<= 255 chars).
    lock_structure: prevent adding/deleting/moving sheets (default True).
    lock_windows: prevent resizing/repositioning workbook windows (default False).

    Returns ``{"ok": True, "protected": True, "lock_structure": ..., "lock_windows": ...,
    "has_password": bool, "message": ...}``.
    Requires EXCEL_ENABLE_WRITE=true and confirm=True.
    """
    _check_write()
    _check_confirm(confirm)
    if password is not None and len(password) > 255:
        raise ValidationError("password must be <= 255 chars")
    resolved = _check_path(path)
    _check_file_not_open_in_excel(resolved)   # LC-2: block if file is open in Excel
    _flush_if_dirty(str(resolved))
    wb = _load_wb(resolved, for_write=True)
    wb.security.lockStructure = lock_structure
    wb.security.lockWindows = lock_windows
    if password is not None:
        wb.security.workbookPassword = password
    _save_and_evict(wb, resolved)
    _audit_log("protect_workbook", str(resolved), None, None, None)
    return {
        "ok": True,
        "protected": True,
        "lock_structure": lock_structure,
        "lock_windows": lock_windows,
        "has_password": password is not None,
        "message": "Workbook protection enabled",
    }


def create_table(
    path: str,
    sheet: str,
    table_range: str,
    table_name: str,
    style: str = "TableStyleMedium9",
    confirm: bool = False,
) -> dict:
    """Create a native Excel table (ListObject) over the specified range.

    style: Excel table style name (default 'TableStyleMedium9'). Excel silently
        ignores unknown style names.
    Returns ``{"table_name": ..., "ref": ..., "sheet": ..., "style": ...}``.
    Requires EXCEL_ENABLE_WRITE=true and confirm=True.
    """
    import re as _re

    from openpyxl.worksheet.table import Table, TableStyleInfo

    _check_write()
    _check_confirm(confirm)

    # Validate table_name: must match ^[A-Za-z_][A-Za-z0-9_]{0,254}$
    if not _re.match(r"^[A-Za-z_][A-Za-z0-9_]{0,254}$", table_name):
        raise ValidationError(
            f"Invalid table_name {table_name!r}: must start with a letter or underscore, "
            f"contain only letters/digits/underscores, and be 1-255 chars"
        )

    # Validate range format: must be A1-notation like "A1:D10"
    range_match = _re.match(r"^([A-Za-z]+)(\d+):([A-Za-z]+)(\d+)$", table_range)
    if not range_match:
        raise ValidationError(
            f"Invalid table_range {table_range!r}: expected A1-notation range like 'A1:D10'"
        )
    start_row = int(range_match.group(2))
    end_row = int(range_match.group(4))
    if end_row - start_row < 1:
        raise ValidationError(
            f"table_range {table_range!r} must span at least 2 rows "
            f"(header + 1 data row); given range has {end_row - start_row + 1} row(s)"
        )

    resolved = _check_path(path)
    _check_file_not_open_in_excel(resolved)   # LC-2: block if file is open in Excel
    _flush_if_dirty(str(resolved))
    wb = _load_wb(resolved, for_write=True)

    if sheet not in wb.sheetnames:
        raise ValidationError(f"Sheet not found: {sheet!r}")

    # Check for duplicate table name across ALL sheets in the workbook
    for ws_iter in wb.worksheets:
        for tbl in ws_iter.tables.values():
            if tbl.displayName == table_name:
                raise ValidationError(
                    f"A table named {table_name!r} already exists in the workbook"
                )

    ws = wb[sheet]
    tbl = Table(displayName=table_name, ref=table_range)
    tbl.tableStyleInfo = TableStyleInfo(
        name=style,
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(tbl)
    # Guard: if a worksheet-level autoFilter shares the same top-left cell as the
    # new table, clear it. A coexisting ws.auto_filter + table autoFilter at the same
    # cell is an OOXML conflict that triggers Excel's recovery dialog.
    if ws.auto_filter.ref:
        try:
            from openpyxl.utils.cell import range_boundaries as _rb
            af_mc, af_mr, _, _ = _rb(ws.auto_filter.ref.upper())
            tbl_mc, tbl_mr, _, _ = _rb(table_range.upper())
            if (af_mc, af_mr) == (tbl_mc, tbl_mr):
                ws.auto_filter.ref = None
        except (ValueError, TypeError) as _exc:
            import logging as _logging
            _logging.getLogger(__name__).warning(
                "auto_filter guard skipped in create_table: %s", _exc
            )
    _save_and_evict(wb, resolved)
    _audit_log("create_table", str(resolved), sheet, table_range, 0)
    return {
        "table_name": table_name,
        "ref": table_range,
        "sheet": sheet,
        "style": style,
    }
