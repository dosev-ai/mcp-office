"""Write helpers: fill_formula_range, write_range_with_format.

Internal sub-module. Import via excelmcp._data (public router).
"""
from __future__ import annotations

from ._core import (
    ValidationError,
    _audit_log,
    _check_confirm,
    _check_path,
    _check_range_size,
    _check_write,
    _flush_if_dirty,
    _load_wb,
    _parse_a1_range,
    _save_and_evict,
    _validate_formula_safe,
    openpyxl,
)

__all__ = [
    "fill_formula_range",
    "write_range_with_format",
]

_ALLOWED_STYLE_KEYS: frozenset[str] = frozenset(
    {"bold", "italic", "underline", "font_color", "bg_color", "font_size", "font_name"}
)
_ALLOWED_ALIGNMENT_KEYS: frozenset[str] = frozenset(
    {"horizontal", "vertical", "wrap_text", "shrink_to_fit", "text_rotation", "indent"}
)


def fill_formula_range(
    path: str,
    sheet: str,
    anchor_cell: str,
    end_cell: str,
    formula: str,
    confirm: bool = False,
) -> dict:
    """Fill a range with a formula, relativizing references for each cell.

    ``anchor_cell``: the cell whose formula is provided as-is (e.g. ``"B2"``).
    ``end_cell``: last cell in the fill range (e.g. ``"B1000"``).
    ``formula``: must start with ``'='``. Row/col references are relativized
    using :class:`openpyxl.formula.translate.Translator` for all cells other
    than the anchor.

    Returns ``{"ok": True, "range": "B2:B1000", "filled_cells": N}``.
    Requires EXCEL_ENABLE_WRITE=true and confirm=True.
    """
    from openpyxl.formula.translate import Translator  # noqa: PLC0415

    _check_write()
    _check_confirm(confirm)
    _validate_formula_safe(formula)

    try:
        anchor_col_str, anchor_row = openpyxl.utils.cell.coordinate_from_string(
            anchor_cell.upper()
        )
        anchor_col = openpyxl.utils.cell.column_index_from_string(anchor_col_str)
    except Exception as exc:
        raise ValidationError(f"Invalid anchor cell: {anchor_cell!r}") from exc

    try:
        end_col_str, end_row = openpyxl.utils.cell.coordinate_from_string(
            end_cell.upper()
        )
        end_col = openpyxl.utils.cell.column_index_from_string(end_col_str)
    except Exception as exc:
        raise ValidationError(f"Invalid end cell: {end_cell!r}") from exc

    n_cells = (end_row - anchor_row + 1) * (end_col - anchor_col + 1)
    if n_cells <= 0:
        raise ValidationError("end_cell must be at or after anchor_cell")
    _check_range_size(n_cells)

    resolved = _check_path(path)
    _flush_if_dirty(str(resolved))
    wb_obj = _load_wb(resolved, for_write=True)
    if sheet not in wb_obj.sheetnames:
        raise ValidationError(f"Sheet not found: {sheet!r}")
    ws = wb_obj[sheet]

    anchor_coord = f"{anchor_col_str}{anchor_row}"
    filled = 0

    for row in ws.iter_rows(
        min_row=anchor_row,
        max_row=end_row,
        min_col=anchor_col,
        max_col=end_col,
    ):
        for cell in row:
            dest_coord = cell.coordinate
            if dest_coord == anchor_coord:
                cell.value = formula
            else:
                cell.value = Translator(formula, origin=anchor_coord).translate_formula(
                    dest_coord
                )
            filled += 1

    range_ref = f"{anchor_coord}:{end_col_str}{end_row}"
    _save_and_evict(wb_obj, resolved)
    _audit_log("fill_formula_range", resolved, sheet, range_ref, filled)
    return {"ok": True, "range": range_ref, "filled_cells": filled}


def write_range_with_format(
    path: str,
    sheet: str,
    address: str,
    data: list,
    style: dict | None = None,
    number_format: str | None = None,
    alignment: dict | None = None,
    confirm: bool = False,
) -> dict:
    """Write a 2-D array of values to a range and apply consistent formatting in one call.

    Requires EXCEL_ENABLE_WRITE=true and confirm=True.
    Returns ``{"cells_written": int, "address": str, "sheet": str}``.
    """
    _check_write()
    _check_confirm(confirm)

    # Early-exit on empty data (before path resolution)
    if not data:
        return {"cells_written": 0}

    # Validate style and alignment keys before any file I/O
    if style is not None:
        unknown_style = set(style) - _ALLOWED_STYLE_KEYS
        if unknown_style:
            raise ValidationError(
                f"Unknown style key(s): {sorted(unknown_style)}. "
                f"Allowed: {sorted(_ALLOWED_STYLE_KEYS)}"
            )
    if alignment is not None:
        unknown_align = set(alignment) - _ALLOWED_ALIGNMENT_KEYS
        if unknown_align:
            raise ValidationError(
                f"Unknown alignment key(s): {sorted(unknown_align)}. "
                f"Allowed: {sorted(_ALLOWED_ALIGNMENT_KEYS)}"
            )

    # Validate address and get start cell
    min_col, min_row, _max_col, _max_row = _parse_a1_range(address)

    # Range-size guard using data dimensions
    n_cols = len(data[0]) if data[0] else 0
    if n_cols == 0:
        return {"cells_written": 0}
    _check_range_size(len(data) * n_cols)

    resolved = _check_path(path)
    _flush_if_dirty(str(resolved))
    wb_obj = _load_wb(resolved, for_write=True)
    if sheet not in wb_obj.sheetnames:
        raise ValidationError(f"Sheet not found: {sheet!r}")
    ws = wb_obj[sheet]

    # Build style objects once (range-constant)
    _font = None
    _fill = None
    if style:
        from openpyxl.styles import Font as _Font  # noqa: PLC0415
        from openpyxl.styles import PatternFill as _PatternFill

        bold = style.get("bold")
        italic = style.get("italic")
        underline = style.get("underline")
        font_color = style.get("font_color")
        font_size = style.get("font_size")
        font_name = style.get("font_name")
        _font = _Font(
            bold=bold,
            italic=italic,
            underline=underline,
            color=font_color,
            size=font_size,
            name=font_name,
        )
        bg_color = style.get("bg_color")
        if bg_color is not None:
            _fill = _PatternFill(fill_type="solid", fgColor=bg_color.lstrip("#"))

    _alignment_obj = None
    if alignment:
        from openpyxl.styles import Alignment as _Alignment  # noqa: PLC0415

        _alignment_obj = _Alignment(**alignment)

    total = 0
    for r_idx, row in enumerate(data):
        for c_idx, val in enumerate(row):
            if isinstance(val, str) and val.lstrip()[:1] in ("=", "+", "-", "@"):
                raise ValidationError(
                    "Formula injection blocked: to write formulas, use fill_formula_range"
                )
            cell = ws.cell(row=min_row + r_idx, column=min_col + c_idx)
            cell.value = val
            if number_format is not None:
                cell.number_format = number_format
            if _font is not None:
                cell.font = _font
            if _fill is not None:
                cell.fill = _fill
            if _alignment_obj is not None:
                cell.alignment = _alignment_obj
            total += 1

    _save_and_evict(wb_obj, resolved)
    _audit_log("write_range_with_format", resolved, sheet, address, total)
    return {"cells_written": total, "address": address, "sheet": sheet}
