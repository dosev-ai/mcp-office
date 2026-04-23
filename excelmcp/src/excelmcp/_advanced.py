"""
Advanced worksheet operations: merge/unmerge cells, freeze panes, auto filter,
sheet protection, and print area.
"""
from __future__ import annotations

from ._core import (
    _RANGE_ADDRESS_RE,
    ValidationError,
    _audit_log,
    _check_confirm,
    _check_path,
    _check_range_size,
    _check_write,
    _flush_if_dirty,
    _load_wb,
    _parse_a1_range,
    _save_and_evict,
    _validate_formula_safe,
    openpyxl,
)


def merge_cells(
    path: str,
    sheet: str,
    address: str,
    confirm: bool = False,
) -> dict:
    """Merge a rectangular cell range on a sheet.

    Returns ``{"ok": True, "sheet": sheet, "address": address, "message": ...}``.
    Requires EXCEL_ENABLE_WRITE=true and confirm=True.
    """
    _check_write()
    _check_confirm(confirm)
    min_col, min_row, max_col, max_row = _parse_a1_range(address)
    n_cells = (max_row - min_row + 1) * (max_col - min_col + 1)
    _check_range_size(n_cells)
    resolved = _check_path(path)
    _flush_if_dirty(str(resolved))
    wb = _load_wb(resolved, for_write=True)
    if sheet not in wb.sheetnames:
        raise ValidationError(f"Sheet not found: {sheet!r}")
    ws = wb[sheet]
    try:
        ws.merge_cells(address)
    except (ValueError, TypeError) as exc:
        raise ValidationError(f"Invalid merge range {address!r}: {exc}") from exc
    _save_and_evict(wb, resolved)
    _audit_log("merge_cells", resolved, sheet, address, None)
    return {"ok": True, "sheet": sheet, "address": address, "message": f"Merged {address}"}


def unmerge_cells(
    path: str,
    sheet: str,
    address: str,
    confirm: bool = False,
) -> dict:
    """Unmerge a previously merged cell range on a sheet.

    Returns ``{"ok": True, "sheet": sheet, "address": address, "message": ...}``.
    Requires EXCEL_ENABLE_WRITE=true and confirm=True.
    """
    _check_write()
    _check_confirm(confirm)
    min_col, min_row, max_col, max_row = _parse_a1_range(address)
    n_cells = (max_row - min_row + 1) * (max_col - min_col + 1)
    _check_range_size(n_cells)
    resolved = _check_path(path)
    _flush_if_dirty(str(resolved))
    wb = _load_wb(resolved, for_write=True)
    if sheet not in wb.sheetnames:
        raise ValidationError(f"Sheet not found: {sheet!r}")
    ws = wb[sheet]
    try:
        ws.unmerge_cells(address)
    except (ValueError, TypeError) as exc:
        raise ValidationError(f"Invalid or unmerged range {address!r}: {exc}") from exc
    _save_and_evict(wb, resolved)
    _audit_log("unmerge_cells", resolved, sheet, address, None)
    return {"ok": True, "sheet": sheet, "address": address, "message": f"Unmerged {address}"}


def freeze_panes(
    path: str,
    sheet: str,
    freeze_at: str | None,
    confirm: bool = False,
) -> dict:
    """Freeze rows/columns above and to the left of freeze_at cell.

    Pass ``freeze_at=None`` to unfreeze.
    Returns ``{"ok": True, "sheet": sheet, "freeze_at": ..., "message": ...}``.
    Requires EXCEL_ENABLE_WRITE=true and confirm=True.
    """
    _check_write()
    _check_confirm(confirm)
    resolved = _check_path(path)
    _flush_if_dirty(str(resolved))
    wb = _load_wb(resolved, for_write=True)
    if sheet not in wb.sheetnames:
        raise ValidationError(f"Sheet not found: {sheet!r}")
    ws = wb[sheet]
    try:
        ws.freeze_panes = freeze_at  # openpyxl validates cell address here
    except (ValueError, TypeError, AttributeError) as exc:
        raise ValidationError(f"Invalid freeze cell {freeze_at!r}: {exc}") from exc
    _save_and_evict(wb, resolved)
    action = f"Frozen at {freeze_at}" if freeze_at else "Unfrozen"
    _audit_log("freeze_panes", resolved, sheet, freeze_at, None)
    return {"ok": True, "sheet": sheet, "freeze_at": freeze_at, "message": action}


def auto_filter(
    path: str,
    sheet: str,
    address: str | None,
    confirm: bool = False,
) -> dict:
    """Apply AutoFilter to the specified range, or clear all filters if address=None.

    When the sheet contains an Excel Table whose top-left cell matches the
    requested address, the worksheet-level AutoFilter is suppressed (set to
    None) instead of written — a coexisting worksheet AutoFilter at the same
    cell as a table's embedded AutoFilter causes an OOXML conflict that
    triggers Excel's recovery dialog.  In this case ``address`` in the return
    value is ``null`` and a ``warning`` key is included in the response.

    Returns ``{"ok": True, "sheet": sheet, "address": ..., "message": ...}``.
    Requires EXCEL_ENABLE_WRITE=true and confirm=True.
    """
    _check_write()
    _check_confirm(confirm)
    if address is not None:
        min_col, min_row, max_col, max_row = _parse_a1_range(address)
        n_cells = (max_row - min_row + 1) * (max_col - min_col + 1)
        _check_range_size(n_cells)
    resolved = _check_path(path)
    _flush_if_dirty(str(resolved))
    wb = _load_wb(resolved, for_write=True)
    if sheet not in wb.sheetnames:
        raise ValidationError(f"Sheet not found: {sheet!r}")
    ws = wb[sheet]
    warning: str | None = None
    if address is None:
        ws.auto_filter.ref = None
        msg = "AutoFilter cleared"
    else:
        # If an Excel Table is present whose top-left cell matches the requested
        # address, suppress the worksheet-level autoFilter entirely.  A coexisting
        # ws.auto_filter + table's embedded autoFilter at the same cell is an
        # OOXML conflict that triggers Excel's recovery dialog.
        effective_address = address
        for tbl in ws.tables.values():
            tbl_ref = tbl.ref or ""
            # Compare top-left cells only (e.g. both start at A1)
            try:
                req_min = openpyxl.utils.cell.range_boundaries(address.upper())
                tbl_min = openpyxl.utils.cell.range_boundaries(tbl_ref.upper())
            except Exception:
                continue
            if req_min[:2] == tbl_min[:2]:
                ws.auto_filter.ref = None
                effective_address = None
                warning = (
                    f"Table '{tbl.displayName}' already provides dropdown filters; "
                    f"worksheet-level AutoFilter suppressed to prevent OOXML conflict."
                )
                break
        if effective_address is not None:
            try:
                ws.auto_filter.ref = effective_address
            except (ValueError, TypeError) as exc:
                raise ValidationError(f"Invalid auto-filter range {address!r}: {exc}") from exc
        msg = (
            f"AutoFilter applied to {effective_address}"
            if effective_address
            else "AutoFilter suppressed (table conflict)"
        )
        if warning:
            msg = f"{msg}. WARNING: {warning}"
    _save_and_evict(wb, resolved)
    _audit_log("auto_filter", resolved, sheet, effective_address if address is not None else None, None)
    return {
        "ok": True,
        "sheet": sheet,
        "address": effective_address if address is not None else None,
        "message": msg,
        **({"warning": warning} if warning else {}),
    }


def protect_sheet(
    path: str,
    sheet: str,
    password: str | None = None,
    confirm: bool = False,
) -> dict:
    """Enable or clear worksheet protection.

    When password is provided, the sheet is protected with that password.
    When password is None (or empty string), any existing protection is cleared.

    Args:
        path: Path to the workbook file.
        sheet: Sheet name.
        password: Password string, or None to clear protection.
        confirm: Must be True to allow writes.

    Returns:
        {"ok": True, "sheet": ..., "protected": bool, "message": ...}
    """
    _check_write()
    _check_confirm(confirm)
    resolved = _check_path(path)
    _flush_if_dirty(str(resolved))
    wb = _load_wb(resolved, for_write=True)
    if sheet not in wb.sheetnames:
        raise ValidationError(f"Sheet not found: {sheet!r}")
    ws = wb[sheet]
    try:
        if password is None:
            ws.protection.sheet = False
            msg = f"Sheet '{sheet}' protection cleared."
        else:
            ws.protection.sheet = True
            ws.protection.set_password(password)
            msg = f"Sheet '{sheet}' is now protected."
    except (ValueError, TypeError, AttributeError) as exc:
        raise ValidationError(f"protect_sheet failed: {exc}") from exc
    _save_and_evict(wb, resolved)
    _audit_log("protect_sheet", str(resolved), sheet, None, None)
    return {"ok": True, "sheet": sheet, "protected": ws.protection.sheet, "has_password": password is not None, "message": msg}


def set_print_area(
    path: str,
    sheet: str,
    address: str | None,
    confirm: bool = False,
) -> dict:
    """Set or clear the print area for a worksheet.

    Args:
        path: Path to the workbook file.
        sheet: Sheet name.
        address: Cell range (e.g. 'A1:F20') or None to clear the print area.
        confirm: Must be True to allow writes.

    Returns:
        {"ok": True, "sheet": ..., "address": ..., "message": ...}
    """
    _check_write()
    _check_confirm(confirm)
    if address is not None:
        min_col, min_row, max_col, max_row = _parse_a1_range(address)
        n_cells = (max_row - min_row + 1) * (max_col - min_col + 1)
        _check_range_size(n_cells)
    resolved = _check_path(path)
    _flush_if_dirty(str(resolved))
    wb = _load_wb(resolved, for_write=True)
    if sheet not in wb.sheetnames:
        raise ValidationError(f"Sheet not found: {sheet!r}")
    if address is not None and not _RANGE_ADDRESS_RE.match(address):
        raise ValidationError(
            f"Invalid print area address {address!r}: must be a valid cell range like 'A1:F20'."
        )
    ws = wb[sheet]
    try:
        if address is None:
            ws.print_area = None
            msg = "Print area cleared."
        else:
            ws.print_area = address
            msg = f"Print area set to {address}."
    except (ValueError, TypeError) as exc:
        raise ValidationError(f"set_print_area failed: {exc}") from exc
    _save_and_evict(wb, resolved)
    _audit_log("set_print_area", str(resolved), sheet, address, None)
    return {"ok": True, "sheet": sheet, "address": address, "message": msg}


_VALID_DV_TYPES: frozenset[str] = frozenset(
    {"list", "whole", "decimal", "date", "time", "textLength", "custom"}
)


def add_data_validation(
    path: str,
    sheet: str,
    rules: "list[dict] | dict | str | None" = None,
    validation_type: str | None = None,  # 4th positional — legacy backward compat
    *,
    # Legacy keyword-only backward compat (used when rules=None or rules is str)
    address: str | None = None,
    formula1: str | None = None,
    formula2: str | None = None,
    operator: str | None = None,
    allow_blank: bool = True,
    show_error: bool = True,
    error_message: str | None = None,
    error_title: str | None = None,
    prompt: str | None = None,
    prompt_title: str | None = None,
    show_input_message: bool = True,
    confirm: bool = False,
) -> dict:
    """Add a data validation rule or multiple rules to a cell or range.

    Unified API (Sprint N Commit C): ``rules`` accepts:

    - A ``list[dict]`` for batch mode (returns ``{"validations_added": N}``).
    - A single ``dict`` for one rule (returns single-rule format).
    - A ``str`` address as the 3rd positional arg (legacy backward compat).
    - ``None`` + keyword args ``address`` + ``validation_type`` (legacy mode).

    Each rule dict keys: address, validation_type, formula1 (opt), formula2
    (opt), operator (opt), allow_blank (default True), show_error (default
    True), error_message (opt), error_title (opt), prompt (opt),
    prompt_title (opt), show_input_message (default True).

    validation_type must be one of: 'list', 'whole', 'decimal', 'date',
    'time', 'textLength', 'custom'.

    Gate order: EXCEL_ENABLE_WRITE → confirm=True → path allowlist.

    Returns:
    - Single rule (legacy or single-dict mode):
      ``{"ok": True, "sheet": ..., "address": ..., "validation_type": ...,
      "message": ...}``
    - List mode: ``{"ok": True, "sheet": ..., "validations_added": N}``

    Requires EXCEL_ENABLE_WRITE=true and confirm=True.
    """
    from openpyxl.worksheet.datavalidation import DataValidation as _DataValidation

    # Gate 1 & 2: write + confirm (before any I/O)
    _check_write()
    _check_confirm(confirm)

    # --- Normalise input to (rules_list: list[dict], _single_rule_mode: bool) ---
    if isinstance(rules, str):
        # Legacy: 3rd positional arg is the address string
        address = rules
        rules = None

    if rules is None:
        # Legacy keyword path: build single-rule dict from kwargs
        # Support both: add_data_validation(path, sheet, "addr", "type", ...) and
        #               add_data_validation(path, sheet, address="addr", validation_type="type", ...)
        eff_address = address
        eff_vtype = validation_type
        if eff_address is None or eff_vtype is None:
            raise ValidationError(
                "Either 'rules' (list[dict] or dict) or both 'address' and "
                "'validation_type' must be provided."
            )
        rules_list: list[dict] = [{
            "address": eff_address,
            "validation_type": eff_vtype,
            "formula1": formula1,
            "formula2": formula2,
            "operator": operator,
            "allow_blank": allow_blank,
            "show_error": show_error,
            "error_message": error_message,
            "error_title": error_title,
            "prompt": prompt,
            "prompt_title": prompt_title,
            "show_input_message": show_input_message,
        }]
        _single_rule_mode = True
    elif isinstance(rules, dict):
        rules_list = [rules]
        _single_rule_mode = True
    else:
        # list[dict] — always multi-rule return format
        rules_list = list(rules)
        _single_rule_mode = False

    # Gate 3: path allowlist (after lightweight normalisation)
    resolved = _check_path(path)

    if not rules_list:
        raise ValidationError("rules must be non-empty")
    if len(rules_list) > 500:
        raise ValidationError(
            f"too many rules — max 500 per call (got {len(rules_list)})"
        )

    # Per-rule validation (fail-fast before any writes)
    for i, item in enumerate(rules_list):
        addr = item.get("address", "")
        if not addr or not isinstance(addr, str) or not addr.strip():
            raise ValidationError(
                f"rules[{i}]: address must be a non-empty string"
            )
        vtype = item.get("validation_type", "")
        if vtype not in _VALID_DV_TYPES:
            raise ValidationError(
                f"rules[{i}]: invalid validation_type {vtype!r}: "
                f"must be one of {sorted(_VALID_DV_TYPES)}"
            )
        f1 = item.get("formula1")
        f2 = item.get("formula2")
        if f1 and isinstance(f1, str) and f1.startswith("="):
            _validate_formula_safe(f1)
        if f2 and isinstance(f2, str) and f2.startswith("="):
            _validate_formula_safe(f2)
        if vtype == "list" and not f1:
            raise ValidationError(
                f"rules[{i}]: formula1 is required for validation_type='list' "
                "(e.g. '\"A,B,C\"' or a range like '$A$1:$A$5')"
            )

    # Load workbook
    _flush_if_dirty(str(resolved))
    wb = _load_wb(resolved, for_write=True)
    if sheet not in wb.sheetnames:
        raise ValidationError(f"Sheet not found: {sheet!r}")
    ws = wb[sheet]

    # Apply rules
    for item in rules_list:
        addr = item["address"].upper()
        vtype = item["validation_type"]
        f1 = item.get("formula1")
        f2 = item.get("formula2")
        op = item.get("operator")
        blank = item.get("allow_blank", True)
        showerr = item.get("show_error", True)
        errmsg = item.get("error_message")
        errtitle = item.get("error_title")
        prpt = item.get("prompt")
        prpt_title = item.get("prompt_title")
        show_input = item.get("show_input_message", True)
        dv = _DataValidation(
            type=vtype,
            formula1=f1,
            formula2=f2,
            operator=op,
            allow_blank=blank,
            showErrorMessage=showerr,
        )
        if errmsg is not None:
            dv.error = errmsg
        if errtitle is not None:
            dv.errorTitle = errtitle
        if prpt is not None:
            dv.prompt = prpt
        if prpt_title is not None:
            dv.promptTitle = prpt_title
        dv.showInputMessage = show_input
        ws.add_data_validation(dv)
        dv.sqref = addr

    _save_and_evict(wb, resolved)

    if _single_rule_mode:
        first = rules_list[0]
        first_addr = first["address"].upper()
        first_vtype = first["validation_type"]
        _audit_log("add_data_validation", str(resolved), sheet, first_addr, None)
        return {
            "ok": True,
            "sheet": sheet,
            "address": first_addr,
            "validation_type": first_vtype,
            "message": f"Data validation ({first_vtype}) added to {first_addr}",
        }
    else:
        _audit_log("add_data_validation", str(resolved), sheet, None, len(rules_list))
        return {
            "ok": True,
            "sheet": sheet,
            "validations_added": len(rules_list),
        }


# ---------------------------------------------------------------------------
# Phase 2.6 — read data validation rules
# ---------------------------------------------------------------------------


def get_data_validation_info(path: str, sheet: str) -> dict:
    """Return all data validation rules configured on a sheet.

    Returns a dict with:
    - ``sheet``: worksheet name
    - ``count``: number of rules found
    - ``validations``: list of rule dicts, each with keys:
      address, validation_type, formula1, formula2, allow_blank,
      show_dropdown, error_title, error_message.
    """
    resolved = _check_path(path)
    wb = _load_wb(resolved)
    if sheet not in wb.sheetnames:
        raise ValidationError(f"Sheet not found: {sheet!r}")
    ws = wb[sheet]
    rules = []
    try:
        dv_list = list(ws.data_validations.dataValidation)
    except AttributeError:
        dv_list = []
    for dv in dv_list:
        try:
            address = str(dv.sqref) if dv.sqref is not None else ""
        except Exception:
            address = ""
        rules.append({
            "address": address,
            "validation_type": getattr(dv, "type", None),
            "formula1": getattr(dv, "formula1", None),
            "formula2": getattr(dv, "formula2", None),
            "allow_blank": bool(getattr(dv, "allow_blank", True)),
            "show_dropdown": not bool(getattr(dv, "showDropDown", False)),
            "error_title": getattr(dv, "errorTitle", None),
            "error_message": getattr(dv, "error", None),
        })
    return {"sheet": sheet, "count": len(rules), "validations": rules}


# ---------------------------------------------------------------------------
# Sprint N Commit C: bulk_add_data_validation — thin alias
# ---------------------------------------------------------------------------


def bulk_add_data_validation(
    path: str,
    sheet: str,
    validations: list[dict],
    confirm: bool = False,
) -> dict:
    """Thin alias → ``add_data_validation`` with list rules (Sprint N Commit C).

    Each dict: {"address": str, "validation_type": str, "formula1": str|None,
    "formula2": str|None, "operator": str|None, "allow_blank": bool,
    "show_error": bool, "error_message": str|None, "error_title": str|None,
    "prompt": str|None, "prompt_title": str|None,
    "show_input_message": bool (default True)}.
    validation_type must be in {"whole","decimal","list","date","time","textLength","custom"}.
    Returns {"ok": True, "sheet": str, "validations_added": int}
    Requires EXCEL_ENABLE_WRITE=true and confirm=True.
    """
    return add_data_validation(path, sheet, validations, confirm=confirm)


# ---------------------------------------------------------------------------
# Sprint D Batch 3 — T-B: apply_conditional_format
# ---------------------------------------------------------------------------

_CF_VALID_RULE_TYPES: frozenset[str] = frozenset({"cell_is", "formula"})
_CF_VALID_FORMAT_KEYS: frozenset[str] = frozenset(
    {"bold", "italic", "font_color", "bg_color", "number_format"}
)
_CF_VALID_OPERATORS: frozenset[str] = frozenset(
    {
        "greaterThan", "greaterThanOrEqual", "lessThan", "lessThanOrEqual",
        "equal", "notEqual", "between", "notBetween",
    }
)


def apply_conditional_format(
    path: str,
    sheet: str,
    address: str,
    rule_type: str,
    condition: str,
    format_spec: dict,
    formula: str | None = None,
    confirm: bool = False,
) -> dict:
    """Apply a conditional-formatting rule to a range.

    rule_type: "cell_is" or "formula".
    condition: For cell_is — "<operator> <value>" e.g. "greaterThan 100",
               "between 10 20". Unused when rule_type="formula".
    format_spec: keys: bold, italic, font_color, bg_color, number_format
                 (number_format is silently ignored — not supported by CF rules API).
    formula: Required when rule_type="formula". Must start with "=".
    Requires EXCEL_ENABLE_WRITE=true and confirm=True.
    Returns ``{"status": "ok", "sheet": sheet, "address": address, "rule_type": rule_type}``.
    """
    _check_write()
    _check_confirm(confirm)

    if rule_type not in _CF_VALID_RULE_TYPES:
        raise ValidationError(
            f"Invalid rule_type {rule_type!r}: must be one of {sorted(_CF_VALID_RULE_TYPES)}"
        )

    unknown_fmt_keys = set(format_spec) - _CF_VALID_FORMAT_KEYS
    if unknown_fmt_keys:
        raise ValidationError(
            f"Unknown format_spec key(s): {sorted(unknown_fmt_keys)}. "
            f"Allowed: {sorted(_CF_VALID_FORMAT_KEYS)}"
        )

    if rule_type == "formula":
        if not formula:
            raise ValidationError("formula is required when rule_type='formula'")
        _validate_formula_safe(formula)

    if not address or not address.strip():
        raise ValidationError("address must be non-empty")

    min_col, min_row, max_col, max_row = _parse_a1_range(address)
    n_cells = (max_row - min_row + 1) * (max_col - min_col + 1)
    _check_range_size(n_cells)

    # Local imports after all pre-load validation
    from openpyxl.formatting.rule import CellIsRule as _CellIsRule
    from openpyxl.formatting.rule import FormulaRule as _FormulaRule
    from openpyxl.styles import Font as _Font
    from openpyxl.styles import PatternFill as _PatternFill

    # Build font / fill from format_spec
    font_kwargs: dict = {}
    if "bold" in format_spec:
        font_kwargs["bold"] = format_spec["bold"]
    if "italic" in format_spec:
        font_kwargs["italic"] = format_spec["italic"]
    if "font_color" in format_spec:
        font_kwargs["color"] = format_spec["font_color"].lstrip("#")
    cf_font = _Font(**font_kwargs) if font_kwargs else None

    cf_fill = None
    if "bg_color" in format_spec:
        cf_fill = _PatternFill(
            fill_type="solid", fgColor=format_spec["bg_color"].lstrip("#")
        )

    if rule_type == "cell_is":
        parts = condition.split()
        if not parts:
            raise ValidationError("condition is empty for rule_type='cell_is'")
        operator = parts[0]
        if operator not in _CF_VALID_OPERATORS:
            raise ValidationError(
                f"Unknown operator {operator!r}. "
                f"Valid: {sorted(_CF_VALID_OPERATORS)}"
            )
        if operator in {"between", "notBetween"}:
            if len(parts) < 3:
                raise ValidationError(
                    f"condition for operator {operator!r} must be "
                    f"'{operator} <val1> <val2>'"
                )
            formula_args = [parts[1], parts[2]]
        else:
            if len(parts) < 2:
                raise ValidationError(
                    f"condition for operator {operator!r} must be '{operator} <value>'"
                )
            formula_args = [parts[1]]

        rule_kwargs: dict = {"operator": operator, "formula": formula_args}
        if cf_font is not None:
            rule_kwargs["font"] = cf_font
        if cf_fill is not None:
            rule_kwargs["fill"] = cf_fill
        rule = _CellIsRule(**rule_kwargs)
    else:
        # formula rule
        rule_kwargs = {"formula": [formula]}
        if cf_font is not None:
            rule_kwargs["font"] = cf_font
        if cf_fill is not None:
            rule_kwargs["fill"] = cf_fill
        rule = _FormulaRule(**rule_kwargs)

    resolved = _check_path(path)
    _flush_if_dirty(str(resolved))
    wb = _load_wb(resolved, for_write=True)
    if sheet not in wb.sheetnames:
        raise ValidationError(f"Sheet not found: {sheet!r}")
    ws = wb[sheet]
    ws.conditional_formatting.add(address, rule)
    _save_and_evict(wb, resolved)
    _audit_log("apply_conditional_format", str(resolved), sheet, address, None)
    return {"status": "ok", "sheet": sheet, "address": address, "rule_type": rule_type}
