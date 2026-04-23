"""Bulk and sheet-level formatting: multi-range style, cross-sheet format, format copy.

Internal sub-module. Import via excelmcp._format (public router).
Depends on _format_cell for cache helpers (_cached_font, _cached_alignment, etc.).
"""
from __future__ import annotations

from copy import copy as _copy

from openpyxl.cell.cell import MergedCell as _MergedCell
from openpyxl.utils import (
    column_index_from_string as _col_idx,
)
from openpyxl.utils import (
    get_column_letter as _col_letter,
)

from ._core import (
    ValidationError,
    _audit_log,
    _check_confirm,
    _check_path,
    _check_range_size,
    _check_write,
    _flush_if_dirty,
    _load_wb,
    _parse_a1_range,
    _save_and_evict,
)
from ._format_cell import (
    _cached_alignment,
    _cached_fill,
    _cached_font,
    _validate_hex_color,
)

__all__ = [
    "apply_style_to_ranges",
    "apply_format_to_sheet_list",
    "copy_range_format",
]


def apply_style_to_ranges(
    path: str,
    sheet: str,
    ranges: list[str],
    bold: bool | None = None,
    italic: bool | None = None,
    font_size: float | None = None,
    font_color: str | None = None,
    bg_color: str | None = None,
    confirm: bool = False,
) -> dict:
    """Apply font and fill styling to multiple ranges in a single call.

    All ranges are validated and their total cell count is checked against
    ``EXCEL_MAX_RANGE_CELLS`` *before* any write occurs (fail-fast atomicity).

    Returns
    ``{"ok": True, "sheet": sheet, "ranges_processed": int, "cells_styled": int}``.
    Requires EXCEL_ENABLE_WRITE=true and confirm=True.
    """
    _check_write()
    _check_confirm(confirm)
    if not ranges:
        raise ValidationError("ranges must be non-empty")
    if all(v is None for v in (bold, italic, font_size, font_color, bg_color)):
        raise ValidationError("at least one style parameter must be provided")
    _validate_hex_color(font_color, "font_color")
    _validate_hex_color(bg_color, "bg_color")

    # Pre-validate all ranges and count total cells before touching workbook
    total_cells = 0
    for r in ranges:
        try:
            min_col, min_row, max_col, max_row = _parse_a1_range(r)
        except ValidationError:
            raise ValidationError(f"Invalid range notation: {r!r}")
        total_cells += (max_row - min_row + 1) * (max_col - min_col + 1)
    _check_range_size(total_cells)

    resolved = _check_path(path)
    _flush_if_dirty(str(resolved))
    wb_obj = _load_wb(resolved, for_write=True)
    if sheet not in wb_obj.sheetnames:
        raise ValidationError(f"Sheet not found: {sheet!r}")
    ws = wb_obj[sheet]

    _fill = _cached_fill(bg_color) if bg_color is not None else None
    _font = (
        _cached_font(bold=bold, italic=italic, size=font_size, color=font_color)
        if all(v is not None for v in (bold, italic, font_size, font_color))
        else None
    )
    cells_styled = 0
    for addr_str in ranges:
        addr = addr_str.upper().strip()
        if ":" not in addr:
            addr = f"{addr}:{addr}"
        for row in ws[addr]:
            for cell in row:
                if any(v is not None for v in (bold, italic, font_size, font_color)):
                    if _font is not None:
                        cell.font = _font
                    else:
                        current = cell.font
                        existing_color = current.color if current.color is not None else None
                        cell.font = _cached_font(
                            bold=bold if bold is not None else current.bold,
                            italic=italic if italic is not None else current.italic,
                            size=font_size if font_size is not None else current.size,
                            color=font_color if font_color is not None else existing_color,
                        )
                if _fill is not None:
                    cell.fill = _fill
                cells_styled += 1

    _save_and_evict(wb_obj, resolved)
    _audit_log("apply_style_to_ranges", resolved, sheet, ",".join(ranges[:3]), cells_styled)
    return {
        "ok": True,
        "sheet": sheet,
        "ranges_processed": len(ranges),
        "cells_styled": cells_styled,
    }


def apply_format_to_sheet_list(
    path: str,
    sheets: list[str] | None = None,
    address: str | None = None,
    style: dict | None = None,
    number_format: str | None = None,
    alignment: dict | None = None,
    column_widths: dict | None = None,
    row_heights: dict | None = None,
    confirm: bool = False,
) -> dict:
    """Apply formatting to the same address across multiple sheets.

    If ``sheets`` is ``None``, ALL sheets in the workbook are targeted EXCEPT
    those whose names start with ``_`` (e.g. ``_Lists``, ``_Config``).
    Passing an explicit list of names (including ``_``-prefixed ones) overrides
    the skip guard — exact name matching takes precedence.

    ``address`` is required when any of ``style``, ``number_format``, or
    ``alignment`` are provided.

    ``style`` dict allowed keys: bold, italic, font_size, font_color, bg_color.
    ``alignment`` dict allowed keys: horizontal, vertical, wrap_text.
    Unknown keys raise ValidationError.

    ``EXCEL_MAX_RANGE_CELLS`` is enforced *per sheet*.

    Returns ``{"ok": True, "sheets_processed": int, "cells_formatted": int}``.
    Requires EXCEL_ENABLE_WRITE=true and confirm=True.
    """
    _check_write()
    _check_confirm(confirm)

    _STYLE_ALLOWED = {"bold", "italic", "font_size", "font_color", "bg_color"}
    _ALIGN_ALLOWED = {"horizontal", "vertical", "wrap_text"}

    # Validate style dict keys
    if style is not None:
        unknown = set(style) - _STYLE_ALLOWED
        if unknown:
            raise ValidationError(
                f"Unknown style key(s): {sorted(unknown)}. "
                f"Allowed: {sorted(_STYLE_ALLOWED)}"
            )
    # Validate alignment dict keys
    if alignment is not None:
        unknown = set(alignment) - _ALIGN_ALLOWED
        if unknown:
            raise ValidationError(
                f"Unknown alignment key(s): {sorted(unknown)}. "
                f"Allowed: {sorted(_ALIGN_ALLOWED)}"
            )

    has_format = style is not None or number_format is not None or alignment is not None
    if has_format and address is None:
        raise ValidationError(
            "address is required when style, number_format, or alignment is provided"
        )

    resolved = _check_path(path)
    _flush_if_dirty(str(resolved))
    wb_obj = _load_wb(resolved, for_write=True)

    # Resolve target sheets
    if sheets is not None:
        target_sheets = sheets
    else:
        target_sheets = [s for s in wb_obj.sheetnames if not s.startswith("_")]

    if not target_sheets:
        return {"ok": True, "sheets_processed": 0, "cells_formatted": 0}

    # Validate all requested sheets exist
    for sn in target_sheets:
        if sn not in wb_obj.sheetnames:
            raise ValidationError(f"Sheet not found: {sn!r}")

    # Pre-build style objects
    bold = style.get("bold") if style else None
    italic = style.get("italic") if style else None
    font_size = style.get("font_size") if style else None
    font_color = style.get("font_color") if style else None
    bg_color = style.get("bg_color") if style else None

    if font_color is not None:
        _validate_hex_color(font_color, "font_color")
    if bg_color is not None:
        _validate_hex_color(bg_color, "bg_color")

    _fill = _cached_fill(bg_color) if bg_color is not None else None
    _font_obj = (
        _cached_font(bold=bold, italic=italic, size=font_size, color=font_color)
        if style is not None and all(
            v is not None for v in (bold, italic, font_size, font_color)
        )
        else None
    )

    horiz = alignment.get("horizontal") if alignment else None
    vert = alignment.get("vertical") if alignment else None
    wrap = alignment.get("wrap_text") if alignment else None
    _align_obj = (
        _cached_alignment(horizontal=horiz, vertical=vert, wrap_text=wrap)
        if alignment is not None and all(
            v is not None for v in (horiz, vert, wrap)
        )
        else None
    )

    sheets_processed = 0
    cells_formatted = 0
    col_widths_set = 0
    row_heights_set = 0

    for sn in target_sheets:
        ws = wb_obj[sn]

        if has_format:
            addr = address.upper().strip()  # type: ignore[union-attr]
            if ":" not in addr:
                addr_range = f"{addr}:{addr}"
            else:
                addr_range = addr

            # Compute cell count for per-sheet MAX_RANGE_CELLS guard
            try:
                min_col, min_row, max_col, max_row = _parse_a1_range(addr_range)
            except ValidationError:
                raise ValidationError(f"Invalid address {address!r}")
            n_cells = (max_row - min_row + 1) * (max_col - min_col + 1)
            _check_range_size(n_cells)

            for row in ws[addr_range]:
                for cell in row:
                    if style is not None:
                        if any(v is not None for v in (bold, italic, font_size, font_color)):
                            if _font_obj is not None:
                                cell.font = _font_obj
                            else:
                                current = cell.font
                                existing_color = (
                                    current.color if current.color is not None else None
                                )
                                cell.font = _cached_font(
                                    bold=bold if bold is not None else current.bold,
                                    italic=italic if italic is not None else current.italic,
                                    size=font_size if font_size is not None else current.size,
                                    color=(
                                        font_color
                                        if font_color is not None
                                        else existing_color
                                    ),
                                )
                        if _fill is not None:
                            cell.fill = _fill
                    if number_format is not None:
                        cell.number_format = number_format
                    if alignment is not None:
                        if _align_obj is not None:
                            cell.alignment = _align_obj
                        else:
                            current_align = cell.alignment
                            cell.alignment = _cached_alignment(
                                horizontal=(
                                    horiz if horiz is not None else current_align.horizontal
                                ),
                                vertical=(
                                    vert if vert is not None else current_align.vertical
                                ),
                                wrap_text=(
                                    wrap if wrap is not None else current_align.wrap_text
                                ),
                            )
                    cells_formatted += 1

        if column_widths:
            for key, val in column_widths.items():
                key_up = key.upper()
                if ":" in key_up:
                    parts = key_up.split(":")
                    start_idx = _col_idx(parts[0].strip())
                    end_idx = _col_idx(parts[1].strip())
                    for idx in range(start_idx, end_idx + 1):
                        ws.column_dimensions[_col_letter(idx)].width = val
                        col_widths_set += 1
                else:
                    ws.column_dimensions[key_up].width = val
                    col_widths_set += 1

        if row_heights:
            for key, val in row_heights.items():
                key_str = str(key).strip()
                if ":" in key_str:
                    parts = key_str.split(":")
                    start_r = int(parts[0].strip())
                    end_r = int(parts[1].strip())
                    for r in range(start_r, end_r + 1):
                        ws.row_dimensions[r].height = val
                        row_heights_set += 1
                else:
                    ws.row_dimensions[int(key_str)].height = val
                    row_heights_set += 1

        sheets_processed += 1

    _save_and_evict(wb_obj, resolved)
    _audit_log(
        "apply_format_to_sheet_list", resolved, ",".join(target_sheets[:3]),
        address, cells_formatted,
    )
    return {
        "ok": True,
        "sheets_processed": sheets_processed,
        "cells_formatted": cells_formatted,
        "column_widths_set": col_widths_set,
        "row_heights_set": row_heights_set,
    }


def copy_range_format(
    path: str,
    src_sheet: str,
    src_address: str,
    dst_sheet: str,
    dst_address: str,
    copy_border: bool = False,
    confirm: bool = False,
) -> dict:
    """Copy ONLY formatting (font, fill, number_format, alignment, optionally border)
    from src to dst. Values are NOT modified. Cross-sheet supported (same workbook only).
    dst_address: if single cell, auto-expand to src dimensions.
    Does NOT copy: values, formulas, comments, hyperlinks, column widths, row heights.
    MergedCell slave cells in src are skipped.
    Returns {"ok": True, "src": {"sheet": str, "address": str},
             "dst": {"sheet": str, "address": str}, "cells_copied": int}
    Requires EXCEL_ENABLE_WRITE=true and confirm=True.
    """
    _check_write()
    _check_confirm(confirm)

    resolved = _check_path(path)
    _flush_if_dirty(str(resolved))
    wb_obj = _load_wb(resolved, for_write=True)

    if src_sheet not in wb_obj.sheetnames:
        raise ValidationError(f"Source sheet not found: {src_sheet!r}")
    if dst_sheet not in wb_obj.sheetnames:
        raise ValidationError(f"Destination sheet not found: {dst_sheet!r}")

    src_min_col, src_min_row, src_max_col, src_max_row = _parse_a1_range(src_address)
    n_src_cells = (src_max_row - src_min_row + 1) * (src_max_col - src_min_col + 1)
    _check_range_size(n_src_cells)

    src_rows = src_max_row - src_min_row + 1
    src_cols = src_max_col - src_min_col + 1

    # Determine effective dst address
    dst_addr_upper = dst_address.upper().strip()
    if ":" not in dst_addr_upper:
        # Single cell — auto-expand to match src dimensions
        dst_min_col, dst_min_row, _, _ = _parse_a1_range(dst_addr_upper)
        dst_max_row_val = dst_min_row + src_rows - 1
        dst_max_col_val = dst_min_col + src_cols - 1
        dst_effective = (
            f"{_col_letter(dst_min_col)}{dst_min_row}"
            f":{_col_letter(dst_max_col_val)}{dst_max_row_val}"
        )
    else:
        dst_min_col, dst_min_row, dst_max_col, dst_max_row = _parse_a1_range(dst_addr_upper)
        dst_rows = dst_max_row - dst_min_row + 1
        dst_cols = dst_max_col - dst_min_col + 1
        if dst_rows != src_rows or dst_cols != src_cols:
            raise ValidationError(
                f"dst_address dimensions ({dst_rows}×{dst_cols}) do not match "
                f"src_address dimensions ({src_rows}×{src_cols})"
            )
        dst_effective = dst_addr_upper

    ws_src = wb_obj[src_sheet]
    ws_dst = wb_obj[dst_sheet]

    cells_copied = 0
    for r_off in range(src_rows):
        for c_off in range(src_cols):
            src_cell = ws_src.cell(
                row=src_min_row + r_off, column=src_min_col + c_off
            )
            if isinstance(src_cell, _MergedCell):
                continue
            dst_cell = ws_dst.cell(
                row=dst_min_row + r_off, column=dst_min_col + c_off
            )
            try:
                dst_cell.font = _copy(src_cell.font)
            except Exception:
                pass
            try:
                dst_cell.fill = _copy(src_cell.fill)
            except Exception:
                pass
            try:
                dst_cell.number_format = src_cell.number_format
            except Exception:
                pass
            try:
                dst_cell.alignment = _copy(src_cell.alignment)
            except Exception:
                pass
            if copy_border:
                try:
                    dst_cell.border = _copy(src_cell.border)
                except Exception:
                    pass
            cells_copied += 1

    _save_and_evict(wb_obj, resolved)
    _audit_log(
        "copy_range_format", resolved, src_sheet,
        f"{src_address}->{dst_effective}", cells_copied,
    )
    return {
        "ok": True,
        "src": {"sheet": src_sheet, "address": src_address.upper()},
        "dst": {"sheet": dst_sheet, "address": dst_effective},
        "cells_copied": cells_copied,
    }
