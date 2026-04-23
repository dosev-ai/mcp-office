"""Snapshot and diff tools for excelmcp gap sprint (US-E-032).

Tools registered on shared mcp instance at import time (side-effect import).
Do not instantiate FastMCP here.

Security:
    snapshot_workbook: _check_write() + _check_confirm(confirm)
                       + _check_path(source) + _check_dest_dir(dest_dir, source_resolved)
                       + shutil.copy2 (no openpyxl; no formula injection risk)
    diff_workbooks:    _check_path(path_a) + _check_path(path_b)
                       + _check_range_size() per sheet pair + openpyxl.load_workbook(data_only=True)

Env vars consumed:
    EXCEL_ENABLE_WRITE    -- gate for snapshot_workbook
    EXCEL_ALLOWLIST_ROOTS -- both source and destination must resolve within these roots
    EXCEL_MAX_RANGE_CELLS -- per-sheet cell count guard for diff_workbooks
"""
from __future__ import annotations

import logging
import os
import shutil
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import openpyxl
from fastmcp.exceptions import ToolError

from excelmcp._core import (
    NotAllowedError,
    ValidationError,
    _check_confirm,
    _check_path,
    _check_range_size,
    _check_write,
)
from excelmcp._server_instance import mcp

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

def _check_dest_dir(dest_dir: str | None, source_resolved: Path) -> Path:
    """Validate dest_dir against EXCEL_ALLOWLIST_ROOTS.

    If dest_dir is None, returns source_resolved.parent (always valid since
    source already passed _check_path — its parent is within an allowed root).

    Raises ValidationError if:
      - dest_dir contains a null byte
      - dest_dir is a UNC path
      - EXCEL_ALLOWLIST_ROOTS is not configured
      - dest_dir resolves outside all configured allowlist roots
    """
    if dest_dir is None:
        return source_resolved.parent

    if "\x00" in str(dest_dir):
        raise ValidationError("dest_dir contains illegal null byte")

    s = str(dest_dir)
    if s.startswith("\\\\") or s.startswith("//"):
        raise ValidationError("UNC paths are not permitted for dest_dir")

    roots_env = os.getenv("EXCEL_ALLOWLIST_ROOTS", "").strip()
    if not roots_env:
        raise ValidationError("No allowlist configured — set EXCEL_ALLOWLIST_ROOTS")

    resolved_dir = Path(dest_dir).resolve()

    roots = [r.strip() for r in roots_env.split(",") if r.strip()]
    for root in roots:
        try:
            if resolved_dir.is_relative_to(Path(root).resolve()):
                return resolved_dir
        except (ValueError, TypeError):
            continue

    raise ValidationError(
        f"dest_dir is not within any configured EXCEL_ALLOWLIST_ROOTS: {resolved_dir}"
    )


# ---------------------------------------------------------------------------
# Tools
# ---------------------------------------------------------------------------

@mcp.tool()
def snapshot_workbook(
    path: str,
    dest_dir: str | None = None,
    filename: str | None = None,
    confirm: bool = False,
) -> dict[str, Any]:
    """Copy a workbook to a timestamped snapshot file.

    Creates a point-in-time copy of the workbook without opening it in Excel
    or openpyxl (shutil.copy2 preserves metadata). The snapshot filename
    defaults to <original_stem>_<YYYYMMDD_HHMMSS><original_suffix>.

    Args:
        path:     Absolute path to source .xlsx/.xlsm workbook.
                  Must be within EXCEL_ALLOWLIST_ROOTS.
        dest_dir: Directory to write the snapshot into.
                  Must be within EXCEL_ALLOWLIST_ROOTS.
                  If None, uses source file's parent directory.
        filename: Override the auto-generated snapshot filename.
                  If provided, must end in .xlsx or .xlsm.
                  If None, auto-generates: <stem>_<YYYYMMDD_HHMMSS><suffix>.
        confirm:  Must be True (write gate).

    Returns:
        {"ok": True, "source": str, "snapshot": str, "size_bytes": int}

    Requires: EXCEL_ENABLE_WRITE=true, confirm=True.
    """
    try:
        # Gate checks first (early rejection before path validation)
        _check_write()
        _check_confirm(confirm)

        # Validate source path against allowlist
        source_resolved = _check_path(path)

        # Validate destination directory against allowlist
        dest_dir_resolved = _check_dest_dir(dest_dir, source_resolved)

        # Compute destination filename
        if filename is not None:
            dest_filename = filename
            dest_suffix = Path(filename).suffix.lower()
            if dest_suffix not in (".xlsx", ".xlsm"):
                raise ValidationError(
                    f"filename must end in .xlsx or .xlsm, got: {dest_suffix!r}"
                )
        else:
            ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
            dest_filename = f"{source_resolved.stem}_{ts}{source_resolved.suffix}"

        # Build full destination path
        dest_path = dest_dir_resolved / dest_filename
        dest_path = dest_path.resolve()
        if not dest_path.is_relative_to(dest_dir_resolved):
            raise ValidationError(
                f"filename must not contain path traversal sequences: {dest_filename!r}"
            )

        # Guard against overwriting an existing file
        if dest_path.exists():
            raise ValidationError(
                f"Snapshot destination already exists: {dest_path}. "
                "Use a unique filename or wait 1 second."
            )

        # Ensure destination directory exists
        dest_dir_resolved.mkdir(parents=True, exist_ok=True)

        # Copy (shutil.copy2 preserves timestamps and metadata)
        shutil.copy2(str(source_resolved), str(dest_path))

        size = dest_path.stat().st_size
        logger.info(
            "[excelmcp] snapshot_workbook: %s -> %s (%d bytes)",
            source_resolved, dest_path, size,
        )
        return {
            "ok": True,
            "source": str(source_resolved),
            "snapshot": str(dest_path),
            "size_bytes": size,
        }

    except (NotAllowedError, ValidationError):
        raise
    except Exception as exc:
        logger.error(
            "[excelmcp] snapshot_workbook unexpected error: %s", exc, exc_info=True
        )
        raise ToolError("snapshot_workbook failed — check server logs for details.") from exc


@mcp.tool()
def diff_workbooks(
    path_a: str,
    path_b: str,
    sheets: list[str] | None = None,
    max_diffs: int = 200,
) -> dict[str, Any]:
    """Compare two workbooks cell-by-cell and return differences.

    Loads both workbooks with openpyxl (data_only=True) and compares the
    values of matching sheets. Only sheets present in BOTH workbooks are
    compared (unless sheets is specified).

    Each difference entry has the form:
        {"sheet": str, "row": int, "col": int, "val_a": Any, "val_b": Any}

    Args:
        path_a:    Absolute path to first workbook (baseline).
                   Must be within EXCEL_ALLOWLIST_ROOTS.
        path_b:    Absolute path to second workbook (comparison).
                   Must be within EXCEL_ALLOWLIST_ROOTS.
        sheets:    Optional list of sheet names to compare.
                   If None, compares sheets present in both workbooks.
        max_diffs: Maximum total differences to return (default 200).
                   Truncates result set with truncated=True flag.

    Returns:
        {
            "ok": True,
            "path_a": str,
            "path_b": str,
            "sheets_compared": [str, ...],
            "sheets_only_in_a": [str, ...],
            "sheets_only_in_b": [str, ...],
            "total_diffs": int,
            "truncated": bool,
            "diffs": [{"sheet": str, "row": int, "col": int,
                       "val_a": Any, "val_b": Any}, ...]
        }

    Does NOT require EXCEL_ENABLE_WRITE (read-only operation).
    Does NOT require confirm (non-mutating).
    """
    try:
        # Validate both source paths against allowlist
        resolved_a = _check_path(path_a)
        resolved_b = _check_path(path_b)

        wb_a = wb_b = None
        try:
            # Load both workbooks read-only with data_only=True
            wb_a = openpyxl.load_workbook(str(resolved_a), data_only=True, read_only=True)
            wb_b = openpyxl.load_workbook(str(resolved_b), data_only=True, read_only=True)

            # Determine sheets to compare
            sheets_a = set(wb_a.sheetnames)
            sheets_b = set(wb_b.sheetnames)

            if sheets is not None:
                unknown = set(sheets) - (sheets_a | sheets_b)
                if unknown:
                    raise ValidationError(
                        f"These sheet names are not in either workbook: {sorted(unknown)}"
                    )
                compare_sheets = [s for s in sheets if s in sheets_a and s in sheets_b]
            else:
                compare_sheets = sorted(sheets_a & sheets_b)

            sheets_only_in_a = sorted(sheets_a - sheets_b)
            sheets_only_in_b = sorted(sheets_b - sheets_a)

            # Compare each sheet pair
            all_diffs: list[dict] = []
            truncated = False

            for sheet_name in compare_sheets:
                ws_a = wb_a[sheet_name]
                ws_b = wb_b[sheet_name]

                # Determine bounding box (union of both sheets' used range)
                max_row = max(ws_a.max_row or 0, ws_b.max_row or 0)
                max_col = max(ws_a.max_column or 0, ws_b.max_column or 0)

                if max_row == 0 or max_col == 0:
                    continue  # both sheets are empty — skip

                # Cell count guard before iterating (W-2)
                n_cells = max_row * max_col
                _check_range_size(n_cells)

                # Read data as value dicts for efficiency with read_only workbooks
                data_a: dict[tuple[int, int], Any] = {}
                data_b: dict[tuple[int, int], Any] = {}

                for ri, row_cells in enumerate(
                    ws_a.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col), start=1
                ):
                    for ci, c in enumerate(row_cells, start=1):
                        data_a[(ri, ci)] = c.value

                for ri, row_cells in enumerate(
                    ws_b.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col), start=1
                ):
                    for ci, c in enumerate(row_cells, start=1):
                        data_b[(ri, ci)] = c.value

                # Find differences
                all_coords = set(data_a.keys()) | set(data_b.keys())
                for (r, c) in sorted(all_coords):
                    val_a = data_a.get((r, c))
                    val_b = data_b.get((r, c))
                    if val_a != val_b:
                        all_diffs.append({
                            "sheet": sheet_name,
                            "row": r,
                            "col": c,
                            "val_a": val_a,
                            "val_b": val_b,
                        })
                        if len(all_diffs) >= max_diffs:
                            truncated = True
                            break
                if truncated:
                    break

            return {
                "ok": True,
                "path_a": str(resolved_a),
                "path_b": str(resolved_b),
                "sheets_compared": compare_sheets,
                "sheets_only_in_a": sheets_only_in_a,
                "sheets_only_in_b": sheets_only_in_b,
                "total_diffs": len(all_diffs),
                "truncated": truncated,
                "diffs": all_diffs,
            }
        finally:
            if wb_a is not None:
                try:
                    wb_a.close()
                except Exception:
                    pass
            if wb_b is not None:
                try:
                    wb_b.close()
                except Exception:
                    pass

    except (NotAllowedError, ValidationError):
        raise
    except Exception as exc:
        logger.error(
            "[excelmcp] diff_workbooks unexpected error: %s", exc, exc_info=True
        )
        raise ToolError("diff_workbooks failed — check server logs for details.") from exc
