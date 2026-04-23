"""WorkbookMetadata contract — single source of truth for all structural metadata fields.

This is a pure module: no FastMCP server imports, no COM imports.
Only stdlib, openpyxl, fastmcp.exceptions, and excelmcp._core are allowed.
"""
from __future__ import annotations

import os
from datetime import datetime, timezone
from typing import TypedDict

import openpyxl
from fastmcp.exceptions import ToolError

from excelmcp._core import _check_path

CONTRACT_VERSION: str = "1.0"

CANONICAL_FIELDS: tuple[str, ...] = (
    "path",
    "workbook_name",
    "sheet_names",
    "active_sheet",
    "sheet_count",
    "has_formulas",
    "last_modified",
    "size_bytes",
    "format",
    "contract_version",
)


class WorkbookMetadata(TypedDict):
    path: str
    workbook_name: str
    sheet_names: list[str]
    active_sheet: str
    sheet_count: int
    has_formulas: bool
    last_modified: str
    size_bytes: int
    format: str
    contract_version: str


def build_workbook_metadata(path: str) -> WorkbookMetadata:
    """Build canonical WorkbookMetadata for the workbook at *path*.

    Raises ToolError if any field cannot be computed.
    Raises ValidationError (from _check_path) if path is not allowlisted.
    """
    # 1. Validate path via allowlist — raises ValidationError on failure
    resolved = _check_path(path)

    # 2. Gather filesystem stats (size, mtime) — raises ToolError on OSError
    try:
        stat = resolved.stat()
        size_bytes: int = stat.st_size
        last_modified: str = datetime.fromtimestamp(
            stat.st_mtime, tz=timezone.utc
        ).isoformat()
    except OSError as exc:
        raise ToolError(f"Cannot stat file '{path}': {exc}") from exc

    # 3. Load workbook read-only, data_only=False so formula strings are preserved
    try:
        wb = openpyxl.load_workbook(str(resolved), read_only=True, data_only=False)
    except Exception as exc:
        raise ToolError(f"Cannot open workbook '{path}': {exc}") from exc

    # Sections 4+5 — wrapped in outer try/finally for fd safety
    try:
        # 4. Extract structural fields
        try:
            sheet_names: list[str] = wb.sheetnames
            sheet_count: int = len(sheet_names)
            active_ws = wb.active
            active_sheet: str = active_ws.title if active_ws is not None else (sheet_names[0] if sheet_names else "")
            fmt: str = resolved.suffix.lower()
            workbook_name: str = os.path.basename(str(resolved))
        except ToolError:
            raise
        except Exception as exc:
            raise ToolError(f"Cannot read workbook structure '{path}': {exc}") from exc

        # 5. Scan all sheets for any cell whose value starts with "="
        has_formulas: bool = False
        try:
            for ws in wb.worksheets:
                for row in ws.iter_rows():
                    for cell in row:
                        if isinstance(cell.value, str) and cell.value.startswith("="):
                            has_formulas = True
                            break           # inner break
                    if has_formulas:
                        break               # row break
                if has_formulas:
                    break                   # sheet break
        except ToolError:
            raise
        except Exception as exc:
            raise ToolError(f"Formula scan failed for '{path}': {exc}") from exc

    finally:
        try:
            wb.close()
        except Exception:
            pass

    return WorkbookMetadata(
        path=str(resolved),
        workbook_name=workbook_name,
        sheet_names=sheet_names,
        active_sheet=active_sheet,
        sheet_count=sheet_count,
        has_formulas=has_formulas,
        last_modified=last_modified,
        size_bytes=size_bytes,
        format=fmt,
        contract_version=CONTRACT_VERSION,
    )


def validate_metadata(meta: dict) -> None:
    """Raise ValueError if any CANONICAL_FIELDS key is absent from *meta*."""
    missing = [f for f in CANONICAL_FIELDS if f not in meta]
    if missing:
        raise ValueError(f"Missing required metadata fields: {missing}")
