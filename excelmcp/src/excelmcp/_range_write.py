"""
Write-path range helpers: create_workbook, write_cell, write_range,
bulk_write_cells, append_rows, save.

No MCP/FastMCP imports. Pure openpyxl + _core.
"""
from __future__ import annotations

import logging
import re
from typing import Any

from openpyxl.utils import column_index_from_string as _col_idx

from excelmcp._active_wb import _check_file_not_open_in_excel

from ._core import (
    ExcelMCPError,
    OfficeCOMError,
    Path,
    ValidationError,
    _audit_log,
    _check_confirm,
    _check_path,
    _check_range_size,
    _check_write,
    _evict_wb,
    _load_wb,
    _resolve_table_last_row,
    _save_and_evict,
    _sync_autofilter_to_table,
    _wb_cache,
    openpyxl,
)

logger = logging.getLogger(__name__)


def create_workbook(
    path: str,
    sheets: list[str] | None = None,
    overwrite: bool = False,
    confirm: bool = False,
) -> dict:
    """Create a new .xlsx workbook at *path*.

    sheets: list of worksheet names to create (default: ["Sheet1"]).
    overwrite: if False (default), raises if the file already exists.
    confirm: must be True — write-gate requirement.
    """
    _check_write()
    _check_confirm(confirm)

    p = Path(path)
    if p.suffix.lower() != ".xlsx":
        raise ExcelMCPError("create_workbook only supports .xlsx files.")

    resolved = _check_path(path)
    _check_file_not_open_in_excel(resolved)  # LC-2: block write if file open in Excel

    if resolved.exists() and not overwrite:
        raise ExcelMCPError(
            f"File already exists: {path}. Pass overwrite=True to replace it."
        )

    sheet_names = sheets if sheets else ["Sheet1"]

    wb = openpyxl.Workbook()
    wb.active.title = sheet_names[0]
    for name in sheet_names[1:]:
        wb.create_sheet(name)
    _save_and_evict(wb, resolved)
    _audit_log("create_workbook", str(resolved), None, None, len(sheet_names))
    return {"ok": True, "path": str(resolved), "sheets": sheet_names, "created": True}


def write_cell(
    path: str,
    sheet: str,
    address: str,
    value: Any,
    confirm: bool = False,
    session_mode: bool = False,
) -> dict:
    """Write a single cell value.

    Returns {ok, previous_value}.
    Requires EXCEL_ENABLE_WRITE=true and confirm=True.
    """
    _check_write()
    _check_confirm(confirm)
    if isinstance(value, str) and value.lstrip()[:1] in ("=", "+", "-", "@"):
        raise ValidationError("Formula injection blocked: to write formulas, use fill_formula_range")

    resolved = _check_path(path)
    if session_mode:
        from excelmcp._active_wb import _write_cell_via_com  # noqa: PLC0415
        return _write_cell_via_com(resolved, sheet, address, value)
    _check_file_not_open_in_excel(resolved)  # LC-2: block write if file open in Excel
    wb = _load_wb(resolved, for_write=True)
    if sheet not in wb.sheetnames:
        raise ValidationError(f"Sheet not found: {sheet!r}")
    ws = wb[sheet]
    cell = ws[address.upper()]
    previous_value = cell.value
    cell.value = value

    _save_and_evict(wb, resolved)
    _audit_log("write_cell", resolved, sheet, address.upper(), 1)
    return {"ok": True, "previous_value": previous_value}


def write_range(
    path: str,
    sheet: str,
    start_cell: str,
    values: list[list[Any]],
    confirm: bool = False,
    session_mode: bool = False,
) -> dict:
    """Write a 2-D values list starting at start_cell.

    Returns {ok, cells_written}.
    Raises ValidationError if the range exceeds EXCEL_MAX_RANGE_CELLS.
    Requires EXCEL_ENABLE_WRITE=true and confirm=True.
    """
    _check_write()
    _check_confirm(confirm)

    if not values:
        return {"ok": True, "cells_written": 0}

    # PRE-VALIDATE all values BEFORE loading workbook
    for row in values:
        for val in row:
            if isinstance(val, str) and val.lstrip()[:1] in ("=", "+", "-", "@"):
                raise ValidationError("Formula injection blocked: to write formulas, use fill_formula_range")

    n_rows = len(values)
    n_cols = max(len(row) for row in values)
    _check_range_size(n_rows * n_cols)

    resolved = _check_path(path)
    if session_mode:
        from excelmcp._active_wb import _write_range_via_com  # noqa: PLC0415
        return _write_range_via_com(resolved, sheet, start_cell, values)
    _check_file_not_open_in_excel(resolved)  # LC-2: block write if file open in Excel
    wb = _load_wb(resolved, for_write=True)
    if sheet not in wb.sheetnames:
        raise ValidationError(f"Sheet not found: {sheet!r}")
    ws = wb[sheet]

    try:
        start_col, start_row, _, _ = openpyxl.utils.cell.range_boundaries(
            f"{start_cell.upper()}:{start_cell.upper()}"
        )
    except Exception as exc:
        raise ValidationError(f"Invalid start_cell reference: {start_cell!r}") from exc

    cells_written = 0
    for r_offset, row in enumerate(values):
        for c_offset, val in enumerate(row):
            ws.cell(row=start_row + r_offset, column=start_col + c_offset, value=val)
            cells_written += 1

    end_col_letter = openpyxl.utils.get_column_letter(start_col + n_cols - 1)
    end_row = start_row + n_rows - 1
    range_ref = f"{start_cell.upper()}:{end_col_letter}{end_row}"
    _save_and_evict(wb, resolved)
    _audit_log("write_range", resolved, sheet, range_ref, n_rows)
    return {"ok": True, "cells_written": cells_written}


def bulk_write_cells(
    path: str,
    sheet: str,
    writes: list[dict],
    confirm: bool = False,
) -> dict:
    """Write multiple non-contiguous cells in a single workbook open/save cycle.

    Each item in ``writes`` must be ``{"address": <cell_ref>, "value": <scalar>}``.
    All values are validated for formula injection BEFORE the workbook is opened.
    Requires ``EXCEL_ENABLE_WRITE=true`` and ``confirm=True``.

    Returns ``{"ok": True, "cells_written": N}``.
    """
    _check_write()
    _check_confirm(confirm)

    if not isinstance(writes, list):
        raise ValidationError("writes must be a list")

    if not writes:
        return {"ok": True, "cells_written": 0}

    # H-002: required-key guard — validate keys before any expensive work
    for i, item in enumerate(writes):
        if not isinstance(item, dict):
            raise ValidationError(f"writes[{i}] must be a dict")
        if "address" not in item:
            raise ValidationError(f"writes[{i}] missing required key 'address'")
        if "value" not in item:
            raise ValidationError(f"writes[{i}] missing required key 'value'")

    # H-006: size check BEFORE injection scan (avoids O(n) work on oversized lists)
    _check_range_size(len(writes))

    # PRE-VALIDATE all values BEFORE loading workbook (formula injection guard)
    for item in writes:
        val = item.get("value")
        if isinstance(val, str) and val.lstrip()[:1] in ("=", "+", "-", "@"):
            raise ValidationError(f"Formula injection blocked in value: {val!r}")

    resolved = _check_path(path)
    _check_file_not_open_in_excel(resolved)
    wb = _load_wb(resolved, for_write=True)
    if sheet not in wb.sheetnames:
        _evict_wb(resolved)
        raise ValidationError(f"Sheet not found: {sheet!r}")
    ws = wb[sheet]
    cells_written = 0
    # H-003: wrap write loop + save in try/except so cache is evicted on error
    try:
        for item in writes:
            address = str(item["address"]).upper()
            # H-001: validate address format before passing to openpyxl
            if not re.fullmatch(r"[A-Z]{1,3}[1-9][0-9]{0,6}", address):
                raise ValidationError(f"Invalid cell address: {address!r}")
            # H2-002: bounds check — reject out-of-range coords before openpyxl silently discards
            _col_str = re.match(r"[A-Z]+", address).group()  # type: ignore[union-attr]
            _row_num = int(address[len(_col_str):])
            if _col_idx(_col_str) > 16384:
                raise ValidationError(f"Column {_col_str!r} exceeds Excel maximum XFD (col 16384)")
            if _row_num > 1048576:
                raise ValidationError(f"Row {_row_num} exceeds Excel maximum (1048576)")
            ws[address].value = item["value"]
            cells_written += 1
        _save_and_evict(wb, resolved)
    except ExcelMCPError:
        _evict_wb(resolved)
        raise
    except Exception as exc:
        _evict_wb(resolved)
        raise ValidationError(f"bulk_write_cells failed: {exc}") from exc
    _audit_log("bulk_write_cells", resolved, sheet, f"{len(writes)} cells", None)
    return {"ok": True, "cells_written": cells_written}


def append_rows(
    path: str,
    sheet: str,
    rows: list[list[Any]],
    confirm: bool = False,
) -> dict:
    """Append rows after the last used row.

    Returns {ok, rows_appended, first_row} where first_row is the 1-based row
    index of the first appended row.
    Requires EXCEL_ENABLE_WRITE=true and confirm=True.
    """
    _check_write()
    _check_confirm(confirm)

    # PRE-VALIDATE all values BEFORE loading workbook
    for row in rows:
        for val in row:
            if isinstance(val, str) and val.lstrip()[:1] in ("=", "+", "-", "@"):
                raise ValidationError("Formula injection blocked: to write formulas, use fill_formula_range")

    resolved = _check_path(path)
    _check_file_not_open_in_excel(resolved)  # LC-2: block write if file open in Excel
    wb = _load_wb(resolved, for_write=True)
    if sheet not in wb.sheetnames:
        raise ValidationError(f"Sheet not found: {sheet!r}")
    ws = wb[sheet]

    # --- T1: table-aware append anchor -----------------------------------
    first_row = _resolve_table_last_row(ws) + 1   # replaces (ws.max_row or 0) + 1

    # Determine the starting column.
    # - For table-backed sheets: use the table's min_col so new rows land
    #   inside the table's column range.
    # - For plain sheets: default to column 1 (replicates ws.append behaviour).
    start_col = 1
    for tbl in ws.tables.values():
        tbl_ref = tbl.ref or ""
        if not tbl_ref:
            continue
        try:
            tbl_min_col, _r0, _c1, _r1 = openpyxl.utils.cell.range_boundaries(
                tbl_ref.upper()
            )
            start_col = tbl_min_col
        except Exception:
            continue
        break  # first valid table sets start_col

    # Explicit cell-by-cell write (replaces ws.append(row) loop).
    # Values are already formula-injection-validated above.
    for r_idx, row in enumerate(rows):
        for c_idx, val in enumerate(row):
            ws.cell(row=first_row + r_idx, column=start_col + c_idx, value=val)

    # Expand tbl.ref for every valid table to encompass the newly written rows.
    # Guard: skip tables with None/empty/malformed ref (BLOCKER 1 pattern).
    for tbl in ws.tables.values():
        tbl_ref = tbl.ref or ""
        if not tbl_ref:
            continue
        try:
            t_min_col, t_min_row, t_max_col, _t_max_row = (
                openpyxl.utils.cell.range_boundaries(tbl_ref.upper())
            )
        except Exception:
            continue
        new_max_row = (first_row + len(rows) - 1)
        min_col_letter = openpyxl.utils.get_column_letter(t_min_col)
        max_col_letter = openpyxl.utils.get_column_letter(t_max_col)
        tbl.ref = f"{min_col_letter}{t_min_row}:{max_col_letter}{new_max_row}"

    # T2: sync autofilter to the (now-expanded) table ref.
    _sync_autofilter_to_table(ws)

    _save_and_evict(wb, resolved)
    _audit_log("append_rows", resolved, sheet, None, len(rows))
    return {"ok": True, "rows_appended": len(rows), "first_row": first_row}


def save(path: str, confirm: bool = False) -> dict:
    """Save the workbook to disk and evict it from the cache.

    Returns {ok, path}.
    Requires EXCEL_ENABLE_WRITE=true and confirm=True.
    """
    _check_write()
    _check_confirm(confirm)

    resolved = _check_path(path)

    cached = _wb_cache.get((str(resolved), True))
    if cached is None:
        # No pending in-memory writes — workbook is already up to date on disk.
        # Return a graceful no-op rather than an error; this is a valid user pattern
        # (defensive save after a structural op that already auto-saved).
        return {"ok": True, "path": str(resolved), "note": "no_pending_writes"}

    # _wb_cache values are (Workbook, mtime) tuples; handle legacy bare-Workbook entries
    wb = cached[0] if isinstance(cached, tuple) else cached
    try:
        _save_and_evict(wb, resolved)
    except PermissionError as exc:
        logger.error("WORKBOOK_LOCKED in save(): %s", exc)
        raise OfficeCOMError("WORKBOOK_LOCKED") from exc
    _audit_log("save", resolved, None, None, None)
    return {"ok": True, "path": str(resolved)}
