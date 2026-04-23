"""
Core infrastructure for Excel MCP openpyxl layer.

Contains path validation, governance guards, workbook cache, and I/O helpers.
Sheet utilities, range helpers, type helpers, and error classes are defined in
_core_helpers.py and re-exported here so all existing callers continue to work
without modification.

No intra-package imports (except _core_helpers) — only stdlib + openpyxl.
No MCP / FastMCP imports.
"""
from __future__ import annotations

import json
import logging
import os
from datetime import datetime, timezone
from pathlib import Path

import openpyxl
import openpyxl.utils
import openpyxl.utils.cell

# ---------------------------------------------------------------------------
# Re-export everything from _core_helpers so that all existing callers that do
#   ``from ._core import ValidationError, _parse_a1_range, …``
# continue to work without any change.
# ---------------------------------------------------------------------------
from ._core_helpers import (  # noqa: F401
    ExcelMCPError,
    NotAllowedError,
    OfficeCOMError,
    ValidationError,
    _HEX_COLOR_RE,
    _NAMED_RANGE_RE,
    _NAMED_RANGE_RESERVED,
    _RANGE_ADDRESS_RE,
    _VALID_BORDER_STYLES,
    _adjust_table_refs_after_delete,
    _adjust_table_refs_after_insert,
    _get_used_range,
    _parse_a1_range,
    _parse_named_range_ref,
    _read_range_cells,
    _resolve_table_last_row,
    _sync_autofilter_to_table,
    _validate_formula_safe,
)

# ---------------------------------------------------------------------------
# Governance helpers
# ---------------------------------------------------------------------------

def _check_path(path: str) -> Path:
    """Validate path against EXCEL_ALLOWLIST_ROOTS and extension rules."""
    if "\x00" in str(path):
        raise ValidationError("Path contains illegal null byte")
    if str(path).startswith("\\\\") or str(path).startswith("//"):
        raise ValidationError("UNC paths are not permitted")
    roots_env = os.getenv("EXCEL_ALLOWLIST_ROOTS", "").strip()
    if not roots_env:
        raise ValidationError("No allowlist configured")

    resolved = Path(path).resolve()

    # File-size pre-flight: reject oversized workbooks before openpyxl loads them.
    if resolved.exists():
        max_bytes = int(os.getenv("EXCEL_MAX_FILE_MB", "50")) * 1024 * 1024
        if resolved.stat().st_size > max_bytes:
            raise ValidationError(
                f"File exceeds maximum allowed size "
                f"({int(os.getenv('EXCEL_MAX_FILE_MB', '50'))} MB): {resolved.name}"
            )

    # NOTE: TOCTOU residual risk — a symlink swap between resolve() and load_workbook()
    # is theoretically possible but requires SeCreateSymbolicLinkPrivilege (admin/dev mode).
    # Mitigated by the allowlist check on the resolved path.

    ext = resolved.suffix.lower()
    if ext not in (".xlsx", ".xlsm"):
        raise ValidationError(f"Unsupported extension: {ext!r}")

    roots = [r.strip() for r in roots_env.split(",") if r.strip()]
    for root in roots:
        try:
            if resolved.is_relative_to(Path(root).resolve()):
                return resolved
        except (ValueError, TypeError):
            continue

    raise ValidationError("Path not in allowlist")


def _check_csv_path(path: str) -> Path:
    """Validate a CSV/TSV file path against EXCEL_ALLOWLIST_ROOTS (allows .csv/.tsv)."""
    if "\x00" in str(path):
        raise ValidationError("Path contains illegal null byte")
    if str(path).startswith("\\\\") or str(path).startswith("//"):
        raise ValidationError("UNC paths are not permitted")
    roots_env = os.getenv("EXCEL_ALLOWLIST_ROOTS", "").strip()
    if not roots_env:
        raise ValidationError(
            "No allowlist configured — set EXCEL_ALLOWLIST_ROOTS to permit CSV access"
        )
    ext = Path(path).suffix.lower()
    if ext not in (".csv", ".tsv"):
        raise ValidationError(f"File must be .csv or .tsv, got: {ext!r}")
    resolved = Path(path).resolve()
    # Security: enforce allowlist symmetrically with _check_path — never allow
    # unrestricted filesystem access even for CSV imports.
    roots = [r.strip() for r in roots_env.split(",") if r.strip()]
    for root in roots:
        try:
            if resolved.is_relative_to(Path(root).resolve()):
                if not resolved.exists():
                    raise ValidationError("File not found")
                return resolved
        except (ValueError, TypeError):
            continue
    raise ValidationError("CSV path not in allowlist")


def _check_write() -> None:
    """Raise NotAllowedError unless EXCEL_ENABLE_WRITE=true."""
    if os.getenv("EXCEL_ENABLE_WRITE", "").strip().lower() != "true":
        raise NotAllowedError("Write operations require EXCEL_ENABLE_WRITE=true")


def _check_confirm(confirm: bool) -> None:
    """Raise ValidationError if confirm is not True."""
    if not confirm:
        raise ValidationError("confirm=True required for mutating operations")


def _check_range_size(n_cells: int) -> None:
    """Raise ValidationError if range exceeds EXCEL_MAX_RANGE_CELLS (default 5000)."""
    limit = int(os.getenv("EXCEL_MAX_RANGE_CELLS", "5000"))
    if n_cells > limit:
        raise ValidationError(
            f"Range too large: {n_cells} cells exceeds limit {limit}"
        )


def _audit_log(
    op: str,
    path: Path | str,
    sheet: str | None,
    range_ref: str | None,
    rows_affected: int | None,
) -> None:
    """Write a JSON audit record to stderr."""
    record = {
        "ts": datetime.now(timezone.utc).isoformat(),
        "op": op,
        "path": str(path),
        "sheet": sheet,
        "range": range_ref,
        "rows_affected": rows_affected,
    }
    logging.getLogger(__name__).info(json.dumps(record))


# ---------------------------------------------------------------------------
# Workbook cache
# ---------------------------------------------------------------------------

# Key is (resolved_path_str, for_write) so read-only and write-mode entries
# are stored separately and never mixed up.
# The _WbCache subclass keeps legacy string-key lookups working for tests
# that do `path_str in _wb_cache`.


class _WbCache(dict):
    """dict subclass whose ``__contains__`` also accepts plain str path keys.

    A bare string ``s`` matches if either ``(s, False)`` or ``(s, True)``
    is present, so code that pre-dates the tuple-key change still works.

    Values are ``(openpyxl.Workbook, float | None)`` tuples — the second
    element is the ``st_mtime`` recorded when the workbook was loaded.
    A ``None`` mtime means the file's stat was unavailable at load time.
    """

    def __contains__(self, key: object) -> bool:
        if isinstance(key, str):
            return (
                super().__contains__((key, False))
                or super().__contains__((key, True))
            )
        return super().__contains__(key)


_wb_cache: _WbCache = _WbCache()


def _load_wb(path: Path, for_write: bool = False) -> openpyxl.Workbook:
    key = (str(path), for_write)
    try:
        current_mtime = path.stat().st_mtime
    except OSError:
        current_mtime = None  # new file or race — let load_workbook handle it

# When a read-only lookup is requested, we must prefer the
    # in-memory write-mode entry if it exists — it holds the most recent (possibly
    # unsaved/dirty) workbook state. This fixes bugs where tools like list_merged_cells
    # would receive stale data if the read-cache was populated before mutations.
    if not for_write:
        write_key = (str(path), True)
        write_entry = _wb_cache.get(write_key)
        if write_entry is not None:
            if not isinstance(write_entry, tuple):
                return write_entry  # legacy bare-Workbook entry — serve unconditionally
            cached_wb, cached_mtime = write_entry
            if current_mtime is None or cached_mtime == current_mtime:
                return cached_wb
            # Write-mode entry is also stale — evict both and fall through to reload
            _evict_wb(path)

    if key in _wb_cache:
        cached_value = _wb_cache[key]
        if not isinstance(cached_value, tuple):
            # Legacy bare-Workbook entry (e.g., injected by tests) — serve unconditionally
            # TODO: remove this guard after all test fixtures are migrated to (wb, mtime) tuple form
            return cached_value
        cached_wb, cached_mtime = cached_value
        if current_mtime is None or cached_mtime == current_mtime:
            return cached_wb
        # File modified externally — evict stale entry and fall through to reload
        _evict_wb(path)

    try:
        wb = openpyxl.load_workbook(str(path), data_only=not for_write)
    except PermissionError as e:
        raise OfficeCOMError("WORKBOOK_LOCKED") from e
    except Exception as exc:
        raise ExcelMCPError(f"Workbook error ({type(exc).__name__})") from None
    _wb_cache[key] = (wb, current_mtime)
    return wb


def _evict_wb(path: Path) -> None:
    """Evict both read-only and write-mode cached entries for *path*."""
    _wb_cache.pop((str(path), False), None)
    _wb_cache.pop((str(path), True), None)


def _save_and_evict(wb, resolved: Path) -> None:
    """Save *wb* to disk, apply the DA-formula OOXML patch, then evict from cache.

    This is the single canonical save path so that every write to disk
    carries the cm attributes and metadata.xml that Excel 365 requires to
    open spill formulas (LET/FILTER/SORT/…) without a repair dialog.
    openpyxl drops these extensions on every round-trip, so the patch must
    run after *every* wb.save() call, not just the explicit save() tool.

    COM Formula2 cells (those with ``aca="1"`` on their ``<f>`` element)
    are snapshotted BEFORE the openpyxl save and their ``t="array" aca ref ca``
    attributes are restored AFTER the DA patcher runs, so that LET+FILTER
    spill formulas written by Excel COM survive MCP round-trips.
    """
    from ._ooxml import restore_formula2_f_attrs, snapshot_formula2_cells

    path_str = str(resolved)
    f2_snap = snapshot_formula2_cells(path_str)  # capture BEFORE openpyxl strips attrs
    wb.save(path_str)
    try:
        _finalize_and_evict(resolved)
        if f2_snap:
            restore_formula2_f_attrs(path_str, f2_snap)
    finally:
        _evict_wb(resolved)


def _flush_if_dirty(resolved_path: str) -> None:
    """Save any pending in-memory write-mode workbook before a structural op."""
    from ._ooxml import restore_formula2_f_attrs, snapshot_formula2_cells

    key = (resolved_path, True)
    if key in _wb_cache:
        f2_snap = snapshot_formula2_cells(resolved_path)  # capture BEFORE openpyxl strips attrs
        cached_entry = _wb_cache[key]
        wb = cached_entry[0] if isinstance(cached_entry, tuple) else cached_entry
        wb.save(resolved_path)
        try:
            _finalize_and_evict(Path(resolved_path))
            if f2_snap:
                restore_formula2_f_attrs(resolved_path, f2_snap)
        finally:
            _wb_cache.pop(key, None)


def _finalize_and_evict(resolved: Path) -> None:
    """Run the canonical DA finalize pipeline for a just-saved workbook path."""
    from ._da_finalize import finalize_workbook_da_artifact

    finalize_workbook_da_artifact(str(resolved))


# ---------------------------------------------------------------------------
# copy_range (bridges read + write; uses both helpers and I/O guards)
# ---------------------------------------------------------------------------


def copy_range(
    path: str,
    sheet: str,
    src_range: str,
    dst_range: str,
    confirm: bool = False,
) -> dict:
    """Copy cells (values and formats) from *src_range* to *dst_range* within the same sheet.

    Copies value, font, fill, number_format, alignment, and border for each cell.
    Requires ``EXCEL_ENABLE_WRITE=true`` and ``confirm=True``.
    Returns ``{"copied": N, "src": src_range, "dst": dst_range}``.
    """
    from copy import copy as _copy

    _check_write()
    _check_confirm(confirm)
    resolved = _check_path(path)
    _flush_if_dirty(str(resolved))
    wb_obj = _load_wb(resolved, for_write=True)
    if sheet not in wb_obj.sheetnames:
        raise ValidationError(f"Sheet not found: {sheet!r}")
    ws = wb_obj[sheet]

    src_min_col, src_min_row, src_max_col, src_max_row = _parse_a1_range(src_range)
    dst_min_col, dst_min_row, _, _ = _parse_a1_range(dst_range)

    src_rows = list(
        ws.iter_rows(
            min_row=src_min_row,
            max_row=src_max_row,
            min_col=src_min_col,
            max_col=src_max_col,
        )
    )

    copied = 0
    for r_idx, row in enumerate(src_rows):
        for c_idx, src_cell in enumerate(row):
            dst_r = dst_min_row + r_idx
            dst_c = dst_min_col + c_idx
            dst_cell = ws.cell(row=dst_r, column=dst_c)
            dst_cell.value = src_cell.value
            # Copy formatting attribute by attribute (openpyxl styles are immutable objects)
            try:
                dst_cell.font = _copy(src_cell.font)
            except Exception:
                pass
            try:
                dst_cell.fill = _copy(src_cell.fill)
            except Exception:
                pass
            try:
                dst_cell.number_format = src_cell.number_format
            except Exception:
                pass
            try:
                dst_cell.alignment = _copy(src_cell.alignment)
            except Exception:
                pass
            try:
                dst_cell.border = _copy(src_cell.border)
            except Exception:
                pass
            copied += 1

    _save_and_evict(wb_obj, resolved)
    _audit_log("copy_range", resolved, sheet, f"{src_range}->{dst_range}", copied)
    return {"copied": copied, "src": src_range, "dst": dst_range}
