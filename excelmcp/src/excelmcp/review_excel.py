"""
review_excel.py — Openpyxl-based render review gate for Excel workbooks.

Mirrors the PPT Phase D review_pptx.py pattern.
No COM, no FastMCP/MCP imports. Pure openpyxl + stdlib only.

Checks (all return findings: list[dict]):
  PRINT_AREA_UNSET        — ws.print_area is None or empty            (severity: medium)
  SHEET_EMPTY             — ws.max_row <= 1 after filter blanks        (severity: low)
  FORMULA_STALE           — workbook has formulas + recalculated=False (severity: medium)
  PRINT_AREA_EXCEEDS_PAGE — estimated col count > 10                  (severity: low)

Finding dict schema:
  {sheet: str, criterion: str, result: "WARN"|"OK", detail: str, severity: "high"|"medium"|"low"}

Main function:
  review_workbook_render(path: str, sheet_names: list[str] | None = None,
                         recalculated: bool = False) -> dict

Returns:
  {
    path: str,
    sheets_reviewed: list[str],
    passed: bool,              # True if no WARN findings with severity >= medium
    findings: list[finding],
    checks_run: int,
    findings_count: int,
  }
"""
from __future__ import annotations

import logging

import openpyxl
from openpyxl.utils.cell import range_boundaries

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Per-sheet check helpers
# ---------------------------------------------------------------------------

def _check_print_area_unset(ws: object, sheet_name: str) -> list[dict]:
    """Return a WARN finding if ws.print_area is None or empty."""
    try:
        pa = ws.print_area  # type: ignore[attr-defined]
    except AttributeError:
        # Chart sheets or other non-standard sheet types
        return [{
            "sheet": sheet_name,
            "criterion": "PRINT_AREA_UNSET",
            "result": "WARN",
            "detail": (
                f"Sheet '{sheet_name}': print_area not accessible "
                "(chart sheet or unsupported sheet type)."
            ),
            "severity": "low",
        }]
    if not pa:
        return [{
            "sheet": sheet_name,
            "criterion": "PRINT_AREA_UNSET",
            "result": "WARN",
            "detail": (
                f"Sheet '{sheet_name}' has no print area set; "
                "PDF export may include unexpected blank areas."
            ),
            "severity": "medium",
        }]
    return []


def _check_sheet_empty(ws: object, sheet_name: str) -> list[dict]:
    """Return a WARN finding if the sheet appears empty (max_row <= 1)."""
    max_row = getattr(ws, "max_row", None)
    if max_row is None or max_row <= 1:
        return [{
            "sheet": sheet_name,
            "criterion": "SHEET_EMPTY",
            "result": "WARN",
            "detail": (
                f"Sheet '{sheet_name}' has {max_row or 0} row(s); "
                "appears empty or contains only a header row."
            ),
            "severity": "low",
        }]
    return []


def _check_print_area_exceeds_page(ws: object, sheet_name: str) -> list[dict]:
    """Return a WARN finding if the print area spans more than 10 columns."""
    try:
        pa = ws.print_area  # type: ignore[attr-defined]
    except AttributeError:
        return []
    if not pa:
        return []
    # Handle multi-range print areas like "A1:J50,A52:J100" or
    # sheet-qualified forms like "'Sheet1'!$A$1:$AA$50"
    first_part = pa.split(",")[0].strip()
    # Strip sheet reference prefix (e.g. "'Sheet1'!" or "Sheet1!")
    if "!" in first_part:
        first_part = first_part.split("!", 1)[1].strip()
    try:
        min_col, _, max_col, _ = range_boundaries(first_part)
        col_count = max_col - min_col + 1
    except Exception:
        return []
    if col_count > 10:
        return [{
            "sheet": sheet_name,
            "criterion": "PRINT_AREA_EXCEEDS_PAGE",
            "result": "WARN",
            "detail": (
                f"Sheet '{sheet_name}' print area spans {col_count} columns "
                f"(> 10); may not fit on a single page in portrait orientation."
            ),
            "severity": "low",
        }]
    return []


# ---------------------------------------------------------------------------
# Workbook-level formula scan (loads separately with data_only=False)
# ---------------------------------------------------------------------------

def _has_formulas(path: str, sheet_names: list[str]) -> tuple[bool, bool]:
    """Return (has_formulas, scan_failed).

    has_formulas: True if any reviewed sheet contains at least one formula cell.
    scan_failed: True if the scan encountered an error and the result is unreliable.

    Loads the workbook with data_only=False so raw formula strings are visible.
    Any cell whose value starts with '=' is treated as a formula.
    """
    wb_raw = None
    try:
        wb_raw = openpyxl.load_workbook(path, read_only=True, data_only=False)
        for name in sheet_names:
            if name not in wb_raw.sheetnames:
                continue
            ws = wb_raw[name]
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        return True, False
    except Exception as exc:
        logger.warning("[review_excel] Formula scan failed for %r: %s", path, exc)
        return False, True
    finally:
        if wb_raw is not None:
            try:
                wb_raw.close()
            except Exception:
                pass
    return False, False


# ---------------------------------------------------------------------------
# Main public function
# ---------------------------------------------------------------------------

def review_workbook_render(
    path: str,
    sheet_names: list[str] | None = None,
    recalculated: bool = False,
) -> dict:
    """Review an Excel workbook's render quality before/after PDF export.

    Runs structural checks (PRINT_AREA_UNSET, SHEET_EMPTY, FORMULA_STALE,
    PRINT_AREA_EXCEEDS_PAGE) on the specified sheets and returns findings.

    Read-only — does NOT modify the workbook.

    Args:
        path: Absolute path to the .xlsx workbook.
        sheet_names: Sheet names to review. None → all non-chart sheets.
        recalculated: Set True if recalculate_workbook() was called before
                      export. Suppresses the FORMULA_STALE finding.

    Returns:
        {
            path: str,
            sheets_reviewed: list[str],
            passed: bool,              # True if no WARN with severity "high" or "medium"
            findings: list[dict],
            checks_run: int,
            findings_count: int,
        }
    """
    # read_only=False required: ReadOnlyWorksheet does not expose ws.print_area
    # (it raises AttributeError). Loading without read_only gives the full
    # Worksheet object with print_area accessible via defined_names.
    # data_only=True is retained so formulas show cached values (formula scan
    # is done separately by _has_formulas with data_only=False).
    wb = openpyxl.load_workbook(path, read_only=False, data_only=True)
    try:
        # Resolve sheets to review — wb.worksheets excludes ChartSheet instances
        if sheet_names is None:
            sheets_to_review: list[str] = [ws.title for ws in wb.worksheets]
        else:
            sheets_to_review = list(sheet_names)

        # FORMULA_STALE — workbook-level scan (loads separately)
        has_formulas, scan_failed = _has_formulas(path, sheets_to_review)

        findings: list[dict] = []
        per_sheet_check_count = 0

        for name in sheets_to_review:
            if name not in wb.sheetnames:
                logger.warning("[review_excel] Sheet %r not in workbook, skipping.", name)
                continue
            ws = wb[name]
            findings.extend(_check_print_area_unset(ws, name))
            findings.extend(_check_sheet_empty(ws, name))
            findings.extend(_check_print_area_exceeds_page(ws, name))
            per_sheet_check_count += 3

        # FORMULA_SCAN_FAILED finding — surfaced when the formula scan errors out
        if scan_failed:
            findings.append({
                "sheet": "ALL",
                "criterion": "FORMULA_SCAN_FAILED",
                "result": "WARN",
                "detail": "Formula scan could not complete — formula staleness unknown",
                "severity": "low",
            })

        # FORMULA_STALE finding (global, one per workbook review call)
        if has_formulas and not recalculated:
            findings.append({
                "sheet": sheets_to_review[0] if sheets_to_review else "workbook",
                "criterion": "FORMULA_STALE",
                "result": "WARN",
                "detail": (
                    "Workbook contains formulas and recalculated=False; "
                    "exported cached values may be stale. "
                    "Call recalculate_workbook(confirm=True) before export."
                ),
                "severity": "medium",
            })

        # +1 for the formula scan (always run, even if no formulas found)
        checks_run = per_sheet_check_count + 1

        # passed = True only if there are no WARN findings with severity "high" or "medium"
        passed = not any(
            f["result"] == "WARN" and f["severity"] in ("high", "medium")
            for f in findings
        )

        return {
            "path": path,
            "sheets_reviewed": sheets_to_review,
            "passed": passed,
            "findings": findings,
            "checks_run": checks_run,
            "findings_count": len(findings),
        }
    finally:
        try:
            wb.close()
        except Exception:
            pass
