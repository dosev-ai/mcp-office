"""Read workbook range routine — openpyxl-based, COM-free.

Public entrypoint: read_workbook_range_v1

Layer: no COM, no MCP imports. Standard library + openpyxl only.
Allowlist: EXCEL_ALLOWLIST_ROOTS (comma-separated directory prefixes).
"""
from __future__ import annotations

import os
from typing import Any

# ---------------------------------------------------------------------------
# Allowlist helpers
# ---------------------------------------------------------------------------


def _check_path_allowlist(workbook_path: str) -> dict | None:
    """Validate workbook_path against EXCEL_ALLOWLIST_ROOTS.

    Returns an error dict if validation fails, None if OK.
    Uses os.path.realpath() to resolve symlinks before comparison.
    """
    roots_env = os.environ.get("EXCEL_ALLOWLIST_ROOTS", "").strip()
    if not roots_env:
        return {"error": "EXCEL_ALLOWLIST_ROOTS not configured", "error_kind": "config_error"}

    roots = [r.strip() for r in roots_env.split(",") if r.strip()]
    if not roots:
        return {"error": "EXCEL_ALLOWLIST_ROOTS not configured", "error_kind": "config_error"}

    real_path = os.path.normcase(os.path.realpath(workbook_path))
    for root in roots:
        real_root = os.path.normcase(os.path.realpath(root))
        # Ensure root ends with sep so "C:\foo" doesn't match "C:\foobar"
        if not real_root.endswith(os.sep):
            real_root = real_root + os.sep
        if real_path.startswith(real_root) or real_path == real_root.rstrip(os.sep):
            return None  # allowed

    return {"error": "path not in allowlist", "error_kind": "allowlist_violation"}


# ---------------------------------------------------------------------------
# Public function
# ---------------------------------------------------------------------------


def read_workbook_range_v1(
    workbook_path: str,
    sheet_name: str,
    range_ref: str | None = None,
    max_rows: int = 1000,
) -> dict[str, Any]:
    """Read a range from an Excel workbook and return rows + headers.

    Uses openpyxl in read_only mode — no COM, no win32com.

    Args:
        workbook_path: Absolute path to the .xlsx file.
        sheet_name:    Sheet to read from.
        range_ref:     Optional Excel range string e.g. "A1:D10".
                       If None, reads the sheet's used range.
        max_rows:      Maximum number of rows to return (default 1000).
                       The header row does NOT count against this limit.

    Returns:
        On success:
            {
                "rows": [[cell_values...]],
                "headers": [str...],
                "row_count": int,
                "sheet_name": str,
                "workbook_path": str,
                "evidence": {
                    "routine": "excel.read_workbook_range_v1",
                    "max_rows": int,
                    "range_ref": str | None,
                },
            }
        On error:
            {"error": str, "error_kind": str}
    """
    try:
        # --- allowlist gate ---
        err = _check_path_allowlist(workbook_path)
        if err is not None:
            return err

        # --- imports deferred so the module-level import never fails ---
        from openpyxl import load_workbook  # noqa: PLC0415
        from openpyxl.utils import column_index_from_string  # noqa: PLC0415
        from openpyxl.utils.cell import coordinate_from_string  # noqa: PLC0415

        wb = load_workbook(workbook_path, read_only=True, data_only=True)
        try:
            if sheet_name not in wb.sheetnames:
                return {
                    "error": f"Sheet {sheet_name!r} not found in workbook",
                    "error_kind": "sheet_not_found",
                }
            ws = wb[sheet_name]

            if range_ref is not None:
                # Parse range_ref like "A1:D10" or "B2:F20"
                try:
                    if ":" in range_ref:
                        top_left, bottom_right = range_ref.split(":", 1)
                        tl_col_str, tl_row = coordinate_from_string(top_left)
                        br_col_str, br_row = coordinate_from_string(bottom_right)
                        min_col = column_index_from_string(tl_col_str)
                        max_col = column_index_from_string(br_col_str)
                        min_row = tl_row
                        max_row = br_row
                    else:
                        col_str, row = coordinate_from_string(range_ref)
                        min_col = max_col = column_index_from_string(col_str)
                        min_row = max_row = row
                except Exception as parse_exc:
                    return {
                        "error": f"Invalid range_ref {range_ref!r}: {parse_exc}",
                        "error_kind": "invalid_range",
                    }
                cell_iter = ws.iter_rows(
                    min_row=min_row,
                    max_row=max_row,
                    min_col=min_col,
                    max_col=max_col,
                    values_only=True,
                )
            else:
                cell_iter = ws.iter_rows(values_only=True)

            all_rows: list[list[Any]] = [list(row) for row in cell_iter]
        finally:
            wb.close()

        if not all_rows:
            return {
                "rows": [],
                "headers": [],
                "row_count": 0,
                "sheet_name": sheet_name,
                "workbook_path": workbook_path,
                "evidence": {
                    "routine": "excel.read_workbook_range_v1",
                    "max_rows": max_rows,
                    "range_ref": range_ref,
                },
            }

        # First row = headers
        headers = [str(v) if v is not None else "" for v in all_rows[0]]
        data_rows = all_rows[1 : max_rows + 1]

        return {
            "rows": data_rows,
            "headers": headers,
            "row_count": len(data_rows),
            "sheet_name": sheet_name,
            "workbook_path": workbook_path,
            "evidence": {
                "routine": "excel.read_workbook_range_v1",
                "max_rows": max_rows,
                "range_ref": range_ref,
            },
        }

    except Exception as exc:  # noqa: BLE001
        return {"error": str(exc), "error_kind": "unexpected_error"}


__all__ = ["read_workbook_range_v1"]
