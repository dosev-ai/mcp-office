"""Pivot table COM helpers for excelmcp.

No MCP imports. No win32com imports at module scope.

Split from _com_template.py — formula evaluation lives in _com_recalc.py.

Public API:
    _create_pivot_table()     -- create a pivot table via Excel COM
    _refresh_pivot_tables()   -- refresh all pivot tables via Excel COM
    _read_pivot_table()       -- read pivot table structure and data via Excel COM
"""
from __future__ import annotations

import logging
import re

from excelmcp._com_gates import (
    _XL_COL_FIELD,
    _XL_COUNT,
    _XL_DATABASE,
    _XL_ROW_FIELD,
)
from excelmcp._com_recalc import _SINGLE_CELL_RE
from excelmcp._core import (
    NotAllowedError,
    OfficeCOMError,
    ValidationError,
    _check_confirm,
    _check_path,
    _check_range_size,
    _check_write,
)

_log = logging.getLogger(__name__)


def _create_pivot_table(
    path: str,
    source_range: str,
    dest_sheet: str,
    dest_cell: str,
    row_fields: list,
    col_fields: list,
    value_fields: list,
    table_name: str = "PivotTable1",
    confirm: bool = False,
) -> dict:
    """Create a pivot table in an existing workbook via Excel COM."""
    from excelmcp._com import _close_workbook, _com_excel_app, _open_workbook  # noqa: PLC0415
    from excelmcp._com_gates import _ensure_com_gate  # noqa: PLC0415

    _ensure_com_gate()
    _check_write()
    _check_confirm(confirm)
    resolved = _check_path(path)

    total_fields = len(row_fields) + len(col_fields) + len(value_fields)
    if total_fields > 50:
        raise ValidationError(
            f"Total field count ({total_fields}) exceeds maximum of 50 — "
            f"reduce row_fields + col_fields + value_fields"
        )

    if "!" not in source_range:
        raise ValidationError(
            f"source_range must include the sheet name (e.g. 'Sheet1!A1:D10'), "
            f"got: {source_range!r}"
        )
    range_part = source_range.rsplit("!", 1)[1]
    _range_only_re = re.compile(
        r"^\$?[A-Z]{1,3}\$?\d+(?::\$?[A-Z]{1,3}\$?\d+)?$", re.IGNORECASE
    )
    if not _range_only_re.match(range_part):
        raise ValidationError(f"source_range has invalid range portion: {range_part!r}")

    if not _SINGLE_CELL_RE.match(dest_cell):
        raise ValidationError(
            f"dest_cell must be A1-notation single cell (e.g. 'F1'), got: {dest_cell!r}"
        )

    # FIX 1 (US-H08): pre-flight dest_sheet check via openpyxl — avoids COM startup
    try:
        import openpyxl as _openpyxl_precheck  # noqa: PLC0415
        _wb_chk = _openpyxl_precheck.load_workbook(resolved, read_only=True, data_only=True)
        _chk_names = _wb_chk.sheetnames
        _wb_chk.close()
        if dest_sheet not in _chk_names:
            raise ValidationError(
                f"dest_sheet '{dest_sheet}' not found in workbook (available: {_chk_names})"
            )
    except ValidationError:
        raise
    except Exception:
        pass  # file not loadable by openpyxl (e.g. xlsb, temp) — let COM handle it

    with _com_excel_app() as excel:
        wb = _open_workbook(excel, resolved, read_only=False)
        try:
            wb.Activate()
        except Exception as _act_exc:
            _log.error("wb.Activate() failed: %s", _act_exc)
            raise OfficeCOMError(f"Failed to activate workbook: {_act_exc}") from _act_exc
        try:
            sheet_names = [wb.Sheets(i + 1).Name for i in range(wb.Sheets.Count)]
            if dest_sheet not in sheet_names:
                raise ValidationError(
                    f"Destination sheet {dest_sheet!r} not found in workbook "
                    f"(available: {sheet_names})"
                )
            try:
                cache = wb.PivotCaches().Create(
                    SourceType=_XL_DATABASE,
                    SourceData=source_range,
                )
            except (ValidationError, NotAllowedError, OfficeCOMError):
                raise
            except Exception as exc:
                _log.error("Pivot COM error [PivotCache create] source=%s: %s", source_range, exc)
                raise OfficeCOMError(
                    f"Failed to create PivotCache for range {source_range!r}"
                ) from exc

            dest_ws = wb.Sheets(dest_sheet)
            try:
                pivot = cache.CreatePivotTable(
                    TableDestination=dest_ws.Range(dest_cell.upper()),
                    TableName=table_name,
                )
            except (ValidationError, NotAllowedError, OfficeCOMError):
                raise
            except Exception as exc:
                _log.error("Pivot COM error [PivotTable create] name=%s: %s", table_name, exc)
                raise OfficeCOMError(
                    f"Failed to create PivotTable {table_name!r}"
                ) from exc

            for field_name in row_fields:
                try:
                    pivot.PivotFields(field_name).Orientation = _XL_ROW_FIELD
                except (ValidationError, NotAllowedError, OfficeCOMError):
                    raise
                except Exception as exc:
                    raise ValidationError(
                        f"Row field {field_name!r} not found in pivot: {exc}"
                    ) from exc

            for field_name in col_fields:
                try:
                    pivot.PivotFields(field_name).Orientation = _XL_COL_FIELD
                except (ValidationError, NotAllowedError, OfficeCOMError):
                    raise
                except Exception as exc:
                    raise ValidationError(
                        f"Column field {field_name!r} not found in pivot: {exc}"
                    ) from exc

            for field_name in value_fields:
                try:
                    pf = pivot.PivotFields(field_name)
                    pivot.AddDataField(pf, field_name, _XL_COUNT)
                except (ValidationError, NotAllowedError, OfficeCOMError):
                    raise
                except Exception as exc:
                    raise ValidationError(
                        f"Value field {field_name!r} not found in pivot: {exc}"
                    ) from exc

            wb.Save()
        finally:
            _close_workbook(wb, save=False)
            del wb

    return {
        "ok": True,
        "path": str(resolved),
        "table_name": table_name,
        "dest_sheet": dest_sheet,
        "row_fields": row_fields,
        "col_fields": col_fields,
        "value_fields": value_fields,
    }


def _refresh_pivot_tables(
    path: str,
    sheet: str | None = None,
    confirm: bool = False,
) -> dict:
    """Refresh all pivot tables (or a single sheet) via Excel COM.

    SSRF guard: only xlDatabase (local range) sources are permitted (RC-2).
    """
    from excelmcp._com import _close_workbook, _com_excel_app, _open_workbook  # noqa: PLC0415
    from excelmcp._com_gates import _ensure_com_gate  # noqa: PLC0415

    _ensure_com_gate()
    _check_write()
    _check_confirm(confirm)
    resolved = _check_path(path)

    refreshed: list = []

    with _com_excel_app() as excel:
        wb = _open_workbook(excel, resolved, read_only=False)
        try:
            wb.Activate()
        except Exception as _act_exc:
            _log.error("wb.Activate() failed: %s", _act_exc)
            raise OfficeCOMError(f"Failed to activate workbook: {_act_exc}") from _act_exc
        try:
            if sheet is not None:
                sheet_names = [wb.Sheets(i + 1).Name for i in range(wb.Sheets.Count)]
                if sheet not in sheet_names:
                    raise ValidationError(
                        f"Sheet {sheet!r} not found in workbook (available: {sheet_names})"
                    )
                sheets_to_process = [wb.Sheets(sheet)]
            else:
                sheets_to_process = [
                    wb.Sheets(i + 1) for i in range(wb.Sheets.Count)
                ]

            for ws in sheets_to_process:
                # FIX 3 (US-H12): PivotTables() can throw on sheets with no pivot tables
                try:
                    pt_count = ws.PivotTables().Count
                except Exception:
                    pt_count = 0
                for i in range(pt_count):
                    pt = ws.PivotTables(i + 1)
                    try:
                        source_type = pt.PivotCache().SourceType
                    except Exception as exc:
                        _log.error(
                            "Pivot COM error [SourceType read] name=%s: %s", pt.Name, exc
                        )
                        raise OfficeCOMError(
                            f"Cannot determine SourceType for pivot '{pt.Name}'"
                        ) from exc

                    if source_type != _XL_DATABASE:
                        raise ValidationError(
                            f"Cannot refresh pivot '{pt.Name}': external data sources are "
                            f"not permitted (SourceType={source_type}, expected {_XL_DATABASE})"
                        )

                    pt.RefreshTable()
                    refreshed.append(pt.Name)

            wb.Save()
        finally:
            _close_workbook(wb, save=False)
            del wb

    return {
        "ok": True,
        "path": str(resolved),
        "sheet": sheet,
        "refreshed": refreshed,
        "count": len(refreshed),
    }


def _read_pivot_table(path: str, sheet: str, pivot_name: str) -> dict:
    """Read the structure and data body of a named pivot table via Excel COM.

    Read-only — does NOT require EXCEL_ENABLE_WRITE.
    """
    from excelmcp._com import _close_workbook, _com_excel_app, _open_workbook  # noqa: PLC0415

    resolved = _check_path(path)

    # FIX 4 (US-H18): pre-flight sheet check via openpyxl — avoids COM startup
    try:
        import openpyxl as _openpyxl_precheck  # noqa: PLC0415
        _wb_chk = _openpyxl_precheck.load_workbook(resolved, read_only=True, data_only=True)
        _chk_names = _wb_chk.sheetnames
        _wb_chk.close()
        if sheet not in _chk_names:
            raise ValidationError(
                f"sheet '{sheet}' not found in workbook (available: {_chk_names})"
            )
    except ValidationError:
        raise
    except Exception:
        pass  # file not loadable by openpyxl (e.g. xlsb, temp) — let COM handle it

    row_fields: list = []
    col_fields: list = []
    value_fields: list = []
    page_fields: list = []
    row_count: int = 0
    col_count: int = 0
    data: list = []

    with _com_excel_app() as excel:
        wb = _open_workbook(excel, resolved, read_only=True)
        try:
            wb.Activate()
        except Exception:
            pass  # non-fatal in hidden Excel instance
        try:
            sheet_names = [wb.Sheets(i + 1).Name for i in range(wb.Sheets.Count)]
            if sheet not in sheet_names:
                raise ValidationError(
                    f"Sheet {sheet!r} not found in workbook (available: {sheet_names})"
                )

            ws = wb.Sheets(sheet)
            pivot_names: list = []
            target_pt = None
            for i in range(ws.PivotTables().Count):
                pt = ws.PivotTables(i + 1)
                pivot_names.append(pt.Name)
                if pt.Name == pivot_name:
                    target_pt = pt

            if target_pt is None:
                raise ValidationError(
                    f"Pivot table {pivot_name!r} not found on sheet {sheet!r} "
                    f"(available: {pivot_names})"
                )

            for i in range(target_pt.PivotFields().Count):
                pf = target_pt.PivotFields(i + 1)
                orient = pf.Orientation
                if orient == 1:
                    row_fields.append(pf.Name)
                elif orient == 2:
                    col_fields.append(pf.Name)
                elif orient == 3:
                    page_fields.append(pf.Name)
                elif orient == 4:
                    value_fields.append(pf.Name)

            data_range = target_pt.DataBodyRange
            if data_range is not None:
                try:
                    _check_range_size(data_range.Count)
                    raw = data_range.Value
                    if isinstance(raw, (list, tuple)):
                        for row in raw:
                            if isinstance(row, (list, tuple)):
                                data.append(list(row))
                            else:
                                data.append([row])
                    elif raw is not None:
                        data = [[raw]]
                    row_count = len(data)
                    col_count = len(data[0]) if data else 0
                except ValidationError:
                    raise
                except Exception:
                    data = []
                    row_count = 0
                    col_count = 0
        finally:
            _close_workbook(wb, save=False)
            del wb

    return {
        "ok": True,
        "path": str(resolved),
        "sheet": sheet,
        "pivot_name": pivot_name,
        "row_fields": row_fields,
        "col_fields": col_fields,
        "value_fields": value_fields,
        "page_fields": page_fields,
        "row_count": row_count,
        "col_count": col_count,
        "data": data,
    }
