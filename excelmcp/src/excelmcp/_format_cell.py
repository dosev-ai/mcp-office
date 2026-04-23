"""Cell-level style helpers: font, fill, alignment, borders, number format.

Internal sub-module. Import via excelmcp._format (public router).
All cache helpers are private (underscore-prefix); only the 4 apply_*
and add_border functions are public.
"""
from __future__ import annotations

import functools

from ._core import (
    _HEX_COLOR_RE,
    _VALID_BORDER_STYLES,
    ExcelMCPError,
    ValidationError,
    _audit_log,
    _check_confirm,
    _check_path,
    _check_write,
    _flush_if_dirty,
    _load_wb,
    _save_and_evict,
)

__all__ = [
    "apply_number_format",
    "apply_cell_style",
    "apply_alignment",
    "add_border",
]


@functools.lru_cache(maxsize=512)
def _cached_font(
    bold: bool | None,
    italic: bool | None,
    size: float | None,
    color: str | None,
) -> "Font":  # noqa: F821
    from openpyxl.styles import Font
    return Font(bold=bold, italic=italic, size=size, color=color)


@functools.lru_cache(maxsize=512)
def _cached_alignment(
    horizontal: str | None,
    vertical: str | None,
    wrap_text: bool | None,
) -> "Alignment":  # noqa: F821
    from openpyxl.styles import Alignment
    return Alignment(horizontal=horizontal, vertical=vertical, wrap_text=wrap_text)


@functools.lru_cache(maxsize=256)
def _cached_fill(bg_color: str) -> "PatternFill":  # noqa: F821
    from openpyxl.styles import PatternFill
    return PatternFill(fill_type="solid", fgColor=bg_color)


def _validate_hex_color(color: str | None, param_name: str) -> None:
    """Validate a 6-digit hex color string (without '#').

    Raises ExcelMCPError if invalid.
    """
    if color is None:
        return
    if not _HEX_COLOR_RE.match(color):
        raise ExcelMCPError(
            f"Invalid hex color for {param_name!r}: {color!r}. "
            "Use 6 hex digits without '#', e.g. 'FF0000'."
        )


def apply_number_format(
    path: str,
    sheet: str,
    address: str,
    format_code: str,
    confirm: bool = False,
) -> dict:
    """Apply a number format code to a cell or range."""
    _check_write()
    _check_confirm(confirm)
    resolved = _check_path(path)
    _flush_if_dirty(str(resolved))
    wb = _load_wb(resolved, for_write=True)
    if sheet not in wb.sheetnames:
        raise ValidationError(f"Sheet not found: {sheet!r}")
    ws = wb[sheet]
    addr = address.upper().strip()
    if ":" not in addr:
        addr = f"{addr}:{addr}"
    for row in ws[addr]:
        for cell in row:
            cell.number_format = format_code
    _save_and_evict(wb, resolved)
    _audit_log("apply_number_format", resolved, sheet, address.upper(), None)
    return {
        "ok": True,
        "sheet": sheet,
        "address": address,
        "message": f"Applied number format '{format_code}' to {address}",
    }


def apply_cell_style(
    path: str,
    sheet: str,
    address: str,
    bold: bool | None = None,
    italic: bool | None = None,
    font_size: float | None = None,
    font_color: str | None = None,
    bg_color: str | None = None,
    confirm: bool = False,
) -> dict:
    """Apply font and fill styling to a cell or range."""
    _check_write()
    _check_confirm(confirm)
    resolved = _check_path(path)
    _validate_hex_color(font_color, "font_color")
    _validate_hex_color(bg_color, "bg_color")
    _flush_if_dirty(str(resolved))
    wb = _load_wb(resolved, for_write=True)
    if sheet not in wb.sheetnames:
        raise ValidationError(f"Sheet not found: {sheet!r}")
    ws = wb[sheet]
    addr = address.upper().strip()
    if ":" not in addr:
        addr = f"{addr}:{addr}"
    # Hoist constant style objects out of the cell loop for performance.
    # PatternFill: bg_color is range-constant — construct once.
    _fill = _cached_fill(bg_color) if bg_color is not None else None
    # Font: only hoist when ALL font params are specified (none are None).
    # When any param is None we must read the existing per-cell value to merge,
    # so we keep the per-cell path to preserve existing behavior.
    _font = (
        _cached_font(bold=bold, italic=italic, size=font_size, color=font_color)
        if all(v is not None for v in (bold, italic, font_size, font_color))
        else None
    )
    for row in ws[addr]:
        for cell in row:
            if any(v is not None for v in (bold, italic, font_size, font_color)):
                if _font is not None:
                    cell.font = _font
                else:
                    current = cell.font
                    # Preserve existing color when font_color is not specified.
                    # Passing color=None would silently drop any previously set color.
                    existing_color = current.color if current.color is not None else None
                    cell.font = _cached_font(
                        bold=bold if bold is not None else current.bold,
                        italic=italic if italic is not None else current.italic,
                        size=font_size if font_size is not None else current.size,
                        color=font_color if font_color is not None else existing_color,
                    )
            if _fill is not None:
                cell.fill = _fill
    _save_and_evict(wb, resolved)
    _audit_log("apply_cell_style", resolved, sheet, address.upper(), None)
    return {
        "ok": True,
        "sheet": sheet,
        "address": address,
        "message": f"Applied style to {address}",
    }


def apply_alignment(
    path: str,
    sheet: str,
    address: str,
    horizontal: str | None = None,
    vertical: str | None = None,
    wrap_text: bool | None = None,
    confirm: bool = False,
) -> dict:
    """Apply alignment to a cell or range."""
    _VALID_HORIZONTAL = {"left", "center", "right", "fill", "justify"}
    _VALID_VERTICAL = {"top", "center", "bottom"}

    _check_write()
    _check_confirm(confirm)
    resolved = _check_path(path)

    if horizontal is not None and horizontal not in _VALID_HORIZONTAL:
        raise ExcelMCPError(
            f"Invalid horizontal alignment: {horizontal!r}. "
            f"Valid values: {sorted(_VALID_HORIZONTAL)}"
        )
    if vertical is not None and vertical not in _VALID_VERTICAL:
        raise ExcelMCPError(
            f"Invalid vertical alignment: {vertical!r}. "
            f"Valid values: {sorted(_VALID_VERTICAL)}"
        )
    _flush_if_dirty(str(resolved))
    wb = _load_wb(resolved, for_write=True)
    if sheet not in wb.sheetnames:
        raise ValidationError(f"Sheet not found: {sheet!r}")
    ws = wb[sheet]
    addr = address.upper().strip()
    if ":" not in addr:
        addr = f"{addr}:{addr}"
    # Hoist Alignment construction when all params are specified (range-constant,
    # no per-cell merge needed).  When any param is None we must read the existing
    # per-cell alignment value to merge, so keep the per-cell fallback path.
    _alignment = (
        _cached_alignment(horizontal=horizontal, vertical=vertical, wrap_text=wrap_text)
        if all(v is not None for v in (horizontal, vertical, wrap_text))
        else None
    )
    for row in ws[addr]:
        for cell in row:
            if _alignment is not None:
                cell.alignment = _alignment
            else:
                current = cell.alignment
                cell.alignment = _cached_alignment(
                    horizontal=horizontal if horizontal is not None else current.horizontal,
                    vertical=vertical if vertical is not None else current.vertical,
                    wrap_text=wrap_text if wrap_text is not None else current.wrap_text,
                )
    _save_and_evict(wb, resolved)
    _audit_log("apply_alignment", resolved, sheet, address.upper(), None)
    return {
        "ok": True,
        "sheet": sheet,
        "address": address,
        "message": f"Applied alignment to {address}",
    }


def add_border(
    path: str,
    sheet: str,
    address: str,
    border_style: str = "thin",
    color: str = "000000",
    confirm: bool = False,
) -> dict:
    """Apply a border to all sides of a cell or range."""
    from openpyxl.styles import Border, Side

    _check_write()
    _check_confirm(confirm)
    resolved = _check_path(path)

    if border_style not in _VALID_BORDER_STYLES:
        raise ExcelMCPError(
            f"Invalid border_style: {border_style!r}. "
            f"Valid values: {sorted(_VALID_BORDER_STYLES)}"
        )
    _validate_hex_color(color, "color")
    _flush_if_dirty(str(resolved))
    wb = _load_wb(resolved, for_write=True)
    if sheet not in wb.sheetnames:
        raise ValidationError(f"Sheet not found: {sheet!r}")
    ws = wb[sheet]
    side = Side(style=border_style, color=color)
    border = Border(left=side, right=side, top=side, bottom=side)
    addr = address.upper().strip()
    if ":" not in addr:
        addr = f"{addr}:{addr}"
    for row in ws[addr]:
        for cell in row:
            cell.border = border
    _save_and_evict(wb, resolved)
    _audit_log("add_border", resolved, sheet, address.upper(), None)
    return {
        "ok": True,
        "sheet": sheet,
        "address": address,
        "border_style": border_style,
        "message": f"Applied {border_style} border to {address}",
    }
