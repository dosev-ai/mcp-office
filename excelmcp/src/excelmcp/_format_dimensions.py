"""Column and row dimension helpers: widths, heights, visibility.

Internal sub-module. Import via excelmcp._format (public router).
"""
from __future__ import annotations

from openpyxl.cell.cell import MergedCell as _MergedCell
from openpyxl.utils import (
    column_index_from_string as _col_idx,
)
from openpyxl.utils import (
    get_column_letter as _col_letter,
)

from ._core import (
    ExcelMCPError,
    ValidationError,
    _audit_log,
    _check_confirm,
    _check_path,
    _check_write,
    _flush_if_dirty,
    _load_wb,
    _save_and_evict,
)

__all__ = [
    "set_column_width",
    "set_column_visibility",
    "set_row_height",
    "set_row_heights",
    "set_column_widths",
]


def set_column_width(
    path: str,
    sheet: str,
    column: str,
    width: float,
    confirm: bool = False,
) -> dict:
    """Set the width of a column."""
    _check_write()
    _check_confirm(confirm)
    if not (0 <= width <= 255):
        raise ExcelMCPError(f"Column width must be 0-255, got {width}")
    resolved = _check_path(path)
    _flush_if_dirty(str(resolved))
    wb = _load_wb(resolved, for_write=True)
    if sheet not in wb.sheetnames:
        raise ValidationError(f"Sheet not found: {sheet!r}")
    ws = wb[sheet]
    ws.column_dimensions[column.upper()].width = width
    _audit_log("set_column_width", resolved, sheet, column.upper(), None)
    _save_and_evict(wb, resolved)
    return {
        "ok": True,
        "sheet": sheet,
        "column": column,
        "width": width,
        "message": f"Set column {column.upper()} width to {width}",
    }


def set_column_visibility(
    path: str,
    sheet: str,
    columns: list[str],
    hidden: bool,
    confirm: bool = False,
) -> dict:
    """Hide or show one or more columns.

    Args:
        path: Path to the workbook.
        sheet: Worksheet name.
        columns: List of column letters, e.g. ["A", "C", "D"].
        hidden: True to hide, False to show.
        confirm: Must be True to apply changes.
    """
    _check_write()
    _check_confirm(confirm)
    resolved = _check_path(path)
    _flush_if_dirty(str(resolved))
    wb = _load_wb(resolved, for_write=True)
    if sheet not in wb.sheetnames:
        raise ValidationError(f"Sheet not found: {sheet!r}")
    ws = wb[sheet]
    for col in columns:
        col_upper = col.upper().strip()
        if not col_upper.isalpha():
            raise ValidationError(f"Invalid column letter: {col!r}")
        ws.column_dimensions[col_upper].hidden = hidden
    _save_and_evict(wb, resolved)
    _audit_log("set_column_visibility", resolved, sheet, ",".join(columns), None)
    return {"sheet": sheet, "columns": columns, "hidden": hidden}


def set_row_height(
    path: str,
    sheet: str,
    row: int,
    height: float,
    confirm: bool = False,
) -> dict:
    """Set the height of a row."""
    _check_write()
    _check_confirm(confirm)
    if not (0 <= height <= 409):
        raise ExcelMCPError(f"Row height must be 0-409, got {height}")
    if row < 1:
        raise ExcelMCPError(f"Row number must be >= 1, got {row}")
    resolved = _check_path(path)
    _flush_if_dirty(str(resolved))
    wb = _load_wb(resolved, for_write=True)
    if sheet not in wb.sheetnames:
        raise ValidationError(f"Sheet not found: {sheet!r}")
    ws = wb[sheet]
    ws.row_dimensions[row].height = height
    _audit_log("set_row_height", resolved, sheet, f"row{row}", None)
    _save_and_evict(wb, resolved)
    return {
        "ok": True,
        "sheet": sheet,
        "row": row,
        "height": height,
        "message": f"Set row {row} height to {height}",
    }


def set_row_heights(
    path: str,
    sheet: str,
    rows: list[int] | None = None,
    row_range: str | None = None,
    height: float | None = None,
    row_specs: list[dict] | None = None,
    confirm: bool = False,
) -> dict:
    """Set heights for multiple rows in a single call.

    Two mutually exclusive modes:

    **Mode 1** (uniform height): Specify rows via ``rows`` (list of 1-based row
    numbers) OR ``row_range`` (e.g. ``"1:20"``), plus ``height``.
    ``rows`` and ``row_range`` are themselves mutually exclusive within Mode 1.

    **Mode 2** (per-row): Specify ``row_specs`` — each dict
    ``{"row": int, "height": float}``.  Duplicate rows use last-write-wins.

    Valid height range: 0.0–409.0 (Excel limit).

    Returns ``{"ok": True, "sheet": sheet, "rows_set": int}``.
    Requires EXCEL_ENABLE_WRITE=true and confirm=True.
    """
    _check_write()
    _check_confirm(confirm)

    mode1_active = rows is not None or row_range is not None or height is not None
    mode2_active = row_specs is not None

    if mode1_active and mode2_active:
        raise ValidationError(
            "row_specs (Mode 2) and rows/row_range/height (Mode 1) are mutually exclusive"
        )
    if not mode1_active and not mode2_active:
        raise ValidationError(
            "Provide either rows/row_range + height (Mode 1) or row_specs (Mode 2)"
        )

    resolved_rows: list[tuple[int, float]]

    if mode2_active:
        if not row_specs:
            raise ValidationError("row_specs must be non-empty")
        resolved_rows = []
        for item in row_specs:
            r = item.get("row")
            h = item.get("height")
            if not isinstance(r, int) or r < 1:
                raise ValidationError(f"row_specs entry has invalid row: {item!r}")
            if not isinstance(h, (int, float)) or not (0.0 <= float(h) <= 409.0):
                raise ValidationError(f"row_specs entry has invalid height: {item!r}")
            resolved_rows.append((r, float(h)))
    else:
        # Mode 1
        if rows is not None and row_range is not None:
            raise ValidationError("rows and row_range are mutually exclusive in Mode 1")
        if height is None:
            raise ValidationError("height is required in Mode 1")
        if not (0.0 <= height <= 409.0):
            raise ValidationError(f"height must be 0.0–409.0, got {height}")
        if rows is not None:
            if not rows:
                raise ValidationError("rows must be non-empty")
            row_list = rows
        else:
            try:
                parts = row_range.split(":")  # type: ignore[union-attr]
                if len(parts) != 2:
                    raise ValueError
                start_r, end_r = int(parts[0].strip()), int(parts[1].strip())
                if start_r < 1 or end_r < 1 or end_r < start_r:
                    raise ValueError
            except (ValueError, AttributeError) as exc:
                raise ValidationError(
                    f"Invalid row_range {row_range!r}: expected 'start:end' positive ints"
                ) from exc
            row_list = list(range(start_r, end_r + 1))
        resolved_rows = [(r, height) for r in row_list]

    resolved = _check_path(path)
    _flush_if_dirty(str(resolved))
    wb = _load_wb(resolved, for_write=True)
    if sheet not in wb.sheetnames:
        raise ValidationError(f"Sheet not found: {sheet!r}")
    ws = wb[sheet]
    for row_num, h in resolved_rows:
        ws.row_dimensions[row_num].height = h
    _save_and_evict(wb, resolved)
    _audit_log("set_row_heights", resolved, sheet, None, len(resolved_rows))
    return {"ok": True, "sheet": sheet, "rows_set": len(resolved_rows)}


def set_column_widths(
    path: str,
    sheet: str,
    columns: list[str] | None = None,
    col_range: str | None = None,
    width: float | None = None,
    autofit: bool = False,
    autofit_padding: float = 2.0,
    max_width: float = 60.0,
    confirm: bool = False,
) -> dict:
    """Set widths for multiple columns in a single call.

    Exactly one of ``columns`` or ``col_range`` must be provided:

    - ``columns``: explicit list of column letters, e.g. ``["A", "C", "D"]``.
    - ``col_range``: range string, e.g. ``"A:E"`` or ``"AA:AZ"``.

    Exactly one of ``width`` or ``autofit=True`` must be specified:

    - ``width``: explicit width (0–255).
    - ``autofit=True``: approximate column widths based on cell content length.
      Bold cells are scaled by 1.12×.  Wrap-text cells are excluded.
      Width is capped at ``max_width`` (default 60).

    Returns
    ``{"ok": True, "sheet": sheet, "columns_set": int, "widths": [{"column": str, "width": float}]}``.
    Requires EXCEL_ENABLE_WRITE=true and confirm=True.
    """
    _check_write()
    _check_confirm(confirm)

    # Mutual exclusion: exactly one source
    if columns is not None and col_range is not None:
        raise ValidationError("columns and col_range are mutually exclusive")
    if columns is None and col_range is None:
        raise ValidationError("Either columns or col_range must be provided")
    if columns is not None and not columns:
        raise ValidationError("columns must be non-empty")

    # Width vs autofit
    if autofit and width is not None:
        raise ValidationError("autofit=True and width are mutually exclusive")
    if not autofit and width is None:
        raise ValidationError("width is required when autofit=False")
    if width is not None and not (0.0 <= width <= 255.0):
        raise ValidationError(f"width must be 0–255, got {width}")

    # Resolve column list
    if col_range is not None:
        try:
            parts = col_range.upper().split(":")
            if len(parts) != 2:
                raise ValueError
            start_idx = _col_idx(parts[0].strip())
            end_idx = _col_idx(parts[1].strip())
            if end_idx < start_idx:
                raise ValueError
            col_list = [_col_letter(i) for i in range(start_idx, end_idx + 1)]
        except (ValueError, AttributeError) as exc:
            raise ValidationError(
                f"Invalid col_range {col_range!r}: expected 'A:E' format"
            ) from exc
    else:
        col_list = [c.upper().strip() for c in columns]  # type: ignore[union-attr]

    if len(col_list) > 500:
        raise ValidationError(
            f"set_column_widths: max 500 columns per call (got {len(col_list)})"
        )

    resolved = _check_path(path)
    _flush_if_dirty(str(resolved))
    wb_obj = _load_wb(resolved, for_write=True)
    if sheet not in wb_obj.sheetnames:
        raise ValidationError(f"Sheet not found: {sheet!r}")
    ws = wb_obj[sheet]

    widths_out: list[dict] = []
    for col_letter in col_list:
        if autofit:
            best = 0.0
            col_idx_val = _col_idx(col_letter)
            for row in ws.iter_rows():
                cell = row[col_idx_val - 1]
                if isinstance(cell, _MergedCell):
                    continue
                if cell.value is None:
                    continue
                try:
                    if cell.alignment and cell.alignment.wrap_text:
                        continue
                except AttributeError:
                    pass
                cell_len = float(len(str(cell.value)))
                try:
                    if cell.font and cell.font.bold:
                        cell_len *= 1.12
                except AttributeError:
                    pass
                if cell_len > best:
                    best = cell_len
            computed = min(best + autofit_padding, max_width)
            ws.column_dimensions[col_letter].width = computed
            widths_out.append({"column": col_letter, "width": computed})
        else:
            ws.column_dimensions[col_letter].width = width  # type: ignore[arg-type]
            widths_out.append({"column": col_letter, "width": width})

    _save_and_evict(wb_obj, resolved)
    _audit_log("set_column_widths", resolved, sheet, ",".join(col_list[:5]), len(col_list))
    return {
        "ok": True,
        "sheet": sheet,
        "columns_set": len(col_list),
        "widths": widths_out,
    }
