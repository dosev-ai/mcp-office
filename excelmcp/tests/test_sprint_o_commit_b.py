"""
Sprint O Commit B — unit tests for 3 new unified tools:
  sheet(), merge(), data_validation()

Tests cover all operations, guard-gates, and error paths.
No COM required; always CI-runnable.
"""
import pytest
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation

from fastmcp.exceptions import ToolError
from excelmcp.workbook_openpyxl import _wb_cache
from excelmcp.server import sheet, merge, data_validation


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture
def wb_path(tmp_path, monkeypatch):
    """Single-sheet workbook with EXCEL_ENABLE_WRITE=true."""
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    path = tmp_path / "test.xlsx"
    _wb = openpyxl.Workbook()
    _wb.active.title = "Sheet1"
    _wb.save(str(path))
    _wb_cache.clear()
    return str(path)


@pytest.fixture
def wb_two_sheets(tmp_path, monkeypatch):
    """Two-sheet workbook."""
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    path = tmp_path / "two.xlsx"
    _wb = openpyxl.Workbook()
    _wb.active.title = "Sheet1"
    _wb.create_sheet("Sheet2")
    _wb.save(str(path))
    _wb_cache.clear()
    return str(path)


@pytest.fixture
def wb_with_merge(tmp_path, monkeypatch):
    """Workbook with A1:B2 already merged."""
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    path = tmp_path / "merged.xlsx"
    _wb = openpyxl.Workbook()
    ws = _wb.active
    ws.title = "Sheet1"
    ws.merge_cells("A1:B2")
    _wb.save(str(path))
    _wb_cache.clear()
    return str(path)


@pytest.fixture
def wb_with_dv(tmp_path, monkeypatch):
    """Workbook with one data validation rule on Sheet1."""
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    path = tmp_path / "dv.xlsx"
    _wb = openpyxl.Workbook()
    ws = _wb.active
    ws.title = "Sheet1"
    dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    dv.sqref = "A1:A10"
    ws.add_data_validation(dv)
    _wb.save(str(path))
    _wb_cache.clear()
    return str(path)


# ---------------------------------------------------------------------------
# sheet() tests
# ---------------------------------------------------------------------------

@pytest.mark.unit
def test_sheet_list_returns_dict(wb_path):
    """sheet(operation='list') returns {'sheets': [...], 'count': N}."""
    result = sheet(operation="list", path=wb_path)
    assert isinstance(result, dict)
    assert "sheets" in result
    assert "count" in result
    assert "Sheet1" in result["sheets"]
    assert result["count"] == len(result["sheets"])


@pytest.mark.unit
def test_sheet_add_success(wb_path):
    """sheet(operation='add') adds a new sheet."""
    result = sheet(operation="add", path=wb_path, sheet_name="NewSheet", confirm=True)
    assert result["ok"] is True
    assert result["sheet_name"] == "NewSheet"


@pytest.mark.unit
def test_sheet_rename_success(wb_two_sheets):
    """sheet(operation='rename') renames an existing sheet."""
    result = sheet(
        operation="rename", path=wb_two_sheets,
        sheet_name="Sheet2", new_name="Renamed", confirm=True,
    )
    assert result["ok"] is True
    assert result["new_name"] == "Renamed"


@pytest.mark.unit
def test_sheet_delete_success(wb_two_sheets):
    """sheet(operation='delete') removes an existing sheet."""
    result = sheet(
        operation="delete", path=wb_two_sheets,
        sheet_name="Sheet2", confirm=True,
    )
    assert result["ok"] is True
    assert result["sheet_name"] == "Sheet2"


@pytest.mark.unit
def test_sheet_delete_no_confirm_raises(wb_two_sheets):
    """sheet(operation='delete') with confirm=False raises ToolError."""
    with pytest.raises(ToolError):
        sheet(operation="delete", path=wb_two_sheets, sheet_name="Sheet2", confirm=False)


@pytest.mark.unit
def test_sheet_rename_no_new_name_raises(wb_two_sheets):
    """sheet(operation='rename') without new_name raises ToolError."""
    with pytest.raises(ToolError, match="new_name is required"):
        sheet(
            operation="rename", path=wb_two_sheets,
            sheet_name="Sheet2", new_name=None, confirm=True,
        )


@pytest.mark.unit
def test_sheet_add_no_sheet_name_raises(wb_path):
    """sheet(operation='add') without sheet_name raises ToolError."""
    with pytest.raises(ToolError, match="sheet_name is required"):
        sheet(operation="add", path=wb_path, sheet_name=None, confirm=True)


@pytest.mark.unit
def test_sheet_unknown_operation_raises(wb_path):
    """sheet() with unknown operation raises ToolError."""
    with pytest.raises(ToolError, match="Unknown operation"):
        sheet(operation="copy", path=wb_path)  # type: ignore[arg-type]


# ---------------------------------------------------------------------------
# merge() tests
# ---------------------------------------------------------------------------

@pytest.mark.unit
def test_merge_list_returns_dict(wb_with_merge):
    """merge(operation='list') returns dict with merged_ranges and count."""
    result = merge(operation="list", path=wb_with_merge, sheet="Sheet1")
    assert isinstance(result, dict)
    assert "merged_ranges" in result
    assert "count" in result
    assert "A1:B2" in result["merged_ranges"]
    assert result["count"] == len(result["merged_ranges"])


@pytest.mark.unit
def test_merge_merge_success(wb_path):
    """merge(operation='merge') merges cells successfully."""
    result = merge(
        operation="merge", path=wb_path,
        sheet="Sheet1", address="C1:D2", confirm=True,
    )
    assert result["ok"] is True


@pytest.mark.unit
def test_merge_unmerge_success(wb_with_merge):
    """merge(operation='unmerge') unmerges an existing merged range."""
    result = merge(
        operation="unmerge", path=wb_with_merge,
        sheet="Sheet1", address="A1:B2", confirm=True,
    )
    assert result["ok"] is True


@pytest.mark.unit
def test_merge_no_address_raises(wb_path):
    """merge(operation='merge') without address raises ToolError."""
    with pytest.raises(ToolError, match="address is required"):
        merge(operation="merge", path=wb_path, sheet="Sheet1", address=None, confirm=True)


@pytest.mark.unit
def test_merge_no_confirm_raises(wb_path):
    """merge(operation='merge') with confirm=False raises ToolError (write gate)."""
    with pytest.raises(ToolError):
        merge(
            operation="merge", path=wb_path,
            sheet="Sheet1", address="A1:B1", confirm=False,
        )


@pytest.mark.unit
def test_merge_unknown_operation_raises(wb_path):
    """merge() with unknown operation raises ToolError."""
    with pytest.raises(ToolError, match="Unknown operation"):
        merge(operation="split", path=wb_path, sheet="Sheet1")  # type: ignore[arg-type]


# ---------------------------------------------------------------------------
# data_validation() tests
# ---------------------------------------------------------------------------

@pytest.mark.unit
def test_data_validation_get_returns_dict(wb_with_dv):
    """data_validation(operation='get') returns sheet, count, validations."""
    result = data_validation(operation="get", path=wb_with_dv, sheet="Sheet1")
    assert isinstance(result, dict)
    assert result["sheet"] == "Sheet1"
    assert "count" in result
    assert "validations" in result
    assert result["count"] >= 1


@pytest.mark.unit
def test_data_validation_add_success(wb_path):
    """data_validation(operation='add') adds a list validation rule."""
    result = data_validation(
        operation="add", path=wb_path, sheet="Sheet1",
        validations=[{"address": "A1:A5", "validation_type": "list", "formula1": '"Yes,No"'}],
        confirm=True,
    )
    assert result["ok"] is True
    assert result["validations_added"] == 1


@pytest.mark.unit
def test_data_validation_add_no_confirm_raises(wb_path):
    """data_validation(operation='add') with confirm=False raises ToolError."""
    with pytest.raises(ToolError):
        data_validation(
            operation="add", path=wb_path, sheet="Sheet1",
            validations=[{"address": "B1:B5", "validation_type": "list",
                          "formula1": '"X,Y"'}],
            confirm=False,
        )


@pytest.mark.unit
def test_data_validation_add_no_validations_raises(wb_path):
    """data_validation(operation='add') with validations=None raises ToolError."""
    with pytest.raises(ToolError, match="validations list is required"):
        data_validation(
            operation="add", path=wb_path, sheet="Sheet1",
            validations=None, confirm=True,
        )


@pytest.mark.unit
def test_data_validation_add_empty_list_raises(wb_path):
    """data_validation(operation='add') with empty validations list raises ToolError."""
    with pytest.raises(ToolError, match="validations list is required"):
        data_validation(
            operation="add", path=wb_path, sheet="Sheet1",
            validations=[], confirm=True,
        )


@pytest.mark.unit
def test_data_validation_unknown_operation_raises(wb_path):
    """data_validation() with unknown operation raises ToolError."""
    with pytest.raises(ToolError, match="Unknown operation"):
        data_validation(operation="delete", path=wb_path, sheet="Sheet1")  # type: ignore[arg-type]
