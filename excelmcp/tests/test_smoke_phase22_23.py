"""
Smoke tests — Phase 2.2 and Phase 2.3 tools.

Exercises MCP dispatch layer for:
  Phase 2.2: write_named_range, merge_cells, unmerge_cells, freeze_panes, auto_filter
  Phase 2.3: insert_rows, insert_cols, protect_sheet, set_print_area

No COM required; always CI-runnable.
"""
import pytest
import openpyxl as _xl
from excelmcp import server

from ._smoke_helpers import _run_tool


# ---------------------------------------------------------------------------
# Phase 2.2 smoke tests
# ---------------------------------------------------------------------------


@pytest.mark.smoke
def test_smoke_write_named_range(tmp_path, monkeypatch):
    """Smoke: named_range(write) happy path through the MCP layer."""
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))

    wb_path = tmp_path / "nr.xlsx"
    wb = _xl.Workbook()
    wb.active.title = "Sheet1"
    wb.save(str(wb_path))

    result = _run_tool(server.mcp, "named_range", {
        "operation": "write",
        "path": str(wb_path),
        "name": "SalesData",
        "sheet": "Sheet1",
        "address": "A1:C5",
        "confirm": True,
    })
    assert result["ok"] is True


@pytest.mark.smoke
def test_smoke_merge_cells(tmp_path, monkeypatch):
    """Smoke: merge_cells happy path through the MCP layer."""
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))

    wb_path = tmp_path / "merge.xlsx"
    wb = _xl.Workbook()
    wb.active.title = "Sheet1"
    wb.save(str(wb_path))

    result = _run_tool(server.mcp, "merge", {
        "operation": "merge",
        "path": str(wb_path),
        "sheet": "Sheet1",
        "address": "B1:C2",
        "confirm": True,
    })
    assert result["ok"] is True


@pytest.mark.smoke
def test_smoke_unmerge_cells(tmp_path, monkeypatch):
    """Smoke: unmerge_cells after merging — happy path through the MCP layer."""
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))

    wb_path = tmp_path / "unmerge.xlsx"
    wb = _xl.Workbook()
    wb.active.title = "Sheet1"
    wb.save(str(wb_path))

    _run_tool(server.mcp, "merge", {
        "operation": "merge",
        "path": str(wb_path), "sheet": "Sheet1", "address": "D1:E1", "confirm": True,
    })
    result = _run_tool(server.mcp, "merge", {
        "operation": "unmerge",
        "path": str(wb_path),
        "sheet": "Sheet1",
        "address": "D1:E1",
        "confirm": True,
    })
    assert result["ok"] is True


@pytest.mark.smoke
def test_smoke_freeze_panes(tmp_path, monkeypatch):
    """Smoke: freeze_panes happy path through the MCP layer."""
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))

    wb_path = tmp_path / "freeze.xlsx"
    wb = _xl.Workbook()
    wb.active.title = "Sheet1"
    wb.save(str(wb_path))

    result = _run_tool(server.mcp, "freeze_panes", {
        "path": str(wb_path),
        "sheet": "Sheet1",
        "address": "B2",
        "confirm": True,
    })
    assert result["ok"] is True


@pytest.mark.smoke
def test_smoke_auto_filter(tmp_path, monkeypatch):
    """Smoke: auto_filter happy path through the MCP layer."""
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))

    wb_path = tmp_path / "filter.xlsx"
    wb = _xl.Workbook()
    wb.active.title = "Sheet1"
    wb.save(str(wb_path))

    result = _run_tool(server.mcp, "auto_filter", {
        "path": str(wb_path),
        "sheet": "Sheet1",
        "address": "A1:D1",
        "confirm": True,
    })
    assert result["ok"] is True


# ---------------------------------------------------------------------------
# Phase 2.3 smoke tests
# ---------------------------------------------------------------------------


@pytest.mark.smoke
def test_smoke_insert_rows(tmp_path, monkeypatch):
    """Smoke: insert_rows happy path through the MCP layer."""
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))

    wb_path = tmp_path / "ins_rows.xlsx"
    wb = _xl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "header"
    ws["A2"] = "data"
    wb.save(str(wb_path))

    result = _run_tool(server.mcp, "rows", {
        "operation": "insert",
        "path": str(wb_path),
        "sheet": "Sheet1",
        "start_row": 2,
        "count": 1,
        "confirm": True,
    })
    assert result["ok"] is True


@pytest.mark.smoke
def test_smoke_insert_cols(tmp_path, monkeypatch):
    """Smoke: insert_cols happy path through the MCP layer."""
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))

    wb_path = tmp_path / "ins_cols.xlsx"
    wb = _xl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "col_a"
    ws["B1"] = "col_b"
    wb.save(str(wb_path))

    result = _run_tool(server.mcp, "cols", {
        "operation": "insert",
        "path": str(wb_path),
        "sheet": "Sheet1",
        "start_col": 2,
        "count": 1,
        "confirm": True,
    })
    assert result["ok"] is True


@pytest.mark.smoke
def test_smoke_protect_sheet(tmp_path, monkeypatch):
    """Smoke: protect_sheet happy path through the MCP layer."""
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))

    wb_path = tmp_path / "protect.xlsx"
    wb = _xl.Workbook()
    wb.active.title = "Sheet1"
    wb.save(str(wb_path))

    result = _run_tool(server.mcp, "protect", {
        "scope": "sheet",
        "path": str(wb_path),
        "sheet": "Sheet1",
        "password": "testpw",
        "confirm": True,
    })
    assert result["ok"] is True


@pytest.mark.smoke
def test_smoke_set_print_area(tmp_path, monkeypatch):
    """Smoke: set_print_area happy path through the MCP layer."""
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))

    wb_path = tmp_path / "print_area.xlsx"
    wb = _xl.Workbook()
    wb.active.title = "Sheet1"
    wb.save(str(wb_path))

    result = _run_tool(server.mcp, "set_print_area", {
        "path": str(wb_path),
        "sheet": "Sheet1",
        "address": "A1:F20",
        "confirm": True,
    })
    assert result["ok"] is True
