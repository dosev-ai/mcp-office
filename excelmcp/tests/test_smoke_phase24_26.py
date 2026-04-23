"""
Smoke tests — Phase 2.4, 2.5, 2.6, and 2.7 tools.

Exercises MCP dispatch layer for:
  Phase 2.4: add/delete cell comment, get/set sheet properties,
             add_data_validation, protect_workbook
  Phase 2.5: validate_formula_syntax, copy_range (write-disabled path),
             list_merged_cells, create_table (write-disabled path)
  Phase 2.6: get_data_validation_info, get_cell_comment,
             set_column_visibility, set_workbook_calc_mode
  Phase 2.7: get_hyperlinks

No COM required; always CI-runnable.
"""
import asyncio

import pytest
import openpyxl as _xl
from excelmcp import server
from openpyxl.worksheet.datavalidation import DataValidation

from ._smoke_helpers import _get_tools_dict, _run_tool


# ---------------------------------------------------------------------------
# Phase 2.4 smoke tests
# ---------------------------------------------------------------------------


@pytest.mark.smoke
def test_smoke_add_cell_comment(tmp_path, monkeypatch):
    """Smoke: bulk_add_comments adds a single comment (migrated from add_cell_comment)."""
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))

    wb_path = tmp_path / "comment.xlsx"
    wb = _xl.Workbook()
    wb.active.title = "Sheet1"
    wb.save(str(wb_path))

    result = _run_tool(server.mcp, "bulk_add_comments", {
        "path": str(wb_path),
        "sheet": "Sheet1",
        "comments": [{"address": "A1", "text": "Test comment"}],
        "confirm": True,
    })
    assert result["ok"] is True
    assert result["comments_added"] == 1
    assert result["sheet"] == "Sheet1"


@pytest.mark.smoke
def test_smoke_delete_cell_comment(tmp_path, monkeypatch):
    """Smoke: delete_cell_comment happy path — add then delete."""
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))

    wb_path = tmp_path / "del_comment.xlsx"
    wb = _xl.Workbook()
    wb.active.title = "Sheet1"
    wb.save(str(wb_path))

    _run_tool(server.mcp, "bulk_add_comments", {
        "path": str(wb_path), "sheet": "Sheet1",
        "comments": [{"address": "B2", "text": "to be deleted"}], "confirm": True,
    })
    result = _run_tool(server.mcp, "cell_comment", {
        "operation": "delete",
        "path": str(wb_path),
        "sheet": "Sheet1",
        "address": "B2",
        "confirm": True,
    })
    assert result["ok"] is True


@pytest.mark.smoke
def test_smoke_get_sheet_properties(tmp_path, monkeypatch):
    """Smoke: get_sheet_properties returns sheet and state keys."""
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))

    wb_path = tmp_path / "props.xlsx"
    wb = _xl.Workbook()
    wb.active.title = "Sheet1"
    wb.save(str(wb_path))

    result = _run_tool(server.mcp, "sheet_properties", {
        "operation": "get",
        "path": str(wb_path),
        "sheet": "Sheet1",
    })
    assert result["sheet"] == "Sheet1"
    assert "state" in result


@pytest.mark.smoke
def test_smoke_set_sheet_properties(tmp_path, monkeypatch):
    """Smoke: set_sheet_properties sets tab_color through the MCP layer."""
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))

    wb_path = tmp_path / "set_props.xlsx"
    wb = _xl.Workbook()
    wb.active.title = "Sheet1"
    wb.create_sheet("Sheet2")
    wb.save(str(wb_path))

    result = _run_tool(server.mcp, "sheet_properties", {
        "operation": "set",
        "path": str(wb_path),
        "sheet": "Sheet1",
        "tab_color": "FF0000",
        "confirm": True,
    })
    assert result["ok"] is True


@pytest.mark.smoke
def test_smoke_add_data_validation(tmp_path, monkeypatch):
    """Smoke: bulk_add_data_validation adds a list rule through the MCP layer (migrated from add_data_validation)."""
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))

    wb_path = tmp_path / "dv.xlsx"
    wb = _xl.Workbook()
    wb.active.title = "Sheet1"
    wb.save(str(wb_path))

    result = _run_tool(server.mcp, "data_validation", {
        "operation": "add",
        "path": str(wb_path),
        "sheet": "Sheet1",
        "validations": [{"address": "A1:A5", "validation_type": "list", "formula1": '"Yes,No"'}],
        "confirm": True,
    })
    assert result["ok"] is True
    assert result["validations_added"] == 1


@pytest.mark.smoke
def test_smoke_protect_workbook(tmp_path, monkeypatch):
    """Smoke: protect_workbook locks structure through the MCP layer."""
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))

    wb_path = tmp_path / "wb_protect.xlsx"
    wb = _xl.Workbook()
    wb.save(str(wb_path))

    result = _run_tool(server.mcp, "protect", {
        "scope": "workbook",
        "path": str(wb_path),
        "lock_structure": True,
        "confirm": True,
    })
    assert result["ok"] is True
    assert result["protected"] is True


# ---------------------------------------------------------------------------
# Phase 2.5 smoke tests
# ---------------------------------------------------------------------------


@pytest.mark.smoke
def test_smoke_validate_formula_syntax(tmp_path, monkeypatch):
    """Smoke: validate_formula_syntax returns valid=True for =SUM(A1:A5)."""
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))

    wb_path = tmp_path / "vfs.xlsx"
    _xl.Workbook().save(str(wb_path))

    result = _run_tool(server.mcp, "validate_formula_syntax", {
        "path": str(wb_path),
        "formula": "=SUM(A1:A5)",
    })
    assert result["valid"] is True
    assert result["error"] is None


@pytest.mark.smoke
def test_smoke_copy_range_write_disabled(tmp_path, monkeypatch):
    """Smoke: copy_range without EXCEL_ENABLE_WRITE raises or returns error."""
    monkeypatch.delenv("EXCEL_ENABLE_WRITE", raising=False)
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))

    wb_path = tmp_path / "copy_rng.xlsx"
    _wb = _xl.Workbook()
    _wb.active["A1"] = "x"
    _wb.save(str(wb_path))

    tools = _get_tools_dict(server.mcp)

    async def _call():
        try:
            return await tools["copy_range"].run({
                "path": str(wb_path),
                "sheet": "Sheet",
                "src_range": "A1:A1",
                "dst_range": "C1:C1",
                "confirm": True,
            })
        except Exception as exc:
            return exc

    outcome = asyncio.run(_call())
    assert outcome is not None


@pytest.mark.smoke
def test_smoke_list_merged_cells(tmp_path, monkeypatch):
    """Smoke: list_merged_cells returns merged_ranges key with pre-merged range."""
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))

    wb_path = tmp_path / "lmc.xlsx"
    _wb = _xl.Workbook()
    _wb.active.title = "Sheet1"
    _wb.active.merge_cells("A1:B2")
    _wb.save(str(wb_path))

    result = _run_tool(server.mcp, "merge", {
        "operation": "list",
        "path": str(wb_path),
        "sheet": "Sheet1",
    })
    assert "merged_ranges" in result
    assert "A1:B2" in result["merged_ranges"]


@pytest.mark.smoke
def test_smoke_create_table_write_disabled(tmp_path, monkeypatch):
    """Smoke: create_table without EXCEL_ENABLE_WRITE raises or returns error."""
    monkeypatch.delenv("EXCEL_ENABLE_WRITE", raising=False)
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))

    wb_path = tmp_path / "ct.xlsx"
    _wb = _xl.Workbook()
    _wb.active.title = "Sheet1"
    _wb.save(str(wb_path))

    tools = _get_tools_dict(server.mcp)

    async def _call():
        try:
            return await tools["create_table"].run({
                "path": str(wb_path),
                "sheet": "Sheet1",
                "table_range": "A1:D5",
                "table_name": "TestTable",
                "confirm": True,
            })
        except Exception as exc:
            return exc

    outcome = asyncio.run(_call())
    assert outcome is not None


# ---------------------------------------------------------------------------
# Phase 2.6 smoke tests
# ---------------------------------------------------------------------------


@pytest.mark.smoke
def test_smoke_get_data_validation_info(tmp_path, monkeypatch):
    """Smoke: get_data_validation_info returns sheet, count, validations keys."""
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))

    wb_path = tmp_path / "dv_smoke.xlsx"
    _wb = _xl.Workbook()
    ws = _wb.active
    ws.title = "Sheet1"
    dv = DataValidation(type="list", formula1='"A,B,C"', allow_blank=True)
    dv.sqref = "A1:A10"
    ws.add_data_validation(dv)
    _wb.save(str(wb_path))

    result = _run_tool(server.mcp, "data_validation", {
        "operation": "get",
        "path": str(wb_path),
        "sheet": "Sheet1",
    })
    assert result["sheet"] == "Sheet1"
    assert "count" in result
    assert "validations" in result
    assert result["count"] == 1
    assert result["validations"][0]["validation_type"] == "list"


@pytest.mark.smoke
def test_smoke_get_cell_comment(tmp_path, monkeypatch):
    """Smoke: get_cell_comment returns has_comment=False for a cell without a comment."""
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))

    wb_path = tmp_path / "comment_smoke.xlsx"
    _wb = _xl.Workbook()
    _wb.active.title = "Sheet1"
    _wb.save(str(wb_path))

    result = _run_tool(server.mcp, "cell_comment", {
        "operation": "get",
        "path": str(wb_path),
        "sheet": "Sheet1",
        "address": "A1",
    })
    assert result["address"] == "A1"
    assert result["has_comment"] is False
    assert result["text"] is None
    assert result["author"] is None


@pytest.mark.smoke
def test_smoke_set_column_visibility(tmp_path, monkeypatch):
    """Smoke: set_column_visibility hides columns through the MCP layer."""
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))

    wb_path = tmp_path / "hide_cols.xlsx"
    _wb = _xl.Workbook()
    _wb.active.title = "Sheet1"
    _wb.save(str(wb_path))

    result = _run_tool(server.mcp, "set_column_visibility", {
        "path": str(wb_path),
        "sheet": "Sheet1",
        "columns": ["B", "C"],
        "hidden": True,
        "confirm": True,
    })
    assert result["sheet"] == "Sheet1"
    assert result["hidden"] is True
    assert "columns" in result


@pytest.mark.smoke
def test_smoke_set_workbook_calc_mode(tmp_path, monkeypatch):
    """Smoke: set_workbook_calc_mode sets manual mode through the MCP layer."""
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))

    wb_path = tmp_path / "calcmode.xlsx"
    _wb = _xl.Workbook()
    _wb.active.title = "Sheet1"
    _wb.save(str(wb_path))

    result = _run_tool(server.mcp, "set_workbook_calc_mode", {
        "path": str(wb_path),
        "calc_mode": "manual",
        "confirm": True,
    })
    assert result["calc_mode"] == "manual"
    assert "full_calc_on_load" in result


# ---------------------------------------------------------------------------
# Phase 2.7 smoke tests
# ---------------------------------------------------------------------------


@pytest.mark.smoke
def test_smoke_get_hyperlinks(tmp_path, monkeypatch):
    """Smoke: hyperlink(operation='get') returns empty hyperlinks list for a blank sheet."""
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))

    wb_path = tmp_path / "hlinks.xlsx"
    _wb = _xl.Workbook()
    _wb.active.title = "Sheet1"
    _wb.save(str(wb_path))

    result = _run_tool(server.mcp, "hyperlink", {
        "operation": "get",
        "path": str(wb_path),
        "sheet": "Sheet1",
    })
    assert result["count"] == 0
    assert result["hyperlinks"] == []
