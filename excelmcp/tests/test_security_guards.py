"""
Regression tests for confirmed security bugs H-001 through H-009.

H-001: Formula injection in write_range_with_format
H-002: Formula injection in import_csv_to_sheet
H-003: Formula injection in display_text (add_hyperlink, bulk_add_hyperlinks)
H-004: File-size pre-flight check in _check_path()
H-007: protect_sheet(password=None) must clear protection
"""
from __future__ import annotations

import csv
from pathlib import Path
from unittest.mock import MagicMock, patch

import openpyxl
import pytest

from excelmcp._core import ValidationError, _check_path, _check_csv_path
from excelmcp._data import (
    add_hyperlink,
    bulk_add_hyperlinks,
    write_range_with_format,
)
from excelmcp._sheet import import_csv_to_sheet
from excelmcp._advanced import protect_sheet


# ---------------------------------------------------------------------------
# Shared fixtures
# ---------------------------------------------------------------------------

@pytest.fixture
def wb_path(tmp_path, monkeypatch):
    """Minimal workbook in a writable allowlisted tmp dir."""
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    p = tmp_path / "test.xlsx"
    wb = openpyxl.Workbook()
    wb.active.title = "Sheet1"
    wb.save(str(p))
    return str(p)


@pytest.fixture
def csv_path(tmp_path, monkeypatch):
    """Return a helper that writes a CSV in the allowlisted dir and returns its str path."""
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")

    wb = openpyxl.Workbook()
    wb.active.title = "Sheet1"
    p = tmp_path / "import_target.xlsx"
    wb.save(str(p))

    def _make_csv(rows, name="data.csv"):
        csv_file = tmp_path / name
        with open(str(csv_file), "w", newline="", encoding="utf-8") as fh:
            writer = csv.writer(fh)
            writer.writerows(rows)
        return str(csv_file), str(p)

    return _make_csv


# ---------------------------------------------------------------------------
# H-001: write_range_with_format formula injection
# ---------------------------------------------------------------------------

@pytest.mark.parametrize("bad_val", ["=SUM(A1:A10)", "+cmd|' /C calc'!A0", "-2+3", "@SUM(1)"])
def test_h001_write_range_with_format_blocks_formula(wb_path, bad_val):
    """write_range_with_format must reject cell values starting with formula prefixes."""
    with pytest.raises(ValidationError, match="Formula injection blocked"):
        write_range_with_format(
            path=wb_path,
            sheet="Sheet1",
            address="A1",
            data=[[bad_val]],
            confirm=True,
        )


def test_h001_write_range_with_format_allows_plain_string(wb_path):
    """write_range_with_format must allow plain string values."""
    result = write_range_with_format(
        path=wb_path,
        sheet="Sheet1",
        address="A1",
        data=[["hello", "world"], [1, 2]],
        confirm=True,
    )
    assert result["cells_written"] == 4


# ---------------------------------------------------------------------------
# H-002: import_csv_to_sheet formula injection
# ---------------------------------------------------------------------------

@pytest.mark.parametrize("bad_val", ["=HYPERLINK(\"http://evil.com\")", "+A1+B1", "-1+2", "@IMPORTDATA"])
def test_h002_import_csv_blocks_formula(csv_path, bad_val):
    """import_csv_to_sheet must reject rows containing formula-prefix values."""
    csv_file, wb_file = csv_path([["name", "value"], ["normal", bad_val]])
    with pytest.raises(ValidationError, match="Formula injection blocked"):
        import_csv_to_sheet(
            path=wb_file,
            csv_path=csv_file,
            sheet_name="Imported",
            confirm=True,
        )


def test_h002_import_csv_allows_clean_data(csv_path):
    """import_csv_to_sheet must allow CSV files with clean data."""
    csv_file, wb_file = csv_path([["name", "value"], ["Alice", "100"], ["Bob", "200"]])
    result = import_csv_to_sheet(
        path=wb_file,
        csv_path=csv_file,
        sheet_name="Imported",
        confirm=True,
    )
    assert result["ok"] is True
    assert result["rows_imported"] == 3


# ---------------------------------------------------------------------------
# H-003a: add_hyperlink display_text formula injection
# ---------------------------------------------------------------------------

@pytest.mark.parametrize("bad_dt", ["=SUM(A1)", "+evil", "-bad", "@cmd"])
def test_h003a_add_hyperlink_blocks_formula_display_text(wb_path, bad_dt):
    """add_hyperlink must reject display_text starting with formula prefixes."""
    with pytest.raises(ValidationError, match="Formula injection blocked in display_text"):
        add_hyperlink(
            path=wb_path,
            sheet="Sheet1",
            address="A1",
            url="https://example.com",
            display_text=bad_dt,
            confirm=True,
        )


def test_h003a_add_hyperlink_allows_clean_display_text(wb_path):
    """add_hyperlink must allow plain display_text."""
    result = add_hyperlink(
        path=wb_path,
        sheet="Sheet1",
        address="B1",
        url="https://example.com",
        display_text="Click here",
        confirm=True,
    )
    assert result["ok"] is True


# ---------------------------------------------------------------------------
# H-003b: bulk_add_hyperlinks display_text formula injection
# ---------------------------------------------------------------------------

@pytest.mark.parametrize("bad_dt", ["=EVIL()", "+exploit", "-inject", "@trigger"])
def test_h003b_bulk_add_hyperlinks_blocks_formula_display_text(wb_path, bad_dt):
    """bulk_add_hyperlinks must reject any item whose display_text starts with formula prefixes."""
    links = [
        {"address": "A1", "url": "https://example.com", "display_text": "safe"},
        {"address": "A2", "url": "https://example.com", "display_text": bad_dt},
    ]
    with pytest.raises(ValidationError, match="Formula injection blocked in display_text"):
        bulk_add_hyperlinks(
            path=wb_path,
            sheet="Sheet1",
            links=links,
            confirm=True,
        )


def test_h003b_bulk_add_hyperlinks_allows_clean(wb_path):
    """bulk_add_hyperlinks must allow items with clean display_text."""
    links = [
        {"address": "A1", "url": "https://example.com", "display_text": "Link 1"},
        {"address": "A2", "url": "https://example.com", "display_text": "Link 2"},
    ]
    result = bulk_add_hyperlinks(path=wb_path, sheet="Sheet1", links=links, confirm=True)
    assert result["ok"] is True
    assert result["links_added"] == 2


# ---------------------------------------------------------------------------
# H-004: _check_path() file-size pre-flight
# ---------------------------------------------------------------------------

def test_h004_check_path_rejects_oversized_file(tmp_path, monkeypatch):
    """_check_path() must raise ValidationError when file exceeds EXCEL_MAX_FILE_MB."""
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    monkeypatch.setenv("EXCEL_MAX_FILE_MB", "1")

    p = tmp_path / "big.xlsx"
    wb = openpyxl.Workbook()
    wb.save(str(p))

    # Patch stat to report a 2 MB file
    fake_stat = MagicMock()
    fake_stat.st_size = 2 * 1024 * 1024  # 2 MB > 1 MB limit

    with patch.object(Path, "stat", return_value=fake_stat):
        with patch.object(Path, "exists", return_value=True):
            with pytest.raises(ValidationError, match="exceeds maximum allowed size"):
                _check_path(str(p))


def test_h004_check_path_allows_small_file(tmp_path, monkeypatch):
    """_check_path() must not raise for a file within the size limit."""
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    monkeypatch.setenv("EXCEL_MAX_FILE_MB", "50")

    p = tmp_path / "small.xlsx"
    wb = openpyxl.Workbook()
    wb.save(str(p))

    result = _check_path(str(p))
    assert result == p.resolve()


# ---------------------------------------------------------------------------
# H-007: protect_sheet(password=None) clears protection
# ---------------------------------------------------------------------------

def test_h007_protect_sheet_clears_protection_when_password_none(tmp_path, monkeypatch):
    """protect_sheet(password=None) must set ws.protection.sheet = False."""
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")

    p = tmp_path / "protected.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    # First, protect it
    ws.protection.sheet = True
    ws.protection.set_password("secret")
    wb.save(str(p))

    result = protect_sheet(path=str(p), sheet="Sheet1", password=None, confirm=True)
    assert result["ok"] is True
    assert result["protected"] is False
    assert "cleared" in result["message"].lower()

    # Verify on disk
    wb2 = openpyxl.load_workbook(str(p))
    assert wb2["Sheet1"].protection.sheet is False


def test_h007_protect_sheet_sets_protection_with_password(tmp_path, monkeypatch):
    """protect_sheet(password='pw') must set ws.protection.sheet = True."""
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")

    p = tmp_path / "unprotected.xlsx"
    wb = openpyxl.Workbook()
    wb.active.title = "Sheet1"
    wb.save(str(p))

    result = protect_sheet(path=str(p), sheet="Sheet1", password="pw", confirm=True)
    assert result["ok"] is True
    assert result["protected"] is True
    assert "protected" in result["message"].lower()


# ---------------------------------------------------------------------------
# R-001 / R-004: import_csv_to_sheet must allow negative numbers
# ---------------------------------------------------------------------------

def test_r001_import_csv_allows_negative_numbers(csv_path):
    """R-001 regression — '-500' and '-1.5' must NOT be blocked as formula injection."""
    csv_file, wb_file = csv_path([["amount"], ["-500"], ["-1.5"], ["+7.0"]])
    result = import_csv_to_sheet(
        path=wb_file,
        csv_path=csv_file,
        sheet_name="Numbers",
        confirm=True,
    )
    assert result.get("ok") is True
    assert result.get("rows_imported") == 4  # header + 3 data rows


def test_r001_import_csv_still_blocks_dde_injection(csv_path):
    """R-001 regression — DDE-style '-1+2+cmd|...' IS still formula injection and must be blocked."""
    csv_file, wb_file = csv_path([["cmd"], ["-1+2+cmd|'/c calc'!K8"]])
    with pytest.raises(ValidationError, match="Formula injection blocked"):
        import_csv_to_sheet(
            path=wb_file,
            csv_path=csv_file,
            sheet_name="Injected",
            confirm=True,
        )


# ---------------------------------------------------------------------------
# Q3 gap: evaluate_formula try/finally closes both workbooks on exception
# ---------------------------------------------------------------------------

def test_evaluate_formula_finally_closes_workbooks_on_exception(tmp_path, monkeypatch):
    """evaluate_formula finally block must close both workbook handles even when the try body raises.

    Regression guard for resource-leak: openpyxl workbooks opened directly (not
    via the cache) must be .close()-d even when an exception interrupts the return.
    """
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))

    p = tmp_path / "ef_close_test.xlsx"
    wb = openpyxl.Workbook()
    wb.active.title = "Sheet1"
    wb.save(str(p))

    # Prime the workbook cache so _load_wb() inside evaluate_formula() is a cache
    # hit — this means openpyxl.load_workbook is only called twice (the two direct
    # calls inside evaluate_formula), not three times.
    from excelmcp._core import _load_wb
    resolved = _check_path(str(p))
    _load_wb(resolved)

    mock_wbf = MagicMock(name="wb_formula")
    mock_wbd = MagicMock(name="wb_data")
    # Make the sheet access raise inside the try block
    mock_wbf.__getitem__ = MagicMock(side_effect=RuntimeError("simulated cell read error"))

    from excelmcp._data import evaluate_formula as _ef
    with patch.object(openpyxl, "load_workbook", side_effect=[mock_wbf, mock_wbd]):
        with pytest.raises(RuntimeError, match="simulated cell read error"):
            _ef(str(p), "Sheet1", "A1")

    mock_wbf.close.assert_called_once()
    mock_wbd.close.assert_called_once()


# ---------------------------------------------------------------------------
# Q4 gap: write_range_with_format blocks plain negative number strings
# (unlike CSV import which allows numeric negatives via float() bypass)
# ---------------------------------------------------------------------------

def test_h001_write_range_with_format_blocks_plain_negative_number_string(wb_path):
    """'-500' must be blocked in write_range_with_format.

    The guard here is intentionally simpler than CSV import: '-' is always
    blocked because callers control what goes into cells — numeric values should
    be passed as int/float, not as negative-prefixed strings.
    """
    with pytest.raises(ValidationError, match="Formula injection blocked"):
        write_range_with_format(
            path=wb_path,
            sheet="Sheet1",
            address="A1",
            data=[["-500"]],
            confirm=True,
        )


def test_h001_write_range_with_format_allows_numeric_types(wb_path):
    """Negative int and float values (not strings) must be allowed in write_range_with_format."""
    result = write_range_with_format(
        path=wb_path,
        sheet="Sheet1",
        address="A1",
        data=[[-500, -1.5, 0, 42]],
        confirm=True,
    )
    assert result["cells_written"] == 4


# ---------------------------------------------------------------------------
# Q5 gap: bulk_add_hyperlinks pre-validation prevents partial write
# (item 1 safe + item 2 formula → neither cell should be written)
# ---------------------------------------------------------------------------

def test_h003b_bulk_add_hyperlinks_prevalidation_no_partial_write(wb_path):
    """After prevalidation failure on item 2, item 1's cell must remain unmodified.

    The pre-validation loop runs BEFORE _load_wb(), so a ValidationError on
    item 2 must leave the workbook entirely untouched — item 1's hyperlink must
    not be present on disk.
    """
    links = [
        {"address": "A1", "url": "https://example.com", "display_text": "Safe text"},
        {"address": "A2", "url": "https://example.com", "display_text": "=EVIL()"},
    ]
    with pytest.raises(ValidationError, match="Formula injection blocked in display_text"):
        bulk_add_hyperlinks(path=wb_path, sheet="Sheet1", links=links, confirm=True)

    # Reload from disk — A1 must have no hyperlink (no partial write occurred)
    wb_loaded = openpyxl.load_workbook(wb_path)
    ws = wb_loaded["Sheet1"]
    assert ws["A1"].hyperlink is None, (
        "Pre-validation failure must leave the workbook unmodified: "
        "item 1 hyperlink must NOT be written when item 2 fails validation"
    )


# ---------------------------------------------------------------------------
# NT-002: _check_csv_path null-byte guard
# ---------------------------------------------------------------------------

class TestCheckCsvPathNullByte:
    """NT-002: _check_csv_path must raise ValidationError (not raw ValueError) for null bytes."""

    def test_null_byte_raises_validation_error(self):
        """NT-002: ValidationError is raised with a message containing 'null byte'."""
        with pytest.raises(ValidationError, match="null byte"):
            _check_csv_path("good_path\x00evil.csv")

    def test_null_byte_not_raw_value_error(self):
        """NT-002: ensure Python's raw ValueError from Path.resolve() does not escape.
        If the guard is absent, Path(...\x00...).resolve() raises ValueError directly.
        With the guard, ValidationError MUST be raised first.
        """
        with pytest.raises(Exception) as exc_info:
            _check_csv_path("good\x00evil.csv")
        assert exc_info.type.__name__ == "ValidationError", (
            f"Expected ValidationError, got {exc_info.type.__name__}. "
            "Null byte guard may be missing or ordered after Path.resolve()."
        )


# ---------------------------------------------------------------------------
# NT-005 + R-005: UNC path rejection in _check_path and _check_csv_path
# ---------------------------------------------------------------------------

class TestUncPathRejection:
    """NT-005 / R-005: UNC path guards in _check_path and _check_csv_path."""

    def test_check_path_rejects_backslash_unc(self):
        """NT-005: _check_path must reject \\\\server\\share style UNC paths."""
        with pytest.raises(ValidationError, match="UNC"):
            _check_path("\\\\server\\share\\evil.xlsx")

    def test_check_csv_path_rejects_backslash_unc(self):
        """NT-005: _check_csv_path must reject \\\\server\\share style UNC paths."""
        with pytest.raises(ValidationError, match="UNC"):
            _check_csv_path("\\\\server\\share\\data.csv")

    def test_check_path_forward_slash_unc_allowlist_backstop(self, monkeypatch, tmp_path):
        """R-005 (W2-005): forward-slash UNC //server/share is caught by the startswith('//')
        UNC guard in _check_path (_core.py: startswith('//') branch), which fires
        before the allowlist check. This test guards against regression of the // prefix
        check in the UNC guard.
        """
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        with pytest.raises(ValidationError):
            _check_path("//server/share/evil.xlsx")

    def test_check_csv_path_forward_slash_unc(self, monkeypatch):
        """NT-005-b: forward-slash UNC rejected in _check_csv_path."""
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", "C:\\Temp")
        with pytest.raises(ValidationError, match="UNC"):
            _check_csv_path("//server/share/data.csv")

