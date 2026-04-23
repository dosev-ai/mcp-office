import pytest
import openpyxl
from excelmcp.workbook_openpyxl import _wb_cache, fill_formula_range, save


class TestDynamicArrayMetadata:
    """Regression tests for the post-save cm/metadata.xml patch.

    openpyxl omits the ``cm="1"`` cell attribute and ``xl/metadata.xml``
    XLDAPR record required by Excel 365 for LET/FILTER/SORT/UNIQUE spill
    formulas.  Without these Excel's repair dialog removes every dynamic-array
    formula on first open.  Covered by ``_ooxml.patch_dynamic_array_metadata``,
    called automatically from ``save()``.

    This file covers methods 1-12 (core patch + prefix rewrite tests).
    Continuation in test_da_finalize.py (methods 13-22 + save/flush paths).
    """

    @pytest.mark.unit
    def test_let_formula_gets_cm_attribute(self, tmp_path, monkeypatch):
        """write_cell + save for a LET formula must produce cm='1' on the cell.

        NOTE: Excel 365 dynamic-array formulas do NOT need t="array" on <f>.
        Only cm="1" + _xlfn. prefixes + metadata.xml are required.
        Adding t="array" causes Excel to treat the formula as a legacy CSE
        array formula and remove it when the result spills to multiple cells.
        """
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        wb_path = str(tmp_path / "da_patch.xlsx")
        _wb = openpyxl.Workbook()
        _wb.active.title = "Sheet1"
        _wb.save(wb_path)
        _wb_cache.clear()

        fill_formula_range(wb_path, "Sheet1", "A1", "A1", "=LET(x,42,x)", confirm=True)
        save(wb_path, confirm=True)

        import zipfile
        with zipfile.ZipFile(wb_path) as z:
            sheet_xml = z.read("xl/worksheets/sheet1.xml").decode("utf-8")
        assert 'cm="1"' in sheet_xml, "cm='1' attribute not found on any cell after save"
        assert 'LET(' in sheet_xml, "LET formula not found in sheet XML after save"

    @pytest.mark.unit
    def test_filter_formula_creates_metadata_xml(self, tmp_path, monkeypatch):
        """save must generate xl/metadata.xml containing XLDAPR when a DA formula is present."""
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        wb_path = str(tmp_path / "da_meta.xlsx")
        _wb = openpyxl.Workbook()
        _wb.active.title = "Sheet1"
        _wb.save(wb_path)
        _wb_cache.clear()

        fill_formula_range(wb_path, "Sheet1", "B2", "B2", "=FILTER(A1:A5,A1:A5>0)", confirm=True)
        save(wb_path, confirm=True)

        import zipfile
        with zipfile.ZipFile(wb_path) as z:
            assert "xl/metadata.xml" in z.namelist(), "xl/metadata.xml not found in zip"
            meta_xml = z.read("xl/metadata.xml").decode("utf-8")
        assert "XLDAPR" in meta_xml, "XLDAPR not found in metadata.xml"
        assert 'cellMeta="1"' in meta_xml, "cellMeta='1' not found in metadata.xml"

    @pytest.mark.unit
    def test_sort_formula_adds_sheet_metadata_rel(self, tmp_path, monkeypatch):
        """Workbook rels must include sheetMetadata and formula must have _xlfn.SORT prefix."""
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        wb_path = str(tmp_path / "da_rel.xlsx")
        _wb = openpyxl.Workbook()
        _wb.active.title = "Sheet1"
        _wb.save(wb_path)
        _wb_cache.clear()

        fill_formula_range(wb_path, "Sheet1", "C3", "C3", "=SORT(A1:A10)", confirm=True)
        save(wb_path, confirm=True)

        import zipfile
        with zipfile.ZipFile(wb_path) as z:
            rels_xml = z.read("xl/_rels/workbook.xml.rels").decode("utf-8")
            sheet_xml = z.read("xl/worksheets/sheet1.xml").decode("utf-8")
        assert "sheetMetadata" in rels_xml, "sheetMetadata relationship not found in workbook rels"
        assert 'cm="1"' in sheet_xml, "cm='1' not found on DA cell"
        assert '_xlfn.SORT(' in sheet_xml, "_xlfn.SORT prefix not found in formula"

    @pytest.mark.unit
    def test_two_pass_handles_v_element_before_f_element(self, tmp_path):
        """XML with <v/> before <f> inside <c> — DA formula must still get cm=1."""
        from excelmcp._ooxml import _add_cm_to_sheet_xml

        xml = (
            '<?xml version="1.0" encoding="UTF-8"?>'
            '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
            '<sheetData>'
            '<row r="1">'
            '<c r="A1" s="0">'
            '<v/>'
            '<f>_xlfn.FILTER(B1:B10,B1:B10&gt;5)</f>'
            '</c>'
            '</row>'
            '</sheetData>'
            '</worksheet>'
        )
        result_bytes, count = _add_cm_to_sheet_xml(xml.encode("utf-8"))
        result = result_bytes.decode("utf-8")

        assert count == 1, "Should patch exactly 1 cell with V-before-F ordering"
        assert 'cm="1"' in result, "Cell must have cm=1 attribute"
        assert "_xlfn.FILTER" in result, "Formula prefix must be present"

    @pytest.mark.unit
    def test_v_before_f_with_cached_value_patched(self):
        """Cell with explicit <v>value</v> before <f> must be patched (v-before-f bug)."""
        from excelmcp._ooxml import _add_cm_to_sheet_xml

        sheet_xml = (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
            '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/sheet">'
            '<sheetData>'
            '<row r="1">'
            '<c r="A1" t="n"><v>3</v><f>_xlfn.FILTER(B1:B10,C1:C10)</f></c>'
            '</row>'
            '</sheetData>'
            '</worksheet>'
        )
        result_bytes, count = _add_cm_to_sheet_xml(sheet_xml.encode("utf-8"))
        result = result_bytes.decode("utf-8")
        assert count >= 1, f"Expected >=1 cells patched, got {count}"
        assert 'cm="1"' in result, 'Expected cm="1" attribute in patched XML'

    @pytest.mark.unit
    def test_static_formula_is_noop(self, tmp_path):
        """patch_dynamic_array_metadata must return 0 and NOT create metadata.xml for static formulas."""
        from excelmcp._ooxml import patch_dynamic_array_metadata
        wb_path = str(tmp_path / "static.xlsx")
        _wb = openpyxl.Workbook()
        ws = _wb.active
        ws.title = "Sheet1"
        ws["A1"] = "=SUM(B1:B5)"
        _wb.save(wb_path)

        patched = patch_dynamic_array_metadata(wb_path)
        assert patched == 0, f"Expected 0 patched cells for static formula, got {patched}"

        import zipfile
        with zipfile.ZipFile(wb_path) as z:
            assert "xl/metadata.xml" not in z.namelist(), \
                "metadata.xml must not be created for non-DA formulas"

    @pytest.mark.unit
    def test_multiple_da_formulas_all_get_cm(self, tmp_path, monkeypatch):
        """Every cell with a DA formula gets cm='1'; metadata.xml uses a single shared entry."""
        import re as _re
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        wb_path = str(tmp_path / "da_multi.xlsx")
        _wb = openpyxl.Workbook()
        _wb.active.title = "Sheet1"
        _wb.save(wb_path)
        _wb_cache.clear()

        fill_formula_range(wb_path, "Sheet1", "A1", "A1", "=LET(x,1,x)", confirm=True)
        fill_formula_range(wb_path, "Sheet1", "B1", "B1", "=UNIQUE(C1:C5)", confirm=True)
        save(wb_path, confirm=True)

        import zipfile
        with zipfile.ZipFile(wb_path) as z:
            sheet_xml = z.read("xl/worksheets/sheet1.xml").decode("utf-8")
            meta_xml = z.read("xl/metadata.xml").decode("utf-8")

        cm_count = len(_re.findall(r'cm="1"', sheet_xml))
        assert cm_count >= 2, f"Expected at least 2 cm='1' attributes, found {cm_count}"
        assert 'LET(' in sheet_xml, "LET formula not found"
        assert '_xlfn.UNIQUE(' in sheet_xml, "_xlfn.UNIQUE prefix not found"
        bk_count = meta_xml.count("<bk>")
        assert bk_count == 2, f"Expected exactly 2 <bk> entries in metadata.xml, found {bk_count}"

    @pytest.mark.unit
    def test_workbook_xml_gets_excel365_flags(self, tmp_path, monkeypatch):
        """patch_dynamic_array_metadata upgrades workbook.xml to Excel 365:
        adds <fileVersion lastEdited=7>, replaces calcId 124519 → 191028,
        and expands bare <workbookPr/> with defaultThemeVersion."""
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        wb_path = str(tmp_path / "da_wb_flags.xlsx")
        _wb = openpyxl.Workbook()
        _wb.active["A1"] = "=LET(x,42,x)"
        _wb.save(wb_path)

        from excelmcp._ooxml import patch_dynamic_array_metadata
        patch_dynamic_array_metadata(wb_path)

        import zipfile
        with zipfile.ZipFile(wb_path) as z:
            wb_xml = z.read("xl/workbook.xml").decode("utf-8")

        assert 'fileVersion' in wb_xml, "fileVersion missing from workbook.xml"
        assert 'lastEdited="7"' in wb_xml, "lastEdited=7 missing"
        assert 'calcId="191028"' in wb_xml, "calcId not upgraded to 191028"
        assert 'calcId="124519"' not in wb_xml, "old calcId 124519 still present"
        assert 'defaultThemeVersion' in wb_xml, "defaultThemeVersion missing from workbookPr"

    @pytest.mark.unit
    def test_filter_formula_gets_xlfn_prefix(self, tmp_path, monkeypatch):
        """FILTER stored as plain FILTER(...) by openpyxl must be rewritten to
        _xlfn.FILTER(...) so Excel's repair engine keeps the formula."""
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        wb_path = str(tmp_path / "da_filter_prefix.xlsx")
        _wb = openpyxl.Workbook()
        ws = _wb.active
        for i in range(1, 6):
            ws.cell(i, 1, i)
        ws["D1"] = "=FILTER(A1:A5,A1:A5>2)"
        _wb.save(wb_path)

        from excelmcp._ooxml import patch_dynamic_array_metadata
        patch_dynamic_array_metadata(wb_path)

        import zipfile
        with zipfile.ZipFile(wb_path) as z:
            sheet_xml = z.read("xl/worksheets/sheet1.xml").decode("utf-8")

        assert '_xlfn.FILTER(' in sheet_xml, \
            "FILTER must be rewritten to _xlfn.FILTER"
        assert '>FILTER(' not in sheet_xml, \
            "Plain FILTER( must not remain in formula body"

    @pytest.mark.unit
    def test_let_formula_stays_unprefixed(self, tmp_path, monkeypatch):
        """LET is native in Excel 365 — it stays as plain LET(...) with no _xlfn. prefix.
        The patch must still add cm='1' to mark the cell as a dynamic-array anchor."""
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        wb_path = str(tmp_path / "da_let_prefix.xlsx")
        _wb = openpyxl.Workbook()
        _wb.active["A1"] = "=LET(x,42,x)"
        _wb.save(wb_path)

        from excelmcp._ooxml import patch_dynamic_array_metadata
        patch_dynamic_array_metadata(wb_path)

        import zipfile
        with zipfile.ZipFile(wb_path) as z:
            sheet_xml = z.read("xl/worksheets/sheet1.xml").decode("utf-8")

        assert 'LET(' in sheet_xml, "LET formula not found in sheet XML"
        assert '_xlfn.LET(' not in sheet_xml, "LET must NOT have _xlfn. prefix in Excel 365"

    @pytest.mark.unit
    def test_already_prefixed_formula_not_double_prefixed(self, tmp_path):
        """_rewrite_formula_body must not double-prefix already-prefixed functions."""
        from excelmcp._ooxml import _rewrite_formula_body
        formula = "_xlfn.FILTER(A1:B5,A1:A5&gt;2)"
        result = _rewrite_formula_body(formula)
        assert result == formula, f"Should be unchanged; got: {result}"

        formula2 = "LET(x,_xlfn.UNIQUE(A1:A10),x)"
        result2 = _rewrite_formula_body(formula2)
        assert result2 == formula2, f"Should be unchanged; got: {result2}"

    @pytest.mark.unit
    def test_nested_let_filter_sort_all_prefixed(self, tmp_path):
        """Complex formula with nested LET(FILTER(SORT())) gets all prefixes."""
        from excelmcp._ooxml import _rewrite_formula_body
        formula = 'IFERROR(LET(x,FILTER(A1:B5,A1:A5&gt;0),SORT(x,1,1)),"none")'
        result = _rewrite_formula_body(formula)
        assert 'LET(' in result
        assert '_xlfn.FILTER(' in result
        assert '_xlfn.SORT(' in result
        assert '>FILTER(' not in result and ',FILTER(' not in result

    @pytest.mark.unit
    @pytest.mark.parametrize("raw,expected_prefix", [
        ("FILTER(A1:A5,B1:B5>0)",           "_xlfn.FILTER("),
        ("SORT(A1:A5)",                      "_xlfn.SORT("),
        ("SORTBY(A:A,B:B)",                  "_xlfn.SORTBY("),
        ("UNIQUE(A1:A10)",                   "_xlfn.UNIQUE("),
        ("SEQUENCE(5,3)",                    "_xlfn.SEQUENCE("),
        ("RANDARRAY(3,2)",                   "_xlfn.RANDARRAY("),
        ("EXPAND(A1:B2,5,5)",                "_xlfn.EXPAND("),
        ("TAKE(A1:C3,2)",                    "_xlfn.TAKE("),
        ("DROP(A1:C3,1)",                    "_xlfn.DROP("),
        ("TOROW(A1:A10)",                    "_xlfn.TOROW("),
        ("TOCOL(A1:J1)",                     "_xlfn.TOCOL("),
        ("WRAPROWS(A1:A12,4)",               "_xlfn.WRAPROWS("),
        ("WRAPCOLS(A1:A12,4)",               "_xlfn.WRAPCOLS("),
        ("LET(x,5,x*2)",                     "LET("),
        ("XLOOKUP(A1,B:B,C:C)",              "_xlfn.XLOOKUP("),
        ("XMATCH(A1,B:B)",                   "_xlfn.XMATCH("),
        ("MAKEARRAY(3,3,LAMBDA(r,c,r*c))",   "_xlfn.MAKEARRAY("),
    ])
    def test_all_da_func_map_entries_get_correct_prefix(self, raw, expected_prefix):
        """Every _DA_FUNC_MAP entry gets its correct prefix (_xlfn. for spill/lookup; LET stays unprefixed)."""
        from excelmcp._ooxml import _rewrite_formula_body
        result = _rewrite_formula_body(raw)
        assert expected_prefix in result, (
            f"Expected prefix {expected_prefix!r} in result {result!r}"
        )
        # Also verify idempotency
        assert _rewrite_formula_body(result) == result, (
            f"Not idempotent: second call changed {result!r}"
        )
