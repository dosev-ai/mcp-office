"""Unit tests for bulk_write_cells (_io.py) and the bulk_range_write MCP tool.

All tests use openpyxl tmp files — no Excel COM or Outlook required.
"""
from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest
from fastmcp.exceptions import ToolError

import excelmcp._io as io_mod
import excelmcp._range_write as rw_mod
import excelmcp.server_batch as server_batch

from excelmcp.workbook_openpyxl import (
    NotAllowedError,
    ValidationError,
    _wb_cache,
)

bulk_write_cells = getattr(io_mod, "bulk_write_cells")
bulk_range_write = getattr(server_batch, "bulk_range_write")


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _make_wb(tmp_path, monkeypatch, sheet_name: str = "Sheet1") -> str:
    """Create a minimal xlsx workbook and return its path string."""
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    path = tmp_path / "bw.xlsx"
    _wb = openpyxl.Workbook()
    ws = _wb.active
    assert ws is not None
    ws.title = sheet_name
    _wb.save(str(path))
    _wb_cache.clear()
    return str(path)


# ---------------------------------------------------------------------------
# Test cases
# ---------------------------------------------------------------------------

class TestBulkWriteCells:
    """Unit tests for bulk_write_cells (Route A, excel/e1-bulk-write)."""

    @pytest.mark.unit
    def test_happy_path_three_cells(self, tmp_path, monkeypatch):
        """Writing 3 non-contiguous cells returns ok=True, cells_written=3."""
        path = _make_wb(tmp_path, monkeypatch)
        writes = [
            {"address": "A1", "value": "Hello"},
            {"address": "C3", "value": 42},
            {"address": "E5", "value": 3.14},
        ]
        result = bulk_write_cells(path, "Sheet1", writes, confirm=True)

        assert result == {"ok": True, "cells_written": 3}

        # Verify on-disk values
        _wb_cache.clear()
        wb_disk = openpyxl.load_workbook(path)
        ws = wb_disk["Sheet1"]
        assert ws["A1"].value == "Hello"
        assert ws["C3"].value == 42
        assert ws["E5"].value == pytest.approx(3.14)

    @pytest.mark.unit
    def test_formula_injection_blocked(self, tmp_path, monkeypatch):
        """A value starting with '=' is rejected BEFORE _load_wb is called."""
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))

        writes = [
            {"address": "A1", "value": 1},
            {"address": "B1", "value": "=SUM(A1:A10)"},
        ]
        # ValidationError must be raised (no workbook touched — path needn't exist)
        with pytest.raises(ValidationError, match="Formula injection"):
            bulk_write_cells(
                str(tmp_path / "nonexistent.xlsx"),
                "Sheet1",
                writes,
                confirm=True,
            )

    @pytest.mark.unit
    def test_formula_injection_other_chars(self, tmp_path, monkeypatch):
        """Values starting with +, -, or @ are also rejected."""
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))

        for bad_prefix in ("+1", "-1", "@CMD"):
            with pytest.raises(ValidationError, match="Formula injection"):
                bulk_write_cells(
                    str(tmp_path / "nonexistent.xlsx"),
                    "Sheet1",
                    [{"address": "A1", "value": bad_prefix}],
                    confirm=True,
                )

    @pytest.mark.unit
    def test_confirm_false_raises(self, tmp_path, monkeypatch):
        """confirm=False raises ValidationError even when write is enabled."""
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))

        with pytest.raises(ValidationError, match="confirm"):
            bulk_write_cells(
                str(tmp_path / "any.xlsx"),
                "Sheet1",
                [{"address": "A1", "value": 1}],
                confirm=False,
            )

    @pytest.mark.unit
    def test_write_disabled_raises(self, tmp_path, monkeypatch):
        """EXCEL_ENABLE_WRITE not set → NotAllowedError before anything else."""
        monkeypatch.delenv("EXCEL_ENABLE_WRITE", raising=False)
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))

        with pytest.raises(NotAllowedError):
            bulk_write_cells(
                str(tmp_path / "any.xlsx"),
                "Sheet1",
                [{"address": "A1", "value": 1}],
                confirm=True,
            )

    @pytest.mark.unit
    def test_empty_writes_list(self, tmp_path, monkeypatch):
        """writes=[] returns {ok: True, cells_written: 0} without touching disk."""
        path = _make_wb(tmp_path, monkeypatch)
        result = bulk_write_cells(path, "Sheet1", [], confirm=True)
        assert result == {"ok": True, "cells_written": 0}

    @pytest.mark.unit
    @pytest.mark.parametrize("writes", [None, {}, "", 0])
    def test_non_list_writes_payload_rejected(self, tmp_path, monkeypatch, writes):
        """Only a list is accepted for the top-level writes payload."""
        path = _make_wb(tmp_path, monkeypatch)

        with pytest.raises(ValidationError, match="writes must be a list"):
            bulk_write_cells(path, "Sheet1", writes, confirm=True)

    @pytest.mark.unit
    @pytest.mark.parametrize("bad_item", [123, None])
    def test_non_dict_write_item_rejected(self, tmp_path, monkeypatch, bad_item):
        """Malformed write items must raise ValidationError instead of raw TypeError."""
        path = _make_wb(tmp_path, monkeypatch)

        with pytest.raises(ValidationError, match="must be a dict"):
            bulk_write_cells(path, "Sheet1", [bad_item], confirm=True)

    @pytest.mark.unit
    @pytest.mark.parametrize(
        ("item", "expected_fragment"),
        [
            ({"value": "x"}, "missing required key 'address'"),
            ({"address": "A1"}, "missing required key 'value'"),
        ],
    )
    def test_missing_required_keys_rejected(
        self,
        tmp_path,
        monkeypatch,
        item,
        expected_fragment,
    ):
        """Each write item must include both address and value keys."""
        path = _make_wb(tmp_path, monkeypatch)

        with pytest.raises(ValidationError, match=expected_fragment):
            bulk_write_cells(path, "Sheet1", [item], confirm=True)

    @pytest.mark.unit
    def test_invalid_sheet_raises(self, tmp_path, monkeypatch):
        """A sheet name not in the workbook raises ValidationError."""
        path = _make_wb(tmp_path, monkeypatch)

        with pytest.raises(ValidationError, match="Sheet not found"):
            bulk_write_cells(
                path,
                "DoesNotExist",
                [{"address": "A1", "value": "x"}],
                confirm=True,
            )
        # H2-001 regression guard: cache must be evicted after sheet-not-found error
        resolved = str(Path(path).resolve())
        assert (resolved, True) not in _wb_cache, (
            "Workbook must be evicted from _wb_cache after sheet-not-found error"
        )

    @pytest.mark.unit
    def test_cell_address_uppercased(self, tmp_path, monkeypatch):
        """Lowercase cell addresses (e.g. 'a1') are normalised to uppercase."""
        path = _make_wb(tmp_path, monkeypatch)
        result = bulk_write_cells(
            path, "Sheet1",
            [{"address": "a1", "value": "lower"}],
            confirm=True,
        )
        assert result["cells_written"] == 1

        _wb_cache.clear()
        wb_disk = openpyxl.load_workbook(path)
        assert wb_disk["Sheet1"]["A1"].value == "lower"

    @pytest.mark.unit
    def test_none_value_allowed(self, tmp_path, monkeypatch):
        """None is a valid value (clears a cell) — must not be rejected."""
        path = _make_wb(tmp_path, monkeypatch)
        result = bulk_write_cells(
            path, "Sheet1",
            [{"address": "A1", "value": None}],
            confirm=True,
        )
        assert result == {"ok": True, "cells_written": 1}

    @pytest.mark.unit
    @pytest.mark.parametrize("address", ["A0", "123", "!A1"])
    def test_invalid_address_format_rejected(self, tmp_path, monkeypatch, address):
        """Invalid-format addresses are rejected before openpyxl cell access."""
        path = _make_wb(tmp_path, monkeypatch)

        with pytest.raises(ValidationError, match="Invalid cell address"):
            bulk_write_cells(
                path,
                "Sheet1",
                [{"address": address, "value": "x"}],
                confirm=True,
            )

    @pytest.mark.unit
    @pytest.mark.parametrize(
        ("address", "expected_fragment"),
        [
            ("XFE1", "exceeds Excel maximum XFD"),
            ("A1048577", "exceeds Excel maximum"),
        ],
    )
    def test_out_of_bounds_coordinates_rejected(
        self,
        tmp_path,
        monkeypatch,
        address,
        expected_fragment,
    ):
        """Addresses beyond Excel's row/column limits are rejected explicitly."""
        path = _make_wb(tmp_path, monkeypatch)

        with pytest.raises(ValidationError, match=expected_fragment):
            bulk_write_cells(
                path,
                "Sheet1",
                [{"address": address, "value": "x"}],
                confirm=True,
            )

    @pytest.mark.unit
    def test_generic_save_failure_wrapped_and_cache_evicted(self, tmp_path, monkeypatch):
        """Unexpected save/write failures are wrapped in ValidationError and evict cache."""
        path = _make_wb(tmp_path, monkeypatch)
        resolved = str(Path(path).resolve())

        def _boom(_wb, _resolved):
            raise ValueError("boom")

        monkeypatch.setattr(rw_mod, "_save_and_evict", _boom)

        with pytest.raises(ValidationError, match="bulk_write_cells failed: boom"):
            bulk_write_cells(
                path,
                "Sheet1",
                [{"address": "A1", "value": "x"}],
                confirm=True,
            )

        assert (resolved, True) not in _wb_cache, (
            "Workbook must be evicted from _wb_cache after wrapped generic failure"
        )


@pytest.mark.unit
def test_bulk_range_write_wrapper_translates_validation_error(monkeypatch):
    """The MCP wrapper must translate workbook-layer ValidationError to ToolError."""

    def _boom(path, sheet, writes, confirm=False):
        raise ValidationError("boom")

    monkeypatch.setattr(server_batch.wb, "bulk_write_cells", _boom)

    with pytest.raises(ToolError, match="boom"):
        bulk_range_write(
            "C:/tmp/book.xlsx",
            "Sheet1",
            [{"address": "A1", "value": "x"}],
            confirm=True,
        )
