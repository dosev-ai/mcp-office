"""
Integration tests for unified Sprint O COM tools.
Requires live Excel on Windows (Office 2016+ or M365).

Preconditions:
    - Excel desktop application installed and licensed
    - EXCEL_ENABLE_COM=true set in shell environment
    - EXCEL_ENABLE_WRITE=true set in shell environment
    - Run with: pytest excel/tests/test_integration_com.py -v -m integration

These tests are excluded from CI (no Excel in CI agents).
"""
from __future__ import annotations

import sys

import pytest
import openpyxl

from excelmcp._core import ValidationError


def _has_live_excel() -> bool:
    """Return True only when Excel COM automation can be started."""
    if sys.platform != "win32":
        return False
    try:
        import pythoncom
        import win32com.client as win32c
    except ImportError:
        return False

    initialized = False
    excel = None
    try:
        pythoncom.CoInitialize()
        initialized = True
        excel = win32c.Dispatch("Excel.Application")
        excel.Quit()
        return True
    except Exception:
        return False
    finally:
        if excel is not None:
            del excel
        if initialized:
            pythoncom.CoUninitialize()


_LIVE_EXCEL_AVAILABLE = _has_live_excel()


@pytest.mark.integration
class TestIntegrationCOMPhase2:

    def test_integration_recalculate_simple_formula(self, monkeypatch, tmp_path):
        """Build a workbook with =SUM formula, recalculate via live COM,
        reload data_only=True and verify the formula evaluates to the expected value.

        Requires: EXCEL_ENABLE_COM=true, EXCEL_ENABLE_WRITE=true in environment.
        """
        if not _LIVE_EXCEL_AVAILABLE:
            pytest.skip("Excel COM automation unavailable on this machine")
        monkeypatch.setenv("EXCEL_ENABLE_COM", "true")
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))

        wb_path = tmp_path / "formula_test.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        assert ws is not None
        ws.title = "Sheet"
        ws["A1"] = 10
        ws["A2"] = 20
        ws["A3"] = "=SUM(A1:A2)"
        wb.save(str(wb_path))

        from excelmcp.server_com import recalculate_workbook
        result = recalculate_workbook(
            path=str(wb_path), full_calc=True, confirm=True
        )

        assert result["ok"] is True
        reloaded = openpyxl.load_workbook(str(wb_path), data_only=True)
        assert reloaded["Sheet"]["A3"].value == 30

    def test_integration_export_as_pdf_sheet_returns_evidence(self, monkeypatch, tmp_path):
        """Export a real worksheet via the unified tool and verify evidence fields.

        Requires: EXCEL_ENABLE_COM=true in environment.
        EXCEL_ENABLE_WRITE is NOT required (source workbook opened read-only).
        """
        if not _LIVE_EXCEL_AVAILABLE:
            pytest.skip("Excel COM automation unavailable on this machine")
        monkeypatch.setenv("EXCEL_ENABLE_COM", "true")
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        monkeypatch.setenv("EXCEL_EXPORT_ROOTS", str(tmp_path))

        wb_path = tmp_path / "export_test.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        assert ws is not None
        ws.title = "Sheet1"
        ws["A1"] = "Sprint G PDF Export Test"
        wb.save(str(wb_path))

        out_pdf = tmp_path / "sheet_export.pdf"

        from excelmcp.server_com import export_as_pdf
        result = export_as_pdf(
            scope="sheet",
            path=str(wb_path),
            output_path=str(out_pdf),
            sheet="Sheet1",
            confirm=True,
        )

        assert result["ok"] is True
        assert result["sheet"] == "Sheet1"
        assert result["output_path"] == str(out_pdf.resolve())
        assert isinstance(result["sha256"], str)
        assert len(result["sha256"]) == 64
        assert result["size_bytes"] == out_pdf.stat().st_size
        assert result["exported_at"].endswith("Z")
        assert isinstance(result["elapsed_ms"], int)
        assert out_pdf.exists()
        assert out_pdf.stat().st_size > 1024, "PDF should be at least 1 KB"

    def test_integration_export_as_pdf_workbook_returns_evidence(self, monkeypatch, tmp_path):
        """Export a real workbook via the unified tool and verify workbook evidence fields."""
        if not _LIVE_EXCEL_AVAILABLE:
            pytest.skip("Excel COM automation unavailable on this machine")
        monkeypatch.setenv("EXCEL_ENABLE_COM", "true")
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        monkeypatch.setenv("EXCEL_EXPORT_ROOTS", str(tmp_path))

        wb_path = tmp_path / "workbook_export_test.xlsx"
        wb = openpyxl.Workbook()
        ws1 = wb.active
        assert ws1 is not None
        ws1.title = "Summary"
        ws1["A1"] = "Workbook export summary"
        ws2 = wb.create_sheet("Detail")
        ws2["A1"] = "Second visible sheet"
        wb.save(str(wb_path))

        out_pdf = tmp_path / "workbook_export.pdf"

        from excelmcp.server_com import export_as_pdf
        result = export_as_pdf(
            scope="workbook",
            path=str(wb_path),
            output_path=str(out_pdf),
            confirm=True,
        )

        assert result["ok"] is True
        assert result["output_path"] == str(out_pdf.resolve())
        assert result["quality"] == 0
        assert result["source"] == str(wb_path.resolve())
        assert isinstance(result["sha256"], str)
        assert len(result["sha256"]) == 64
        assert result["size_bytes"] == out_pdf.stat().st_size
        assert result["exported_at"].endswith("Z")
        assert isinstance(result["elapsed_ms"], int)
        assert out_pdf.exists()
        assert out_pdf.stat().st_size > 1024, "Workbook PDF should be at least 1 KB"

    def test_integration_export_as_pdf_clears_stale_target_before_live_export(self, monkeypatch, tmp_path):
        """A stale PDF target is replaced by fresh bytes during a live export."""
        if not _LIVE_EXCEL_AVAILABLE:
            pytest.skip("Excel COM automation unavailable on this machine")
        monkeypatch.setenv("EXCEL_ENABLE_COM", "true")
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        monkeypatch.setenv("EXCEL_EXPORT_ROOTS", str(tmp_path))

        wb_path = tmp_path / "export_replace.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        assert ws is not None
        ws.title = "Sheet1"
        ws["A1"] = "Fresh export replaces stale target"
        ws["A2"] = 123
        wb.save(str(wb_path))

        out_pdf = tmp_path / "stale_then_fresh.pdf"
        stale_bytes = b"%PDF-1.7\nstale-target\n"
        out_pdf.write_bytes(stale_bytes)

        from excelmcp.server_com import export_as_pdf
        result = export_as_pdf(
            scope="sheet",
            path=str(wb_path),
            output_path=str(out_pdf),
            sheet="Sheet1",
            confirm=True,
        )

        fresh_bytes = out_pdf.read_bytes()
        assert result["ok"] is True
        assert fresh_bytes != stale_bytes
        assert result["size_bytes"] == len(fresh_bytes)
        assert len(fresh_bytes) > len(stale_bytes)

    def test_integration_export_as_pdf_invalid_sheet_preserves_existing_target(self, monkeypatch, tmp_path):
        """Validation failures do not delete a pre-existing export target."""
        if not _LIVE_EXCEL_AVAILABLE:
            pytest.skip("Excel COM automation unavailable on this machine")
        monkeypatch.setenv("EXCEL_ENABLE_COM", "true")
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        monkeypatch.setenv("EXCEL_EXPORT_ROOTS", str(tmp_path))

        wb_path = tmp_path / "export_preserve.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        assert ws is not None
        ws.title = "Sheet1"
        ws["A1"] = "Preserve stale target on invalid request"
        wb.save(str(wb_path))

        out_pdf = tmp_path / "preserved_stale.pdf"
        stale_bytes = b"%PDF-1.7\npreserve-me\n"
        out_pdf.write_bytes(stale_bytes)

        from excelmcp.server_com import export_as_pdf
        with pytest.raises(ValidationError, match="not found in workbook"):
            export_as_pdf(
                scope="sheet",
                path=str(wb_path),
                output_path=str(out_pdf),
                sheet="MissingSheet",
                confirm=True,
            )

        assert out_pdf.read_bytes() == stale_bytes

    def test_integration_export_as_pdf_workbook_scope_rejects_stray_sheet_preserves_existing_target(self, monkeypatch, tmp_path):
        """Workbook exports reject stray sheet args before touching a pre-existing target."""
        if not _LIVE_EXCEL_AVAILABLE:
            pytest.skip("Excel COM automation unavailable on this machine")
        monkeypatch.setenv("EXCEL_ENABLE_COM", "true")
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        monkeypatch.setenv("EXCEL_EXPORT_ROOTS", str(tmp_path))

        wb_path = tmp_path / "workbook_scope_invalid.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        assert ws is not None
        ws.title = "Sheet1"
        ws["A1"] = "Reject stray sheet on workbook export"
        wb.save(str(wb_path))

        out_pdf = tmp_path / "workbook_scope_invalid.pdf"
        stale_bytes = b"%PDF-1.7\npreserve-workbook-invalid\n"
        out_pdf.write_bytes(stale_bytes)

        from excelmcp.server_com import export_as_pdf
        with pytest.raises(ValidationError, match="sheet is not allowed when scope='workbook'"):
            export_as_pdf(
                scope="workbook",
                path=str(wb_path),
                output_path=str(out_pdf),
                sheet="Sheet1",
                confirm=True,
            )

        assert out_pdf.read_bytes() == stale_bytes

    def test_integration_hydrate_template(self, monkeypatch, tmp_path):
        """Build a template workbook, hydrate cell A1 via COM SaveAs, reload
        the output file with openpyxl and verify the written value persists.

        Requires: EXCEL_ENABLE_COM=true, EXCEL_ENABLE_WRITE=true in environment.
        """
        if not _LIVE_EXCEL_AVAILABLE:
            pytest.skip("Excel COM automation unavailable on this machine")
        monkeypatch.setenv("EXCEL_ENABLE_COM", "true")
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))

        tmpl_path = tmp_path / "template.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        assert ws is not None
        ws.title = "Data"
        ws["A1"] = "PLACEHOLDER"
        wb.save(str(tmpl_path))

        out_path = tmp_path / "hydrated_output.xlsx"

        from excelmcp.server_com import hydrate_template
        result = hydrate_template(
            template_path=str(tmpl_path),
            cell_writes=[{"ref": "A1", "value": "Hydrated Value"}],
            output_path=str(out_path),
            sheet="Data",
            recalculate=True,
            confirm=True,
        )

        assert result["ok"] is True
        assert out_path.exists()
        reloaded = openpyxl.load_workbook(str(out_path))
        assert reloaded["Data"]["A1"].value == "Hydrated Value"
