"""Tests for governed write to active workbook via COM session.

Covers:
  D-series: _write_cell_via_com direct tests
  E-series: _write_range_via_com direct tests
  F-series: _io.py routing tests
  G-series: server_io.py smoke integration tests

All tests run without a live Excel instance.
"""
from __future__ import annotations

from pathlib import Path
from unittest.mock import MagicMock, PropertyMock, patch

import pytest

from excelmcp._core import OfficeCOMError, ValidationError, NotAllowedError

# ---------------------------------------------------------------------------
# pywintypes surrogate (allows tests to run without pywin32 installed)
# ---------------------------------------------------------------------------
try:
    import pywintypes as _pywintypes_mod
    _FakeCOMError = _pywintypes_mod.com_error
except ImportError:
    class _FakeCOMError(Exception):  # type: ignore[misc]
        """Surrogate for pywintypes.com_error in non-Windows / CI environments."""


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _make_mock_ws(previous_value=None):
    """Return (mock_ws, mock_range) with Range(...).Value = previous_value."""
    mock_range = MagicMock()
    mock_range.Value = previous_value
    mock_ws = MagicMock()
    mock_ws.Range.return_value = mock_range
    return mock_ws, mock_range


def _make_mock_wb(full_name: str, read_only: bool = False, sheets: dict | None = None):
    """Return a mock COM workbook with FullName, ReadOnly, Worksheets."""
    mock_wb = MagicMock()
    mock_wb.FullName = full_name
    mock_wb.Name = Path(full_name).name
    mock_wb.ReadOnly = read_only
    if sheets:
        def _worksheet_lookup(name):
            if name in sheets:
                return sheets[name]
            raise Exception(f"Sheet '{name}' not found")
        mock_wb.Worksheets.side_effect = _worksheet_lookup
    else:
        mock_wb.Worksheets.side_effect = Exception("Sheet not found")
    return mock_wb


def _mock_com_tuple(mock_xl_app):
    """Return (mock_pythoncom, mock_win32c) wired to yield mock_xl_app."""
    mock_pythoncom = MagicMock()
    mock_win32c = MagicMock()
    mock_win32c.GetActiveObject.return_value = mock_xl_app
    return mock_pythoncom, mock_win32c


# ---------------------------------------------------------------------------
# D-series: _write_cell_via_com direct tests
# ---------------------------------------------------------------------------

class TestWriteCellViaCom:

    def test_d01_com_disabled_raises(self, monkeypatch, tmp_path):
        """D-01: EXCEL_ENABLE_COM not 'true' → OfficeCOMError."""
        monkeypatch.setenv("EXCEL_ENABLE_COM", "false")
        from excelmcp._active_wb import _write_cell_via_com

        with pytest.raises(OfficeCOMError, match="EXCEL_ENABLE_COM"):
            _write_cell_via_com(tmp_path / "wb.xlsx", "Sheet1", "A1", "hello")

    def test_d02_pywin32_not_installed_raises(self, monkeypatch, tmp_path):
        """D-02: pywin32 not importable → OfficeCOMError."""
        monkeypatch.setenv("EXCEL_ENABLE_COM", "true")
        with patch("excelmcp._active_wb._ensure_com_available", side_effect=ImportError("no pywin32")):
            from excelmcp._active_wb import _write_cell_via_com

            with pytest.raises(OfficeCOMError, match="pywin32"):
                _write_cell_via_com(tmp_path / "wb.xlsx", "Sheet1", "A1", "hello")

    def test_d03_no_running_excel_raises(self, monkeypatch, tmp_path):
        """D-03: No running Excel → OfficeCOMError."""
        monkeypatch.setenv("EXCEL_ENABLE_COM", "true")
        mock_pythoncom = MagicMock()
        mock_win32c = MagicMock()
        mock_win32c.GetActiveObject.side_effect = Exception("not running")

        with patch("excelmcp._active_wb._ensure_com_available", return_value=(mock_pythoncom, mock_win32c)):
            from excelmcp._active_wb import _write_cell_via_com

            with pytest.raises(OfficeCOMError, match="No running Excel"):
                _write_cell_via_com(tmp_path / "wb.xlsx", "Sheet1", "A1", "hello")

    def test_d04_excel_not_ready_raises(self, monkeypatch, tmp_path):
        """D-04: xl_app.Ready=False → OfficeCOMError."""
        monkeypatch.setenv("EXCEL_ENABLE_COM", "true")
        mock_xl_app = MagicMock()
        mock_xl_app.Ready = False
        mock_pythoncom, mock_win32c = _mock_com_tuple(mock_xl_app)

        with patch("excelmcp._active_wb._ensure_com_available", return_value=(mock_pythoncom, mock_win32c)):
            from excelmcp._active_wb import _write_cell_via_com

            with pytest.raises(OfficeCOMError, match="busy"):
                _write_cell_via_com(tmp_path / "wb.xlsx", "Sheet1", "A1", "hello")

    def test_d05_no_active_workbook_raises(self, monkeypatch, tmp_path):
        """D-05: ActiveWorkbook is None → OfficeCOMError."""
        monkeypatch.setenv("EXCEL_ENABLE_COM", "true")
        mock_xl_app = MagicMock()
        mock_xl_app.Ready = True
        mock_xl_app.ActiveWorkbook = None
        mock_pythoncom, mock_win32c = _mock_com_tuple(mock_xl_app)

        with patch("excelmcp._active_wb._ensure_com_available", return_value=(mock_pythoncom, mock_win32c)):
            from excelmcp._active_wb import _write_cell_via_com

            with pytest.raises(OfficeCOMError, match="No active workbook"):
                _write_cell_via_com(tmp_path / "wb.xlsx", "Sheet1", "A1", "hello")

    def test_d06_path_mismatch_raises(self, monkeypatch, tmp_path):
        """D-06: Active wb FullName != resolved path → ValidationError."""
        monkeypatch.setenv("EXCEL_ENABLE_COM", "true")
        other_path = str(tmp_path / "other.xlsx")
        mock_wb = _make_mock_wb(other_path)
        mock_xl_app = MagicMock()
        mock_xl_app.Ready = True
        mock_xl_app.ActiveWorkbook = mock_wb
        mock_pythoncom, mock_win32c = _mock_com_tuple(mock_xl_app)

        with patch("excelmcp._active_wb._ensure_com_available", return_value=(mock_pythoncom, mock_win32c)):
            from excelmcp._active_wb import _write_cell_via_com

            with pytest.raises(ValidationError, match="does not match"):
                _write_cell_via_com(tmp_path / "target.xlsx", "Sheet1", "A1", "hello")

    def test_d07_readonly_raises(self, monkeypatch, tmp_path):
        """D-07: Workbook ReadOnly=True → ValidationError."""
        monkeypatch.setenv("EXCEL_ENABLE_COM", "true")
        wb_path = str(tmp_path / "wb.xlsx")
        mock_wb = _make_mock_wb(wb_path, read_only=True)
        mock_xl_app = MagicMock()
        mock_xl_app.Ready = True
        mock_xl_app.ActiveWorkbook = mock_wb
        mock_pythoncom, mock_win32c = _mock_com_tuple(mock_xl_app)

        with patch("excelmcp._active_wb._ensure_com_available", return_value=(mock_pythoncom, mock_win32c)):
            from excelmcp._active_wb import _write_cell_via_com

            with pytest.raises(ValidationError, match="ReadOnly"):
                _write_cell_via_com(tmp_path / "wb.xlsx", "Sheet1", "A1", "hello")

    def test_d08_sheet_not_found_raises(self, monkeypatch, tmp_path):
        """D-08: Sheet not in workbook → ValidationError."""
        monkeypatch.setenv("EXCEL_ENABLE_COM", "true")
        wb_path = str((tmp_path / "wb.xlsx").resolve())
        # No sheets= arg → Worksheets() raises by default
        mock_wb = _make_mock_wb(wb_path, read_only=False)
        mock_xl_app = MagicMock()
        mock_xl_app.Ready = True
        mock_xl_app.ActiveWorkbook = mock_wb
        mock_pythoncom, mock_win32c = _mock_com_tuple(mock_xl_app)

        with patch("excelmcp._active_wb._ensure_com_available", return_value=(mock_pythoncom, mock_win32c)):
            from excelmcp._active_wb import _write_cell_via_com

            with pytest.raises(ValidationError, match="Sheet not found"):
                _write_cell_via_com(tmp_path / "wb.xlsx", "MissingSheet", "A1", "hello")

    def test_d09_successful_write_returns_ok(self, monkeypatch, tmp_path):
        """D-09: Successful write → {"ok": True, "previous_value": ...}."""
        monkeypatch.setenv("EXCEL_ENABLE_COM", "true")
        wb_path = str((tmp_path / "wb.xlsx").resolve())
        mock_ws, mock_range = _make_mock_ws(previous_value="old_value")
        mock_wb = _make_mock_wb(wb_path, read_only=False, sheets={"Sheet1": mock_ws})
        mock_xl_app = MagicMock()
        mock_xl_app.Ready = True
        mock_xl_app.ActiveWorkbook = mock_wb
        mock_pythoncom, mock_win32c = _mock_com_tuple(mock_xl_app)

        with patch("excelmcp._active_wb._ensure_com_available", return_value=(mock_pythoncom, mock_win32c)):
            from excelmcp._active_wb import _write_cell_via_com

            result = _write_cell_via_com(tmp_path / "wb.xlsx", "Sheet1", "A1", "new_value")

        assert result["ok"] is True
        assert result["previous_value"] == "old_value"

    def test_d10_previous_value_captured_before_overwrite(self, monkeypatch, tmp_path):
        """D-10: previous_value reflects the cell value BEFORE the write."""
        monkeypatch.setenv("EXCEL_ENABLE_COM", "true")
        wb_path = str((tmp_path / "wb.xlsx").resolve())
        mock_ws, mock_range = _make_mock_ws(previous_value=42)
        mock_wb = _make_mock_wb(wb_path, read_only=False, sheets={"Sheet1": mock_ws})
        mock_xl_app = MagicMock()
        mock_xl_app.Ready = True
        mock_xl_app.ActiveWorkbook = mock_wb
        mock_pythoncom, mock_win32c = _mock_com_tuple(mock_xl_app)

        with patch("excelmcp._active_wb._ensure_com_available", return_value=(mock_pythoncom, mock_win32c)):
            from excelmcp._active_wb import _write_cell_via_com

            result = _write_cell_via_com(tmp_path / "wb.xlsx", "Sheet1", "B2", 99)

        # previous_value was 42 (from mock_range.Value before assignment)
        assert result["previous_value"] == 42
        # The new value was assigned to the COM range object
        assert mock_range.Value == 99

    def test_d11_com_error_during_write_raises_office_com_error(self, monkeypatch, tmp_path):
        """D-11: pywintypes.com_error at the write SET is caught and re-raised as OfficeCOMError.

        Advisory-3 fix: use side_effect=[42, _FakeCOMError(...)] so that the GET (previous_value)
        returns 42 and the SET raises — testing the correct write-assignment path.
        """
        monkeypatch.setenv("EXCEL_ENABLE_COM", "true")
        wb_path = str((tmp_path / "wb.xlsx").resolve())
        mock_ws = MagicMock()
        mock_range = MagicMock()
        # GET returns 42 (previous_value read), SET raises com_error
        type(mock_range).Value = PropertyMock(side_effect=[42, _FakeCOMError("Cell is protected.")])
        mock_ws.Range.return_value = mock_range
        mock_wb = _make_mock_wb(wb_path, read_only=False, sheets={"Sheet1": mock_ws})
        mock_xl_app = MagicMock()
        mock_xl_app.Ready = True
        mock_xl_app.ActiveWorkbook = mock_wb
        mock_pythoncom, mock_win32c = _mock_com_tuple(mock_xl_app)

        with patch("excelmcp._active_wb._ensure_com_available", return_value=(mock_pythoncom, mock_win32c)):
            from excelmcp._active_wb import _write_cell_via_com

            with pytest.raises(OfficeCOMError):
                _write_cell_via_com(tmp_path / "wb.xlsx", "Sheet1", "B2", 99)

    def test_d12_formula_value_rejected_before_com(self, monkeypatch, tmp_path):
        """D-12 (Advisory-1): Formula string as value raises ValidationError before COM is touched."""
        monkeypatch.setenv("EXCEL_ENABLE_COM", "true")
        mock_pythoncom = MagicMock()
        mock_win32c = MagicMock()

        with patch("excelmcp._active_wb._ensure_com_available", return_value=(mock_pythoncom, mock_win32c)):
            from excelmcp._active_wb import _write_cell_via_com

            with pytest.raises(ValidationError, match="Formula values rejected"):
                _write_cell_via_com(tmp_path / "wb.xlsx", "Sheet1", "A1", "=SUM(A1:A10)")

        # CoInitialize must NOT have been called
        mock_pythoncom.CoInitialize.assert_not_called()

    def test_d13_finally_block_runs_and_couninitialize_called_on_exception(self, monkeypatch, tmp_path):
        """D-13: finally block is entered on exception path — CoUninitialize is called and
        null-proxy assignments (xl_app=None, wb=None, etc.) do not raise even though the
        locals may have never been bound. Exercises the null-before-CoUninitialize fix.
        Simulates GetActiveObject raising to drive the exception path through the finally."""
        monkeypatch.setenv("EXCEL_ENABLE_COM", "true")
        mock_pythoncom = MagicMock()
        mock_win32c = MagicMock()
        mock_win32c.GetActiveObject.side_effect = Exception("Excel not running")

        with patch("excelmcp._active_wb._ensure_com_available", return_value=(mock_pythoncom, mock_win32c)):
            from excelmcp._active_wb import _write_cell_via_com

            with pytest.raises(OfficeCOMError, match="No running Excel"):
                _write_cell_via_com(tmp_path / "wb.xlsx", "Sheet1", "A1", "hello")

        # The finally block MUST have been entered — CoUninitialize must have been called
        mock_pythoncom.CoUninitialize.assert_called_once()


class TestWriteRangeViaCom:

    def test_e01_com_disabled_raises(self, monkeypatch, tmp_path):
        """E-01: EXCEL_ENABLE_COM not 'true' → OfficeCOMError."""
        monkeypatch.setenv("EXCEL_ENABLE_COM", "false")
        from excelmcp._active_wb import _write_range_via_com

        with pytest.raises(OfficeCOMError, match="EXCEL_ENABLE_COM"):
            _write_range_via_com(tmp_path / "wb.xlsx", "Sheet1", "A1", [[1, 2], [3, 4]])

    def test_e02_excel_not_ready_raises(self, monkeypatch, tmp_path):
        """E-02: xl_app.Ready=False → OfficeCOMError."""
        monkeypatch.setenv("EXCEL_ENABLE_COM", "true")
        mock_xl_app = MagicMock()
        mock_xl_app.Ready = False
        mock_pythoncom, mock_win32c = _mock_com_tuple(mock_xl_app)

        with patch("excelmcp._active_wb._ensure_com_available", return_value=(mock_pythoncom, mock_win32c)):
            from excelmcp._active_wb import _write_range_via_com

            with pytest.raises(OfficeCOMError, match="busy"):
                _write_range_via_com(tmp_path / "wb.xlsx", "Sheet1", "A1", [[1, 2]])

    def test_e03_path_mismatch_raises(self, monkeypatch, tmp_path):
        """E-03: Active wb path != resolved → ValidationError."""
        monkeypatch.setenv("EXCEL_ENABLE_COM", "true")
        other_path = str(tmp_path / "other.xlsx")
        mock_wb = _make_mock_wb(other_path)
        mock_xl_app = MagicMock()
        mock_xl_app.Ready = True
        mock_xl_app.ActiveWorkbook = mock_wb
        mock_pythoncom, mock_win32c = _mock_com_tuple(mock_xl_app)

        with patch("excelmcp._active_wb._ensure_com_available", return_value=(mock_pythoncom, mock_win32c)):
            from excelmcp._active_wb import _write_range_via_com

            with pytest.raises(ValidationError, match="does not match"):
                _write_range_via_com(tmp_path / "target.xlsx", "Sheet1", "A1", [[1]])

    def test_e04_readonly_raises(self, monkeypatch, tmp_path):
        """E-04: ReadOnly workbook → ValidationError."""
        monkeypatch.setenv("EXCEL_ENABLE_COM", "true")
        wb_path = str((tmp_path / "wb.xlsx").resolve())
        mock_wb = _make_mock_wb(wb_path, read_only=True)
        mock_xl_app = MagicMock()
        mock_xl_app.Ready = True
        mock_xl_app.ActiveWorkbook = mock_wb
        mock_pythoncom, mock_win32c = _mock_com_tuple(mock_xl_app)

        with patch("excelmcp._active_wb._ensure_com_available", return_value=(mock_pythoncom, mock_win32c)):
            from excelmcp._active_wb import _write_range_via_com

            with pytest.raises(ValidationError, match="ReadOnly"):
                _write_range_via_com(tmp_path / "wb.xlsx", "Sheet1", "A1", [[1, 2]])

    def test_e05_successful_range_write_returns_cells_written(self, monkeypatch, tmp_path):
        """E-05: Successful range write → {"ok": True, "cells_written": N}."""
        monkeypatch.setenv("EXCEL_ENABLE_COM", "true")
        wb_path = str((tmp_path / "wb.xlsx").resolve())
        mock_ws = MagicMock()
        mock_resize = MagicMock()
        mock_ws.Range.return_value.Resize.return_value = mock_resize
        mock_wb = _make_mock_wb(wb_path, read_only=False, sheets={"Sheet1": mock_ws})
        mock_xl_app = MagicMock()
        mock_xl_app.Ready = True
        mock_xl_app.ActiveWorkbook = mock_wb
        mock_pythoncom, mock_win32c = _mock_com_tuple(mock_xl_app)

        values = [[1, 2, 3], [4, 5, 6]]
        with patch("excelmcp._active_wb._ensure_com_available", return_value=(mock_pythoncom, mock_win32c)):
            from excelmcp._active_wb import _write_range_via_com

            result = _write_range_via_com(tmp_path / "wb.xlsx", "Sheet1", "B2", values)

        assert result["ok"] is True
        assert result["cells_written"] == 6

    def test_e06_resize_value_called_with_tuple_of_tuples(self, monkeypatch, tmp_path):
        """E-06: ws.Range(start).Resize(rows, cols).Value = tuple-of-tuples."""
        monkeypatch.setenv("EXCEL_ENABLE_COM", "true")
        wb_path = str((tmp_path / "wb.xlsx").resolve())
        mock_ws = MagicMock()
        mock_resize_range = MagicMock()
        mock_ws.Range.return_value.Resize.return_value = mock_resize_range
        mock_wb = _make_mock_wb(wb_path, read_only=False, sheets={"Sheet1": mock_ws})
        mock_xl_app = MagicMock()
        mock_xl_app.Ready = True
        mock_xl_app.ActiveWorkbook = mock_wb
        mock_pythoncom, mock_win32c = _mock_com_tuple(mock_xl_app)

        values = [[10, 20], [30, 40]]
        with patch("excelmcp._active_wb._ensure_com_available", return_value=(mock_pythoncom, mock_win32c)):
            from excelmcp._active_wb import _write_range_via_com

            _write_range_via_com(tmp_path / "wb.xlsx", "Sheet1", "A1", values)

        # Verify Resize was called correctly
        mock_ws.Range.assert_called_with("A1")
        mock_ws.Range.return_value.Resize.assert_called_with(2, 2)
        # Value was assigned as tuple-of-tuples
        assert mock_resize_range.Value == ((10, 20), (30, 40))

    def test_e07_com_error_during_range_write_raises_office_com_error(self, monkeypatch, tmp_path):
        """E-07: pywintypes.com_error during target_range.Value write is caught as OfficeCOMError."""
        monkeypatch.setenv("EXCEL_ENABLE_COM", "true")
        wb_path = str((tmp_path / "wb.xlsx").resolve())
        mock_ws = MagicMock()
        mock_resize_range = MagicMock()
        type(mock_resize_range).Value = PropertyMock(side_effect=_FakeCOMError("Sheet is protected."))
        mock_ws.Range.return_value.Resize.return_value = mock_resize_range
        mock_wb = _make_mock_wb(wb_path, read_only=False, sheets={"Sheet1": mock_ws})
        mock_xl_app = MagicMock()
        mock_xl_app.Ready = True
        mock_xl_app.ActiveWorkbook = mock_wb
        mock_pythoncom, mock_win32c = _mock_com_tuple(mock_xl_app)

        with patch("excelmcp._active_wb._ensure_com_available", return_value=(mock_pythoncom, mock_win32c)):
            from excelmcp._active_wb import _write_range_via_com

            with pytest.raises(OfficeCOMError):
                _write_range_via_com(tmp_path / "wb.xlsx", "Sheet1", "A1", [[1, 2], [3, 4]])

    def test_e08_ragged_values_pads_with_none(self, monkeypatch, tmp_path):
        """E-08: Ragged values matrix is padded to rectangular before COM write."""
        monkeypatch.setenv("EXCEL_ENABLE_COM", "true")
        wb_path = str((tmp_path / "wb.xlsx").resolve())
        mock_ws = MagicMock()
        mock_resize_range = MagicMock()
        mock_ws.Range.return_value.Resize.return_value = mock_resize_range
        mock_wb = _make_mock_wb(wb_path, read_only=False, sheets={"Sheet1": mock_ws})
        mock_xl_app = MagicMock()
        mock_xl_app.Ready = True
        mock_xl_app.ActiveWorkbook = mock_wb
        mock_pythoncom, mock_win32c = _mock_com_tuple(mock_xl_app)

        with patch("excelmcp._active_wb._ensure_com_available", return_value=(mock_pythoncom, mock_win32c)):
            from excelmcp._active_wb import _write_range_via_com

            result = _write_range_via_com(tmp_path / "wb.xlsx", "Sheet1", "A1", [[1, 2, 3], [4, 5]])

        assert result["ok"] is True
        assert result["cells_written"] == 5   # sum of actual cells, not n_rows * n_cols
        # Value was set with ragged row padded to ((1,2,3),(4,5,None))
        assert mock_resize_range.Value == ((1, 2, 3), (4, 5, None))

    def test_e09_invalid_start_cell_raises_validation_error(self, monkeypatch, tmp_path):
        """E-09 (CBR-R01): Invalid start_cell address → ValidationError before COM is touched."""
        monkeypatch.setenv("EXCEL_ENABLE_COM", "true")
        mock_pythoncom = MagicMock()
        mock_win32c = MagicMock()

        with patch("excelmcp._active_wb._ensure_com_available", return_value=(mock_pythoncom, mock_win32c)):
            from excelmcp._active_wb import _write_range_via_com

            with pytest.raises(ValidationError, match="Invalid start_cell"):
                _write_range_via_com(tmp_path / "wb.xlsx", "Sheet1", "A1:B2", [[1, 2]])

    def test_e10_formula_in_values_rejected_before_com(self, monkeypatch, tmp_path):
        """E-10 (Advisory-1): Formula string in values matrix raises ValidationError before COM is touched."""
        monkeypatch.setenv("EXCEL_ENABLE_COM", "true")
        mock_pythoncom = MagicMock()
        mock_win32c = MagicMock()

        with patch("excelmcp._active_wb._ensure_com_available", return_value=(mock_pythoncom, mock_win32c)):
            from excelmcp._active_wb import _write_range_via_com

            with pytest.raises(ValidationError, match="Formula values rejected in range"):
                _write_range_via_com(tmp_path / "wb.xlsx", "Sheet1", "A1", [["=MALICIOUS()", 1]])

        mock_pythoncom.CoInitialize.assert_not_called()

    def test_e11_empty_row_values_raises_validation_error(self, monkeypatch, tmp_path):
        """E-11: values=[[]] (single empty row) hits n_cols==0 guard → ValidationError before COM."""
        monkeypatch.setenv("EXCEL_ENABLE_COM", "true")
        mock_pythoncom = MagicMock()
        mock_win32c = MagicMock()

        with patch("excelmcp._active_wb._ensure_com_available", return_value=(mock_pythoncom, mock_win32c)):
            from excelmcp._active_wb import _write_range_via_com

            with pytest.raises(ValidationError, match="at least one non-empty row"):
                _write_range_via_com(tmp_path / "wb.xlsx", "Sheet1", "A1", [[]])

        mock_pythoncom.CoInitialize.assert_not_called()


# ---------------------------------------------------------------------------
# F-series: _io.py routing tests
# ---------------------------------------------------------------------------

class TestIoRouting:

    def test_f01_write_cell_session_false_uses_openpyxl(self, monkeypatch, tmp_path):
        """F-01: write_cell(session_mode=False) takes the openpyxl path (regression)."""
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))

        wb_file = tmp_path / "test.xlsx"
        import openpyxl
        _wb = openpyxl.Workbook()
        _wb.active.title = "Sheet1"
        _wb.save(str(wb_file))

        # session_mode=False — must NOT call _write_cell_via_com
        with patch("excelmcp._active_wb._write_cell_via_com") as mock_com_write:
            from excelmcp._io import write_cell

            result = write_cell(str(wb_file), "Sheet1", "A1", "hello", confirm=True, session_mode=False)

        assert result["ok"] is True
        mock_com_write.assert_not_called()

    def test_f02_write_cell_session_true_calls_com(self, monkeypatch, tmp_path):
        """F-02: write_cell(session_mode=True) calls _write_cell_via_com."""
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))

        wb_file = tmp_path / "test.xlsx"

        with patch("excelmcp._active_wb._write_cell_via_com", return_value={"ok": True, "previous_value": None}) as mock_com:
            from excelmcp._io import write_cell

            result = write_cell(str(wb_file), "Sheet1", "A1", "hello", confirm=True, session_mode=True)

        mock_com.assert_called_once()
        assert result["ok"] is True

    def test_f03_write_range_session_false_uses_openpyxl(self, monkeypatch, tmp_path):
        """F-03: write_range(session_mode=False) takes the openpyxl path (regression)."""
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))

        wb_file = tmp_path / "test.xlsx"
        import openpyxl
        _wb = openpyxl.Workbook()
        _wb.active.title = "Sheet1"
        _wb.save(str(wb_file))

        with patch("excelmcp._active_wb._write_range_via_com") as mock_com_write:
            from excelmcp._io import write_range

            result = write_range(str(wb_file), "Sheet1", "A1", [["x", "y"]], confirm=True, session_mode=False)

        assert result["ok"] is True
        mock_com_write.assert_not_called()

    def test_f04_write_range_session_true_calls_com(self, monkeypatch, tmp_path):
        """F-04: write_range(session_mode=True) calls _write_range_via_com."""
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))

        wb_file = tmp_path / "test.xlsx"

        with patch("excelmcp._active_wb._write_range_via_com", return_value={"ok": True, "cells_written": 2}) as mock_com:
            from excelmcp._io import write_range

            result = write_range(str(wb_file), "Sheet1", "A1", [["a", "b"]], confirm=True, session_mode=True)

        mock_com.assert_called_once()
        assert result["cells_written"] == 2

    def test_f05_formula_injection_blocked_before_session_routing(self, monkeypatch, tmp_path):
        """F-05: Formula injection blocked even with session_mode=True."""
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))

        from excelmcp._io import write_cell

        with pytest.raises(ValidationError, match="Formula injection"):
            write_cell(str(tmp_path / "wb.xlsx"), "Sheet1", "A1", "=SUM(A1:A99)", confirm=True, session_mode=True)

    def test_f06_enable_write_gate_blocks_session_mode(self, monkeypatch, tmp_path):
        """F-06: EXCEL_ENABLE_WRITE not set → NotAllowedError even with session_mode=True."""
        monkeypatch.delenv("EXCEL_ENABLE_WRITE", raising=False)

        from excelmcp._io import write_cell

        with pytest.raises(NotAllowedError):
            write_cell(str(tmp_path / "wb.xlsx"), "Sheet1", "A1", "hello", confirm=True, session_mode=True)

    def test_f07_confirm_false_blocks_session_mode(self, monkeypatch, tmp_path):
        """F-07: confirm=False → ValidationError even with session_mode=True."""
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")

        from excelmcp._io import write_cell

        with pytest.raises(ValidationError, match="confirm=True required"):
            write_cell(str(tmp_path / "wb.xlsx"), "Sheet1", "A1", "hello", confirm=False, session_mode=True)


# ---------------------------------------------------------------------------
# G-series: server_io.py integration smoke tests
# ---------------------------------------------------------------------------

class TestServerIoSmoke:

    def test_g01_cell_write_session_mode_removed(self, monkeypatch, tmp_path):
        """G-01: cell() no longer accepts session_mode — Phase 3D stub removal.

        Verifies that 'session_mode' is not in cell's signature.
        """
        import inspect
        from excelmcp.server_io import cell
        params = list(inspect.signature(cell).parameters.keys())
        assert "session_mode" not in params, (
            "session_mode stub must be removed from cell() per Phase 3D spec"
        )

    def test_g02_range_io_session_mode_removed(self, monkeypatch, tmp_path):
        """G-02: range_io() no longer accepts session_mode — Phase 3D stub removal."""
        import inspect
        from excelmcp.server_io import range_io
        params = list(inspect.signature(range_io).parameters.keys())
        assert "session_mode" not in params, (
            "session_mode stub must be removed from range_io() per Phase 3D spec"
        )

    def test_g03_cell_read_uses_openpyxl(self, monkeypatch, tmp_path):
        """G-03: cell(operation='read') uses openpyxl path."""
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))

        wb_file = tmp_path / "wb.xlsx"
        import openpyxl
        _wb = openpyxl.Workbook()
        ws = _wb.active
        ws.title = "Sheet1"
        ws["B3"] = 777
        _wb.save(str(wb_file))

        with patch("excelmcp._active_wb._write_cell_via_com") as mock_com:
            from excelmcp.server_io import cell

            result = cell(operation="read", path=str(wb_file), sheet="Sheet1", address="B3")

        mock_com.assert_not_called()
        assert result["value"] == 777

    def test_g04_create_workbook_session_mode_removed(self, monkeypatch, tmp_path):
        """G-04: create_workbook() no longer accepts session_mode — Phase 3D stub removal."""
        import inspect
        from excelmcp.server_io import create_workbook
        params = list(inspect.signature(create_workbook).parameters.keys())
        assert "session_mode" not in params, (
            "session_mode stub must be removed from create_workbook() per Phase 3D spec"
        )

    def test_g05_capabilities_contains_enable_com_env_key(self):
        """G-05: capabilities() returns a dict with governance.enable_com_env key."""
        from excelmcp._io import capabilities

        result = capabilities()

        assert "governance" in result
        assert "enable_com_env" in result["governance"]
        assert result["governance"]["enable_com_env"] == "EXCEL_ENABLE_COM"
