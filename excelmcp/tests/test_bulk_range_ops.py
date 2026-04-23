import pytest
import openpyxl

from excelmcp.workbook_openpyxl import (
    bulk_copy_sheet,
    set_row_heights,
    apply_style_to_ranges,
    set_column_widths,
    apply_format_to_sheet_list,
    ValidationError,
    _wb_cache,
)


class TestBulkCopySheet:
    """Phase 2.7 Sprint C — bulk_copy_sheet."""

    def _make_wb(self, tmp_path, monkeypatch, sheets=("Sheet1",)):
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        wb_path = tmp_path / "bcs.xlsx"
        _wb = openpyxl.Workbook()
        _wb.active.title = sheets[0]
        for s in sheets[1:]:
            _wb.create_sheet(s)
        _wb.save(str(wb_path))
        return str(wb_path)

    @pytest.mark.unit
    def test_bulk_copy_sheet_single_sheet(self, tmp_path, monkeypatch):
        path = self._make_wb(tmp_path, monkeypatch)
        result = bulk_copy_sheet(path, "Sheet1", ["Copy1"], confirm=True)
        assert result["ok"] is True
        assert result["sheets_created"] == 1
        assert result["names"] == ["Copy1"]
        _wb_cache.clear()
        wb = openpyxl.load_workbook(path)
        assert "Copy1" in wb.sheetnames

    @pytest.mark.unit
    def test_bulk_copy_sheet_multiple_sheets(self, tmp_path, monkeypatch):
        path = self._make_wb(tmp_path, monkeypatch)
        result = bulk_copy_sheet(path, "Sheet1", ["CopyA", "CopyB", "CopyC"], confirm=True)
        assert result["sheets_created"] == 3
        assert result["names"] == ["CopyA", "CopyB", "CopyC"]
        _wb_cache.clear()
        wb = openpyxl.load_workbook(path)
        for name in ["CopyA", "CopyB", "CopyC"]:
            assert name in wb.sheetnames

    @pytest.mark.unit
    def test_bulk_copy_sheet_insert_after(self, tmp_path, monkeypatch):
        path = self._make_wb(tmp_path, monkeypatch, sheets=("Sheet1", "Sheet2", "Sheet3"))
        result = bulk_copy_sheet(path, "Sheet1", ["New1", "New2"], insert_after="Sheet1", confirm=True)
        assert result["ok"] is True
        _wb_cache.clear()
        wb = openpyxl.load_workbook(path)
        idx_s1 = wb.sheetnames.index("Sheet1")
        idx_n1 = wb.sheetnames.index("New1")
        idx_n2 = wb.sheetnames.index("New2")
        assert idx_n1 == idx_s1 + 1
        assert idx_n2 == idx_s1 + 2

    @pytest.mark.unit
    def test_bulk_copy_sheet_source_not_found(self, tmp_path, monkeypatch):
        path = self._make_wb(tmp_path, monkeypatch)
        with pytest.raises(ValidationError, match="Source sheet not found"):
            bulk_copy_sheet(path, "NoSuch", ["Copy1"], confirm=True)

    @pytest.mark.unit
    def test_bulk_copy_sheet_empty_new_names(self, tmp_path, monkeypatch):
        path = self._make_wb(tmp_path, monkeypatch)
        with pytest.raises(ValidationError):
            bulk_copy_sheet(path, "Sheet1", [], confirm=True)

    @pytest.mark.unit
    def test_bulk_copy_sheet_duplicate_name_in_list(self, tmp_path, monkeypatch):
        path = self._make_wb(tmp_path, monkeypatch)
        with pytest.raises(ValidationError, match="[Dd]uplicate"):
            bulk_copy_sheet(path, "Sheet1", ["CopyA", "CopyA"], confirm=True)

    @pytest.mark.unit
    def test_bulk_copy_sheet_name_already_exists(self, tmp_path, monkeypatch):
        path = self._make_wb(tmp_path, monkeypatch, sheets=("Sheet1", "Existing"))
        with pytest.raises(ValidationError, match="already exists"):
            bulk_copy_sheet(path, "Sheet1", ["Existing"], confirm=True)

    @pytest.mark.unit
    def test_bulk_copy_sheet_atomicity_second_name_exists(self, tmp_path, monkeypatch):
        """Fail-fast: second name conflicts → first copy must NOT be created."""
        path = self._make_wb(tmp_path, monkeypatch, sheets=("Sheet1", "Existing"))
        with pytest.raises(ValidationError, match="already exists"):
            bulk_copy_sheet(path, "Sheet1", ["NewCopy", "Existing"], confirm=True)
        # Workbook must be unchanged — NewCopy must not have been created
        _wb_cache.clear()
        wb = openpyxl.load_workbook(path)
        assert "NewCopy" not in wb.sheetnames

    @pytest.mark.unit
    def test_bulk_copy_sheet_name_too_long(self, tmp_path, monkeypatch):
        path = self._make_wb(tmp_path, monkeypatch)
        long_name = "A" * 32
        with pytest.raises(ValidationError, match="31"):
            bulk_copy_sheet(path, "Sheet1", [long_name], confirm=True)


class TestSetRowHeights:
    """Phase 2.7 Sprint C — set_row_heights."""

    def _make_wb(self, tmp_path, monkeypatch):
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        wb_path = tmp_path / "srh.xlsx"
        _wb = openpyxl.Workbook()
        _wb.active.title = "Sheet1"
        _wb.save(str(wb_path))
        return str(wb_path)

    @pytest.mark.unit
    def test_set_row_heights_uniform_rows(self, tmp_path, monkeypatch):
        path = self._make_wb(tmp_path, monkeypatch)
        result = set_row_heights(path, "Sheet1", rows=[1, 2, 3], height=30.0, confirm=True)
        assert result["ok"] is True
        assert result["rows_set"] == 3
        _wb_cache.clear()
        wb = openpyxl.load_workbook(path)
        assert wb["Sheet1"].row_dimensions[1].height == 30.0

    @pytest.mark.unit
    def test_set_row_heights_row_range(self, tmp_path, monkeypatch):
        path = self._make_wb(tmp_path, monkeypatch)
        result = set_row_heights(path, "Sheet1", row_range="1:5", height=25.0, confirm=True)
        assert result["rows_set"] == 5
        _wb_cache.clear()
        wb = openpyxl.load_workbook(path)
        assert wb["Sheet1"].row_dimensions[3].height == 25.0

    @pytest.mark.unit
    def test_set_row_heights_per_row_specs(self, tmp_path, monkeypatch):
        path = self._make_wb(tmp_path, monkeypatch)
        specs = [{"row": 1, "height": 20.0}, {"row": 5, "height": 40.0}]
        result = set_row_heights(path, "Sheet1", row_specs=specs, confirm=True)
        assert result["rows_set"] == 2
        _wb_cache.clear()
        wb = openpyxl.load_workbook(path)
        assert wb["Sheet1"].row_dimensions[1].height == 20.0
        assert wb["Sheet1"].row_dimensions[5].height == 40.0

    @pytest.mark.unit
    def test_set_row_heights_both_modes_error(self, tmp_path, monkeypatch):
        path = self._make_wb(tmp_path, monkeypatch)
        with pytest.raises(ValidationError, match="mutually exclusive"):
            set_row_heights(path, "Sheet1", rows=[1], height=20.0,
                            row_specs=[{"row": 1, "height": 20.0}], confirm=True)

    @pytest.mark.unit
    def test_set_row_heights_neither_mode_error(self, tmp_path, monkeypatch):
        path = self._make_wb(tmp_path, monkeypatch)
        with pytest.raises(ValidationError):
            set_row_heights(path, "Sheet1", confirm=True)

    @pytest.mark.unit
    def test_set_row_heights_empty_rows(self, tmp_path, monkeypatch):
        path = self._make_wb(tmp_path, monkeypatch)
        with pytest.raises(ValidationError):
            set_row_heights(path, "Sheet1", rows=[], height=20.0, confirm=True)


class TestApplyStyleToRanges:
    """Phase 2.7 Sprint C — apply_style_to_ranges."""

    def _make_wb(self, tmp_path, monkeypatch):
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        wb_path = tmp_path / "astr.xlsx"
        _wb = openpyxl.Workbook()
        ws = _wb.active
        ws.title = "Sheet1"
        ws["A1"] = "hello"
        ws["B2"] = "world"
        _wb.save(str(wb_path))
        return str(wb_path)

    @pytest.mark.unit
    def test_apply_style_to_ranges_bold_multiple(self, tmp_path, monkeypatch):
        path = self._make_wb(tmp_path, monkeypatch)
        result = apply_style_to_ranges(path, "Sheet1", ["A1", "B2"], bold=True, confirm=True)
        assert result["ok"] is True
        assert result["ranges_processed"] == 2
        _wb_cache.clear()
        wb = openpyxl.load_workbook(path)
        assert wb["Sheet1"]["A1"].font.bold is True

    @pytest.mark.unit
    def test_apply_style_to_ranges_bg_color(self, tmp_path, monkeypatch):
        path = self._make_wb(tmp_path, monkeypatch)
        result = apply_style_to_ranges(path, "Sheet1", ["A1:B2"], bg_color="FF0000", confirm=True)
        assert result["cells_styled"] == 4
        _wb_cache.clear()
        wb = openpyxl.load_workbook(path)
        assert wb["Sheet1"]["A1"].fill.fgColor.rgb.endswith("FF0000")

    @pytest.mark.unit
    def test_apply_style_to_ranges_empty_ranges(self, tmp_path, monkeypatch):
        path = self._make_wb(tmp_path, monkeypatch)
        with pytest.raises(ValidationError):
            apply_style_to_ranges(path, "Sheet1", [], bold=True, confirm=True)

    @pytest.mark.unit
    def test_apply_style_to_ranges_all_none_style(self, tmp_path, monkeypatch):
        path = self._make_wb(tmp_path, monkeypatch)
        with pytest.raises(ValidationError, match="at least one style"):
            apply_style_to_ranges(path, "Sheet1", ["A1"], confirm=True)

    @pytest.mark.unit
    def test_apply_style_to_ranges_counts_cells(self, tmp_path, monkeypatch):
        path = self._make_wb(tmp_path, monkeypatch)
        result = apply_style_to_ranges(path, "Sheet1", ["A1:C3"], bold=True, confirm=True)
        assert result["cells_styled"] == 9

    @pytest.mark.unit
    def test_apply_style_to_ranges_invalid_range(self, tmp_path, monkeypatch):
        path = self._make_wb(tmp_path, monkeypatch)
        with pytest.raises(ValidationError, match="[Ii]nvalid range"):
            apply_style_to_ranges(path, "Sheet1", ["NOTARANGE!!!"], bold=True, confirm=True)


class TestSetColumnWidths:
    """Phase 2.7 Sprint C — set_column_widths."""

    def _make_wb(self, tmp_path, monkeypatch):
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        wb_path = tmp_path / "scw.xlsx"
        _wb = openpyxl.Workbook()
        ws = _wb.active
        ws.title = "Sheet1"
        ws["A1"] = "hello world"
        ws["B1"] = "short"
        _wb.save(str(wb_path))
        return str(wb_path)

    @pytest.mark.unit
    def test_set_column_widths_explicit_list(self, tmp_path, monkeypatch):
        path = self._make_wb(tmp_path, monkeypatch)
        result = set_column_widths(path, "Sheet1", columns=["A", "B"], width=15, confirm=True)
        assert result["ok"] is True
        assert result["columns_set"] == 2
        assert all(w["width"] == 15 for w in result["widths"])
        _wb_cache.clear()
        wb = openpyxl.load_workbook(path)
        assert wb["Sheet1"].column_dimensions["A"].width == 15

    @pytest.mark.unit
    def test_set_column_widths_col_range(self, tmp_path, monkeypatch):
        path = self._make_wb(tmp_path, monkeypatch)
        result = set_column_widths(path, "Sheet1", col_range="A:C", width=12, confirm=True)
        assert result["columns_set"] == 3
        assert [w["column"] for w in result["widths"]] == ["A", "B", "C"]

    @pytest.mark.unit
    def test_set_column_widths_autofit(self, tmp_path, monkeypatch):
        path = self._make_wb(tmp_path, monkeypatch)
        result = set_column_widths(path, "Sheet1", columns=["A"], autofit=True, confirm=True)
        assert result["ok"] is True
        w = result["widths"][0]["width"]
        assert 0 < w <= 60.0

    @pytest.mark.unit
    def test_set_column_widths_both_sources_error(self, tmp_path, monkeypatch):
        path = self._make_wb(tmp_path, monkeypatch)
        with pytest.raises(ValidationError, match="mutually exclusive"):
            set_column_widths(path, "Sheet1", columns=["A"], col_range="A:B", width=10, confirm=True)

    @pytest.mark.unit
    def test_set_column_widths_no_source_error(self, tmp_path, monkeypatch):
        path = self._make_wb(tmp_path, monkeypatch)
        with pytest.raises(ValidationError):
            set_column_widths(path, "Sheet1", width=10, confirm=True)

    @pytest.mark.unit
    def test_set_column_widths_autofit_with_width_error(self, tmp_path, monkeypatch):
        path = self._make_wb(tmp_path, monkeypatch)
        with pytest.raises(ValidationError, match="mutually exclusive"):
            set_column_widths(path, "Sheet1", columns=["A"], autofit=True, width=10, confirm=True)

    @pytest.mark.unit
    def test_set_column_widths_empty_columns(self, tmp_path, monkeypatch):
        path = self._make_wb(tmp_path, monkeypatch)
        with pytest.raises(ValidationError):
            set_column_widths(path, "Sheet1", columns=[], width=10, confirm=True)


class TestApplyFormatToSheetList:
    """Phase 2.7 Sprint C — apply_format_to_sheet_list."""

    def _make_wb(self, tmp_path, monkeypatch, sheets=("Sheet1", "Sheet2")):
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        wb_path = tmp_path / "afsl.xlsx"
        _wb = openpyxl.Workbook()
        _wb.active.title = sheets[0]
        for s in sheets[1:]:
            _wb.create_sheet(s)
        # Add data for the address
        for s in _wb.sheetnames:
            _wb[s]["A1"] = "data"
        _wb.save(str(wb_path))
        return str(wb_path)

    @pytest.mark.unit
    def test_apply_format_to_sheet_list_all_sheets(self, tmp_path, monkeypatch):
        path = self._make_wb(tmp_path, monkeypatch, sheets=("Sheet1", "Sheet2"))
        result = apply_format_to_sheet_list(
            path, address="A1", style={"bold": True}, confirm=True
        )
        assert result["ok"] is True
        assert result["sheets_processed"] == 2

    @pytest.mark.unit
    def test_apply_format_to_sheet_list_explicit_sheets(self, tmp_path, monkeypatch):
        path = self._make_wb(tmp_path, monkeypatch, sheets=("Sheet1", "Sheet2", "Sheet3"))
        result = apply_format_to_sheet_list(
            path, sheets=["Sheet1", "Sheet2"], address="A1",
            style={"bold": True}, confirm=True
        )
        assert result["sheets_processed"] == 2

    @pytest.mark.unit
    def test_apply_format_to_sheet_list_underscore_prefix_skipped(self, tmp_path, monkeypatch):
        path = self._make_wb(tmp_path, monkeypatch, sheets=("Sheet1", "_Lists"))
        result = apply_format_to_sheet_list(
            path, address="A1", style={"bold": True}, confirm=True
        )
        # _Lists should be skipped when sheets=None
        assert result["sheets_processed"] == 1

    @pytest.mark.unit
    def test_apply_format_to_sheet_list_underscore_explicit(self, tmp_path, monkeypatch):
        path = self._make_wb(tmp_path, monkeypatch, sheets=("Sheet1", "_Lists"))
        result = apply_format_to_sheet_list(
            path, sheets=["_Lists"], address="A1", style={"bold": True}, confirm=True
        )
        # Explicit naming includes _Lists
        assert result["sheets_processed"] == 1
        _wb_cache.clear()
        wb = openpyxl.load_workbook(path)
        assert wb["_Lists"]["A1"].font.bold is True

    @pytest.mark.unit
    def test_apply_format_to_sheet_list_zero_sheets(self, tmp_path, monkeypatch):
        # Workbook with only underscore-prefixed sheets → graceful zero result
        path = self._make_wb(tmp_path, monkeypatch, sheets=("_Config",))
        result = apply_format_to_sheet_list(
            path, address="A1", style={"bold": True}, confirm=True
        )
        assert result["ok"] is True
        assert result["sheets_processed"] == 0
        assert result["cells_formatted"] == 0

    @pytest.mark.unit
    def test_apply_format_to_sheet_list_no_params_error(self, tmp_path, monkeypatch):
        path = self._make_wb(tmp_path, monkeypatch)
        with pytest.raises(ValidationError, match="address"):
            apply_format_to_sheet_list(
                path, style={"bold": True}, confirm=True  # no address
            )

    @pytest.mark.unit
    def test_apply_format_to_sheet_list_unknown_style_key(self, tmp_path, monkeypatch):
        path = self._make_wb(tmp_path, monkeypatch)
        with pytest.raises(ValidationError, match="Unknown style key"):
            apply_format_to_sheet_list(
                path, address="A1", style={"unknown_key": True}, confirm=True
            )


class TestApplyFormatToSheetListV2:
    """Unit tests for column_widths and row_heights params (T9v2, Sprint D Batch 2)."""

    def _make_wb(self, tmp_path, monkeypatch, sheets=("Sheet1", "Sheet2")):
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        path = tmp_path / "afsl_v2.xlsx"
        import openpyxl as _xl
        _wb = _xl.Workbook()
        _wb.active.title = sheets[0]
        for s in sheets[1:]:
            _wb.create_sheet(s)
        _wb.save(str(path))
        return str(path)

    @pytest.mark.unit
    def test_column_widths_single(self, tmp_path, monkeypatch):
        """{'A': 15} sets column A width on 2 sheets; column_widths_set == 2."""
        path = self._make_wb(tmp_path, monkeypatch, sheets=("Sheet1", "Sheet2"))
        result = apply_format_to_sheet_list(
            path, column_widths={"A": 15}, confirm=True
        )
        assert result["ok"] is True
        assert result["column_widths_set"] == 2

        _wb_cache.clear()
        import openpyxl as _xl
        wb_disk = _xl.load_workbook(path)
        for sn in ("Sheet1", "Sheet2"):
            assert wb_disk[sn].column_dimensions["A"].width == 15

    @pytest.mark.unit
    def test_column_widths_range(self, tmp_path, monkeypatch):
        """{'A:C': 10} sets columns A, B, C on each of 2 sheets; total == 6."""
        path = self._make_wb(tmp_path, monkeypatch, sheets=("Sheet1", "Sheet2"))
        result = apply_format_to_sheet_list(
            path, column_widths={"A:C": 10}, confirm=True
        )
        assert result["column_widths_set"] == 6  # 3 cols × 2 sheets

        _wb_cache.clear()
        import openpyxl as _xl
        wb_disk = _xl.load_workbook(path)
        for sn in ("Sheet1", "Sheet2"):
            for col in ("A", "B", "C"):
                assert wb_disk[sn].column_dimensions[col].width == 10

    @pytest.mark.unit
    def test_row_heights_single(self, tmp_path, monkeypatch):
        """{'1': 25} sets row 1 height on each of 2 sheets; row_heights_set == 2."""
        path = self._make_wb(tmp_path, monkeypatch, sheets=("Sheet1", "Sheet2"))
        result = apply_format_to_sheet_list(
            path, row_heights={"1": 25}, confirm=True
        )
        assert result["ok"] is True
        assert result["row_heights_set"] == 2

        _wb_cache.clear()
        import openpyxl as _xl
        wb_disk = _xl.load_workbook(path)
        for sn in ("Sheet1", "Sheet2"):
            assert wb_disk[sn].row_dimensions[1].height == 25

    @pytest.mark.unit
    def test_row_heights_range(self, tmp_path, monkeypatch):
        """{'2:5': 18} sets rows 2–5 on each of 2 sheets; total == 8."""
        path = self._make_wb(tmp_path, monkeypatch, sheets=("Sheet1", "Sheet2"))
        result = apply_format_to_sheet_list(
            path, row_heights={"2:5": 18}, confirm=True
        )
        assert result["row_heights_set"] == 8  # 4 rows × 2 sheets

        _wb_cache.clear()
        import openpyxl as _xl
        wb_disk = _xl.load_workbook(path)
        for sn in ("Sheet1", "Sheet2"):
            for r in range(2, 6):
                assert wb_disk[sn].row_dimensions[r].height == 18
