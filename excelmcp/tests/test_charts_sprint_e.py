"""Sprint E chart tests: update_chart_data + create_chart anchor/size extensions.

Tests for the new update_chart_data function and the anchor/width_px/height_px
extensions to create_chart added in Sprint E.
"""
from __future__ import annotations

import pytest
import openpyxl
from openpyxl.chart import BarChart, Reference

from excelmcp._core import ExcelMCPError, ValidationError, NotAllowedError
from excelmcp._chart import create_chart, update_chart_data


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture()
def wb_with_chart(tmp_path, monkeypatch):
    """Workbook with numeric data and one BarChart anchored at D1."""
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")

    path = str(tmp_path / "chart_test.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for row in range(1, 6):
        ws.cell(row=row, column=1, value=row)
        ws.cell(row=row, column=2, value=row * 10)
    # Add a BarChart manually so we can update it
    chart = BarChart()
    data_ref = Reference(ws, min_col=1, min_row=1, max_col=2, max_row=5)
    chart.add_data(data_ref, titles_from_data=True)
    ws.add_chart(chart, "D1")
    wb.save(path)

    # Clear cache so the saved file is freshly loaded
    from excelmcp._core import _wb_cache
    _wb_cache.clear()

    return path


@pytest.fixture()
def wb_empty(tmp_path, monkeypatch):
    """Workbook with numeric data but no charts."""
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")

    path = str(tmp_path / "empty_chart.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for row in range(1, 6):
        ws.cell(row=row, column=1, value=row)
        ws.cell(row=row, column=2, value=row * 10)
    wb.save(path)

    from excelmcp._core import _wb_cache
    _wb_cache.clear()

    return path


# ---------------------------------------------------------------------------
# TestUpdateChartData
# ---------------------------------------------------------------------------

class TestUpdateChartData:
    """Tests for update_chart_data."""

    @pytest.mark.unit
    def test_ok_replaces_data_range(self, wb_with_chart):
        """Happy path: update_chart_data returns ok result with expected fields."""
        result = update_chart_data(wb_with_chart, "Sheet1", 0, "A1:B5", confirm=True)
        assert result["ok"] is True
        assert result["chart_index"] == 0
        assert "A1:B5" in result["data_range"]
        assert "BarChart" in result["chart_type"]

    @pytest.mark.unit
    def test_requires_write_env(self, wb_with_chart, monkeypatch):
        """EXCEL_ENABLE_WRITE not set → NotAllowedError before touching disk."""
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "false")
        with pytest.raises(NotAllowedError):
            update_chart_data(wb_with_chart, "Sheet1", 0, "A1:B5", confirm=True)

    @pytest.mark.unit
    def test_requires_confirm(self, wb_with_chart):
        """confirm=False → ValidationError."""
        with pytest.raises(ValidationError):
            update_chart_data(wb_with_chart, "Sheet1", 0, "A1:B5", confirm=False)

    @pytest.mark.unit
    def test_single_cell_data_range_rejected(self, wb_with_chart):
        """data_range without ':' → ExcelMCPError matching 'must be a range'."""
        with pytest.raises(ExcelMCPError, match="must be a range"):
            update_chart_data(wb_with_chart, "Sheet1", 0, "A1", confirm=True)

    @pytest.mark.unit
    def test_invalid_data_range_raises(self, wb_with_chart):
        """Unparseable data_range → ExcelMCPError matching 'Invalid data_range'."""
        with pytest.raises(ExcelMCPError, match="Invalid data_range"):
            update_chart_data(wb_with_chart, "Sheet1", 0, "!@#:$%^", confirm=True)

    @pytest.mark.unit
    def test_chart_index_out_of_bounds(self, wb_empty):
        """chart_index=99 on sheet with no charts → ExcelMCPError matching 'out of range'."""
        with pytest.raises(ExcelMCPError, match="out of range"):
            update_chart_data(wb_empty, "Sheet1", 99, "A1:B5", confirm=True)

    @pytest.mark.unit
    def test_sheet_not_found(self, wb_with_chart):
        """Unknown sheet name → ValidationError matching 'Sheet not found'."""
        with pytest.raises(ValidationError, match="Sheet not found"):
            update_chart_data(wb_with_chart, "NoSuchSheet", 0, "A1:B5", confirm=True)


# ---------------------------------------------------------------------------
# TestCreateChartExtensions
# ---------------------------------------------------------------------------

class TestCreateChartExtensions:
    """Tests for the anchor, width_px, height_px extensions to create_chart."""

    @pytest.fixture()
    def workbook(self, tmp_path, monkeypatch):
        """Simple workbook with data."""
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")

        path = str(tmp_path / "ext_test.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        for row in range(1, 6):
            ws.cell(row=row, column=1, value=row)
            ws.cell(row=row, column=2, value=row * 5)
        wb.save(path)

        from excelmcp._core import _wb_cache
        _wb_cache.clear()

        return path

    @pytest.mark.unit
    def test_default_params_backward_compat(self, workbook):
        """No new params → ok=True, target_cell='D1' (original default)."""
        result = create_chart(workbook, "Sheet1", "bar", "A1:B5", confirm=True)
        assert result["ok"] is True
        assert result["target_cell"] == "D1"

    @pytest.mark.unit
    def test_anchor_overrides_target_cell(self, workbook):
        """anchor='F5' → result target_cell is 'F5', not 'D1'."""
        result = create_chart(
            workbook, "Sheet1", "bar", "A1:B5", anchor="F5", confirm=True
        )
        assert result["ok"] is True
        assert result["target_cell"] == "F5"

    @pytest.mark.unit
    def test_custom_dimensions(self, workbook):
        """width_px=400, height_px=300 → width_cm and height_cm match conversion."""
        expected_w = round(400 * 2.54 / 96, 4)
        expected_h = round(300 * 2.54 / 96, 4)
        result = create_chart(
            workbook, "Sheet1", "bar", "A1:B5",
            width_px=400, height_px=300, confirm=True,
        )
        assert result["ok"] is True
        assert result["width_cm"] == expected_w
        assert result["height_cm"] == expected_h

    @pytest.mark.unit
    def test_invalid_anchor_format_raises(self, workbook):
        """anchor='NOT!VALID' → ExcelMCPError about bare cell reference."""
        with pytest.raises(ExcelMCPError, match="bare cell reference"):
            create_chart(
                workbook, "Sheet1", "bar", "A1:B5",
                anchor="NOT!VALID", confirm=True,
            )

    @pytest.mark.unit
    def test_invalid_px_values_raise(self, workbook):
        """width_px=0 → ExcelMCPError; height_px=-5 → ExcelMCPError."""
        with pytest.raises(ExcelMCPError, match="width_px"):
            create_chart(
                workbook, "Sheet1", "bar", "A1:B5",
                width_px=0, confirm=True,
            )
        with pytest.raises(ExcelMCPError, match="height_px"):
            create_chart(
                workbook, "Sheet1", "bar", "A1:B5",
                height_px=-5, confirm=True,
            )
