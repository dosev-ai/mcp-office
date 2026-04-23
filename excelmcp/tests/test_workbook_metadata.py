import json
from typing import cast

import openpyxl
import pytest
from fastmcp.exceptions import ToolError

from excelmcp.server_io import workbook_metadata
from excelmcp.workbook_openpyxl import (
    ValidationError,
    _wb_cache,
    read_workbook_metadata,
    resolve_workbook_metadata_resource,
    write_workbook_metadata,
)


def _create_workbook(path):
    wb = openpyxl.Workbook()
    try:
        wb.save(path)
    finally:
        wb.close()


def _write_raw_metadata(path, raw_payload: str):
    wb = openpyxl.load_workbook(path)
    try:
        if "_MCP_META" in wb.sheetnames:
            del wb["_MCP_META"]
        ws = wb.create_sheet("_MCP_META")
        ws.sheet_state = "hidden"
        ws["A1"] = "EXCELMCP_WORKBOOK_METADATA"
        ws["A2"] = "v1"
        ws["A3"] = raw_payload
        wb.save(path)
    finally:
        wb.close()


@pytest.mark.unit
def test_read_workbook_metadata_absent_sheet_returns_not_present(tmp_path, monkeypatch):
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    path = tmp_path / "meta_absent.xlsx"
    _create_workbook(path)

    result = read_workbook_metadata(str(path))

    assert result == {
        "present": False,
        "sheet_name": "_MCP_META",
        "visibility": None,
        "version": None,
        "metadata": None,
        "bytes": 0,
    }
    wb = openpyxl.load_workbook(path)
    try:
        assert "_MCP_META" not in wb.sheetnames
    finally:
        wb.close()


@pytest.mark.unit
def test_write_workbook_metadata_round_trip(tmp_path, monkeypatch):
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    path = tmp_path / "meta_roundtrip.xlsx"
    _create_workbook(path)
    payload = {"owner": "excel", "flags": {"approved": True}}
    expected_bytes = len(
        json.dumps(payload, separators=(",", ":"), ensure_ascii=False, allow_nan=False).encode("utf-8")
    )

    write_result = write_workbook_metadata(str(path), payload, confirm=True)
    read_result = read_workbook_metadata(str(path))

    assert write_result == {
        "ok": True,
        "path": str(path.resolve()),
        "sheet_name": "_MCP_META",
        "created": True,
        "replaced": False,
        "visibility": "hidden",
        "version": "v1",
        "bytes_written": expected_bytes,
    }
    assert read_result == {
        "present": True,
        "sheet_name": "_MCP_META",
        "visibility": "hidden",
        "version": "v1",
        "metadata": payload,
        "bytes": expected_bytes,
    }

    wb = openpyxl.load_workbook(path)
    try:
        ws = wb["_MCP_META"]
        assert ws.sheet_state == "hidden"
        assert ws["A1"].value == "EXCELMCP_WORKBOOK_METADATA"
        assert ws["A2"].value == "v1"
        assert ws["A3"].value == json.dumps(payload, separators=(",", ":"), ensure_ascii=False, allow_nan=False)
    finally:
        wb.close()


@pytest.mark.unit
def test_write_workbook_metadata_replace_existing_sheet(tmp_path, monkeypatch):
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    path = tmp_path / "meta_replace.xlsx"
    _create_workbook(path)

    write_workbook_metadata(str(path), {"version": 1}, confirm=True)
    result = write_workbook_metadata(str(path), {"version": 2}, confirm=True)

    assert result["created"] is False
    assert result["replaced"] is True
    assert read_workbook_metadata(str(path))["metadata"] == {"version": 2}


@pytest.mark.unit
def test_write_workbook_metadata_collision_fails_closed(tmp_path, monkeypatch):
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    path = tmp_path / "meta_collision_write.xlsx"
    wb = openpyxl.Workbook()
    try:
        ws = wb.create_sheet("_MCP_META")
        ws["A1"] = "NOT_EXCELMCP"
        wb.save(path)
    finally:
        wb.close()

    with pytest.raises(ValidationError, match="collision"):
        write_workbook_metadata(str(path), {"x": 1}, confirm=True)


@pytest.mark.unit
def test_write_workbook_metadata_collision_evicts_write_cache(tmp_path, monkeypatch):
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    path = tmp_path / "meta_collision_evict.xlsx"
    wb = openpyxl.Workbook()
    try:
        ws = wb.create_sheet("_MCP_META")
        ws["A1"] = "NOT_EXCELMCP"
        wb.save(path)
    finally:
        wb.close()

    with pytest.raises(ValidationError, match="collision"):
        write_workbook_metadata(str(path), {"x": 1}, confirm=True)

    assert (str(path.resolve()), True) not in _wb_cache


@pytest.mark.unit
def test_read_workbook_metadata_collision_fails_closed(tmp_path, monkeypatch):
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    path = tmp_path / "meta_collision_read.xlsx"
    wb = openpyxl.Workbook()
    try:
        ws = wb.create_sheet("_MCP_META")
        ws["A1"] = "NOT_EXCELMCP"
        wb.save(path)
    finally:
        wb.close()

    with pytest.raises(ValidationError, match="collision"):
        read_workbook_metadata(str(path))


@pytest.mark.unit
def test_write_workbook_metadata_confirm_checked_before_payload(tmp_path, monkeypatch):
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))

    with pytest.raises(ValidationError, match="confirm=True required"):
        write_workbook_metadata(
            str(tmp_path / "ordered.xlsx"),
            cast(dict, []),
            confirm=False,
        )


@pytest.mark.unit
def test_write_workbook_metadata_payload_validated_before_path(monkeypatch, tmp_path):
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    monkeypatch.delenv("EXCEL_ALLOWLIST_ROOTS", raising=False)

    with pytest.raises(ValidationError, match="top-level JSON object"):
        write_workbook_metadata(
            str(tmp_path / "outside.xlsx"),
            cast(dict, []),
            confirm=True,
        )


@pytest.mark.unit
@pytest.mark.parametrize("non_finite_value", [float("nan"), float("inf"), float("-inf")])
def test_write_workbook_metadata_rejects_non_finite_payload(monkeypatch, tmp_path, non_finite_value):
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    monkeypatch.delenv("EXCEL_ALLOWLIST_ROOTS", raising=False)

    with pytest.raises(ValidationError, match="Invalid workbook metadata payload"):
        write_workbook_metadata(
            str(tmp_path / "outside.xlsx"),
            {"value": non_finite_value},
            confirm=True,
        )


@pytest.mark.unit
def test_write_workbook_metadata_size_limit_checked_before_path(monkeypatch, tmp_path):
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    monkeypatch.setenv("EXCEL_MAX_METADATA_BYTES", "8")
    monkeypatch.delenv("EXCEL_ALLOWLIST_ROOTS", raising=False)

    with pytest.raises(ValidationError, match="exceeds limit 8 bytes"):
        write_workbook_metadata(str(tmp_path / "outside.xlsx"), {"abcdef": "ghijkl"}, confirm=True)


@pytest.mark.unit
def test_write_workbook_metadata_single_cell_limit_checked_before_path(monkeypatch, tmp_path):
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    monkeypatch.setenv("EXCEL_MAX_METADATA_BYTES", "50000")
    monkeypatch.delenv("EXCEL_ALLOWLIST_ROOTS", raising=False)

    with pytest.raises(ValidationError, match="single-cell limit 32767 chars"):
        write_workbook_metadata(
            str(tmp_path / "outside.xlsx"),
            {"payload": "a" * 40000},
            confirm=True,
        )


@pytest.mark.unit
def test_write_workbook_metadata_invalid_size_limit_env_rejected(monkeypatch, tmp_path):
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    monkeypatch.setenv("EXCEL_MAX_METADATA_BYTES", "not-an-int")
    monkeypatch.delenv("EXCEL_ALLOWLIST_ROOTS", raising=False)

    with pytest.raises(ValidationError, match="Invalid EXCEL_MAX_METADATA_BYTES"):
        write_workbook_metadata(
            str(tmp_path / "outside.xlsx"),
            {"ok": True},
            confirm=True,
        )


@pytest.mark.unit
def test_read_workbook_metadata_enforces_size_limit(tmp_path, monkeypatch):
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    monkeypatch.setenv("EXCEL_MAX_METADATA_BYTES", "8")
    path = tmp_path / "meta_oversized.xlsx"
    _create_workbook(path)
    _write_raw_metadata(path, json.dumps({"abcdef": "ghijkl"}, separators=(",", ":")))

    with pytest.raises(ValidationError, match="exceeds limit 8 bytes"):
        read_workbook_metadata(str(path))


@pytest.mark.unit
def test_workbook_metadata_tool_write_single_cell_limit_uses_normal_error_flow(monkeypatch, tmp_path):
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    monkeypatch.setenv("EXCEL_MAX_METADATA_BYTES", "50000")
    monkeypatch.delenv("EXCEL_ALLOWLIST_ROOTS", raising=False)

    with pytest.raises(ToolError, match="single-cell limit 32767 chars"):
        workbook_metadata(
            operation="write",
            path=str(tmp_path / "outside.xlsx"),
            payload={"payload": "a" * 40000},
            confirm=True,
        )


@pytest.mark.unit
def test_workbook_metadata_tool_read_oversized_payload_uses_normal_error_flow(tmp_path, monkeypatch):
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    monkeypatch.setenv("EXCEL_MAX_METADATA_BYTES", "8")
    path = tmp_path / "meta_tool_oversized.xlsx"
    _create_workbook(path)
    _write_raw_metadata(path, json.dumps({"abcdef": "ghijkl"}, separators=(",", ":")))

    with pytest.raises(ToolError, match="exceeds limit 8 bytes"):
        workbook_metadata(operation="read", path=str(path))


@pytest.mark.unit
def test_workbook_metadata_resource_invalid_size_limit_uses_normal_error_flow(tmp_path, monkeypatch):
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    path = tmp_path / "meta_resource_invalid_env.xlsx"
    _create_workbook(path)
    _write_raw_metadata(path, json.dumps({"ok": True}, separators=(",", ":")))
    monkeypatch.setenv("EXCEL_MAX_METADATA_BYTES", "not-an-int")

    with pytest.raises((ToolError, ValidationError), match="Invalid EXCEL_MAX_METADATA_BYTES"):
        resolve_workbook_metadata_resource(str(path))


@pytest.mark.unit
def test_resolve_workbook_metadata_resource_matches_read_contract(tmp_path, monkeypatch):
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    path = tmp_path / "meta_resource.xlsx"
    _create_workbook(path)
    payload = {"hello": "world"}
    write_workbook_metadata(str(path), payload, confirm=True)

    assert resolve_workbook_metadata_resource(str(path)) == read_workbook_metadata(str(path))
