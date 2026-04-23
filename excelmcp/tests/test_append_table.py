"""Tests for Sprint I T1: _resolve_table_last_row helper and table-aware
append_rows behaviour.

All 7 tests run with `pytest -v` without a live COM object (openpyxl only).
No @pytest.mark.integration marker is required -- no COM path is touched.
"""
import pytest
import openpyxl
from openpyxl.worksheet.table import Table
from openpyxl.utils.cell import range_boundaries

from excelmcp._core import _resolve_table_last_row   # new helper -- Sprint I
from excelmcp.workbook_openpyxl import append_rows
from excelmcp.server import append_rows as append_rows_tool  # for .fn() smoke


# ---------------------------------------------------------------------------
# Module-level helper factory
# Used by tests #4, #5, #6, #7 that need a full file-based workbook.
# ---------------------------------------------------------------------------

def _make_table_workbook(tmp_path, monkeypatch, *, table_ref="A1:C5", stray_row=None):
    """Create a workbook with Sheet1 containing an Excel Table at `table_ref`.

    If stray_row is given, write a non-empty value at that row to simulate
    trailing whitespace below the table -- this is the append_rows BUG trigger
    (ws.max_row != table last row when stray data exists below the table).

    Returns the absolute path as a str.
    """
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    path = tmp_path / "table_test.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Write header + data rows to match the table ref
    min_col, min_row, max_col, max_row = range_boundaries(table_ref.upper())
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            ws.cell(row=r, column=c).value = f"R{r}C{c}"

    # Register the Excel Table
    tbl = Table(displayName="DataTable", ref=table_ref)
    ws.add_table(tbl)

    # Optionally create stray data below the table to inflate ws.max_row
    if stray_row is not None:
        ws.cell(row=stray_row, column=1).value = "stray"

    wb.save(str(path))
    return str(path)


# ---------------------------------------------------------------------------
# TEST #1
# Function:  test_resolve_table_last_row_returns_table_max_row_not_sheet_max_row
# Group:     T1 -- helper unit test
# W2 Blocker: (none -- primary regression test for the T1 fix)
# ---------------------------------------------------------------------------

@pytest.mark.unit
def test_resolve_table_last_row_returns_table_max_row_not_sheet_max_row():
    """_resolve_table_last_row(ws) returns the Excel Table's last row (5),
    NOT ws.max_row (500), when stray data exists below the table.

    This is the primary regression test proving the W1-identified BUG
    at _io.py:355 (`first_row = (ws.max_row or 0) + 1`) is fixed in the
    _resolve_table_last_row helper.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Table at A1:C5 -> table last row = 5
    for r in range(1, 6):
        for c in range(1, 4):
            ws.cell(row=r, column=c).value = f"v{r}{c}"
    tbl = Table(displayName="TestTable", ref="A1:C5")
    ws.add_table(tbl)

    # Stray data at row 500 -- causes ws.max_row == 500 (the BUG trigger)
    ws.cell(row=500, column=1).value = "stray"

    result = _resolve_table_last_row(ws)

    assert result == 5          # table max_row per tbl.ref, NOT 500
    assert result != ws.max_row  # confirm ws.max_row would give wrong answer


# ---------------------------------------------------------------------------
# TEST #2
# Function:  test_resolve_table_last_row_plain_sheet_fallback
# Group:     T1 -- helper unit test (backward compat)
# W2 Blocker: (none -- backward-compat acceptance criterion)
# ---------------------------------------------------------------------------

@pytest.mark.unit
def test_resolve_table_last_row_plain_sheet_fallback():
    """_resolve_table_last_row(ws) falls back to (ws.max_row or 0) when the
    sheet has NO Excel Table, preserving pre-Sprint-I behaviour for plain sheets.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    # Deliberately NO ws.add_table(...) call
    ws["A1"] = "row1"
    ws["A2"] = "row2"
    ws["A3"] = "row3"   # ws.max_row == 3

    result = _resolve_table_last_row(ws)

    assert result == 3   # equals ws.max_row; no table present


# ---------------------------------------------------------------------------
# TEST #3  * COVERS W2 BLOCKER 1 *
# Function:  test_resolve_table_last_row_none_ref_falls_back_to_max_row
# Group:     T1 -- BLOCKER regression test
# W2 Blocker: BLOCKER 1 (unguarded None/empty tbl.ref crashes _resolve_table_last_row)
# ---------------------------------------------------------------------------

@pytest.mark.unit
def test_resolve_table_last_row_none_ref_falls_back_to_max_row():
    """_resolve_table_last_row(ws) must NOT raise AttributeError when an Excel
    Table on the sheet has tbl.ref = None (malformed workbook).

    Validates the required guard in the helper:
        tbl_ref = tbl.ref or ""
        if not tbl_ref: continue

    Without the guard, `range_boundaries(None)` raises AttributeError.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Malformed table -- tbl.ref is None (exact W2 BLOCKER 1 scenario)
    tbl = Table(displayName="BadTable", ref=None)
    ws.add_table(tbl)

    # ws.max_row is set by real data; fallback should land here
    ws.cell(row=5, column=1).value = "data"   # ws.max_row == 5

    result = _resolve_table_last_row(ws)  # must NOT raise AttributeError

    assert result == 5   # fallback to ws.max_row


# ---------------------------------------------------------------------------
# TEST #4
# Function:  test_append_rows_table_sheet_anchors_after_table_last_row
# Group:     T1 -- end-to-end via workbook_openpyxl direct call + .fn() smoke
# W2 Blocker: (none -- primary T1 acceptance criterion)
# ---------------------------------------------------------------------------

@pytest.mark.unit
def test_append_rows_table_sheet_anchors_after_table_last_row(tmp_path, monkeypatch):
    """append_rows on a TABLE-backed sheet appends immediately after the last data
    row of the table (row 6), NOT after ws.max_row (row 500) when stray data exists
    below the table.

    Verifies result["first_row"] == 6 and that the new cell is at row 6 on disk.
    """
    # Table at A1:C5; stray data at row 500 -> ws.max_row == 500
    path = _make_table_workbook(tmp_path, monkeypatch,
                                table_ref="A1:C5", stray_row=500)

    # Primary assertion: direct workbook_openpyxl call
    result = append_rows(path, "Sheet1", [["X", "Y", "Z"]], confirm=True)

    assert result["ok"] is True
    assert result["rows_appended"] == 1
    assert result["first_row"] == 6    # table last row (5) + 1; NOT 501

    # Verify cell was actually written at row 6 on disk (not row 501)
    wb2 = openpyxl.load_workbook(path)
    ws2 = wb2["Sheet1"]
    assert ws2.cell(row=6, column=1).value == "X"
    assert ws2.cell(row=6, column=2).value == "Y"
    assert ws2.cell(row=6, column=3).value == "Z"

    # Optional server-layer smoke: verify .fn() path produces ok=True
    # (appends at row 7 -- after the row 6 already written above)
    # FastMCP 3.x: @mcp.tool() returns the original function; call it directly.
    result2 = append_rows_tool(
        path=path, sheet="Sheet1", rows=[["A", "B", "C"]], confirm=True)
    assert result2["ok"] is True


# ---------------------------------------------------------------------------
# TEST #5
# Function:  test_append_rows_plain_sheet_uses_ws_max_row_backward_compat
# Group:     T1 -- backward compat end-to-end
# W2 Blocker: (none -- backward-compat check)
# ---------------------------------------------------------------------------

@pytest.mark.unit
def test_append_rows_plain_sheet_uses_ws_max_row_backward_compat(tmp_path, monkeypatch):
    """append_rows on a PLAIN sheet (no Excel Table) continues to append at
    ws.max_row + 1, exactly as before Sprint I.  No regression introduced.
    """
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    path = tmp_path / "plain.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "row1"
    ws["A2"] = "row2"   # ws.max_row == 2; no table registered
    wb.save(str(path))

    result = append_rows(str(path), "Sheet1", [["new"]], confirm=True)

    assert result["ok"] is True
    assert result["first_row"] == 3   # ws.max_row (2) + 1 = 3

    # Confirm cell written at row 3 on disk
    wb2 = openpyxl.load_workbook(str(path))
    assert wb2["Sheet1"].cell(row=3, column=1).value == "new"


# ---------------------------------------------------------------------------
# TEST #6
# Function:  test_append_rows_expands_table_ref_after_write
# Group:     T1 -- tbl.ref expansion acceptance criterion
# W2 Blocker: (none -- table boundary must expand to avoid Excel corruption)
# ---------------------------------------------------------------------------

@pytest.mark.unit
def test_append_rows_expands_table_ref_after_write(tmp_path, monkeypatch):
    """After append_rows writes new rows, tbl.ref is expanded to encompass the
    newly appended rows.

    Without this expansion, Excel opens the file with a "content not readable /
    repaired" dialog because data sits outside the table's defined reference.

    Table starts at A1:C5.  3 rows appended -> expect tbl.ref == "A1:C8".
    """
    path = _make_table_workbook(tmp_path, monkeypatch, table_ref="A1:C5")

    append_rows(path, "Sheet1",
                [["a", "b", "c"], ["d", "e", "f"], ["g", "h", "i"]],
                confirm=True)

    wb2 = openpyxl.load_workbook(path)
    ws2 = wb2["Sheet1"]
    tables = list(ws2.tables.values())

    assert len(tables) == 1
    assert tables[0].ref == "A1:C8"   # expanded from A1:C5 by 3 rows


# ---------------------------------------------------------------------------
# TEST #7
# Function:  test_append_rows_row_count_matches_input
# Group:     T1 -- return value correctness
# W2 Blocker: (none -- rows_appended must always equal len(rows))
# ---------------------------------------------------------------------------

@pytest.mark.unit
def test_append_rows_row_count_matches_input(tmp_path, monkeypatch):
    """result["rows_appended"] always equals len(rows) for any input size.

    Verifies the return value is not accidentally hard-coded or off-by-one.
    4 rows are input; expected rows_appended == 4.
    """
    path = _make_table_workbook(tmp_path, monkeypatch, table_ref="A1:C3")

    rows_input = [
        ["r1c1", "r1c2", "r1c3"],
        ["r2c1", "r2c2", "r2c3"],
        ["r3c1", "r3c2", "r3c3"],
        ["r4c1", "r4c2", "r4c3"],
    ]
    result = append_rows(path, "Sheet1", rows_input, confirm=True)

    assert result["ok"] is True
    assert result["rows_appended"] == 4
