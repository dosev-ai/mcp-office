"""
Sprint N - Commit A tests.

US-N01: set_column_widths replaces set_column_width (single-col removed from MCP)
US-N02: set_row_heights replaces set_row_height (single-row removed from MCP)
US-N03: bulk_copy_sheet replaces copy_sheet (single-copy removed from MCP)
US-N04: stale tool names removed from prompt strings
US-N05: bulk_copy_sheet has 100-item cap

No COM required — openpyxl only.
"""
import openpyxl
import pytest
import re

from excelmcp._core import _wb_cache
from excelmcp._format import set_column_widths
from excelmcp._format import set_row_heights
from excelmcp._sheet import bulk_copy_sheet
from excelmcp._core import ValidationError, NotAllowedError

from ._smoke_helpers import _get_capabilities_snapshot, _get_tool_names, _render_prompt


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _make_workbook(tmp_path, monkeypatch, *, name="test.xlsx"):
    """Create a minimal xlsx in tmp_path, configure env, return str path."""
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    path = tmp_path / name
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    wb.save(str(path))
    return str(path)
# ---------------------------------------------------------------------------
# TestCommitAColumnWidth (US-N01)
# ---------------------------------------------------------------------------

class TestCommitAColumnWidth:
    """set_column_widths covers the single-column use-case correctly."""

    @pytest.mark.unit
    def test_set_column_widths_single_col_mode(self, tmp_path, monkeypatch):
        """columns=['A'], width=20 → ok=True, columns_set=1 (US-N01)."""
        path = _make_workbook(tmp_path, monkeypatch)
        _wb_cache.clear()
        result = set_column_widths(path, "Sheet1", columns=["A"], width=20, confirm=True)
        assert result.get("ok") is True
        assert result.get("columns_set") == 1

    @pytest.mark.unit
    def test_set_column_widths_multi_col_mode(self, tmp_path, monkeypatch):
        """columns=['A','B','C'], width=15 → columns_set=3 (US-N01)."""
        path = _make_workbook(tmp_path, monkeypatch)
        _wb_cache.clear()
        result = set_column_widths(
            path, "Sheet1", columns=["A", "B", "C"], width=15, confirm=True
        )
        assert result.get("ok") is True
        assert result.get("columns_set") == 3

    @pytest.mark.unit
    def test_set_column_widths_write_gate_blocks(self, tmp_path, monkeypatch):
        """EXCEL_ENABLE_WRITE unset → NotAllowedError (US-N17)."""
        path = _make_workbook(tmp_path, monkeypatch)
        _wb_cache.clear()
        monkeypatch.delenv("EXCEL_ENABLE_WRITE", raising=False)
        with pytest.raises(NotAllowedError):
            set_column_widths(path, "Sheet1", columns=["A"], width=20, confirm=True)

    @pytest.mark.unit
    def test_set_column_widths_confirm_false_blocks(self, tmp_path, monkeypatch):
        """confirm=False → ValidationError (US-N18)."""
        path = _make_workbook(tmp_path, monkeypatch)
        _wb_cache.clear()
        with pytest.raises((ValidationError, Exception)):
            set_column_widths(path, "Sheet1", columns=["A"], width=20, confirm=False)

    @pytest.mark.unit
    def test_set_column_widths_500_col_cap_ok(self, tmp_path, monkeypatch):
        """Exactly 26 columns is well below cap — just checks no error raised."""
        path = _make_workbook(tmp_path, monkeypatch)
        _wb_cache.clear()
        result = set_column_widths(path, "Sheet1", col_range="A:Z", width=8, confirm=True)
        assert result.get("ok") is True

    @pytest.mark.unit
    def test_set_column_widths_501_col_cap_raises(self, tmp_path, monkeypatch):
        """501 explicit columns → ValidationError: max 500 (US-N05 analogous)."""
        path = _make_workbook(tmp_path, monkeypatch)
        _wb_cache.clear()
        # Build 501 column letters (A..ZZ is 702, so use explicit list)
        from openpyxl.utils import get_column_letter
        cols_501 = [get_column_letter(i) for i in range(1, 502)]
        with pytest.raises(ValidationError, match="500"):
            set_column_widths(path, "Sheet1", columns=cols_501, width=8, confirm=True)


# ---------------------------------------------------------------------------
# TestCommitARowHeight (US-N02)
# ---------------------------------------------------------------------------

class TestCommitARowHeight:
    """set_row_heights covers the single-row use-case correctly."""

    @pytest.mark.unit
    def test_set_row_heights_single_row_mode(self, tmp_path, monkeypatch):
        """rows=[3], height=25 → ok=True, rows_set=1 (US-N02)."""
        path = _make_workbook(tmp_path, monkeypatch)
        _wb_cache.clear()
        result = set_row_heights(path, "Sheet1", rows=[3], height=25, confirm=True)
        assert result.get("ok") is True
        assert result.get("rows_set") == 1

    @pytest.mark.unit
    def test_set_row_heights_multi_row_mode(self, tmp_path, monkeypatch):
        """rows=[1,2,3], height=20 → rows_set=3 (US-N02)."""
        path = _make_workbook(tmp_path, monkeypatch)
        _wb_cache.clear()
        result = set_row_heights(path, "Sheet1", rows=[1, 2, 3], height=20, confirm=True)
        assert result.get("ok") is True
        assert result.get("rows_set") == 3

    @pytest.mark.unit
    def test_set_row_heights_write_gate_blocks(self, tmp_path, monkeypatch):
        """EXCEL_ENABLE_WRITE unset → NotAllowedError (US-N17)."""
        path = _make_workbook(tmp_path, monkeypatch)
        _wb_cache.clear()
        monkeypatch.delenv("EXCEL_ENABLE_WRITE", raising=False)
        with pytest.raises(NotAllowedError):
            set_row_heights(path, "Sheet1", rows=[1], height=20, confirm=True)

    @pytest.mark.unit
    def test_set_row_heights_confirm_false_blocks(self, tmp_path, monkeypatch):
        """confirm=False → ValidationError (US-N18)."""
        path = _make_workbook(tmp_path, monkeypatch)
        _wb_cache.clear()
        with pytest.raises((ValidationError, Exception)):
            set_row_heights(path, "Sheet1", rows=[1], height=20, confirm=False)


# ---------------------------------------------------------------------------
# TestCommitACopySheet (US-N03 + US-N05)
# ---------------------------------------------------------------------------

class TestCommitACopySheet:
    """bulk_copy_sheet covers the single-copy use-case and 100-item cap."""

    @pytest.mark.unit
    def test_bulk_copy_sheet_single_name_via_unified(self, tmp_path, monkeypatch):
        """new_names=['OnlyCopy'] → ok=True, sheets_created=1 (US-N03)."""
        path = _make_workbook(tmp_path, monkeypatch)
        _wb_cache.clear()
        result = bulk_copy_sheet(
            path, source_sheet="Sheet1", new_names=["OnlyCopy"], confirm=True
        )
        assert result.get("ok") is True
        assert result.get("sheets_created") == 1
        assert "OnlyCopy" in result.get("names", [])

    @pytest.mark.unit
    def test_bulk_copy_sheet_multi_name_via_unified(self, tmp_path, monkeypatch):
        """new_names=['C1','C2'] → sheets_created=2 (US-N03)."""
        path = _make_workbook(tmp_path, monkeypatch)
        _wb_cache.clear()
        result = bulk_copy_sheet(
            path, source_sheet="Sheet1", new_names=["C1", "C2"], confirm=True
        )
        assert result.get("ok") is True
        assert result.get("sheets_created") == 2

    @pytest.mark.unit
    def test_bulk_copy_sheet_write_gate_via_unified(self, tmp_path, monkeypatch):
        """EXCEL_ENABLE_WRITE unset → NotAllowedError (US-N17)."""
        path = _make_workbook(tmp_path, monkeypatch)
        _wb_cache.clear()
        monkeypatch.delenv("EXCEL_ENABLE_WRITE", raising=False)
        with pytest.raises(NotAllowedError):
            bulk_copy_sheet(
                path, source_sheet="Sheet1", new_names=["ShouldFail"], confirm=True
            )

    @pytest.mark.unit
    def test_bulk_copy_sheet_cap_101_raises(self, tmp_path, monkeypatch):
        """101 new_names → ValidationError: max 100 copies (US-N05)."""
        path = _make_workbook(tmp_path, monkeypatch)
        _wb_cache.clear()
        too_many = [f"Copy{i}" for i in range(101)]
        with pytest.raises(ValidationError, match="100"):
            bulk_copy_sheet(
                path, source_sheet="Sheet1", new_names=too_many, confirm=True
            )


# ---------------------------------------------------------------------------
# TestCommitAToolRegistry (MCP registry checks)
# ---------------------------------------------------------------------------

class TestCommitAToolRegistry:
    """Verify tool registry: 3 removed, 2 bulk equivalents still present."""

    @pytest.mark.smoke
    def test_set_column_width_not_registered(self):
        """set_column_width (singular) must NOT be in the MCP tool registry."""
        from excelmcp import server

        tool_names = _get_tool_names(server.mcp)
        assert "set_column_width" not in tool_names

    @pytest.mark.smoke
    def test_set_row_height_not_registered(self):
        """set_row_height (singular) must NOT be in the MCP tool registry."""
        from excelmcp import server

        tool_names = _get_tool_names(server.mcp)
        assert "set_row_height" not in tool_names

    @pytest.mark.smoke
    def test_copy_sheet_not_registered(self):
        """copy_sheet (singular) must NOT be in the MCP tool registry."""
        from excelmcp import server

        tool_names = _get_tool_names(server.mcp)
        assert "copy_sheet" not in tool_names

    @pytest.mark.smoke
    def test_set_column_widths_registered(self):
        """set_column_widths (bulk) must still be registered."""
        from excelmcp import server

        tool_names = _get_tool_names(server.mcp)
        caps = _get_capabilities_snapshot(server.mcp)
        assert "set_column_widths" in tool_names
        assert "set_column_widths" in caps["tools"]

    @pytest.mark.smoke
    def test_bulk_copy_sheet_registered(self):
        """bulk_copy_sheet must still be registered."""
        from excelmcp import server

        tool_names = _get_tool_names(server.mcp)
        caps = _get_capabilities_snapshot(server.mcp)
        assert "bulk_copy_sheet" in tool_names
        assert "bulk_copy_sheet" in caps["tools"]


# ---------------------------------------------------------------------------
# TestCommitAPrompts (US-N04 — stale tool name removal)
# ---------------------------------------------------------------------------

class TestCommitAPrompts:
    """Prompt strings must not reference the removed single-item tool names."""

    @pytest.mark.smoke
    def test_smoke_prompts_no_stale_single_tool_names(self):
        """suggest_formatting and prepare_for_print must not contain
        'set_column_width ' or 'set_row_height ' (old singular names, US-N04)."""
        from excelmcp import server

        suggest_text = _render_prompt(
            server.mcp,
            "suggest_formatting",
            {"sheet_json": "{}", "report_purpose": "status update"},
        )
        print_text = _render_prompt(
            server.mcp,
            "prepare_for_print",
            {"sheet_description": "single-sheet report", "page_size": "A4"},
        )

        stale_names = {"set_column_width", "set_row_height"}
        for stale_name in stale_names:
            assert not re.search(rf"\b{re.escape(stale_name)}\b", suggest_text), (
                f"suggest_formatting still references {stale_name}"
            )
            assert not re.search(rf"\b{re.escape(stale_name)}\b", print_text), (
                f"prepare_for_print still references {stale_name}"
            )


# ===========================================================================
# COMMIT B — US-N06–N10
# ===========================================================================

# ---------------------------------------------------------------------------
# TestCommitBHyperlinks (US-N06, US-N07)
# ---------------------------------------------------------------------------

