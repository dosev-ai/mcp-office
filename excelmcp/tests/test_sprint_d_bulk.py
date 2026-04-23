import pytest

from excelmcp.workbook_openpyxl import (
    bulk_add_comments,
    fill_formula_range,
    bulk_add_data_validation,
    ValidationError,
    NotAllowedError,
    _wb_cache,
)


class TestBulkAddComments:
    """Unit tests for bulk_add_comments (T7, Sprint D Batch 1)."""

    @pytest.mark.unit
    def test_bulk_add_comments_basic(self, tmp_path, monkeypatch):
        """Adding 2 comments returns comments_added=2 and correct cells list."""
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        path = tmp_path / "comments.xlsx"
        import openpyxl as _xl
        _wb = _xl.Workbook()
        _wb.active.title = "Sheet1"
        _wb.save(str(path))

        result = bulk_add_comments(
            str(path),
            "Sheet1",
            comments=[
                {"address": "A1", "text": "First comment", "author": "Alice"},
                {"address": "B2", "text": "Second comment", "author": "Bob"},
            ],
            confirm=True,
        )
        assert result["ok"] is True
        assert result["comments_added"] == 2
        assert result["skipped"] == 0
        assert "A1" in result["cells"]
        assert "B2" in result["cells"]

    @pytest.mark.unit
    def test_bulk_add_comments_overwrite_true(self, tmp_path, monkeypatch):
        """Second call with overwrite=True replaces the first comment on A1."""
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        path = tmp_path / "overwrite.xlsx"
        import openpyxl as _xl
        _wb = _xl.Workbook()
        _wb.active.title = "Sheet1"
        _wb.save(str(path))

        bulk_add_comments(
            str(path), "Sheet1",
            comments=[{"address": "A1", "text": "original", "author": "Alice"}],
            confirm=True,
        )
        _wb_cache.clear()
        result = bulk_add_comments(
            str(path), "Sheet1",
            comments=[{"address": "A1", "text": "replaced", "author": "Alice"}],
            overwrite=True,
            confirm=True,
        )
        assert result["comments_added"] == 1
        assert result["skipped"] == 0
        _wb_cache.clear()
        from openpyxl.comments import Comment  # noqa: F401
        wb_disk = _xl.load_workbook(str(path))
        assert wb_disk["Sheet1"]["A1"].comment.text == "replaced"

    @pytest.mark.unit
    def test_bulk_add_comments_overwrite_false(self, tmp_path, monkeypatch):
        """Second call with overwrite=False skips cells that already have a comment."""
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        path = tmp_path / "no_overwrite.xlsx"
        import openpyxl as _xl
        _wb = _xl.Workbook()
        _wb.active.title = "Sheet1"
        _wb.save(str(path))

        bulk_add_comments(
            str(path), "Sheet1",
            comments=[{"address": "A1", "text": "keep me", "author": "Alice"}],
            confirm=True,
        )
        _wb_cache.clear()
        result = bulk_add_comments(
            str(path), "Sheet1",
            comments=[
                {"address": "A1", "text": "skip me", "author": "Alice"},
                {"address": "B1", "text": "new one", "author": "Bob"},
            ],
            overwrite=False,
            confirm=True,
        )
        assert result["comments_added"] == 1
        assert result["skipped"] == 1

    @pytest.mark.unit
    def test_bulk_add_comments_empty_list_raises(self, sample_workbook, monkeypatch):
        """Empty comments list raises ValidationError."""
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        with pytest.raises(ValidationError, match="non-empty"):
            bulk_add_comments(sample_workbook, "Sheet1", comments=[], confirm=True)

    @pytest.mark.unit
    def test_bulk_add_comments_too_many_raises(self, sample_workbook, monkeypatch):
        """501 comments raises ValidationError (max is 500)."""
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        comments = [{"address": f"A{i+1}", "text": "x", "author": "A"} for i in range(501)]
        with pytest.raises(ValidationError, match="500"):
            bulk_add_comments(sample_workbook, "Sheet1", comments=comments, confirm=True)

    @pytest.mark.unit
    def test_bulk_add_comments_invalid_address_raises(self, sample_workbook, monkeypatch):
        """An unparseable cell address raises ValidationError."""
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        with pytest.raises(ValidationError, match="[Ii]nvalid"):
            bulk_add_comments(
                sample_workbook, "Sheet1",
                comments=[{"address": "NOTACELL!", "text": "hello", "author": "A"}],
                confirm=True,
            )

    @pytest.mark.unit
    def test_bulk_add_comments_empty_text_raises(self, sample_workbook, monkeypatch):
        """Empty text (after strip) raises ValidationError before any write."""
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        with pytest.raises(ValidationError, match="non-empty"):
            bulk_add_comments(
                sample_workbook, "Sheet1",
                comments=[
                    {"address": "A1", "text": "", "author": "Alice"},
                ],
                confirm=True,
            )

    @pytest.mark.unit
    def test_bulk_add_comments_text_too_long_raises(self, sample_workbook, monkeypatch):
        """Text over 32767 chars raises ValidationError before any write."""
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        long_text = "x" * 32768
        with pytest.raises(ValidationError, match="32767"):
            bulk_add_comments(
                sample_workbook, "Sheet1",
                comments=[{"address": "A1", "text": long_text, "author": "Alice"}],
                confirm=True,
            )

    @pytest.mark.unit
    def test_bulk_add_comments_confirm_false_blocked(self, sample_workbook, monkeypatch):
        """bulk_add_comments with confirm=False raises ValidationError (gap01)."""
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        with pytest.raises(ValidationError, match="confirm=True"):
            bulk_add_comments(
                sample_workbook, "Sheet1",
                comments=[{"address": "A1", "text": "hello", "author": "Alice"}],
                confirm=False,
            )


class TestFillFormulaRange:
    """Unit tests for fill_formula_range (T5, Sprint D Batch 1)."""

    def _make_wb(self, tmp_path, monkeypatch, rows=10, cols=3):
        """Create a workbook with numeric data in A2:A11 for formula tests."""
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        import openpyxl as _xl
        path = tmp_path / "formulas.xlsx"
        _wb = _xl.Workbook()
        ws = _wb.active
        ws.title = "Sheet1"
        for r in range(1, rows + 1):
            ws.cell(row=r, column=1, value=r)
        _wb.save(str(path))
        return str(path)

    @pytest.mark.unit
    def test_fill_formula_range_basic(self, tmp_path, monkeypatch):
        """Fill B2:B5 with =A2*2; B3 must equal the Translator-relativized formula."""
        from openpyxl.formula.translate import Translator as _T
        path = self._make_wb(tmp_path, monkeypatch)

        result = fill_formula_range(
            path, "Sheet1", anchor_cell="B2", end_cell="B5", formula="=A2*2", confirm=True
        )
        assert result["ok"] is True
        assert result["filled_cells"] == 4
        assert result["range"] == "B2:B5"

        _wb_cache.clear()
        import openpyxl as _xl
        wb_disk = _xl.load_workbook(path, data_only=False)
        expected_b3 = _T("=A2*2", origin="B2").translate_formula("B3")
        assert wb_disk["Sheet1"]["B3"].value == expected_b3

    @pytest.mark.unit
    def test_fill_formula_range_anchor_unchanged(self, tmp_path, monkeypatch):
        """The anchor cell must receive the literal formula without translation."""
        path = self._make_wb(tmp_path, monkeypatch)

        fill_formula_range(
            path, "Sheet1", anchor_cell="B2", end_cell="B5", formula="=A2*2", confirm=True
        )
        _wb_cache.clear()
        import openpyxl as _xl
        wb_disk = _xl.load_workbook(path, data_only=False)
        assert wb_disk["Sheet1"]["B2"].value == "=A2*2"

    @pytest.mark.unit
    def test_fill_formula_range_no_eq_raises(self, tmp_path, monkeypatch):
        """Formula without '=' prefix raises ValidationError."""
        path = self._make_wb(tmp_path, monkeypatch)
        with pytest.raises(ValidationError, match="'='"):
            fill_formula_range(
                path, "Sheet1", anchor_cell="B2", end_cell="B5",
                formula="A2*2", confirm=True
            )

    @pytest.mark.unit
    def test_fill_formula_range_max_cells_raises(self, tmp_path, monkeypatch):
        """Range exceeding EXCEL_MAX_RANGE_CELLS raises ValidationError."""
        monkeypatch.setenv("EXCEL_MAX_RANGE_CELLS", "5")
        path = self._make_wb(tmp_path, monkeypatch)
        with pytest.raises(ValidationError, match="[Ee]xceeds|[Tt]oo large|limit"):
            fill_formula_range(
                path, "Sheet1", anchor_cell="B2", end_cell="B10",
                formula="=A2*2", confirm=True
            )

    @pytest.mark.unit
    def test_fill_formula_range_absolute_ref(self, tmp_path, monkeypatch):
        """$A$1 in formula stays as-is in all cells (absolute reference unchanged)."""
        from openpyxl.formula.translate import Translator as _T
        path = self._make_wb(tmp_path, monkeypatch)

        fill_formula_range(
            path, "Sheet1", anchor_cell="B2", end_cell="B5",
            formula="=A2+$A$1", confirm=True
        )
        _wb_cache.clear()
        import openpyxl as _xl
        wb_disk = _xl.load_workbook(path, data_only=False)
        b4 = wb_disk["Sheet1"]["B4"].value
        expected = _T("=A2+$A$1", origin="B2").translate_formula("B4")
        assert b4 == expected
        assert "$A$1" in b4

    @pytest.mark.unit
    def test_fill_formula_range_mixed_ref(self, tmp_path, monkeypatch):
        """$A2 (mixed) relativizes the row but keeps the column fixed."""
        from openpyxl.formula.translate import Translator as _T
        path = self._make_wb(tmp_path, monkeypatch)

        fill_formula_range(
            path, "Sheet1", anchor_cell="B2", end_cell="C4",
            formula="=$A2*2", confirm=True
        )
        _wb_cache.clear()
        import openpyxl as _xl
        wb_disk = _xl.load_workbook(path, data_only=False)
        c3 = wb_disk["Sheet1"]["C3"].value
        expected = _T("=$A2*2", origin="B2").translate_formula("C3")
        assert c3 == expected

    @pytest.mark.unit
    def test_fill_formula_range_single_cell(self, tmp_path, monkeypatch):
        """anchor==end (single-cell range) → filled_cells=1, formula unchanged."""
        path = self._make_wb(tmp_path, monkeypatch)

        result = fill_formula_range(
            path, "Sheet1", anchor_cell="B2", end_cell="B2",
            formula="=A2*10", confirm=True
        )
        assert result["filled_cells"] == 1
        assert result["range"] == "B2:B2"

        _wb_cache.clear()
        import openpyxl as _xl
        wb_disk = _xl.load_workbook(path, data_only=False)
        assert wb_disk["Sheet1"]["B2"].value == "=A2*10"

    @pytest.mark.unit
    def test_fill_formula_range_multi_column(self, tmp_path, monkeypatch):
        """2D range B2:C5 fills 8 cells; each cell gets a relativized formula."""
        from openpyxl.formula.translate import Translator as _T
        path = self._make_wb(tmp_path, monkeypatch)

        result = fill_formula_range(
            path, "Sheet1", anchor_cell="B2", end_cell="C5",
            formula="=A2+1", confirm=True
        )
        assert result["filled_cells"] == 8
        assert result["range"] == "B2:C5"

        _wb_cache.clear()
        import openpyxl as _xl
        wb_disk = _xl.load_workbook(path, data_only=False)
        c3 = wb_disk["Sheet1"]["C3"].value
        expected = _T("=A2+1", origin="B2").translate_formula("C3")
        assert c3 == expected


class TestBulkAddDataValidation:
    """Unit tests for bulk_add_data_validation (T8, Sprint D Batch 2)."""

    def _make_wb(self, tmp_path, monkeypatch):
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        path = tmp_path / "dv_bulk.xlsx"
        import openpyxl as _xl
        _wb = _xl.Workbook()
        _wb.active.title = "Sheet1"
        _wb.save(str(path))
        return str(path)

    @pytest.mark.unit
    def test_bulk_add_dv_basic(self, tmp_path, monkeypatch):
        """Adding 2 validations returns validations_added=2."""
        path = self._make_wb(tmp_path, monkeypatch)
        result = bulk_add_data_validation(
            path, "Sheet1",
            validations=[
                {"address": "A1:A10", "validation_type": "list",
                 "formula1": '"Red,Green,Blue"'},
                {"address": "B1:B10", "validation_type": "whole",
                 "formula1": "1", "formula2": "100", "operator": "between"},
            ],
            confirm=True,
        )
        assert result["ok"] is True
        assert result["validations_added"] == 2
        assert result["sheet"] == "Sheet1"

    @pytest.mark.unit
    def test_bulk_add_dv_empty_list_raises(self, tmp_path, monkeypatch):
        """Empty validations list raises ValidationError."""
        path = self._make_wb(tmp_path, monkeypatch)
        with pytest.raises(ValidationError, match="non-empty"):
            bulk_add_data_validation(path, "Sheet1", validations=[], confirm=True)

    @pytest.mark.unit
    def test_bulk_add_dv_invalid_type_raises(self, tmp_path, monkeypatch):
        """Unknown validation_type raises ValidationError (fail-fast)."""
        path = self._make_wb(tmp_path, monkeypatch)
        with pytest.raises(ValidationError, match="invalid validation_type"):
            bulk_add_data_validation(
                path, "Sheet1",
                validations=[{"address": "A1", "validation_type": "badtype"}],
                confirm=True,
            )

    @pytest.mark.unit
    def test_bulk_add_dv_sqref_syntax(self, tmp_path, monkeypatch):
        """address with space-separated sqref syntax is accepted and written."""
        path = self._make_wb(tmp_path, monkeypatch)
        result = bulk_add_data_validation(
            path, "Sheet1",
            validations=[
                {"address": "A1:A10 C1:C10", "validation_type": "list",
                 "formula1": '"X,Y"'},
            ],
            confirm=True,
        )
        assert result["ok"] is True
        assert result["validations_added"] == 1

        _wb_cache.clear()
        import openpyxl as _xl
        wb_disk = _xl.load_workbook(path)
        dvs = list(wb_disk["Sheet1"].data_validations.dataValidation)
        assert len(dvs) == 1
        assert "A1" in str(dvs[0].sqref) or "C1" in str(dvs[0].sqref)

    @pytest.mark.unit
    def test_bulk_add_dv_formula_injection_guard(self, tmp_path, monkeypatch):
        """formula1 starting with '=' that exceeds 8192 chars raises ValidationError."""
        path = self._make_wb(tmp_path, monkeypatch)
        long_formula = "=" + "A1+" * 2731  # > 8192 chars
        with pytest.raises(ValidationError, match="[Ff]ormula too long|[Tt]oo long"):
            bulk_add_data_validation(
                path, "Sheet1",
                validations=[
                    {"address": "A1", "validation_type": "custom",
                     "formula1": long_formula},
                ],
                confirm=True,
            )

    @pytest.mark.unit
    def test_bulk_add_dv_allow_blank_propagated(self, tmp_path, monkeypatch):
        """allow_blank=False is reflected in the created DataValidation object."""
        path = self._make_wb(tmp_path, monkeypatch)
        bulk_add_data_validation(
            path, "Sheet1",
            validations=[
                {"address": "A1:A5", "validation_type": "list",
                 "formula1": '"Yes,No"', "allow_blank": False},
            ],
            confirm=True,
        )
        _wb_cache.clear()
        import openpyxl as _xl
        wb_disk = _xl.load_workbook(path)
        dvs = list(wb_disk["Sheet1"].data_validations.dataValidation)
        assert len(dvs) == 1
        assert dvs[0].allow_blank is False

    @pytest.mark.unit
    def test_bulk_add_dv_overwrite_existing(self, tmp_path, monkeypatch):
        """Adding DV to a cell that already has one accumulates rules (openpyxl allows multiple)."""
        path = self._make_wb(tmp_path, monkeypatch)
        # First call
        bulk_add_data_validation(
            path, "Sheet1",
            validations=[{"address": "A1:A5", "validation_type": "list",
                          "formula1": '"Red,Blue"'}],
            confirm=True,
        )
        _wb_cache.clear()
        # Second call on same range
        result = bulk_add_data_validation(
            path, "Sheet1",
            validations=[{"address": "A1:A5", "validation_type": "list",
                          "formula1": '"Green,Yellow"'}],
            confirm=True,
        )
        assert result["ok"] is True
        assert result["validations_added"] == 1

    @pytest.mark.unit
    def test_bulk_add_dv_fail_fast(self, tmp_path, monkeypatch):
        """If item 2 is invalid, item 1 is not written (fail-fast validation)."""
        path = self._make_wb(tmp_path, monkeypatch)
        with pytest.raises(ValidationError, match="invalid validation_type"):
            bulk_add_data_validation(
                path, "Sheet1",
                validations=[
                    {"address": "A1:A5", "validation_type": "list",
                     "formula1": '"Red,Blue"'},   # valid
                    {"address": "B1:B5", "validation_type": "BOGUS"},  # invalid
                ],
                confirm=True,
            )
        # Item 1 should NOT have been written
        _wb_cache.clear()
        import openpyxl as _xl
        wb_disk = _xl.load_workbook(path)
        dvs = list(wb_disk["Sheet1"].data_validations.dataValidation)
        assert dvs == [], "No DV should be written when fail-fast triggers"

    @pytest.mark.unit
    def test_bulk_add_dv_missing_address_raises(self, tmp_path, monkeypatch):
        """Missing 'address' in a validation item raises ValidationError."""
        path = self._make_wb(tmp_path, monkeypatch)
        with pytest.raises(ValidationError, match="address"):
            bulk_add_data_validation(
                path, "Sheet1",
                validations=[{"validation_type": "list", "formula1": '"Yes,No"'}],
                confirm=True,
            )

    @pytest.mark.unit
    def test_bulk_add_dv_write_guard(self, tmp_path, monkeypatch):
        """EXCEL_ENABLE_WRITE not set → NotAllowedError before touching disk."""
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        monkeypatch.delenv("EXCEL_ENABLE_WRITE", raising=False)
        with pytest.raises(NotAllowedError):
            bulk_add_data_validation(
                str(tmp_path / "ignored.xlsx"), "Sheet1",
                validations=[{"address": "A1", "validation_type": "list",
                              "formula1": '"X,Y"'}],
                confirm=True,
            )
