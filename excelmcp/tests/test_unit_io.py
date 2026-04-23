"""Unit tests for consolidate_ranges tool in server_io.py (US-E-030).

All tests use real openpyxl and tmp_path — no COM required.
BLOCKER-2: consolidate_ranges must honour EXCEL_ALLOWLIST_ROOTS via wb.read_range.
"""
from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl
import pytest

from excelmcp._core import ValidationError
from excelmcp.server_io import consolidate_ranges


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture()
def multi_wb_allowlist(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> Path:
    """Create 3 .xlsx files in tmp_path (the configured allowlist root).

    Each file has a 'Sheet1' with 5 rows and columns A-C.
    Returns tmp_path for use as a base directory.
    """
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))

    for wb_idx in range(1, 4):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        for row in range(1, 6):  # 5 data rows
            ws.cell(row=row, column=1).value = f"wb{wb_idx}_r{row}_A"
            ws.cell(row=row, column=2).value = f"wb{wb_idx}_r{row}_B"
            ws.cell(row=row, column=3).value = wb_idx * 10 + row
        wb.save(str(tmp_path / f"wb{wb_idx}.xlsx"))

    return tmp_path


# ---------------------------------------------------------------------------
# TestConsolidateRanges
# ---------------------------------------------------------------------------

class TestConsolidateRanges:

    def test_happy_path(self, multi_wb_allowlist: Path) -> None:
        """3 workbooks × 5 rows = 15 total rows, ok=True, truncated=False."""
        base = multi_wb_allowlist
        ranges: list[dict[str, Any]] = [
            {"path": str(base / f"wb{i}.xlsx"), "sheet": "Sheet1", "range": "A1:C5"}
            for i in range(1, 4)
        ]

        result = consolidate_ranges(ranges)

        assert result["ok"] is True
        assert result["source_count"] == 3
        assert result["total_rows"] == 15
        assert result["truncated"] is False
        assert len(result["rows"]) == 15
        # Verify source label is present in each row
        for row in result["rows"]:
            assert "source" in row
            assert "row_index" in row
            assert "data" in row

    def test_too_many_ranges(self, multi_wb_allowlist: Path) -> None:
        """21 ranges (> max 20) raises ValidationError."""
        base = multi_wb_allowlist
        ranges: list[dict[str, Any]] = [
            {"path": str(base / "wb1.xlsx"), "sheet": "Sheet1", "range": "A1:A1"}
            for _ in range(21)
        ]

        with pytest.raises(ValidationError, match="too many source ranges"):
            consolidate_ranges(ranges)

    def test_path_not_in_allowlist(
        self, multi_wb_allowlist: Path, tmp_path: Path
    ) -> None:
        """BLOCKER-2: a path outside the allowlist raises ValidationError (not ToolError)."""
        outside = tmp_path.parent / "outside_consolidate"
        outside.mkdir(exist_ok=True)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "bad"
        wb.save(str(outside / "bad.xlsx"))

        ranges = [
            {"path": str(outside / "bad.xlsx"), "sheet": "Sheet1", "range": "A1:A1"}
        ]

        with pytest.raises(ValidationError):
            consolidate_ranges(ranges)

    def test_missing_required_key(self, multi_wb_allowlist: Path) -> None:
        """A range entry missing 'path', 'sheet', or 'range' raises ValidationError."""
        bad_ranges: list[dict[str, Any]] = [
            {"sheet": "Sheet1", "range": "A1:A1"},  # no 'path'
        ]

        with pytest.raises((ValidationError, KeyError)):
            consolidate_ranges(bad_ranges)

    def test_max_rows_exceeded(self, multi_wb_allowlist: Path, monkeypatch: pytest.MonkeyPatch) -> None:
        """Outer-loop break fires after truncation: read_range called only for ranges 1+2."""
        import excelmcp.workbook_openpyxl as _wb_mod

        base = multi_wb_allowlist
        # Create wb4 so we have 4 ranges (fixture only makes 3)
        wb4 = openpyxl.Workbook()
        ws4 = wb4.active
        ws4.title = "Sheet1"
        for row in range(1, 6):
            ws4.cell(row=row, column=1).value = f"wb4_r{row}_A"
            ws4.cell(row=row, column=2).value = f"wb4_r{row}_B"
            ws4.cell(row=row, column=3).value = 40 + row
        wb4.save(str(base / "wb4.xlsx"))

        # 4 ranges × 5 rows each; max_rows=8 → truncation fires mid range-2 (5+3=8)
        ranges: list[dict[str, Any]] = [
            {"path": str(base / f"wb{i}.xlsx"), "sheet": "Sheet1", "range": "A1:C5"}
            for i in range(1, 5)
        ]

        _real_read_range = _wb_mod.read_range
        call_count: list[int] = [0]

        def _spy(path: str, sheet: str, a1_range: str) -> list:
            call_count[0] += 1
            return _real_read_range(path=path, sheet=sheet, a1_range=a1_range)

        monkeypatch.setattr(_wb_mod, "read_range", _spy)

        result = consolidate_ranges(ranges, max_rows=8)

        assert result["ok"] is True
        assert result["truncated"] is True
        assert result["total_rows"] == 8
        assert call_count[0] == 2, (
            f"Expected read_range called exactly 2 times (ranges 1+2 only), got {call_count[0]}"
        )

    def test_consolidate_reserved_label_key_raises(self) -> None:
        """source_label_key matching a reserved output key raises ValidationError."""
        with pytest.raises(ValidationError, match="conflicts with reserved"):
            consolidate_ranges(
                ranges=[{"path": "x.xlsx", "sheet": "s", "range": "A1"}],
                source_label_key="row_index",
            )

    def test_empty_range_list(self, multi_wb_allowlist: Path) -> None:
        """Empty ranges list raises ValidationError immediately."""
        with pytest.raises(ValidationError, match="must not be empty"):
            consolidate_ranges([])
