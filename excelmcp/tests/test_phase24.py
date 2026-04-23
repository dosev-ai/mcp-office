import pytest
import openpyxl

from excelmcp.workbook_openpyxl import (
    ValidationError,
    add_cell_comment,
    delete_cell_comment,
    get_cell_comment,
    bulk_add_comments,
    get_sheet_properties,
    set_sheet_properties,
    add_data_validation,
    protect_workbook,
    _wb_cache,
)


class TestPhase24Tools:
    """Phase 2.4 — add/delete_cell_comment, get/set_sheet_properties,
    add_data_validation, protect_workbook."""

    # ── add_cell_comment ─────────────────────────────────────────────────

    @pytest.mark.unit
    def test_add_cell_comment_success(self, tmp_path, monkeypatch):
        """add_cell_comment writes a Comment and returns ok=True."""
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        wb_path = tmp_path / "comments.xlsx"
        _wb = openpyxl.Workbook()
        _wb.active.title = "Sheet1"
        _wb.save(str(wb_path))

        result = add_cell_comment(str(wb_path), "Sheet1", "B2", "My note", confirm=True)
        assert result["ok"] is True
        assert result["address"] == "B2"
        assert result["author"] == "ExcelMCP"
        assert "B2" in result["message"]

        _wb_cache.clear()
        _wb_v = openpyxl.load_workbook(str(wb_path))
        assert _wb_v["Sheet1"]["B2"].comment is not None
        assert _wb_v["Sheet1"]["B2"].comment.text == "My note"

    @pytest.mark.unit
    def test_add_cell_comment_requires_confirm(self, sample_workbook, monkeypatch):
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        with pytest.raises(ValidationError, match="confirm=True"):
            add_cell_comment(sample_workbook, "Sheet1", "A1", "text", confirm=False)

    @pytest.mark.unit
    def test_add_cell_comment_invalid_sheet(self, sample_workbook, monkeypatch):
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        with pytest.raises(ValidationError, match="[Ss]heet"):
            add_cell_comment(sample_workbook, "NoSuchSheet", "A1", "text", confirm=True)

    # ── delete_cell_comment ──────────────────────────────────────────────

    @pytest.mark.unit
    def test_delete_cell_comment_success(self, tmp_path, monkeypatch):
        """Add then delete a comment; verify it is gone after delete."""
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        wb_path = tmp_path / "del_comment.xlsx"
        _wb = openpyxl.Workbook()
        _wb.active.title = "Sheet1"
        _wb.save(str(wb_path))

        add_cell_comment(str(wb_path), "Sheet1", "C3", "delete me", confirm=True)
        _wb_cache.clear()

        result = delete_cell_comment(str(wb_path), "Sheet1", "C3", confirm=True)
        assert result["ok"] is True
        assert result["address"] == "C3"
        assert "C3" in result["message"]

        _wb_cache.clear()
        _wb_v = openpyxl.load_workbook(str(wb_path))
        assert _wb_v["Sheet1"]["C3"].comment is None

    @pytest.mark.unit
    def test_delete_cell_comment_requires_confirm(self, sample_workbook, monkeypatch):
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        with pytest.raises(ValidationError, match="confirm=True"):
            delete_cell_comment(sample_workbook, "Sheet1", "A1", confirm=False)

    @pytest.mark.unit
    def test_delete_cell_comment_invalid_sheet(self, sample_workbook, monkeypatch):
        """delete_cell_comment must raise ValidationError for an unknown sheet."""
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        with pytest.raises(ValidationError, match="[Ss]heet"):
            delete_cell_comment(sample_workbook, "NoSuchSheet", "A1", confirm=True)

    # ── get_sheet_properties ─────────────────────────────────────────────

    @pytest.mark.unit
    def test_get_sheet_properties_returns_visible_state(self, sample_workbook):
        """Fresh workbook sheet is visible; state must be 'visible'."""
        result = get_sheet_properties(sample_workbook, "Sheet1")
        assert result["sheet"] == "Sheet1"
        assert result["state"] == "visible"
        assert result["tab_color"] is None

    @pytest.mark.unit
    def test_get_sheet_properties_invalid_sheet(self, sample_workbook):
        with pytest.raises(ValidationError, match="[Ss]heet"):
            get_sheet_properties(sample_workbook, "NoSuchSheet")

    @pytest.mark.unit
    def test_get_sheet_properties_returns_hidden_state(self, tmp_path, monkeypatch):
        """get_sheet_properties reflects hidden sheet_state correctly."""
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        wb_path = tmp_path / "hidden.xlsx"
        _wb = openpyxl.Workbook()
        _wb.active.title = "Sheet1"
        _wb.create_sheet("Sheet2")
        _wb["Sheet1"].sheet_state = "hidden"
        _wb.save(str(wb_path))
        result = get_sheet_properties(str(wb_path), "Sheet1")
        assert result["state"] == "hidden"
        assert result["sheet"] == "Sheet1"

    @pytest.mark.unit
    def test_get_sheet_properties_returns_tab_color(self, tmp_path, monkeypatch):
        """get_sheet_properties returns the 6-digit hex tab_color when set."""
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        wb_path = tmp_path / "tabcolor.xlsx"
        _wb = openpyxl.Workbook()
        _wb.active.title = "Sheet1"
        _wb.save(str(wb_path))
        # Use set_sheet_properties to set the colour, then read it back
        set_sheet_properties(str(wb_path), "Sheet1", tab_color="FF0000", confirm=True)
        _wb_cache.clear()
        result = get_sheet_properties(str(wb_path), "Sheet1")
        assert result["tab_color"] == "FF0000"

    # ── set_sheet_properties ─────────────────────────────────────────────

    @pytest.mark.unit
    def test_set_sheet_properties_set_state(self, tmp_path, monkeypatch):
        """set_sheet_properties can mark a sheet as hidden."""
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        wb_path = tmp_path / "props.xlsx"
        _wb = openpyxl.Workbook()
        _wb.active.title = "Sheet1"
        _wb.create_sheet("Sheet2")
        _wb.save(str(wb_path))

        result = set_sheet_properties(str(wb_path), "Sheet1", state="hidden", confirm=True)
        assert result["ok"] is True
        assert result["state"] == "hidden"
        assert "updated" in result["message"].lower()

        _wb_cache.clear()
        _wb_v = openpyxl.load_workbook(str(wb_path))
        assert _wb_v["Sheet1"].sheet_state == "hidden"

    @pytest.mark.unit
    def test_set_sheet_properties_invalid_tab_color(self, sample_workbook, monkeypatch):
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        with pytest.raises(ValidationError, match="[Ii]nvalid tab_color|hex|6"):
            set_sheet_properties(sample_workbook, "Sheet1", tab_color="ZZZZZZ", confirm=True)

    @pytest.mark.unit
    def test_set_sheet_properties_invalid_state(self, sample_workbook, monkeypatch):
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        with pytest.raises(ValidationError, match="[Ii]nvalid state"):
            set_sheet_properties(sample_workbook, "Sheet1", state="invisible", confirm=True)

    @pytest.mark.unit
    def test_set_sheet_properties_requires_confirm(self, sample_workbook, monkeypatch):
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        with pytest.raises(ValidationError, match="confirm=True"):
            set_sheet_properties(sample_workbook, "Sheet1", state="hidden", confirm=False)

    @pytest.mark.unit
    def test_set_sheet_properties_set_tab_color(self, tmp_path, monkeypatch):
        """set_sheet_properties with a valid 6-digit hex tab_color succeeds."""
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        wb_path = tmp_path / "setcolor.xlsx"
        _wb = openpyxl.Workbook()
        _wb.active.title = "Sheet1"
        _wb.save(str(wb_path))
        result = set_sheet_properties(str(wb_path), "Sheet1", tab_color="00B050", confirm=True)
        assert result["ok"] is True
        assert result["tab_color"] == "00B050"
        assert "updated" in result["message"].lower()
        _wb_cache.clear()
        _wb_v = openpyxl.load_workbook(str(wb_path))
        color = _wb_v["Sheet1"].sheet_properties.tabColor
        assert color is not None and color.rgb[-6:] == "00B050"

    # ── add_data_validation ──────────────────────────────────────────────

    @pytest.mark.unit
    def test_add_data_validation_list_type(self, tmp_path, monkeypatch):
        """Add a list validation rule; verify it is serialised in the workbook."""
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        wb_path = tmp_path / "dv.xlsx"
        _wb = openpyxl.Workbook()
        _wb.active.title = "Sheet1"
        _wb.save(str(wb_path))

        result = add_data_validation(
            str(wb_path), "Sheet1", "A1:A10",
            validation_type="list", formula1='"Yes,No,Maybe"', confirm=True,
        )
        assert result["ok"] is True
        assert result["validation_type"] == "list"
        assert result["address"] == "A1:A10"
        assert "list" in result["message"]

        _wb_cache.clear()
        _wb_v = openpyxl.load_workbook(str(wb_path))
        dvs = list(_wb_v["Sheet1"].data_validations.dataValidation)
        assert any(dv.type == "list" for dv in dvs)

    @pytest.mark.unit
    def test_add_data_validation_invalid_type(self, sample_workbook, monkeypatch):
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        with pytest.raises(ValidationError, match="[Ii]nvalid validation_type"):
            add_data_validation(
                sample_workbook, "Sheet1", "A1",
                validation_type="bogus", confirm=True,
            )

    @pytest.mark.unit
    def test_add_data_validation_requires_confirm(self, sample_workbook, monkeypatch):
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        with pytest.raises(ValidationError, match="confirm=True"):
            add_data_validation(
                sample_workbook, "Sheet1", "A1",
                validation_type="whole", formula1="1", confirm=False,
            )

    @pytest.mark.unit
    def test_add_data_validation_missing_formula1_for_list(self, sample_workbook, monkeypatch):
        """add_data_validation with type='list' and no formula1 must raise ValidationError."""
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        with pytest.raises(ValidationError, match="formula1.*required|required.*formula1"):
            add_data_validation(
                sample_workbook, "Sheet1", "A1",
                validation_type="list", formula1=None, confirm=True,
            )

    # ── protect_workbook ─────────────────────────────────────────────────

    @pytest.mark.unit
    def test_protect_workbook_success_no_password(self, tmp_path, monkeypatch):
        """Protect workbook structure without password."""
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        wb_path = tmp_path / "wb_protect.xlsx"
        _wb = openpyxl.Workbook()
        _wb.save(str(wb_path))

        result = protect_workbook(str(wb_path), confirm=True)
        assert result["ok"] is True
        assert result["protected"] is True
        assert result["lock_structure"] is True
        assert result["has_password"] is False

        _wb_cache.clear()
        _wb_v = openpyxl.load_workbook(str(wb_path))
        assert _wb_v.security.lockStructure is True

    @pytest.mark.unit
    def test_protect_workbook_with_password(self, tmp_path, monkeypatch):
        """Protect workbook with a password; has_password must be True."""
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        wb_path = tmp_path / "wb_protect_pw.xlsx"
        _wb = openpyxl.Workbook()
        _wb.save(str(wb_path))

        result = protect_workbook(str(wb_path), password="Secret123", confirm=True)
        assert result["ok"] is True
        assert result["has_password"] is True
        assert result["protected"] is True

    @pytest.mark.unit
    def test_protect_workbook_password_too_long(self, sample_workbook, monkeypatch):
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        with pytest.raises(ValidationError, match="255"):
            protect_workbook(sample_workbook, password="x" * 256, confirm=True)

    @pytest.mark.unit
    def test_protect_workbook_requires_confirm(self, sample_workbook, monkeypatch):
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        with pytest.raises(ValidationError, match="confirm=True"):
            protect_workbook(sample_workbook, confirm=False)


class TestUS21CommentAuditCycle:
    """US-21 — Comment audit cycle: bulk add, read, delete, deduplication."""

    @pytest.mark.unit
    def test_us21_bulk_add_three_comments(self, tmp_path, monkeypatch):
        """bulk_add_comments with 3 distinct addresses writes all 3."""
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        wb_path = tmp_path / "bulk3.xlsx"
        _wb = openpyxl.Workbook()
        _wb.active.title = "Sheet1"
        _wb.save(str(wb_path))

        result = bulk_add_comments(
            str(wb_path), "Sheet1",
            comments=[
                {"address": "A1", "text": "Alpha", "author": "Test"},
                {"address": "B2", "text": "Beta", "author": "Test"},
                {"address": "C3", "text": "Gamma", "author": "Test"},
            ],
            confirm=True,
        )
        assert result["ok"] is True
        assert result["comments_added"] == 3
        assert result["skipped"] == 0
        assert set(result["cells"]) == {"A1", "B2", "C3"}

        # MCP read — no _wb_cache.clear() needed (_save_and_evict handles eviction)
        a1 = get_cell_comment(str(wb_path), "Sheet1", "A1")
        assert a1["has_comment"] is True
        assert a1["text"] == "Alpha"

        _wb_cache.clear()  # only before direct openpyxl load
        wb_v = openpyxl.load_workbook(str(wb_path))
        assert wb_v["Sheet1"]["C3"].comment.text == "Gamma"

    @pytest.mark.unit
    def test_us21_read_individual_comment(self, sample_workbook, monkeypatch):
        """add_cell_comment → get_cell_comment round-trip returns correct fields."""
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        add_cell_comment(sample_workbook, "Sheet1", "A1", "round-trip note",
                         author="Tester", confirm=True)
        result = get_cell_comment(sample_workbook, "Sheet1", "A1")
        assert result["has_comment"] is True
        assert result["text"] == "round-trip note"
        assert result["author"] == "Tester"
        assert result["address"] == "A1"

    @pytest.mark.unit
    def test_us21_delete_comment(self, tmp_path, monkeypatch):
        """add → delete → get_cell_comment returns has_comment=False."""
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        wb_path = tmp_path / "del_test.xlsx"
        _wb = openpyxl.Workbook()
        _wb.active.title = "Sheet1"
        _wb.save(str(wb_path))

        add_cell_comment(str(wb_path), "Sheet1", "D4", "to be deleted", confirm=True)
        delete_cell_comment(str(wb_path), "Sheet1", "D4", confirm=True)
        result = get_cell_comment(str(wb_path), "Sheet1", "D4")
        assert result["has_comment"] is False
        assert result["text"] is None

    @pytest.mark.unit
    def test_us21_bulk_then_delete_two(self, tmp_path, monkeypatch):
        """bulk_add×3 → delete A1 + B2 → only C3 survives."""
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        wb_path = tmp_path / "bulk_del.xlsx"
        _wb = openpyxl.Workbook()
        _wb.active.title = "Sheet1"
        _wb.save(str(wb_path))

        bulk_add_comments(
            str(wb_path), "Sheet1",
            comments=[
                {"address": "A1", "text": "doomed-A1", "author": "T"},
                {"address": "B2", "text": "doomed-B2", "author": "T"},
                {"address": "C3", "text": "survivor", "author": "T"},
            ],
            confirm=True,
        )
        delete_cell_comment(str(wb_path), "Sheet1", "A1", confirm=True)
        delete_cell_comment(str(wb_path), "Sheet1", "B2", confirm=True)

        assert get_cell_comment(str(wb_path), "Sheet1", "A1")["has_comment"] is False
        assert get_cell_comment(str(wb_path), "Sheet1", "B2")["has_comment"] is False
        c3 = get_cell_comment(str(wb_path), "Sheet1", "C3")
        assert c3["has_comment"] is True
        assert c3["text"] == "survivor"

    @pytest.mark.unit
    def test_us21_bulk_add_dedup_same_address(self, tmp_path, monkeypatch):
        """Duplicate address in bulk list: last entry wins (deduped dict _data.py ~L725)."""
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        wb_path = tmp_path / "dedup.xlsx"
        _wb = openpyxl.Workbook()
        _wb.active.title = "Sheet1"
        _wb.save(str(wb_path))

        result = bulk_add_comments(
            str(wb_path), "Sheet1",
            comments=[
                {"address": "A1", "text": "first-A1", "author": "T"},
                {"address": "B2", "text": "only-B2", "author": "T"},
                {"address": "A1", "text": "last-A1", "author": "T"},
            ],
            confirm=True,
        )
        # Deduplication: A1 appears twice → 2 unique addresses written
        assert result["comments_added"] == 2
        assert get_cell_comment(str(wb_path), "Sheet1", "A1")["text"] == "last-A1"
        assert get_cell_comment(str(wb_path), "Sheet1", "B2")["text"] == "only-B2"


@pytest.mark.unit
def test_add_data_validation_formula_too_long_rejected(tmp_path, monkeypatch):
    """add_data_validation rejects formula1 starting with '=' that exceeds 8192 chars."""
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    path = tmp_path / "dv_injection.xlsx"
    from excelmcp._io import create_workbook
    create_workbook(str(path), confirm=True)
    long_formula = "=" + "A1+" * 2731  # > 8192 chars - rejected by _validate_formula_safe
    with pytest.raises(ValidationError, match="[Ff]ormula too long|[Tt]oo long"):
        add_data_validation(
            str(path), "Sheet1", "A1:A10", "custom",
            formula1=long_formula, confirm=True,
        )
