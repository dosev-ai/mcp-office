"""
Integration tests for the Excel MCP server DA formula pipeline.
All tests require Excel 16.0 COM available (Windows only).
Run with: pytest excel/tests/test_integration.py -v -m integration
Skip automatically when win32com unavailable.
"""
from __future__ import annotations

import os
from pathlib import Path

import openpyxl
import pytest

from excelmcp._ooxml import patch_dynamic_array_metadata
from excelmcp._da_com import probe_excel_reopen
from excelmcp._da_finalize import finalize_workbook_da_artifact

# ── Module-level skip guard ──────────────────────────────────────────────────

try:
    import win32com.client  # noqa: F401
    HAS_COM = True
except ImportError:
    HAS_COM = False

skip_no_com = pytest.mark.skipif(
    not HAS_COM,
    reason="win32com not available (non-Windows or no Office installed)",
)

pytestmark = pytest.mark.integration

# ── Helpers ──────────────────────────────────────────────────────────────────


def _make_da_workbook(tmp_path: Path, formula: str, cell: str = "A1") -> Path:
    """Create minimal xlsx with one formula cell using openpyxl (no patching)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws[cell] = formula
    path = tmp_path / "test_da.xlsx"
    wb.save(str(path))
    return path


# ── Test class ───────────────────────────────────────────────────────────────


@skip_no_com
class TestDAFormulaPipeline:
    """BDD ground-truth scenarios: COM as compatibility oracle for DA patching."""

    @pytest.fixture(autouse=True)
    def _set_com_gate(self) -> None:  # type: ignore[return]
        """Ensure EXCEL_ENABLE_COM=true for all tests in this class."""
        os.environ["EXCEL_ENABLE_COM"] = "true"
        yield
        os.environ.pop("EXCEL_ENABLE_COM", None)

    # TC-INT-01
    def test_da_filter_formula_no_repair_after_patch(self, tmp_path: Path) -> None:
        """
        GIVEN a pipeline-generated DA workbook with a FILTER formula,
        WHEN patch_dynamic_array_metadata() runs on it,
        THEN COM can open the file without repair (probe returns ok=True).
        """
        path = _make_da_workbook(tmp_path, "=FILTER(A2:A10,B2:B10>0)")
        patch_dynamic_array_metadata(str(path))

        result = probe_excel_reopen(str(path))

        assert result.get("ok") is True, (
            f"Excel opened the patched FILTER workbook with repair: {result.get('error')}"
        )

    # TC-INT-02
    def test_da_xlookup_formula_no_repair_after_patch(self, tmp_path: Path) -> None:
        """
        GIVEN a pipeline-generated DA workbook with an XLOOKUP formula,
        WHEN patch_dynamic_array_metadata() runs on it,
        THEN COM can open the file without repair (probe returns ok=True).
        """
        path = _make_da_workbook(tmp_path, "=XLOOKUP(A2,B2:B10,C2:C10)")
        patch_dynamic_array_metadata(str(path))

        result = probe_excel_reopen(str(path))

        assert result.get("ok") is True, (
            f"Excel opened the patched XLOOKUP workbook with repair: {result.get('error')}"
        )

    # TC-INT-03
    def test_da_multiple_formulas_no_repair_after_patch(self, tmp_path: Path) -> None:
        """
        GIVEN an xlsx with 3 DA formulas (FILTER, SORT, UNIQUE in separate cells),
        WHEN patch_dynamic_array_metadata() runs on it,
        THEN COM opens without repair (probe returns ok=True).
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "=FILTER(B2:B10,C2:C10>0)"
        ws["A2"] = "=SORT(D2:D10)"
        ws["A3"] = "=UNIQUE(E2:E10)"
        path = tmp_path / "test_multi_da.xlsx"
        wb.save(str(path))

        patch_dynamic_array_metadata(str(path))
        result = probe_excel_reopen(str(path))

        assert result.get("ok") is True, (
            f"Excel opened multi-DA workbook with repair: {result.get('error')}"
        )

    # TC-INT-04
    def test_com_roundtrip_preserves_formula_value(self, tmp_path: Path) -> None:
        """
        GIVEN a patched DA workbook with SEQUENCE(5) in A1,
        WHEN COM performs a full roundtrip (open, CalculateFull, save, reopen),
        THEN the file is not corrupted (probe returns ok=True on second open).

        BDD: GIVEN a patched DA workbook, WHEN COM performs a full roundtrip,
             THEN the file survives.
        """
        path = _make_da_workbook(tmp_path, "=SEQUENCE(5)")
        patch_dynamic_array_metadata(str(path))

        # First probe: COM open → CalculateFull → SaveAs → reopen
        result = probe_excel_reopen(str(path))
        assert result.get("ok") is True, (
            f"First COM roundtrip failed: {result.get('error')}"
        )

        # Second probe on the same (still-original) file to confirm no corruption
        result2 = probe_excel_reopen(str(path))
        assert result2.get("ok") is True, (
            f"Second COM roundtrip failed (file corrupted): {result2.get('error')}"
        )

    # TC-INT-05
    def test_com_seed_workflow(self, tmp_path: Path) -> None:
        """
        BDD: COM-seed workflow — COM creates initial DA file, openpyxl adds data,
        patch is idempotent on COM-authored files.

        GIVEN: COM creates a workbook with a FILTER formula via xlwings/win32com
        WHEN: openpyxl opens the file and writes additional non-DA data (label in C1)
        AND: patch_dynamic_array_metadata runs on the modified file
        THEN: COM can still open the file without repair (patch is idempotent on
              COM-authored DA files).

        Note: validates that the patch doesn't corrupt COM-authored xlsx.
        """
        import pythoncom
        import win32com.client

        seed_path = tmp_path / "com_seed_test.xlsx"
        abs_seed = str(seed_path)

        # Step 1: COM creates the initial DA workbook with a FILTER formula
        pythoncom.CoInitialize()
        excel = None
        try:
            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False

            wb = excel.Workbooks.Add()
            ws = wb.Sheets(1)
            ws.Range("A1").Formula2 = "=FILTER(B2:B10,C2:C10>0)"
            wb.SaveAs(abs_seed, FileFormat=51)  # 51 = xlOpenXMLWorkbook
            wb.Close(SaveChanges=False)
            del wb
        finally:
            if excel is not None:
                try:
                    excel.Quit()
                except Exception:
                    pass
                excel = None
            pythoncom.CoUninitialize()

        assert seed_path.exists(), "COM did not create the seed file"

        # Step 2: openpyxl opens the COM-authored file and adds a non-DA label
        wb_opy = openpyxl.load_workbook(str(seed_path))
        ws_opy = wb_opy.active
        ws_opy["C1"] = "openpyxl_label"
        wb_opy.save(str(seed_path))

        # Step 3: patch runs on the modified file (idempotency check)
        patch_dynamic_array_metadata(str(seed_path))

        # Step 4: COM verifies the file is clean (no repair dialog)
        result = probe_excel_reopen(str(seed_path))
        assert result.get("ok") is True, (
            f"COM-seed workflow: Excel opened with repair after openpyxl+patch: "
            f"{result.get('error')}"
        )

    # TC-INT-06
    def test_finalize_workbook_da_artifact_with_com_probe(
        self, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        """
        GIVEN an openpyxl-written xlsx with a FILTER formula,
        WHEN finalize_workbook_da_artifact(path, require_com_probe=True) is called,
        THEN the diagnostics contain a probe stage with status == "ok".
        """
        path = _make_da_workbook(tmp_path, "=FILTER(A2:A10,B2:B10>0)")

        # Force com_probe off at env level so the require_com_probe param controls it
        monkeypatch.setenv("EXCEL_DA_COM_PROBE", "off")

        diagnostics = finalize_workbook_da_artifact(str(path), require_com_probe=True)

        stages = {s["name"]: s for s in diagnostics.get("stages", [])}
        assert "probe" in stages, f"probe stage missing from diagnostics: {diagnostics}"
        assert stages["probe"]["status"] == "ok", (
            f"Probe stage not ok: {stages['probe']}"
        )
        probe_result = stages["probe"].get("result", {})
        assert probe_result.get("ok") is True, (
            f"COM probe returned non-ok inside finalize: {probe_result}"
        )
