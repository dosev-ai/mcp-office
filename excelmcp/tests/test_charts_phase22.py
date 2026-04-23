import os
import pytest
import openpyxl

from excelmcp.workbook_openpyxl import (
    ExcelMCPError,
    ValidationError,
    NotAllowedError,
    create_workbook,
    create_chart,
    list_charts,
    delete_chart,
    write_named_range,
    merge_cells,
    unmerge_cells,
    freeze_panes,
    auto_filter,
    get_named_ranges,
    _wb_cache,
)


# ── GROUP 27: create_workbook() ──────────────────────────────────────────────────

class TestCreateWorkbook:
    @pytest.mark.unit
    def test_create_workbook_basic(self, tmp_path, monkeypatch):
        """Creates a valid .xlsx with default Sheet1."""
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        p = str(tmp_path / "new.xlsx")
        result = create_workbook(path=p, confirm=True)
        assert result["ok"] is True
        assert result["created"] is True
        assert result["sheets"] == ["Sheet1"]
        wb = openpyxl.load_workbook(p)
        assert wb.sheetnames == ["Sheet1"]

    @pytest.mark.unit
    def test_create_workbook_custom_sheets(self, tmp_path, monkeypatch):
        """Creates workbook with multiple custom sheets."""
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        p = str(tmp_path / "multi.xlsx")
        result = create_workbook(path=p, sheets=["Sales", "Products", "Summary"], confirm=True)
        assert result["sheets"] == ["Sales", "Products", "Summary"]
        wb = openpyxl.load_workbook(p)
        assert wb.sheetnames == ["Sales", "Products", "Summary"]

    @pytest.mark.unit
    def test_create_workbook_overwrite_false_raises(self, tmp_path, monkeypatch):
        """Raises if file exists and overwrite=False."""
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        p = str(tmp_path / "existing.xlsx")
        create_workbook(path=p, confirm=True)
        with pytest.raises(ExcelMCPError, match="already exists"):
            create_workbook(path=p, confirm=True)  # overwrite defaults to False

    @pytest.mark.unit
    def test_create_workbook_overwrite_true(self, tmp_path, monkeypatch):
        """Replaces existing file when overwrite=True."""
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        p = str(tmp_path / "replace.xlsx")
        create_workbook(path=p, confirm=True)
        result = create_workbook(path=p, sheets=["New"], overwrite=True, confirm=True)
        assert result["ok"] is True
        wb = openpyxl.load_workbook(p)
        assert wb.sheetnames == ["New"]

    @pytest.mark.unit
    def test_create_workbook_confirm_false_raises(self, tmp_path, monkeypatch):
        """Raises ValidationError when confirm=False."""
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        p = str(tmp_path / "no_confirm.xlsx")
        with pytest.raises(ExcelMCPError, match="confirm"):
            create_workbook(path=p, confirm=False)

    @pytest.mark.unit
    def test_create_workbook_non_xlsx_raises(self, tmp_path, monkeypatch):
        """Raises for non-.xlsx extension."""
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        p = str(tmp_path / "bad.xls")
        with pytest.raises(ExcelMCPError, match=".xlsx"):
            create_workbook(path=p, confirm=True)

    @pytest.mark.unit
    def test_create_workbook_path_not_in_allowlist(self, tmp_path, monkeypatch):
        """Raises for path outside allowlist."""
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path / "allowed_only"))
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        p = str(tmp_path / "blocked.xlsx")
        with pytest.raises(ExcelMCPError, match="allowlist"):
            create_workbook(path=p, confirm=True)

    @pytest.mark.unit
    def test_create_workbook_write_gate_raises(self, tmp_path, monkeypatch):
        """EXCEL_ENABLE_WRITE not set → NotAllowedError (or ExcelMCPError) before touching disk."""
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        monkeypatch.delenv("EXCEL_ENABLE_WRITE", raising=False)
        p = str(tmp_path / "wb.xlsx")
        with pytest.raises((ExcelMCPError, NotAllowedError)):
            create_workbook(p, confirm=True)
        assert not os.path.exists(p), "File must NOT be created when write gate blocks"


# ── GROUP 27: Chart tools ──────────────────────────────────────────────────

@pytest.mark.unit
def test_create_chart_happy_path(sample_workbook, monkeypatch):
    """create_chart writes a bar chart and list_charts finds it."""
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    import openpyxl as _opx
    # Overwrite sample with numeric data so reference is valid
    wb = _opx.load_workbook(sample_workbook)
    ws = wb["Sheet1"]
    ws["A1"] = "Value"
    ws["A2"] = 10
    ws["A3"] = 20
    wb.save(sample_workbook)
    _wb_cache.clear()

    result = create_chart(sample_workbook, "Sheet1", "bar", "A1:A3", "Test Chart", "C1", confirm=True)
    assert result["ok"] is True
    assert "message" in result
    assert "bar" in result["message"]
    assert result["chart_type"] == "bar"
    assert result["target_cell"] == "C1"

    _wb_cache.clear()
    charts_result = list_charts(sample_workbook, "Sheet1")
    assert charts_result["count"] >= 1
    assert charts_result["charts"][0]["type"] == "BarChart"


@pytest.mark.unit
def test_create_chart_invalid_type_raises(sample_workbook, monkeypatch):
    """Unknown chart type must raise ExcelMCPError before touching disk."""
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    with pytest.raises(ExcelMCPError, match="Invalid chart_type"):
        create_chart(sample_workbook, "Sheet1", "unknown_type", "A1:A3", confirm=True)


@pytest.mark.unit
def test_create_chart_invalid_range_raises(sample_workbook, monkeypatch):
    """Non-range data_range (no colon) must raise ExcelMCPError."""
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    with pytest.raises(ExcelMCPError, match="data_range must be a range"):
        create_chart(sample_workbook, "Sheet1", "bar", "A1", confirm=True)


@pytest.mark.unit
def test_list_charts_empty(sample_workbook):
    """Fresh workbook sheet has no charts."""
    result = list_charts(sample_workbook, "Sheet1")
    assert "charts" in result
    assert isinstance(result["charts"], list)
    assert result["count"] == 0


@pytest.mark.unit
def test_delete_chart_out_of_range_raises(sample_workbook, monkeypatch):
    """delete_chart with no charts raises ExcelMCPError."""
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    with pytest.raises(ExcelMCPError, match="out of range"):
        delete_chart(sample_workbook, "Sheet1", 0, confirm=True)


@pytest.mark.unit
def test_delete_chart_happy_path(sample_workbook, monkeypatch):
    """Create then delete a chart; count drops to 0."""
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    import openpyxl as _opx
    wb = _opx.load_workbook(sample_workbook)
    ws = wb["Sheet1"]
    ws["A1"] = "Value"
    ws["A2"] = 10
    wb.save(sample_workbook)
    _wb_cache.clear()

    create_chart(sample_workbook, "Sheet1", "line", "A1:A2", "DeleteMe", "D1", confirm=True)
    _wb_cache.clear()

    charts_before = list_charts(sample_workbook, "Sheet1")
    assert charts_before["count"] >= 1
    _wb_cache.clear()

    result = delete_chart(sample_workbook, "Sheet1", 0, confirm=True)
    assert result["ok"] is True
    assert "message" in result
    assert result["chart_index"] == 0
    _wb_cache.clear()

    charts_after = list_charts(sample_workbook, "Sheet1")
    assert charts_after["count"] == 0


@pytest.mark.unit
def test_create_chart_requires_write(sample_workbook, monkeypatch):
    """create_chart blocked when EXCEL_ENABLE_WRITE is not set."""
    monkeypatch.delenv("EXCEL_ENABLE_WRITE", raising=False)
    with pytest.raises(ExcelMCPError):
        create_chart(sample_workbook, "Sheet1", "bar", "A1:A2", confirm=True)


@pytest.mark.unit
def test_delete_chart_requires_write(sample_workbook, monkeypatch):
    """delete_chart blocked when EXCEL_ENABLE_WRITE is not set."""
    monkeypatch.delenv("EXCEL_ENABLE_WRITE", raising=False)
    with pytest.raises(ExcelMCPError):
        delete_chart(sample_workbook, "Sheet1", 0, confirm=True)


# ── GROUP — Phase 2.2: write_named_range, merge_cells, unmerge_cells,
#                       freeze_panes, auto_filter ─────────────────────────

class TestPhase22Tools:
    """Phase 2.2 — write_named_range, merge_cells, unmerge_cells, freeze_panes, auto_filter."""

    # ── write_named_range ───────────────────────────────────────────────

    @pytest.mark.unit
    def test_write_named_range_happy(self, sample_workbook, monkeypatch):
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        result = write_named_range(
            sample_workbook, name="MyRange", sheet="Sheet1",
            range_address="A1:A2", confirm=True,
        )
        assert result["ok"] is True
        assert result["name"] == "MyRange"
        assert "refers_to" in result
        assert "Sheet1" in result["refers_to"]

    @pytest.mark.unit
    def test_write_named_range_confirm_false_raises(self, sample_workbook, monkeypatch):
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        with pytest.raises(ValidationError, match="confirm=True"):
            write_named_range(
                sample_workbook, name="X", sheet="Sheet1",
                range_address="A1", confirm=False,
            )

    @pytest.mark.unit
    def test_write_named_range_write_gate_raises(self, sample_workbook, monkeypatch):
        monkeypatch.delenv("EXCEL_ENABLE_WRITE", raising=False)
        with pytest.raises(NotAllowedError):
            write_named_range(
                sample_workbook, name="X", sheet="Sheet1",
                range_address="A1", confirm=True,
            )

    @pytest.mark.unit
    def test_write_named_range_persisted(self, sample_workbook, monkeypatch):
        """Named range is readable back via get_named_ranges after save."""
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        write_named_range(
            sample_workbook, name="PersistTest", sheet="Sheet1",
            range_address="A1:A2", confirm=True,
        )
        _wb_cache.clear()
        names = [r["name"] for r in get_named_ranges(sample_workbook)]
        assert "PersistTest" in names

    # ── merge_cells ─────────────────────────────────────────────────────

    @pytest.mark.unit
    def test_merge_cells_happy(self, sample_workbook, monkeypatch):
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        result = merge_cells(sample_workbook, sheet="Sheet1", address="B1:C1", confirm=True)
        assert result["ok"] is True
        assert result["address"] == "B1:C1"
        assert "Merged" in result["message"]

    @pytest.mark.unit
    def test_merge_cells_confirm_false_raises(self, sample_workbook, monkeypatch):
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        with pytest.raises(ValidationError, match="confirm=True"):
            merge_cells(sample_workbook, sheet="Sheet1", address="B1:C1", confirm=False)

    @pytest.mark.unit
    def test_merge_cells_write_gate_raises(self, sample_workbook, monkeypatch):
        monkeypatch.delenv("EXCEL_ENABLE_WRITE", raising=False)
        with pytest.raises(NotAllowedError):
            merge_cells(sample_workbook, sheet="Sheet1", address="B1:C1", confirm=True)

    @pytest.mark.unit
    def test_merge_cells_persisted(self, sample_workbook, monkeypatch):
        """After merge, the openpyxl merged_cells set contains the range."""
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        merge_cells(sample_workbook, sheet="Sheet1", address="D1:E2", confirm=True)
        _wb_cache.clear()
        wb_check = openpyxl.load_workbook(sample_workbook)
        merged = [str(m) for m in wb_check["Sheet1"].merged_cells.ranges]
        assert "D1:E2" in merged

    # ── unmerge_cells ────────────────────────────────────────────────────

    @pytest.mark.unit
    def test_unmerge_cells_happy(self, sample_workbook, monkeypatch):
        """Merge then unmerge; result is ok and message says Unmerged."""
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        merge_cells(sample_workbook, sheet="Sheet1", address="B2:C2", confirm=True)
        _wb_cache.clear()
        result = unmerge_cells(sample_workbook, sheet="Sheet1", address="B2:C2", confirm=True)
        assert result["ok"] is True
        assert "Unmerged" in result["message"]

    @pytest.mark.unit
    def test_unmerge_cells_noop_range(self, tmp_path, monkeypatch):
        """unmerge_cells on a range that was never merged raises ValidationError (openpyxl raises ValueError)."""
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        p = tmp_path / "noop.xlsx"
        wb = openpyxl.Workbook()
        wb.active.title = "Sheet1"
        wb.save(str(p))
        with pytest.raises((ValidationError, ExcelMCPError)):
            unmerge_cells(str(p), sheet="Sheet1", address="B2:C3", confirm=True)

    @pytest.mark.unit
    def test_unmerge_cells_confirm_false_raises(self, sample_workbook, monkeypatch):
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        with pytest.raises(ValidationError, match="confirm=True"):
            unmerge_cells(sample_workbook, sheet="Sheet1", address="A1:A2", confirm=False)

    @pytest.mark.unit
    def test_unmerge_cells_write_gate_raises(self, sample_workbook, monkeypatch):
        monkeypatch.delenv("EXCEL_ENABLE_WRITE", raising=False)
        with pytest.raises(NotAllowedError):
            unmerge_cells(sample_workbook, sheet="Sheet1", address="A1:A2", confirm=True)

    # ── freeze_panes ────────────────────────────────────────────────────

    @pytest.mark.unit
    def test_freeze_panes_happy(self, sample_workbook, monkeypatch):
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        result = freeze_panes(sample_workbook, sheet="Sheet1", freeze_at="B2", confirm=True)
        assert result["ok"] is True
        assert result["freeze_at"] == "B2"
        assert "Frozen" in result["message"]

    @pytest.mark.unit
    def test_freeze_panes_unfreeze(self, sample_workbook, monkeypatch):
        """Passing freeze_at=None clears freeze; message says Unfrozen."""
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        freeze_panes(sample_workbook, sheet="Sheet1", freeze_at="B2", confirm=True)
        _wb_cache.clear()
        result = freeze_panes(sample_workbook, sheet="Sheet1", freeze_at=None, confirm=True)
        assert result["ok"] is True
        assert result["message"] == "Unfrozen"
        assert result.get("freeze_at") is None

    @pytest.mark.unit
    def test_freeze_panes_confirm_false_raises(self, sample_workbook, monkeypatch):
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        with pytest.raises(ValidationError, match="confirm=True"):
            freeze_panes(sample_workbook, sheet="Sheet1", freeze_at="A2", confirm=False)

    @pytest.mark.unit
    def test_freeze_panes_write_gate_raises(self, sample_workbook, monkeypatch):
        monkeypatch.delenv("EXCEL_ENABLE_WRITE", raising=False)
        with pytest.raises(NotAllowedError):
            freeze_panes(sample_workbook, sheet="Sheet1", freeze_at="A2", confirm=True)

    # ── auto_filter ──────────────────────────────────────────────────────

    @pytest.mark.unit
    def test_auto_filter_happy(self, sample_workbook, monkeypatch):
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        result = auto_filter(sample_workbook, sheet="Sheet1", address="A1:A2", confirm=True)
        assert result["ok"] is True
        assert result["address"] == "A1:A2"
        assert "AutoFilter applied" in result["message"]

    @pytest.mark.unit
    def test_auto_filter_clear(self, sample_workbook, monkeypatch):
        """Passing address=None clears the AutoFilter."""
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        auto_filter(sample_workbook, sheet="Sheet1", address="A1:A2", confirm=True)
        _wb_cache.clear()
        result = auto_filter(sample_workbook, sheet="Sheet1", address=None, confirm=True)
        assert result["ok"] is True
        assert "cleared" in result["message"]

    @pytest.mark.unit
    def test_auto_filter_confirm_false_raises(self, sample_workbook, monkeypatch):
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        with pytest.raises(ValidationError, match="confirm=True"):
            auto_filter(sample_workbook, sheet="Sheet1", address="A1:A2", confirm=False)

    @pytest.mark.unit
    def test_auto_filter_write_gate_raises(self, sample_workbook, monkeypatch):
        monkeypatch.delenv("EXCEL_ENABLE_WRITE", raising=False)
        with pytest.raises(NotAllowedError):
            auto_filter(sample_workbook, sheet="Sheet1", address="A1:A2", confirm=True)

    # ── N1: wrong-sheet guards (all 5 Phase 2.2 tools) ──────────────────

    @pytest.mark.unit
    def test_write_named_range_wrong_sheet_raises(self, sample_workbook, monkeypatch):
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        with pytest.raises(ValidationError, match="[Ss]heet"):
            write_named_range(sample_workbook, name="X", sheet="NoSuchSheet",
                              range_address="A1:A2", confirm=True)

    @pytest.mark.unit
    def test_merge_cells_wrong_sheet_raises(self, sample_workbook, monkeypatch):
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        with pytest.raises(ValidationError, match="[Ss]heet"):
            merge_cells(sample_workbook, sheet="NoSuchSheet", address="A1:B1", confirm=True)

    @pytest.mark.unit
    def test_unmerge_cells_wrong_sheet_raises(self, sample_workbook, monkeypatch):
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        with pytest.raises(ValidationError, match="[Ss]heet"):
            unmerge_cells(sample_workbook, sheet="NoSuchSheet", address="A1:B1", confirm=True)

    @pytest.mark.unit
    def test_freeze_panes_wrong_sheet_raises(self, sample_workbook, monkeypatch):
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        with pytest.raises(ValidationError, match="[Ss]heet"):
            freeze_panes(sample_workbook, sheet="NoSuchSheet", freeze_at="B2", confirm=True)

    @pytest.mark.unit
    def test_auto_filter_wrong_sheet_raises(self, sample_workbook, monkeypatch):
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        with pytest.raises(ValidationError, match="[Ss]heet"):
            auto_filter(sample_workbook, sheet="NoSuchSheet", address="A1:A2", confirm=True)

    # ── N2: invalid named-range name validation ──────────────────────────

    @pytest.mark.unit
    @pytest.mark.parametrize("bad_name", [
        "1startsWithDigit",   # must start with letter/underscore
        "has space",          # spaces not allowed
        "</xml>",             # XML injection payload rejected
        "foo;drop",           # semicolon not allowed
    ])
    def test_write_named_range_invalid_name_raises(self, sample_workbook, monkeypatch, bad_name):
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        with pytest.raises((ValidationError, ExcelMCPError)):
            write_named_range(sample_workbook, name=bad_name, sheet="Sheet1",
                              range_address="A1:A2", confirm=True)

    @pytest.mark.unit
    @pytest.mark.parametrize("reserved_name", [
        "__xlnm._FilterDatabase",
        "_xlfn.Reserved",
    ])
    def test_write_named_range_reserved_name_raises(self, sample_workbook, monkeypatch, reserved_name):
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        with pytest.raises((ValidationError, ExcelMCPError)):
            write_named_range(sample_workbook, name=reserved_name, sheet="Sheet1",
                              range_address="A1:A2", confirm=True)
