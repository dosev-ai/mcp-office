"""Tests for the create_table + auto_filter dual-<autoFilter> fix.

Root cause: after ``create_table`` adds an Excel Table, calling ``auto_filter``
on the same (or overlapping top-left) range writes a *second* ``<autoFilter>``
element to ``xl/worksheets/sheet1.xml``, conflicting with the one openpyxl
already embedded inside ``xl/tables/table1.xml``.  Excel's OOXML reader detects
the conflict, shows the "We found a problem ... content not readable" recovery
dialog, and silently drops both the Table *and* the AutoFilter.

Fix 1 – ``_advanced.py::auto_filter``
    When the requested address shares its top-left cell with any Excel Table on
    the sheet (condition: ``req_min[:2] == tbl_min[:2]``), suppress the
    worksheet-level autoFilter entirely (``ws.auto_filter.ref = None``) rather
    than writing it.  Return ``{"address": None, "warning": "<message>"}`` to
    signal the suppression to callers.

Fix 2 – ``_sheet.py::create_table``
    After ``ws.add_table(tbl)``, if ``ws.auto_filter.ref`` already occupies the
    same top-left cell as the new table range, clear it
    (``ws.auto_filter.ref = None``) to remove any pre-existing worksheet-level
    AutoFilter that would conflict with the new table's embedded one.

Test matrix
-----------
TC-AF-01  create_table then auto_filter (exact range)  → sheet autoFilter suppressed
TC-AF-02  create_table then auto_filter (header row only, same top-left)  → suppressed
TC-AF-03  auto_filter pre-set, then create_table  → sheet autoFilter cleared by Fix 2
TC-AF-04  auto_filter with NO table  → sheet autoFilter written (regression guard)
TC-AF-05  auto_filter on range with DIFFERENT top-left to existing table  → written
TC-AF-06  create_table golden path  → workbook opens clean, table accessible

All 6 tests run without a live COM object (openpyxl only).
No ``@pytest.mark.integration`` marker required.

NOTE: lxml is not in the project's dev dependencies.  XML assertions use the
stdlib ``xml.etree.ElementTree`` module, which provides identical element-search
semantics for the namespace-qualified lookups performed here.
"""

import xml.etree.ElementTree as ET
import zipfile
from pathlib import Path

import openpyxl
import pytest

from excelmcp.workbook_openpyxl import auto_filter, create_table

# ---------------------------------------------------------------------------
# OOXML namespace constant shared by worksheet and table XML parts
# ---------------------------------------------------------------------------

_SS_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_AF_TAG = f"{{{_SS_NS}}}autoFilter"


# ---------------------------------------------------------------------------
# XML inspection helpers (zipfile + ElementTree — no COM, no openpyxl load)
# ---------------------------------------------------------------------------

def _sheet_autofilter_elements(xlsx_path: str) -> list:
    """Return all ``<autoFilter>`` elements found in xl/worksheets/sheet1.xml.

    Returns an empty list when the element is absent; each item is an
    ``xml.etree.ElementTree.Element`` whose ``ref`` attribute can be inspected.
    """
    with zipfile.ZipFile(xlsx_path) as z:
        raw = z.read("xl/worksheets/sheet1.xml")
    root = ET.fromstring(raw)
    return root.findall(f".//{_AF_TAG}")


def _table_autofilter_elements(xlsx_path: str, table_index: int = 1) -> list:
    """Return all ``<autoFilter>`` elements from xl/tables/table{n}.xml."""
    with zipfile.ZipFile(xlsx_path) as z:
        raw = z.read(f"xl/tables/table{table_index}.xml")
    root = ET.fromstring(raw)
    return root.findall(f".//{_AF_TAG}")


def _table_file_exists(xlsx_path: str, table_index: int = 1) -> bool:
    """True when xl/tables/table{n}.xml is present in the archive."""
    with zipfile.ZipFile(xlsx_path) as z:
        return f"xl/tables/table{table_index}.xml" in z.namelist()


# ---------------------------------------------------------------------------
# Workbook factory helper
# ---------------------------------------------------------------------------

def _make_data_workbook(
    tmp_path: Path,
    monkeypatch,
    *,
    filename: str = "af_test.xlsx",
    first_row: int = 1,
    first_col: int = 1,
    n_cols: int = 3,
    n_rows: int = 5,
) -> str:
    """Create a minimal openpyxl workbook at ``tmp_path/filename``.

    Writes string headers at ``(first_row, first_col..first_col+n_cols-1)``
    and numeric filler values for ``n_rows-1`` data rows below.  Configures
    ``EXCEL_ALLOWLIST_ROOTS`` and ``EXCEL_ENABLE_WRITE`` via monkeypatch so
    MCP tools can operate on the file without validation errors.

    Returns the absolute path as a ``str`` suitable for passing to MCP tools.
    """
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")

    path = tmp_path / filename
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    for c_off in range(n_cols):
        ws.cell(first_row, first_col + c_off).value = f"H{c_off + 1}"
    for r_off in range(1, n_rows):
        for c_off in range(n_cols):
            ws.cell(first_row + r_off, first_col + c_off).value = (
                (first_row + r_off) * 10 + (first_col + c_off)
            )

    wb.save(str(path))
    return str(path)


# ===========================================================================
# TC-AF-01
# ``create_table`` → ``auto_filter`` (exact matching range)
# Expected: Fix 1 suppresses worksheet-level <autoFilter>
# ===========================================================================

@pytest.mark.unit
def test_create_table_then_auto_filter_suppresses_sheet_autofilter(
    tmp_path, monkeypatch
):
    """Sequence: create_table(A1:C5) → auto_filter(address=A1:C5).

    After Fix 1, ``auto_filter`` detects the table at the same top-left cell
    (A1 == A1) and must suppress the worksheet-level ``<autoFilter>`` entirely
    rather than writing it into the sheet XML.

    OOXML assertions (via zipfile, not openpyxl round-trip):
      - xl/worksheets/sheet1.xml : ZERO <autoFilter> elements
      - xl/tables/table1.xml     : at least ONE <autoFilter> element
                                   (openpyxl always embeds one in table XML)
    Return-value assertions:
      - result["ok"]      is True
      - result["address"] is None  (suppressed, not the requested address)
      - "warning"         in result (caller-visible suppression signal)
    """
    path = _make_data_workbook(tmp_path, monkeypatch)
    create_table(path, "Sheet1", table_range="A1:C5", table_name="T1", confirm=True)

    result = auto_filter(path, "Sheet1", address="A1:C5", confirm=True)

    # --- return-value assertions ---
    assert result["ok"] is True
    assert result["address"] is None, (
        f"Fix 1 must suppress address to None when a Table occupies the same "
        f"top-left cell; got address={result['address']!r}"
    )
    assert "warning" in result, (
        "Fix 1 must include a 'warning' key in the result when autoFilter is suppressed"
    )
    assert result["warning"], "warning value must be a non-empty string"

    # --- OOXML (XML) assertions ---
    sheet_afs = _sheet_autofilter_elements(path)
    assert sheet_afs == [], (
        f"xl/worksheets/sheet1.xml must contain NO <autoFilter> when a Table "
        f"covers the same top-left; found {len(sheet_afs)} element(s)"
    )
    assert _table_file_exists(path), "xl/tables/table1.xml not found in archive"
    table_afs = _table_autofilter_elements(path)
    assert table_afs != [], (
        "xl/tables/table1.xml must still contain its own embedded <autoFilter>"
    )


# ===========================================================================
# TC-AF-02
# ``create_table`` → ``auto_filter`` (header-row only / narrower range,
#                                     same top-left cell)
# Expected: Fix 1 suppresses even when the requested range doesn't exactly
#           match the table range — the top-left cell match is sufficient.
# ===========================================================================

@pytest.mark.unit
def test_create_table_then_auto_filter_header_row_only_suppresses_sheet_autofilter(
    tmp_path, monkeypatch
):
    """Sequence: create_table(A5:D15) → auto_filter(address=A5:A5).

    The requested range ``A5:A5`` is narrower than the table's ``A5:D15``
    but they share the same top-left cell (A5 == A5).  After Fix 1 the
    worksheet-level ``<autoFilter>`` must still be suppressed.

    This covers the common real-world call pattern where the caller passes
    only the header row to ``auto_filter`` rather than the full table range.
    """
    # Table anchored at A5; 4 cols × 11 rows (header + 10 data rows)
    path = _make_data_workbook(
        tmp_path, monkeypatch,
        first_row=5, first_col=1, n_cols=4, n_rows=11,
    )
    create_table(path, "Sheet1", table_range="A5:D15", table_name="T2", confirm=True)

    # Address = header row only (A5:A5) — same top-left, narrower
    result = auto_filter(path, "Sheet1", address="A5:A5", confirm=True)

    assert result["ok"] is True
    assert result["address"] is None, (
        f"Fix 1 must suppress when top-left cells match even if ranges differ; "
        f"got address={result['address']!r}"
    )
    assert "warning" in result

    sheet_afs = _sheet_autofilter_elements(path)
    assert sheet_afs == [], (
        "xl/worksheets/sheet1.xml must contain NO <autoFilter> (top-left match → suppressed)"
    )


# ===========================================================================
# TC-AF-03
# Reverse sequence: ``auto_filter`` pre-set → ``create_table`` (same range)
# Expected: Fix 2 clears the pre-existing worksheet-level <autoFilter>
# ===========================================================================

@pytest.mark.unit
def test_create_table_after_auto_filter_clears_sheet_autofilter(
    tmp_path, monkeypatch
):
    """Sequence: set auto_filter manually → create_table(A1:C5).

    When the workbook already has a worksheet-level ``<autoFilter ref="A1:C5"/>``
    and then ``create_table`` is called for the same range, Fix 2 must clear
    the worksheet-level autoFilter so only the table-embedded one remains.

    This guards the case where an earlier workflow set an autoFilter and now a
    table is being added to formalise the data range.
    """
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")

    path = tmp_path / "af_pre.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    # Write headers + data for A1:C5
    for c in range(1, 4):
        ws.cell(1, c).value = f"H{c}"
    for r in range(2, 6):
        for c in range(1, 4):
            ws.cell(r, c).value = r * 10 + c
    # Pre-set the worksheet-level autoFilter (simulates a prior auto_filter call)
    ws.auto_filter.ref = "A1:C5"
    wb.save(str(path))

    # Precondition: sheet XML has <autoFilter> before create_table
    pre_afs = _sheet_autofilter_elements(str(path))
    assert pre_afs != [], "Precondition failed: sheet must have <autoFilter> before create_table"

    result = create_table(
        str(path), "Sheet1", table_range="A1:C5", table_name="T3", confirm=True
    )

    assert result["table_name"] == "T3"

    # --- OOXML assertion: Fix 2 must have cleared the worksheet-level autoFilter ---
    sheet_afs = _sheet_autofilter_elements(str(path))
    assert sheet_afs == [], (
        f"Fix 2: create_table must clear pre-existing worksheet-level <autoFilter> "
        f"when it shares the table's top-left cell; found {len(sheet_afs)} element(s)"
    )


# ===========================================================================
# TC-AF-04
# ``auto_filter`` with no Excel Table present — normal path unchanged
# Expected: worksheet-level <autoFilter> IS written (regression guard)
# ===========================================================================

@pytest.mark.unit
def test_auto_filter_no_table_still_sets_sheet_autofilter(tmp_path, monkeypatch):
    """``auto_filter`` on a sheet with NO Excel Table must behave as before Fix 1.

    The fix must only suppress when a Table is present at the same top-left.
    Normal (non-table) worksheets must continue to receive a worksheet-level
    ``<autoFilter>`` element as they always did.
    """
    path = _make_data_workbook(tmp_path, monkeypatch)
    # Deliberately do NOT call create_table

    result = auto_filter(path, "Sheet1", address="A1:C5", confirm=True)

    # --- return-value assertions ---
    assert result["ok"] is True
    assert result["address"] == "A1:C5", (
        f"Without a Table, address must be echoed unchanged; got {result['address']!r}"
    )
    # "warning" must be absent or None — no suppression took place
    assert result.get("warning") is None, (
        f"No 'warning' expected when no Table is present; got {result.get('warning')!r}"
    )

    # --- OOXML assertion: autoFilter IS present ---
    sheet_afs = _sheet_autofilter_elements(path)
    assert len(sheet_afs) == 1, (
        f"xl/worksheets/sheet1.xml must contain exactly 1 <autoFilter> "
        f"when no Table is present; found {len(sheet_afs)}"
    )
    ref_attr = sheet_afs[0].get("ref")
    assert ref_attr == "A1:C5", (
        f"<autoFilter ref> must equal the requested address 'A1:C5'; got {ref_attr!r}"
    )


# ===========================================================================
# TC-AF-05
# ``create_table`` → ``auto_filter`` on a DIFFERENT top-left range
# Expected: worksheet-level <autoFilter> IS written (regression guard)
# ===========================================================================

@pytest.mark.unit
def test_create_table_then_auto_filter_different_topleft_still_sets_sheet_autofilter(
    tmp_path, monkeypatch
):
    """Table at A1:C5; ``auto_filter`` called with E1:G5 (top-left = E1 ≠ A1).

    Fix 1's condition is ``req_min[:2] == tbl_min[:2]`` (top-left columns AND
    rows must both match).  When they differ, no suppression must occur and the
    worksheet-level ``<autoFilter>`` must be written normally.

    This guards against over-eager suppression when an autoFilter is applied to
    a range that is not covered by any table on the sheet.
    """
    # Build a workbook wide enough to host data at both A1:C5 (table) and E1:G5
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")

    path = tmp_path / "af_diff.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    # Data at A1:C5 (for the table)
    for c in range(1, 4):
        ws.cell(1, c).value = f"TH{c}"
    for r in range(2, 6):
        for c in range(1, 4):
            ws.cell(r, c).value = r * 10 + c
    # Data at E1:G5 (for the separate autoFilter, column E=5, G=7)
    for c in range(5, 8):
        ws.cell(1, c).value = f"AH{c}"
    for r in range(2, 6):
        for c in range(5, 8):
            ws.cell(r, c).value = r * 10 + c
    wb.save(str(path))

    create_table(str(path), "Sheet1", table_range="A1:C5", table_name="T4", confirm=True)

    # Apply autoFilter to E1:G5 — an entirely different top-left (E1 vs A1)
    result = auto_filter(str(path), "Sheet1", address="E1:G5", confirm=True)

    # --- return-value assertions ---
    assert result["ok"] is True
    assert result["address"] == "E1:G5", (
        f"AutoFilter on a range with a different top-left must not be suppressed; "
        f"got address={result['address']!r}"
    )
    assert result.get("warning") is None, (
        "No 'warning' expected when autoFilter top-left differs from any table top-left"
    )

    # --- OOXML assertion: autoFilter IS present at E1:G5 ---
    sheet_afs = _sheet_autofilter_elements(str(path))
    assert len(sheet_afs) == 1, (
        f"Expected exactly 1 <autoFilter> in sheet XML (for E1:G5); "
        f"found {len(sheet_afs)}"
    )
    ref_attr = sheet_afs[0].get("ref")
    assert ref_attr == "E1:G5", (
        f"<autoFilter ref> must be 'E1:G5'; got {ref_attr!r}"
    )


# ===========================================================================
# TC-AF-06
# ``create_table`` golden path — workbook opens correctly after Fix 2
# Expected: no corruption; table is accessible; no <autoFilter> conflict
# ===========================================================================

@pytest.mark.unit
def test_create_table_does_not_corrupt_on_workbook_open(tmp_path, monkeypatch):
    """End-to-end golden path: write headers + data → create_table → reload.

    Verifies that a workbook containing an Excel Table written purely through
    MCP tools can be re-opened by openpyxl without raising any exception and
    that the table is accessible via ``ws.tables``.

    This is the primary smoke test confirming Fix 2 has not introduced any
    new serialisation issue: if the workbook were corrupt, ``openpyxl.load_workbook``
    would raise ``InvalidFileException``, ``zipfile.BadZipFile``, or similar.
    """
    path = _make_data_workbook(tmp_path, monkeypatch)
    result = create_table(
        path, "Sheet1", table_range="A1:C5", table_name="GoldenTable", confirm=True
    )
    assert result["table_name"] == "GoldenTable"
    assert result["ref"] == "A1:C5"

    # Reload from disk — must not raise
    wb2 = openpyxl.load_workbook(path)
    ws2 = wb2["Sheet1"]

    # Table must be accessible
    tables = list(ws2.tables.values())
    assert len(tables) == 1, f"Expected 1 table after reload; found {len(tables)}"
    assert tables[0].displayName == "GoldenTable"
    assert tables[0].ref == "A1:C5"

    # Sheet must contain NO worksheet-level <autoFilter> (only the table-embedded one)
    sheet_afs = _sheet_autofilter_elements(path)
    assert sheet_afs == [], (
        "Freshly created table workbook must not have a worksheet-level <autoFilter> "
        "that could conflict with the table-embedded one"
    )
