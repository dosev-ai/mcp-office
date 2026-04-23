"""Unit tests for excelmcp.review_excel — pure openpyxl, no COM, no live server.

All tests use tmp_path to create temporary workbooks with no filesystem side-effects.
"""
from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from excelmcp.review_excel import review_workbook_render


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _make_workbook(tmp_path: Path, filename: str = "test.xlsx") -> tuple[openpyxl.Workbook, Path]:
    """Create a workbook and return (wb, path) without saving."""
    wb = openpyxl.Workbook()
    p = tmp_path / filename
    return wb, p


def _save(wb: openpyxl.Workbook, path: Path) -> str:
    wb.save(str(path))
    return str(path)


# ---------------------------------------------------------------------------
# PRINT_AREA_UNSET
# ---------------------------------------------------------------------------

def test_print_area_unset_warns(tmp_path: Path) -> None:
    """Sheet with no print area should produce PRINT_AREA_UNSET WARN finding."""
    wb, path = _make_workbook(tmp_path)
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Header"
    ws["A2"] = "Data"
    # print_area intentionally not set (default = None)
    _save(wb, path)

    result = review_workbook_render(str(path))

    warn_findings = [
        f for f in result["findings"]
        if f["criterion"] == "PRINT_AREA_UNSET" and f["result"] == "WARN"
    ]
    assert warn_findings, "Expected PRINT_AREA_UNSET WARN finding but got none"
    assert warn_findings[0]["severity"] == "medium"
    assert warn_findings[0]["sheet"] == "Sheet1"


def test_print_area_set_passes(tmp_path: Path) -> None:
    """Sheet with print area set should NOT produce PRINT_AREA_UNSET WARN."""
    wb, path = _make_workbook(tmp_path)
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Header"
    ws["A2"] = "Data"
    ws.print_area = "A1:J10"
    _save(wb, path)

    result = review_workbook_render(str(path))

    warn_findings = [
        f for f in result["findings"]
        if f["criterion"] == "PRINT_AREA_UNSET" and f["result"] == "WARN"
    ]
    assert not warn_findings, f"Unexpected PRINT_AREA_UNSET WARN: {warn_findings}"


# ---------------------------------------------------------------------------
# SHEET_EMPTY
# ---------------------------------------------------------------------------

def test_sheet_empty_warns(tmp_path: Path) -> None:
    """Empty sheet should produce SHEET_EMPTY WARN finding."""
    wb, path = _make_workbook(tmp_path)
    ws = wb.active
    ws.title = "EmptySheet"
    # No data written — ws.max_row will be None or 1
    _save(wb, path)

    result = review_workbook_render(str(path))

    warn_findings = [
        f for f in result["findings"]
        if f["criterion"] == "SHEET_EMPTY" and f["result"] == "WARN"
    ]
    assert warn_findings, "Expected SHEET_EMPTY WARN finding for empty sheet"
    assert warn_findings[0]["severity"] == "low"
    assert warn_findings[0]["sheet"] == "EmptySheet"


def test_sheet_nonempty_passes(tmp_path: Path) -> None:
    """Sheet with data rows should NOT produce SHEET_EMPTY WARN."""
    wb, path = _make_workbook(tmp_path)
    ws = wb.active
    ws.title = "DataSheet"
    ws["A1"] = "Header"
    ws["A2"] = "Value1"
    ws["A3"] = "Value2"
    _save(wb, path)

    result = review_workbook_render(str(path))

    warn_findings = [
        f for f in result["findings"]
        if f["criterion"] == "SHEET_EMPTY" and f["result"] == "WARN"
    ]
    assert not warn_findings, f"Unexpected SHEET_EMPTY WARN: {warn_findings}"


# ---------------------------------------------------------------------------
# FORMULA_STALE
# ---------------------------------------------------------------------------

def test_formula_stale_warns(tmp_path: Path) -> None:
    """Workbook with formulas + recalculated=False should produce FORMULA_STALE WARN."""
    wb, path = _make_workbook(tmp_path)
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = 10
    ws["A2"] = 20
    ws["A3"] = "=SUM(A1:A2)"  # formula — will be detected by data_only=False scan
    _save(wb, path)

    result = review_workbook_render(str(path), recalculated=False)

    warn_findings = [
        f for f in result["findings"]
        if f["criterion"] == "FORMULA_STALE" and f["result"] == "WARN"
    ]
    assert warn_findings, "Expected FORMULA_STALE WARN for formula workbook with recalculated=False"
    assert warn_findings[0]["severity"] == "medium"


def test_formula_stale_suppressed(tmp_path: Path) -> None:
    """With recalculated=True the FORMULA_STALE finding should be suppressed."""
    wb, path = _make_workbook(tmp_path)
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = 10
    ws["A2"] = 20
    ws["A3"] = "=SUM(A1:A2)"
    _save(wb, path)

    result = review_workbook_render(str(path), recalculated=True)

    warn_findings = [
        f for f in result["findings"]
        if f["criterion"] == "FORMULA_STALE" and f["result"] == "WARN"
    ]
    assert not warn_findings, "FORMULA_STALE should be suppressed when recalculated=True"


# ---------------------------------------------------------------------------
# PRINT_AREA_EXCEEDS_PAGE
# ---------------------------------------------------------------------------

def test_print_area_exceeds_page_warns(tmp_path: Path) -> None:
    """Print area spanning > 10 columns should produce PRINT_AREA_EXCEEDS_PAGE WARN."""
    wb, path = _make_workbook(tmp_path)
    ws = wb.active
    ws.title = "WideSheet"
    ws["A1"] = "Data"
    ws["A2"] = "More"
    ws.print_area = "A1:AA50"  # 27 columns — exceeds 10
    _save(wb, path)

    result = review_workbook_render(str(path))

    warn_findings = [
        f for f in result["findings"]
        if f["criterion"] == "PRINT_AREA_EXCEEDS_PAGE" and f["result"] == "WARN"
    ]
    assert warn_findings, "Expected PRINT_AREA_EXCEEDS_PAGE WARN for 27-column print area"
    assert warn_findings[0]["severity"] == "low"
    assert "27" in warn_findings[0]["detail"]


def test_print_area_fits_page_no_warn(tmp_path: Path) -> None:
    """Print area <= 10 columns should NOT produce PRINT_AREA_EXCEEDS_PAGE WARN."""
    wb, path = _make_workbook(tmp_path)
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Data"
    ws["A2"] = "Row"
    ws.print_area = "A1:J10"  # exactly 10 columns — within limit
    _save(wb, path)

    result = review_workbook_render(str(path))

    warn_findings = [
        f for f in result["findings"]
        if f["criterion"] == "PRINT_AREA_EXCEEDS_PAGE" and f["result"] == "WARN"
    ]
    assert not warn_findings, f"Unexpected PRINT_AREA_EXCEEDS_PAGE WARN: {warn_findings}"


# ---------------------------------------------------------------------------
# passed flag
# ---------------------------------------------------------------------------

def test_review_passed_flag(tmp_path: Path) -> None:
    """Workbook with print area, data, and recalculated=True should return passed=True."""
    wb, path = _make_workbook(tmp_path)
    ws = wb.active
    ws.title = "Report"
    ws["A1"] = "Header"
    ws["A2"] = "Row1"
    ws["B2"] = "Row1B"
    ws.print_area = "A1:J20"
    _save(wb, path)

    result = review_workbook_render(str(path), recalculated=True)

    assert result["passed"] is True, (
        f"Expected passed=True but got False. Findings: {result['findings']}"
    )


def test_review_passed_false_on_medium_warn(tmp_path: Path) -> None:
    """Workbook with PRINT_AREA_UNSET (medium severity) should return passed=False."""
    wb, path = _make_workbook(tmp_path)
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Header"
    ws["A2"] = "Data"
    # Intentionally no print_area → PRINT_AREA_UNSET WARN medium
    _save(wb, path)

    result = review_workbook_render(str(path), recalculated=True)

    assert result["passed"] is False, (
        "Expected passed=False due to PRINT_AREA_UNSET medium WARN"
    )


# ---------------------------------------------------------------------------
# Result schema
# ---------------------------------------------------------------------------

def test_review_result_schema(tmp_path: Path) -> None:
    """Result dict must contain all required keys."""
    wb, path = _make_workbook(tmp_path)
    ws = wb.active
    ws["A1"] = "X"
    ws["A2"] = "Y"
    ws.print_area = "A1:E10"
    _save(wb, path)

    result = review_workbook_render(str(path))

    required_keys = {"path", "sheets_reviewed", "passed", "findings", "checks_run", "findings_count"}
    missing = required_keys - set(result.keys())
    assert not missing, f"Result missing keys: {missing}"
    assert isinstance(result["path"], str)
    assert isinstance(result["sheets_reviewed"], list)
    assert isinstance(result["passed"], bool)
    assert isinstance(result["findings"], list)
    assert isinstance(result["checks_run"], int)
    assert isinstance(result["findings_count"], int)
    assert result["findings_count"] == len(result["findings"])


def test_finding_schema(tmp_path: Path) -> None:
    """Each finding in the findings list must contain all required keys."""
    wb, path = _make_workbook(tmp_path)
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Header"
    # No print area, no data beyond row 1 — produces SHEET_EMPTY and PRINT_AREA_UNSET findings
    _save(wb, path)

    result = review_workbook_render(str(path))

    required_finding_keys = {"sheet", "criterion", "result", "detail", "severity"}
    for finding in result["findings"]:
        missing = required_finding_keys - set(finding.keys())
        assert not missing, f"Finding missing keys {missing}: {finding}"
        assert finding["result"] in ("WARN", "OK"), f"Unexpected result value: {finding['result']}"
        assert finding["severity"] in ("high", "medium", "low"), (
            f"Unexpected severity: {finding['severity']}"
        )


# ---------------------------------------------------------------------------
# sheet_names parameter
# ---------------------------------------------------------------------------

def test_none_sheet_names_reviews_all(tmp_path: Path) -> None:
    """sheet_names=None should review all non-chart sheets."""
    wb, path = _make_workbook(tmp_path)
    ws1 = wb.active
    ws1.title = "Alpha"
    ws1["A1"] = "a"
    ws1["A2"] = "b"
    ws1.print_area = "A1:E5"
    ws2 = wb.create_sheet("Beta")
    ws2["A1"] = "x"
    ws2["A2"] = "y"
    ws2.print_area = "A1:D5"
    _save(wb, path)

    result = review_workbook_render(str(path), sheet_names=None)

    assert "Alpha" in result["sheets_reviewed"]
    assert "Beta" in result["sheets_reviewed"]
    assert len(result["sheets_reviewed"]) == 2


def test_specific_sheet_names_filters(tmp_path: Path) -> None:
    """Passing specific sheet_names should review only those sheets."""
    wb, path = _make_workbook(tmp_path)
    ws1 = wb.active
    ws1.title = "Alpha"
    ws1["A1"] = "a"
    ws1["A2"] = "b"
    ws1.print_area = "A1:E5"
    ws2 = wb.create_sheet("Beta")
    ws2["A1"] = "x"
    ws2["A2"] = "y"
    ws2.print_area = "A1:D5"
    _save(wb, path)

    result = review_workbook_render(str(path), sheet_names=["Alpha"])

    assert result["sheets_reviewed"] == ["Alpha"]
    # Beta was not reviewed — no findings should reference it
    beta_findings = [f for f in result["findings"] if f.get("sheet") == "Beta"]
    assert not beta_findings, f"Found findings for Beta sheet (should be excluded): {beta_findings}"


# ---------------------------------------------------------------------------
# L-001: produce_export_evidence_bundle happy path
# ---------------------------------------------------------------------------

def test_produce_export_evidence_bundle_happy_path() -> None:
    """produce_export_evidence_bundle returns a well-formed bundle on valid inputs."""
    from excelmcp.server_review import produce_export_evidence_bundle
    export_results = [{"sha256": "abc123", "exported_at": "2026-01-01T00:00:00Z", "success": True}]
    review_results = [{"passed": True, "findings_count": 0, "findings": []}]
    # _check_path not called in produce_export_evidence_bundle; path is metadata-only
    result = produce_export_evidence_bundle(
        path="fake.xlsx",
        export_results=export_results,
        review_results=review_results,
    )
    assert result["overall_passed"] is True
    assert result["bundle_version"] == "1.0"
    assert "generated_at" in result
    assert result["total_exports"] == 1
    assert result["total_review_findings"] == 0


# ---------------------------------------------------------------------------
# L-002: produce_export_evidence_bundle — review_results not a list
# ---------------------------------------------------------------------------

def test_produce_evidence_bundle_review_results_not_list() -> None:
    """produce_export_evidence_bundle raises ToolError if review_results is not a list."""
    from excelmcp.server_review import produce_export_evidence_bundle
    from fastmcp.exceptions import ToolError
    # _check_path not called in produce_export_evidence_bundle; path is metadata-only
    with pytest.raises(ToolError, match="review_results must be a list"):
        produce_export_evidence_bundle(
            path="fake.xlsx",
            export_results=[],
            review_results="not-a-list",
        )


# ---------------------------------------------------------------------------
# L-003: review_workbook_render gracefully handles chart-sheet AttributeError
# ---------------------------------------------------------------------------

# ---------------------------------------------------------------------------
# W4: export_changed_sheets_only
# ---------------------------------------------------------------------------

def test_export_changed_sheets_only_requires_confirm(monkeypatch) -> None:
    """export_changed_sheets_only raises ValidationError when confirm is not True."""
    from excelmcp.server_review import export_changed_sheets_only
    from excelmcp._core import ValidationError

    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    with pytest.raises(ValidationError, match="confirm=True required"):
        export_changed_sheets_only(
            path="/any/path.xlsx",
            base_dir="/any/dir",
            previous_hashes={},
            confirm=False,
        )


def test_export_changed_sheets_only_graceful_without_com(tmp_path: Path) -> None:
    """With EXCEL_ENABLE_COM unset, returns graceful fallback dict with hashes."""
    import os
    from unittest.mock import patch
    from excelmcp.server_review import export_changed_sheets_only

    # Create a minimal workbook for hashing
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "test_value"
    ws["A2"] = "row2"
    p = tmp_path / "graceful.xlsx"
    wb.save(str(p))

    # Build env: EXCEL_ENABLE_WRITE=true, EXCEL_ALLOWLIST_ROOTS=tmp_path, no EXCEL_ENABLE_COM
    env = {k: v for k, v in os.environ.items() if k != "EXCEL_ENABLE_COM"}
    env["EXCEL_ENABLE_WRITE"] = "true"
    env["EXCEL_ALLOWLIST_ROOTS"] = str(tmp_path)

    with patch.dict(os.environ, env, clear=True):
        result = export_changed_sheets_only(
            path=str(p),
            base_dir=str(tmp_path),
            previous_hashes={},
            confirm=True,
        )

    assert result["ok"] is False
    assert "EXCEL_ENABLE_COM" in result["error"]
    assert "new_hashes" in result
    assert isinstance(result["new_hashes"], dict)
    assert "Sheet1" in result["new_hashes"]
    assert result["exported_sheets"] == []
    assert result["run_dir"] is None
    assert result["export_details"] == []


def test_review_workbook_render_no_crash_on_chartsheet(tmp_path: Path) -> None:
    """Chart-sheet AttributeError in print_area is handled without crashing."""
    from unittest.mock import patch, PropertyMock
    p = tmp_path / "test.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "data"
    wb.save(str(p))
    # Patch print_area property on the worksheet class to raise AttributeError
    with patch(
        "openpyxl.worksheet.worksheet.Worksheet.print_area",
        new_callable=PropertyMock,
        side_effect=AttributeError,
    ):
        result = review_workbook_render(str(p), recalculated=True)
    # Should not raise; PRINT_AREA_UNSET finding should appear
    assert result is not None
    assert isinstance(result["findings"], list)


# ---------------------------------------------------------------------------
# L-004: review_workbook_render raises on nonexistent path
# ---------------------------------------------------------------------------

def test_review_workbook_render_invalid_path_raises() -> None:
    """review_workbook_render raises FileNotFoundError for a nonexistent path."""
    with pytest.raises(FileNotFoundError):
        review_workbook_render("/nonexistent/path/that/does/not/exist.xlsx")
