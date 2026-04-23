import pytest
import openpyxl

from excelmcp.workbook_openpyxl import (
    get_hyperlinks,
    add_hyperlink,
    remove_hyperlink,
    set_sheet_properties,
    bulk_add_hyperlinks,
    ExcelMCPError,
    ValidationError,
    NotAllowedError,
    _wb_cache,
)


class TestPhase27Hyperlinks:
    """Phase 2.7 — add_hyperlink, remove_hyperlink, get_hyperlinks."""

    @pytest.mark.unit
    def test_get_hyperlinks_empty(self, tmp_path, monkeypatch):
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        wb_path = tmp_path / "hlinks.xlsx"
        _wb = openpyxl.Workbook()
        _wb.active.title = "Sheet1"
        _wb.save(str(wb_path))
        result = get_hyperlinks(str(wb_path), "Sheet1")
        assert result["count"] == 0
        assert result["hyperlinks"] == []

    @pytest.mark.unit
    def test_get_hyperlinks_with_link(self, tmp_path, monkeypatch):
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        from openpyxl.worksheet.hyperlink import Hyperlink
        wb_path = tmp_path / "hlinks2.xlsx"
        _wb = openpyxl.Workbook()
        ws = _wb.active
        ws.title = "Sheet1"
        ws["A1"].hyperlink = Hyperlink(ref="A1", target="https://example.com")
        ws["A1"].value = "Click here"
        _wb.save(str(wb_path))
        result = get_hyperlinks(str(wb_path), "Sheet1")
        assert result["count"] == 1
        assert result["hyperlinks"][0]["url"] == "https://example.com"

    @pytest.mark.unit
    def test_get_hyperlinks_sheet_not_found(self, tmp_path, monkeypatch):
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        wb_path = tmp_path / "hlinks3.xlsx"
        openpyxl.Workbook().save(str(wb_path))
        with pytest.raises(ValidationError, match="Sheet not found"):
            get_hyperlinks(str(wb_path), "NoSuch")

    @pytest.mark.unit
    def test_get_hyperlinks_not_allowed(self, tmp_path, monkeypatch):
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        with pytest.raises((ExcelMCPError, NotAllowedError, ValidationError)):
            get_hyperlinks("C:\\other\\file.xlsx", "Sheet1")

    @pytest.mark.unit
    def test_add_hyperlink_happy(self, tmp_path, monkeypatch):
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        wb_path = tmp_path / "add_hlink.xlsx"
        _wb = openpyxl.Workbook()
        _wb.active.title = "Sheet1"
        _wb.save(str(wb_path))
        result = add_hyperlink(str(wb_path), "Sheet1", "A1", "https://example.com",
                               display_text="Click", confirm=True)
        assert result["ok"] is True
        assert result["url"] == "https://example.com"
        # Verify persisted
        _wb_cache.clear()
        _wb2 = openpyxl.load_workbook(str(wb_path))
        assert _wb2["Sheet1"]["A1"].hyperlink is not None

    @pytest.mark.unit
    def test_add_hyperlink_not_allowed(self, tmp_path, monkeypatch):
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        with pytest.raises((ExcelMCPError, NotAllowedError, ValidationError)):
            add_hyperlink("C:\\other\\file.xlsx", "Sheet1", "A1",
                          "https://example.com", confirm=True)

    @pytest.mark.unit
    def test_add_hyperlink_sheet_not_found(self, tmp_path, monkeypatch):
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        wb_path = tmp_path / "add_hlink2.xlsx"
        openpyxl.Workbook().save(str(wb_path))
        with pytest.raises(ValidationError, match="Sheet not found"):
            add_hyperlink(str(wb_path), "NoSuch", "A1", "https://example.com",
                          confirm=True)

    @pytest.mark.unit
    def test_remove_hyperlink_happy(self, tmp_path, monkeypatch):
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        from openpyxl.worksheet.hyperlink import Hyperlink
        wb_path = tmp_path / "rem_hlink.xlsx"
        _wb = openpyxl.Workbook()
        ws = _wb.active
        ws.title = "Sheet1"
        ws["A1"].hyperlink = Hyperlink(ref="A1", target="https://example.com")
        _wb.save(str(wb_path))
        result = remove_hyperlink(str(wb_path), "Sheet1", "A1", confirm=True)
        assert result["ok"] is True
        # Verify removed
        _wb_cache.clear()
        _wb2 = openpyxl.load_workbook(str(wb_path))
        assert _wb2["Sheet1"]["A1"].hyperlink is None

    @pytest.mark.unit
    def test_remove_hyperlink_not_allowed(self, tmp_path, monkeypatch):
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        with pytest.raises((ExcelMCPError, NotAllowedError, ValidationError)):
            remove_hyperlink("C:\\other\\file.xlsx", "Sheet1", "A1", confirm=True)

    @pytest.mark.unit
    def test_remove_hyperlink_sheet_not_found(self, tmp_path, monkeypatch):
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        wb_path = tmp_path / "rem_hlink2.xlsx"
        openpyxl.Workbook().save(str(wb_path))
        with pytest.raises(ValidationError, match="Sheet not found"):
            remove_hyperlink(str(wb_path), "NoSuch", "A1", confirm=True)

    @pytest.mark.unit
    def test_add_hyperlink_write_disabled(self, tmp_path, monkeypatch):
        monkeypatch.delenv("EXCEL_ENABLE_WRITE", raising=False)
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        with pytest.raises(NotAllowedError, match="EXCEL_ENABLE_WRITE"):
            add_hyperlink(str(tmp_path / "x.xlsx"), "Sheet1", "A1",
                          "https://example.com", confirm=True)

    @pytest.mark.unit
    def test_remove_hyperlink_write_disabled(self, tmp_path, monkeypatch):
        monkeypatch.delenv("EXCEL_ENABLE_WRITE", raising=False)
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        with pytest.raises(NotAllowedError, match="EXCEL_ENABLE_WRITE"):
            remove_hyperlink(str(tmp_path / "x.xlsx"), "Sheet1", "A1", confirm=True)


class TestPhase27ZoomScale:
    """Phase 2.7 — set_sheet_properties zoom_scale extension."""

    @pytest.mark.unit
    def test_zoom_scale_valid(self, tmp_path, monkeypatch):
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        wb_path = tmp_path / "zoom.xlsx"
        _wb = openpyxl.Workbook()
        _wb.active.title = "Sheet1"
        _wb.save(str(wb_path))
        result = set_sheet_properties(str(wb_path), "Sheet1", zoom_scale=150, confirm=True)
        assert result["ok"] is True
        assert result["zoom_scale"] == 150

    @pytest.mark.unit
    def test_zoom_scale_out_of_range_low(self, tmp_path, monkeypatch):
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        wb_path = tmp_path / "zoom_low.xlsx"
        _wb = openpyxl.Workbook()
        _wb.active.title = "Sheet1"
        _wb.save(str(wb_path))
        with pytest.raises((ValidationError, ExcelMCPError)):
            set_sheet_properties(str(wb_path), "Sheet1", zoom_scale=5, confirm=True)

    @pytest.mark.unit
    def test_zoom_scale_out_of_range_high(self, tmp_path, monkeypatch):
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        wb_path = tmp_path / "zoom_high.xlsx"
        _wb = openpyxl.Workbook()
        _wb.active.title = "Sheet1"
        _wb.save(str(wb_path))
        with pytest.raises((ValidationError, ExcelMCPError)):
            set_sheet_properties(str(wb_path), "Sheet1", zoom_scale=500, confirm=True)

    @pytest.mark.unit
    def test_zoom_scale_none_no_op(self, tmp_path, monkeypatch):
        """zoom_scale=None should not raise or change anything."""
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        wb_path = tmp_path / "zoom_none.xlsx"
        _wb = openpyxl.Workbook()
        _wb.active.title = "Sheet1"
        _wb.save(str(wb_path))
        result = set_sheet_properties(str(wb_path), "Sheet1", zoom_scale=None, confirm=True)
        assert result["ok"] is True


class TestHyperlinkUrlValidation:
    """Security backport tests for _validate_hyperlink_url and add_hyperlink."""

    @pytest.mark.unit
    def test_add_hyperlink_rejects_javascript(self, tmp_path, monkeypatch):
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        wb_path = tmp_path / "sec.xlsx"
        _wb = openpyxl.Workbook()
        _wb.active.title = "Sheet1"
        _wb.save(str(wb_path))
        with pytest.raises(ValidationError, match="not allowed"):
            add_hyperlink(str(wb_path), "Sheet1", "A1", "javascript:alert(1)", confirm=True)

    @pytest.mark.unit
    def test_add_hyperlink_rejects_vbscript(self, tmp_path, monkeypatch):
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        wb_path = tmp_path / "sec2.xlsx"
        _wb = openpyxl.Workbook()
        _wb.active.title = "Sheet1"
        _wb.save(str(wb_path))
        with pytest.raises(ValidationError, match="not allowed"):
            add_hyperlink(str(wb_path), "Sheet1", "A1", "vbscript:msgbox(1)", confirm=True)

    @pytest.mark.unit
    def test_add_hyperlink_rejects_data_url(self, tmp_path, monkeypatch):
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        wb_path = tmp_path / "sec3.xlsx"
        _wb = openpyxl.Workbook()
        _wb.active.title = "Sheet1"
        _wb.save(str(wb_path))
        with pytest.raises(ValidationError, match="not allowed"):
            add_hyperlink(str(wb_path), "Sheet1", "A1", "data:text/html,<h1>x</h1>", confirm=True)

    @pytest.mark.unit
    def test_add_hyperlink_rejects_file_url_by_default(self, tmp_path, monkeypatch):
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        monkeypatch.delenv("EXCEL_ALLOW_FILE_HYPERLINKS", raising=False)
        wb_path = tmp_path / "sec4.xlsx"
        _wb = openpyxl.Workbook()
        _wb.active.title = "Sheet1"
        _wb.save(str(wb_path))
        with pytest.raises(ValidationError, match="file"):
            add_hyperlink(str(wb_path), "Sheet1", "A1", "file:///etc/passwd", confirm=True)

    @pytest.mark.unit
    def test_add_hyperlink_allows_https(self, tmp_path, monkeypatch):
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        wb_path = tmp_path / "sec5.xlsx"
        _wb = openpyxl.Workbook()
        _wb.active.title = "Sheet1"
        _wb.save(str(wb_path))
        result = add_hyperlink(str(wb_path), "Sheet1", "A1", "https://example.com", confirm=True)
        assert result["ok"] is True

    @pytest.mark.unit
    def test_add_hyperlink_allows_mailto(self, tmp_path, monkeypatch):
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        wb_path = tmp_path / "sec6.xlsx"
        _wb = openpyxl.Workbook()
        _wb.active.title = "Sheet1"
        _wb.save(str(wb_path))
        result = add_hyperlink(str(wb_path), "Sheet1", "A1", "mailto:test@example.com", confirm=True)
        assert result["ok"] is True


class TestBulkAddHyperlinks:
    """Phase 2.7 Sprint C — bulk_add_hyperlinks (includes security tests)."""

    def _make_wb(self, tmp_path, monkeypatch, sheet="Sheet1"):
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        monkeypatch.delenv("EXCEL_ALLOW_FILE_HYPERLINKS", raising=False)
        wb_path = tmp_path / "bah.xlsx"
        _wb = openpyxl.Workbook()
        _wb.active.title = sheet
        _wb.save(str(wb_path))
        return str(wb_path)

    @pytest.mark.unit
    def test_bulk_add_hyperlinks_single_link(self, tmp_path, monkeypatch):
        path = self._make_wb(tmp_path, monkeypatch)
        result = bulk_add_hyperlinks(
            path, "Sheet1",
            [{"address": "A1", "url": "https://example.com"}],
            confirm=True
        )
        assert result["ok"] is True
        assert result["links_added"] == 1
        _wb_cache.clear()
        wb = openpyxl.load_workbook(path)
        assert wb["Sheet1"]["A1"].hyperlink is not None

    @pytest.mark.unit
    def test_bulk_add_hyperlinks_multiple_links(self, tmp_path, monkeypatch):
        path = self._make_wb(tmp_path, monkeypatch)
        links = [
            {"address": "A1", "url": "https://a.com"},
            {"address": "B2", "url": "https://b.com"},
            {"address": "C3", "url": "https://c.com"},
        ]
        result = bulk_add_hyperlinks(path, "Sheet1", links, confirm=True)
        assert result["links_added"] == 3

    @pytest.mark.unit
    def test_bulk_add_hyperlinks_empty_list(self, tmp_path, monkeypatch):
        path = self._make_wb(tmp_path, monkeypatch)
        with pytest.raises(ValidationError):
            bulk_add_hyperlinks(path, "Sheet1", [], confirm=True)

    @pytest.mark.unit
    def test_bulk_add_hyperlinks_reject_javascript(self, tmp_path, monkeypatch):
        path = self._make_wb(tmp_path, monkeypatch)
        with pytest.raises(ValidationError, match="not allowed"):
            bulk_add_hyperlinks(
                path, "Sheet1",
                [{"address": "A1", "url": "javascript:alert(1)"}],
                confirm=True,
            )

    @pytest.mark.unit
    def test_bulk_add_hyperlinks_reject_vbscript(self, tmp_path, monkeypatch):
        path = self._make_wb(tmp_path, monkeypatch)
        with pytest.raises(ValidationError, match="not allowed"):
            bulk_add_hyperlinks(
                path, "Sheet1",
                [{"address": "A1", "url": "vbscript:msgbox(1)"}],
                confirm=True,
            )

    @pytest.mark.unit
    def test_bulk_add_hyperlinks_reject_data_url(self, tmp_path, monkeypatch):
        path = self._make_wb(tmp_path, monkeypatch)
        with pytest.raises(ValidationError, match="not allowed"):
            bulk_add_hyperlinks(
                path, "Sheet1",
                [{"address": "A1", "url": "data:text/html,<h1>x</h1>"}],
                confirm=True,
            )

    @pytest.mark.unit
    def test_bulk_add_hyperlinks_reject_file_url_default(self, tmp_path, monkeypatch):
        path = self._make_wb(tmp_path, monkeypatch)
        with pytest.raises(ValidationError, match="file"):
            bulk_add_hyperlinks(
                path, "Sheet1",
                [{"address": "A1", "url": "file:///etc/passwd"}],
                confirm=True,
            )

    @pytest.mark.unit
    def test_bulk_add_hyperlinks_allow_file_url_with_env(self, tmp_path, monkeypatch):
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        monkeypatch.setenv("EXCEL_ALLOW_FILE_HYPERLINKS", "true")
        path = tmp_path / "bah_file.xlsx"
        _wb = openpyxl.Workbook()
        _wb.active.title = "Sheet1"
        _wb.save(str(path))
        result = bulk_add_hyperlinks(
            str(path), "Sheet1",
            [{"address": "A1", "url": "file:///C:/data/report.xlsx"}],
            confirm=True,
        )
        assert result["ok"] is True

    @pytest.mark.unit
    def test_bulk_add_hyperlinks_display_text_sets_cell_value(self, tmp_path, monkeypatch):
        path = self._make_wb(tmp_path, monkeypatch)
        bulk_add_hyperlinks(
            path, "Sheet1",
            [{"address": "A1", "url": "https://example.com", "display_text": "Click here"}],
            confirm=True,
        )
        _wb_cache.clear()
        wb = openpyxl.load_workbook(path)
        assert wb["Sheet1"]["A1"].value == "Click here"

    @pytest.mark.unit
    def test_bulk_add_hyperlinks_atomicity_partial_invalid(self, tmp_path, monkeypatch):
        """All items validated before any write; one bad URL means zero links added."""
        path = self._make_wb(tmp_path, monkeypatch)
        with pytest.raises(ValidationError, match="not allowed"):
            bulk_add_hyperlinks(
                path, "Sheet1",
                [
                    {"address": "A1", "url": "https://example.com"},
                    {"address": "B2", "url": "javascript:alert(1)"},
                ],
                confirm=True,
            )
        # Workbook must be unchanged — no hyperlink on A1
        _wb_cache.clear()
        wb = openpyxl.load_workbook(path)
        assert wb["Sheet1"]["A1"].hyperlink is None
