"""
Smoke tests — Phase 1 and Phase 2.0 tools (direct call-through).

Tests every Phase-1/2.0 tool by calling workbook_openpyxl functions directly
(not via MCP dispatch) to verify happy-path behaviour and blocked-path gates.

Previously these tools had only registry verification; this file provides
per-tool call-through coverage.

No COM required; always CI-runnable.
"""
import pytest
import openpyxl
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.table import Table

import excelmcp.workbook_openpyxl as wb
from excelmcp.workbook_openpyxl import ExcelMCPError


# ---------------------------------------------------------------------------
# Discovery
# ---------------------------------------------------------------------------


@pytest.mark.smoke
def test_smoke_p1_capabilities():
    result = wb.capabilities()
    assert result["phase"] == "2.0"
    assert "range_io" in result["tools"]
    assert "workbook_metadata" in result["tools"]
    assert len(result["tools"]) == len(set(result["tools"]))


# ---------------------------------------------------------------------------
# Read tools
# ---------------------------------------------------------------------------


@pytest.mark.smoke
def test_smoke_p1_list_sheets(sample_workbook):
    result = wb.list_sheets(sample_workbook)
    assert "Sheet1" in result


@pytest.mark.smoke
def test_smoke_p1_get_used_range(sample_workbook):
    result = wb.get_used_range(sample_workbook, "Sheet1")
    assert result["min_row"] == 1
    assert "address" in result


@pytest.mark.smoke
def test_smoke_p1_read_cell(sample_workbook):
    result = wb.read_cell(sample_workbook, "Sheet1", "A1")
    assert result["address"] == "A1"
    assert result["value"] == "hello"


@pytest.mark.smoke
def test_smoke_p1_read_range(sample_workbook):
    result = wb.read_range(sample_workbook, "Sheet1", "A1:A2")
    assert isinstance(result, list)
    assert len(result) == 2
    assert result[0][0]["value"] == "hello"


@pytest.mark.smoke
def test_smoke_p1_read_sheet_all(sample_workbook):
    result = wb.read_sheet_all(sample_workbook, "Sheet1")
    assert "rows" in result
    assert "data" in result
    assert result["rows"] >= 2


@pytest.mark.smoke
def test_smoke_p1_get_named_ranges_empty(sample_workbook):
    result = wb.get_named_ranges(sample_workbook)
    assert isinstance(result, list)


@pytest.mark.smoke
def test_smoke_p1_read_named_range(tmp_path, monkeypatch):
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    path = str(tmp_path / "nr.xlsx")
    _wb = openpyxl.Workbook()
    ws = _wb.active
    ws.title = "Sheet1"
    ws["A1"] = 10
    ws["A2"] = 20
    dn = DefinedName("MyRange", attr_text="Sheet1!$A$1:$A$2")
    _wb.defined_names.add(dn)
    _wb.save(path)

    result = wb.read_named_range(path, "MyRange")
    assert result["name"] == "MyRange"
    assert "data" in result


@pytest.mark.smoke
def test_smoke_p1_get_cell_metadata(sample_workbook):
    result = wb.get_cell_metadata(sample_workbook, "Sheet1", "A1")
    assert result["address"] == "A1"
    assert "font" in result


@pytest.mark.smoke
def test_smoke_p1_evaluate_formula(sample_workbook):
    result = wb.evaluate_formula(sample_workbook, "Sheet1", "A1")
    assert result["address"] == "A1"
    assert "formula" in result
    assert "cached_value" in result


@pytest.mark.smoke
def test_smoke_p1_list_tables_empty(sample_workbook):
    result = wb.list_tables(sample_workbook, "Sheet1")
    assert isinstance(result, list)
    assert result == []


@pytest.mark.smoke
def test_smoke_p1_read_table(tmp_path, monkeypatch):
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    path = str(tmp_path / "tbl.xlsx")
    _wb = openpyxl.Workbook()
    ws = _wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Name"
    ws["B1"] = "Value"
    ws["A2"] = "Alpha"
    ws["B2"] = 1
    ws.add_table(Table(displayName="MyTable", ref="A1:B2"))
    _wb.save(path)

    result = wb.read_table(path, "Sheet1", "MyTable")
    assert result["table_name"] == "MyTable"
    assert "headers" in result


# ---------------------------------------------------------------------------
# Write tools
# ---------------------------------------------------------------------------


@pytest.mark.smoke
def test_smoke_p1_write_cell(sample_workbook, monkeypatch):
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    result = wb.write_cell(sample_workbook, "Sheet1", "C1", 999, confirm=True)
    assert result["ok"] is True


@pytest.mark.smoke
def test_smoke_p1_write_cell_no_confirm_raises(sample_workbook, monkeypatch):
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    with pytest.raises(ExcelMCPError):
        wb.write_cell(sample_workbook, "Sheet1", "C1", 1, confirm=False)


@pytest.mark.smoke
def test_smoke_p1_write_range(sample_workbook, monkeypatch):
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    result = wb.write_range(sample_workbook, "Sheet1", "D1", [[1, 2], [3, 4]], confirm=True)
    assert result["ok"] is True


@pytest.mark.smoke
def test_smoke_p1_append_rows(sample_workbook, monkeypatch):
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    result = wb.append_rows(sample_workbook, "Sheet1", [["x", "y"], ["a", "b"]], confirm=True)
    assert result["ok"] is True
    assert result["rows_appended"] == 2


@pytest.mark.smoke
def test_smoke_p1_find_replace(sample_workbook, monkeypatch):
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    result = wb.find_replace(
        sample_workbook, "Sheet1", "hello", "world", confirm=True
    )
    assert result["ok"] is True
    assert result["replacements_made"] >= 1


# ---------------------------------------------------------------------------
# Sheet management
# ---------------------------------------------------------------------------


@pytest.mark.smoke
def test_smoke_p1_add_sheet(sample_workbook, monkeypatch):
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    result = wb.add_sheet(sample_workbook, "NewSheet", confirm=True)
    assert result["ok"] is True
    assert result["sheet_name"] == "NewSheet"


@pytest.mark.smoke
def test_smoke_p1_rename_sheet(sample_workbook, monkeypatch):
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    wb.add_sheet(sample_workbook, "TempSheet", confirm=True)
    result = wb.rename_sheet(sample_workbook, "TempSheet", "RenamedSheet", confirm=True)
    assert result["ok"] is True
    assert result["new_name"] == "RenamedSheet"


@pytest.mark.smoke
def test_smoke_p1_delete_sheet(tmp_path, monkeypatch):
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    path = str(tmp_path / "del.xlsx")
    _wb = openpyxl.Workbook()
    _wb.active.title = "Keep"
    _wb.create_sheet("Remove")
    _wb.save(path)
    result = wb.delete_sheet(path, "Remove", confirm=True)
    assert result["ok"] is True


@pytest.mark.smoke
def test_smoke_p1_copy_sheet(sample_workbook, monkeypatch):
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    result = wb.copy_sheet(sample_workbook, "Sheet1", "Sheet1_Copy", confirm=True)
    assert result["ok"] is True
    assert result["new_name"] == "Sheet1_Copy"


# ---------------------------------------------------------------------------
# Row / column operations
# ---------------------------------------------------------------------------


@pytest.mark.smoke
def test_smoke_p1_delete_rows(sample_workbook, monkeypatch):
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    result = wb.delete_rows(sample_workbook, "Sheet1", 2, confirm=True)
    assert result["ok"] is True


@pytest.mark.smoke
def test_smoke_p1_delete_cols(sample_workbook, monkeypatch):
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    result = wb.delete_cols(sample_workbook, "Sheet1", 1, confirm=True)
    assert result["ok"] is True


@pytest.mark.smoke
def test_smoke_p1_insert_rows(sample_workbook, monkeypatch):
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    result = wb.insert_rows(sample_workbook, "Sheet1", 1, count=2, confirm=True)
    assert result["ok"] is True


@pytest.mark.smoke
def test_smoke_p1_insert_cols(sample_workbook, monkeypatch):
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    result = wb.insert_cols(sample_workbook, "Sheet1", 1, count=1, confirm=True)
    assert result["ok"] is True


# ---------------------------------------------------------------------------
# Import / save / create
# ---------------------------------------------------------------------------


@pytest.mark.smoke
def test_smoke_p1_import_csv_to_sheet(tmp_path, monkeypatch):
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    csv_path = tmp_path / "data.csv"
    csv_path.write_text("1,foo\n2,bar\n")
    path = str(tmp_path / "csv.xlsx")
    _wb = openpyxl.Workbook()
    _wb.active.title = "Data"
    _wb.save(path)

    result = wb.import_csv_to_sheet(path, str(csv_path), "Data", overwrite=True, confirm=True)
    assert result["ok"] is True
    assert result["rows_imported"] == 2


@pytest.mark.smoke
def test_smoke_p1_save(sample_workbook, monkeypatch):
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    result = wb.save(sample_workbook, confirm=True)
    assert result["ok"] is True
    assert result["path"] == sample_workbook


@pytest.mark.smoke
def test_smoke_p1_create_workbook(tmp_path, monkeypatch):
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    path = str(tmp_path / "new.xlsx")
    result = wb.create_workbook(path, confirm=True)
    assert result["ok"] is True
    assert result["created"] is True
    assert result["sheets"] == ["Sheet1"]


# ---------------------------------------------------------------------------
# Formatting tools (Phase 2.0)
# ---------------------------------------------------------------------------


@pytest.mark.smoke
def test_smoke_p1_apply_number_format(sample_workbook, monkeypatch):
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    result = wb.apply_number_format(sample_workbook, "Sheet1", "A2", "#,##0.00", confirm=True)
    assert result["ok"] is True
    assert result["address"] == "A2"


@pytest.mark.smoke
def test_smoke_p1_apply_cell_style(sample_workbook, monkeypatch):
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    result = wb.apply_cell_style(sample_workbook, "Sheet1", "A1", bold=True, confirm=True)
    assert result["ok"] is True


@pytest.mark.smoke
def test_smoke_p1_apply_alignment(sample_workbook, monkeypatch):
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    result = wb.apply_alignment(
        sample_workbook, "Sheet1", "A1", horizontal="center", confirm=True
    )
    assert result["ok"] is True


@pytest.mark.smoke
def test_smoke_p1_set_column_width(sample_workbook, monkeypatch):
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    result = wb.set_column_width(sample_workbook, "Sheet1", "A", 20.0, confirm=True)
    assert result["ok"] is True
    assert result["column"] == "A"


@pytest.mark.smoke
def test_smoke_p1_set_row_height(sample_workbook, monkeypatch):
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    result = wb.set_row_height(sample_workbook, "Sheet1", 1, 30.0, confirm=True)
    assert result["ok"] is True
    assert result["row"] == 1


@pytest.mark.smoke
def test_smoke_p1_add_border(sample_workbook, monkeypatch):
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    result = wb.add_border(sample_workbook, "Sheet1", "A1:B2", confirm=True)
    assert result["ok"] is True


# ---------------------------------------------------------------------------
# Chart tools (Phase 2.0)
# ---------------------------------------------------------------------------


@pytest.mark.smoke
def test_smoke_p1_create_chart(tmp_path, monkeypatch):
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    path = str(tmp_path / "chart.xlsx")
    _wb = openpyxl.Workbook()
    ws = _wb.active
    ws.title = "Sheet1"
    for i, v in enumerate([10, 20, 30], 1):
        ws.cell(row=i, column=1, value=v)
    _wb.save(path)

    result = wb.create_chart(path, "Sheet1", "bar", "A1:A3", confirm=True)
    assert result["ok"] is True


@pytest.mark.smoke
def test_smoke_p1_list_charts(tmp_path, monkeypatch):
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    path = str(tmp_path / "lc.xlsx")
    _wb = openpyxl.Workbook()
    ws = _wb.active
    ws.title = "Sheet1"
    for i, v in enumerate([10, 20, 30], 1):
        ws.cell(row=i, column=1, value=v)
    _wb.save(path)
    wb.create_chart(path, "Sheet1", "bar", "A1:A3", confirm=True)

    result = wb.list_charts(path, "Sheet1")
    assert result["count"] >= 1


@pytest.mark.smoke
def test_smoke_p1_delete_chart(tmp_path, monkeypatch):
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    path = str(tmp_path / "dc.xlsx")
    _wb = openpyxl.Workbook()
    ws = _wb.active
    ws.title = "Sheet1"
    for i, v in enumerate([10, 20, 30], 1):
        ws.cell(row=i, column=1, value=v)
    _wb.save(path)
    wb.create_chart(path, "Sheet1", "bar", "A1:A3", confirm=True)

    result = wb.delete_chart(path, "Sheet1", 0, confirm=True)
    assert result["ok"] is True


# ---------------------------------------------------------------------------
# NC-001: pivot_table exception leak — structured error return
# ---------------------------------------------------------------------------


@pytest.mark.unit
def test_pivot_table_read_exception_raises_tool_error():
    """NC-001: pivot_table raises ToolError for unexpected exceptions.
    Patches _read_pivot_table to raise generic Exception."""
    from unittest.mock import patch
    from excelmcp.server_com import pivot_table
    from fastmcp.exceptions import ToolError

    with patch(
        "excelmcp.server_com_pivot._read_pivot_table",
        side_effect=Exception("openpyxl internal error"),
    ):
        with pytest.raises(ToolError, match="Pivot table read failed"):
            pivot_table(
                operation="read",
                path="C:/fake/ignored.xlsx",
                sheet="Sheet1",
                pivot_name="PivotTable1",
            )



# ---------------------------------------------------------------------------
# Phase 3 split: module-load sanity checks
# ---------------------------------------------------------------------------

def test_server_module_loads_cleanly_after_prompt_split() -> None:
    """server.py must import cleanly after removing all @mcp.prompt() decorators."""
    import importlib
    # Force a fresh resolve (module may already be cached, that is fine)
    mod = importlib.import_module("excelmcp.server")
    # Must NOT have any @mcp.prompt()-registered names as direct attributes
    # (they now live in server_prompts_* sub-modules)
    import inspect
    src = inspect.getsource(mod)
    assert "@mcp.prompt()" not in src, (
        "server.py must not contain @mcp.prompt() decorators after Phase 3 split"
    )
    # Must still have main() and the 6 resource functions
    assert hasattr(mod, "main"), "server.py must still define main()"
    assert hasattr(mod, "resource_sheets"), "server.py must still define resource_sheets"


def test_server_com_module_loads_cleanly_after_split() -> None:
    """server_com.py thin dispatcher must import and expose all 12 public functions."""
    import importlib
    com = importlib.import_module("excelmcp.server_com")
    expected = [
        "export_as_pdf",
        "export_range_as_csv",
        "export_sheet_as_pdf",
        "export_workbook_as_pdf",
        "recalculate_workbook",
        "hydrate_template",
        "pivot_table",
        "open_template_and_recalculate",
        "evaluate_formula_live",
        "create_pivot_table",
        "refresh_pivot_tables",
        "read_pivot_table",
    ]
    for name in expected:
        assert hasattr(com, name), (
            "server_com thin dispatcher missing re-export: %s" % name
        )
    # W5 ADJUSTMENT-1: exception classes must be re-exported
    for exc in ("NotAllowedError", "OfficeCOMError", "ValidationError"):
        assert hasattr(com, exc), (
            "server_com must re-export exception class: %s" % exc
        )
