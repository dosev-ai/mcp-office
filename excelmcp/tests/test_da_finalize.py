import pytest
import openpyxl
from pathlib import Path
from excelmcp.workbook_openpyxl import (
    _wb_cache,
    fill_formula_range,
    save,
    write_cell,
    add_sheet,
)


class TestDynamicArrayMetadata:
    """DA finalize-path regression tests (methods 13-22).

    Continuation of TestDynamicArrayMetadata from test_da_metadata.py.
    Covers patch-when-cm-already-set, legacy-attr stripping, the finalize
    pipeline, save/evict, flush-if-dirty, and the round-trip survival test.
    No COM required.
    """

    @pytest.mark.unit
    def test_patch_updates_formula_body_when_cm_already_set(self, tmp_path):
        """patch_dynamic_array_metadata updates unprefixed formulas even when cm="1" is already set."""
        import zipfile
        from excelmcp._ooxml import patch_dynamic_array_metadata

        sheet_xml = (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
            '<sheetData>'
            '<row r="7"><c r="B7" s="0" cm="1">'
            '<f t="array" aca="1" ref="B7" ca="1">'
            'IFERROR(LET(x,1,FILTER(A1:A5,A1:A5&gt;"")),"no data")'
            '</f>'
            '</c></row>'
            '</sheetData>'
            '</worksheet>'
        )
        wb_xml = (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"'
            ' xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
            '<sheets><sheet name="Sheet1" sheetId="1" r:id="rId1"/></sheets>'
            '</workbook>'
        )
        wb_rels_xml = (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            '<Relationship Id="rId1"'
            ' Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet"'
            ' Target="worksheets/sheet1.xml"/>'
            '</Relationships>'
        )
        ct_xml = (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
            '<Default Extension="rels"'
            ' ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
            '<Default Extension="xml" ContentType="application/xml"/>'
            '<Override PartName="/xl/workbook.xml"'
            ' ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>'
            '<Override PartName="/xl/worksheets/sheet1.xml"'
            ' ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>'
            '</Types>'
        )

        xlsx_path = tmp_path / "test_cm_already_set.xlsx"
        with zipfile.ZipFile(str(xlsx_path), "w") as zf:
            zf.writestr("[Content_Types].xml", ct_xml)
            zf.writestr("xl/workbook.xml", wb_xml)
            zf.writestr("xl/_rels/workbook.xml.rels", wb_rels_xml)
            zf.writestr("xl/worksheets/sheet1.xml", sheet_xml)

        patch_dynamic_array_metadata(str(xlsx_path))

        with zipfile.ZipFile(str(xlsx_path), "r") as zf:
            patched_sheet = zf.read("xl/worksheets/sheet1.xml").decode("utf-8")

        assert "LET(" in patched_sheet, f"LET formula not found. Sheet XML:\n{patched_sheet}"
        assert "_xlfn.FILTER(" in patched_sheet, f"FILTER not prefixed. Sheet XML:\n{patched_sheet}"
        import re as _re
        unprefixed_matches = _re.findall(r'(?<![._])\b(FILTER)\s*\(', patched_sheet)
        assert not unprefixed_matches, f"Unprefixed functions remain: {unprefixed_matches}"

    @pytest.mark.unit
    def test_patch_strips_legacy_array_attrs_when_cm_missing(self, tmp_path):
        """Regression: DA formulas add cm="1" when absent; Formula2 (aca="1") attrs preserved.

        Post-Formula2-preservation: cells with aca="1" are NOT stripped by
        _strip_legacy_f_attrs — the snapshot/restore guard preserves them.
        cm="1" and _xlfn. prefix are still added by the patcher as expected.
        """
        import zipfile
        from excelmcp._ooxml import patch_dynamic_array_metadata

        sheet_xml = (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
            '<sheetData><row r="2"><c r="B2">'
            '<f t="array" aca="1" ref="B2" ca="1">FILTER(A1:A5,A1:A5&gt;0)</f>'
            '</c></row></sheetData></worksheet>'
        )
        wb_xml = (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"'
            ' xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
            '<sheets><sheet name="Sheet1" sheetId="1" r:id="rId1"/></sheets>'
            '</workbook>'
        )
        wb_rels_xml = (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            '<Relationship Id="rId1"'
            ' Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet"'
            ' Target="worksheets/sheet1.xml"/>'
            '</Relationships>'
        )
        ct_xml = (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
            '<Default Extension="rels"'
            ' ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
            '<Default Extension="xml" ContentType="application/xml"/>'
            '<Override PartName="/xl/workbook.xml"'
            ' ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>'
            '<Override PartName="/xl/worksheets/sheet1.xml"'
            ' ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>'
            '</Types>'
        )

        xlsx_path = tmp_path / "test_strip_legacy_no_cm.xlsx"
        with zipfile.ZipFile(str(xlsx_path), "w") as zf:
            zf.writestr("[Content_Types].xml", ct_xml)
            zf.writestr("xl/workbook.xml", wb_xml)
            zf.writestr("xl/_rels/workbook.xml.rels", wb_rels_xml)
            zf.writestr("xl/worksheets/sheet1.xml", sheet_xml)

        patch_dynamic_array_metadata(str(xlsx_path))

        with zipfile.ZipFile(str(xlsx_path), "r") as zf:
            patched_sheet = zf.read("xl/worksheets/sheet1.xml").decode("utf-8")

        assert 'cm="1"' in patched_sheet
        assert '_xlfn.FILTER(' in patched_sheet
        # Formula2 attrs preserved by the aca="1" guard in _strip_legacy_f_attrs
        assert 'aca="1"' in patched_sheet
        assert 't="array"' in patched_sheet
        assert 'ref="B2"' in patched_sheet

    # ── Scenario A: _save_and_evict ──────────────────────────────────────

    @pytest.mark.unit
    def test_finalize_workbook_da_artifact_uses_deterministic_stage_order(
        self, tmp_path, monkeypatch
    ):
        """Finalize pipeline runs patch->classify->validate->probe in order."""
        from excelmcp._da_finalize import finalize_workbook_da_artifact
        from excelmcp import _da_finalize

        wb_path = tmp_path / "finalize_order.xlsx"
        wb = openpyxl.Workbook()
        wb.save(str(wb_path))

        calls: list[str] = []

        def _fake_patch(path: str) -> int:
            calls.append("patch")
            assert path == str(wb_path)
            return 1

        def _fake_classify(path: str, patched_cells: int) -> str:
            calls.append("classify")
            assert path == str(wb_path)
            assert patched_cells == 1
            return "patched"

        def _fake_validate(path: str) -> dict:
            calls.append("validate")
            assert path == str(wb_path)
            return {"ok": True, "failures": []}

        monkeypatch.setattr(
            _da_finalize._ooxml,
            "patch_dynamic_array_metadata",
            _fake_patch,
        )
        monkeypatch.setattr(
            _da_finalize._ooxml,
            "classify_patch_outcome",
            _fake_classify,
            raising=False,
        )
        monkeypatch.setattr(
            _da_finalize._ooxml,
            "validate_dynamic_array_invariants",
            _fake_validate,
            raising=False,
        )

        result = finalize_workbook_da_artifact(str(wb_path))

        assert calls == ["patch", "classify", "validate"]
        assert [stage["name"] for stage in result["stages"]] == [
            "patch",
            "classify",
            "validate",
            "probe",
        ]
        assert result["stages"][-1]["status"] == "skipped"
        assert result["ok"] is True

    @pytest.mark.unit
    def test_finalize_strict_mode_raises_when_stage_fails(self, tmp_path, monkeypatch):
        """Strict finalize mode must raise when any stage reports failure."""
        from excelmcp._da_finalize import DAFinalizeError, finalize_workbook_da_artifact
        from excelmcp import _da_finalize

        monkeypatch.setenv("EXCEL_DA_FINALIZE_MODE", "strict")

        wb_path = tmp_path / "finalize_strict_fail.xlsx"
        wb = openpyxl.Workbook()
        wb.save(str(wb_path))

        monkeypatch.setattr(
            _da_finalize._ooxml,
            "patch_dynamic_array_metadata",
            lambda _path: 0,
        )
        monkeypatch.setattr(
            _da_finalize._ooxml,
            "validate_dynamic_array_invariants",
            lambda _path: {"ok": False, "failures": ["bad_invariant"]},
            raising=False,
        )

        with pytest.raises(DAFinalizeError, match="DA finalization failed"):
            finalize_workbook_da_artifact(str(wb_path))

    @pytest.mark.unit
    def test_finalize_required_probe_not_implemented_reports_failure(
        self, tmp_path, monkeypatch
    ):
        """Required probe without implementation must fail with probe-stage diagnostics."""
        from excelmcp._da_finalize import finalize_workbook_da_artifact
        from excelmcp import _da_finalize

        monkeypatch.setenv("EXCEL_DA_FINALIZE_MODE", "warn")
        monkeypatch.setenv("EXCEL_DA_COM_PROBE", "required")

        wb_path = tmp_path / "finalize_probe_required_missing.xlsx"
        wb = openpyxl.Workbook()
        wb.save(str(wb_path))

        monkeypatch.setattr(
            _da_finalize._ooxml,
            "patch_dynamic_array_metadata",
            lambda _path: 0,
        )
        monkeypatch.delattr(_da_finalize._da_com, "probe_excel_reopen", raising=False)

        result = finalize_workbook_da_artifact(str(wb_path))
        probe_stage = next(stage for stage in result["stages"] if stage["name"] == "probe")

        assert result["ok"] is False
        assert result["com_probe_policy"] == "required"
        assert probe_stage["status"] == "error"
        assert probe_stage["reason"] == "required_but_not_implemented"

    @pytest.mark.unit
    def test_finalize_probe_ok_false_marks_probe_stage_error(
        self, tmp_path, monkeypatch
    ):
        """A probe response with ok=False must mark diagnostics as failed."""
        from excelmcp._da_finalize import finalize_workbook_da_artifact
        from excelmcp import _da_finalize

        monkeypatch.setenv("EXCEL_DA_FINALIZE_MODE", "warn")
        monkeypatch.setenv("EXCEL_DA_COM_PROBE", "off")

        wb_path = tmp_path / "finalize_probe_ok_false.xlsx"
        wb = openpyxl.Workbook()
        wb.save(str(wb_path))

        monkeypatch.setattr(
            _da_finalize._ooxml,
            "patch_dynamic_array_metadata",
            lambda _path: 0,
        )
        monkeypatch.setattr(
            _da_finalize._da_com,
            "probe_excel_reopen",
            lambda _path: {"ok": False, "error": "open_failed"},
            raising=False,
        )

        result = finalize_workbook_da_artifact(str(wb_path), require_com_probe=True)
        probe_stage = next(stage for stage in result["stages"] if stage["name"] == "probe")

        assert result["ok"] is False
        assert probe_stage["status"] == "error"
        assert probe_stage["result"]["ok"] is False

    @pytest.mark.unit
    def test_save_and_evict_calls_patch_and_evicts_cache(self, tmp_path, monkeypatch):
        """_save_and_evict must call patch_dynamic_array_metadata and evict
        both cache keys for the resolved path."""
        from unittest.mock import patch as _mock_patch
        from excelmcp._core import _save_and_evict

        wb_path = tmp_path / "save_evict.xlsx"
        _wb = openpyxl.Workbook()
        _wb.active.title = "Sheet1"
        _wb.save(str(wb_path))

        # Pre-populate both cache keys to verify eviction
        _wb_cache[(str(wb_path), False)] = _wb
        _wb_cache[(str(wb_path), True)] = _wb

        call_log: list[str] = []

        def _fake_patch(path_str: str) -> int:
            call_log.append(path_str)
            return 0

        with _mock_patch("excelmcp._ooxml.patch_dynamic_array_metadata", _fake_patch):
            _save_and_evict(_wb, wb_path)

        assert call_log == [str(wb_path)], "patch_dynamic_array_metadata must be called once with the correct path"
        assert (str(wb_path), False) not in _wb_cache, "read-only cache entry must be evicted"
        assert (str(wb_path), True) not in _wb_cache, "write-mode cache entry must be evicted"
        assert wb_path.exists(), "workbook file must be written to disk"

    # ── Scenario B: _flush_if_dirty ──────────────────────────────────────

    @pytest.mark.unit
    def test_flush_if_dirty_calls_patch_and_removes_cache_entry(self, tmp_path, monkeypatch):
        """_flush_if_dirty must save the dirty workbook, call patch, and drop
        the (path, True) cache key. A clean (no write-mode entry) path is a no-op."""
        from unittest.mock import patch as _mock_patch
        from excelmcp._core import _flush_if_dirty

        wb_path = tmp_path / "flush_dirty.xlsx"
        _wb = openpyxl.Workbook()
        _wb.active.title = "Sheet1"
        _wb.save(str(wb_path))

        # Simulate a dirty write-mode workbook in the cache
        _wb.active["A1"] = 999
        _wb_cache[(str(wb_path), True)] = _wb

        call_log: list[str] = []

        def _fake_patch(path_str: str) -> int:
            call_log.append(path_str)
            return 0

        with _mock_patch("excelmcp._ooxml.patch_dynamic_array_metadata", _fake_patch):
            _flush_if_dirty(str(wb_path))

        assert call_log == [str(wb_path)], "patch_dynamic_array_metadata must be called once"
        assert (str(wb_path), True) not in _wb_cache, "write-mode cache entry must be removed after flush"

        # Calling again on a clean path must be a no-op (no patch, no error)
        call_log.clear()
        with _mock_patch("excelmcp._ooxml.patch_dynamic_array_metadata", _fake_patch):
            _flush_if_dirty(str(wb_path))
        assert call_log == [], "second flush on a clean path must not call the patcher"

    @pytest.mark.unit
    def test_flush_if_dirty_routes_through_finalize_boundary(self, tmp_path):
        """Dirty flush must call _finalize_and_evict once for that workbook path."""
        from unittest.mock import patch as _mock_patch
        from excelmcp._core import _flush_if_dirty

        wb_path = tmp_path / "flush_finalize.xlsx"
        wb = openpyxl.Workbook()
        wb.save(str(wb_path))
        _wb_cache[(str(wb_path), True)] = wb

        seen: list[str] = []

        def _fake_finalize(resolved: Path) -> None:
            seen.append(str(resolved))

        with _mock_patch("excelmcp._core._finalize_and_evict", _fake_finalize):
            _flush_if_dirty(str(wb_path))

        assert seen == [str(wb_path)]
        assert (str(wb_path), True) not in _wb_cache

    # ── Scenario C: round-trip DA formula survives structural op ─────────

    @pytest.mark.unit
    def test_da_metadata_survives_after_flush_and_reload(self, tmp_path, monkeypatch):
        """Round-trip: write a DA formula → save (via write_cell+save) →
        simulate cache eviction → reload from disk → verify cm='1' is still
        present in the OOXML (i.e. not destroyed by the second save path)."""
        import zipfile
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))

        wb_path = str(tmp_path / "da_roundtrip.xlsx")
        _wb = openpyxl.Workbook()
        _wb.active.title = "Sheet1"
        _wb.save(wb_path)
        _wb_cache.clear()

        fill_formula_range(wb_path, "Sheet1", "A1", "A1", "=LET(x,42,x)", confirm=True)
        save(wb_path, confirm=True)
        _wb_cache.clear()

        with zipfile.ZipFile(wb_path) as z:
            names_after_save = z.namelist()
            sheet_xml_after_save = z.read("xl/worksheets/sheet1.xml").decode("utf-8")

        assert "xl/metadata.xml" in names_after_save, \
            "metadata.xml must exist after initial save of DA formula"
        assert 'cm="1"' in sheet_xml_after_save, \
            "cm='1' must be present in sheet XML after initial save"

        write_cell(wb_path, "Sheet1", "B1", "static value", confirm=True)
        add_sheet(wb_path, "ExtraSheet", confirm=True)
        _wb_cache.clear()

        with zipfile.ZipFile(wb_path) as z:
            names_after_op = z.namelist()
            sheet_xml_after_op = z.read("xl/worksheets/sheet1.xml").decode("utf-8")

        assert "xl/metadata.xml" in names_after_op, \
            "metadata.xml must still exist after structural operation"
        assert 'cm="1"' in sheet_xml_after_op, \
            "cm='1' must survive after a structural operation flushes the workbook"

    # ── BUG 1 regression: append_rows must call the DA patch ─────────────

    @pytest.mark.unit
    def test_append_rows_triggers_da_patch(self, tmp_path, monkeypatch):
        """append_rows calls patch_dynamic_array_metadata after openpyxl save.

        Regression test for the dynamic-array patch routing bug: append_rows must
        route through _save_and_evict → _finalize_and_evict →
        finalize_workbook_da_artifact → patch_dynamic_array_metadata so that
        dynamic-array formulas already in the workbook keep their cm="1" and
        _xlfn. prefix annotations intact after new rows are appended.
        """
        import zipfile
        from excelmcp.workbook_openpyxl import append_rows

        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")

        # Build a minimal xlsx with an existing FILTER formula (unprefixed, as
        # openpyxl would write it) so the patcher has something to stamp cm="1" on.
        sheet_xml = (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
            '<sheetData>'
            '<row r="1"><c r="A1"><v>1</v></c></row>'
            '<row r="2"><c r="A2"><f>FILTER(A:A,A:A&gt;0)</f></c></row>'
            '</sheetData>'
            '</worksheet>'
        )
        wb_xml = (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"'
            ' xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
            '<sheets><sheet name="Sheet1" sheetId="1" r:id="rId1"/></sheets>'
            '</workbook>'
        )
        wb_rels_xml = (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            '<Relationship Id="rId1"'
            ' Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet"'
            ' Target="worksheets/sheet1.xml"/>'
            '</Relationships>'
        )
        ct_xml = (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
            '<Default Extension="rels"'
            ' ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
            '<Default Extension="xml" ContentType="application/xml"/>'
            '<Override PartName="/xl/workbook.xml"'
            ' ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>'
            '<Override PartName="/xl/worksheets/sheet1.xml"'
            ' ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>'
            '</Types>'
        )

        xlsx_path = tmp_path / "append_rows_da.xlsx"
        with zipfile.ZipFile(str(xlsx_path), "w") as zf:
            zf.writestr("[Content_Types].xml", ct_xml)
            zf.writestr("xl/workbook.xml", wb_xml)
            zf.writestr("xl/_rels/workbook.xml.rels", wb_rels_xml)
            zf.writestr("xl/worksheets/sheet1.xml", sheet_xml)

        # append_rows must trigger the DA patch (not skip it)
        result = append_rows(str(xlsx_path), "Sheet1", [["appended_value"]], confirm=True)
        assert result["ok"] is True

        with zipfile.ZipFile(str(xlsx_path), "r") as zf:
            names_in_zip = zf.namelist()
            patched_sheet = zf.read("xl/worksheets/sheet1.xml").decode("utf-8")

        # DA patch must have stamped cm="1" and added _xlfn. prefix
        assert 'cm="1"' in patched_sheet, \
            f"cm='1' missing after append_rows — DA patch was skipped. sheet XML:\n{patched_sheet}"
        assert "_xlfn.FILTER(" in patched_sheet, \
            f"_xlfn.FILTER( missing — formula body not rewritten. sheet XML:\n{patched_sheet}"
        assert "xl/metadata.xml" in names_in_zip, \
            "xl/metadata.xml missing — metadata.xml not injected by DA patch"

    @pytest.mark.unit
    def test_write_cell_triggers_da_patch(self, tmp_path, monkeypatch):
        """write_cell (and by extension write_range) calls the DA patch.

        All openpyxl write paths route through _save_and_evict which calls
        _finalize_and_evict → patch_dynamic_array_metadata.  This test
        verifies write_cell as the representative path for all scalar writes.
        """
        import zipfile
        from excelmcp.workbook_openpyxl import write_cell

        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")

        # Same starter workbook with an unprefixed FILTER formula
        sheet_xml = (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
            '<sheetData>'
            '<row r="1"><c r="A1"><f>FILTER(C:C,C:C&gt;10)</f></c></row>'
            '</sheetData>'
            '</worksheet>'
        )
        wb_xml = (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"'
            ' xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
            '<sheets><sheet name="Sheet1" sheetId="1" r:id="rId1"/></sheets>'
            '</workbook>'
        )
        wb_rels_xml = (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            '<Relationship Id="rId1"'
            ' Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet"'
            ' Target="worksheets/sheet1.xml"/>'
            '</Relationships>'
        )
        ct_xml = (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
            '<Default Extension="rels"'
            ' ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
            '<Default Extension="xml" ContentType="application/xml"/>'
            '<Override PartName="/xl/workbook.xml"'
            ' ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>'
            '<Override PartName="/xl/worksheets/sheet1.xml"'
            ' ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>'
            '</Types>'
        )

        xlsx_path = tmp_path / "write_cell_da.xlsx"
        with zipfile.ZipFile(str(xlsx_path), "w") as zf:
            zf.writestr("[Content_Types].xml", ct_xml)
            zf.writestr("xl/workbook.xml", wb_xml)
            zf.writestr("xl/_rels/workbook.xml.rels", wb_rels_xml)
            zf.writestr("xl/worksheets/sheet1.xml", sheet_xml)

        result = write_cell(str(xlsx_path), "Sheet1", "B1", "hello", confirm=True)
        assert result["ok"] is True

        with zipfile.ZipFile(str(xlsx_path), "r") as zf:
            patched_sheet = zf.read("xl/worksheets/sheet1.xml").decode("utf-8")

        assert 'cm="1"' in patched_sheet, \
            "cm='1' missing after write_cell — DA patch was not applied"
        assert "_xlfn.FILTER(" in patched_sheet, \
            "_xlfn.FILTER( missing after write_cell — formula body not rewritten"
