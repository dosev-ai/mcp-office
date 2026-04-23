import pytest
import openpyxl

from excelmcp.workbook_openpyxl import (
    set_column_visibility,
    set_workbook_calc_mode,
    set_sheet_properties,
    create_table,
    add_data_validation,
    ExcelMCPError,
    ValidationError,
    NotAllowedError,
)


class TestPhase26Tools:
    """Tests for Phase 2.6 gap-closing tools:
    set_column_visibility, set_workbook_calc_mode, gridlines extension to
    set_sheet_properties, style extension to create_table, and error_title/
    prompt extensions to add_data_validation.
    """

    # ── set_column_visibility ─────────────────────────────────────────────

    @pytest.mark.unit
    def test_set_column_visibility_hide_column_pass(self, tmp_path, monkeypatch):
        """Hiding a column sets column_dimensions.hidden=True."""
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        wb_path = tmp_path / "hide.xlsx"
        _wb = openpyxl.Workbook()
        _wb.active.title = "Sheet1"
        _wb.save(str(wb_path))

        result = set_column_visibility(str(wb_path), "Sheet1", ["B", "C"], hidden=True, confirm=True)
        assert result["hidden"] is True
        assert result["columns"] == ["B", "C"]
        assert result["sheet"] == "Sheet1"

        # Verify the hidden flag was persisted
        _wb2 = openpyxl.load_workbook(str(wb_path))
        ws = _wb2["Sheet1"]
        assert ws.column_dimensions["B"].hidden is True
        assert ws.column_dimensions["C"].hidden is True

    @pytest.mark.unit
    def test_set_column_visibility_show_column_pass(self, tmp_path, monkeypatch):
        """Un-hiding a column sets column_dimensions.hidden=False."""
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        wb_path = tmp_path / "show.xlsx"
        _wb = openpyxl.Workbook()
        ws = _wb.active
        ws.title = "Sheet1"
        ws.column_dimensions["A"].hidden = True
        _wb.save(str(wb_path))

        result = set_column_visibility(str(wb_path), "Sheet1", ["A"], hidden=False, confirm=True)
        assert result["hidden"] is False

        _wb2 = openpyxl.load_workbook(str(wb_path))
        assert _wb2["Sheet1"].column_dimensions["A"].hidden is False

    @pytest.mark.unit
    def test_set_column_visibility_invalid_column_raises(self, tmp_path, monkeypatch):
        """A non-alphabetic column string raises ValidationError."""
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        wb_path = tmp_path / "badcol.xlsx"
        _wb = openpyxl.Workbook()
        _wb.active.title = "Sheet1"
        _wb.save(str(wb_path))

        with pytest.raises(ValidationError, match="Invalid column letter"):
            set_column_visibility(str(wb_path), "Sheet1", ["A1"], hidden=True, confirm=True)

    @pytest.mark.unit
    def test_set_column_visibility_confirm_false_raises(self, tmp_path, monkeypatch):
        """confirm=False raises NotAllowedError."""
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        wb_path = tmp_path / "noconfirm.xlsx"
        _wb = openpyxl.Workbook()
        _wb.active.title = "Sheet1"
        _wb.save(str(wb_path))

        with pytest.raises((NotAllowedError, ExcelMCPError)):
            set_column_visibility(str(wb_path), "Sheet1", ["A"], hidden=True, confirm=False)

    # ── set_sheet_properties — show_gridlines / show_row_col_headers ─────

    @pytest.mark.unit
    def test_set_sheet_props_show_gridlines_false(self, tmp_path, monkeypatch):
        """Setting show_gridlines=False persists in the workbook XML."""
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        wb_path = tmp_path / "gridlines.xlsx"
        _wb = openpyxl.Workbook()
        _wb.active.title = "Sheet1"
        _wb.save(str(wb_path))

        result = set_sheet_properties(str(wb_path), "Sheet1", show_gridlines=False, confirm=True)
        assert result["show_gridlines"] is False

        _wb2 = openpyxl.load_workbook(str(wb_path))
        assert _wb2["Sheet1"].sheet_view.showGridLines is False

    @pytest.mark.unit
    def test_set_sheet_props_show_gridlines_true(self, tmp_path, monkeypatch):
        """Re-enabling gridlines sets showGridLines back to True."""
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        wb_path = tmp_path / "gridlines2.xlsx"
        _wb = openpyxl.Workbook()
        ws = _wb.active
        ws.title = "Sheet1"
        ws.sheet_view.showGridLines = False
        _wb.save(str(wb_path))

        result = set_sheet_properties(str(wb_path), "Sheet1", show_gridlines=True, confirm=True)
        assert result["show_gridlines"] is True

        _wb2 = openpyxl.load_workbook(str(wb_path))
        assert _wb2["Sheet1"].sheet_view.showGridLines is True

    # ── set_workbook_calc_mode ────────────────────────────────────────────

    @pytest.mark.unit
    def test_set_workbook_calc_mode_manual_pass(self, tmp_path, monkeypatch):
        """Setting calc_mode='manual' is persisted to the workbook."""
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        wb_path = tmp_path / "calcmode.xlsx"
        _wb = openpyxl.Workbook()
        _wb.active.title = "Sheet1"
        _wb.save(str(wb_path))

        result = set_workbook_calc_mode(str(wb_path), calc_mode="manual", confirm=True)
        assert result["calc_mode"] == "manual"

    @pytest.mark.unit
    def test_set_workbook_calc_mode_invalid_mode_raises(self, tmp_path, monkeypatch):
        """An unrecognised calc_mode raises ValidationError."""
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        wb_path = tmp_path / "calcmode_bad.xlsx"
        _wb = openpyxl.Workbook()
        _wb.active.title = "Sheet1"
        _wb.save(str(wb_path))

        with pytest.raises(ValidationError, match="calc_mode"):
            set_workbook_calc_mode(str(wb_path), calc_mode="bogus", confirm=True)

    @pytest.mark.unit
    def test_set_workbook_calc_mode_confirm_false_raises(self, tmp_path, monkeypatch):
        """confirm=False raises NotAllowedError."""
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        wb_path = tmp_path / "calcmode_noconfirm.xlsx"
        _wb = openpyxl.Workbook()
        _wb.active.title = "Sheet1"
        _wb.save(str(wb_path))

        with pytest.raises((NotAllowedError, ExcelMCPError)):
            set_workbook_calc_mode(str(wb_path), calc_mode="manual", confirm=False)

    @pytest.mark.unit
    def test_set_workbook_calc_mode_full_calc_on_load_pass(self, tmp_path, monkeypatch):
        """Setting full_calc_on_load=True is reflected in the return dict."""
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        wb_path = tmp_path / "calcmode_full.xlsx"
        _wb = openpyxl.Workbook()
        _wb.active.title = "Sheet1"
        _wb.save(str(wb_path))

        result = set_workbook_calc_mode(str(wb_path), full_calc_on_load=True, confirm=True)
        assert result["full_calc_on_load"] is True

    # ── create_table with custom style ────────────────────────────────────

    @pytest.mark.unit
    def test_create_table_with_custom_style_pass(self, tmp_path, monkeypatch):
        """create_table accepts a custom style and returns it in the result."""
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        wb_path = tmp_path / "styled_table.xlsx"
        _wb = openpyxl.Workbook()
        ws = _wb.active
        ws.title = "Sheet1"
        for r in range(1, 4):
            ws.cell(r, 1).value = f"R{r}C1"
            ws.cell(r, 2).value = r
        _wb.save(str(wb_path))

        result = create_table(
            str(wb_path), "Sheet1", "A1:B3", "StyledTable",
            style="TableStyleLight1", confirm=True
        )
        assert result["style"] == "TableStyleLight1"
        assert result["table_name"] == "StyledTable"

    # ── add_data_validation with error_title / prompt ─────────────────────

    @pytest.mark.unit
    def test_add_data_validation_with_error_title_pass(self, tmp_path, monkeypatch):
        """error_title is stored on the DataValidation object."""
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))

        wb_path = tmp_path / "dv_et.xlsx"
        _wb = openpyxl.Workbook()
        _wb.active.title = "Sheet1"
        _wb.save(str(wb_path))

        result = add_data_validation(
            str(wb_path), "Sheet1", "A1:A5",
            validation_type="list", formula1='"Yes,No"',
            error_title="Bad Input", confirm=True
        )
        assert result["ok"] is True

        _wb2 = openpyxl.load_workbook(str(wb_path))
        dv_list = list(_wb2["Sheet1"].data_validations.dataValidation)
        assert any(getattr(dv, "errorTitle", None) == "Bad Input" for dv in dv_list)

    @pytest.mark.unit
    def test_add_data_validation_with_prompt_pass(self, tmp_path, monkeypatch):
        """prompt and prompt_title are stored on the DataValidation object."""
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))

        wb_path = tmp_path / "dv_prompt.xlsx"
        _wb = openpyxl.Workbook()
        _wb.active.title = "Sheet1"
        _wb.save(str(wb_path))

        result = add_data_validation(
            str(wb_path), "Sheet1", "B1:B5",
            validation_type="whole", formula1="1", formula2="10",
            operator="between",
            prompt="Enter a value between 1 and 10",
            prompt_title="Input Guide",
            confirm=True,
        )
        assert result["ok"] is True

        _wb2 = openpyxl.load_workbook(str(wb_path))
        dv_list = list(_wb2["Sheet1"].data_validations.dataValidation)
        assert any(getattr(dv, "prompt", None) == "Enter a value between 1 and 10" for dv in dv_list)

    @pytest.mark.unit
    def test_add_data_validation_with_show_input_message_false(self, tmp_path, monkeypatch):
        """show_input_message=False is honoured."""
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))

        wb_path = tmp_path / "dv_noinput.xlsx"
        _wb = openpyxl.Workbook()
        _wb.active.title = "Sheet1"
        _wb.save(str(wb_path))

        result = add_data_validation(
            str(wb_path), "Sheet1", "C1:C5",
            validation_type="list", formula1='"A,B,C"',
            show_input_message=False,
            confirm=True,
        )
        assert result["ok"] is True

        _wb2 = openpyxl.load_workbook(str(wb_path))
        dv_list = list(_wb2["Sheet1"].data_validations.dataValidation)
        assert any(not bool(getattr(dv, "showInputMessage", True)) for dv in dv_list)

    # ── set_sheet_properties — show_row_col_headers (blocking gap) ────────

    @pytest.mark.unit
    def test_set_sheet_props_show_row_col_headers_false(self, tmp_path, monkeypatch):
        """Setting show_row_col_headers=False persists in the workbook XML."""
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        wb_path = tmp_path / "rowcolheaders.xlsx"
        _wb = openpyxl.Workbook()
        _wb.active.title = "Sheet1"
        _wb.save(str(wb_path))

        result = set_sheet_properties(
            str(wb_path), "Sheet1", show_row_col_headers=False, confirm=True
        )
        assert result["ok"] is True
        assert result["show_row_col_headers"] is False

        _wb2 = openpyxl.load_workbook(str(wb_path))
        assert _wb2["Sheet1"].sheet_view.showRowColHeaders is False

    @pytest.mark.unit
    def test_set_sheet_props_show_row_col_headers_true(self, tmp_path, monkeypatch):
        """Re-enabling row/col headers sets showRowColHeaders back to True."""
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        wb_path = tmp_path / "rowcolheaders2.xlsx"
        _wb = openpyxl.Workbook()
        ws = _wb.active
        ws.title = "Sheet1"
        ws.sheet_view.showRowColHeaders = False
        _wb.save(str(wb_path))

        result = set_sheet_properties(
            str(wb_path), "Sheet1", show_row_col_headers=True, confirm=True
        )
        assert result["show_row_col_headers"] is True

        _wb2 = openpyxl.load_workbook(str(wb_path))
        assert _wb2["Sheet1"].sheet_view.showRowColHeaders is True

    @pytest.mark.unit
    def test_set_sheet_props_none_params_are_no_op(self, tmp_path, monkeypatch):
        """Passing None for show_gridlines and show_row_col_headers leaves them unchanged."""
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        wb_path = tmp_path / "noop.xlsx"
        _wb = openpyxl.Workbook()
        ws = _wb.active
        ws.title = "Sheet1"
        ws.sheet_view.showGridLines = False
        ws.sheet_view.showRowColHeaders = False
        _wb.save(str(wb_path))

        result = set_sheet_properties(
            str(wb_path), "Sheet1", show_gridlines=None, show_row_col_headers=None, confirm=True
        )
        assert result["show_gridlines"] is None
        assert result["show_row_col_headers"] is None

        # Values must be unchanged in the persisted file
        _wb2 = openpyxl.load_workbook(str(wb_path))
        assert _wb2["Sheet1"].sheet_view.showGridLines is False
        assert _wb2["Sheet1"].sheet_view.showRowColHeaders is False
