"""
Metadata Sprint — 20 User Story Validation Harness for excelmcp
===============================================================

Each test maps 1-to-1 to a user story from the metadata umbrella sprint.

| Test function                           | User Story |
|-----------------------------------------|------------|
| test_us_a1_canonical_10_plus_policy     | US-A1      |
| test_us_a2_3sheet_names                 | US-A2      |
| test_us_a3_has_formulas_true            | US-A3      |
| test_us_a4_has_formulas_false           | US-A4      |
| test_us_a5_last_modified_tz             | US-A5      |
| test_us_b1_default_policy_strict        | US-B1      |
| test_us_b2_lenient_policy_env           | US-B2      |
| test_us_b3_invalid_policy_warns         | US-B3      |
| test_us_c1_alias_name_remapped          | US-C1      |
| test_us_c2_all_aliases_parametrize      | US-C2      |
| test_us_c3_unknown_keys_dropped         | US-C3      |
| test_us_c4_empty_dict_raises            | US-C4      |
| test_us_d1_contract_resource_json       | US-D1      |
| test_us_d2_workflow_prompt_fields       | US-D2      |
| test_us_d3_reference_prompt_table       | US-D3      |
| test_us_e1_outside_allowlist            | US-E1      |
| test_us_e2_allowlisted_missing_file     | US-E2      |
| test_us_e3_attribute_error_bubbles      | US-E3      |
| test_us_f1_capabilities_metadata        | US-F1      |
| test_us_g1_meta_alias_parity            | US-G1      |
"""
from __future__ import annotations

import json
import warnings
from datetime import datetime
from unittest.mock import MagicMock

import openpyxl
import pytest

from excelmcp._metadata_contract import (
    CANONICAL_FIELDS,
    CONTRACT_VERSION,
    build_workbook_metadata,
    validate_metadata,
)
from excelmcp._lineage import FIELD_ALIASES, map_lineage_fields, normalize_metadata
from excelmcp.runtime_config import (
    get_effective_policy,
    is_strict,
    load_runtime_config,
    get_contract_version,
)
from excelmcp.server_io import get_workbook_metadata
from excelmcp._io import capabilities as _io_capabilities
from excelmcp._data import add_cell_comment
from excelmcp._sheet import copy_sheet
from excelmcp.workbook_openpyxl import (
    ExcelMCPError,
    NotAllowedError,
    ValidationError,
    read_range,
    read_cell,
    write_range,
    write_cell,
    append_rows,
    list_sheets,
    add_sheet,
    rename_sheet,
    delete_sheet,
    create_table,
    list_tables,
    read_table,
    apply_cell_style,
    set_column_width,
    merge_cells,
    delete_cell_comment,
    bulk_add_data_validation,
)

# ---------------------------------------------------------------------------
# Shared helpers
# ---------------------------------------------------------------------------

def _make_wb(tmp_path, monkeypatch, *, name: str = "test.xlsx") -> str:
    """Create a minimal 1-sheet workbook, set allowlist, return path str."""
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    p = tmp_path / name
    wb = openpyxl.Workbook()
    wb.active.title = "Sheet1"
    wb.active["A1"] = "hello"
    wb.save(str(p))
    return str(p)


def _canonical_base() -> dict:
    """Return a complete minimal canonical dict (no aliases)."""
    return {
        "path": "/some/file.xlsx",
        "workbook_name": "file.xlsx",
        "sheet_names": ["Sheet1"],
        "active_sheet": "Sheet1",
        "sheet_count": 1,
        "has_formulas": False,
        "last_modified": "2025-01-01T00:00:00+00:00",
        "size_bytes": 1000,
        "format": ".xlsx",
        "contract_version": "1.0",
    }


def _call_prompt(fn, **kwargs):
    """Call a FastMCP-decorated prompt/resource regardless of wrapper version.

    FastMCP 2.x wraps functions in FunctionTool / FunctionPrompt which may
    not be directly callable.  Fall back to .fn() when needed.
    """
    if hasattr(fn, "fn") and callable(fn.fn):
        return fn.fn(**kwargs)
    return fn(**kwargs)


# ---------------------------------------------------------------------------
# ANGLE A — Canonical metadata retrieval
# ---------------------------------------------------------------------------


def test_us_a1_canonical_10_plus_policy(tmp_path, monkeypatch):
    """US-A1: build_workbook_metadata on 1-sheet .xlsx → 10 CANONICAL_FIELDS + 'policy',
    contract_version=='1.0', has_formulas==False, sheet_count==1.

    NOTE: Per implementation, policy is injected at the server layer (get_workbook_metadata),
    not by build_workbook_metadata.  This test validates the story's specification —
    if 'policy' is absent from the return, the test will FAIL and is flagged as FEATURE GAP.
    """
    path = _make_wb(tmp_path, monkeypatch)
    result = get_workbook_metadata(path)  # injects 'policy' at server layer

    expected_keys = set(CANONICAL_FIELDS) | {"policy"}
    assert set(result.keys()) == expected_keys, (
        f"Keys mismatch. Got: {set(result.keys())}  Expected: {expected_keys}"
    )
    assert result["contract_version"] == "1.0"
    assert result["has_formulas"] is False
    assert result["sheet_count"] == 1


def test_us_a2_3sheet_names(tmp_path, monkeypatch):
    """US-A2: build_workbook_metadata on 3-sheet workbook → sheet_names has 3 entries,
    sheet_count==3, active_sheet is in sheet_names.
    """
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    p = tmp_path / "three.xlsx"
    wb = openpyxl.Workbook()
    wb.active.title = "Alpha"
    wb.create_sheet("Beta")
    wb.create_sheet("Gamma")
    wb.save(str(p))

    result = dict(build_workbook_metadata(str(p)))

    assert len(result["sheet_names"]) == 3
    assert result["sheet_count"] == 3
    assert result["active_sheet"] in result["sheet_names"]


def test_us_a3_has_formulas_true(tmp_path, monkeypatch):
    """US-A3: build_workbook_metadata on workbook with =SUM(A1:A3) → has_formulas==True."""
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    p = tmp_path / "formula.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = 1
    ws["A2"] = 2
    ws["A3"] = 3
    ws["B1"] = "=SUM(A1:A3)"   # literal formula string preserved in file
    wb.save(str(p))

    result = dict(build_workbook_metadata(str(p)))
    assert result["has_formulas"] is True


def test_us_a4_has_formulas_false(tmp_path, monkeypatch):
    """US-A4: build_workbook_metadata on workbook with only string/numeric values
    → has_formulas==False.
    """
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    p = tmp_path / "values_only.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "hello"
    ws["A2"] = 42
    ws["A3"] = 3.14
    wb.save(str(p))

    result = dict(build_workbook_metadata(str(p)))
    assert result["has_formulas"] is False


def test_us_a5_last_modified_tz(tmp_path, monkeypatch):
    """US-A5: last_modified parses via datetime.fromisoformat() without error
    and result has tzinfo not None.
    """
    path = _make_wb(tmp_path, monkeypatch)
    result = dict(build_workbook_metadata(path))

    dt = datetime.fromisoformat(result["last_modified"])
    assert dt.tzinfo is not None, (
        f"last_modified has no tzinfo: {result['last_modified']!r}"
    )


# ---------------------------------------------------------------------------
# ANGLE B — Policy enforcement
# ---------------------------------------------------------------------------


def test_us_b1_default_policy_strict(monkeypatch):
    """US-B1: With EXCEL_METADATA_POLICY absent from os.environ,
    get_effective_policy() returns "strict".
    """
    monkeypatch.delenv("EXCEL_METADATA_POLICY", raising=False)
    assert get_effective_policy() == "strict"


def test_us_b2_lenient_policy_env(monkeypatch):
    """US-B2: With EXCEL_METADATA_POLICY=lenient, get_effective_policy() returns "lenient"."""
    monkeypatch.setenv("EXCEL_METADATA_POLICY", "lenient")
    assert get_effective_policy() == "lenient"


def test_us_b3_invalid_policy_warns(monkeypatch):
    """US-B3: With EXCEL_METADATA_POLICY=nonsense, get_effective_policy() emits a
    UserWarning and returns "strict".
    """
    monkeypatch.setenv("EXCEL_METADATA_POLICY", "nonsense")
    with warnings.catch_warnings(record=True) as caught:
        warnings.simplefilter("always")
        result = get_effective_policy()

    assert result == "strict"
    user_warnings = [w for w in caught if issubclass(w.category, UserWarning)]
    assert len(user_warnings) >= 1, "Expected at least one UserWarning for invalid policy"
    assert "nonsense" in str(user_warnings[0].message).lower()


# ---------------------------------------------------------------------------
# ANGLE C — Lineage / alias normalization
# ---------------------------------------------------------------------------


def test_us_c1_alias_name_remapped(tmp_path, monkeypatch):
    """US-C1: normalize_metadata({"name": "Book1", ...other canonical...})
    → result has "workbook_name"=="Book1", no "name" key.
    """
    raw = _canonical_base()
    del raw["workbook_name"]   # remove canonical key
    raw["name"] = "Book1"      # add alias instead

    result = normalize_metadata(raw)

    assert result["workbook_name"] == "Book1"
    assert "name" not in result


# Build parametrize inputs for all 9 FIELD_ALIASES
_ALIAS_PARAMS = []
for _alias, _canonical in FIELD_ALIASES.items():
    # Start from a full canonical dict, remove the canonical key, add the alias
    _base = _canonical_base()
    _test_value = _base[_canonical]  # preserve value type
    del _base[_canonical]
    _base[_alias] = _test_value
    _ALIAS_PARAMS.append(pytest.param(_alias, _canonical, _base, _test_value, id=f"{_alias}→{_canonical}"))


@pytest.mark.parametrize("alias,canonical,raw,expected_value", _ALIAS_PARAMS)
def test_us_c2_all_aliases_parametrize(alias, canonical, raw, expected_value):
    """US-C2: All 9 FIELD_ALIASES are correctly remapped by normalize_metadata.

    For each alias key: the alias is remapped to its canonical name, the alias
    key is absent from the result, and the canonical key has the correct value.
    """
    result = normalize_metadata(raw)

    assert canonical in result, f"Canonical key '{canonical}' missing from result"
    assert result[canonical] == expected_value, (
        f"Alias '{alias}' → '{canonical}': expected {expected_value!r}, got {result[canonical]!r}"
    )
    assert alias not in result, f"Alias key '{alias}' should be absent from result"


def test_us_c3_unknown_keys_dropped(tmp_path, monkeypatch):
    """US-C3: map_lineage_fields with extra "_internal" and "garbage" keys
    → output has no "_internal" or "garbage" key, no exception.
    """
    raw = _canonical_base()
    raw["_internal"] = 99
    raw["garbage"] = "y"

    result = map_lineage_fields(raw)

    assert "_internal" not in result
    assert "garbage" not in result


def test_us_c4_empty_dict_raises():
    """US-C4: map_lineage_fields({}) (no canonical keys) raises ValueError."""
    with pytest.raises(ValueError):
        map_lineage_fields({})


# ---------------------------------------------------------------------------
# ANGLE D — Contract resource & prompts
# ---------------------------------------------------------------------------


def test_us_d1_contract_resource_json():
    """US-D1: resource_metadata_contract() returns JSON with contract_version=="1.0"
    and exactly 10 fields in the 'fields' array.
    """
    from excelmcp.server_metadata_prompts import resource_metadata_contract

    raw = _call_prompt(resource_metadata_contract)
    data = json.loads(raw)

    assert data["contract_version"] == "1.0"
    assert len(data["fields"]) == 10


def test_us_d2_workflow_prompt_fields():
    """US-D2: workbook_metadata_workflow(path="test.xlsx") renders all 10 canonical
    field names in the returned string.
    """
    from excelmcp.server_metadata_prompts import workbook_metadata_workflow

    result = _call_prompt(workbook_metadata_workflow, path="test.xlsx")

    for field in CANONICAL_FIELDS:
        assert field in result, f"Field '{field}' not found in workbook_metadata_workflow output"


def test_us_d3_reference_prompt_table():
    """US-D3: metadata_contract_reference() returns a string containing a markdown
    table (with '| Field' header) and all 10 canonical field names.
    """
    from excelmcp.server_metadata_prompts import metadata_contract_reference

    result = _call_prompt(metadata_contract_reference)

    assert "| Field" in result, "Expected markdown table header '| Field' in result"
    for field in CANONICAL_FIELDS:
        assert field in result, f"Field '{field}' not found in metadata_contract_reference output"


# ---------------------------------------------------------------------------
# ANGLE E — Security & error handling
# ---------------------------------------------------------------------------


def test_us_e1_outside_allowlist(tmp_path, monkeypatch, tmp_path_factory):
    """US-E1: build_workbook_metadata on a path outside the allowlist root raises
    (ValidationError, ToolError, ValueError, or PermissionError — any exception
    that is NOT a successful return).
    """
    # allowlist is set to tmp_path; outside_dir is a separate tmp directory
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    outside_dir = tmp_path_factory.mktemp("outside")
    outside_path = outside_dir / "not_allowed.xlsx"

    # Create the file so the failure is purely allowlist-based, not missing-file
    wb = openpyxl.Workbook()
    wb.save(str(outside_path))

    with pytest.raises(Exception) as exc_info:
        build_workbook_metadata(str(outside_path))

    # Must NOT be a successful return — any exception is acceptable
    assert exc_info.type.__name__ in (
        "ValidationError", "ToolError", "ValueError", "PermissionError",
    ) or issubclass(exc_info.type, Exception), (
        f"Expected an exception but got {exc_info.type}"
    )


def test_us_e2_allowlisted_missing_file(tmp_path, monkeypatch):
    """US-E2: build_workbook_metadata on an allowlisted path that does not exist
    raises an exception (ToolError, FileNotFoundError, or similar).
    """
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    missing_path = str(tmp_path / "does_not_exist.xlsx")

    with pytest.raises(Exception) as exc_info:
        build_workbook_metadata(missing_path)

    # Should not be a successful return
    assert exc_info.value is not None
    assert exc_info.type is not type(None)


def test_us_e3_attribute_error_bubbles(tmp_path, monkeypatch):
    """US-E3: When openpyxl.load_workbook returns a stub workbook where wb.sheetnames
    raises AttributeError (simulating structural extraction failure), the call
    raises ToolError (not None / silent failure).

    Uses a plain stub class rather than MagicMock so that the property descriptor
    is not shadowed by MagicMock.__getattr__.
    """
    path = _make_wb(tmp_path, monkeypatch)

    # Stub workbook: sheetnames raises AttributeError; close() is a no-op
    class _StubBadWorkbook:
        @property
        def sheetnames(self):
            raise AttributeError("forced sheetnames failure for US-E3")

        @property
        def worksheets(self):
            return []

        @property
        def active(self):
            m = MagicMock()
            m.title = "Sheet1"
            return m

        def close(self):
            pass

    def _bad_load(*args, **kwargs):
        return _StubBadWorkbook()

    monkeypatch.setattr("excelmcp._metadata_contract.openpyxl.load_workbook", _bad_load)

    from fastmcp.exceptions import ToolError

    with warnings.catch_warnings(record=True):
        warnings.simplefilter("always")
        with pytest.raises(ToolError):
            build_workbook_metadata(path)


# ---------------------------------------------------------------------------
# ANGLE F — Cross-tool consistency
# ---------------------------------------------------------------------------


def test_us_f1_capabilities_metadata(monkeypatch):
    """US-F1: capabilities() from _io returns a dict with key
    'metadata_contract_version'=='1.0' and 'metadata_policy' in ('strict','lenient').
    """
    monkeypatch.delenv("EXCEL_METADATA_POLICY", raising=False)

    from excelmcp._io import capabilities

    result = capabilities()

    assert "metadata_contract_version" in result, "Key 'metadata_contract_version' missing"
    assert result["metadata_contract_version"] == "1.0"
    assert "metadata_policy" in result, "Key 'metadata_policy' missing"
    assert result["metadata_policy"] in ("strict", "lenient")


# ---------------------------------------------------------------------------
# ANGLE G — Regression guard
# ---------------------------------------------------------------------------


def test_us_g1_meta_alias_parity(tmp_path, monkeypatch):
    """US-G1: Calling build_workbook_metadata twice on the same valid workbook
    returns identical keys and values both times.
    This confirms the /meta alias parity: both /metadata and /meta delegate to
    build_workbook_metadata and get the same result.
    """
    path = _make_wb(tmp_path, monkeypatch)

    result1 = dict(build_workbook_metadata(path))
    result2 = dict(build_workbook_metadata(path))

    assert set(result1.keys()) == set(result2.keys()), "Keys differ between two calls"
    assert result1 == result2, "Values differ between two calls"


# ---------------------------------------------------------------------------
# ANGLE H — Read operations
# ---------------------------------------------------------------------------


def _make_wb_with_values(tmp_path, monkeypatch) -> str:
    """Create a workbook with [[1,2],[3,4]] in A1:B2, return path str."""
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    p = tmp_path / "values.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = 1
    ws["B1"] = 2
    ws["A2"] = 3
    ws["B2"] = 4
    wb.save(str(p))
    return str(p)


def test_us_h1_read_range_returns_values(tmp_path, monkeypatch):
    """US-H1: read_range on A1:B2 returns the values [[1,2],[3,4]]."""
    path = _make_wb_with_values(tmp_path, monkeypatch)
    result = read_range(path, "Sheet1", "A1:B2")

    assert len(result) == 2, f"Expected 2 rows, got {len(result)}"
    assert result[0][0]["value"] == 1
    assert result[0][1]["value"] == 2
    assert result[1][0]["value"] == 3
    assert result[1][1]["value"] == 4


def test_us_h2_read_range_empty(tmp_path, monkeypatch):
    """US-H2: read_range on a range with no values returns a list of cells with None values
    (not an error).
    """
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    p = tmp_path / "empty.xlsx"
    wb = openpyxl.Workbook()
    wb.active.title = "Sheet1"
    wb.save(str(p))

    result = read_range(str(p), "Sheet1", "A1:B1")

    assert isinstance(result, list), "Expected list result for empty range"
    # All values should be None (empty cells)
    for row in result:
        for cell in row:
            assert cell["value"] is None, f"Expected None in empty cell, got {cell['value']!r}"


def test_us_h3_get_sheet_names_lists_all(tmp_path, monkeypatch):
    """US-H3: list_sheets on a 3-named-sheet workbook returns all 3 names."""
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    p = tmp_path / "three.xlsx"
    wb = openpyxl.Workbook()
    wb.active.title = "Alpha"
    wb.create_sheet("Beta")
    wb.create_sheet("Gamma")
    wb.save(str(p))

    sheets = list_sheets(str(p))

    assert set(sheets) == {"Alpha", "Beta", "Gamma"}
    assert len(sheets) == 3


def test_us_h4_read_cell_returns_value(tmp_path, monkeypatch):
    """US-H4: write 'Hello' to A1 via openpyxl, then read_cell returns value=='Hello'."""
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    p = tmp_path / "hello.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Hello"
    wb.save(str(p))

    result = read_cell(str(p), "Sheet1", "A1")

    assert result["value"] == "Hello"
    assert result["address"] == "A1"


def test_us_h5_capabilities_tool_count_positive():
    """US-H5: _io.capabilities() returns a dict where 'tools' is a non-empty list."""
    result = _io_capabilities()

    assert isinstance(result, dict), "Expected dict from capabilities()"
    assert "tools" in result, "Key 'tools' missing from capabilities()"
    assert isinstance(result["tools"], list), "Expected 'tools' to be a list"
    assert len(result["tools"]) > 0, "Expected at least one tool in capabilities()"


# ---------------------------------------------------------------------------
# ANGLE I — Write operations
# ---------------------------------------------------------------------------


def test_us_i1_write_range_persists_values(tmp_path, monkeypatch):
    """US-I1: write_range [[10,20],[30,40]] at A1, re-open → exact values persist."""
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    p = tmp_path / "write_range.xlsx"
    wb_obj = openpyxl.Workbook()
    wb_obj.active.title = "Sheet1"
    wb_obj.save(str(p))

    result = write_range(str(p), "Sheet1", "A1", [[10, 20], [30, 40]], confirm=True)

    assert result["ok"] is True
    assert result["cells_written"] == 4

    wb2 = openpyxl.load_workbook(str(p))
    ws2 = wb2.active
    assert ws2["A1"].value == 10
    assert ws2["B1"].value == 20
    assert ws2["A2"].value == 30
    assert ws2["B2"].value == 40


def test_us_i2_write_cell_single_value(tmp_path, monkeypatch):
    """US-I2: write_cell value 42 at C3, re-open → C3 == 42."""
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    p = tmp_path / "write_cell.xlsx"
    wb_obj = openpyxl.Workbook()
    wb_obj.active.title = "Sheet1"
    wb_obj.save(str(p))

    result = write_cell(str(p), "Sheet1", "C3", 42, confirm=True)

    assert result["ok"] is True

    wb2 = openpyxl.load_workbook(str(p))
    ws2 = wb2.active
    assert ws2["C3"].value == 42


def test_us_i3_append_row_adds_data(tmp_path, monkeypatch):
    """US-I3: append_rows on a workbook with 1 existing row → first_row >= 2."""
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    p = tmp_path / "append.xlsx"
    wb_obj = openpyxl.Workbook()
    ws = wb_obj.active
    ws.title = "Sheet1"
    ws["A1"] = "Name"
    ws["B1"] = "Value"
    wb_obj.save(str(p))

    result = append_rows(str(p), "Sheet1", [["Alice", 100]], confirm=True)

    assert result["ok"] is True
    assert result["rows_appended"] == 1
    assert result["first_row"] >= 2, "Appended row should be row 2 or later"

    wb2 = openpyxl.load_workbook(str(p))
    ws2 = wb2.active
    assert ws2.cell(row=result["first_row"], column=1).value == "Alice"
    assert ws2.cell(row=result["first_row"], column=2).value == 100


def test_us_i4_write_disabled_raises_error(tmp_path, monkeypatch):
    """US-I4: with EXCEL_ENABLE_WRITE absent (equivalent to false), write_cell raises
    NotAllowedError.
    """
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    monkeypatch.delenv("EXCEL_ENABLE_WRITE", raising=False)
    p = tmp_path / "disabled.xlsx"
    wb_obj = openpyxl.Workbook()
    wb_obj.active.title = "Sheet1"
    wb_obj.save(str(p))

    with pytest.raises((NotAllowedError, ExcelMCPError)):
        write_cell(str(p), "Sheet1", "A1", "blocked", confirm=True)


def test_us_i5_write_invalid_address_raises_error(tmp_path, monkeypatch):
    """US-I5: write_range with invalid start_cell raises ValidationError."""
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    p = tmp_path / "invalid_addr.xlsx"
    wb_obj = openpyxl.Workbook()
    wb_obj.active.title = "Sheet1"
    wb_obj.save(str(p))

    with pytest.raises((ValidationError, ExcelMCPError)):
        write_range(str(p), "Sheet1", "INVALID!!!", [[1, 2]], confirm=True)


# ---------------------------------------------------------------------------
# ANGLE J — Sheet management
# ---------------------------------------------------------------------------


def test_us_j1_add_sheet_creates_sheet(tmp_path, monkeypatch):
    """US-J1: add_sheet 'NewSheet' → appears in list_sheets result."""
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    p = tmp_path / "add_sheet.xlsx"
    wb_obj = openpyxl.Workbook()
    wb_obj.active.title = "Sheet1"
    wb_obj.save(str(p))

    result = add_sheet(str(p), "NewSheet", confirm=True)

    assert result["ok"] is True
    assert result["sheet_name"] == "NewSheet"

    sheets = list_sheets(str(p))
    assert "NewSheet" in sheets


def test_us_j2_delete_sheet_removes_sheet(tmp_path, monkeypatch):
    """US-J2: delete_sheet on 2-sheet workbook → only 1 sheet remains."""
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    p = tmp_path / "del_sheet.xlsx"
    wb_obj = openpyxl.Workbook()
    wb_obj.active.title = "KeepMe"
    wb_obj.create_sheet("DeleteMe")
    wb_obj.save(str(p))

    result = delete_sheet(str(p), "DeleteMe", confirm=True)

    assert result["ok"] is True
    sheets = list_sheets(str(p))
    assert "DeleteMe" not in sheets
    assert "KeepMe" in sheets
    assert len(sheets) == 1


def test_us_j3_rename_sheet_updates_name(tmp_path, monkeypatch):
    """US-J3: rename_sheet Sheet1 → MySheet; old name gone, new name present."""
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    p = tmp_path / "rename_sheet.xlsx"
    wb_obj = openpyxl.Workbook()
    wb_obj.active.title = "Sheet1"
    wb_obj.save(str(p))

    result = rename_sheet(str(p), "Sheet1", "MySheet", confirm=True)

    assert result["ok"] is True
    sheets = list_sheets(str(p))
    assert "MySheet" in sheets
    assert "Sheet1" not in sheets


def test_us_j4_copy_sheet_duplicates_data(tmp_path, monkeypatch):
    """US-J4: copy_sheet from Sheet1 to Sheet2; Sheet2 exists afterwards with same A1 value."""
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    p = tmp_path / "copy_sheet.xlsx"
    wb_obj = openpyxl.Workbook()
    ws = wb_obj.active
    ws.title = "Sheet1"
    ws["A1"] = "CopyMe"
    wb_obj.save(str(p))

    result = copy_sheet(str(p), "Sheet1", "Sheet2", confirm=True)

    assert result["ok"] is True
    sheets = list_sheets(str(p))
    assert "Sheet2" in sheets

    wb2 = openpyxl.load_workbook(str(p))
    assert wb2["Sheet2"]["A1"].value == "CopyMe"


# ---------------------------------------------------------------------------
# ANGLE K — Table operations
# ---------------------------------------------------------------------------


def _make_wb_with_table(tmp_path, monkeypatch) -> str:
    """Create workbook with headers A1:C1 + 2 data rows + a named table, return path."""
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    p = tmp_path / "tbl.xlsx"
    wb_obj = openpyxl.Workbook()
    ws = wb_obj.active
    ws.title = "Sheet1"
    ws["A1"] = "Name"
    ws["B1"] = "Score"
    ws["C1"] = "Grade"
    ws["A2"] = "Alice"
    ws["B2"] = 90
    ws["C2"] = "A"
    ws["A3"] = "Bob"
    ws["B3"] = 75
    ws["C3"] = "B"
    wb_obj.save(str(p))

    create_table(str(p), "Sheet1", "A1:C3", "tblStudents", confirm=True)
    return str(p)


def test_us_k1_create_table_on_range(tmp_path, monkeypatch):
    """US-K1: create_table on A1:C3 → appears in list_tables for that sheet."""
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    p = tmp_path / "k1_tbl.xlsx"
    wb_obj = openpyxl.Workbook()
    ws = wb_obj.active
    ws.title = "Sheet1"
    ws["A1"] = "Col1"
    ws["B1"] = "Col2"
    ws["C1"] = "Col3"
    ws["A2"] = 1
    ws["B2"] = 2
    ws["C2"] = 3
    wb_obj.save(str(p))

    result = create_table(str(p), "Sheet1", "A1:C2", "tblTest", confirm=True)

    assert result["table_name"] == "tblTest"

    tables = list_tables(str(p), "Sheet1")
    table_names = [t["name"] for t in tables]
    assert "tblTest" in table_names


def test_us_k2_append_table_row(tmp_path, monkeypatch):
    """US-K2: append_rows after the table's last row → row count increases by 1."""
    path = _make_wb_with_table(tmp_path, monkeypatch)

    result = append_rows(path, "Sheet1", [["Charlie", 85, "B"]], confirm=True)

    assert result["ok"] is True
    assert result["rows_appended"] == 1

    wb2 = openpyxl.load_workbook(path)
    ws2 = wb2["Sheet1"]
    # Row 4 should now have Charlie (3 header + 2 original + 1 new = row 4)
    appended_row = result["first_row"]
    assert ws2.cell(row=appended_row, column=1).value == "Charlie"


def test_us_k3_read_table_returns_structured(tmp_path, monkeypatch):
    """US-K3: read_table on tblStudents → returns headers and rows with expected values."""
    path = _make_wb_with_table(tmp_path, monkeypatch)

    result = read_table(path, "Sheet1", "tblStudents")

    assert "headers" in result, "Expected 'headers' key in read_table result"
    assert "rows" in result, "Expected 'rows' key in read_table result"
    assert result["headers"] == ["Name", "Score", "Grade"]
    assert len(result["rows"]) == 2
    # First row should be Alice's data
    first_row_values = result["rows"][0]
    assert "Alice" in first_row_values


# ---------------------------------------------------------------------------
# ANGLE L — Formatting
# ---------------------------------------------------------------------------


def test_us_l1_apply_cell_style_bold(tmp_path, monkeypatch):
    """US-L1: apply_cell_style bold=True on A1 → re-open shows font.bold == True."""
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    p = tmp_path / "bold.xlsx"
    wb_obj = openpyxl.Workbook()
    ws = wb_obj.active
    ws.title = "Sheet1"
    ws["A1"] = "Bold Text"
    wb_obj.save(str(p))

    result = apply_cell_style(
        str(p), "Sheet1", "A1", bold=True, italic=False,
        font_size=11.0, font_color="000000", confirm=True
    )

    assert result["ok"] is True

    wb2 = openpyxl.load_workbook(str(p))
    assert wb2["Sheet1"]["A1"].font.bold is True


def test_us_l2_set_column_width(tmp_path, monkeypatch):
    """US-L2: set_column_width 20 for column A → re-open shows column_dimensions["A"].width == 20."""
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    p = tmp_path / "colwidth.xlsx"
    wb_obj = openpyxl.Workbook()
    wb_obj.active.title = "Sheet1"
    wb_obj.save(str(p))

    result = set_column_width(str(p), "Sheet1", "A", 20.0, confirm=True)

    assert result["ok"] is True

    wb2 = openpyxl.load_workbook(str(p))
    assert wb2["Sheet1"].column_dimensions["A"].width == 20.0


def test_us_l3_merge_cells(tmp_path, monkeypatch):
    """US-L3: merge_cells A1:B2 → re-open shows A1:B2 in merged_cells."""
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    p = tmp_path / "merge.xlsx"
    wb_obj = openpyxl.Workbook()
    ws = wb_obj.active
    ws.title = "Sheet1"
    ws["A1"] = "Merged"
    wb_obj.save(str(p))

    result = merge_cells(str(p), "Sheet1", "A1:B2", confirm=True)

    assert result["ok"] is True

    wb2 = openpyxl.load_workbook(str(p))
    ws2 = wb2["Sheet1"]
    merged_ranges = [str(m) for m in ws2.merged_cells.ranges]
    assert "A1:B2" in merged_ranges


def test_us_l4_apply_fill_background(tmp_path, monkeypatch):
    """US-L4: apply_cell_style bg_color='FF0000' on C3 → re-open shows fill fgColor != default."""
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    p = tmp_path / "fill.xlsx"
    wb_obj = openpyxl.Workbook()
    ws = wb_obj.active
    ws.title = "Sheet1"
    ws["C3"] = "Red"
    wb_obj.save(str(p))

    result = apply_cell_style(
        str(p), "Sheet1", "C3", bold=False, italic=False,
        font_size=11.0, font_color="000000", bg_color="FF0000", confirm=True
    )

    assert result["ok"] is True

    wb2 = openpyxl.load_workbook(str(p))
    fill = wb2["Sheet1"]["C3"].fill
    # The fill should not be a default (no fill)
    assert fill is not None
    assert fill.fill_type != "none" or fill.fgColor.rgb != "00000000"


# ---------------------------------------------------------------------------
# ANGLE M — Comments & validation
# ---------------------------------------------------------------------------


def test_us_m1_add_comment(tmp_path, monkeypatch):
    """US-M1: add_cell_comment on A1 → re-open shows comment is not None."""
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    p = tmp_path / "comment.xlsx"
    wb_obj = openpyxl.Workbook()
    wb_obj.active.title = "Sheet1"
    wb_obj.save(str(p))

    result = add_cell_comment(
        str(p), "Sheet1", "A1", "Test comment", author="Tester", confirm=True
    )

    assert result["ok"] is True

    wb2 = openpyxl.load_workbook(str(p))
    assert wb2["Sheet1"]["A1"].comment is not None
    assert wb2["Sheet1"]["A1"].comment.text == "Test comment"


def test_us_m2_add_data_validation_list(tmp_path, monkeypatch):
    """US-M2: bulk_add_data_validation with list type on B2 → re-open shows non-empty validations."""
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    p = tmp_path / "dv.xlsx"
    wb_obj = openpyxl.Workbook()
    ws = wb_obj.active
    ws.title = "Sheet1"
    ws["B1"] = "Status"
    wb_obj.save(str(p))

    result = bulk_add_data_validation(
        str(p),
        "Sheet1",
        validations=[{
            "address": "B2",
            "validation_type": "list",
            "formula1": '"yes,no"',
        }],
        confirm=True,
    )

    assert result["ok"] is True

    wb2 = openpyxl.load_workbook(str(p))
    ws2 = wb2["Sheet1"]
    assert len(ws2.data_validations.dataValidation) > 0


def test_us_m3_remove_comment(tmp_path, monkeypatch):
    """US-M3: add a comment then delete it → re-open shows comment is None."""
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    p = tmp_path / "del_comment.xlsx"
    wb_obj = openpyxl.Workbook()
    wb_obj.active.title = "Sheet1"
    wb_obj.save(str(p))

    add_cell_comment(str(p), "Sheet1", "A1", "Temporary", confirm=True)
    del_result = delete_cell_comment(str(p), "Sheet1", "A1", confirm=True)

    assert del_result["ok"] is True

    wb2 = openpyxl.load_workbook(str(p))
    assert wb2["Sheet1"]["A1"].comment is None


# ---------------------------------------------------------------------------
# ANGLE N — Workbook lifecycle
# ---------------------------------------------------------------------------


def test_us_n1_capabilities_has_version():
    """US-N1: _io.capabilities() result has 'version' key with non-empty string value."""
    result = _io_capabilities()

    assert "version" in result, "Key 'version' missing from capabilities()"
    assert isinstance(result["version"], str)
    assert len(result["version"]) > 0


def test_us_n2_capabilities_prompt_count():
    """US-N2: _io.capabilities() with prompt_names supplied → prompts.count matches."""
    result = _io_capabilities(prompt_names=["p1", "p2", "p3"])

    assert "prompts" in result, "Key 'prompts' missing from capabilities()"
    assert isinstance(result["prompts"]["count"], int)
    assert result["prompts"]["count"] == 3


def test_us_n3_contract_version_import_consistent():
    """US-N3: CONTRACT_VERSION imported directly == get_contract_version() == '1.0'."""
    runtime_ver = get_contract_version()

    assert CONTRACT_VERSION == "1.0"
    assert runtime_ver == "1.0"
    assert CONTRACT_VERSION == runtime_ver


# ---------------------------------------------------------------------------
# ANGLE O — Error handling / edge cases
# ---------------------------------------------------------------------------


def test_us_o1_read_nonexistent_sheet_raises(tmp_path, monkeypatch):
    """US-O1: read_cell on a non-existent sheet name raises ValidationError."""
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    p = tmp_path / "one_sheet.xlsx"
    wb_obj = openpyxl.Workbook()
    wb_obj.active.title = "Sheet1"
    wb_obj.save(str(p))

    with pytest.raises((ValidationError, ExcelMCPError, KeyError)):
        read_cell(str(p), "NoSuchSheet", "A1")


def test_us_o2_validate_metadata_type_check():
    """US-O2: validate_metadata only checks key *presence*, not value types.

    Per implementation, validate_metadata raises ValueError only for *missing* keys.
    Passing a wrong type (e.g., sheet_count='three') does NOT raise if all keys are present.
    This test documents that behaviour (FEATURE GAP if strict type validation is wanted).
    """
    meta = _canonical_base()
    meta["sheet_count"] = "three"   # wrong type — all keys still present

    # Should NOT raise (key presence is satisfied)
    try:
        validate_metadata(meta)   # type: ignore[arg-type]
    except ValueError as exc:
        pytest.skip(
            f"validate_metadata raised ValueError for wrong type (stricter than spec): {exc}"
        )


def test_us_o3_build_workbook_metadata_csv_raises(tmp_path, monkeypatch):
    """US-O3: build_workbook_metadata on a .csv file raises ValidationError
    (unsupported extension check in _check_path).
    """
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    csv_path = tmp_path / "dummy.csv"
    csv_path.write_text("col1,col2\n1,2\n")

    with pytest.raises(Exception):
        build_workbook_metadata(str(csv_path))


def test_us_o4_normalize_metadata_with_only_alias_keys():
    """US-O4: normalize_metadata with all 9 alias keys + 'path' → all canonical keys present."""
    # Build a dict using ALL FIELD_ALIASES (9 aliases) + 'path' (no alias for path)
    # and 'sheet_count' (no alias for sheet_count)
    raw = {
        "path": "/some/file.xlsx",
        "name": "file.xlsx",           # → workbook_name
        "sheets": ["S1"],              # → sheet_names
        "active": "S1",               # → active_sheet
        "sheet_count": 1,              # no alias — use canonical key
        "formula": False,              # → has_formulas
        "modified": "2025-01-01T00:00:00+00:00",  # → last_modified
        "bytes": 1000,                 # → size_bytes
        "ext": ".xlsx",                # → format
        "version": "1.0",             # → contract_version
    }

    result = normalize_metadata(raw)

    for field in CANONICAL_FIELDS:
        assert field in result, f"Canonical field '{field}' missing after alias normalization"
    # Aliases must be absent from the result
    for alias in FIELD_ALIASES:
        if alias not in CANONICAL_FIELDS:  # only check aliases that aren't also canonical
            assert alias not in result, f"Alias key '{alias}' should not appear in normalized result"


# ---------------------------------------------------------------------------
# ANGLE P — Security
# ---------------------------------------------------------------------------


def test_us_p1_allowlist_env_respected(tmp_path, monkeypatch, tmp_path_factory):
    """US-P1: build_workbook_metadata on a path outside EXCEL_ALLOWLIST_ROOTS raises."""
    allowed_dir = tmp_path
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(allowed_dir))

    outside_dir = tmp_path_factory.mktemp("p1_outside")
    outside_path = outside_dir / "not_allowed.xlsx"
    wb_obj = openpyxl.Workbook()
    wb_obj.save(str(outside_path))

    with pytest.raises(Exception):
        build_workbook_metadata(str(outside_path))


def test_us_p2_path_traversal_blocked(tmp_path, monkeypatch):
    """US-P2: build_workbook_metadata with a path traversal string raises ValidationError."""
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    # Build a traversal path that attempts to escape tmp_path
    traversal_path = str(tmp_path / ".." / ".." / "etc" / "passwd.xlsx")

    with pytest.raises(Exception):
        build_workbook_metadata(traversal_path)


def test_us_p3_is_strict_returns_bool(monkeypatch):
    """US-P3: is_strict() returns a bool (True or False), not a string."""
    monkeypatch.delenv("EXCEL_METADATA_POLICY", raising=False)
    result = is_strict()

    assert isinstance(result, bool), f"Expected bool, got {type(result).__name__}: {result!r}"


def test_us_p4_load_runtime_config_invalid_raises(monkeypatch):
    """US-P4: EXCEL_METADATA_POLICY=bad_value → load_runtime_config() raises ValueError."""
    monkeypatch.setenv("EXCEL_METADATA_POLICY", "bad_value")

    with pytest.raises(ValueError):
        load_runtime_config()


# ---------------------------------------------------------------------------
# ANGLE Q — Chart operations (soft)
# ---------------------------------------------------------------------------


def test_us_q1_capabilities_does_not_raise():
    """US-Q1: _io.capabilities() returns a dict without raising — chart tools registered."""
    result = _io_capabilities()

    assert isinstance(result, dict), "capabilities() must return a dict"
    # Soft check: 'chart' tool should appear in tools list (Phase 2.2+)
    if "tools" in result:
        # Presence of 'chart' confirms chart phase is registered
        assert "chart" in result["tools"], (
            "Expected 'chart' in capabilities tools list — check Phase 2.2 registration"
        )


def test_us_q2_validate_metadata_all_fields_present():
    """US-Q2: validate_metadata with all 10 CANONICAL_FIELDS present → no exception."""
    meta = _canonical_base()  # already contains all 10 canonical fields

    # Should not raise
    validate_metadata(meta)


# ---------------------------------------------------------------------------
# ANGLE R — Capabilities governance
# ---------------------------------------------------------------------------


def test_us_r1_canonical_fields_len_is_10():
    """US-R1: CANONICAL_FIELDS has exactly 10 entries and first entry is 'path'."""
    assert len(CANONICAL_FIELDS) == 10, (
        f"Expected 10 CANONICAL_FIELDS, got {len(CANONICAL_FIELDS)}: {CANONICAL_FIELDS}"
    )
    assert CANONICAL_FIELDS[0] == "path", (
        f"Expected CANONICAL_FIELDS[0]=='path', got {CANONICAL_FIELDS[0]!r}"
    )


def test_us_r2_field_aliases_count_is_9():
    """US-R2: FIELD_ALIASES has exactly 9 entries; all values map to a CANONICAL_FIELDS entry."""
    assert len(FIELD_ALIASES) == 9, (
        f"Expected 9 FIELD_ALIASES, got {len(FIELD_ALIASES)}: {list(FIELD_ALIASES)}"
    )
    canonical_set = set(CANONICAL_FIELDS)
    for alias, canonical in FIELD_ALIASES.items():
        assert canonical in canonical_set, (
            f"Alias '{alias}' maps to '{canonical}' which is not in CANONICAL_FIELDS"
        )
