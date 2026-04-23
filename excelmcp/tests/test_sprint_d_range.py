import pytest

from excelmcp.workbook_openpyxl import (
    copy_range_format,
    write_range_with_format,
    apply_conditional_format,
    ValidationError,
    NotAllowedError,
    _wb_cache,
)


class TestCopyRangeFormat:
    """Unit tests for copy_range_format (T10, Sprint D Batch 2)."""

    def _make_wb(self, tmp_path, monkeypatch, extra_sheets=None):
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        path = tmp_path / "crf.xlsx"
        import openpyxl as _xl
        from openpyxl.styles import Font, PatternFill
        _wb = _xl.Workbook()
        ws = _wb.active
        ws.title = "Sheet1"
        for c in range(1, 4):
            for r in range(1, 4):
                cell = ws.cell(row=r, column=c)
                cell.value = f"R{r}C{c}"
                cell.font = Font(bold=True, size=12)
                cell.fill = PatternFill(fill_type="solid", fgColor="FF0000")
        if extra_sheets:
            for s in extra_sheets:
                _wb.create_sheet(s)
        _wb.save(str(path))
        return str(path)

    @pytest.mark.unit
    def test_crf_basic_same_sheet(self, tmp_path, monkeypatch):
        """Copy A1:C3 format to E1:G3 same sheet — cells_copied==9, fonts match."""
        path = self._make_wb(tmp_path, monkeypatch)
        result = copy_range_format(
            path, src_sheet="Sheet1", src_address="A1:C3",
            dst_sheet="Sheet1", dst_address="E1:G3", confirm=True
        )
        assert result["ok"] is True
        assert result["cells_copied"] == 9
        assert result["src"]["sheet"] == "Sheet1"
        assert result["dst"]["sheet"] == "Sheet1"

        _wb_cache.clear()
        import openpyxl as _xl
        wb_disk = _xl.load_workbook(path)
        assert wb_disk["Sheet1"]["E1"].font.bold is True
        assert wb_disk["Sheet1"]["E1"].font.size == 12

    @pytest.mark.unit
    def test_crf_cross_sheet(self, tmp_path, monkeypatch):
        """Copy format from Sheet1 to Sheet2 — values on Sheet2 unchanged."""
        path = self._make_wb(tmp_path, monkeypatch, extra_sheets=["Sheet2"])
        import openpyxl as _xl
        _wb = _xl.load_workbook(path)
        _wb["Sheet2"]["A1"].value = "OriginalValue"
        _wb.save(path)
        _wb_cache.clear()

        result = copy_range_format(
            path, src_sheet="Sheet1", src_address="A1:C3",
            dst_sheet="Sheet2", dst_address="A1:C3", confirm=True
        )
        assert result["ok"] is True
        assert result["cells_copied"] == 9

        _wb_cache.clear()
        wb_disk = _xl.load_workbook(path)
        assert wb_disk["Sheet2"]["A1"].value == "OriginalValue"
        assert wb_disk["Sheet2"]["A1"].font.bold is True

    @pytest.mark.unit
    def test_crf_single_cell_dst_expands(self, tmp_path, monkeypatch):
        """dst='E1' auto-expands to E1:G3 matching src A1:C3 (3×3)."""
        path = self._make_wb(tmp_path, monkeypatch)
        result = copy_range_format(
            path, src_sheet="Sheet1", src_address="A1:C3",
            dst_sheet="Sheet1", dst_address="E1", confirm=True
        )
        assert result["ok"] is True
        assert result["cells_copied"] == 9
        assert result["dst"]["address"] == "E1:G3"

    @pytest.mark.unit
    def test_crf_dst_wrong_dims_raises(self, tmp_path, monkeypatch):
        """dst_address with wrong dimensions raises ValidationError."""
        path = self._make_wb(tmp_path, monkeypatch)
        with pytest.raises(ValidationError, match="[Dd]imensions|mismatch"):
            copy_range_format(
                path, src_sheet="Sheet1", src_address="A1:C3",
                dst_sheet="Sheet1", dst_address="E1:F2",  # 2×2 ≠ 3×3
                confirm=True,
            )

    @pytest.mark.unit
    def test_crf_values_not_copied(self, tmp_path, monkeypatch):
        """Values in dst remain unchanged after format copy."""
        path = self._make_wb(tmp_path, monkeypatch, extra_sheets=["Sheet2"])
        import openpyxl as _xl
        _wb = _xl.load_workbook(path)
        _wb["Sheet2"]["A1"].value = "KeepMe"
        _wb["Sheet2"]["B1"].value = 42
        _wb.save(path)
        _wb_cache.clear()

        copy_range_format(
            path, src_sheet="Sheet1", src_address="A1:B1",
            dst_sheet="Sheet2", dst_address="A1:B1", confirm=True
        )
        _wb_cache.clear()
        wb_disk = _xl.load_workbook(path)
        assert wb_disk["Sheet2"]["A1"].value == "KeepMe"
        assert wb_disk["Sheet2"]["B1"].value == 42

    @pytest.mark.unit
    def test_crf_merged_cell_skipped(self, tmp_path, monkeypatch):
        """No error when src contains merged cell slave; cells_copied is reduced."""
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        import openpyxl as _xl
        from openpyxl.styles import Font
        path = tmp_path / "crf_merged.xlsx"
        _wb = _xl.Workbook()
        ws = _wb.active
        ws.title = "Sheet1"
        ws["A1"].font = Font(bold=True)
        ws["B1"].font = Font(bold=True)
        ws.merge_cells("A1:B1")  # B1 becomes a MergedCell slave
        _wb.create_sheet("Sheet2")
        _wb.save(str(path))

        result = copy_range_format(
            str(path), src_sheet="Sheet1", src_address="A1:B1",
            dst_sheet="Sheet2", dst_address="A1:B1", confirm=True
        )
        # A1 copied (1), B1 slave skipped → cells_copied == 1
        assert result["ok"] is True
        assert result["cells_copied"] == 1

    @pytest.mark.unit
    def test_crf_nonexistent_src_sheet_raises(self, tmp_path, monkeypatch):
        """Non-existent src_sheet raises ValidationError."""
        path = self._make_wb(tmp_path, monkeypatch)
        with pytest.raises(ValidationError, match="[Ss]ource sheet not found|[Ss]heet not found"):
            copy_range_format(
                path, src_sheet="NoSuchSheet", src_address="A1:C3",
                dst_sheet="Sheet1", dst_address="E1:G3", confirm=True
            )

    @pytest.mark.unit
    def test_crf_copy_border_false_default(self, tmp_path, monkeypatch):
        """Border is NOT copied when copy_border=False (default)."""
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        import openpyxl as _xl
        from openpyxl.styles import Border, Side, Font
        path = tmp_path / "crf_border.xlsx"
        _wb = _xl.Workbook()
        ws = _wb.active
        ws.title = "Sheet1"
        side = Side(style="thin")
        ws["A1"].border = Border(left=side, right=side, top=side, bottom=side)
        ws["A1"].font = Font(bold=True)
        _wb.save(str(path))

        copy_range_format(
            str(path), src_sheet="Sheet1", src_address="A1",
            dst_sheet="Sheet1", dst_address="C1", confirm=True, copy_border=False
        )
        _wb_cache.clear()
        wb_disk = _xl.load_workbook(path)
        dst = wb_disk["Sheet1"]["C1"]
        # Font should be copied
        assert dst.font.bold is True
        # Border should NOT be copied (default thin border object has no style)
        assert dst.border.left.style is None

    @pytest.mark.unit
    def test_crf_max_range_cells_rejected(self, tmp_path, monkeypatch):
        """src range exceeding EXCEL_MAX_RANGE_CELLS raises ValidationError."""
        path = self._make_wb(tmp_path, monkeypatch)
        monkeypatch.setenv("EXCEL_MAX_RANGE_CELLS", "4")  # 3×3=9 > 4
        with pytest.raises(ValidationError, match="[Ee]xceeds|[Tt]oo large|limit"):
            copy_range_format(
                path, src_sheet="Sheet1", src_address="A1:C3",
                dst_sheet="Sheet1", dst_address="E1:G3", confirm=True,
            )

    @pytest.mark.unit
    def test_crf_write_guard(self, tmp_path, monkeypatch):
        """EXCEL_ENABLE_WRITE not set → NotAllowedError before touching disk."""
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        monkeypatch.delenv("EXCEL_ENABLE_WRITE", raising=False)
        with pytest.raises(NotAllowedError):
            copy_range_format(
                str(tmp_path / "ignored.xlsx"), src_sheet="Sheet1", src_address="A1",
                dst_sheet="Sheet1", dst_address="B1", confirm=True,
            )


class TestWriteRangeWithFormat:
    """Unit tests for write_range_with_format."""

    def _make_wb(self, tmp_path, monkeypatch):
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        path = tmp_path / "wrf.xlsx"
        import openpyxl as _xl
        _wb = _xl.Workbook()
        _wb.active.title = "Sheet1"
        _wb.save(str(path))
        return str(path)

    @pytest.mark.unit
    def test_write_basic(self, tmp_path, monkeypatch):
        """Writes 2×2 data; reads back values match."""
        path = self._make_wb(tmp_path, monkeypatch)
        result = write_range_with_format(
            path, sheet="Sheet1", address="B2",
            data=[[1, 2], [3, 4]], confirm=True,
        )
        assert result["cells_written"] == 4
        _wb_cache.clear()
        import openpyxl as _xl
        wb_disk = _xl.load_workbook(path)
        assert wb_disk["Sheet1"]["B2"].value == 1
        assert wb_disk["Sheet1"]["C2"].value == 2
        assert wb_disk["Sheet1"]["B3"].value == 3
        assert wb_disk["Sheet1"]["C3"].value == 4

    @pytest.mark.unit
    def test_write_number_format(self, tmp_path, monkeypatch):
        """Writes with number_format; format persisted on disk."""
        path = self._make_wb(tmp_path, monkeypatch)
        write_range_with_format(
            path, sheet="Sheet1", address="A1",
            data=[[12345.678]], number_format="#,##0.00", confirm=True,
        )
        _wb_cache.clear()
        import openpyxl as _xl
        wb_disk = _xl.load_workbook(path)
        assert wb_disk["Sheet1"]["A1"].number_format == "#,##0.00"

    @pytest.mark.unit
    def test_write_alignment(self, tmp_path, monkeypatch):
        """Writes with wrap_text=True; alignment persisted."""
        path = self._make_wb(tmp_path, monkeypatch)
        write_range_with_format(
            path, sheet="Sheet1", address="A1",
            data=[["hello"]], alignment={"wrap_text": True}, confirm=True,
        )
        _wb_cache.clear()
        import openpyxl as _xl
        wb_disk = _xl.load_workbook(path)
        assert wb_disk["Sheet1"]["A1"].alignment.wrap_text is True

    @pytest.mark.unit
    def test_write_invalid_style_key_raises(self, tmp_path, monkeypatch):
        """style with unknown key raises ValidationError."""
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        with pytest.raises(ValidationError, match="[Uu]nknown style"):
            write_range_with_format(
                str(tmp_path / "x.xlsx"), sheet="Sheet1", address="A1",
                data=[[1]], style={"unknown_key": 1}, confirm=True,
            )

    @pytest.mark.unit
    def test_write_invalid_alignment_key_raises(self, tmp_path, monkeypatch):
        """alignment with unknown key raises ValidationError."""
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        with pytest.raises(ValidationError, match="[Uu]nknown alignment"):
            write_range_with_format(
                str(tmp_path / "x.xlsx"), sheet="Sheet1", address="A1",
                data=[[1]], alignment={"unknown_key": 1}, confirm=True,
            )

    @pytest.mark.unit
    def test_write_empty_data_returns_zero(self, tmp_path, monkeypatch):
        """data=[] returns {"cells_written": 0} without opening workbook."""
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        result = write_range_with_format(
            str(tmp_path / "nonexistent.xlsx"), sheet="Sheet1", address="A1",
            data=[], confirm=True,
        )
        assert result == {"cells_written": 0}

    @pytest.mark.unit
    def test_write_no_confirm_blocked(self, tmp_path, monkeypatch):
        """confirm=False raises ValidationError."""
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        with pytest.raises(ValidationError, match="[Cc]onfirm"):
            write_range_with_format(
                str(tmp_path / "x.xlsx"), sheet="Sheet1", address="A1",
                data=[[1]], confirm=False,
            )

    @pytest.mark.unit
    def test_write_guard(self, tmp_path, monkeypatch):
        """EXCEL_ENABLE_WRITE not set → NotAllowedError."""
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        monkeypatch.delenv("EXCEL_ENABLE_WRITE", raising=False)
        with pytest.raises(NotAllowedError):
            write_range_with_format(
                str(tmp_path / "x.xlsx"), sheet="Sheet1", address="A1",
                data=[[1]], confirm=True,
            )


class TestApplyConditionalFormat:
    """Unit tests for apply_conditional_format."""

    def _make_wb(self, tmp_path, monkeypatch):
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        path = tmp_path / "cf.xlsx"
        import openpyxl as _xl
        _wb = _xl.Workbook()
        ws = _wb.active
        ws.title = "Sheet1"
        for r in range(1, 6):
            ws.cell(row=r, column=1).value = r * 10
        _wb.save(str(path))
        return str(path)

    @pytest.mark.unit
    def test_cf_cell_is_greater_than(self, tmp_path, monkeypatch):
        """CF greaterThan rule applied; conditional_formatting non-empty."""
        path = self._make_wb(tmp_path, monkeypatch)
        result = apply_conditional_format(
            path, sheet="Sheet1", address="A1:A5",
            rule_type="cell_is", condition="greaterThan 30",
            format_spec={"bold": True, "bg_color": "FFFF00"},
            confirm=True,
        )
        assert result["status"] == "ok"
        assert result["rule_type"] == "cell_is"
        _wb_cache.clear()
        import openpyxl as _xl
        wb_disk = _xl.load_workbook(path)
        cf_rules = list(wb_disk["Sheet1"].conditional_formatting)
        assert len(cf_rules) > 0

    @pytest.mark.unit
    def test_cf_formula_rule(self, tmp_path, monkeypatch):
        """CF formula rule applied; conditional_formatting non-empty."""
        path = self._make_wb(tmp_path, monkeypatch)
        result = apply_conditional_format(
            path, sheet="Sheet1", address="A1:A5",
            rule_type="formula", condition="",
            format_spec={"italic": True},
            formula="=A1>30",
            confirm=True,
        )
        assert result["status"] == "ok"
        assert result["rule_type"] == "formula"
        _wb_cache.clear()
        import openpyxl as _xl
        wb_disk = _xl.load_workbook(path)
        cf_rules = list(wb_disk["Sheet1"].conditional_formatting)
        assert len(cf_rules) > 0

    @pytest.mark.unit
    def test_cf_invalid_rule_type_raises(self, tmp_path, monkeypatch):
        """rule_type='bad' raises ValidationError."""
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        with pytest.raises(ValidationError, match="[Ii]nvalid rule_type"):
            apply_conditional_format(
                str(tmp_path / "x.xlsx"), sheet="Sheet1", address="A1",
                rule_type="bad", condition="greaterThan 1",
                format_spec={}, confirm=True,
            )

    @pytest.mark.unit
    def test_cf_invalid_format_spec_key_raises(self, tmp_path, monkeypatch):
        """format_spec with unknown key raises ValidationError."""
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        with pytest.raises(ValidationError, match="[Uu]nknown format_spec"):
            apply_conditional_format(
                str(tmp_path / "x.xlsx"), sheet="Sheet1", address="A1",
                rule_type="cell_is", condition="greaterThan 1",
                format_spec={"unknown": 1}, confirm=True,
            )

    @pytest.mark.unit
    def test_cf_formula_injection_guard(self, tmp_path, monkeypatch):
        """Oversized formula raises ValidationError (injection guard)."""
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        with pytest.raises(ValidationError):
            apply_conditional_format(
                str(tmp_path / "x.xlsx"), sheet="Sheet1", address="A1",
                rule_type="formula", condition="",
                format_spec={},
                formula="=" + "x" * 8193,
                confirm=True,
            )

    @pytest.mark.unit
    def test_cf_formula_missing_raises(self, tmp_path, monkeypatch):
        """rule_type='formula' with formula=None raises ValidationError."""
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        with pytest.raises(ValidationError, match="[Ff]ormula"):
            apply_conditional_format(
                str(tmp_path / "x.xlsx"), sheet="Sheet1", address="A1",
                rule_type="formula", condition="",
                format_spec={}, formula=None, confirm=True,
            )

    @pytest.mark.unit
    def test_cf_no_confirm_blocked(self, tmp_path, monkeypatch):
        """confirm=False raises ValidationError."""
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        with pytest.raises(ValidationError, match="[Cc]onfirm"):
            apply_conditional_format(
                str(tmp_path / "x.xlsx"), sheet="Sheet1", address="A1",
                rule_type="cell_is", condition="greaterThan 1",
                format_spec={}, confirm=False,
            )

    @pytest.mark.unit
    def test_cf_write_guard(self, tmp_path, monkeypatch):
        """EXCEL_ENABLE_WRITE not set → NotAllowedError."""
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        monkeypatch.delenv("EXCEL_ENABLE_WRITE", raising=False)
        with pytest.raises(NotAllowedError):
            apply_conditional_format(
                str(tmp_path / "x.xlsx"), sheet="Sheet1", address="A1",
                rule_type="cell_is", condition="greaterThan 1",
                format_spec={}, confirm=True,
            )
