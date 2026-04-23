import pytest
import openpyxl

from excelmcp.workbook_openpyxl import (
    write_named_range,
    read_named_range,
    get_named_ranges,
    add_data_validation,
    get_data_validation_info,
)


class TestUS22NamedRangeDropdown:
    """US-22 — Named range creation and use as dropdown validation source."""

    @pytest.mark.unit
    def test_us22_write_and_read_named_range(self, tmp_path, monkeypatch):
        """write_named_range creates a named range; read_named_range returns its data."""
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        wb_path = tmp_path / "nr.xlsx"
        _wb = openpyxl.Workbook()
        ws = _wb.active
        ws.title = "Sheet1"
        for i, v in enumerate(["Apple", "Banana", "Cherry"], start=1):
            ws.cell(row=i, column=1, value=v)
        _wb.save(str(wb_path))

        result = write_named_range(str(wb_path), "Fruits", "Sheet1", "A1:A3", confirm=True)
        assert result["ok"] is True
        assert result["name"] == "Fruits"
        assert "Sheet1" in result["refers_to"]
        assert "A1:A3" in result["refers_to"]

        read_result = read_named_range(str(wb_path), "Fruits")
        assert read_result["name"] == "Fruits"
        assert len(read_result["data"]) == 3

    @pytest.mark.unit
    def test_us22_list_named_ranges(self, tmp_path, monkeypatch):
        """get_named_ranges returns all defined names after two write_named_range calls."""
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        wb_path = tmp_path / "nr_list.xlsx"
        _wb = openpyxl.Workbook()
        _wb.active.title = "Sheet1"
        _wb.save(str(wb_path))

        write_named_range(str(wb_path), "RangeA", "Sheet1", "A1:A5", confirm=True)
        write_named_range(str(wb_path), "RangeB", "Sheet1", "B1:B5", confirm=True)

        ranges = get_named_ranges(str(wb_path))
        names = [r["name"] for r in ranges]
        assert "RangeA" in names
        assert "RangeB" in names

    @pytest.mark.unit
    def test_us22_dvlist_with_namedrange_formula1(self, tmp_path, monkeypatch):
        """Dropdown formula1 can reference a named range by name (not an inline string)."""
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        wb_path = tmp_path / "dv_namedrange.xlsx"
        _wb = openpyxl.Workbook()
        ws = _wb.active
        ws.title = "Sheet1"
        for i, v in enumerate(["Yes", "No", "Maybe"], start=1):
            ws.cell(row=i, column=5, value=v)  # E1:E3
        _wb.save(str(wb_path))

        write_named_range(str(wb_path), "Options", "Sheet1", "E1:E3", confirm=True)
        add_data_validation(
            str(wb_path), "Sheet1", "A1:A10",
            validation_type="list", formula1="Options", confirm=True,
        )
        dvinfo = get_data_validation_info(str(wb_path), "Sheet1")
        assert dvinfo["count"] >= 1
        formulas = [v["formula1"] for v in dvinfo["validations"]]
        assert "Options" in formulas

    @pytest.mark.unit
    def test_us22_named_range_as_dropdown_source(self, tmp_path, monkeypatch):
        """Named range 'Choices' pre-created by write_named_range is used as formula1."""
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        wb_path = tmp_path / "dv_choices.xlsx"
        _wb = openpyxl.Workbook()
        ws = _wb.active
        ws.title = "Sheet1"
        for i, v in enumerate(["P1", "P2", "P3"], start=1):
            ws.cell(row=i, column=4, value=v)  # D1:D3
        _wb.save(str(wb_path))

        # C1: write_named_range MUST precede add_data_validation
        write_named_range(str(wb_path), "Choices", "Sheet1", "D1:D3", confirm=True)
        add_data_validation(
            str(wb_path), "Sheet1", "A1:A5",
            validation_type="list", formula1="Choices", confirm=True,
        )
        dvinfo = get_data_validation_info(str(wb_path), "Sheet1")
        formulas = [v["formula1"] for v in dvinfo["validations"]]
        assert "Choices" in formulas

    @pytest.mark.unit
    def test_us22_overwrite_existing_named_range(self, tmp_path, monkeypatch):
        """write_named_range on an existing name updates the refers_to reference."""
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        wb_path = tmp_path / "nr_overwrite.xlsx"
        _wb = openpyxl.Workbook()
        _wb.active.title = "Sheet1"
        _wb.save(str(wb_path))

        write_named_range(str(wb_path), "MyRange", "Sheet1", "A1:A5", confirm=True)
        result = write_named_range(str(wb_path), "MyRange", "Sheet1", "B1:B10", confirm=True)
        assert result["ok"] is True
        assert "B1:B10" in result["refers_to"]

        read_result = read_named_range(str(wb_path), "MyRange")
        assert "B1" in read_result["refers_to"]
