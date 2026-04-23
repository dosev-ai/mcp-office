"""Sprint O Commit A — unified tool tests.

Covers the 5 new tools that replace 10 old tools:
  rows()          replaces insert_rows + delete_rows
  cols()          replaces insert_cols + delete_cols
  sheet_properties()  replaces get_sheet_properties + set_sheet_properties
  cell_comment()  replaces get_cell_comment + delete_cell_comment
  protect()       replaces protect_sheet + protect_workbook

All tests are unit-level; no COM required.
"""
import pytest
import openpyxl

from fastmcp.exceptions import ToolError

from excelmcp.server import rows
from excelmcp.server import cols
from excelmcp.server import sheet_properties
from excelmcp.server import cell_comment
from excelmcp.server import protect


# ---------------------------------------------------------------------------
# rows() — insert
# ---------------------------------------------------------------------------

@pytest.mark.unit
def test_rows_insert_success(tmp_path, monkeypatch):
    """rows(operation='insert') inserts a blank row and returns ok=True."""
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    wb_path = tmp_path / "rows_insert.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "header"
    ws["A2"] = "data"
    wb.save(str(wb_path))

    result = rows(operation="insert", path=str(wb_path), sheet="Sheet1",
                  start_row=2, count=1, confirm=True)
    assert result["ok"] is True

    wb2 = openpyxl.load_workbook(str(wb_path))
    ws2 = wb2["Sheet1"]
    assert ws2["A2"].value is None        # newly inserted blank row
    assert ws2["A3"].value == "data"      # original row shifted down


@pytest.mark.unit
def test_rows_insert_confirm_false_raises(tmp_path, monkeypatch):
    """rows(operation='insert', confirm=False) must raise ToolError."""
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    wb_path = tmp_path / "rows_noconf.xlsx"
    openpyxl.Workbook().save(str(wb_path))
    with pytest.raises(ToolError):
        rows(operation="insert", path=str(wb_path), sheet="Sheet1",
             start_row=1, confirm=False)


@pytest.mark.unit
def test_rows_insert_no_write_gate_raises(tmp_path, monkeypatch):
    """rows(operation='insert') without EXCEL_ENABLE_WRITE must raise ToolError."""
    monkeypatch.delenv("EXCEL_ENABLE_WRITE", raising=False)
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    wb_path = tmp_path / "rows_nowrite.xlsx"
    openpyxl.Workbook().save(str(wb_path))
    with pytest.raises(ToolError):
        rows(operation="insert", path=str(wb_path), sheet="Sheet1",
             start_row=1, confirm=True)


# ---------------------------------------------------------------------------
# rows() — delete
# ---------------------------------------------------------------------------

@pytest.mark.unit
def test_rows_delete_success(tmp_path, monkeypatch):
    """rows(operation='delete') deletes a row and returns ok=True."""
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    wb_path = tmp_path / "rows_delete.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "keep"
    ws["A2"] = "delete_me"
    ws["A3"] = "keep2"
    wb.save(str(wb_path))

    result = rows(operation="delete", path=str(wb_path), sheet="Sheet1",
                  start_row=2, confirm=True)
    assert result["ok"] is True

    wb2 = openpyxl.load_workbook(str(wb_path))
    ws2 = wb2["Sheet1"]
    assert ws2["A1"].value == "keep"
    assert ws2["A2"].value == "keep2"


@pytest.mark.unit
def test_rows_delete_end_row_less_than_start_raises(tmp_path, monkeypatch):
    """rows(operation='delete', end_row < start_row) must raise ToolError."""
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    wb_path = tmp_path / "rows_del_bad.xlsx"
    openpyxl.Workbook().save(str(wb_path))
    with pytest.raises(ToolError, match="end_row"):
        rows(operation="delete", path=str(wb_path), sheet="Sheet1",
             start_row=5, end_row=3, confirm=True)


# ---------------------------------------------------------------------------
# cols() — insert
# ---------------------------------------------------------------------------

@pytest.mark.unit
def test_cols_insert_success(tmp_path, monkeypatch):
    """cols(operation='insert') inserts a blank column and returns ok=True."""
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    wb_path = tmp_path / "cols_insert.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "col_a"
    ws["B1"] = "col_b"
    wb.save(str(wb_path))

    result = cols(operation="insert", path=str(wb_path), sheet="Sheet1",
                  start_col=2, count=1, confirm=True)
    assert result["ok"] is True

    wb2 = openpyxl.load_workbook(str(wb_path))
    ws2 = wb2["Sheet1"]
    assert ws2["B1"].value is None        # newly inserted blank col
    assert ws2["C1"].value == "col_b"     # original col B shifted right


@pytest.mark.unit
def test_cols_insert_confirm_false_raises(tmp_path, monkeypatch):
    """cols(operation='insert', confirm=False) must raise ToolError."""
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    wb_path = tmp_path / "cols_noconf.xlsx"
    openpyxl.Workbook().save(str(wb_path))
    with pytest.raises(ToolError):
        cols(operation="insert", path=str(wb_path), sheet="Sheet1",
             start_col=1, confirm=False)


@pytest.mark.unit
def test_cols_insert_no_write_gate_raises(tmp_path, monkeypatch):
    """cols(operation='insert') without EXCEL_ENABLE_WRITE must raise ToolError."""
    monkeypatch.delenv("EXCEL_ENABLE_WRITE", raising=False)
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    wb_path = tmp_path / "cols_nowrite.xlsx"
    openpyxl.Workbook().save(str(wb_path))
    with pytest.raises(ToolError):
        cols(operation="insert", path=str(wb_path), sheet="Sheet1",
             start_col=1, confirm=True)


# ---------------------------------------------------------------------------
# cols() — delete
# ---------------------------------------------------------------------------

@pytest.mark.unit
def test_cols_delete_success(tmp_path, monkeypatch):
    """cols(operation='delete') deletes a column and returns ok=True."""
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    wb_path = tmp_path / "cols_delete.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "keep"
    ws["B1"] = "delete_me"
    ws["C1"] = "keep2"
    wb.save(str(wb_path))

    result = cols(operation="delete", path=str(wb_path), sheet="Sheet1",
                  start_col=2, confirm=True)
    assert result["ok"] is True

    wb2 = openpyxl.load_workbook(str(wb_path))
    ws2 = wb2["Sheet1"]
    assert ws2["A1"].value == "keep"
    assert ws2["B1"].value == "keep2"


@pytest.mark.unit
def test_cols_delete_end_col_less_than_start_raises(tmp_path, monkeypatch):
    """cols(operation='delete', end_col < start_col) must raise ToolError."""
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    wb_path = tmp_path / "cols_del_bad.xlsx"
    openpyxl.Workbook().save(str(wb_path))
    with pytest.raises(ToolError, match="end_col"):
        cols(operation="delete", path=str(wb_path), sheet="Sheet1",
             start_col=5, end_col=3, confirm=True)


# ---------------------------------------------------------------------------
# sheet_properties() — get
# ---------------------------------------------------------------------------

@pytest.mark.unit
def test_sheet_properties_get_returns_dict(tmp_path, monkeypatch):
    """sheet_properties(operation='get') returns a dict with sheet and state keys."""
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    wb_path = tmp_path / "sp_get.xlsx"
    wb = openpyxl.Workbook()
    wb.active.title = "Sheet1"
    wb.save(str(wb_path))

    result = sheet_properties(operation="get", path=str(wb_path), sheet="Sheet1")
    assert isinstance(result, dict)
    assert result["sheet"] == "Sheet1"
    assert "state" in result


# ---------------------------------------------------------------------------
# sheet_properties() — set
# ---------------------------------------------------------------------------

@pytest.mark.unit
def test_sheet_properties_set_success(tmp_path, monkeypatch):
    """sheet_properties(operation='set') sets tab_color and returns ok=True."""
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    wb_path = tmp_path / "sp_set.xlsx"
    wb = openpyxl.Workbook()
    wb.active.title = "Sheet1"
    wb.create_sheet("Sheet2")
    wb.save(str(wb_path))

    result = sheet_properties(operation="set", path=str(wb_path), sheet="Sheet1",
                               tab_color="FF0000", confirm=True)
    assert result["ok"] is True


@pytest.mark.unit
def test_sheet_properties_set_confirm_false_raises(tmp_path, monkeypatch):
    """sheet_properties(operation='set', confirm=False) must raise ToolError."""
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    wb_path = tmp_path / "sp_noconf.xlsx"
    wb = openpyxl.Workbook()
    wb.active.title = "Sheet1"
    wb.save(str(wb_path))
    with pytest.raises(ToolError):
        sheet_properties(operation="set", path=str(wb_path), sheet="Sheet1",
                         tab_color="00FF00", confirm=False)


# ---------------------------------------------------------------------------
# cell_comment() — get
# ---------------------------------------------------------------------------

@pytest.mark.unit
def test_cell_comment_get_returns_dict(tmp_path, monkeypatch):
    """cell_comment(operation='get') returns dict with has_comment key."""
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    wb_path = tmp_path / "cc_get.xlsx"
    wb = openpyxl.Workbook()
    wb.active.title = "Sheet1"
    wb.save(str(wb_path))

    result = cell_comment(operation="get", path=str(wb_path),
                          sheet="Sheet1", address="A1")
    assert isinstance(result, dict)
    assert result["address"] == "A1"
    assert result["has_comment"] is False
    assert result["text"] is None


# ---------------------------------------------------------------------------
# cell_comment() — delete
# ---------------------------------------------------------------------------

@pytest.mark.unit
def test_cell_comment_delete_success(tmp_path, monkeypatch):
    """cell_comment(operation='delete') removes a comment and returns ok=True."""
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    wb_path = tmp_path / "cc_del.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    from openpyxl.comments import Comment
    ws["A1"].comment = Comment("test comment", "Author")
    wb.save(str(wb_path))

    result = cell_comment(operation="delete", path=str(wb_path),
                          sheet="Sheet1", address="A1", confirm=True)
    assert result["ok"] is True


@pytest.mark.unit
def test_cell_comment_delete_confirm_false_raises(tmp_path, monkeypatch):
    """cell_comment(operation='delete', confirm=False) must raise ToolError."""
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    wb_path = tmp_path / "cc_noconf.xlsx"
    openpyxl.Workbook().save(str(wb_path))
    with pytest.raises(ToolError):
        cell_comment(operation="delete", path=str(wb_path),
                     sheet="Sheet1", address="A1", confirm=False)


# ---------------------------------------------------------------------------
# protect() — scope='sheet'
# ---------------------------------------------------------------------------

@pytest.mark.unit
def test_protect_sheet_success(tmp_path, monkeypatch):
    """protect(scope='sheet') protects the sheet and returns ok=True."""
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    wb_path = tmp_path / "prot_sheet.xlsx"
    wb = openpyxl.Workbook()
    wb.active.title = "Sheet1"
    wb.save(str(wb_path))

    result = protect(scope="sheet", path=str(wb_path), sheet="Sheet1",
                     password="pw", confirm=True)
    assert result["ok"] is True
    assert result["protected"] is True


@pytest.mark.unit
def test_protect_sheet_none_raises(tmp_path, monkeypatch):
    """protect(scope='sheet', sheet=None) must raise ToolError."""
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    wb_path = tmp_path / "prot_nosheet.xlsx"
    openpyxl.Workbook().save(str(wb_path))
    with pytest.raises(ToolError, match="sheet is required"):
        protect(scope="sheet", path=str(wb_path), sheet=None, confirm=True)


# ---------------------------------------------------------------------------
# protect() — scope='workbook'
# ---------------------------------------------------------------------------

@pytest.mark.unit
def test_protect_workbook_success(tmp_path, monkeypatch):
    """protect(scope='workbook') protects workbook structure and returns ok=True."""
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    wb_path = tmp_path / "prot_wb.xlsx"
    openpyxl.Workbook().save(str(wb_path))

    result = protect(scope="workbook", path=str(wb_path),
                     lock_structure=True, confirm=True)
    assert result["ok"] is True
    assert result["protected"] is True


@pytest.mark.unit
def test_protect_confirm_false_raises(tmp_path, monkeypatch):
    """protect(confirm=False) must raise ToolError."""
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    wb_path = tmp_path / "prot_noconf.xlsx"
    openpyxl.Workbook().save(str(wb_path))
    with pytest.raises(ToolError):
        protect(scope="workbook", path=str(wb_path), confirm=False)


# ---------------------------------------------------------------------------
# Invalid operation/scope guards
# ---------------------------------------------------------------------------

@pytest.mark.unit
def test_rows_unknown_operation_raises(tmp_path, monkeypatch):
    """rows(operation='bogus') must raise ToolError."""
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    wb_path = tmp_path / "rows_bogus.xlsx"
    openpyxl.Workbook().save(str(wb_path))
    with pytest.raises(ToolError, match="Unknown operation"):
        rows(operation="bogus", path=str(wb_path), sheet="Sheet1",  # type: ignore[arg-type]
             start_row=1, confirm=True)


@pytest.mark.unit
def test_protect_unknown_scope_raises(tmp_path, monkeypatch):
    """protect(scope='bogus') must raise ToolError."""
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    wb_path = tmp_path / "prot_bogus.xlsx"
    openpyxl.Workbook().save(str(wb_path))
    with pytest.raises(ToolError, match="Unknown scope"):
        protect(scope="bogus", path=str(wb_path), confirm=True)  # type: ignore[arg-type]
