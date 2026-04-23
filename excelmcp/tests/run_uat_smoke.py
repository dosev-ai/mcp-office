#!/usr/bin/env python3
"""
UAT Smoke Suite — excelmcp (Excel MCP Phase 1, openpyxl)
Date : 2026-02-23
Covers: S-01 stdout discipline, S-02 initialize, S-03 tools/list,
        S-04 read_range (allowlisted), S-05 path-traversal / outside allowlist
"""

import subprocess
import json
import sys
import os
import pathlib
import datetime
import tempfile
import shutil
import threading
import time

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
_DEFAULT_SRC = pathlib.Path(__file__).resolve().parent.parent / "src"
PYTHONPATH = os.environ.get("EXCELMCP_PYTHONPATH", str(_DEFAULT_SRC))
SERVER_CMD  = [sys.executable, "-m", "excelmcp.server"]

EXPECTED_TOOLS = {
    "capabilities", "sheet", "get_used_range",
    "cell", "range_io", "read_sheet_all",
    "named_range",
    "get_cell_metadata", "evaluate_formula",
    "list_tables", "read_table",
    "append_rows",
    "find_replace", "copy_range",
    "rows", "cols",
    "import_csv_to_sheet", "save",
    "apply_style", "hyperlink", "chart",
}

LOG_PATH = pathlib.Path(f"uat_evidence_excelmcp_{datetime.date.today()}.jsonl")

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def log(entry: dict) -> None:
    with LOG_PATH.open("a", encoding="utf-8") as fh:
        fh.write(json.dumps(entry) + "\n")


def start_server(allowlist_roots: str) -> subprocess.Popen:
    env = {**os.environ,
           "PYTHONPATH": PYTHONPATH,
           "EXCEL_ALLOWLIST_ROOTS": allowlist_roots}
    return subprocess.Popen(
        SERVER_CMD,
        stdin=subprocess.PIPE,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        env=env,
    )


def send(proc: subprocess.Popen, method: str,
         params: dict = None, req_id: int = 1) -> tuple:
    msg = {"jsonrpc": "2.0", "id": req_id, "method": method,
           "params": params or {}}
    raw_req = json.dumps(msg) + "\n"
    proc.stdin.write(raw_req.encode())
    proc.stdin.flush()
    raw_resp = proc.stdout.readline()
    if not raw_resp:
        raise RuntimeError("Server closed stdout unexpectedly")
    return json.loads(raw_resp.decode()), msg


def terminate(proc: subprocess.Popen) -> None:
    try:
        proc.stdin.close()
    except Exception:
        pass
    proc.terminate()
    try:
        proc.wait(timeout=5)
    except subprocess.TimeoutExpired:
        proc.kill()
        proc.wait()


def setup_workbook(parent_dir: pathlib.Path):
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "hello"
    ws["A2"] = 42
    path = parent_dir / "test.xlsx"
    wb.save(path)
    wb2 = openpyxl.load_workbook(path, read_only=True)
    actual_sheet = wb2.sheetnames[0]
    wb2.close()
    print(f"  [SETUP] workbook: {path}")
    print(f"  [SETUP] sheet   : {actual_sheet!r}")
    return path, actual_sheet


def main():
    results  = {}
    evidence = {}
    proc     = None

    sep = "=" * 64
    print(sep)
    print("excelmcp UAT Smoke Suite")
    print(f"Date      : {datetime.date.today()}")
    print(f"Server    : {' '.join(SERVER_CMD)}")
    print(f"PYTHONPATH: {PYTHONPATH}")
    print(sep)

    tmp_dir = pathlib.Path(tempfile.mkdtemp(prefix="excelmcp_uat_"))
    print(f"\n[SETUP] temp dir (allowlist root): {tmp_dir}\n")

    try:
        wb_path, sheet_name = setup_workbook(tmp_dir)

        # ── S-01: Stdout discipline (separate short-lived process) ────────
        print("\n--- S-01: stdout discipline ---")
        proc_s01 = start_server(allowlist_roots=str(tmp_dir))
        time.sleep(0.3)

        spontaneous_lines = []
        buf = b""

        def _try_read():
            nonlocal buf
            buf = proc_s01.stdout.readline()

        t = threading.Thread(target=_try_read, daemon=True)
        t.start()
        t.join(timeout=0.5)

        if t.is_alive():
            s01_clean = True
            print("  Server emitted NO spontaneous stdout before first request [OK]")
        else:
            spontaneous_lines.append(buf.decode(errors="replace").strip())
            try:
                json.loads(buf)
                s01_clean = True
                print(f"  Spontaneous line (JSON): {buf.decode().strip()!r}")
            except json.JSONDecodeError:
                s01_clean = False
                print(f"  FAIL – non-JSON line on stdout: {buf.decode()!r}")

        s01_status = "PASS" if s01_clean else "FAIL"
        print(f"  [S-01] Status: {s01_status}")
        terminate(proc_s01)
        results["S-01"] = s01_status
        evidence["S-01"] = {"spontaneous_lines": spontaneous_lines, "clean": s01_clean}
        log({"tc": "S-01", "status": s01_status, "spontaneous_lines": spontaneous_lines})

        # ── Start shared server for S-02 … S-05 ──────────────────────────
        proc = start_server(allowlist_roots=str(tmp_dir))
        time.sleep(0.3)

        # ── S-02: Initialize handshake ────────────────────────────────────
        print("\n--- S-02: initialize handshake ---")
        init_params = {
            "protocolVersion": "2024-11-05",
            "capabilities": {},
            "clientInfo": {"name": "uat-tester", "version": "1.0"},
        }
        init_resp, init_req = send(proc, "initialize", init_params, req_id=1)

        pv = init_resp.get("result", {}).get("protocolVersion", "MISSING")
        s02_pass = (
            init_resp.get("jsonrpc") == "2.0"
            and init_resp.get("id") == 1
            and "result" in init_resp
            and "protocolVersion" in init_resp.get("result", {})
        )
        s02_status = "PASS" if s02_pass else "FAIL"
        print(f"  protocolVersion : {pv}")
        print(f"  serverInfo      : {init_resp.get('result', {}).get('serverInfo', {})}")
        print(f"  [S-02] Status   : {s02_status}")
        results["S-02"] = s02_status
        evidence["S-02"] = {"request": init_req, "response": init_resp}
        log({"tc": "S-02", "status": s02_status, "request": init_req, "response": init_resp})

        # Send initialized notification
        initialized_notif = {"jsonrpc": "2.0", "method": "notifications/initialized", "params": {}}
        proc.stdin.write((json.dumps(initialized_notif) + "\n").encode())
        proc.stdin.flush()

        # ── S-03: tools/list ──────────────────────────────────────────────
        print("\n--- S-03: tools/list ---")
        tl_resp, tl_req = send(proc, "tools/list", {}, req_id=2)

        found   = {t["name"] for t in tl_resp.get("result", {}).get("tools", [])}
        missing = EXPECTED_TOOLS - found
        extra   = found - EXPECTED_TOOLS
        s03_pass   = len(missing) == 0
        s03_status = "PASS" if s03_pass else "FAIL"

        print(f"  Found tools  : {sorted(found)}")
        if missing:
            print(f"  MISSING      : {sorted(missing)}")
        if extra:
            print(f"  Extra (OK)   : {sorted(extra)}")
        print(f"  [S-03] Status: {s03_status}")
        results["S-03"] = s03_status
        evidence["S-03"] = {"request": tl_req, "response": tl_resp,
                             "found": sorted(found), "missing": sorted(missing)}
        log({"tc": "S-03", "status": s03_status, "request": tl_req, "response": tl_resp,
             "found_tools": sorted(found), "missing_tools": sorted(missing)})

        # ── S-04: read_range on allowlisted workbook ──────────────────────
        print("\n--- S-04: read_range (allowlisted workbook) ---")
        rr_params = {
            "name": "read_range",
            "arguments": {
                "path": str(wb_path),
                "sheet": sheet_name,
                "a1_range": "A1:A2",
            },
        }
        rr_resp, rr_req = send(proc, "tools/call", rr_params, req_id=3)

        resp_str  = json.dumps(rr_resp).lower()
        has_hello = "hello" in resp_str
        has_42    = "42"    in resp_str
        is_err    = (rr_resp.get("error") is not None or
                     rr_resp.get("result", {}).get("isError") is True)
        s04_pass   = has_hello and has_42 and not is_err
        s04_status = "PASS" if s04_pass else "FAIL"

        print(f"  Contains 'hello': {has_hello}")
        print(f"  Contains '42'   : {has_42}")
        print(f"  isError flag    : {is_err}")
        print(f"  Full response   : {json.dumps(rr_resp)}")
        print(f"  [S-04] Status   : {s04_status}")
        results["S-04"] = s04_status
        evidence["S-04"] = {"request": rr_req, "response": rr_resp}
        log({"tc": "S-04", "status": s04_status, "request": rr_req, "response": rr_resp})

        # ── S-05: Path outside allowlist ──────────────────────────────────
        print("\n--- S-05: path outside allowlist (security block) ---")
        oa_params = {
            "name": "sheet",
            "arguments": {"operation": "list", "path": r"C:\Windows\notepad.exe"},
        }
        oa_resp, oa_req = send(proc, "tools/call", oa_params, req_id=4)

        oa_str = json.dumps(oa_resp).lower()
        is_blocked = (
            oa_resp.get("error") is not None
            or oa_resp.get("result", {}).get("isError") is True
            or any(kw in oa_str for kw in
                   ("not allowed", "denied", "allowlist", "outside",
                    "permission", "unauthorized", "access", "forbidden"))
        )
        print(f"  Structured error/block received: {is_blocked}")
        print(f"  Full response: {json.dumps(oa_resp)}")

        # Verify server still alive
        ping_resp, _ = send(proc, "tools/list", {}, req_id=5)
        still_alive = "result" in ping_resp
        print(f"  Server still alive after block : {still_alive}")

        s05_pass   = is_blocked and still_alive
        s05_status = "PASS" if s05_pass else "FAIL"
        print(f"  [S-05] Status  : {s05_status}")
        results["S-05"] = s05_status
        evidence["S-05"] = {"request": oa_req, "response": oa_resp,
                             "ping_response": ping_resp,
                             "server_still_alive": still_alive}
        log({"tc": "S-05", "status": s05_status, "request": oa_req, "response": oa_resp,
             "server_still_alive": still_alive})

    except Exception as exc:
        print(f"\n[FATAL] Unhandled exception: {exc}")
        import traceback
        traceback.print_exc()
    finally:
        if proc is not None:
            terminate(proc)
        shutil.rmtree(tmp_dir, ignore_errors=True)
        print(f"\n[CLEANUP] temp dir removed: {tmp_dir}")

    # ── Summary ───────────────────────────────────────────────────────────
    print("\n" + sep)
    print("UAT Smoke Suite — Summary")
    print(sep)
    all_pass = all(v == "PASS" for v in results.values()) and len(results) == 5
    for tc in ["S-01", "S-02", "S-03", "S-04", "S-05"]:
        status = results.get(tc, "NOT RUN")
        mark   = "+" if status == "PASS" else "!"
        print(f"  {tc}: {status} [{mark}]")
    print()
    verdict = "PASS" if all_pass else "FAIL"
    print(f"Overall Verdict: {verdict}")
    print(f"Evidence log   : {LOG_PATH.absolute()}")
    print(sep)

    print("\n--- Full JSON-RPC Evidence ---\n")
    for tc, ev in evidence.items():
        print(f"=== {tc} ===")
        for k, v in ev.items():
            if isinstance(v, (dict, list)):
                print(f"  {k}:\n    {json.dumps(v, indent=2)}")
            else:
                print(f"  {k}: {v}")
        print()

    return 0 if all_pass else 1


if __name__ == "__main__":
    sys.exit(main())
