import pytest
import openpyxl

from excelmcp.workbook_openpyxl import (
    ValidationError,
    NotAllowedError,
    capabilities,
    list_sheets,
    get_used_range,
    read_cell,
    read_range,
    write_cell,
    write_range,
    append_rows,
    save,
    read_sheet_all,
)


# ── GROUP 1: capabilities() ───────────────────────────────────────────────

@pytest.mark.unit
def test_capabilities_returns_phase2_manifest():
    result = capabilities()
    assert result["phase"] == "2.0"
    assert result["backend"] == "openpyxl"
    assert "range_io" in result["tools"]
    assert "workbook_metadata" in result["tools"]
    assert len(result["tools"]) == len(set(result["tools"]))
    assert result["prompts"] == {"count": 0, "names": []}


@pytest.mark.unit
def test_capabilities_uses_supplied_prompt_inventory():
    result = capabilities(prompt_names=["alpha_prompt", "beta_prompt"])
    assert result["prompts"]["names"] == ["alpha_prompt", "beta_prompt"]
    assert result["prompts"]["count"] == 2


@pytest.mark.unit
def test_capabilities_copies_supplied_prompt_inventory():
    prompt_names = ["alpha_prompt"]

    result = capabilities(prompt_names=prompt_names)
    prompt_names.append("mutated_after_call")

    assert result["prompts"]["names"] == ["alpha_prompt"]


# ── GROUP 2: _check_path / allowlist ──────────────────────────────────────

@pytest.mark.unit
def test_path_not_in_allowlist(tmp_path, monkeypatch):
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    outside = tmp_path.parent / "outside_dir" / "file.xlsx"
    with pytest.raises(ValidationError, match="not in allowlist"):
        list_sheets(str(outside))


@pytest.mark.unit
def test_no_allowlist_configured(monkeypatch, tmp_path):
    monkeypatch.delenv("EXCEL_ALLOWLIST_ROOTS", raising=False)
    with pytest.raises(ValidationError, match="No allowlist"):
        list_sheets(str(tmp_path / "test.xlsx"))


@pytest.mark.unit
def test_unsupported_extension(tmp_path, monkeypatch):
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    csv_path = tmp_path / "data.csv"
    csv_path.write_text("a,b\n1,2\n")
    with pytest.raises(ValidationError, match="Unsupported extension"):
        list_sheets(str(csv_path))


@pytest.mark.unit
def test_check_path_null_byte_raises(monkeypatch):
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", "C:\\Temp")
    with pytest.raises(ValidationError, match="null byte"):
        list_sheets("C:\\Temp\\file\x00.xlsx")


# ── GROUP 3: list_sheets() ────────────────────────────────────────────────

@pytest.mark.unit
def test_list_sheets_happy(sample_workbook):
    result = list_sheets(sample_workbook)
    assert result == ["Sheet1"]


# ── GROUP 4: get_used_range() ─────────────────────────────────────────────

@pytest.mark.unit
def test_get_used_range_happy(sample_workbook):
    result = get_used_range(sample_workbook, "Sheet1")
    assert result["min_row"] == 1
    assert result["max_row"] == 2
    assert "address" in result


@pytest.mark.unit
def test_get_used_range_invalid_sheet(sample_workbook):
    with pytest.raises(ValidationError, match="Sheet not found"):
        get_used_range(sample_workbook, "NoSuchSheet")


@pytest.mark.unit
def test_get_used_range_empty_sheet(tmp_path, monkeypatch):
    """CODE-GAP-01 fix: get_used_range must not crash on a sheet with no data.
    openpyxl 3.x returns min_row=1/min_col=1 for empty sheets; the None guard
    protects against older openpyxl versions that return None instead.
    """
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    import openpyxl as _opx
    wb_path = tmp_path / "empty.xlsx"
    wb = _opx.Workbook()
    wb.create_sheet("EmptySheet")
    wb.save(str(wb_path))
    result = get_used_range(str(wb_path), "EmptySheet")
    assert isinstance(result, dict)
    assert result.get("empty") is True or result.get("address") is not None


# ── GROUP 5: read_cell() ──────────────────────────────────────────────────

@pytest.mark.unit
def test_read_cell_string(sample_workbook):
    result = read_cell(sample_workbook, "Sheet1", "A1")
    assert result["value"] == "hello"


@pytest.mark.unit
def test_read_cell_number(sample_workbook):
    result = read_cell(sample_workbook, "Sheet1", "A2")
    assert result["value"] == 42


@pytest.mark.unit
def test_read_cell_invalid_sheet(sample_workbook):
    with pytest.raises(ValidationError, match="Sheet not found"):
        read_cell(sample_workbook, "NoSuchSheet", "A1")


# ── GROUP 6: read_range() ─────────────────────────────────────────────────

@pytest.mark.unit
def test_read_range_happy(sample_workbook):
    # read_range returns list[list[dict]] — 2 rows x 1 col for A1:A2
    result = read_range(sample_workbook, "Sheet1", "A1:A2")
    assert isinstance(result, list)
    assert len(result) == 2
    assert result[0][0]["value"] == "hello"
    assert result[1][0]["value"] == 42


@pytest.mark.unit
def test_read_range_too_large(sample_workbook, monkeypatch):
    monkeypatch.setenv("EXCEL_MAX_RANGE_CELLS", "1")
    with pytest.raises(ValidationError, match="exceeds limit"):
        read_range(sample_workbook, "Sheet1", "A1:B5")


@pytest.mark.unit
def test_read_range_invalid_ref(sample_workbook):
    with pytest.raises(ValidationError, match="Invalid range"):
        read_range(sample_workbook, "Sheet1", "NOTARANGE")


# ── GROUP 7: write_cell() ─────────────────────────────────────────────────

@pytest.mark.unit
def test_write_cell_requires_enable_write(sample_workbook, monkeypatch):
    monkeypatch.delenv("EXCEL_ENABLE_WRITE", raising=False)
    with pytest.raises(NotAllowedError, match="EXCEL_ENABLE_WRITE"):
        write_cell(sample_workbook, "Sheet1", "A1", "new", confirm=True)


@pytest.mark.unit
def test_write_cell_requires_confirm(sample_workbook, monkeypatch):
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    with pytest.raises(ValidationError, match="confirm=True"):
        write_cell(sample_workbook, "Sheet1", "A1", "new", confirm=False)


@pytest.mark.unit
def test_write_cell_happy(sample_workbook, monkeypatch):
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    result = write_cell(sample_workbook, "Sheet1", "A1", "updated", confirm=True)
    assert result["ok"] is True
    # After _save_and_evict the workbook is flushed to disk and evicted from cache.
    # Verify persistence by reloading from disk.
    wb_check = openpyxl.load_workbook(sample_workbook)
    assert wb_check["Sheet1"]["A1"].value == "updated"


# ── GROUP 8: write_range() ────────────────────────────────────────────────

@pytest.mark.unit
def test_write_range_requires_enable_write(sample_workbook, monkeypatch):
    monkeypatch.delenv("EXCEL_ENABLE_WRITE", raising=False)
    with pytest.raises(NotAllowedError):
        write_range(sample_workbook, "Sheet1", "A1", [[1, 2], [3, 4]], confirm=True)


@pytest.mark.unit
def test_write_range_requires_confirm(sample_workbook, monkeypatch):
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    with pytest.raises(ValidationError):
        write_range(sample_workbook, "Sheet1", "A1", [[1, 2], [3, 4]], confirm=False)


@pytest.mark.unit
def test_write_range_happy(sample_workbook, monkeypatch):
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    result = write_range(sample_workbook, "Sheet1", "A1", [[1, 2], [3, 4]], confirm=True)
    assert result["ok"] is True
    assert result["cells_written"] == 4


@pytest.mark.unit
def test_write_range_too_large(sample_workbook, monkeypatch):
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    monkeypatch.setenv("EXCEL_MAX_RANGE_CELLS", "1")
    with pytest.raises(ValidationError, match="exceeds limit"):
        write_range(
            sample_workbook, "Sheet1", "A1",
            [[1, 2, 3], [4, 5, 6], [7, 8, 9]], confirm=True,
        )


# ── GROUP 9: append_rows() ────────────────────────────────────────────────

@pytest.mark.unit
def test_append_rows_requires_enable_write(sample_workbook, monkeypatch):
    monkeypatch.delenv("EXCEL_ENABLE_WRITE", raising=False)
    with pytest.raises(NotAllowedError):
        append_rows(sample_workbook, "Sheet1", [[1, 2]], confirm=True)


@pytest.mark.unit
def test_append_rows_requires_confirm(sample_workbook, monkeypatch):
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    with pytest.raises(ValidationError):
        append_rows(sample_workbook, "Sheet1", [[1, 2]], confirm=False)


@pytest.mark.unit
def test_append_rows_happy(sample_workbook, monkeypatch):
    # sample_workbook has rows 1 (A1="hello") and 2 (A2=42) → first appended row is 3
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    result = append_rows(sample_workbook, "Sheet1", [[1, 2], [3, 4]], confirm=True)
    assert result["ok"] is True
    assert result["rows_appended"] == 2
    assert result["first_row"] == 3


# ── GROUP 10: save() ──────────────────────────────────────────────────────

@pytest.mark.unit
def test_save_requires_enable_write(sample_workbook, monkeypatch):
    monkeypatch.delenv("EXCEL_ENABLE_WRITE", raising=False)
    with pytest.raises(NotAllowedError):
        save(sample_workbook, confirm=True)


@pytest.mark.unit
def test_save_requires_confirm(sample_workbook, monkeypatch):
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    # Populate the cache first so the "not loaded" check doesn't fire before confirm check
    write_cell(sample_workbook, "Sheet1", "A1", "x", confirm=True)
    with pytest.raises(ValidationError):
        save(sample_workbook, confirm=False)


@pytest.mark.unit
def test_save_happy(sample_workbook, monkeypatch):
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    write_cell(sample_workbook, "Sheet1", "B1", "saved_value", confirm=True)
    result = save(sample_workbook, confirm=True)
    assert result["ok"] is True
    # Verify value was flushed to disk
    wb_disk = openpyxl.load_workbook(sample_workbook)
    assert wb_disk["Sheet1"]["B1"].value == "saved_value"


@pytest.mark.unit
def test_save_no_pending_writes_returns_noop(sample_workbook, monkeypatch):
    """save() with no pending in-memory writes returns a graceful no-op (GAP-02 fix)."""
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    # _clear_wb_cache autouse fixture ensures cache is empty at test start
    result = save(sample_workbook, confirm=True)
    assert result["ok"] is True
    assert result.get("note") == "no_pending_writes"


# ── GROUP 11: read_sheet_all() ────────────────────────────────────────────

@pytest.mark.unit
def test_read_sheet_all_happy(sample_workbook):
    result = read_sheet_all(sample_workbook, "Sheet1")
    assert result["sheet"] == "Sheet1"
    assert result["rows"] == 2
    assert result["cols"] == 1
    assert isinstance(result["data"], list)
    assert result["data"][0][0]["value"] == "hello"
    assert result["data"][1][0]["value"] == 42


@pytest.mark.unit
def test_read_sheet_all_invalid_sheet(sample_workbook):
    with pytest.raises(ValidationError, match="Sheet not found"):
        read_sheet_all(sample_workbook, "NoSuch")
