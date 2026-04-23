"""
Sprint O Commit C — unified tool tests.

Covers: hyperlink(), apply_style(), cell(), range_io()
All 4 unified tools replacing 9 old tools.
"""
import pytest
import openpyxl

from excelmcp.server import cell, range_io, apply_style, hyperlink
from fastmcp.exceptions import ToolError


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture
def wb_path(tmp_path, monkeypatch):
    """Create a minimal test workbook with Sheet1 containing A1='hello', A2=42."""
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    path = tmp_path / "test.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "hello"
    ws["A2"] = 42
    wb.save(str(path))
    return str(path)


# ---------------------------------------------------------------------------
# US-17: cell() — read + write
# ---------------------------------------------------------------------------

@pytest.mark.unit
def test_cell_read_returns_value(wb_path):
    """cell(operation='read') returns dict with value key."""
    result = cell(operation="read", path=wb_path, sheet="Sheet1", address="A1")
    assert isinstance(result, dict)
    assert result["value"] == "hello"


@pytest.mark.unit
def test_cell_read_number(wb_path):
    """cell(operation='read') returns numeric value."""
    result = cell(operation="read", path=wb_path, sheet="Sheet1", address="A2")
    assert result["value"] == 42


@pytest.mark.unit
def test_cell_write_happy(wb_path, monkeypatch):
    """cell(operation='write') writes value and returns ok."""
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    result = cell(operation="write", path=wb_path, sheet="Sheet1", address="A1",
                  value="updated", confirm=True)
    assert result["ok"] is True
    # Verify persistence
    wb_check = openpyxl.load_workbook(wb_path)
    assert wb_check["Sheet1"]["A1"].value == "updated"


@pytest.mark.unit
def test_cell_write_requires_confirm(wb_path, monkeypatch):
    """cell(operation='write') raises ToolError when confirm=False."""
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    with pytest.raises(ToolError):
        cell(operation="write", path=wb_path, sheet="Sheet1", address="A1",
             value="x", confirm=False)


@pytest.mark.unit
def test_cell_write_value_none_raises(wb_path, monkeypatch):
    """cell(operation='write') raises ToolError when value=None."""
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    with pytest.raises(ToolError, match="value is required"):
        cell(operation="write", path=wb_path, sheet="Sheet1", address="A1",
             value=None, confirm=True)


@pytest.mark.unit
def test_cell_unknown_operation_raises(wb_path):
    """cell() raises ToolError for unknown operation."""
    with pytest.raises(ToolError, match="Unknown operation"):
        cell(operation="upsert", path=wb_path, sheet="Sheet1", address="A1")


# ---------------------------------------------------------------------------
# US-18: range_io() — read + write
# ---------------------------------------------------------------------------

@pytest.mark.unit
def test_range_io_read_returns_wrapped_dict(wb_path):
    """range_io(operation='read') returns {'data': ..., 'rows': N, 'cols': M}."""
    result = range_io(operation="read", path=wb_path, sheet="Sheet1", address="A1:A2")
    assert isinstance(result, dict)
    assert "data" in result
    assert "rows" in result
    assert "cols" in result
    assert result["rows"] == 2
    assert result["cols"] == 1
    assert result["data"][0][0]["value"] == "hello"
    assert result["data"][1][0]["value"] == 42


@pytest.mark.unit
def test_range_io_read_empty_range(wb_path):
    """range_io(operation='read') on empty range still returns dict with rows/cols."""
    result = range_io(operation="read", path=wb_path, sheet="Sheet1", address="C10:D11")
    assert isinstance(result, dict)
    assert result["rows"] == 2
    assert result["cols"] == 2


@pytest.mark.unit
def test_range_io_write_happy(wb_path, monkeypatch):
    """range_io(operation='write') writes 2-D values array."""
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    result = range_io(operation="write", path=wb_path, sheet="Sheet1", address="B1",
                      values=[[10, 20], [30, 40]], confirm=True)
    assert result["ok"] is True
    assert result["cells_written"] == 4


@pytest.mark.unit
def test_range_io_write_requires_confirm(wb_path, monkeypatch):
    """range_io(operation='write') raises ToolError when confirm=False."""
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    with pytest.raises(ToolError):
        range_io(operation="write", path=wb_path, sheet="Sheet1", address="A1",
                 values=[[1, 2]], confirm=False)


@pytest.mark.unit
def test_range_io_write_values_none_raises(wb_path, monkeypatch):
    """range_io(operation='write') raises ToolError when values=None."""
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    with pytest.raises(ToolError, match="values is required"):
        range_io(operation="write", path=wb_path, sheet="Sheet1", address="A1",
                 values=None, confirm=True)


@pytest.mark.unit
def test_range_io_unknown_operation_raises(wb_path):
    """range_io() raises ToolError for unknown operation."""
    with pytest.raises(ToolError, match="Unknown operation"):
        range_io(operation="patch", path=wb_path, sheet="Sheet1", address="A1")


# ---------------------------------------------------------------------------
# US-15: apply_style() — single + multi range
# ---------------------------------------------------------------------------

@pytest.mark.unit
def test_apply_style_single_range(wb_path, monkeypatch):
    """apply_style with ranges=['A1'] applies bold to single cell."""
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    result = apply_style(path=wb_path, sheet="Sheet1", ranges=["A1"],
                         bold=True, confirm=True)
    assert result["ok"] is True
    assert result["ranges_processed"] == 1


@pytest.mark.unit
def test_apply_style_multi_range(wb_path, monkeypatch):
    """apply_style with multiple ranges applies italic to all."""
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    result = apply_style(path=wb_path, sheet="Sheet1", ranges=["A1:A2", "B1:B2"],
                         italic=True, confirm=True)
    assert result["ok"] is True
    assert result["ranges_processed"] == 2


@pytest.mark.unit
def test_apply_style_requires_confirm(wb_path, monkeypatch):
    """apply_style raises ToolError when confirm=False."""
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    with pytest.raises(ToolError):
        apply_style(path=wb_path, sheet="Sheet1", ranges=["A1"],
                    bold=True, confirm=False)


# ---------------------------------------------------------------------------
# US-12: hyperlink() — get, add, remove + security
# ---------------------------------------------------------------------------

@pytest.mark.unit
def test_hyperlink_get_empty(wb_path):
    """hyperlink(operation='get') returns empty list for sheet with no links."""
    result = hyperlink(operation="get", path=wb_path, sheet="Sheet1")
    assert isinstance(result, dict)
    assert result["count"] == 0
    assert result["hyperlinks"] == []


@pytest.mark.unit
def test_hyperlink_add_single(wb_path, monkeypatch):
    """hyperlink(operation='add') adds one hyperlink."""
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    result = hyperlink(
        operation="add", path=wb_path, sheet="Sheet1",
        links=[{"address": "A1", "url": "https://example.com"}],
        confirm=True,
    )
    assert result["ok"] is True
    assert result["links_added"] == 1


@pytest.mark.unit
def test_hyperlink_add_multiple(wb_path, monkeypatch):
    """hyperlink(operation='add') adds multiple hyperlinks."""
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    result = hyperlink(
        operation="add", path=wb_path, sheet="Sheet1",
        links=[
            {"address": "A1", "url": "https://example.com"},
            {"address": "B2", "url": "https://another.org"},
        ],
        confirm=True,
    )
    assert result["ok"] is True
    assert result["links_added"] == 2


@pytest.mark.unit
def test_hyperlink_add_requires_confirm(wb_path, monkeypatch):
    """hyperlink(operation='add') raises ToolError when confirm=False."""
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    with pytest.raises(ToolError):
        hyperlink(
            operation="add", path=wb_path, sheet="Sheet1",
            links=[{"address": "A1", "url": "https://example.com"}],
            confirm=False,
        )


@pytest.mark.unit
def test_hyperlink_add_empty_links_raises(wb_path, monkeypatch):
    """hyperlink(operation='add') raises ToolError when links is empty."""
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    with pytest.raises(ToolError, match="links list"):
        hyperlink(operation="add", path=wb_path, sheet="Sheet1",
                  links=[], confirm=True)


@pytest.mark.unit
def test_hyperlink_add_blocks_javascript_url(wb_path, monkeypatch):
    """SECURITY: hyperlink(operation='add') BLOCKS javascript: URLs (OWASP A03)."""
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    with pytest.raises(ToolError):
        hyperlink(
            operation="add", path=wb_path, sheet="Sheet1",
            links=[{"address": "A1", "url": "javascript:alert(1)"}],
            confirm=True,
        )


@pytest.mark.unit
def test_hyperlink_add_blocks_data_url(wb_path, monkeypatch):
    """SECURITY: hyperlink(operation='add') BLOCKS data: URLs (OWASP A03)."""
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    with pytest.raises(ToolError):
        hyperlink(
            operation="add", path=wb_path, sheet="Sheet1",
            links=[{"address": "A1", "url": "data:text/html,<script>alert(1)</script>"}],
            confirm=True,
        )


@pytest.mark.unit
def test_hyperlink_add_blocks_vbscript_url(wb_path, monkeypatch):
    """SECURITY: hyperlink(operation='add') BLOCKS vbscript: URLs (OWASP A03)."""
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    with pytest.raises(ToolError):
        hyperlink(
            operation="add", path=wb_path, sheet="Sheet1",
            links=[{"address": "A1", "url": "vbscript:MsgBox(1)"}],
            confirm=True,
        )


@pytest.mark.unit
def test_hyperlink_remove_happy(tmp_path, monkeypatch):
    """hyperlink(operation='remove') removes an existing hyperlink."""
    from openpyxl.worksheet.hyperlink import Hyperlink as OxHyperlink
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))

    path = tmp_path / "hlink.xlsx"
    wb_obj = openpyxl.Workbook()
    ws = wb_obj.active
    ws.title = "Sheet1"
    ws["A1"].hyperlink = OxHyperlink(ref="A1", target="https://example.com")
    wb_obj.save(str(path))

    result = hyperlink(operation="remove", path=str(path), sheet="Sheet1",
                       address="A1", confirm=True)
    assert result["ok"] is True


@pytest.mark.unit
def test_hyperlink_remove_requires_address(wb_path, monkeypatch):
    """hyperlink(operation='remove') raises ToolError when address is None."""
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    with pytest.raises(ToolError, match="address is required"):
        hyperlink(operation="remove", path=wb_path, sheet="Sheet1",
                  address=None, confirm=True)


@pytest.mark.unit
def test_hyperlink_remove_requires_confirm(wb_path, monkeypatch):
    """hyperlink(operation='remove') raises ToolError when confirm=False."""
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    with pytest.raises(ToolError):
        hyperlink(operation="remove", path=wb_path, sheet="Sheet1",
                  address="A1", confirm=False)


@pytest.mark.unit
def test_hyperlink_unknown_operation_raises(wb_path):
    """hyperlink() raises ToolError for unknown operation."""
    with pytest.raises(ToolError, match="Unknown operation"):
        hyperlink(operation="edit", path=wb_path, sheet="Sheet1")


# ---------------------------------------------------------------------------
# Tool names removed — verify old names no longer registered
# ---------------------------------------------------------------------------

@pytest.mark.unit
def test_old_tools_not_registered():
    """Old tool names must NOT appear in the registered tool set."""
    import asyncio
    from excelmcp import server
    mcp = server.mcp

    tools_dict = None
    if hasattr(mcp, "_tool_manager") and hasattr(mcp._tool_manager, "_tools"):
        tools_dict = mcp._tool_manager._tools
    elif hasattr(mcp, "_tools"):
        tools_dict = mcp._tools

    if tools_dict is not None:
        tool_names = {name for name in tools_dict}
    else:
        async def _get_names():
            tools = await mcp.list_tools()
            return {t.name for t in tools}
        tool_names = asyncio.run(_get_names())

    removed = {
        "read_cell", "read_range", "write_cell", "write_range",
        "apply_cell_style", "apply_style_to_ranges",
        "get_hyperlinks", "bulk_add_hyperlinks", "remove_hyperlink",
    }
    still_present = removed & tool_names
    assert not still_present, f"Old tools still registered: {still_present}"

    new_tools = {"cell", "range_io", "apply_style", "hyperlink"}
    missing_new = new_tools - tool_names
    assert not missing_new, f"New unified tools not registered: {missing_new}"
