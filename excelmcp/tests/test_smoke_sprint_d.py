"""
Smoke tests — Sprint D Batch 1-3 and Phase 2.7 Sprint C tools.

Exercises MCP dispatch layer for:
  Sprint D Batch 1: bulk_add_comments, fill_formula_range
  Sprint D Batch 2: bulk_add_data_validation, apply_format_to_sheet_list
                    (column_widths / row_heights), copy_range_format
  Sprint D Batch 3: write_range_with_format (blocked), apply_conditional_format
                    (blocked), add_hyperlink, remove_hyperlink
  Sprint C:         bulk_copy_sheet, set_row_heights, apply_style_to_ranges,
                    set_column_widths, bulk_add_hyperlinks, apply_format_to_sheet_list

No COM required; always CI-runnable.
"""
import pytest
import openpyxl as _xl
from openpyxl.styles import Font
from openpyxl.worksheet.hyperlink import Hyperlink

from excelmcp import server
from fastmcp.exceptions import ToolError

from ._smoke_helpers import _run_tool


# ---------------------------------------------------------------------------
# Sprint D Batch 1 — bulk_add_comments, fill_formula_range
# ---------------------------------------------------------------------------


@pytest.mark.smoke
def test_smoke_bulk_add_comments_happy(tmp_path, monkeypatch):
    """Smoke: bulk_add_comments adds 2 comments through the MCP layer."""
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))

    wb_path = tmp_path / "batch_comments.xlsx"
    _wb = _xl.Workbook()
    _wb.active.title = "Sheet1"
    _wb.save(str(wb_path))

    result = _run_tool(server.mcp, "bulk_add_comments", {
        "path": str(wb_path),
        "sheet": "Sheet1",
        "comments": [
            {"address": "A1", "text": "Hello", "author": "Tester"},
            {"address": "B1", "text": "World", "author": "Tester"},
        ],
        "confirm": True,
    })
    assert result["ok"] is True
    assert result["comments_added"] == 2


@pytest.mark.smoke
def test_smoke_bulk_add_comments_no_confirm(tmp_path, monkeypatch):
    """Smoke: bulk_add_comments with confirm=False must raise ToolError."""
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))

    wb_path = tmp_path / "no_confirm.xlsx"
    _wb = _xl.Workbook()
    _wb.active.title = "Sheet1"
    _wb.save(str(wb_path))

    with pytest.raises(ToolError):
        _run_tool(server.mcp, "bulk_add_comments", {
            "path": str(wb_path),
            "sheet": "Sheet1",
            "comments": [{"address": "A1", "text": "Hi"}],
            "confirm": False,
        })


@pytest.mark.smoke
def test_smoke_fill_formula_range_happy(tmp_path, monkeypatch):
    """Smoke: fill_formula_range fills B2:B5 through the MCP layer."""
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))

    wb_path = tmp_path / "fill_formula.xlsx"
    _wb = _xl.Workbook()
    ws = _wb.active
    ws.title = "Sheet1"
    for i in range(1, 6):
        ws.cell(row=i, column=1, value=i)
    _wb.save(str(wb_path))

    result = _run_tool(server.mcp, "fill_formula_range", {
        "path": str(wb_path),
        "sheet": "Sheet1",
        "start_address": "B2",
        "end_address": "B5",
        "formula": "=A2*2",
        "confirm": True,
    })
    assert result["ok"] is True
    assert result["filled_cells"] == 4


@pytest.mark.smoke
def test_smoke_fill_formula_range_no_eq_raises(tmp_path, monkeypatch):
    """Smoke: fill_formula_range with formula not starting '=' raises ToolError."""
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))

    wb_path = tmp_path / "no_eq.xlsx"
    _wb = _xl.Workbook()
    _wb.active.title = "Sheet1"
    _wb.save(str(wb_path))

    with pytest.raises(ToolError):
        _run_tool(server.mcp, "fill_formula_range", {
            "path": str(wb_path),
            "sheet": "Sheet1",
            "start_address": "B2",
            "end_address": "B5",
            "formula": "A2*2",
            "confirm": True,
        })


# ---------------------------------------------------------------------------
# Sprint D Batch 2 — bulk_add_data_validation, apply_format_to_sheet_list,
#                    copy_range_format
# ---------------------------------------------------------------------------


@pytest.mark.smoke
def test_smoke_bulk_add_data_validation_happy(tmp_path, monkeypatch):
    """Smoke: bulk_add_data_validation adds 2 rules through the MCP layer."""
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))

    wb_path = tmp_path / "bulk_dv.xlsx"
    _wb = _xl.Workbook()
    _wb.active.title = "Sheet1"
    _wb.save(str(wb_path))

    result = _run_tool(server.mcp, "data_validation", {
        "operation": "add",
        "path": str(wb_path),
        "sheet": "Sheet1",
        "validations": [
            {"address": "A1:A10", "validation_type": "list", "formula1": '"Red,Green,Blue"'},
            {"address": "B1:B10", "validation_type": "whole", "formula1": "1",
             "formula2": "100", "operator": "between"},
        ],
        "confirm": True,
    })
    assert result["ok"] is True
    assert result["validations_added"] == 2


@pytest.mark.smoke
def test_smoke_bulk_add_data_validation_no_confirm(tmp_path, monkeypatch):
    """Smoke: bulk_add_data_validation with confirm=False raises ToolError."""
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))

    wb_path = tmp_path / "dv_no_confirm.xlsx"
    _wb = _xl.Workbook()
    _wb.active.title = "Sheet1"
    _wb.save(str(wb_path))

    with pytest.raises(ToolError):
        _run_tool(server.mcp, "data_validation", {
            "operation": "add",
            "path": str(wb_path),
            "sheet": "Sheet1",
            "validations": [{"address": "A1", "validation_type": "list",
                             "formula1": '"X,Y"'}],
            "confirm": False,
        })


@pytest.mark.smoke
def test_smoke_apply_format_to_sheet_list_column_widths(tmp_path, monkeypatch):
    """Smoke: apply_format_to_sheet_list with column_widths sets widths on all sheets."""
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))

    wb_path = tmp_path / "afsl_cw.xlsx"
    _wb = _xl.Workbook()
    _wb.active.title = "Sheet1"
    _wb.create_sheet("Sheet2")
    _wb.save(str(wb_path))

    result = _run_tool(server.mcp, "apply_format_to_sheet_list", {
        "path": str(wb_path),
        "column_widths": {"A": 20, "B": 15},
        "confirm": True,
    })
    assert result["ok"] is True
    assert result["column_widths_set"] == 4  # 2 cols × 2 sheets


@pytest.mark.smoke
def test_smoke_apply_format_to_sheet_list_row_heights(tmp_path, monkeypatch):
    """Smoke: apply_format_to_sheet_list with row_heights sets heights on all sheets."""
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))

    wb_path = tmp_path / "afsl_rh.xlsx"
    _wb = _xl.Workbook()
    _wb.active.title = "Sheet1"
    _wb.create_sheet("Sheet2")
    _wb.save(str(wb_path))

    result = _run_tool(server.mcp, "apply_format_to_sheet_list", {
        "path": str(wb_path),
        "row_heights": {"1": 30},
        "confirm": True,
    })
    assert result["ok"] is True
    assert result["row_heights_set"] == 2  # 1 row × 2 sheets


@pytest.mark.smoke
def test_smoke_copy_range_format_happy(tmp_path, monkeypatch):
    """Smoke: copy_range_format copies formatting (not values) to dst range."""
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))

    wb_path = tmp_path / "crf.xlsx"
    _wb = _xl.Workbook()
    ws = _wb.active
    ws.title = "Sheet1"
    ws["A1"].value = "Hello"
    ws["A1"].font = Font(bold=True, size=14)
    ws["B1"].value = "World"
    _wb.save(str(wb_path))

    result = _run_tool(server.mcp, "copy_range_format", {
        "path": str(wb_path),
        "src_sheet": "Sheet1",
        "src_address": "A1:B1",
        "dst_sheet": "Sheet1",
        "dst_address": "A3:B3",
        "confirm": True,
    })
    assert result["ok"] is True
    assert result["cells_copied"] == 2


@pytest.mark.smoke
def test_smoke_copy_range_format_no_confirm(tmp_path, monkeypatch):
    """Smoke: copy_range_format with confirm=False raises ToolError."""
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))

    wb_path = tmp_path / "crf_nc.xlsx"
    _wb = _xl.Workbook()
    _wb.active.title = "Sheet1"
    _wb.save(str(wb_path))

    with pytest.raises(ToolError):
        _run_tool(server.mcp, "copy_range_format", {
            "path": str(wb_path),
            "src_sheet": "Sheet1",
            "src_address": "A1:B1",
            "dst_sheet": "Sheet1",
            "dst_address": "A3:B3",
            "confirm": False,
        })


# ---------------------------------------------------------------------------
# Sprint D Batch 3 — write_range_with_format (blocked), apply_conditional_format
#                    (blocked), add_hyperlink, remove_hyperlink
# ---------------------------------------------------------------------------


@pytest.mark.smoke
def test_smoke_write_range_with_format_no_confirm(tmp_path, monkeypatch):
    """Smoke: write_range_with_format with confirm=False raises ToolError."""
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))

    wb_path = tmp_path / "wrf_nc.xlsx"
    _wb = _xl.Workbook()
    _wb.active.title = "Sheet1"
    _wb.save(str(wb_path))

    with pytest.raises(ToolError):
        _run_tool(server.mcp, "write_range_with_format", {
            "path": str(wb_path),
            "sheet": "Sheet1",
            "address": "A1",
            "data": [[1, 2], [3, 4]],
            "confirm": False,
        })


@pytest.mark.smoke
def test_smoke_apply_conditional_format_no_confirm(tmp_path, monkeypatch):
    """Smoke: apply_conditional_format with confirm=False raises ToolError."""
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))

    wb_path = tmp_path / "cf_nc.xlsx"
    _wb = _xl.Workbook()
    _wb.active.title = "Sheet1"
    _wb.save(str(wb_path))

    with pytest.raises(ToolError):
        _run_tool(server.mcp, "apply_conditional_format", {
            "path": str(wb_path),
            "sheet": "Sheet1",
            "address": "A1:A5",
            "rule_type": "cell_is",
            "condition": "greaterThan 10",
            "format_spec": {"bold": True},
            "confirm": False,
        })


@pytest.mark.smoke
def test_smoke_add_hyperlink(tmp_path, monkeypatch):
    """Smoke: hyperlink(operation='add') adds a single hyperlink."""
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))

    wb_path = tmp_path / "add_hlink.xlsx"
    _wb = _xl.Workbook()
    _wb.active.title = "Sheet1"
    _wb.save(str(wb_path))

    result = _run_tool(server.mcp, "hyperlink", {
        "operation": "add",
        "path": str(wb_path),
        "sheet": "Sheet1",
        "links": [{"address": "A1", "url": "https://example.com"}],
        "confirm": True,
    })
    assert result["ok"] is True
    assert result["links_added"] == 1


@pytest.mark.smoke
def test_smoke_remove_hyperlink(tmp_path, monkeypatch):
    """Smoke: hyperlink(operation='remove') removes a hyperlink from a cell."""
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))

    wb_path = tmp_path / "rem_hlink.xlsx"
    _wb = _xl.Workbook()
    ws = _wb.active
    ws.title = "Sheet1"
    ws["A1"].hyperlink = Hyperlink(ref="A1", target="https://example.com")
    _wb.save(str(wb_path))

    result = _run_tool(server.mcp, "hyperlink", {
        "operation": "remove",
        "path": str(wb_path),
        "sheet": "Sheet1",
        "address": "A1",
        "confirm": True,
    })
    assert result["ok"] is True


# ---------------------------------------------------------------------------
# Phase 2.7 Sprint C smoke tests
# ---------------------------------------------------------------------------


@pytest.mark.smoke
def test_smoke_bulk_copy_sheet(tmp_path, monkeypatch):
    """Smoke: bulk_copy_sheet creates named copies of a sheet."""
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))

    wb_path = tmp_path / "bcs.xlsx"
    _wb = _xl.Workbook()
    _wb.active.title = "Sheet1"
    _wb.save(str(wb_path))

    result = _run_tool(server.mcp, "bulk_copy_sheet", {
        "path": str(wb_path),
        "source_sheet": "Sheet1",
        "new_names": ["CopyA", "CopyB"],
        "confirm": True,
    })
    assert result["ok"] is True
    assert result["sheets_created"] == 2


@pytest.mark.smoke
def test_smoke_set_row_heights(tmp_path, monkeypatch):
    """Smoke: set_row_heights sets uniform height for a list of rows."""
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))

    wb_path = tmp_path / "srh.xlsx"
    _wb = _xl.Workbook()
    _wb.active.title = "Sheet1"
    _wb.save(str(wb_path))

    result = _run_tool(server.mcp, "set_row_heights", {
        "path": str(wb_path),
        "sheet": "Sheet1",
        "rows": [1, 2, 3],
        "height": 30.0,
        "confirm": True,
    })
    assert result["ok"] is True
    assert result["rows_set"] == 3


@pytest.mark.smoke
def test_smoke_apply_style_to_ranges(tmp_path, monkeypatch):
    """Smoke: apply_style applies bold to multiple ranges."""
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))

    wb_path = tmp_path / "asr.xlsx"
    _wb = _xl.Workbook()
    _wb.active.title = "Sheet1"
    _wb.save(str(wb_path))

    result = _run_tool(server.mcp, "apply_style", {
        "path": str(wb_path),
        "sheet": "Sheet1",
        "ranges": ["A1:B2", "C3:D4"],
        "bold": True,
        "confirm": True,
    })
    assert result["ok"] is True
    assert result["ranges_processed"] == 2


@pytest.mark.smoke
def test_smoke_set_column_widths(tmp_path, monkeypatch):
    """Smoke: set_column_widths sets explicit width for listed columns."""
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))

    wb_path = tmp_path / "scw.xlsx"
    _wb = _xl.Workbook()
    _wb.active.title = "Sheet1"
    _wb.save(str(wb_path))

    result = _run_tool(server.mcp, "set_column_widths", {
        "path": str(wb_path),
        "sheet": "Sheet1",
        "columns": ["A", "B", "C"],
        "width": 20.0,
        "confirm": True,
    })
    assert result["ok"] is True
    assert result["columns_set"] == 3


@pytest.mark.smoke
def test_smoke_bulk_add_hyperlinks(tmp_path, monkeypatch):
    """Smoke: hyperlink(operation='add') adds multiple hyperlinks in one call."""
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))

    wb_path = tmp_path / "bah.xlsx"
    _wb = _xl.Workbook()
    _wb.active.title = "Sheet1"
    _wb.save(str(wb_path))

    result = _run_tool(server.mcp, "hyperlink", {
        "operation": "add",
        "path": str(wb_path),
        "sheet": "Sheet1",
        "links": [
            {"address": "A1", "url": "https://example.com"},
            {"address": "B2", "url": "https://another.com"},
        ],
        "confirm": True,
    })
    assert result["ok"] is True
    assert result["links_added"] == 2


@pytest.mark.smoke
def test_smoke_apply_format_to_sheet_list(tmp_path, monkeypatch):
    """Smoke: apply_format_to_sheet_list applies bold to A1 across all sheets."""
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))

    wb_path = tmp_path / "afsl.xlsx"
    _wb = _xl.Workbook()
    _wb.active.title = "Sheet1"
    _wb.create_sheet("Sheet2")
    _wb.save(str(wb_path))

    result = _run_tool(server.mcp, "apply_format_to_sheet_list", {
        "path": str(wb_path),
        "address": "A1",
        "style": {"bold": True},
        "confirm": True,
    })
    assert result["ok"] is True
    assert result["sheets_processed"] == 2
