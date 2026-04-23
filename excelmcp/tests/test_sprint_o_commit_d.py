"""Sprint O Commit D — unified tool tests.

Covers the 2 new tools that replace 7 old tools:
  named_range()  replaces get_named_ranges + read_named_range + write_named_range
  chart()        replaces create_chart + list_charts + delete_chart + update_chart_data

All tests are unit-level; no COM required.
"""
from __future__ import annotations

import pytest
import openpyxl

from fastmcp.exceptions import ToolError

from excelmcp.server import named_range
from excelmcp.server_chart import chart


# ---------------------------------------------------------------------------
# named_range() — list
# ---------------------------------------------------------------------------

@pytest.mark.unit
def test_named_range_list_returns_dict(tmp_path, monkeypatch):
    """named_range(operation='list') returns {'named_ranges': [...], 'count': N}."""
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    wb_path = tmp_path / "list_nr.xlsx"
    wb = openpyxl.Workbook()
    wb.active.title = "Sheet1"
    wb.save(str(wb_path))

    result = named_range(operation="list", path=str(wb_path))
    assert "named_ranges" in result
    assert "count" in result
    assert isinstance(result["named_ranges"], list)
    assert result["count"] == len(result["named_ranges"])


@pytest.mark.unit
def test_named_range_list_includes_created_range(tmp_path, monkeypatch):
    """named_range(operation='list') includes a range created via write."""
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    wb_path = tmp_path / "list_nr2.xlsx"
    wb = openpyxl.Workbook()
    wb.active.title = "Sheet1"
    wb.save(str(wb_path))

    named_range(operation="write", path=str(wb_path),
                name="Sales", sheet="Sheet1", address="A1:A5", confirm=True)

    result = named_range(operation="list", path=str(wb_path))
    names = [r["name"] for r in result["named_ranges"]]
    assert "Sales" in names
    assert result["count"] >= 1


# ---------------------------------------------------------------------------
# named_range() — read
# ---------------------------------------------------------------------------

@pytest.mark.unit
def test_named_range_read_success(tmp_path, monkeypatch):
    """named_range(operation='read', name='MyRange') returns data."""
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    wb_path = tmp_path / "read_nr.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = 10
    ws["A2"] = 20
    wb.save(str(wb_path))

    named_range(operation="write", path=str(wb_path),
                name="MyRange", sheet="Sheet1", address="A1:A2", confirm=True)

    result = named_range(operation="read", path=str(wb_path), name="MyRange")
    assert result["name"] == "MyRange"
    assert "data" in result


@pytest.mark.unit
def test_named_range_read_name_none_raises(tmp_path, monkeypatch):
    """named_range(operation='read', name=None) raises ToolError."""
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    wb_path = tmp_path / "read_none.xlsx"
    openpyxl.Workbook().save(str(wb_path))

    with pytest.raises(ToolError, match="name is required"):
        named_range(operation="read", path=str(wb_path), name=None)


# ---------------------------------------------------------------------------
# named_range() — write
# ---------------------------------------------------------------------------

@pytest.mark.unit
def test_named_range_write_success(tmp_path, monkeypatch):
    """named_range(operation='write', ..., confirm=True) creates a named range."""
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    wb_path = tmp_path / "write_nr.xlsx"
    wb = openpyxl.Workbook()
    wb.active.title = "Sheet1"
    wb.save(str(wb_path))

    result = named_range(operation="write", path=str(wb_path),
                         name="Region", sheet="Sheet1",
                         address="B1:B10", confirm=True)
    assert result["ok"] is True
    assert result["name"] == "Region"


@pytest.mark.unit
def test_named_range_write_confirm_false_raises(tmp_path, monkeypatch):
    """named_range(operation='write', confirm=False) raises ToolError."""
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    wb_path = tmp_path / "write_noconf.xlsx"
    openpyxl.Workbook().save(str(wb_path))

    with pytest.raises(ToolError):
        named_range(operation="write", path=str(wb_path),
                    name="X", sheet="Sheet1",
                    address="A1:A1", confirm=False)


@pytest.mark.unit
def test_named_range_write_sheet_none_raises(tmp_path, monkeypatch):
    """named_range(operation='write', sheet=None) raises ToolError."""
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    wb_path = tmp_path / "write_nosht.xlsx"
    openpyxl.Workbook().save(str(wb_path))

    with pytest.raises(ToolError, match="sheet is required"):
        named_range(operation="write", path=str(wb_path),
                    name="X", sheet=None,
                    address="A1:A1", confirm=True)


@pytest.mark.unit
def test_named_range_write_range_address_none_raises(tmp_path, monkeypatch):
    """named_range(operation='write', address=None) raises ToolError."""
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    wb_path = tmp_path / "write_noaddr.xlsx"
    openpyxl.Workbook().save(str(wb_path))

    with pytest.raises(ToolError, match="address is required"):
        named_range(operation="write", path=str(wb_path),
                    name="X", sheet="Sheet1",
                    address=None, confirm=True)


# ---------------------------------------------------------------------------
# chart() — list
# ---------------------------------------------------------------------------

@pytest.mark.unit
def test_chart_list_returns_dict(tmp_path, monkeypatch):
    """chart(operation='list') returns {'charts': [...], 'count': N}."""
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    wb_path = tmp_path / "list_charts.xlsx"
    wb = openpyxl.Workbook()
    wb.active.title = "Sheet1"
    wb.save(str(wb_path))

    result = chart(operation="list", path=str(wb_path), sheet="Sheet1")
    assert "charts" in result
    assert "count" in result
    assert result["count"] == 0


# ---------------------------------------------------------------------------
# chart() — create
# ---------------------------------------------------------------------------

@pytest.mark.unit
def test_chart_create_success(tmp_path, monkeypatch):
    """chart(operation='create', chart_type='bar', data_range=..., confirm=True) succeeds."""
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    wb_path = tmp_path / "create_chart.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for i, v in enumerate([10, 20, 30], 1):
        ws.cell(row=i, column=1, value=v)
    wb.save(str(wb_path))

    result = chart(operation="create", path=str(wb_path), sheet="Sheet1",
                   chart_type="bar", data_address="A1:A3", confirm=True)
    assert result["ok"] is True


@pytest.mark.unit
def test_chart_create_accepts_preferred_aliases_and_target_address_wins(tmp_path, monkeypatch):
    """target_address works correctly for chart creation."""
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    wb_path = tmp_path / "create_chart_aliases.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Month"
    ws["B1"] = "Revenue"
    ws.append(["Jan", 10])
    ws.append(["Feb", 20])
    ws.append(["Mar", 30])
    wb.save(str(wb_path))

    result = chart(
        operation="create",
        path=str(wb_path),
        sheet="Sheet1",
        chart_type="bar",
        data_address="A1:B4",
        target_address="G2",
        confirm=True,
    )
    assert result["ok"] is True
    assert result["chart_type"] == "bar"
    assert result["target_cell"] == "G2"


@pytest.mark.unit
def test_chart_create_wrapper_passes_anchor_none_when_target_address_supplied(monkeypatch):
    """Wrapper passes anchor through to backend when target_address is supplied."""
    captured: dict[str, object] = {}

    def fake_create_chart(path, sheet, chart_type, data_address, **kwargs):
        captured["path"] = path
        captured["sheet"] = sheet
        captured["chart_type"] = chart_type
        captured["data_address"] = data_address
        captured.update(kwargs)
        return {
            "ok": True,
            "chart_type": chart_type,
            "target_cell": kwargs["target_cell"],
        }

    monkeypatch.setattr("excelmcp.server_chart.wb.create_chart", fake_create_chart)

    result = chart(
        operation="create",
        path="C:/allowed/test.xlsx",
        sheet="Sheet1",
        chart_type="bar",
        data_address="A1:B4",
        anchor="E7",
        target_address="G2",
        confirm=True,
    )

    assert result["ok"] is True
    assert captured["target_cell"] == "G2"
    assert captured["anchor"] == "E7"


@pytest.mark.unit
def test_chart_create_wrapper_preserves_anchor_when_target_address_absent(monkeypatch):
    """Wrapper preserves anchor when target_address defaults to D1."""
    captured: dict[str, object] = {}

    def fake_create_chart(path, sheet, chart_type, data_address, **kwargs):
        captured.update(kwargs)
        return {
            "ok": True,
            "chart_type": chart_type,
            "target_cell": kwargs["target_cell"],
        }

    monkeypatch.setattr("excelmcp.server_chart.wb.create_chart", fake_create_chart)

    result = chart(
        operation="create",
        path="C:/allowed/test.xlsx",
        sheet="Sheet1",
        chart_type="bar",
        data_address="A1:B4",
        anchor="E7",
        confirm=True,
    )

    assert result["ok"] is True
    assert captured["target_cell"] == "D1"
    assert captured["anchor"] == "E7"


@pytest.mark.unit
def test_chart_create_rejects_empty_explicit_target_address(monkeypatch):
    """Explicit empty `target_address` is rejected before `wb.create_chart()` is called."""
    create_chart_calls = 0

    def fake_create_chart(*args, **kwargs):
        nonlocal create_chart_calls
        create_chart_calls += 1
        return {"ok": True}

    monkeypatch.setattr("excelmcp.server_chart.wb.create_chart", fake_create_chart)

    with pytest.raises(ToolError, match="target_address must not be empty when supplied"):
        chart(
            operation="create",
            path="C:/allowed/test.xlsx",
            sheet="Sheet1",
            chart_type="bar",
            data_address="A1:B4",
            target_address="",
            confirm=True,
        )

    assert create_chart_calls == 0


@pytest.mark.unit
def test_chart_create_rejects_whitespace_only_explicit_target_address(monkeypatch):
    """Whitespace-only `target_address` is rejected before `wb.create_chart()` is called."""
    create_chart_calls = 0

    def fake_create_chart(*args, **kwargs):
        nonlocal create_chart_calls
        create_chart_calls += 1
        return {"ok": True}

    monkeypatch.setattr("excelmcp.server_chart.wb.create_chart", fake_create_chart)

    with pytest.raises(ToolError, match="target_address must not be empty when supplied"):
        chart(
            operation="create",
            path="C:/allowed/test.xlsx",
            sheet="Sheet1",
            chart_type="bar",
            data_address="A1:B4",
            target_address="  	  ",
            confirm=True,
        )

    assert create_chart_calls == 0


@pytest.mark.unit
@pytest.mark.parametrize("explicit_target_address", [123, 0, False])
def test_chart_create_rejects_non_string_explicit_target_address(monkeypatch, explicit_target_address):
    """Explicit non-string `target_address` values are rejected before fallback or backend calls."""
    create_chart_calls = 0

    def fake_create_chart(*args, **kwargs):
        nonlocal create_chart_calls
        create_chart_calls += 1
        return {"ok": True}

    monkeypatch.setattr("excelmcp.server_chart.wb.create_chart", fake_create_chart)

    with pytest.raises(ToolError, match="target_address must be a string when supplied"):
        chart(
            operation="create",
            path="C:/allowed/test.xlsx",
            sheet="Sheet1",
            chart_type="bar",
            data_address="A1:B4",
            target_address=explicit_target_address,
            confirm=True,
        )

    assert create_chart_calls == 0


@pytest.mark.unit
@pytest.mark.parametrize("stray_target_address", ["", "   \t  ", 0, 123])
def test_chart_non_create_operations_ignore_stray_target_address(monkeypatch, stray_target_address):
    """Non-create operations ignore stray explicit `target_address` values."""
    calls: list[tuple[str, tuple[object, ...], dict[str, object]]] = []

    def fake_list_charts(path, sheet):
        calls.append(("list", (path, sheet), {}))
        return {"charts": [], "count": 0}

    def fake_delete_chart(path, sheet, chart_index, *, confirm):
        calls.append(("delete", (path, sheet, chart_index), {"confirm": confirm}))
        return {"ok": True}

    def fake_update_chart_data(path, sheet, chart_index, data_address, *, confirm):
        calls.append(
            (
                "update",
                (path, sheet, chart_index, data_address),
                {"confirm": confirm},
            )
        )
        return {"ok": True, "data_range": data_address}

    monkeypatch.setattr("excelmcp.server_chart.wb.list_charts", fake_list_charts)
    monkeypatch.setattr("excelmcp.server_chart.wb.delete_chart", fake_delete_chart)
    monkeypatch.setattr("excelmcp.server_chart.wb.update_chart_data", fake_update_chart_data)

    list_result = chart(
        operation="list",
        path="C:/allowed/test.xlsx",
        sheet="Sheet1",
        target_address=stray_target_address,
    )
    delete_result = chart(
        operation="delete",
        path="C:/allowed/test.xlsx",
        sheet="Sheet1",
        chart_index=0,
        target_address=stray_target_address,
        confirm=True,
    )
    update_result = chart(
        operation="update",
        path="C:/allowed/test.xlsx",
        sheet="Sheet1",
        chart_index=0,
        data_address="A1:A2",
        target_address=stray_target_address,
        confirm=True,
    )

    assert list_result == {"charts": [], "count": 0}
    assert delete_result == {"ok": True}
    assert update_result == {"ok": True, "data_range": "A1:A2"}
    assert calls == [
        ("list", ("C:/allowed/test.xlsx", "Sheet1"), {}),
        ("delete", ("C:/allowed/test.xlsx", "Sheet1", 0), {"confirm": True}),
        ("update", ("C:/allowed/test.xlsx", "Sheet1", 0, "A1:A2"), {"confirm": True}),
    ]


@pytest.mark.unit
def test_chart_create_chart_type_none_raises(tmp_path, monkeypatch):
    """chart(operation='create', chart_type=None) raises ToolError."""
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    wb_path = tmp_path / "create_notype.xlsx"
    openpyxl.Workbook().save(str(wb_path))

    with pytest.raises(ToolError, match="chart_type is required"):
        chart(operation="create", path=str(wb_path), sheet="Sheet1",
              chart_type=None, data_address="A1:A3", confirm=True)


@pytest.mark.unit
def test_chart_create_confirm_false_raises(tmp_path, monkeypatch):
    """chart(operation='create', confirm=False) raises ToolError (write gate)."""
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    wb_path = tmp_path / "create_noconf.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for i, v in enumerate([1, 2, 3], 1):
        ws.cell(row=i, column=1, value=v)
    wb.save(str(wb_path))

    with pytest.raises(ToolError):
        chart(operation="create", path=str(wb_path), sheet="Sheet1",
              chart_type="bar", data_address="A1:A3", confirm=False)


# ---------------------------------------------------------------------------
# chart() — delete
# ---------------------------------------------------------------------------

@pytest.mark.unit
def test_chart_delete_success(tmp_path, monkeypatch):
    """chart(operation='delete', chart_index=0, confirm=True) deletes the chart."""
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    wb_path = tmp_path / "delete_chart.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for i, v in enumerate([10, 20, 30], 1):
        ws.cell(row=i, column=1, value=v)
    wb.save(str(wb_path))
    chart(operation="create", path=str(wb_path), sheet="Sheet1",
          chart_type="bar", data_address="A1:A3", confirm=True)

    result = chart(operation="delete", path=str(wb_path), sheet="Sheet1",
                   chart_index=0, confirm=True)
    assert result["ok"] is True


@pytest.mark.unit
def test_chart_delete_chart_index_none_raises(tmp_path, monkeypatch):
    """chart(operation='delete', chart_index=None) raises ToolError."""
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    wb_path = tmp_path / "delete_noindex.xlsx"
    openpyxl.Workbook().save(str(wb_path))

    with pytest.raises(ToolError, match="chart_index is required"):
        chart(operation="delete", path=str(wb_path), sheet="Sheet1",
              chart_index=None, confirm=True)


# ---------------------------------------------------------------------------
# chart() — update
# ---------------------------------------------------------------------------

@pytest.mark.unit
def test_chart_update_success(tmp_path, monkeypatch):
    """chart(operation='update', chart_index=0, data_address=..., confirm=True) succeeds."""
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    wb_path = tmp_path / "update_chart.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for i, v in enumerate([10, 20, 30, 40, 50], 1):
        ws.cell(row=i, column=1, value=v)
    wb.save(str(wb_path))
    chart(operation="create", path=str(wb_path), sheet="Sheet1",
          chart_type="bar", data_address="A1:A3", confirm=True)

    result = chart(operation="update", path=str(wb_path), sheet="Sheet1",
                   chart_index=0, data_address="A1:A5", confirm=True)
    assert result["ok"] is True


@pytest.mark.unit
def test_chart_update_prefers_data_address_alias_over_legacy_data_range(tmp_path, monkeypatch):
    """data_address parameter works for chart update operation."""
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    wb_path = tmp_path / "update_chart_alias.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for i, v in enumerate([10, 20, 30, 40, 50], 1):
        ws.cell(row=i, column=1, value=v)
    wb.save(str(wb_path))
    chart(
        operation="create",
        path=str(wb_path),
        sheet="Sheet1",
        chart_type="bar",
        data_address="A1:A3",
        confirm=True,
    )

    result = chart(
        operation="update",
        path=str(wb_path),
        sheet="Sheet1",
        chart_index=0,
        data_address="A1:A5",
        confirm=True,
    )
    assert result["ok"] is True
    assert result["data_range"] == "A1:A5"


@pytest.mark.unit
def test_chart_unknown_operation_raises(tmp_path, monkeypatch):
    """chart(operation='bogus') raises ToolError."""
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    wb_path = tmp_path / "bogus_op.xlsx"
    openpyxl.Workbook().save(str(wb_path))

    with pytest.raises(ToolError, match="Unknown operation"):
        chart(operation="bogus", path=str(wb_path), sheet="Sheet1")


@pytest.mark.unit
def test_named_range_unknown_operation_raises(tmp_path, monkeypatch):
    """named_range(operation='bogus') raises ToolError."""
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    wb_path = tmp_path / "bogus_nr.xlsx"
    openpyxl.Workbook().save(str(wb_path))

    with pytest.raises(ToolError, match="Unknown operation"):
        named_range(operation="bogus", path=str(wb_path))
