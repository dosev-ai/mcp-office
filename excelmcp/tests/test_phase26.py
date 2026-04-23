import pytest
import openpyxl

from excelmcp.workbook_openpyxl import (
    ExcelMCPError,
    ValidationError,
    NotAllowedError,
    create_table,
    get_data_validation_info,
    get_cell_comment,
    validate_formula_syntax,
    _wb_cache,
)


class TestCreateTable:
    """Unit tests for create_table() — native Excel table (ListObject) support."""

    @pytest.mark.unit
    def test_create_table_success(self, sample_workbook, monkeypatch):
        """create_table creates a native Excel table and returns expected metadata."""
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        # populate 5 rows of data so the table range is valid
        _wb = openpyxl.load_workbook(sample_workbook)
        ws = _wb.active
        ws["A1"] = "Name"
        ws["B1"] = "Value"
        ws["A2"] = "Alpha"
        ws["B2"] = 1
        ws["A3"] = "Beta"
        ws["B3"] = 2
        ws["A4"] = "Gamma"
        ws["B4"] = 3
        ws["A5"] = "Delta"
        ws["B5"] = 4
        _wb.save(sample_workbook)

        result = create_table(sample_workbook, "Sheet1", "A1:B5", "SalesTable", confirm=True)
        assert result["table_name"] == "SalesTable"
        assert result["ref"] == "A1:B5"
        assert result["sheet"] == "Sheet1"
        assert result["style"] == "TableStyleMedium9"

        # verify persisted to disk
        _wb_cache.clear()
        _wb2 = openpyxl.load_workbook(sample_workbook)
        table_names = [t.displayName for t in _wb2["Sheet1"].tables.values()]
        assert "SalesTable" in table_names

    @pytest.mark.unit
    def test_create_table_write_disabled(self, sample_workbook, monkeypatch):
        """create_table raises NotAllowedError when EXCEL_ENABLE_WRITE is not set."""
        monkeypatch.delenv("EXCEL_ENABLE_WRITE", raising=False)
        with pytest.raises(NotAllowedError, match="EXCEL_ENABLE_WRITE"):
            create_table(sample_workbook, "Sheet1", "A1:B5", "TestTable", confirm=True)

    @pytest.mark.unit
    def test_create_table_requires_confirm(self, sample_workbook, monkeypatch):
        """create_table raises ValidationError when confirm=False."""
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        with pytest.raises(ValidationError, match="confirm"):
            create_table(sample_workbook, "Sheet1", "A1:B5", "TestTable", confirm=False)

    @pytest.mark.unit
    def test_create_table_invalid_sheet(self, sample_workbook, monkeypatch):
        """Non-existent sheet name raises ValidationError."""
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        with pytest.raises(ValidationError, match="[Ss]heet"):
            create_table(sample_workbook, "NoSuchSheet", "A1:D5", "MyTable", confirm=True)

    @pytest.mark.unit
    def test_create_table_invalid_table_name(self, sample_workbook, monkeypatch):
        """Table name with spaces raises ValidationError."""
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        with pytest.raises(ValidationError, match="table_name"):
            create_table(sample_workbook, "Sheet1", "A1:B5", "My Table", confirm=True)

    @pytest.mark.unit
    def test_create_table_duplicate_name(self, tmp_path, monkeypatch):
        """Creating a table whose name already exists in the workbook raises ValidationError."""
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        from openpyxl.worksheet.table import Table as _XlTable

        wb_path = tmp_path / "dup.xlsx"
        _wb = openpyxl.Workbook()
        ws = _wb.active
        ws.title = "Sheet1"
        for r in range(1, 6):
            ws.cell(r, 1).value = f"Row{r}A"
            ws.cell(r, 2).value = r
        existing = _XlTable(displayName="DupTable", ref="A1:B5")
        ws.add_table(existing)
        _wb.save(str(wb_path))

        with pytest.raises(ValidationError, match="already exists"):
            create_table(str(wb_path), "Sheet1", "A1:B5", "DupTable", confirm=True)

    @pytest.mark.unit
    def test_create_table_single_row_range(self, sample_workbook, monkeypatch):
        """A single-row range (A1:D1) raises ValidationError — minimum 2 rows required."""
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        with pytest.raises(ValidationError, match="[Rr]ow"):
            create_table(sample_workbook, "Sheet1", "A1:D1", "OneRow", confirm=True)

    @pytest.mark.unit
    def test_create_table_invalid_range_format(self, sample_workbook, monkeypatch):
        """A malformed range string raises ValidationError."""
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        with pytest.raises(ValidationError, match="table_range"):
            create_table(sample_workbook, "Sheet1", "NOT_RANGE", "GoodName", confirm=True)

    @pytest.mark.unit
    def test_create_table_bad_name_starts_with_digit(self, sample_workbook, monkeypatch):
        """Table name starting with a digit raises ValidationError."""
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        with pytest.raises(ValidationError, match="table_name"):
            create_table(sample_workbook, "Sheet1", "A1:B5", "1Table", confirm=True)

    @pytest.mark.unit
    def test_create_table_path_outside_allowlist(self, tmp_path, monkeypatch):
        """Path outside EXCEL_ALLOWLIST_ROOTS raises ValidationError (allowlist gate)."""
        allowed_dir = tmp_path / "allowed"
        allowed_dir.mkdir()
        outside_dir = tmp_path / "outside"
        outside_dir.mkdir()

        wb_outside = outside_dir / "outside.xlsx"
        _wb = openpyxl.Workbook()
        ws = _wb.active
        ws.title = "Sheet1"
        _wb.save(str(wb_outside))

        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(allowed_dir))
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        with pytest.raises(ValidationError, match="not in allowlist"):
            create_table(str(wb_outside), "Sheet1", "A1:B5", "TestTable", confirm=True)


class TestPhase26:
    """Unit tests for Phase 2.6: get_data_validation_info, get_cell_comment,
    and validate_formula_syntax path-optional fix."""

    # ── get_data_validation_info ──────────────────────────────────────────

    @pytest.mark.unit
    def test_get_data_validation_info_no_rules(self, tmp_path, monkeypatch):
        """Sheet with no DV rules returns count=0 and empty validations list."""
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        wb_path = tmp_path / "no_dv.xlsx"
        wb_obj = openpyxl.Workbook()
        wb_obj.active.title = "Sheet1"
        wb_obj.save(str(wb_path))

        result = get_data_validation_info(str(wb_path), "Sheet1")
        assert result["sheet"] == "Sheet1"
        assert result["count"] == 0
        assert result["validations"] == []

    @pytest.mark.unit
    def test_get_data_validation_info_with_list_rule(self, tmp_path, monkeypatch):
        """Sheet with a list DV rule returns count=1 and correct rule fields."""
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        from openpyxl.worksheet.datavalidation import DataValidation

        wb_path = tmp_path / "dv_test.xlsx"
        wb_obj = openpyxl.Workbook()
        ws = wb_obj.active
        ws.title = "Sheet1"
        dv = DataValidation(type="list", formula1='"Yes,No,Maybe"', allow_blank=True)
        dv.sqref = "B2:B10"
        ws.add_data_validation(dv)
        wb_obj.save(str(wb_path))

        result = get_data_validation_info(str(wb_path), "Sheet1")
        assert result["count"] == 1
        rule = result["validations"][0]
        assert rule["validation_type"] == "list"
        assert rule["formula1"] == '"Yes,No,Maybe"'
        assert rule["allow_blank"] is True
        assert "B2:B10" in rule["address"]

    @pytest.mark.unit
    def test_get_data_validation_info_not_allowed(self, tmp_path, monkeypatch):
        """Path outside EXCEL_ALLOWLIST_ROOTS raises an error."""
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        with pytest.raises((ExcelMCPError, NotAllowedError, ValidationError)):
            get_data_validation_info("C:\\other\\file.xlsx", "Sheet1")

    @pytest.mark.unit
    def test_get_data_validation_info_sheet_not_found(self, tmp_path, monkeypatch):
        """Non-existent sheet name raises ValidationError."""
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        wb_path = tmp_path / "nosheet.xlsx"
        wb_obj = openpyxl.Workbook()
        wb_obj.active.title = "Sheet1"
        wb_obj.save(str(wb_path))

        with pytest.raises((ExcelMCPError, ValidationError)):
            get_data_validation_info(str(wb_path), "NoSuchSheet")

    # ── get_cell_comment ──────────────────────────────────────────────────

    @pytest.mark.unit
    def test_get_cell_comment_no_comment(self, tmp_path, monkeypatch):
        """Cell with no comment returns has_comment=False and None fields."""
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        wb_path = tmp_path / "no_comment.xlsx"
        wb_obj = openpyxl.Workbook()
        wb_obj.active.title = "Sheet1"
        wb_obj.save(str(wb_path))

        result = get_cell_comment(str(wb_path), "Sheet1", "A1")
        assert result["address"] == "A1"
        assert result["has_comment"] is False
        assert result["text"] is None
        assert result["author"] is None

    @pytest.mark.unit
    def test_get_cell_comment_with_comment(self, tmp_path, monkeypatch):
        """Cell with a comment returns has_comment=True and correct text/author."""
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        from openpyxl.comments import Comment

        wb_path = tmp_path / "with_comment.xlsx"
        wb_obj = openpyxl.Workbook()
        ws = wb_obj.active
        ws.title = "Sheet1"
        ws["B3"].comment = Comment("Hello world", "TestAuthor")
        wb_obj.save(str(wb_path))

        result = get_cell_comment(str(wb_path), "Sheet1", "B3")
        assert result["address"] == "B3"
        assert result["has_comment"] is True
        assert result["text"] == "Hello world"
        assert result["author"] == "TestAuthor"

    @pytest.mark.unit
    def test_get_cell_comment_not_allowed(self, tmp_path, monkeypatch):
        """Path outside allowlist raises an error."""
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        with pytest.raises((ExcelMCPError, NotAllowedError, ValidationError)):
            get_cell_comment("C:\\other\\file.xlsx", "Sheet1", "A1")

    @pytest.mark.unit
    def test_get_cell_comment_sheet_not_found(self, tmp_path, monkeypatch):
        """Non-existent sheet name raises ValidationError with 'Sheet not found' message."""
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        wb_path = tmp_path / "nosheet_comment.xlsx"
        wb_obj = openpyxl.Workbook()
        wb_obj.active.title = "Sheet1"
        wb_obj.save(str(wb_path))

        with pytest.raises(ValidationError, match="Sheet not found"):
            get_cell_comment(str(wb_path), "NoSuchSheet", "A1")

    # ── validate_formula_syntax path-optional fix ─────────────────────────

    @pytest.mark.unit
    def test_validate_formula_syntax_path_optional(self, tmp_path, monkeypatch):
        """validate_formula_syntax works when path argument is omitted (defaults to '')."""
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        result = validate_formula_syntax("=SUM(A1:A5)")
        assert result["valid"] is True
        assert result["error"] is None
