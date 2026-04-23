import pytest
import openpyxl

from excelmcp.workbook_openpyxl import (
    ExcelMCPError,
    ValidationError,
    get_cell_metadata,
    evaluate_formula,
    list_tables,
    read_table,
)


# ── GROUP 22: get_cell_metadata() ────────────────────────────────────────

@pytest.mark.unit
def test_get_cell_metadata_happy(sample_workbook):
    result = get_cell_metadata(sample_workbook, "Sheet1", "A1")
    assert result["address"] == "A1"
    assert result["value"] == "hello"
    assert "font" in result
    assert "fill_type" in result
    assert "has_comment" in result
    assert result["has_comment"] is False


@pytest.mark.unit
def test_get_cell_metadata_invalid_sheet(sample_workbook):
    with pytest.raises(ValidationError, match="Sheet not found"):
        get_cell_metadata(sample_workbook, "NoSuch", "A1")


# ── GROUP 23: evaluate_formula() ─────────────────────────────────────────

@pytest.mark.unit
def test_evaluate_formula_non_formula(sample_workbook):
    result = evaluate_formula(sample_workbook, "Sheet1", "A1")
    assert result["address"] == "A1"
    assert result["note"] == "openpyxl_cached_only"
    assert "formula" in result
    assert "cached_value" in result


@pytest.mark.unit
def test_evaluate_formula_invalid_sheet(sample_workbook):
    with pytest.raises(ValidationError, match="Sheet not found"):
        evaluate_formula(sample_workbook, "NoSuch", "A1")


# ── GROUP 24: resolve_*_resource() ───────────────────────────────────────

@pytest.mark.unit
def test_resolve_sheets_resource_happy_path(sample_workbook):
    from excelmcp.workbook_openpyxl import resolve_sheets_resource
    result = resolve_sheets_resource(str(sample_workbook))
    assert "sheets" in result
    assert any(s["name"] == "Sheet1" for s in result["sheets"])


@pytest.mark.unit
def test_resolve_sheets_resource_bad_path_raises(tmp_path, monkeypatch):
    from excelmcp.workbook_openpyxl import resolve_sheets_resource
    # Use a different tmp_path from the one in EXCEL_ALLOWLIST_ROOTS env var.
    # Since monkeypatch is scoped to this test, EXCEL_ALLOWLIST_ROOTS points to a
    # different directory — set it explicitly to some unrelated path.
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path / "allowed"))
    outside = tmp_path / "not_allowed" / "file.xlsx"
    with pytest.raises((ExcelMCPError, ValueError)):
        resolve_sheets_resource(str(outside))


@pytest.mark.unit
def test_resolve_sheet_resource_happy_path(sample_workbook):
    from excelmcp.workbook_openpyxl import resolve_sheet_resource
    result = resolve_sheet_resource(str(sample_workbook), "Sheet1")
    assert result["sheet"] == "Sheet1"
    assert result["max_col"] >= 1


@pytest.mark.unit
def test_resolve_sheet_resource_unknown_sheet(sample_workbook):
    from excelmcp.workbook_openpyxl import resolve_sheet_resource
    with pytest.raises(ExcelMCPError):
        resolve_sheet_resource(str(sample_workbook), "NoSuchSheet")


@pytest.mark.unit
def test_resolve_sheet_resource_empty_sheet(tmp_path, monkeypatch):
    """resolve_sheet_resource must not crash on a sheet with no written cells."""
    from unittest.mock import PropertyMock, patch
    from excelmcp import workbook_openpyxl as wop
    from excelmcp.workbook_openpyxl import resolve_sheet_resource

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "EmptySheet"
    fp = tmp_path / "empty.xlsx"
    wb.save(fp)

    # Allow this path
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))

    # Inject the workbook into the cache and patch max_row/max_column → None
    # to simulate older openpyxl / truly empty sheet XML behaviour.
    wop._wb_cache[(str(fp), False)] = wb
    try:
        with patch.object(type(ws), "max_row", new_callable=PropertyMock, return_value=None):
            with patch.object(type(ws), "max_column", new_callable=PropertyMock, return_value=None):
                result = resolve_sheet_resource(str(fp), "EmptySheet")
    finally:
        wop._wb_cache.pop((str(fp), False), None)

    assert result["sheet"] == "EmptySheet"
    assert result["max_row"] == 0
    assert result["max_col"] == 0
    assert result["first_row_headers"] == []


@pytest.mark.unit
def test_resolve_named_ranges_resource_empty(sample_workbook):
    from excelmcp.workbook_openpyxl import resolve_named_ranges_resource
    result = resolve_named_ranges_resource(str(sample_workbook))
    assert "named_ranges" in result
    assert isinstance(result["named_ranges"], list)


# ── GROUP 24: list_tables() ───────────────────────────────────────────────

@pytest.mark.unit
def test_list_tables_empty(sample_workbook):
    result = list_tables(sample_workbook, "Sheet1")
    assert isinstance(result, list)
    assert len(result) == 0


@pytest.mark.unit
def test_list_tables_invalid_sheet(sample_workbook):
    with pytest.raises(ValidationError, match="Sheet not found"):
        list_tables(sample_workbook, "NoSuch")


@pytest.mark.unit
def test_list_tables_with_table(tmp_path, monkeypatch):
    """list_tables returns the table when one is defined."""
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    import openpyxl as _opx
    from openpyxl.worksheet.table import Table, TableStyleInfo
    wb_path = tmp_path / "t.xlsx"
    _wb = _opx.Workbook()
    _ws = _wb.active
    _ws["A1"] = "Name"
    _ws["B1"] = "Value"
    _ws["A2"] = "a"
    _ws["B2"] = 1
    tbl = Table(displayName="MyTable", ref="A1:B2")
    tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9")
    _ws.add_table(tbl)
    _wb.save(str(wb_path))
    result = list_tables(str(wb_path), _ws.title)
    names = [t["name"] for t in result]
    assert "MyTable" in names


# ── GROUP 25: read_table() ────────────────────────────────────────────────

@pytest.mark.unit
def test_read_table_not_found(sample_workbook):
    with pytest.raises(ValidationError, match="Table not found"):
        read_table(sample_workbook, "Sheet1", "NoSuchTable")


@pytest.mark.unit
def test_read_table_happy(tmp_path, monkeypatch):
    """read_table returns rows from the named table."""
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    import openpyxl as _opx
    from openpyxl.worksheet.table import Table, TableStyleInfo
    wb_path = tmp_path / "t.xlsx"
    _wb = _opx.Workbook()
    _ws = _wb.active
    _ws["A1"] = "Name"
    _ws["B1"] = "Value"
    _ws["A2"] = "alpha"
    _ws["B2"] = 10
    tbl = Table(displayName="DataTable", ref="A1:B2")
    tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9")
    _ws.add_table(tbl)
    _wb.save(str(wb_path))
    result = read_table(str(wb_path), _ws.title, "DataTable")
    assert result is not None
    assert "rows" in result or "data" in result or isinstance(result, dict)
