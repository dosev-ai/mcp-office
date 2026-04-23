import pytest
import openpyxl

from excelmcp.workbook_openpyxl import (
    ValidationError,
    NotAllowedError,
    OfficeCOMError,
    find_replace,
    get_named_ranges,
    read_named_range,
    add_sheet,
    rename_sheet,
    delete_sheet,
    copy_sheet,
    delete_rows,
    delete_cols,
    import_csv_to_sheet,
    list_tables,
    _wb_cache,
)


# ── GROUP 12: find_replace() ──────────────────────────────────────────────

@pytest.mark.unit
def test_find_replace_whole_cell(sample_workbook, monkeypatch):
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    result = find_replace(
        sample_workbook, "Sheet1", "hello", "world",
        whole_cell=True, confirm=True
    )
    assert result["ok"] is True
    assert result["replacements_made"] == 1
    assert "A1" in result["cells_modified"]


@pytest.mark.unit
def test_find_replace_no_match(sample_workbook, monkeypatch):
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    result = find_replace(
        sample_workbook, "Sheet1", "zzz", "nope",
        whole_cell=True, confirm=True
    )
    assert result["replacements_made"] == 0


@pytest.mark.unit
def test_find_replace_requires_write(sample_workbook, monkeypatch):
    monkeypatch.delenv("EXCEL_ENABLE_WRITE", raising=False)
    with pytest.raises(NotAllowedError):
        find_replace(sample_workbook, "Sheet1", "hello", "x", confirm=True)


@pytest.mark.unit
def test_find_replace_requires_confirm(sample_workbook, monkeypatch):
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    with pytest.raises(ValidationError):
        find_replace(sample_workbook, "Sheet1", "hello", "x", confirm=False)


@pytest.mark.unit
def test_find_replace_substring(sample_workbook, monkeypatch):
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    result = find_replace(
        sample_workbook, "Sheet1", "ell", "ELL",
        whole_cell=False, match_case=True, confirm=True
    )
    assert result["replacements_made"] == 1
    assert "A1" in result["cells_modified"]


@pytest.mark.unit
def test_find_replace_rejects_formula_prefix_replace_value(sample_workbook, monkeypatch):
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    with pytest.raises(ValidationError):
        find_replace(sample_workbook, "Sheet1", "hello", "=DANGER()", confirm=True)


@pytest.mark.unit
def test_find_replace_allows_negative_replace_value_r006(sample_workbook, monkeypatch):
    """R-006 regression — negative numbers must NOT be blocked as formula injection."""
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    result = find_replace(sample_workbook, "Sheet1", "hello", "-500", confirm=True)
    assert result.get("ok") is True  # "-500" is not a formula

    result2 = find_replace(sample_workbook, "Sheet1", "-500", "+7", confirm=True)
    assert result2.get("ok") is True  # "+7" parses as float, not a formula

    # Non-numeric "-" prefix still blocked:
    with pytest.raises(ValidationError):
        find_replace(sample_workbook, "Sheet1", "+7", "-cmd|calc", confirm=True)


# ── GROUP 13: get_named_ranges() ──────────────────────────────────────────

@pytest.mark.unit
def test_get_named_ranges_empty(sample_workbook):
    # A fresh workbook has no named ranges
    result = get_named_ranges(sample_workbook)
    assert isinstance(result, list)


@pytest.mark.unit
def test_get_named_ranges_with_range(tmp_path, monkeypatch):
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    path = tmp_path / "named.xlsx"
    from openpyxl.workbook.defined_name import DefinedName
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "data"
    dn = DefinedName("MyRange", attr_text="Sheet1!$A$1")
    wb.defined_names.add(dn)
    wb.save(str(path))
    result = get_named_ranges(str(path))
    assert any(r["name"] == "MyRange" for r in result)


# ── GROUP 14: read_named_range() ──────────────────────────────────────────

@pytest.mark.unit
def test_read_named_range_not_found(sample_workbook):
    with pytest.raises(ValidationError, match="Named range not found"):
        read_named_range(sample_workbook, "DoesNotExist")


# ── GROUP 15: add_sheet() ─────────────────────────────────────────────────

@pytest.mark.unit
def test_add_sheet_happy(sample_workbook, monkeypatch):
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    result = add_sheet(sample_workbook, "NewSheet", confirm=True)
    assert result["ok"] is True
    assert result["sheet_name"] == "NewSheet"


@pytest.mark.unit
def test_add_sheet_duplicate(sample_workbook, monkeypatch):
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    with pytest.raises(ValidationError, match="already exists"):
        add_sheet(sample_workbook, "Sheet1", confirm=True)


@pytest.mark.unit
def test_add_sheet_requires_write(sample_workbook, monkeypatch):
    monkeypatch.delenv("EXCEL_ENABLE_WRITE", raising=False)
    with pytest.raises(NotAllowedError):
        add_sheet(sample_workbook, "X", confirm=True)


# ── GROUP 16: rename_sheet() ──────────────────────────────────────────────

@pytest.mark.unit
def test_rename_sheet_happy(sample_workbook, monkeypatch):
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    result = rename_sheet(sample_workbook, "Sheet1", "Renamed", confirm=True)
    assert result["ok"] is True
    assert result["new_name"] == "Renamed"


@pytest.mark.unit
def test_rename_sheet_not_found(sample_workbook, monkeypatch):
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    with pytest.raises(ValidationError, match="Sheet not found"):
        rename_sheet(sample_workbook, "NoSuch", "NewName", confirm=True)


@pytest.mark.unit
def test_rename_sheet_requires_write(sample_workbook, monkeypatch):
    monkeypatch.delenv("EXCEL_ENABLE_WRITE", raising=False)
    with pytest.raises(NotAllowedError):
        rename_sheet(sample_workbook, "Sheet1", "X", confirm=True)


# ── GROUP 17: delete_sheet() ──────────────────────────────────────────────

@pytest.mark.unit
def test_delete_sheet_only_sheet(sample_workbook, monkeypatch):
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    with pytest.raises(ValidationError, match="only sheet"):
        delete_sheet(sample_workbook, "Sheet1", confirm=True)


@pytest.mark.unit
def test_delete_sheet_happy(sample_workbook, monkeypatch):
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    add_sheet(sample_workbook, "Extra", confirm=True)
    _wb_cache.clear()
    result = delete_sheet(sample_workbook, "Extra", confirm=True)
    assert result["ok"] is True


@pytest.mark.unit
def test_delete_sheet_not_found(sample_workbook, monkeypatch):
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    with pytest.raises(ValidationError, match="Sheet not found"):
        delete_sheet(sample_workbook, "NoSuch", confirm=True)


# ── GROUP 18: copy_sheet() ────────────────────────────────────────────────

@pytest.mark.unit
def test_copy_sheet_happy(sample_workbook, monkeypatch):
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    result = copy_sheet(sample_workbook, "Sheet1", "Sheet1_copy", confirm=True)
    assert result["ok"] is True
    assert result["new_name"] == "Sheet1_copy"


@pytest.mark.unit
def test_copy_sheet_source_not_found(sample_workbook, monkeypatch):
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    with pytest.raises(ValidationError, match="Source sheet not found"):
        copy_sheet(sample_workbook, "NoSuch", "Copy", confirm=True)


@pytest.mark.unit
def test_copy_sheet_name_exists(sample_workbook, monkeypatch):
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    with pytest.raises(ValidationError, match="already exists"):
        copy_sheet(sample_workbook, "Sheet1", "Sheet1", confirm=True)


# ── GROUP 19: delete_rows() ───────────────────────────────────────────────

@pytest.mark.unit
def test_delete_rows_happy(sample_workbook, monkeypatch):
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    result = delete_rows(sample_workbook, "Sheet1", 1, confirm=True)
    assert result["ok"] is True
    assert result["rows_deleted"] == 1


@pytest.mark.unit
def test_delete_rows_range(sample_workbook, monkeypatch):
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    result = delete_rows(sample_workbook, "Sheet1", 1, end_row=2, confirm=True)
    assert result["rows_deleted"] == 2


@pytest.mark.unit
def test_delete_rows_invalid_start(sample_workbook, monkeypatch):
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    with pytest.raises(ValidationError, match="start_row"):
        delete_rows(sample_workbook, "Sheet1", 0, confirm=True)


# ── GROUP 20: delete_cols() ───────────────────────────────────────────────

@pytest.mark.unit
def test_delete_cols_happy(sample_workbook, monkeypatch):
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    result = delete_cols(sample_workbook, "Sheet1", 1, confirm=True)
    assert result["ok"] is True
    assert result["cols_deleted"] == 1


@pytest.mark.unit
def test_delete_cols_invalid_start(sample_workbook, monkeypatch):
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    with pytest.raises(ValidationError, match="start_col"):
        delete_cols(sample_workbook, "Sheet1", 0, confirm=True)


# ── GROUP 21: import_csv_to_sheet() ──────────────────────────────────────

@pytest.mark.unit
def test_import_csv_happy(tmp_path, monkeypatch):
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    # Create xlsx
    wb_path = tmp_path / "test.xlsx"
    import openpyxl as _opx
    temp_wb = _opx.Workbook()
    temp_wb.save(str(wb_path))
    # Create csv (also in tmp_path so it's in allowlist)
    csv_path = tmp_path / "data.csv"
    csv_path.write_text("col1,col2\n1,2\n3,4\n")
    result = import_csv_to_sheet(str(wb_path), str(csv_path), "Imported", confirm=True)
    assert result["ok"] is True
    assert result["rows_imported"] == 3  # header + 2 data rows
    assert result["sheet_name"] == "Imported"


@pytest.mark.unit
def test_import_csv_sheet_exists_no_overwrite(tmp_path, monkeypatch):
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    wb_path = tmp_path / "test.xlsx"
    import openpyxl as _opx
    temp_wb = _opx.Workbook()
    temp_wb.save(str(wb_path))
    csv_path = tmp_path / "data.csv"
    csv_path.write_text("a\n1\n")
    # Import once to a NEW sheet name (default Workbook already has "Sheet")
    import_csv_to_sheet(str(wb_path), str(csv_path), "Imported", confirm=True)
    _wb_cache.clear()
    # Import again to the same name without overwrite=True → should fail
    with pytest.raises(ValidationError, match="already exists"):
        import_csv_to_sheet(str(wb_path), str(csv_path), "Imported", confirm=True)


@pytest.mark.unit
def test_import_csv_bad_extension(tmp_path, monkeypatch):
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    wb_path = tmp_path / "test.xlsx"
    import openpyxl as _opx
    temp_wb = _opx.Workbook()
    temp_wb.save(str(wb_path))
    bad_path = tmp_path / "data.txt"
    bad_path.write_text("a,b\n")
    with pytest.raises(ValidationError, match=".csv or .tsv"):
        import_csv_to_sheet(str(wb_path), str(bad_path), "Sheet", confirm=True)


# ── Group 26: coverage gaps ─────────────────────────────────────────────────

@pytest.mark.unit
def test_find_replace_match_case(tmp_path, monkeypatch):
    """match_case=True should NOT replace 'Hello' when searching for 'hello'."""
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    wb_path = tmp_path / "t.xlsx"
    import openpyxl as _opx
    _wb = _opx.Workbook()
    _ws = _wb.active
    _ws["A1"] = "Hello"
    _ws["A2"] = "hello"
    _wb.save(str(wb_path))
    result = find_replace(str(wb_path), _ws.title, "hello", "X", match_case=True, confirm=True)
    _wb_cache.clear()
    wb2 = _opx.load_workbook(str(wb_path))
    assert wb2.active["A1"].value == "Hello"   # unchanged — wrong case
    assert wb2.active["A2"].value == "X"        # replaced
    assert result["replacements_made"] == 1


@pytest.mark.unit
def test_read_named_range_happy(tmp_path, monkeypatch):
    """Happy path: resolve a named range to cell data."""
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    import openpyxl as _opx
    from openpyxl.workbook.defined_name import DefinedName
    wb_path = tmp_path / "nr.xlsx"
    _wb = _opx.Workbook()
    _ws = _wb.active
    _ws["A1"] = 42
    dn = DefinedName("MyRange", attr_text=f"'{_ws.title}'!$A$1")
    _wb.defined_names["MyRange"] = dn
    _wb.save(str(wb_path))
    result = read_named_range(str(wb_path), "MyRange")
    assert result is not None


@pytest.mark.unit
def test_add_sheet_at_position(tmp_path, monkeypatch):
    """add_sheet with position=0 inserts at front."""
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    import openpyxl as _opx
    wb_path = tmp_path / "t.xlsx"
    _wb = _opx.Workbook()
    _wb.save(str(wb_path))
    add_sheet(str(wb_path), "NewFirst", position=0, confirm=True)
    _wb_cache.clear()
    wb2 = _opx.load_workbook(str(wb_path))
    assert "NewFirst" in wb2.sheetnames
    assert wb2.sheetnames[0] == "NewFirst"


@pytest.mark.unit
def test_copy_sheet_write_gate(tmp_path, monkeypatch):
    """copy_sheet blocked if EXCEL_ENABLE_WRITE not set."""
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    monkeypatch.delenv("EXCEL_ENABLE_WRITE", raising=False)
    import openpyxl as _opx
    wb_path = tmp_path / "t.xlsx"
    _opx.Workbook().save(str(wb_path))
    with pytest.raises(NotAllowedError):
        copy_sheet(str(wb_path), "Sheet", "Copy", confirm=True)


@pytest.mark.unit
def test_delete_cols_write_gate(tmp_path, monkeypatch):
    """delete_cols blocked if EXCEL_ENABLE_WRITE not set."""
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    monkeypatch.delenv("EXCEL_ENABLE_WRITE", raising=False)
    import openpyxl as _opx
    wb_path = tmp_path / "t.xlsx"
    _opx.Workbook().save(str(wb_path))
    with pytest.raises(NotAllowedError):
        delete_cols(str(wb_path), "Sheet", 1, confirm=True)


@pytest.mark.unit
def test_import_csv_overwrite_true(tmp_path, monkeypatch):
    """import_csv_to_sheet with overwrite=True succeeds when sheet already exists."""
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    import openpyxl as _opx
    wb_path = tmp_path / "t.xlsx"
    _wb = _opx.Workbook()
    _wb.save(str(wb_path))
    csv_path = tmp_path / "d.csv"
    csv_path.write_text("a,b\n1,2\n")
    import_csv_to_sheet(str(wb_path), str(csv_path), "Imported", confirm=True)
    _wb_cache.clear()
    import_csv_to_sheet(str(wb_path), str(csv_path), "Imported", overwrite=True, confirm=True)
    _wb_cache.clear()
    wb2 = _opx.load_workbook(str(wb_path))
    assert "Imported" in wb2.sheetnames


@pytest.mark.unit
def test_import_csv_header_written_by_default(tmp_path, monkeypatch):
    """GAP-03 fix: import_csv_to_sheet writes the header row to the sheet by default."""
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    import openpyxl as _opx
    wb_path = tmp_path / "test.xlsx"
    _opx.Workbook().save(str(wb_path))
    csv_path = tmp_path / "data.csv"
    csv_path.write_text("Name,Value\nalpha,1\nbeta,2\n")
    result = import_csv_to_sheet(str(wb_path), str(csv_path), "Sheet1", confirm=True)
    assert result["ok"] is True
    assert result["rows_imported"] == 3  # all 3 rows (header + 2 data) written
    wb_disk = _opx.load_workbook(str(wb_path))
    ws = wb_disk["Sheet1"]
    assert ws["A1"].value == "Name"  # header must be row 1
    assert ws["A2"].value == "alpha"
    assert ws.max_row == 3


@pytest.mark.unit
def test_import_csv_skip_header_true(tmp_path, monkeypatch):
    """GAP-03 fix: skip_header=True discards only the first row (explicit opt-in)."""
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    import openpyxl as _opx
    wb_path = tmp_path / "test.xlsx"
    _opx.Workbook().save(str(wb_path))
    csv_path = tmp_path / "data.csv"
    csv_path.write_text("Name,Value\nalpha,1\nbeta,2\n")
    result = import_csv_to_sheet(
        str(wb_path), str(csv_path), "Sheet1", skip_header=True, confirm=True
    )
    assert result["rows_imported"] == 2  # only data rows counted
    wb_disk = _opx.load_workbook(str(wb_path))
    ws = wb_disk["Sheet1"]
    assert ws["A1"].value == "alpha"  # first data row at row 1
    assert ws.max_row == 2


@pytest.mark.unit
def test_check_csv_path_no_allowlist_raises(tmp_path, monkeypatch):
    """Security fix: _check_csv_path raises when EXCEL_ALLOWLIST_ROOTS is unset."""
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", "")
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    import openpyxl as _opx
    wb_path = tmp_path / "test.xlsx"
    _opx.Workbook().save(str(wb_path))
    csv_path = tmp_path / "data.csv"
    csv_path.write_text("a,b\n1,2\n")
    with pytest.raises(ValidationError, match="No allowlist configured"):
        import_csv_to_sheet(str(wb_path), str(csv_path), "Sheet", confirm=True)


# ── GROUP 22: LC-2 conflict guard in _sheet.py write paths ─────────────────

@pytest.mark.unit
def test_add_sheet_lc2_guard_raises_office_com_error(sample_workbook, monkeypatch):
    """LC-2-S1: OfficeCOMError raised by the LC-2 guard propagates from add_sheet.

    Mirrors C-06 (write_range in _io.py).  Verifies that the guard wired at
    _sheet.py:148 is not silently swallowed.
    """
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")

    def _raise_com(path):
        raise OfficeCOMError("Could not verify file lock state via COM")

    monkeypatch.setattr("excelmcp._sheet._check_file_not_open_in_excel", _raise_com)

    with pytest.raises(OfficeCOMError):
        add_sheet(sample_workbook, "NewSheet", confirm=True)


@pytest.mark.unit
def test_rename_sheet_lc2_guard_raises_validation_error(sample_workbook, monkeypatch):
    """LC-2-S2: ValidationError raised by the LC-2 guard propagates from rename_sheet.

    Mirrors C-05 (write_cell in _io.py).  Verifies that the guard wired at
    _sheet.py:175 is not silently swallowed.
    """
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")

    def _raise_val(path):
        raise ValidationError("Workbook is open in Excel. Close it first.")

    monkeypatch.setattr("excelmcp._sheet._check_file_not_open_in_excel", _raise_val)

    with pytest.raises(ValidationError):
        rename_sheet(sample_workbook, "Sheet1", "Renamed", confirm=True)


@pytest.mark.unit
def test_list_tables_workbook_wide(tmp_path, monkeypatch):
    """GAP-07 fix: list_tables without sheet= returns tables across all sheets."""
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    import openpyxl as _opx
    from openpyxl.worksheet.table import Table, TableStyleInfo
    wb_path = tmp_path / "t.xlsx"
    _wb = _opx.Workbook()
    _ws1 = _wb.active
    _ws1.title = "Sheet1"
    _ws1["A1"] = "X"
    _ws1["A2"] = 1
    tbl1 = Table(displayName="T1", ref="A1:A2")
    tbl1.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9")
    _ws1.add_table(tbl1)
    _ws2 = _wb.create_sheet("Sheet2")
    _ws2["A1"] = "Y"
    _ws2["A2"] = 2
    tbl2 = Table(displayName="T2", ref="A1:A2")
    tbl2.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9")
    _ws2.add_table(tbl2)
    _wb.save(str(wb_path))
    # Call without sheet= should return both tables
    result = list_tables(str(wb_path))
    names = {t["name"] for t in result}
    sheets = {t["sheet"] for t in result}
    assert "T1" in names
    assert "T2" in names
    assert "Sheet1" in sheets
    assert "Sheet2" in sheets
