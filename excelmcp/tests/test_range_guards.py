import pytest
from excelmcp.workbook_openpyxl import (
    merge_cells,
    unmerge_cells,
    auto_filter,
    set_print_area,
    ValidationError
)

@pytest.fixture
def sample_xlsx(tmp_path):
    import openpyxl
    path = tmp_path / "test.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    wb.save(str(path))
    return str(path)

@pytest.mark.unit
def test_merge_cells_range_guard(sample_xlsx, monkeypatch):
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    monkeypatch.setenv("EXCEL_MAX_RANGE_CELLS", "10")
    # A1:B10 is 20 cells > 10
    with pytest.raises(ValidationError, match="Range too large"):
        merge_cells(sample_xlsx, "Sheet1", "A1:B10", confirm=True)

@pytest.mark.unit
def test_unmerge_cells_range_guard(sample_xlsx, monkeypatch):
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    monkeypatch.setenv("EXCEL_MAX_RANGE_CELLS", "10")
    with pytest.raises(ValidationError, match="Range too large"):
        unmerge_cells(sample_xlsx, "Sheet1", "A1:B10", confirm=True)

@pytest.mark.unit
def test_auto_filter_range_guard(sample_xlsx, monkeypatch):
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    monkeypatch.setenv("EXCEL_MAX_RANGE_CELLS", "10")
    with pytest.raises(ValidationError, match="Range too large"):
        auto_filter(sample_xlsx, "Sheet1", "A1:B10", confirm=True)

@pytest.mark.unit
def test_set_print_area_range_guard(sample_xlsx, monkeypatch):
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    monkeypatch.setenv("EXCEL_MAX_RANGE_CELLS", "10")
    with pytest.raises(ValidationError, match="Range too large"):
        set_print_area(sample_xlsx, "Sheet1", "A1:B10", confirm=True)
