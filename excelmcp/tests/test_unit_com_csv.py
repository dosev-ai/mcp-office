"""Unit tests for export_range_as_csv + CSV helper functions.

Sprint W3+W4 — Excel MCP Phase 2 COM.

Classes:
    TestSanitizeCsvValue        — _sanitize_csv_value() pure-function tests (4)
    TestCheckCsvExportPath      — _check_csv_export_path() security tests (2)
    TestExportRangeAsCsvOpenpyxl — export_range_as_csv live=False path (8)
    TestExportRangeAsCsvSizeCap  — 5 MB inline cap (2)

All tests run without live Excel (no COM, no EXCEL_ENABLE_COM required for
live=False path). CI-safe.
"""
from __future__ import annotations

from unittest.mock import patch

import openpyxl
import pytest

from excelmcp._core import ValidationError


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _make_workbook(tmp_path, sheet_name: str = "Sheet1", data: list[list] | None = None):
    """Create a minimal workbook at tmp_path/test.xlsx and return its str path."""
    path = tmp_path / "test.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    if data:
        for row in data:
            ws.append(row)
    wb.save(str(path))
    return str(path)


# ---------------------------------------------------------------------------
# Class 1: TestSanitizeCsvValue
# ---------------------------------------------------------------------------

class TestSanitizeCsvValue:
    """Tests for _sanitize_csv_value() in excelmcp._com."""

    @pytest.mark.unit
    def test_sanitize_csv_value_safe(self):
        """Plain strings are returned unchanged."""
        from excelmcp._com import _sanitize_csv_value
        assert _sanitize_csv_value("hello") == "hello"
        assert _sanitize_csv_value("world 123") == "world 123"
        assert _sanitize_csv_value("") == ""

    @pytest.mark.unit
    def test_sanitize_csv_value_formula_prefix(self):
        """Values starting with '=' are prefixed with a single quote."""
        from excelmcp._com import _sanitize_csv_value
        assert _sanitize_csv_value("=SUM(A1:B1)") == "'=SUM(A1:B1)"
        assert _sanitize_csv_value("=1+1") == "'=1+1"

    @pytest.mark.unit
    def test_sanitize_csv_value_plus_minus_at_prefix(self):
        """Values starting with +, -, or @ are also prefixed."""
        from excelmcp._com import _sanitize_csv_value
        assert _sanitize_csv_value("+44207000000") == "'+44207000000"
        assert _sanitize_csv_value("-5") == "'-5"
        assert _sanitize_csv_value("@user") == "'@user"

    @pytest.mark.unit
    def test_sanitize_csv_value_none(self):
        """None is coerced to empty string."""
        from excelmcp._com import _sanitize_csv_value
        assert _sanitize_csv_value(None) == ""

    @pytest.mark.unit
    def test_sanitize_csv_value_numeric(self):
        """Numeric values (int, float) are stringified without prefix."""
        from excelmcp._com import _sanitize_csv_value
        assert _sanitize_csv_value(42) == "42"
        assert _sanitize_csv_value(3.14) == "3.14"
        # Negative numbers as int/float: str("-5") → "-5" starts with "-" → prefixed
        # But int -5 → str is "-5" which starts with "-" → prefix applied
        result = _sanitize_csv_value(-5)
        assert result == "'-5"


# ---------------------------------------------------------------------------
# Class 2: TestCheckCsvExportPath
# ---------------------------------------------------------------------------

class TestCheckCsvExportPath:
    """Tests for _check_csv_export_path() in excelmcp._com."""

    @pytest.mark.unit
    def test_bad_suffix_raises_validation_error(self, monkeypatch, tmp_path):
        """output_path with non-.csv extension raises ValidationError."""
        monkeypatch.setenv("EXCEL_CSV_EXPORT_ROOTS", str(tmp_path))
        from excelmcp._com import _check_csv_export_path
        with pytest.raises(ValidationError, match=r"\.csv|suffix"):
            _check_csv_export_path(str(tmp_path / "out.txt"))

    @pytest.mark.unit
    def test_outside_allowlist_raises_validation_error(self, monkeypatch, tmp_path):
        """output_path outside EXCEL_CSV_EXPORT_ROOTS raises ValidationError."""
        allowed = tmp_path / "allowed"
        allowed.mkdir()
        monkeypatch.setenv("EXCEL_CSV_EXPORT_ROOTS", str(allowed))
        from excelmcp._com import _check_csv_export_path
        with pytest.raises(ValidationError):
            _check_csv_export_path(str(tmp_path / "other" / "out.csv"))

    @pytest.mark.unit
    def test_valid_path_returns_resolved_path(self, monkeypatch, tmp_path):
        """Valid path inside allowlist is returned as resolved Path."""
        monkeypatch.setenv("EXCEL_CSV_EXPORT_ROOTS", str(tmp_path))
        from excelmcp._com import _check_csv_export_path
        from pathlib import Path
        result = _check_csv_export_path(str(tmp_path / "out.csv"))
        assert isinstance(result, Path)
        assert result.suffix == ".csv"

    @pytest.mark.unit
    def test_unset_roots_raises_validation_error(self, monkeypatch, tmp_path):
        """Unset EXCEL_CSV_EXPORT_ROOTS raises ValidationError."""
        monkeypatch.delenv("EXCEL_CSV_EXPORT_ROOTS", raising=False)
        from excelmcp._com import _check_csv_export_path
        with pytest.raises(ValidationError, match=r"EXCEL_CSV_EXPORT_ROOTS"):
            _check_csv_export_path(str(tmp_path / "out.csv"))

    @pytest.mark.unit
    def test_relative_path_raises_validation_error(self, monkeypatch, tmp_path):
        """Relative output_path raises ValidationError (CBR-001)."""
        monkeypatch.setenv("EXCEL_CSV_EXPORT_ROOTS", str(tmp_path))
        from excelmcp._com import _check_csv_export_path
        with pytest.raises(ValidationError, match=r"absolute"):
            _check_csv_export_path("relative/path/out.csv")


# ---------------------------------------------------------------------------
# Class 3: TestExportRangeAsCsvOpenpyxl
# ---------------------------------------------------------------------------

class TestExportRangeAsCsvOpenpyxl:
    """Tests for export_range_as_csv with live=False (openpyxl path).

    No COM required — safe for CI on any platform.
    """

    @pytest.mark.unit
    def test_happy_path_3x3(self, monkeypatch, tmp_path):
        """Happy path: 3×3 range, headers included, CSV rows correct."""
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        data = [["Name", "Score", "Grade"], ["Alice", 95, "A"], ["Bob", 82, "B"]]
        wb_path = _make_workbook(tmp_path, data=data)

        from excelmcp.server_com import export_range_as_csv
        result = export_range_as_csv(
            path=wb_path,
            sheet="Sheet1",
            address="A1:C3",
            live=False,
        )

        assert result["rows"] == 3
        assert result["cols"] == 3
        assert "Name,Score,Grade" in result["csv"]
        assert result["output_path"] is None

    @pytest.mark.unit
    def test_no_headers_skips_first_row(self, monkeypatch, tmp_path):
        """include_headers=False: first (header) row is omitted from CSV output."""
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        data = [["Name", "Score", "Grade"], ["Alice", 95, "A"], ["Bob", 82, "B"]]
        wb_path = _make_workbook(tmp_path, data=data)

        from excelmcp.server_com import export_range_as_csv
        result = export_range_as_csv(
            path=wb_path,
            sheet="Sheet1",
            address="A1:C3",
            include_headers=False,
            live=False,
        )

        assert result["rows"] == 2
        assert "Name" not in result["csv"]
        assert "Alice" in result["csv"]

    @pytest.mark.unit
    def test_custom_delimiter_tab(self, monkeypatch, tmp_path):
        """delimiter=\\t produces tab-separated output."""
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        data = [["A", "B"], ["1", "2"]]
        wb_path = _make_workbook(tmp_path, data=data)

        from excelmcp.server_com import export_range_as_csv
        result = export_range_as_csv(
            path=wb_path,
            sheet="Sheet1",
            address="A1:B2",
            delimiter="\t",
            live=False,
        )

        assert "\t" in result["csv"]

    @pytest.mark.unit
    def test_csv_injection_sanitized(self, monkeypatch, tmp_path):
        """Cell values with injection-risk prefixes (+, @) are escaped in output CSV.

        Note: openpyxl treats '=...' as a formula so we test with '+' prefix
        (stored as a plain string) and '@' prefix. The _sanitize_csv_value unit
        tests cover '=' prefix exhaustively.
        """
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        path = tmp_path / "inj.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"] = "+44207000000"   # plus-prefix injection risk
        ws["A2"] = "@user"          # at-prefix injection risk
        wb.save(str(path))

        from excelmcp.server_com import export_range_as_csv
        result = export_range_as_csv(
            path=str(path),
            sheet="Sheet1",
            address="A1:A2",
            live=False,
        )

        assert "'+44207000000" in result["csv"]
        assert "'@user" in result["csv"]
        # Raw injection prefix must not appear unescaped at start of a CSV field
        assert "\n+44207000000" not in result["csv"]
        assert "\n@user" not in result["csv"]

    @pytest.mark.unit
    def test_output_path_writes_file(self, monkeypatch, tmp_path):
        """output_path provided: CSV written to file, csv key is None."""
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        monkeypatch.setenv("EXCEL_CSV_EXPORT_ROOTS", str(tmp_path))
        data = [["X", "Y"], [1, 2]]
        wb_path = _make_workbook(tmp_path, data=data)
        out = tmp_path / "out.csv"

        from excelmcp.server_com import export_range_as_csv
        result = export_range_as_csv(
            path=wb_path,
            sheet="Sheet1",
            address="A1:B2",
            output_path=str(out),
            confirm=True,
            live=False,
        )

        assert out.exists()
        assert result["output_path"] == str(out.resolve())
        assert result["csv"] is None

    @pytest.mark.unit
    def test_output_path_bad_suffix_raises(self, monkeypatch, tmp_path):
        """output_path with .txt extension raises ValidationError."""
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        monkeypatch.setenv("EXCEL_CSV_EXPORT_ROOTS", str(tmp_path))
        data = [["A"]]
        wb_path = _make_workbook(tmp_path, data=data)

        from excelmcp.server_com import export_range_as_csv
        with pytest.raises(ValidationError, match=r"\.csv|suffix"):
            export_range_as_csv(
                path=wb_path,
                sheet="Sheet1",
                address="A1:A1",
                output_path=str(tmp_path / "bad.txt"),
                confirm=True,
                live=False,
            )

    @pytest.mark.unit
    def test_output_path_outside_allowlist_raises(self, monkeypatch, tmp_path):
        """output_path outside EXCEL_CSV_EXPORT_ROOTS raises ValidationError."""
        csv_only = tmp_path / "csv_allowed"
        csv_only.mkdir()
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        monkeypatch.setenv("EXCEL_CSV_EXPORT_ROOTS", str(csv_only))
        data = [["A"]]
        wb_path = _make_workbook(tmp_path, data=data)

        from excelmcp.server_com import export_range_as_csv
        with pytest.raises(ValidationError):
            export_range_as_csv(
                path=wb_path,
                sheet="Sheet1",
                address="A1:A1",
                output_path=str(tmp_path / "escaped.csv"),  # outside csv_only
                confirm=True,
                live=False,
            )

    @pytest.mark.unit
    def test_address_invalid_raises(self, monkeypatch, tmp_path):
        """Non-A1-notation address raises ValidationError."""
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        wb_path = _make_workbook(tmp_path, data=[["A"]])

        from excelmcp.server_com import export_range_as_csv
        with pytest.raises(ValidationError):
            export_range_as_csv(
                path=wb_path,
                sheet="Sheet1",
                address="NOTVALID",
                live=False,
            )

    @pytest.mark.unit
    def test_invalid_delimiter_raises(self, monkeypatch, tmp_path):
        """Unsupported delimiter raises ValidationError."""
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        wb_path = _make_workbook(tmp_path, data=[["A"]])

        from excelmcp.server_com import export_range_as_csv
        with pytest.raises(ValidationError):
            export_range_as_csv(
                path=wb_path,
                sheet="Sheet1",
                address="A1:A1",
                delimiter="~",
                live=False,
            )

    @pytest.mark.unit
    def test_sheet_prefix_mismatch_raises(self, monkeypatch, tmp_path):
        """Address sheet prefix that does not match sheet param raises ValidationError."""
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        wb_path = _make_workbook(tmp_path, data=[["A"]])

        from excelmcp.server_com import export_range_as_csv
        with pytest.raises(ValidationError, match=r"sheet prefix|does not match"):
            export_range_as_csv(
                path=wb_path,
                sheet="Sheet1",
                address="WrongSheet!A1:A1",  # prefix ≠ sheet param
                live=False,
            )

    @pytest.mark.unit
    def test_sheet_prefix_matching_is_normalised(self, monkeypatch, tmp_path):
        """Address with correct sheet prefix is normalised and processed normally."""
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        data = [["X"], [1]]
        wb_path = _make_workbook(tmp_path, data=data)

        from excelmcp.server_com import export_range_as_csv
        result = export_range_as_csv(
            path=wb_path,
            sheet="Sheet1",
            address="Sheet1!A1:A2",  # prefix matches
            live=False,
        )

        assert result["rows"] == 2


# ---------------------------------------------------------------------------
# Class 4: TestExportRangeAsCsvSizeCap
# ---------------------------------------------------------------------------

class TestExportRangeAsCsvSizeCap:
    """Tests for 5 MB inline CSV cap."""

    @pytest.mark.unit
    def test_size_cap_raises_tool_error(self, monkeypatch, tmp_path):
        """CSV output exceeding 5 MB raises ToolError with '5MB' or 'limit' in message."""
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))

        # Produce a fake read_range that returns data that serialises > 5MB
        big_cell_value = "X" * 1024  # 1 KB per cell
        big_row = [{"address": f"A{i}", "value": big_cell_value, "data_type": "s"}
                   for i in range(1024)]
        fake_data = [big_row] * 6  # 6 rows × 1024 × 1024 bytes > 5MB

        def _fake_read_range(path, sheet, address):
            return fake_data

        monkeypatch.setattr("excelmcp._io.read_range", _fake_read_range)

        path = tmp_path / "test.xlsx"
        openpyxl.Workbook().save(str(path))

        from fastmcp.exceptions import ToolError
        from excelmcp.server_com import export_range_as_csv
        with pytest.raises(ToolError, match=r"5.MB|5MB|limit"):
            export_range_as_csv(
                path=str(path),
                sheet="Sheet1",
                address="A1:A1",
                live=False,
            )

    @pytest.mark.unit
    def test_size_cap_not_raised_with_output_path(self, monkeypatch, tmp_path):
        """CBR-002: CSV exceeding 5 MB must NOT raise when output_path is set."""
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        monkeypatch.setenv("EXCEL_CSV_EXPORT_ROOTS", str(tmp_path))

        # Create a fake workbook file (path must exist for _check_path allowlist guard)
        wb_path = tmp_path / "big.xlsx"
        wb_path.write_bytes(b"fake")

        out = tmp_path / "big.csv"

        # Data that produces > 5 MB of CSV content
        LARGE_VALUE = "X" * 1000  # 1000 chars per cell
        fake_rows = [
            [{"address": f"A{r}", "value": LARGE_VALUE, "data_type": "s"} for _ in range(100)]
            for r in range(60)
        ]  # 100 cols × 60 rows × ~1000 bytes ≈ 6 MB

        from excelmcp.server_com import export_range_as_csv
        with patch("excelmcp._io.read_range", return_value=fake_rows):
            result = export_range_as_csv(
                path=str(wb_path),
                sheet="Sheet1",
                address="A1:CV60",
                output_path=str(out),
                confirm=True,
                live=False,
            )

        assert out.exists()
        assert result["csv"] is None
        assert result["output_path"] == str(out.resolve())
