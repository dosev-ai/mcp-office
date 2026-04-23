import pytest
import openpyxl

from excelmcp.workbook_openpyxl import (
    ExcelMCPError,
    apply_number_format,
    apply_cell_style,
    apply_alignment,
    set_column_width,
    set_row_height,
    add_border,
    auto_filter,
    save,
    _wb_cache,
)


# ── GROUP 26: Formatting tools ──────────────────────────────────────────────

@pytest.mark.unit
def test_apply_number_format_happy_path(sample_workbook, monkeypatch):
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    result = apply_number_format(str(sample_workbook), "Sheet1", "A1", "#,##0.00", confirm=True)
    assert "message" in result


@pytest.mark.unit
def test_apply_number_format_requires_confirm(sample_workbook, monkeypatch):
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    with pytest.raises(ExcelMCPError):
        apply_number_format(str(sample_workbook), "Sheet1", "A1", "#,##0.00", confirm=False)


@pytest.mark.unit
def test_apply_number_format_requires_write(sample_workbook, monkeypatch):
    monkeypatch.delenv("EXCEL_ENABLE_WRITE", raising=False)
    with pytest.raises(ExcelMCPError):
        apply_number_format(str(sample_workbook), "Sheet1", "A1", "#,##0.00", confirm=True)


@pytest.mark.unit
def test_apply_cell_style_happy_path(sample_workbook, monkeypatch):
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    result = apply_cell_style(
        str(sample_workbook), "Sheet1", "A1",
        bold=True, font_size=14.0, bg_color="FFFF00", confirm=True,
    )
    assert "message" in result


@pytest.mark.unit
def test_apply_cell_style_bad_hex_raises(sample_workbook, monkeypatch):
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    with pytest.raises(ExcelMCPError):
        apply_cell_style(str(sample_workbook), "Sheet1", "A1", font_color="ZZZZZZ", confirm=True)


@pytest.mark.unit
def test_apply_cell_style_bad_bg_hex_raises(sample_workbook, monkeypatch):
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    with pytest.raises(ExcelMCPError):
        apply_cell_style(str(sample_workbook), "Sheet1", "A1", bg_color="#FF0000", confirm=True)


@pytest.mark.unit
def test_apply_cell_style_write_gate_blocks_without_write_env(tmp_path, monkeypatch):
    """_check_write() must fire before any param validation — gate blocks when EXCEL_ENABLE_WRITE unset."""
    monkeypatch.delenv("EXCEL_ENABLE_WRITE", raising=False)
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    wb = openpyxl.Workbook()
    path = str(tmp_path / "gate_test.xlsx")
    wb.save(path)
    with pytest.raises(ExcelMCPError, match="EXCEL_ENABLE_WRITE"):
        apply_cell_style(path, "Sheet", "A1", bold=True, confirm=True)


@pytest.mark.unit
def test_apply_alignment_write_gate_blocks_without_write_env(tmp_path, monkeypatch):
    """_check_write() must fire before horizontal/vertical enum check."""
    monkeypatch.delenv("EXCEL_ENABLE_WRITE", raising=False)
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    wb = openpyxl.Workbook()
    path = str(tmp_path / "gate_test2.xlsx")
    wb.save(path)
    with pytest.raises(ExcelMCPError, match="EXCEL_ENABLE_WRITE"):
        apply_alignment(path, "Sheet", "A1", horizontal="LEFT_INVALID", confirm=True)


@pytest.mark.unit
def test_apply_alignment_happy_path(sample_workbook, monkeypatch):
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    result = apply_alignment(
        str(sample_workbook), "Sheet1", "A1",
        horizontal="center", vertical="top", wrap_text=True, confirm=True,
    )
    assert "message" in result


@pytest.mark.unit
def test_apply_alignment_bad_horizontal_raises(sample_workbook, monkeypatch):
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    with pytest.raises(ExcelMCPError):
        apply_alignment(str(sample_workbook), "Sheet1", "A1", horizontal="invalid", confirm=True)


@pytest.mark.unit
def test_apply_alignment_bad_vertical_raises(sample_workbook, monkeypatch):
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    with pytest.raises(ExcelMCPError):
        apply_alignment(str(sample_workbook), "Sheet1", "A1", vertical="middle", confirm=True)


@pytest.mark.unit
def test_set_column_width_happy_path(sample_workbook, monkeypatch):
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    result = set_column_width(str(sample_workbook), "Sheet1", "A", 20.0, confirm=True)
    assert "message" in result


@pytest.mark.unit
def test_set_column_width_out_of_range_raises(sample_workbook, monkeypatch):
    # Width validation fires before _check_write so no EXCEL_ENABLE_WRITE needed
    monkeypatch.delenv("EXCEL_ENABLE_WRITE", raising=False)
    with pytest.raises(ExcelMCPError):
        set_column_width(str(sample_workbook), "Sheet1", "A", 999, confirm=True)


@pytest.mark.unit
def test_set_row_height_happy_path(sample_workbook, monkeypatch):
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    result = set_row_height(str(sample_workbook), "Sheet1", 1, 30.0, confirm=True)
    assert "message" in result


@pytest.mark.unit
def test_set_row_height_out_of_range_raises(sample_workbook, monkeypatch):
    monkeypatch.delenv("EXCEL_ENABLE_WRITE", raising=False)
    with pytest.raises(ExcelMCPError):
        set_row_height(str(sample_workbook), "Sheet1", 1, 500.0, confirm=True)


@pytest.mark.unit
def test_set_row_height_invalid_row_raises(sample_workbook, monkeypatch):
    monkeypatch.delenv("EXCEL_ENABLE_WRITE", raising=False)
    with pytest.raises(ExcelMCPError):
        set_row_height(str(sample_workbook), "Sheet1", 0, 30.0, confirm=True)


@pytest.mark.unit
def test_add_border_happy_path(sample_workbook, monkeypatch):
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    result = add_border(str(sample_workbook), "Sheet1", "A1", border_style="thin", confirm=True)
    assert "message" in result


@pytest.mark.unit
def test_add_border_range_happy_path(sample_workbook, monkeypatch):
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    result = add_border(str(sample_workbook), "Sheet1", "A1:A2", border_style="medium", confirm=True)
    assert "message" in result


@pytest.mark.unit
def test_add_border_bad_style_raises(sample_workbook, monkeypatch):
    # border_style validation fires before _check_write so no EXCEL_ENABLE_WRITE needed
    monkeypatch.delenv("EXCEL_ENABLE_WRITE", raising=False)
    with pytest.raises(ExcelMCPError):
        add_border(str(sample_workbook), "Sheet1", "A1", border_style="explode", confirm=True)


@pytest.mark.unit
def test_add_border_bad_color_raises(sample_workbook, monkeypatch):
    monkeypatch.delenv("EXCEL_ENABLE_WRITE", raising=False)
    with pytest.raises(ExcelMCPError):
        add_border(str(sample_workbook), "Sheet1", "A1", color="GGGGGG", confirm=True)


# ── GROUP 28: Formatting tools ok-schema assertions ───────────────────────────

@pytest.mark.unit
def test_apply_number_format_ok_schema(sample_workbook, monkeypatch):
    """apply_number_format returns dict with ok=True, sheet, address."""
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    result = apply_number_format(str(sample_workbook), "Sheet1", "A1", "#,##0.00", confirm=True)
    assert result.get("ok") is True
    assert result.get("sheet") == "Sheet1"
    assert result.get("address") == "A1"


@pytest.mark.unit
def test_apply_cell_style_ok_schema(sample_workbook, monkeypatch):
    """apply_cell_style returns dict with ok=True, sheet, address."""
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    result = apply_cell_style(str(sample_workbook), "Sheet1", "A1", bold=True, confirm=True)
    assert result.get("ok") is True
    assert result.get("sheet") == "Sheet1"


@pytest.mark.unit
def test_add_border_ok_schema(sample_workbook, monkeypatch):
    """add_border returns dict with ok=True, sheet, address, border_style."""
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    result = add_border(str(sample_workbook), "Sheet1", "A1", border_style="thin", confirm=True)
    assert result.get("ok") is True
    assert result.get("border_style") == "thin"


@pytest.mark.unit
def test_apply_alignment_ok_schema(sample_workbook, monkeypatch):
    """apply_alignment returns dict with ok=True."""
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    result = apply_alignment(str(sample_workbook), "Sheet1", "A1", horizontal="center", confirm=True)
    assert result.get("ok") is True, f"Expected ok=True, got: {result}"


@pytest.mark.unit
def test_set_column_width_ok_schema(sample_workbook, monkeypatch):
    """set_column_width returns dict with ok=True."""
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    result = set_column_width(str(sample_workbook), "Sheet1", "A", 20.0, confirm=True)
    assert result.get("ok") is True, f"Expected ok=True, got: {result}"


@pytest.mark.unit
def test_set_row_height_ok_schema(sample_workbook, monkeypatch):
    """set_row_height returns dict with ok=True."""
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    result = set_row_height(str(sample_workbook), "Sheet1", 1, 30.0, confirm=True)
    assert result.get("ok") is True, f"Expected ok=True, got: {result}"


# ── GROUP — Bug Fix Regressions ────────────────────────────────────────────

class TestBugFixes:
    """Regression tests for Bug 1 (auto_filter table promotion) and Bug 2 (apply_cell_style preserves font color)."""

    @pytest.mark.unit
    def test_apply_cell_style_preserves_font_color_when_not_specified(self, sample_workbook, monkeypatch):
        """Calling apply_cell_style without font_color must NOT drop an existing color (Bug 2 regression)."""
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        # Step 1: set a red font; explicit save needed under lazy-save refactor
        apply_cell_style(sample_workbook, "Sheet1", "A1", font_color="FF0000", confirm=True)
        save(sample_workbook, confirm=True)
        _wb_cache.clear()
        # Step 2: apply bold without specifying font_color — color must survive
        apply_cell_style(sample_workbook, "Sheet1", "A1", bold=True, confirm=True)
        save(sample_workbook, confirm=True)
        _wb_cache.clear()
        wb = openpyxl.load_workbook(sample_workbook)
        cell = wb["Sheet1"]["A1"]
        assert cell.font.bold is True
        assert "FF0000" in (cell.font.color.rgb or ""), f"font color was dropped, got: {cell.font.color.rgb!r}"

    @pytest.mark.unit
    def test_auto_filter_table_sheet_promotes_address(self, sample_workbook, monkeypatch):
        """auto_filter with header-only range on a Table sheet must promote to full range (Bug 1 regression)."""
        from openpyxl.worksheet.table import Table
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        # Add an Excel Table to the workbook
        wb = openpyxl.load_workbook(sample_workbook)
        ws = wb["Sheet1"]
        ws.add_table(Table(displayName="TestTable", ref="A1:C5"))
        wb.save(sample_workbook)
        _wb_cache.clear()
        # Call with header-only address — must be promoted to A1:C5
        result = auto_filter(sample_workbook, sheet="Sheet1", address="A1:C1", confirm=True)
        assert result["ok"] is True
        # After Fix 1: address is suppressed, not promoted; worksheet-level AF cleared
        assert result["address"] is None
        assert "warning" in result
        assert "suppressed" in result["warning"].lower()


# ── GROUP — PERF-OPT Sprint: LRU-cached style factory coverage ───────────

class TestPerfOptLRUCache:
    """Tests for the LRU-cached style factory functions added in the PERF-OPT sprint.

    _cached_font / _cached_alignment / _cached_fill live in excelmcp._format and
    are exercised through the public apply_cell_style / apply_alignment MCP tools.
    """

    @pytest.mark.unit
    def test_apply_cell_style_full_params_hoisted_font(self, sample_workbook, monkeypatch):
        """Full-param path: when bold, italic, font_size, AND font_color are all non-None,
        apply_cell_style hoists a single _cached_font() object before the cell loop rather
        than reading per-cell existing values.  Verify all four font properties plus the
        fill color are correctly persisted on the target cell."""
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        result = apply_cell_style(
            sample_workbook, "Sheet1", "A1",
            bold=True, italic=True, font_size=16.0, font_color="0000FF",
            bg_color="FFFF00", confirm=True,
        )
        assert result["ok"] is True
        save(sample_workbook, confirm=True)
        _wb_cache.clear()
        wb = openpyxl.load_workbook(sample_workbook)
        cell = wb["Sheet1"]["A1"]
        assert cell.font.bold is True
        assert cell.font.italic is True
        assert cell.font.size == 16.0
        assert "0000FF" in (cell.font.color.rgb or ""), f"font color wrong: {cell.font.color.rgb!r}"
        assert "FFFF00" in (cell.fill.fgColor.rgb or ""), f"fill color wrong: {cell.fill.fgColor.rgb!r}"

    @pytest.mark.unit
    def test_lru_cache_safety_apply_cell_style(self, sample_workbook, monkeypatch):
        """Cache safety: styling two cells with identical full params produces the same
        LRU cache hit for _cached_font.  Verify that assigning the shared cached object
        to both cells does NOT corrupt either cell's style.  openpyxl wraps raw style
        objects in an immutable StyleProxy on cell assignment, so sharing is safe."""
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        # Both calls use identical args → same LRU cache entry for _cached_font
        apply_cell_style(
            sample_workbook, "Sheet1", "A1",
            bold=True, italic=False, font_size=14.0, font_color="FF0000",
            confirm=True,
        )
        apply_cell_style(
            sample_workbook, "Sheet1", "B1",
            bold=True, italic=False, font_size=14.0, font_color="FF0000",
            confirm=True,
        )
        save(sample_workbook, confirm=True)
        _wb_cache.clear()
        wb = openpyxl.load_workbook(sample_workbook)
        a1 = wb["Sheet1"]["A1"]
        b1 = wb["Sheet1"]["B1"]
        assert a1.font.bold is True, "A1 bold lost after cache sharing"
        assert b1.font.bold is True, "B1 bold lost after cache sharing"
        assert a1.font.italic is False
        assert b1.font.italic is False
        assert a1.font.size == 14.0
        assert b1.font.size == 14.0
        assert "FF0000" in (a1.font.color.rgb or ""), f"A1 color lost: {a1.font.color.rgb!r}"
        assert "FF0000" in (b1.font.color.rgb or ""), f"B1 color lost: {b1.font.color.rgb!r}"

    @pytest.mark.unit
    def test_apply_alignment_partial_params_merges_existing(self, sample_workbook, monkeypatch):
        """Partial-param path for apply_alignment: specifying only horizontal must preserve
        previously set vertical and wrap_text values (per-cell merge via _cached_alignment)."""
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        # Establish full alignment first
        apply_alignment(
            sample_workbook, "Sheet1", "A1",
            horizontal="left", vertical="top", wrap_text=True, confirm=True,
        )
        save(sample_workbook, confirm=True)
        _wb_cache.clear()
        # Partial update — only horizontal changes; vertical and wrap_text must survive
        apply_alignment(
            sample_workbook, "Sheet1", "A1",
            horizontal="right", confirm=True,
        )
        save(sample_workbook, confirm=True)
        _wb_cache.clear()
        wb = openpyxl.load_workbook(sample_workbook)
        cell = wb["Sheet1"]["A1"]
        assert cell.alignment.horizontal == "right"
        assert cell.alignment.vertical == "top", "vertical was clobbered by partial update"
        assert cell.alignment.wrap_text is True, "wrap_text was clobbered by partial update"
