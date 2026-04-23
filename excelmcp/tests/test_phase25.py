import pytest

from excelmcp.workbook_openpyxl import (
    ValidationError,
    NotAllowedError,
    validate_formula_syntax,
    copy_range,
    list_merged_cells,
    _wb_cache,
)


class TestPhase25QuickWins:
    """Phase 2.5 — validate_formula_syntax, copy_range, list_merged_cells."""

    # ── validate_formula_syntax ──────────────────────────────────────────

    @pytest.mark.unit
    def test_valid_formula(self):
        """=SUM(A1:A10) is syntactically valid."""
        result = validate_formula_syntax("=SUM(A1:A10)")
        assert result["valid"] is True
        assert result["error"] is None

    @pytest.mark.unit
    def test_valid_formula_simple(self):
        """=A1+B1 is syntactically valid."""
        result = validate_formula_syntax("=A1+B1")
        assert result["valid"] is True
        assert result["error"] is None

    @pytest.mark.unit
    def test_empty_formula_is_invalid(self):
        """Empty string is not a valid formula."""
        result = validate_formula_syntax("")
        assert result["valid"] is False
        assert result["error"] is not None

    @pytest.mark.unit
    def test_non_formula_is_invalid(self):
        """A plain string without '=' is not a valid formula."""
        result = validate_formula_syntax("hello")
        assert result["valid"] is False
        assert result["error"] is not None

    # ── copy_range ───────────────────────────────────────────────────────

    @pytest.mark.unit
    def test_copy_range_disabled_by_default(self, sample_workbook, monkeypatch):
        """copy_range must fail when EXCEL_ENABLE_WRITE is not set."""
        monkeypatch.delenv("EXCEL_ENABLE_WRITE", raising=False)
        with pytest.raises(NotAllowedError, match="EXCEL_ENABLE_WRITE"):
            copy_range(sample_workbook, "Sheet1", "A1:A2", "C1:C2", confirm=True)

    @pytest.mark.unit
    def test_copy_range_requires_confirm(self, sample_workbook, monkeypatch):
        """copy_range must fail when confirm=False even if write is enabled."""
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        with pytest.raises(ValidationError, match="confirm=True"):
            copy_range(sample_workbook, "Sheet1", "A1:A2", "C1:C2", confirm=False)

    @pytest.mark.unit
    def test_copy_range_success(self, sample_workbook, monkeypatch):
        """copy_range copies values; C1 must equal A1 after the call."""
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        result = copy_range(sample_workbook, "Sheet1", "A1:A2", "C1:C2", confirm=True)
        assert result["copied"] == 2
        assert result["src"] == "A1:A2"
        assert result["dst"] == "C1:C2"

        _wb_cache.clear()
        import openpyxl as _xl
        wb_v = _xl.load_workbook(sample_workbook)
        assert wb_v["Sheet1"]["C1"].value == "hello"
        assert wb_v["Sheet1"]["C2"].value == 42

    @pytest.mark.unit
    def test_copy_range_path_not_in_allowlist(self, tmp_path, monkeypatch):
        """copy_range must raise ValidationError when path is outside allowlist."""
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        outside = tmp_path.parent / "other_dir" / "file.xlsx"
        with pytest.raises(ValidationError, match="not in allowlist"):
            copy_range(str(outside), "Sheet1", "A1:A2", "C1:C2", confirm=True)

    @pytest.mark.unit
    def test_copy_range_nonexistent_sheet(self, sample_workbook, monkeypatch):
        """copy_range with a non-existent sheet name must raise ValidationError."""
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        with pytest.raises(ValidationError, match="[Ss]heet"):
            copy_range(sample_workbook, "NoSuchSheet", "A1:A2", "C1:C2", confirm=True)

    @pytest.mark.unit
    def test_copy_range_single_cell(self, sample_workbook, monkeypatch):
        """Single-cell src=A1:A1 → dst=C1:C1 copies exactly 1 cell; no off-by-one."""
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        result = copy_range(sample_workbook, "Sheet1", "A1:A1", "C1:C1", confirm=True)
        assert result["copied"] == 1
        assert result["src"] == "A1:A1"
        assert result["dst"] == "C1:C1"

        _wb_cache.clear()
        import openpyxl as _xl
        wb_v = _xl.load_workbook(sample_workbook)
        assert wb_v["Sheet1"]["C1"].value == "hello"

    @pytest.mark.unit
    def test_copy_range_invalid_src_range(self, sample_workbook, monkeypatch):
        """copy_range with a malformed src_range must raise ValidationError."""
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        with pytest.raises(ValidationError, match="[Ii]nvalid range"):
            copy_range(sample_workbook, "Sheet1", "NOT_A_RANGE", "C1:C2", confirm=True)

    # ── list_merged_cells ────────────────────────────────────────────────

    @pytest.mark.unit
    def test_list_merged_cells_no_merges(self, sample_workbook):
        """A fresh workbook sheet with no merged cells returns an empty list."""
        result = list_merged_cells(sample_workbook, "Sheet1")
        assert isinstance(result, list)
        assert result == []

    @pytest.mark.unit
    def test_list_merged_cells_with_merge(self, tmp_path, monkeypatch):
        """After programmatically merging A1:B2, list_merged_cells returns that range."""
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        import openpyxl as _xl
        wb_path = tmp_path / "merged.xlsx"
        _wb = _xl.Workbook()
        _wb.active.title = "Sheet1"
        _wb.active.merge_cells("A1:B2")
        _wb.save(str(wb_path))

        result = list_merged_cells(str(wb_path), "Sheet1")
        assert "A1:B2" in result

    @pytest.mark.unit
    def test_list_merged_cells_invalid_sheet(self, sample_workbook):
        """Non-existent sheet name must raise ValidationError."""
        with pytest.raises(ValidationError, match="[Ss]heet"):
            list_merged_cells(sample_workbook, "NoSuchSheet")

    @pytest.mark.unit
    def test_list_merged_cells_path_not_in_allowlist(self, tmp_path, monkeypatch):
        """list_merged_cells with a path outside the allowlist must raise ValidationError."""
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        outside = tmp_path.parent / "other_dir" / "file.xlsx"
        with pytest.raises(ValidationError, match="not in allowlist"):
            list_merged_cells(str(outside), "Sheet1")

    @pytest.mark.unit
    def test_list_merged_cells_multiple_merges(self, tmp_path, monkeypatch):
        """A sheet with 2+ merged ranges returns all of them (not just the first)."""
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        import openpyxl as _xl
        wb_path = tmp_path / "multi_merged.xlsx"
        _wb = _xl.Workbook()
        _wb.active.title = "Sheet1"
        _wb.active.merge_cells("A1:B2")
        _wb.active.merge_cells("D5:E6")
        _wb.save(str(wb_path))

        result = list_merged_cells(str(wb_path), "Sheet1")
        assert len(result) == 2
        assert "A1:B2" in result
        assert "D5:E6" in result

    @pytest.mark.unit
    def test_list_merged_cells_on_dirty_workbook(self, tmp_path, monkeypatch):
        """list_merged_cells must read from the in-memory write-mode cache entry when
        available, so that merges applied in-session (before save()) are visible.

        Regression for US-13: previously list_merged_cells called _load_wb(for_write=False)
        which missed the write-mode cache key (path, True) and reloaded from disk,
        returning stale (pre-merge) data instead of the dirty in-memory state.
        """
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        import openpyxl as _xl
        wb_path = tmp_path / "dirty.xlsx"
        # Disk copy has NO merged cells
        _wb_disk = _xl.Workbook()
        _wb_disk.active.title = "Sheet1"
        _wb_disk.save(str(wb_path))

        # Build a dirty in-memory workbook WITH merged cells and inject it as the
        # write-mode cache entry — simulating merge_cells + write_cell without save().
        _wb_dirty = _xl.Workbook()
        _wb_dirty.active.title = "Sheet1"
        _wb_dirty.active.merge_cells("A1:D1")
        path_str = str(wb_path.resolve())
        disk_mtime = wb_path.stat().st_mtime
        _wb_cache[(path_str, True)] = (_wb_dirty, disk_mtime)

        # list_merged_cells must return the in-memory merged range, NOT reload from disk
        result = list_merged_cells(str(wb_path), "Sheet1")
        assert "A1:D1" in result, (
            "list_merged_cells must see in-memory (unsaved/dirty) merged ranges "
            "when a write-mode cache entry exists (regression for US-13 cache bypass bug)"
        )

    # ── validate_formula_syntax — supplemental coverage ──────────────────

    @pytest.mark.unit
    def test_validate_formula_syntax_allowlist_path(self, tmp_path, monkeypatch):
        """validate_formula_syntax with path outside the allowlist must raise ValidationError."""
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        outside = tmp_path.parent / "other_dir" / "file.xlsx"
        with pytest.raises(ValidationError, match="not in allowlist"):
            validate_formula_syntax("=SUM(A1)", path=str(outside))

    @pytest.mark.unit
    def test_validate_formula_syntax_openpyxl_false_positive_known_limitation(self):
        """KNOWN LIMITATION: openpyxl stores formulas as strings without real syntax parsing.

        =SUMX((( is syntactically invalid in Excel but openpyxl does NOT validate formula
        syntax — it accepts any '=...' string and returns valid=True here. This test
        documents the known false-positive behaviour (peer review finding #1, non-critical).
        Do NOT change the assertion to False; if openpyxl ever adds real parsing, this test
        will start failing and validate_formula_syntax (and its docstring) should be updated.
        See also: _data.py validate_formula_syntax docstring note.
        """
        result = validate_formula_syntax("=SUMX(((")
        # openpyxl accepts any '=...' string without parsing — intentional false positive
        assert result["valid"] is True  # documents the openpyxl limitation, NOT desired Excel behaviour
