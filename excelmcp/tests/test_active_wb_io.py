"""Tests that the LC-2 conflict guard is wired into all _io.py write paths.

7 tests: guard is called by write_cell, write_range, append_rows, create_workbook;
ValidationError and OfficeCOMError from guard propagate cleanly; guard runs before
openpyxl workbook load.
"""
from __future__ import annotations

from pathlib import Path
from unittest.mock import MagicMock

import openpyxl
import pytest

from excelmcp._core import OfficeCOMError, ValidationError


# ---------------------------------------------------------------------------
# Shared fixture
# ---------------------------------------------------------------------------

@pytest.fixture
def wb_path(tmp_path, monkeypatch):
    """Minimal workbook in a writable, allowlisted tmp directory."""
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    p = tmp_path / "test.xlsx"
    wb = openpyxl.Workbook()
    wb.active.title = "Sheet1"
    wb["Sheet1"]["A1"] = "seed"
    wb.save(str(p))
    return str(p)


# ---------------------------------------------------------------------------
# C-01: write_cell calls the conflict guard
# ---------------------------------------------------------------------------

def test_write_cell_calls_conflict_guard(wb_path, monkeypatch):
    """C-01: write_cell invokes _check_file_not_open_in_excel with the resolved path."""
    guard = MagicMock()
    monkeypatch.setattr("excelmcp._io._check_file_not_open_in_excel", guard)

    from excelmcp._io import write_cell
    write_cell(wb_path, "Sheet1", "B1", "value", confirm=True)

    guard.assert_called_once()
    called_path = guard.call_args[0][0]
    assert isinstance(called_path, Path), f"Guard must receive a Path, got: {type(called_path)}"


# ---------------------------------------------------------------------------
# C-02: write_range calls the conflict guard
# ---------------------------------------------------------------------------

def test_write_range_calls_conflict_guard(wb_path, monkeypatch):
    """C-02: write_range invokes _check_file_not_open_in_excel with the resolved path."""
    guard = MagicMock()
    monkeypatch.setattr("excelmcp._io._check_file_not_open_in_excel", guard)

    from excelmcp._io import write_range
    write_range(wb_path, "Sheet1", "A1", [["a", "b"]], confirm=True)

    guard.assert_called_once()
    called_path = guard.call_args[0][0]
    assert isinstance(called_path, Path)


# ---------------------------------------------------------------------------
# C-03: append_rows calls the conflict guard
# ---------------------------------------------------------------------------

def test_append_rows_calls_conflict_guard(wb_path, monkeypatch):
    """C-03: append_rows invokes _check_file_not_open_in_excel with the resolved path."""
    guard = MagicMock()
    monkeypatch.setattr("excelmcp._io._check_file_not_open_in_excel", guard)

    from excelmcp._io import append_rows
    append_rows(wb_path, "Sheet1", [["row1col1"]], confirm=True)

    guard.assert_called_once()
    called_path = guard.call_args[0][0]
    assert isinstance(called_path, Path)


# ---------------------------------------------------------------------------
# C-04: create_workbook calls the conflict guard
# ---------------------------------------------------------------------------

def test_create_workbook_calls_conflict_guard(tmp_path, monkeypatch):
    """C-04: create_workbook invokes _check_file_not_open_in_excel for the new file path."""
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")

    guard = MagicMock()
    monkeypatch.setattr("excelmcp._io._check_file_not_open_in_excel", guard)

    new_path = str(tmp_path / "new.xlsx")
    from excelmcp._io import create_workbook
    create_workbook(new_path, confirm=True)

    guard.assert_called_once()
    called_path = guard.call_args[0][0]
    assert isinstance(called_path, Path)


# ---------------------------------------------------------------------------
# C-05: ValidationError from guard propagates out of write_cell
# ---------------------------------------------------------------------------

def test_write_cell_blocked_by_guard_validation_error(wb_path, monkeypatch):
    """C-05: ValidationError raised by the guard propagates; file contents unchanged."""
    original_wb = openpyxl.load_workbook(wb_path)
    original_a1 = original_wb.active["A1"].value

    def _raise_guard(path: Path) -> None:
        raise ValidationError("Workbook is open in Excel. Close it first.")

    monkeypatch.setattr("excelmcp._io._check_file_not_open_in_excel", _raise_guard)

    from excelmcp._io import write_cell
    with pytest.raises(ValidationError):
        write_cell(wb_path, "Sheet1", "A1", "SHOULD_NOT_WRITE", confirm=True)

    # File must be unchanged
    after_wb = openpyxl.load_workbook(wb_path)
    assert after_wb.active["A1"].value == original_a1


# ---------------------------------------------------------------------------
# C-06: OfficeCOMError from guard propagates out of write_range
# ---------------------------------------------------------------------------

def test_write_range_blocked_by_guard_office_com_error(wb_path, monkeypatch):
    """C-06: OfficeCOMError raised by the guard propagates from write_range."""
    def _raise_guard(path: Path) -> None:
        raise OfficeCOMError("Could not verify file lock state via COM")

    monkeypatch.setattr("excelmcp._io._check_file_not_open_in_excel", _raise_guard)

    from excelmcp._io import write_range
    with pytest.raises(OfficeCOMError):
        write_range(wb_path, "Sheet1", "A1", [["x"]], confirm=True)


# ---------------------------------------------------------------------------
# C-07: Guard is called BEFORE openpyxl workbook load
# ---------------------------------------------------------------------------

def test_guard_called_before_file_load(wb_path, monkeypatch):
    """C-07: _check_file_not_open_in_excel is invoked before _load_wb opens the file."""
    call_order: list[str] = []

    def _tracking_guard(path: Path) -> None:
        call_order.append("guard")

    real_load_wb = None

    def _tracking_load(resolved, for_write: bool = False):
        call_order.append("load")
        # Delegate to the real implementation
        return real_load_wb(resolved, for_write=for_write)

    # Capture real _load_wb before patching
    import excelmcp._io as _io_mod
    real_load_wb = _io_mod._load_wb

    monkeypatch.setattr("excelmcp._io._check_file_not_open_in_excel", _tracking_guard)
    monkeypatch.setattr("excelmcp._io._load_wb", _tracking_load)

    from excelmcp._io import write_cell
    write_cell(wb_path, "Sheet1", "B2", "test_order", confirm=True)

    assert "guard" in call_order, "Guard must be called"
    assert "load" in call_order, "_load_wb must be called"
    assert call_order.index("guard") < call_order.index("load"), (
        "Guard must be called BEFORE workbook load, "
        f"but call order was: {call_order}"
    )
