import pytest
import openpyxl

from excelmcp.workbook_openpyxl import (
    ExcelMCPError,
    ValidationError,
    NotAllowedError,
    insert_rows,
    insert_cols,
    protect_sheet,
    set_print_area,
    write_named_range,
)


class TestPhase23Tools:
    """Phase 2.3 — insert_rows, insert_cols, protect_sheet, set_print_area, N3 validation."""

    # ── N3: range_address validation for write_named_range ──────────────────

    @pytest.mark.unit
    def test_write_named_range_invalid_range_address(self, sample_workbook, monkeypatch):
        """N3: garbage range_address must raise ValidationError."""
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        with pytest.raises(ValidationError, match="[Ii]nvalid range address"):
            write_named_range(
                sample_workbook, name="TestRange", sheet="Sheet1",
                range_address="NOTARANGE", confirm=True,
            )

    @pytest.mark.unit
    def test_write_named_range_valid_range_address_cell(self, sample_workbook, monkeypatch):
        """N3: simple cell ref A1 must succeed."""
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        result = write_named_range(
            sample_workbook, name="SingleCell", sheet="Sheet1",
            range_address="A1", confirm=True,
        )
        assert result["ok"] is True

    @pytest.mark.unit
    def test_write_named_range_valid_range_address_range(self, sample_workbook, monkeypatch):
        """N3: standard A1:D4 range must succeed."""
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        result = write_named_range(
            sample_workbook, name="MultiCell", sheet="Sheet1",
            range_address="A1:D4", confirm=True,
        )
        assert result["ok"] is True

    # ── insert_rows ────────────────────────────────────────────

    @pytest.mark.unit
    def test_insert_rows_happy(self, tmp_path, monkeypatch):
        """Insert 1 row before row 2; old row 2 data shifts to row 3."""
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        wb_path = tmp_path / "ins_row.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"] = "header"
        ws["A2"] = "original_row2"
        wb.save(str(wb_path))

        result = insert_rows(str(wb_path), "Sheet1", row=2, count=1, confirm=True)
        assert result["ok"] is True
        assert result["row"] == 2
        assert result["count"] == 1

        wb_verify = openpyxl.load_workbook(str(wb_path))
        ws_verify = wb_verify["Sheet1"]
        assert ws_verify["A2"].value is None  # newly inserted blank row
        assert ws_verify["A3"].value == "original_row2"  # old row 2 shifted down

    @pytest.mark.unit
    def test_insert_rows_confirm_false_raises(self, sample_workbook, monkeypatch):
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        with pytest.raises(ValidationError):
            insert_rows(sample_workbook, "Sheet1", row=1, confirm=False)

    @pytest.mark.unit
    def test_insert_rows_write_gate_raises(self, sample_workbook, monkeypatch):
        monkeypatch.delenv("EXCEL_ENABLE_WRITE", raising=False)
        with pytest.raises(NotAllowedError):
            insert_rows(sample_workbook, "Sheet1", row=1, confirm=True)

    @pytest.mark.unit
    def test_insert_rows_wrong_sheet_raises(self, sample_workbook, monkeypatch):
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        with pytest.raises(ValidationError, match="[Ss]heet"):
            insert_rows(sample_workbook, "NoSuchSheet", row=1, confirm=True)

    # ── insert_cols ────────────────────────────────────────────

    @pytest.mark.unit
    def test_insert_cols_happy(self, tmp_path, monkeypatch):
        """Insert 1 col before col 2; old col B data shifts to col C."""
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        wb_path = tmp_path / "ins_col.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"] = "col_a"
        ws["B1"] = "original_col_b"
        wb.save(str(wb_path))

        result = insert_cols(str(wb_path), "Sheet1", col=2, count=1, confirm=True)
        assert result["ok"] is True
        assert result["col"] == 2
        assert result["count"] == 1

        wb_verify = openpyxl.load_workbook(str(wb_path))
        ws_verify = wb_verify["Sheet1"]
        assert ws_verify["B1"].value is None  # newly inserted blank col
        assert ws_verify["C1"].value == "original_col_b"  # old col B shifted right

    @pytest.mark.unit
    def test_insert_cols_confirm_false_raises(self, sample_workbook, monkeypatch):
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        with pytest.raises(ValidationError):
            insert_cols(sample_workbook, "Sheet1", col=1, confirm=False)

    @pytest.mark.unit
    def test_insert_cols_write_gate_raises(self, sample_workbook, monkeypatch):
        monkeypatch.delenv("EXCEL_ENABLE_WRITE", raising=False)
        with pytest.raises(NotAllowedError):
            insert_cols(sample_workbook, "Sheet1", col=1, confirm=True)

    # ── protect_sheet ──────────────────────────────────────────

    @pytest.mark.unit
    def test_protect_sheet_happy(self, tmp_path, monkeypatch):
        """Set protection with a password; verify ws.protection.sheet is True."""
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        wb_path = tmp_path / "protect.xlsx"
        wb = openpyxl.Workbook()
        wb.active.title = "Sheet1"
        wb.save(str(wb_path))

        result = protect_sheet(str(wb_path), "Sheet1", password="secret", confirm=True)
        assert result["ok"] is True
        assert result["protected"] is True
        assert result["has_password"] is True

        wb_verify = openpyxl.load_workbook(str(wb_path))
        assert wb_verify["Sheet1"].protection.sheet is True

    @pytest.mark.unit
    def test_protect_sheet_no_password(self, tmp_path, monkeypatch):
        """Protect without password; sheet must be protected, has_password=False."""
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        wb_path = tmp_path / "protect_no_pw.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        wb.save(str(wb_path))

        result = protect_sheet(str(wb_path), "Sheet1", password=None, confirm=True)
        assert result["ok"] is True
        assert result["protected"] is False
        assert result.get("has_password") is False

        wb_verify = openpyxl.load_workbook(str(wb_path))
        assert wb_verify["Sheet1"].protection.sheet is False

    @pytest.mark.unit
    def test_protect_sheet_confirm_false_raises(self, sample_workbook, monkeypatch):
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        with pytest.raises(ValidationError):
            protect_sheet(sample_workbook, "Sheet1", password="pw", confirm=False)

    @pytest.mark.unit
    def test_protect_sheet_write_gate_raises(self, sample_workbook, monkeypatch):
        monkeypatch.delenv("EXCEL_ENABLE_WRITE", raising=False)
        with pytest.raises(NotAllowedError):
            protect_sheet(sample_workbook, "Sheet1", password="pw", confirm=True)

    # ── set_print_area ─────────────────────────────────────────

    @pytest.mark.unit
    def test_set_print_area_happy(self, tmp_path, monkeypatch):
        """Set print area; verify ws.print_area after load."""
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        wb_path = tmp_path / "print_area.xlsx"
        wb = openpyxl.Workbook()
        wb.active.title = "Sheet1"
        wb.save(str(wb_path))

        result = set_print_area(str(wb_path), "Sheet1", address="A1:F10", confirm=True)
        assert result["ok"] is True
        assert result["address"] == "A1:F10"

        wb_verify = openpyxl.load_workbook(str(wb_path))
        # openpyxl normalizes print_area on reload (e.g. "'Sheet1'!$A$1:$F$10")
        pa = wb_verify["Sheet1"].print_area or ""
        pa_normalized = pa.split("!")[-1].replace("$", "").upper()
        assert pa_normalized == "A1:F10"

    @pytest.mark.unit
    def test_set_print_area_clear(self, tmp_path, monkeypatch):
        """Clear print area; verify ws.print_area is None or empty."""
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        wb_path = tmp_path / "print_area_clear.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.print_area = "A1:B2"
        wb.save(str(wb_path))

        result = set_print_area(str(wb_path), "Sheet1", address=None, confirm=True)
        assert result["ok"] is True
        assert result["address"] is None

        wb_verify = openpyxl.load_workbook(str(wb_path))
        assert not wb_verify["Sheet1"].print_area  # None or empty string

    @pytest.mark.unit
    def test_set_print_area_confirm_false_raises(self, sample_workbook, monkeypatch):
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        with pytest.raises(ValidationError):
            set_print_area(sample_workbook, "Sheet1", address="A1:B2", confirm=False)

    @pytest.mark.unit
    def test_set_print_area_write_gate_raises(self, sample_workbook, monkeypatch):
        monkeypatch.delenv("EXCEL_ENABLE_WRITE", raising=False)
        with pytest.raises(NotAllowedError):
            set_print_area(sample_workbook, "Sheet1", address="A1:B2", confirm=True)

    # ── G1: insert_cols wrong-sheet ────────────────────────────────────────

    @pytest.mark.unit
    def test_insert_cols_wrong_sheet_raises(self, sample_workbook, monkeypatch):
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        with pytest.raises(ValidationError, match="[Ss]heet"):
            insert_cols(sample_workbook, sheet="NoSuchSheet", col=1, confirm=True)

    # ── G2: protect_sheet wrong-sheet ─────────────────────────────────────

    @pytest.mark.unit
    def test_protect_sheet_wrong_sheet_raises(self, sample_workbook, monkeypatch):
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        with pytest.raises(ValidationError, match="[Ss]heet"):
            protect_sheet(sample_workbook, sheet="NoSuchSheet", confirm=True)

    # ── G3: set_print_area wrong-sheet ────────────────────────────────────

    @pytest.mark.unit
    def test_set_print_area_wrong_sheet_raises(self, sample_workbook, monkeypatch):
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        with pytest.raises(ValidationError, match="[Ss]heet"):
            set_print_area(sample_workbook, sheet="NoSuchSheet", address="A1:B2", confirm=True)

    # ── Bounds tests for insert_rows/insert_cols ──────────────────────────

    @pytest.mark.unit
    def test_insert_rows_count_too_large_raises(self, tmp_path, monkeypatch):
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        p = tmp_path / "wb.xlsx"
        openpyxl.Workbook().save(str(p))
        with pytest.raises(ExcelMCPError):
            insert_rows(str(p), sheet="Sheet", row=1, count=9999, confirm=True)

    @pytest.mark.unit
    def test_insert_cols_count_too_large_raises(self, tmp_path, monkeypatch):
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        p = tmp_path / "wb.xlsx"
        openpyxl.Workbook().save(str(p))
        with pytest.raises(ExcelMCPError):
            insert_cols(str(p), sheet="Sheet", col=1, count=9999, confirm=True)
