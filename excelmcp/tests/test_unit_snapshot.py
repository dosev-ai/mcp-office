"""Unit tests for server_snapshot.py (US-E-032: snapshot_workbook + diff_workbooks).

All tests use real openpyxl and tmp_path — no COM required.
"""
from __future__ import annotations

import shutil
from pathlib import Path
from unittest.mock import MagicMock, patch

import openpyxl
import pytest

from excelmcp._core import NotAllowedError, ValidationError
from excelmcp.server_snapshot import diff_workbooks, snapshot_workbook


# ---------------------------------------------------------------------------
# Shared fixtures
# ---------------------------------------------------------------------------

@pytest.fixture()
def wb_in_allowlist(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> Path:
    """Return a real .xlsx file path that is inside EXCEL_ALLOWLIST_ROOTS."""
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")

    wb_path = tmp_path / "test_wb.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "hello"
    ws["B1"] = 42
    wb.save(str(wb_path))
    return wb_path


@pytest.fixture()
def xlsm_in_allowlist(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> Path:
    """Return a real .xlsm-named file inside EXCEL_ALLOWLIST_ROOTS."""
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")

    # openpyxl can't create real macro-enabled workbooks, so copy an xlsx as xlsm
    wb_path = tmp_path / "test_macro.xlsm"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "macro_wkbk"
    wb.save(str(wb_path))
    return wb_path


# ---------------------------------------------------------------------------
# TestSnapshotWorkbook
# ---------------------------------------------------------------------------

class TestSnapshotWorkbook:

    def test_happy_path_with_dest_dir(self, wb_in_allowlist: Path, tmp_path: Path) -> None:
        """Snapshot written to an explicit dest_dir inside allowlist."""
        dest = tmp_path / "snapshots"
        result = snapshot_workbook(
            path=str(wb_in_allowlist),
            dest_dir=str(dest),
            confirm=True,
        )
        assert result["ok"] is True
        snap_path = Path(result["snapshot"])
        assert snap_path.exists()
        assert snap_path.parent == dest
        assert result["size_bytes"] > 0

    def test_happy_path_default_dest_dir(self, wb_in_allowlist: Path) -> None:
        """Snapshot written to source parent when dest_dir is None."""
        result = snapshot_workbook(
            path=str(wb_in_allowlist),
            confirm=True,
        )
        assert result["ok"] is True
        snap_path = Path(result["snapshot"])
        assert snap_path.parent == wb_in_allowlist.parent
        assert snap_path.name != wb_in_allowlist.name  # timestamped name
        assert snap_path.exists()

    def test_confirm_false_raises(
        self, wb_in_allowlist: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        """confirm=False must raise NotAllowedError before any file I/O."""
        with pytest.raises((NotAllowedError, ValidationError)):
            snapshot_workbook(path=str(wb_in_allowlist), confirm=False)

    def test_source_not_in_allowlist(
        self, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        """Source file outside allowlist raises ValidationError."""
        other = tmp_path / "other"
        other.mkdir()
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(other))
        monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")

        wb_path = tmp_path / "outside.xlsx"
        openpyxl.Workbook().save(str(wb_path))

        with pytest.raises(ValidationError):
            snapshot_workbook(path=str(wb_path), confirm=True)

    def test_dest_dir_outside_allowlist(
        self, wb_in_allowlist: Path, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        """dest_dir outside allowlist raises ValidationError (BLOCKER-1)."""
        outside = tmp_path.parent / "outside_snap"
        outside.mkdir(exist_ok=True)

        with pytest.raises(ValidationError):
            snapshot_workbook(
                path=str(wb_in_allowlist),
                dest_dir=str(outside),
                confirm=True,
            )

    def test_dest_file_already_exists(
        self, wb_in_allowlist: Path, tmp_path: Path
    ) -> None:
        """Overwriting an existing snapshot file raises ValidationError."""
        existing = tmp_path / "existing_snap.xlsx"
        shutil.copy2(str(wb_in_allowlist), str(existing))

        with pytest.raises(ValidationError, match="already exists"):
            snapshot_workbook(
                path=str(wb_in_allowlist),
                filename="existing_snap.xlsx",
                confirm=True,
            )

    def test_unsupported_filename_extension(self, wb_in_allowlist: Path) -> None:
        """filename with a non-xlsx/xlsm extension raises ValidationError."""
        with pytest.raises(ValidationError, match=".xlsx or .xlsm"):
            snapshot_workbook(
                path=str(wb_in_allowlist),
                filename="snapshot.csv",
                confirm=True,
            )

    def test_xlsm_source_accepted(self, xlsm_in_allowlist: Path) -> None:
        """An .xlsm source file is accepted and produces an .xlsm snapshot."""
        result = snapshot_workbook(path=str(xlsm_in_allowlist), confirm=True)
        assert result["ok"] is True
        assert Path(result["snapshot"]).suffix == ".xlsm"

    def test_filename_path_traversal_rejected(self, wb_in_allowlist: Path) -> None:
        """filename with .. path components must be rejected (BUG-001 regression guard)."""
        with pytest.raises(ValidationError, match="traversal"):
            snapshot_workbook(
                path=str(wb_in_allowlist),
                filename="../../secrets.xlsx",
                confirm=True,
            )


# ---------------------------------------------------------------------------
# TestDiffWorkbooks
# ---------------------------------------------------------------------------

def _make_wb(path: Path, data: dict) -> Path:
    """Helper: create a workbook at *path* with data in Sheet1."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for (row, col), val in data.items():
        ws.cell(row=row, column=col).value = val
    wb.save(str(path))
    return path


class TestDiffWorkbooks:

    @pytest.fixture()
    def diff_env(self, tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> Path:
        monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
        return tmp_path

    def test_identical_workbooks(self, diff_env: Path) -> None:
        """Identical workbooks produce zero diffs."""
        data = {(1, 1): "A", (1, 2): 10}
        a = _make_wb(diff_env / "a.xlsx", data)
        b = _make_wb(diff_env / "b.xlsx", data)

        result = diff_workbooks(str(a), str(b))
        assert result["ok"] is True
        assert result["total_diffs"] == 0
        assert result["diffs"] == []

    def test_modified_cell_detected(self, diff_env: Path) -> None:
        """A single changed cell should appear in the diff list."""
        a = _make_wb(diff_env / "a.xlsx", {(1, 1): "original"})
        b = _make_wb(diff_env / "b.xlsx", {(1, 1): "changed"})

        result = diff_workbooks(str(a), str(b))
        assert result["total_diffs"] == 1
        diff = result["diffs"][0]
        assert diff["val_a"] == "original"
        assert diff["val_b"] == "changed"

    def test_added_cell_detected(self, diff_env: Path) -> None:
        """A value added in b but absent in a should appear as diff."""
        a = _make_wb(diff_env / "a.xlsx", {})
        b = _make_wb(diff_env / "b.xlsx", {(3, 5): "new_val"})

        result = diff_workbooks(str(a), str(b))
        assert result["total_diffs"] >= 1
        diff = result["diffs"][0]
        assert diff["val_b"] == "new_val"

    def test_scope_limited_to_named_sheet(self, diff_env: Path) -> None:
        """sheets kwarg limits comparison to the named sheets only."""
        # Both workbooks have Sheet1 and Sheet2.
        # Only Sheet2 differs; comparing Sheet1 only should produce 0 diffs.
        wb_a = openpyxl.Workbook()
        ws_a1 = wb_a.active
        ws_a1.title = "Sheet1"
        ws_a1["A1"] = "same"
        ws_a2 = wb_a.create_sheet("Sheet2")
        ws_a2["A1"] = "only_a"
        wb_a.save(str(diff_env / "a.xlsx"))

        wb_b = openpyxl.Workbook()
        ws_b1 = wb_b.active
        ws_b1.title = "Sheet1"
        ws_b1["A1"] = "same"
        ws_b2 = wb_b.create_sheet("Sheet2")
        ws_b2["A1"] = "only_b"
        wb_b.save(str(diff_env / "b.xlsx"))

        result = diff_workbooks(
            str(diff_env / "a.xlsx"),
            str(diff_env / "b.xlsx"),
            sheets=["Sheet1"],
        )
        assert result["total_diffs"] == 0
        assert result["sheets_compared"] == ["Sheet1"]

    @pytest.mark.parametrize("which", ["a", "b"])
    def test_path_not_in_allowlist(
        self, diff_env: Path, tmp_path: Path, which: str
    ) -> None:
        """A path outside the allowlist raises ValidationError."""
        good = _make_wb(diff_env / "good.xlsx", {(1, 1): 1})
        outside_dir = tmp_path.parent / "outside_diff"
        outside_dir.mkdir(exist_ok=True)
        bad = _make_wb(outside_dir / "bad.xlsx", {(1, 1): 1})

        with pytest.raises(ValidationError):
            if which == "a":
                diff_workbooks(str(bad), str(good))
            else:
                diff_workbooks(str(good), str(bad))

    def test_large_workbook_exceeds_cell_limit(self, diff_env: Path) -> None:
        """A sheet whose cell count exceeds EXCEL_MAX_RANGE_CELLS raises ValidationError."""
        a = _make_wb(diff_env / "a.xlsx", {(1, 1): 1})
        b = _make_wb(diff_env / "b.xlsx", {(1, 1): 1})

        # Mock openpyxl.load_workbook to return a workbook whose sheet has huge dimensions.
        mock_ws = MagicMock()
        mock_ws.max_row = 5000
        mock_ws.max_column = 200  # 5000 * 200 = 1 000 000 >> default 5 000

        mock_wb = MagicMock()
        mock_wb.sheetnames = ["Sheet1"]
        mock_wb.__getitem__ = lambda self, key: mock_ws

        with patch("excelmcp.server_snapshot.openpyxl.load_workbook", return_value=mock_wb):
            with pytest.raises(ValidationError):
                diff_workbooks(str(a), str(b))
