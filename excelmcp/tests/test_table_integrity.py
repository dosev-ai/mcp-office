"""Tests for Sprint I T2: _sync_autofilter_to_table helper and its call sites
in delete_rows / insert_rows.

All 5 tests run with `pytest -v` without a live COM object (openpyxl only).
No @pytest.mark.integration marker is required.
"""
import pytest
import openpyxl
from openpyxl.worksheet.table import Table
from openpyxl.utils.cell import range_boundaries

from excelmcp._core import _sync_autofilter_to_table  # new helper -- Sprint I
from excelmcp.workbook_openpyxl import delete_rows, insert_rows
from excelmcp.server import rows as rows_tool


# ---------------------------------------------------------------------------
# Module-level helper factory
# Used by tests #11 and #12 that need a file with both a Table and autoFilter.
# ---------------------------------------------------------------------------

def _make_af_table_workbook(tmp_path, monkeypatch, *, table_ref="A1:C5"):
    """Create a workbook with Sheet1 containing an Excel Table AND a matching
    worksheet-level autoFilter, both referencing `table_ref`.

    This mirrors the real-world case where create_table (or Excel itself) sets
    both objects to the same range.  After a delete_rows or insert_rows call,
    openpyxl auto-updates tbl.ref but leaves auto_filter.ref stale --
    the T2 sync fix must reconcile them.

    Returns the absolute path as a str.
    """
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    path = tmp_path / "af_table.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    min_col, min_row, max_col, max_row = range_boundaries(table_ref.upper())
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            ws.cell(row=r, column=c).value = f"r{r}c{c}"

    tbl = Table(displayName="MainTable", ref=table_ref)
    ws.add_table(tbl)
    ws.auto_filter.ref = table_ref   # autoFilter matches table exactly

    wb.save(str(path))
    return str(path)


# ---------------------------------------------------------------------------
# TEST #8
# Function:  test_sync_autofilter_updates_af_ref_to_match_table_ref
# Group:     T2 -- helper happy-path unit test
# W2 Blocker: (none -- positive acceptance criterion for T2)
# ---------------------------------------------------------------------------

@pytest.mark.unit
def test_sync_autofilter_updates_af_ref_to_match_table_ref():
    """_sync_autofilter_to_table(ws) aligns ws.auto_filter.ref with tbl.ref
    when they share the same top-left cell but have diverged.

    Simulates the post-append scenario: tbl.ref was expanded to A1:C15 (by T1
    code), but auto_filter.ref still shows the pre-expansion extent A1:C10.
    After calling the helper, auto_filter.ref must equal tbl.ref.
    """
    wb = openpyxl.Workbook()
    ws = wb.active

    # tbl.ref is the new expanded extent (set by T1 code after append)
    tbl = Table(displayName="T1", ref="A1:C15")
    ws.add_table(tbl)

    # auto_filter.ref is stale -- still shows the pre-expansion extent
    ws.auto_filter.ref = "A1:C10"

    _sync_autofilter_to_table(ws)

    assert ws.auto_filter.ref == "A1:C15"   # synced to table's new extent


# ---------------------------------------------------------------------------
# TEST #9  * COVERS W2 BLOCKER 2-A *
# Function:  test_sync_autofilter_none_autofilter_ref_is_noop
# Group:     T2 -- BLOCKER regression test (Case A)
# W2 Blocker: BLOCKER 2-A (AttributeError when ws.auto_filter.ref is None)
# ---------------------------------------------------------------------------

@pytest.mark.unit
def test_sync_autofilter_none_autofilter_ref_is_noop():
    """_sync_autofilter_to_table(ws) must NOT crash when ws.auto_filter.ref is
    None -- the openpyxl default for any workbook where no explicit autoFilter
    has ever been set, even when the sheet contains an Excel Table.

    Without the guard (`if not af_ref: return`), calling `.upper()` on None
    raises: AttributeError: 'NoneType' object has no attribute 'upper'.

    This BLOCKER affects every append_rows/delete_rows/insert_rows call on a
    table-backed sheet that was never manually filtered.
    """
    wb = openpyxl.Workbook()
    ws = wb.active

    tbl = Table(displayName="T1", ref="A1:C10")
    ws.add_table(tbl)
    # ws.auto_filter.ref is None by default -- deliberately NOT set

    # Precondition: confirm auto_filter.ref is indeed None
    assert ws.auto_filter.ref is None

    _sync_autofilter_to_table(ws)   # must NOT raise AttributeError

    # Postcondition: auto_filter.ref left untouched (still None)
    assert ws.auto_filter.ref is None


# ---------------------------------------------------------------------------
# TEST #10  * COVERS W2 BLOCKER 2-B *
# Function:  test_sync_autofilter_table_none_tbl_ref_is_skipped
# Group:     T2 -- BLOCKER regression test (Case B)
# W2 Blocker: BLOCKER 2-B (AttributeError when tbl.ref is None inside helper)
# ---------------------------------------------------------------------------

@pytest.mark.unit
def test_sync_autofilter_table_none_tbl_ref_is_skipped():
    """_sync_autofilter_to_table(ws) must NOT crash when a table on the sheet
    has tbl.ref = None.  The malformed table is silently skipped and
    ws.auto_filter.ref is left unmodified.

    Validates the per-table guard:
        tbl_ref = tbl.ref or ""
        if not tbl_ref: continue

    Without the guard, `range_boundaries(None)` raises AttributeError.
    """
    wb = openpyxl.Workbook()
    ws = wb.active

    ws.auto_filter.ref = "A1:C5"                       # af is set (not None)
    tbl = Table(displayName="BadTable", ref=None)       # malformed -- ref is None
    ws.add_table(tbl)

    _sync_autofilter_to_table(ws)           # must NOT raise

    # auto_filter.ref left unmodified -- the bad table was skipped
    assert ws.auto_filter.ref == "A1:C5"


# ---------------------------------------------------------------------------
# TEST #11  * COVERS W2 BLOCKER 3 *
# Function:  test_delete_rows_table_autofilter_synced
# Group:     T2 -- call-site regression via workbook_openpyxl.delete_rows
# W2 Blocker: BLOCKER 3 (T2 call-site must be in _sheet.py::delete_rows)
# ---------------------------------------------------------------------------

@pytest.mark.unit
def test_delete_rows_table_autofilter_synced(tmp_path, monkeypatch):
    """After delete_rows removes a row from a table-backed sheet, the saved
    workbook must have ws.auto_filter.ref == tbl.ref (both contracted by the
    deletion via the T2 sync call).

    openpyxl auto-contracts tbl.ref when a row inside the table is deleted
    (A1:C5 -> A1:C4 when row 5 is removed).  Without BLOCKER 3 fix, the
    auto_filter.ref stays at the old A1:C5 extent, causing Excel to open
    the file with "content not readable / repaired" dialog.

    Fix location: _sheet.py::delete_rows must call _sync_autofilter_to_table(ws)
    after ws.delete_rows().
    """
    # Table at A1:C5; autoFilter.ref = "A1:C5"
    path = _make_af_table_workbook(tmp_path, monkeypatch, table_ref="A1:C5")

    # Delete the last table row (row 5) -- table contracts to A1:C4
    result = delete_rows(path, "Sheet1", 5, confirm=True)

    assert result["ok"] is True
    assert result["rows_deleted"] == 1

    # Read back from disk and verify both refs contracted together
    wb2 = openpyxl.load_workbook(path)
    ws2 = wb2["Sheet1"]
    tables = list(ws2.tables.values())
    assert len(tables) == 1

    new_tbl_ref = tables[0].ref          # openpyxl shifts tbl.ref after deletion
    assert new_tbl_ref == "A1:C4"        # contracted by 1 row

    # KEY T2 assertion: auto_filter.ref must match tbl.ref via _sync call
    assert ws2.auto_filter.ref == new_tbl_ref

    # Server-layer smoke: verify .fn() path also returns ok=True
    # Recreate a fresh workbook at the same path (_save_and_evict already cleared cache)
    path2 = _make_af_table_workbook(tmp_path, monkeypatch, table_ref="A1:C5")
    r = rows_tool(operation="delete", path=path2, sheet="Sheet1", start_row=5, confirm=True)
    assert r["ok"] is True


# ---------------------------------------------------------------------------
# TEST #12  * COVERS W2 BLOCKER 3 *
# Function:  test_insert_rows_table_autofilter_synced
# Group:     T2 -- call-site regression via workbook_openpyxl.insert_rows
# W2 Blocker: BLOCKER 3 (T2 call-site must be in _sheet.py::insert_rows)
# ---------------------------------------------------------------------------

@pytest.mark.unit
def test_insert_rows_table_autofilter_synced(tmp_path, monkeypatch):
    """After insert_rows inserts rows inside a table-backed sheet, the saved
    workbook must have ws.auto_filter.ref == tbl.ref (both expanded by the
    insertion via the T2 sync call).

    openpyxl auto-expands tbl.ref when rows are inserted inside the table
    (A1:C5 -> A1:C7 when 2 rows are inserted at row 3).  Without BLOCKER 3
    fix, the auto_filter.ref stays at the old A1:C5 extent, causing the
    same Excel "repaired content" corruption.

    Fix location: _sheet.py::insert_rows must call _sync_autofilter_to_table(ws)
    after ws.insert_rows().
    """
    # Table at A1:C5; autoFilter.ref = "A1:C5"
    path = _make_af_table_workbook(tmp_path, monkeypatch, table_ref="A1:C5")

    # Insert 2 rows at position 3 (inside the table) -- rows shift down by 2
    result = insert_rows(path, "Sheet1", 3, count=2, confirm=True)

    assert result["ok"] is True

    # Read back from disk and verify both refs expanded together
    wb2 = openpyxl.load_workbook(path)
    ws2 = wb2["Sheet1"]
    tables = list(ws2.tables.values())
    assert len(tables) == 1

    new_tbl_ref = tables[0].ref          # openpyxl expands tbl.ref after insertion
    assert new_tbl_ref == "A1:C7"        # 5 original rows + 2 inserted = 7 rows

    # KEY T2 assertion: auto_filter.ref must match tbl.ref via _sync call
    assert ws2.auto_filter.ref == new_tbl_ref

    # Server-layer smoke: verify .fn() path also returns ok=True
    # Recreate a fresh workbook at the same path (_save_and_evict already cleared cache)
    path2 = _make_af_table_workbook(tmp_path, monkeypatch, table_ref="A1:C5")
    r = rows_tool(operation="insert", path=path2, sheet="Sheet1", start_row=3, count=2, confirm=True)
    assert r["ok"] is True
