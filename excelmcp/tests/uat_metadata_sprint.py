#!/usr/bin/env python3
"""UAT evidence collector — Excel MCP metadata umbrella sprint (Phase 7).

Test cases:
  TC1 — tools/call "capabilities"  → metadata_contract_version="1.0" + metadata_policy present
  TC2 — resources/read /metadata   → 10 canonical fields + policy in response
  TC3 — resources/read /meta       → alias parity with TC2 (same structure)
  TC4 — resources/read /meta/contract → contract_version="1.0" + 10-entry fields array

Evidence saved to: tests/uat_metadata_sprint_evidence.jsonl
"""
from __future__ import annotations

import json
import os
import pathlib
import subprocess
import sys
import tempfile
import time
from urllib.parse import quote

import openpyxl

# ---------------------------------------------------------------------------
# Paths & configuration
# ---------------------------------------------------------------------------
_REPO_ROOT = pathlib.Path(__file__).resolve().parent.parent.parent   # mcp-office/
_EXCEL_SRC  = pathlib.Path(__file__).resolve().parent.parent / "src"  # excel/src
PYTHONPATH  = os.environ.get("EXCELMCP_PYTHONPATH", str(_EXCEL_SRC))
SERVER_CMD  = [sys.executable, "-m", "excelmcp.server"]
EVIDENCE_PATH = pathlib.Path(__file__).with_name("uat_metadata_sprint_evidence.jsonl")

CANONICAL_FIELDS = (
    "path", "workbook_name", "sheet_names", "active_sheet", "sheet_count",
    "has_formulas", "last_modified", "size_bytes", "format", "contract_version",
)

# ---------------------------------------------------------------------------
# Subprocess helpers
# ---------------------------------------------------------------------------

def _start_server(allowlist_roots: str, enable_write: bool = True) -> subprocess.Popen:
    env = {
        **os.environ,
        "PYTHONPATH": PYTHONPATH,
        "EXCEL_ALLOWLIST_ROOTS": allowlist_roots,
        "EXCEL_ENABLE_WRITE": "true" if enable_write else "false",
        "FASTMCP_SHOW_CLI_BANNER": "false",
    }
    return subprocess.Popen(
        SERVER_CMD,
        stdin=subprocess.PIPE,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        env=env,
    )


def _send(proc: subprocess.Popen, method: str, params: dict, req_id: int) -> tuple[dict, dict]:
    req = {"jsonrpc": "2.0", "id": req_id, "method": method, "params": params}
    raw = json.dumps(req) + "\n"
    proc.stdin.write(raw.encode())
    proc.stdin.flush()
    raw_resp = proc.stdout.readline()
    if not raw_resp:
        raise RuntimeError("Server closed stdout unexpectedly")
    return json.loads(raw_resp.decode()), req


def _notify(proc: subprocess.Popen, method: str, params: dict | None = None) -> None:
    msg = {"jsonrpc": "2.0", "method": method, "params": params or {}}
    proc.stdin.write((json.dumps(msg) + "\n").encode())
    proc.stdin.flush()


def _terminate(proc: subprocess.Popen) -> None:
    try:
        proc.stdin.close()
    except Exception:
        pass
    proc.terminate()
    try:
        proc.wait(timeout=5)
    except subprocess.TimeoutExpired:
        proc.kill()
        proc.wait()


# ---------------------------------------------------------------------------
# Fixture builder
# ---------------------------------------------------------------------------

def _make_workbook(tmp_dir: pathlib.Path) -> pathlib.Path:
    wb_path = tmp_dir / "uat_meta.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = "Revenue"
    ws["B1"] = "=SUM(A2:A5)"
    ws["A2"] = 100
    ws["A3"] = 200
    wb.create_sheet("Summary")
    wb.save(str(wb_path))
    return wb_path


# ---------------------------------------------------------------------------
# Evidence logger
# ---------------------------------------------------------------------------

def _log(entry: dict) -> None:
    with EVIDENCE_PATH.open("a", encoding="utf-8") as fh:
        fh.write(json.dumps(entry) + "\n")


# ---------------------------------------------------------------------------
# Response helpers
# ---------------------------------------------------------------------------

def _extract_tool_result(response: dict) -> dict | None:
    """Parse a tools/call response — returns the decoded result dict or None."""
    if response.get("error"):
        return None
    result = response.get("result", {})
    if result.get("isError"):
        return None
    content = result.get("content", [])
    if not content:
        return None
    first = content[0]
    text = first.get("text") or first.get("json_text") or ""
    try:
        return json.loads(text)
    except (json.JSONDecodeError, TypeError):
        return None


def _extract_resource_result(response: dict) -> dict | None:
    """Parse a resources/read response — returns the decoded resource dict or None."""
    if response.get("error"):
        return None
    result = response.get("result", {})
    contents = result.get("contents", [])
    if not contents:
        return None
    first = contents[0]
    text = first.get("text", "")
    try:
        return json.loads(text)
    except (json.JSONDecodeError, TypeError):
        return None


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> int:
    sep = "=" * 70
    print(sep)
    print("Excel MCP — Metadata Sprint UAT Evidence Collector (Phase 7)")
    print(sep)

    # Clear stale evidence
    if EVIDENCE_PATH.exists():
        EVIDENCE_PATH.unlink()

    # Build fixture workbook
    tmp_dir = pathlib.Path(tempfile.mkdtemp(prefix="excelmcp_meta_uat_"))
    wb_path = _make_workbook(tmp_dir)
    print(f"\n[FIXTURE] workbook : {wb_path}")
    print(f"[FIXTURE] allowlist: {tmp_dir}\n")

    # URL-encode the Windows path for embedding in a URI
    # On Windows: C:\path\to\file -> C%3A%5Cpath%5Cto%5Cfile
    wb_uri_path = quote(str(wb_path), safe="")

    results: dict[str, str] = {}
    proc: subprocess.Popen | None = None

    try:
        proc = _start_server(str(tmp_dir))
        time.sleep(0.5)

        # ── Handshake ────────────────────────────────────────────────────
        init_resp, init_req = _send(
            proc, "initialize",
            {
                "protocolVersion": "2024-11-05",
                "capabilities": {},
                "clientInfo": {"name": "uat-metadata-sprint", "version": "7.0"},
            },
            req_id=1,
        )
        _log({"step": "initialize", "request": init_req, "response": init_resp})
        _notify(proc, "notifications/initialized")

        handshake_ok = (
            init_resp.get("jsonrpc") == "2.0"
            and "result" in init_resp
            and "protocolVersion" in init_resp.get("result", {})
        )
        print(f"[HANDSHAKE] {'OK' if handshake_ok else 'FAIL'}  "
              f"protocol={init_resp.get('result', {}).get('protocolVersion', '?')}")

        # ── TC1: capabilities → metadata_contract_version + metadata_policy ──
        print(f"\n{'─'*60}")
        print("TC1: tools/call 'capabilities'")
        caps_resp, caps_req = _send(
            proc, "tools/call",
            {"name": "capabilities", "arguments": {}},
            req_id=2,
        )
        caps_data = _extract_tool_result(caps_resp)
        tc1_cv   = (caps_data or {}).get("metadata_contract_version")
        tc1_pol  = (caps_data or {}).get("metadata_policy")
        tc1_pass = (tc1_cv == "1.0") and (tc1_pol is not None)

        print(f"  metadata_contract_version : {tc1_cv!r}  (expect '1.0')")
        print(f"  metadata_policy           : {tc1_pol!r}  (expect 'strict' or 'lenient')")
        print(f"  → TC1: {'PASS' if tc1_pass else 'FAIL'}")

        _log({
            "tc": "TC1",
            "description": "capabilities tool — metadata_contract_version + metadata_policy",
            "status": "PASS" if tc1_pass else "FAIL",
            "request": caps_req,
            "response": caps_resp,
            "extracted": {"metadata_contract_version": tc1_cv, "metadata_policy": tc1_pol},
        })
        results["TC1"] = "PASS" if tc1_pass else "FAIL"

        # ── TC2: /metadata resource → 10 canonical fields + policy ───────
        print(f"\n{'─'*60}")
        print("TC2: resources/read excelmcp://workbook/{path}/metadata")
        meta_uri   = f"excelmcp://workbook/{wb_uri_path}/metadata"
        print(f"  URI: {meta_uri}")
        meta_resp, meta_req = _send(
            proc, "resources/read",
            {"uri": meta_uri},
            req_id=3,
        )
        meta_data = _extract_resource_result(meta_resp)

        if meta_data is None:
            # Attempt fallback: some FastMCP versions nest text differently
            raw_result = meta_resp.get("result", {})
            if isinstance(raw_result, str):
                try:
                    meta_data = json.loads(raw_result)
                except Exception:
                    pass

        tc2_missing = [f for f in CANONICAL_FIELDS if f not in (meta_data or {})]
        tc2_has_policy = "policy" in (meta_data or {})
        tc2_pass = (len(tc2_missing) == 0) and tc2_has_policy

        print(f"  missing canonical fields : {tc2_missing or 'none'}")
        print(f"  policy present           : {tc2_has_policy}")
        if meta_data:
            print(f"  sheet_names              : {meta_data.get('sheet_names')}")
            print(f"  has_formulas             : {meta_data.get('has_formulas')}")
            print(f"  contract_version         : {meta_data.get('contract_version')}")
        print(f"  → TC2: {'PASS' if tc2_pass else 'FAIL'}")

        _log({
            "tc": "TC2",
            "description": "/metadata resource — 10 canonical fields + policy",
            "status": "PASS" if tc2_pass else "FAIL",
            "request": meta_req,
            "response": meta_resp,
            "extracted": meta_data,
            "missing_fields": tc2_missing,
            "has_policy": tc2_has_policy,
        })
        results["TC2"] = "PASS" if tc2_pass else "FAIL"

        # ── TC3: /meta resource (alias) → same structure as TC2 ──────────
        print(f"\n{'─'*60}")
        print("TC3: resources/read excelmcp://workbook/{path}/meta  (alias parity)")
        alias_uri = f"excelmcp://workbook/{wb_uri_path}/meta"
        print(f"  URI: {alias_uri}")
        alias_resp, alias_req = _send(
            proc, "resources/read",
            {"uri": alias_uri},
            req_id=4,
        )
        alias_data = _extract_resource_result(alias_resp)

        tc3_missing = [f for f in CANONICAL_FIELDS if f not in (alias_data or {})]
        tc3_has_policy = "policy" in (alias_data or {})
        # Parity check — same keys as TC2 result
        tc3_parity = (
            set((alias_data or {}).keys()) == set((meta_data or {}).keys())
            if (alias_data and meta_data) else False
        )
        tc3_pass = (len(tc3_missing) == 0) and tc3_has_policy and tc3_parity

        print(f"  missing canonical fields : {tc3_missing or 'none'}")
        print(f"  policy present           : {tc3_has_policy}")
        print(f"  parity with /metadata    : {tc3_parity}")
        print(f"  → TC3: {'PASS' if tc3_pass else 'FAIL'}")

        _log({
            "tc": "TC3",
            "description": "/meta alias resource — parity with /metadata",
            "status": "PASS" if tc3_pass else "FAIL",
            "request": alias_req,
            "response": alias_resp,
            "extracted": alias_data,
            "missing_fields": tc3_missing,
            "has_policy": tc3_has_policy,
            "parity_with_metadata": tc3_parity,
        })
        results["TC3"] = "PASS" if tc3_pass else "FAIL"

        # ── TC4: /meta/contract → contract_version + 10-entry fields array ─
        print(f"\n{'─'*60}")
        print("TC4: resources/read excelmcp://meta/contract")
        contract_resp, contract_req = _send(
            proc, "resources/read",
            {"uri": "excelmcp://meta/contract"},
            req_id=5,
        )
        contract_data = _extract_resource_result(contract_resp)

        tc4_cv     = (contract_data or {}).get("contract_version")
        tc4_fields = (contract_data or {}).get("fields", [])
        tc4_count  = len(tc4_fields)
        tc4_pass   = (tc4_cv == "1.0") and (tc4_count == 10)

        print(f"  contract_version : {tc4_cv!r}  (expect '1.0')")
        print(f"  fields count     : {tc4_count}  (expect 10)")
        if tc4_fields:
            field_names = [f.get("name") for f in tc4_fields]
            print(f"  field names      : {field_names}")
        print(f"  → TC4: {'PASS' if tc4_pass else 'FAIL'}")

        _log({
            "tc": "TC4",
            "description": "/meta/contract resource — contract_version + 10-field schema",
            "status": "PASS" if tc4_pass else "FAIL",
            "request": contract_req,
            "response": contract_resp,
            "extracted": contract_data,
            "contract_version": tc4_cv,
            "fields_count": tc4_count,
        })
        results["TC4"] = "PASS" if tc4_pass else "FAIL"

    except Exception as exc:
        import traceback
        print(f"\n[FATAL] {exc}")
        traceback.print_exc()
        _log({"step": "FATAL", "error": str(exc)})
    finally:
        if proc is not None:
            _terminate(proc)
        # Do NOT remove tmp_dir — leave fixture for inspection

    # ── Summary ──────────────────────────────────────────────────────────
    print(f"\n{sep}")
    print("UAT Metadata Sprint — Summary")
    print(sep)
    all_pass = all(v == "PASS" for v in results.values()) and len(results) == 4
    for tc in ["TC1", "TC2", "TC3", "TC4"]:
        status = results.get(tc, "NOT RUN")
        mark   = "+" if status == "PASS" else "!"
        print(f"  {tc}: {status} [{mark}]")
    print()
    verdict = "ALL PASS" if all_pass else "FAILURES DETECTED"
    print(f"Overall Verdict : {verdict}")
    print(f"Evidence JSONL  : {EVIDENCE_PATH.absolute()}")
    print(f"Fixture workbook: {wb_path}")
    print(sep)

    return 0 if all_pass else 1


if __name__ == "__main__":
    sys.exit(main())
