"""Tests for excelmcp._metadata_contract — 16 tests."""
from __future__ import annotations

import os
from datetime import datetime

import openpyxl
import pytest

from excelmcp._metadata_contract import (
    CONTRACT_VERSION,
    CANONICAL_FIELDS,
    build_workbook_metadata,
    validate_metadata,
)
from excelmcp.workbook_openpyxl import write_cell


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture
def wb_path(tmp_path, monkeypatch):
    """Minimal .xlsx at tmp_path with EXCEL_ALLOWLIST_ROOTS set."""
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    p = tmp_path / "test.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "hello"
    ws["B1"] = 42
    wb.save(str(p))
    return str(p)


@pytest.fixture
def wb_path_with_formula(tmp_path, monkeypatch):
    """Workbook containing =SUM(A1:A3)."""
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    p = tmp_path / "formulas.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = 10
    ws["A2"] = 20
    ws["A3"] = "=SUM(A1:A2)"
    wb.save(str(p))
    return str(p)


@pytest.fixture
def wb_path_no_formula(tmp_path, monkeypatch):
    """Workbook with no formula cells."""
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    p = tmp_path / "noformula.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Plain"
    ws["A1"] = "text"
    ws["B1"] = 99
    wb.save(str(p))
    return str(p)


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------

def test_build_metadata_happy_path(wb_path):
    """All 10 canonical fields are present with correct types."""
    meta = build_workbook_metadata(wb_path)
    assert isinstance(meta["path"], str)
    assert isinstance(meta["workbook_name"], str)
    assert isinstance(meta["sheet_names"], list)
    assert isinstance(meta["active_sheet"], str)
    assert isinstance(meta["sheet_count"], int)
    assert isinstance(meta["has_formulas"], bool)
    assert isinstance(meta["last_modified"], str)
    assert isinstance(meta["size_bytes"], int)
    assert isinstance(meta["format"], str)
    assert isinstance(meta["contract_version"], str)
    assert set(meta.keys()) >= set(CANONICAL_FIELDS)


def test_build_metadata_invalid_path(tmp_path, monkeypatch):
    """Non-existent path raises ToolError or ValidationError."""
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    from fastmcp.exceptions import ToolError
    from excelmcp._core import ValidationError
    with pytest.raises((ToolError, ValidationError)):
        build_workbook_metadata(str(tmp_path / "does_not_exist.xlsx"))


def test_build_metadata_has_formulas_true(wb_path_with_formula):
    """Workbook with =SUM(...) has_formulas is True."""
    meta = build_workbook_metadata(wb_path_with_formula)
    assert meta["has_formulas"] is True


def test_build_metadata_has_formulas_false(wb_path_no_formula):
    """Workbook with no formulas has_formulas is False."""
    meta = build_workbook_metadata(wb_path_no_formula)
    assert meta["has_formulas"] is False


def test_build_metadata_contract_version(wb_path):
    """contract_version field equals the CONTRACT_VERSION constant."""
    meta = build_workbook_metadata(wb_path)
    assert meta["contract_version"] == CONTRACT_VERSION
    assert meta["contract_version"] == "1.0"


def test_build_metadata_last_modified_iso(wb_path):
    """last_modified is an ISO-8601 string with UTC timezone offset."""
    meta = build_workbook_metadata(wb_path)
    lm = meta["last_modified"]
    assert isinstance(lm, str)
    from datetime import datetime
    dt = datetime.fromisoformat(lm)
    assert dt.tzinfo is not None


def test_build_metadata_sheet_names_list(wb_path):
    """sheet_names is a list of strings."""
    meta = build_workbook_metadata(wb_path)
    assert isinstance(meta["sheet_names"], list)
    assert all(isinstance(s, str) for s in meta["sheet_names"])
    assert "Sheet1" in meta["sheet_names"]


def test_build_metadata_size_bytes_int(wb_path):
    """size_bytes is a positive integer."""
    meta = build_workbook_metadata(wb_path)
    assert isinstance(meta["size_bytes"], int)
    assert meta["size_bytes"] > 0


def test_validate_metadata_missing_field(wb_path):
    """validate_metadata raises ValueError when a required field is absent."""
    meta = dict(build_workbook_metadata(wb_path))
    del meta["has_formulas"]
    with pytest.raises(ValueError, match="has_formulas"):
        validate_metadata(meta)


def test_validate_metadata_all_fields_pass(wb_path):
    """validate_metadata does not raise when all required fields are present."""
    meta = dict(build_workbook_metadata(wb_path))
    validate_metadata(meta)   # must not raise


def test_canonical_fields_constant():
    """CANONICAL_FIELDS is a tuple with exactly 10 elements."""
    assert isinstance(CANONICAL_FIELDS, tuple)
    assert len(CANONICAL_FIELDS) == 10
    assert "path" in CANONICAL_FIELDS
    assert "contract_version" in CANONICAL_FIELDS


def test_build_metadata_active_sheet(wb_path):
    """active_sheet matches wb.active.title."""
    wb = openpyxl.load_workbook(wb_path)
    expected = wb.active.title
    wb.close()
    meta = build_workbook_metadata(wb_path)
    assert meta["active_sheet"] == expected


def test_build_metadata_format(wb_path):
    """format field is '.xlsx' for a .xlsx file."""
    meta = build_workbook_metadata(wb_path)
    assert meta["format"] == ".xlsx"


def test_build_metadata_workbook_name(wb_path):
    """workbook_name equals os.path.basename(path)."""
    meta = build_workbook_metadata(wb_path)
    assert meta["workbook_name"] == os.path.basename(wb_path)


def test_build_metadata_after_write_cell_preserves_lineage_and_refreshes_mtime(tmp_path, monkeypatch):
    """A normal write_cell path preserves lineage fields and refreshes file metadata."""
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    path = tmp_path / "lineage_after_write.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "seed"
    wb.save(str(path))
    wb.close()

    stale_epoch = 946684800
    os.utime(path, (stale_epoch, stale_epoch))

    before = dict(build_workbook_metadata(str(path)))
    result = write_cell(str(path), "Sheet1", "C3", 42, confirm=True)

    persisted_wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    try:
        assert persisted_wb["Sheet1"]["C3"].value == 42
    finally:
        persisted_wb.close()

    after = dict(build_workbook_metadata(str(path)))

    assert result["ok"] is True
    assert after["path"] == before["path"]
    assert after["workbook_name"] == before["workbook_name"]
    assert after["sheet_names"] == before["sheet_names"]
    assert after["active_sheet"] == before["active_sheet"]
    assert after["sheet_count"] == before["sheet_count"]
    assert after["has_formulas"] is before["has_formulas"] is False
    assert after["format"] == before["format"]
    assert after["contract_version"] == before["contract_version"]
    assert after["size_bytes"] == path.stat().st_size
    assert datetime.fromisoformat(after["last_modified"]) > datetime.fromisoformat(before["last_modified"])


def test_no_fastmcp_import_in_contract_module():
    """_metadata_contract must not import fastmcp.server (the MCP app object)."""
    import ast
    import inspect
    import excelmcp._metadata_contract as mod
    source = inspect.getsource(mod)
    tree = ast.parse(source)
    for node in ast.walk(tree):
        if isinstance(node, ast.Import):
            for alias in node.names:
                assert not alias.name.startswith("fastmcp") or alias.name == "fastmcp.exceptions", (
                    f"_metadata_contract must not import fastmcp modules except fastmcp.exceptions; found: {alias.name}"
                )
        elif isinstance(node, ast.ImportFrom) and node.module:
            assert node.module != "fastmcp.server", (
                "_metadata_contract must not import fastmcp.server"
            )
            if node.module.startswith("fastmcp") and node.module != "fastmcp.exceptions":
                assert False, (
                    f"_metadata_contract must not import fastmcp modules except fastmcp.exceptions; found: {node.module}"
                )
