"""Delta performance benchmark for excelmcp formatting hot-paths.

Measures BOTH the old inline-allocation pattern (pre-optimization) and the new
hoisted approach via ``apply_cell_style`` / ``apply_alignment`` from ``_format.py``
at three cell counts (1 000, 5 000, 10 000).

Usage::

    python excel/tests/benchmarks/excel_perf_baseline.py

Writes:
    ``excel/tests/benchmarks/pre_optimization.json``  – old inline pattern
    ``excel/tests/benchmarks/post_optimization.json`` – optimised _format.py

Prints a delta summary table to stderr.
"""
from __future__ import annotations

import json
import os
import sys
import tempfile
import time
import tracemalloc
from datetime import datetime, timezone
from pathlib import Path

# ---------------------------------------------------------------------------
# Make excelmcp importable without an editable install when run standalone.
# ---------------------------------------------------------------------------
_src = Path(__file__).resolve().parents[3] / "src"
if str(_src) not in sys.path:
    sys.path.insert(0, str(_src))

# Pre-set the write-gate env var before importing _format so the import itself
# does not fail on any module-level guard.
os.environ["EXCEL_ENABLE_WRITE"] = "true"

import openpyxl  # noqa: E402  (after sys.path fix)
from openpyxl.styles import Alignment, Font, PatternFill  # noqa: E402

from excelmcp._format import apply_cell_style, apply_alignment  # noqa: E402
from excelmcp._core import _wb_cache  # noqa: E402


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _build_workbook(path: str, n_cells: int) -> str:
    """Create a fresh .xlsx with *n_cells* rows of data and return *path*."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for i in range(1, n_cells + 1):
        ws.cell(row=i, column=1, value=f"value_{i}")
    wb.save(path)
    return path


def _run_one(
    func,
    path: str,
    sheet: str,
    address: str,
    func_name: str,
    n_cells: int,
    **kwargs,
) -> dict:
    """Time *func* with tracemalloc; return a result record."""
    # Ensure a clean cache entry so we always load from disk.
    _wb_cache.clear()

    tracemalloc.start()
    t0 = time.perf_counter()
    func(path=path, sheet=sheet, address=address, confirm=True, **kwargs)
    t1 = time.perf_counter()
    _, peak = tracemalloc.get_traced_memory()
    tracemalloc.stop()

    # Evict dirty entry so subsequent test passes won't reuse stale state.
    _wb_cache.clear()

    return {
        "function": func_name,
        "cells": n_cells,
        "time_seconds": round(t1 - t0, 4),
        "peak_memory_mb": round(peak / (1024 * 1024), 2),
    }


# ---------------------------------------------------------------------------
# Pre-optimization inline-allocation benchmarks
# These replicate the exact pattern that existed BEFORE the hoist refactor:
# Font, PatternFill, and Alignment objects are constructed INSIDE the cell loop.
# ---------------------------------------------------------------------------


def _inline_apply_cell_style(
    path: str,
    sheet: str,
    address: str,
    bold: bool | None = None,
    italic: bool | None = None,
    font_size: float | None = None,
    font_color: str | None = None,
    bg_color: str | None = None,
    confirm: bool = False,  # noqa: ARG001  (signature parity)
) -> None:
    """Old inline pattern: creates Font/PatternFill INSIDE the cell loop."""
    wb = openpyxl.load_workbook(path)
    ws = wb[sheet]
    addr = address.upper().strip()
    if ":" not in addr:
        addr = f"{addr}:{addr}"
    for row in ws[addr]:
        for cell in row:
            if any(v is not None for v in (bold, italic, font_size, font_color)):
                current = cell.font
                existing_color = (
                    current.color if current.color is not None else None
                )
                # Inline construction — a new Font object every iteration.
                cell.font = Font(
                    bold=bold if bold is not None else current.bold,
                    italic=italic if italic is not None else current.italic,
                    size=font_size if font_size is not None else current.size,
                    color=(
                        font_color if font_color is not None else existing_color
                    ),
                )
            if bg_color is not None:
                # Inline construction — a new PatternFill object every iteration.
                cell.fill = PatternFill(fill_type="solid", fgColor=bg_color)
    wb.save(path)


def _inline_apply_alignment(
    path: str,
    sheet: str,
    address: str,
    horizontal: str | None = None,
    vertical: str | None = None,
    wrap_text: bool | None = None,
    confirm: bool = False,  # noqa: ARG001  (signature parity)
) -> None:
    """Old inline pattern: creates Alignment INSIDE the cell loop."""
    wb = openpyxl.load_workbook(path)
    ws = wb[sheet]
    addr = address.upper().strip()
    if ":" not in addr:
        addr = f"{addr}:{addr}"
    for row in ws[addr]:
        for cell in row:
            current = cell.alignment
            # Inline construction — a new Alignment object every iteration.
            cell.alignment = Alignment(
                horizontal=(
                    horizontal if horizontal is not None else current.horizontal
                ),
                vertical=(
                    vertical if vertical is not None else current.vertical
                ),
                wrap_text=(
                    wrap_text if wrap_text is not None else current.wrap_text
                ),
            )
    wb.save(path)


def _run_inline(
    func,
    path: str,
    sheet: str,
    address: str,
    func_name: str,
    n_cells: int,
    **kwargs,
) -> dict:
    """Time an inline (pre-opt) benchmark function; no cache involved."""
    tracemalloc.start()
    t0 = time.perf_counter()
    func(path=path, sheet=sheet, address=address, **kwargs)
    t1 = time.perf_counter()
    _, peak = tracemalloc.get_traced_memory()
    tracemalloc.stop()

    return {
        "function": func_name,
        "cells": n_cells,
        "time_seconds": round(t1 - t0, 4),
        "peak_memory_mb": round(peak / (1024 * 1024), 2),
    }


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

CELL_COUNTS = [1_000, 5_000, 10_000]

BENCHMARKS = [
    {
        "name": "apply_cell_style",
        "new_func": apply_cell_style,
        "old_func": _inline_apply_cell_style,
        "kwargs": {"bold": True, "font_color": "FF0000", "bg_color": "FFFF00"},
    },
    {
        "name": "apply_alignment",
        "new_func": apply_alignment,
        "old_func": _inline_apply_alignment,
        "kwargs": {
            "horizontal": "center",
            "vertical": "top",
            "wrap_text": True,
        },
    },
]


def main() -> None:
    pre_results: list[dict] = []
    post_results: list[dict] = []

    with tempfile.TemporaryDirectory() as tmpdir:
        os.environ["EXCEL_ALLOWLIST_ROOTS"] = tmpdir

        # ------------------------------------------------------------------
        # Pre-optimization (inline) pass
        # ------------------------------------------------------------------
        print("\n--- PRE-OPTIMIZATION (inline allocation) ---")
        for bm in BENCHMARKS:
            for n in CELL_COUNTS:
                wbpath = os.path.join(
                    tmpdir, f"pre_{bm['name']}_{n}.xlsx"
                )
                _build_workbook(wbpath, n)
                address = f"A1:A{n}"

                record = _run_inline(
                    bm["old_func"],
                    path=wbpath,
                    sheet="Sheet1",
                    address=address,
                    func_name=bm["name"],
                    n_cells=n,
                    **bm["kwargs"],
                )
                pre_results.append(record)
                print(
                    f"  {bm['name']:20s}  cells={n:>6d}"
                    f"  time={record['time_seconds']:.4f}s"
                    f"  peak={record['peak_memory_mb']:.1f} MB"
                )

        # ------------------------------------------------------------------
        # Post-optimization (hoisted) pass
        # ------------------------------------------------------------------
        print("\n--- POST-OPTIMIZATION (hoisted allocation) ---")
        for bm in BENCHMARKS:
            for n in CELL_COUNTS:
                wbpath = os.path.join(
                    tmpdir, f"post_{bm['name']}_{n}.xlsx"
                )
                _build_workbook(wbpath, n)
                address = f"A1:A{n}"

                record = _run_one(
                    bm["new_func"],
                    path=wbpath,
                    sheet="Sheet1",
                    address=address,
                    func_name=bm["name"],
                    n_cells=n,
                    **bm["kwargs"],
                )
                post_results.append(record)
                print(
                    f"  {bm['name']:20s}  cells={n:>6d}"
                    f"  time={record['time_seconds']:.4f}s"
                    f"  peak={record['peak_memory_mb']:.1f} MB"
                )

    # ------------------------------------------------------------------
    # Write JSON output files
    # ------------------------------------------------------------------
    benchmark_dir = Path(__file__).parent

    pre_output = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "label": "pre_optimization_simulated",
        "results": pre_results,
    }
    pre_path = benchmark_dir / "pre_optimization.json"
    pre_path.write_text(json.dumps(pre_output, indent=2))

    post_output = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "label": "post_optimization",
        "results": post_results,
    }
    post_path = benchmark_dir / "post_optimization.json"
    post_path.write_text(json.dumps(post_output, indent=2))

    # Replace baseline.json with post_optimization to correct its label.
    baseline_path = benchmark_dir / "baseline.json"
    baseline_path.write_text(json.dumps(post_output, indent=2))

    # ------------------------------------------------------------------
    # Delta summary table → stderr
    # ------------------------------------------------------------------
    print("\n--- DELTA SUMMARY (pre vs post optimization) ---", file=sys.stderr)
    header = (
        f"{'function':<22} {'cells':>7}  {'pre_time_s':>12}  "
        f"{'post_time_s':>12}  {'speedup_ratio':>14}"
    )
    sep = "-" * len(header)
    print(header, file=sys.stderr)
    print(sep, file=sys.stderr)

    # Build lookup: (function, cells) → record
    pre_map = {(r["function"], r["cells"]): r for r in pre_results}
    post_map = {(r["function"], r["cells"]): r for r in post_results}

    for bm in BENCHMARKS:
        for n in CELL_COUNTS:
            key = (bm["name"], n)
            pre_t = pre_map[key]["time_seconds"]
            post_t = post_map[key]["time_seconds"]
            ratio = round(pre_t / post_t, 3) if post_t > 0 else float("inf")
            print(
                f"{bm['name']:<22} {n:>7d}  {pre_t:>12.4f}  "
                f"{post_t:>12.4f}  {ratio:>14.3f}",
                file=sys.stderr,
            )

    print(sep, file=sys.stderr)
    print(f"\npre_optimization.json  → {pre_path}", file=sys.stderr)
    print(f"post_optimization.json → {post_path}", file=sys.stderr)
    print(
        f"baseline.json          → {baseline_path} (updated to post_optimization)",
        file=sys.stderr,
    )


def _run_profiling(output_path: str) -> None:
    """Profile apply_cell_style over 5k cells and write a report."""
    import cProfile
    import pstats
    import io as _io_module
    import os as _os
    import tempfile as _tempfile
    from openpyxl import Workbook

    _os.environ["EXCEL_ENABLE_WRITE"] = "true"
    _os.environ["EXCEL_ALLOWLIST_ROOTS"] = _tempfile.gettempdir()

    wb = Workbook()
    ws = wb.active
    tf = _tempfile.mktemp(suffix=".xlsx", dir=_tempfile.gettempdir())
    wb.save(tf)

    pr = cProfile.Profile()
    pr.enable()
    from excelmcp._format import apply_cell_style  # noqa: PLC0415
    apply_cell_style(
        path=tf,
        sheet=ws.title,
        address="A1:E1000",   # 5k cells (5 × 1000)
        bold=True,
        italic=False,
        font_size=10,
        font_color="000000",
        bg_color="FFFFFF",
        confirm=True,
    )
    pr.disable()

    s = _io_module.StringIO()
    ps = pstats.Stats(pr, stream=s).sort_stats(pstats.SortKey.CUMULATIVE)
    ps.print_stats(20)

    # Governance overhead measurement
    import timeit
    from excelmcp._core import _check_write, _check_confirm, _check_path  # noqa: PLC0415

    n_gov = 10000
    t_write = timeit.timeit(_check_write, number=n_gov) / n_gov * 1e6
    t_confirm = timeit.timeit(lambda: _check_confirm(True), number=n_gov) / n_gov * 1e6
    t_path = timeit.timeit(
        lambda: _check_path(tf),
        number=100,
    ) / 100 * 1e6

    gov_section = (
        "\n--- Governance Gate Overhead (microseconds/call, amortized) ---\n"
        f"  _check_write():   {t_write:.2f} µs\n"
        f"  _check_confirm(): {t_confirm:.2f} µs\n"
        f"  _check_path():    {t_path:.2f} µs\n"
    )

    with open(output_path, "w", encoding="utf-8") as f:
        f.write("=== excelmcp PERF-OPT cProfile Report ===\n")
        f.write("Profiled: apply_cell_style over A1:E1000 (5000 cells)\n\n")
        f.write(s.getvalue())
        f.write(gov_section)

    try:
        _os.unlink(tf)
    except OSError:
        pass

    print(f"Profile report written to: {output_path}")  # noqa: T201 — CLI script, not server code


if __name__ == "__main__":
    main()
    _run_profiling(os.path.join(os.path.dirname(__file__), "profile_report.txt"))
