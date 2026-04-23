"""COM integration tests for dynamic-array workbook artifacts.

These tests use Excel COM as a ground-truth oracle to verify that:
- MCP-generated DA workbooks can be opened by Excel without repair dialogs
- DA formulas recalculate correctly
- The probe_excel_reopen helper works as expected

Run with:  pytest excel/tests/test_integration_da_com.py -v -m integration
CI: Windows runner only. Skip in non-Windows / no-COM environments.
"""
import os

import pytest
import openpyxl
from pathlib import Path

# Skip entire module if win32com is not available
win32com = pytest.importorskip(
    "win32com.client", reason="win32com not installed — COM tests skipped"
)

from excelmcp._ooxml import patch_dynamic_array_metadata  # noqa: E402
from excelmcp._da_com import probe_excel_reopen  # noqa: E402
from excelmcp._da_finalize import finalize_workbook_da_artifact  # noqa: E402


class TestDAComIntegration:
    """COM-based ground-truth oracle tests for DA workbook artifacts."""

    @pytest.fixture(autouse=True)
    def com_sta_init(self) -> None:  # type: ignore[return]
        """Initialise COM STA apartment for this test.

        Excel is an STA COM server.  pytest's main thread may not be in STA
        by default; CoInitialize ensures the calling thread is in STA mode
        before any COM dispatch.

        CoUninitialize is intentionally NOT called here: in pytest, calling it
        during test teardown triggers 0x80010108 noise because pytest's logging
        infrastructure still holds COM-connected state when teardown fires.
        The STA apartment is released cleanly when the process exits.
        """
        import pythoncom
        pythoncom.CoInitialize()
        os.environ["EXCEL_ENABLE_COM"] = "true"
        yield
        os.environ.pop("EXCEL_ENABLE_COM", None)

    @pytest.mark.integration
    def test_probe_returns_ok_for_plain_workbook(self, tmp_path: Path) -> None:
        """probe_excel_reopen returns ok=True for a plain workbook with no formulas.

        Establishes baseline: the COM probe cycle works end-to-end for simple files,
        ruling out environment or COM setup issues when DA-specific tests fail.
        """
        path = tmp_path / "plain.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "hello"
        ws["A2"] = 42
        ws["A3"] = 3.14
        wb.save(str(path))

        result = probe_excel_reopen(str(path))

        assert result["ok"] is True, f"Plain workbook probe failed: {result}"

    @pytest.mark.integration
    def test_probe_returns_ok_after_da_patch(self, tmp_path: Path) -> None:
        """probe_excel_reopen returns ok=True for a DA-patched workbook.

        KEY test: proves that patch_dynamic_array_metadata produces an artifact
        that Excel can open without triggering the repair dialog.
        FILTER formula is chosen as the canonical DA trigger because FILTER
        requires both _xlfn._xlws. prefix AND cm='1' metadata to be preserved.
        """
        path = tmp_path / "da_patched.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        # Fill input data for FILTER
        for i in range(1, 6):
            ws.cell(row=i, column=1, value=i)
        # openpyxl stores formulas as strings; patch step adds cm + _xlfn prefix
        ws["B1"].value = "=FILTER(A1:A5,A1:A5>2)"
        wb.save(str(path))

        patch_dynamic_array_metadata(str(path))
        result = probe_excel_reopen(str(path))

        assert result["ok"] is True, f"DA-patched workbook probe failed: {result}"

    @pytest.mark.integration
    def test_probe_sequence_matches_finalize_pipeline(
        self, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        """finalize_workbook_da_artifact with require_com_probe=True produces ok=True + probe stage ok.

        Validates the full 4-stage pipeline: patch → classify → validate → probe.
        Also checks auto-mode guard: when EXCEL_DA_COM_PROBE=auto, probe is currently
        skipped (should_probe=False in this implementation) but finalize still succeeds.
        """
        path = tmp_path / "sequence.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"].value = "=SEQUENCE(5)"
        wb.save(str(path))

        # First pass: auto policy (probe skipped per current implementation)
        monkeypatch.setenv("EXCEL_DA_COM_PROBE", "auto")
        diag_auto = finalize_workbook_da_artifact(str(path))
        assert diag_auto["ok"] is True, f"auto-policy finalize failed: {diag_auto}"

        # Second pass: force probe via require_com_probe=True
        diag_probe = finalize_workbook_da_artifact(str(path), require_com_probe=True)
        assert diag_probe["ok"] is True, f"require_com_probe finalize failed: {diag_probe}"

        probe_stage = next(
            (s for s in diag_probe["stages"] if s["name"] == "probe"), None
        )
        assert probe_stage is not None, "probe stage missing from diagnostics"
        assert probe_stage["status"] == "ok", f"probe stage not ok: {probe_stage}"

    @pytest.mark.integration
    def test_probe_returns_false_for_corrupt_zip(self, tmp_path: Path) -> None:
        """Negative test: probe_excel_reopen returns ok=False (never raises) for a corrupt file.

        Validates fault-tolerance: the probe helper must absorb COM errors and
        return a structured failure dict rather than propagating exceptions to callers.
        """
        corrupt_path = tmp_path / "corrupt.xlsx"
        # Write bytes that look like a ZIP signature but are otherwise invalid
        corrupt_path.write_bytes(b"PK\x03\x04" + b"\x00" * 18 + b"not a valid xlsx")

        result = probe_excel_reopen(str(corrupt_path))

        assert result["ok"] is False, "Expected ok=False for corrupt file"
        assert result.get("error") is not None, "Expected error message in result"

    @pytest.mark.integration
    def test_com_control_workbook_reopens_ok(self, tmp_path: Path) -> None:
        """Ground-truth BDD: a COM-native DA workbook must pass probe_excel_reopen.

        Creates a workbook using Excel COM directly (the gold standard),
        then verifies it passes the probe cycle. Any MCP-generated DA artifact
        should eventually match this standard — this test is the reference baseline.
        """
        com_path = tmp_path / "com_control.xlsx"

        # Create the control workbook via COM.
        # win32com is the win32com.client module (from pytest.importorskip),
        # so win32com.Dispatch(...) == win32com.client.Dispatch(...).
        excel = win32com.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        try:
            wb = excel.Workbooks.Add()
            ws = wb.Worksheets(1)
            for i in range(1, 6):
                ws.Cells(i, 1).Value = i
            # Formula2 is the dynamic-array-aware property (Excel 365)
            ws.Cells(1, 2).Formula2 = "=FILTER(A1:A5,A1:A5>2)"
            wb.SaveAs(str(com_path), 51)  # 51 = xlOpenXMLWorkbook
            wb.Close(SaveChanges=False)
        finally:
            excel.Quit()
            excel = None

        result = probe_excel_reopen(str(com_path))

        assert result["ok"] is True, f"COM-native control workbook probe failed: {result}"

    @pytest.mark.integration
    @pytest.mark.xfail(
        reason="Expected to fail until COM finalize path is fully wired; "
        "passes when openpyxl-patched OOXML matches COM-native output quality",
        strict=False,
    )
    def test_probe_com_created_vs_openpyxl_patched(self, tmp_path: Path) -> None:
        """BDD reconciliation: COM-created and openpyxl-patched DA workbooks both pass probe.

        Given:  a COM-created DA workbook (ground truth)
               AND an openpyxl-generated+patched DA workbook (MCP output path)
        When:   probe_excel_reopen is called on each
        Then:   both should return ok=True

        The openpyxl path may produce OOXML differences vs COM native output that
        trigger Excel repair. This xfail test tracks the DA fix: when both assertions
        pass (xpass with strict=False), the MCP output matches COM quality.
        """
        # COM-created workbook (ground truth).
        # win32com here is the win32com.client module from pytest.importorskip.
        com_path = tmp_path / "com_bdd.xlsx"
        excel = win32com.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        try:
            wb = excel.Workbooks.Add()
            ws = wb.Worksheets(1)
            for i in range(1, 6):
                ws.Cells(i, 1).Value = i
            ws.Cells(1, 2).Formula2 = "=FILTER(A1:A5,A1:A5>2)"
            wb.SaveAs(str(com_path), 51)
            wb.Close(SaveChanges=False)
        finally:
            excel.Quit()
            excel = None

        # openpyxl-generated + DA-patched workbook (MCP output path)
        openpyxl_path = tmp_path / "openpyxl_bdd.xlsx"
        wb2 = openpyxl.Workbook()
        ws2 = wb2.active
        for i in range(1, 6):
            ws2.cell(row=i, column=1, value=i)
        ws2["B1"].value = "=FILTER(A1:A5,A1:A5>2)"
        wb2.save(str(openpyxl_path))
        patch_dynamic_array_metadata(str(openpyxl_path))

        com_result = probe_excel_reopen(str(com_path))
        openpyxl_result = probe_excel_reopen(str(openpyxl_path))

        assert com_result["ok"] is True, f"COM workbook failed probe: {com_result}"
        assert openpyxl_result["ok"] is True, (
            f"openpyxl-patched workbook failed probe: {openpyxl_result}"
        )

    @pytest.mark.integration
    def test_let_filter_formula2_survives_mcp_save(self, tmp_path: Path) -> None:
        """COM Formula2 LET+FILTER survives an MCP (openpyxl) save round-trip.

        Regression test for Known Open Gap: LET() wrapping a spill function
        written via COM .Formula2 must keep its ``t="array" aca="1" ref ca``
        OOXML attributes after any Excel MCP save operation.

        Scenario:
        1. Write LET+FILTER via COM (.Formula2) to cell A1 → saves file
        2. Call apply_number_format via MCP (triggers openpyxl save)
        3. Verify OOXML still contains ``aca="1"`` on A1's <f> element
        4. probe_excel_reopen confirms Excel opens without repair
        """
        import os
        import zipfile
        from excelmcp.workbook_openpyxl import apply_number_format

        path = tmp_path / "let_filter_mcp_roundtrip.xlsx"

        # Step 1: write LET+FILTER via COM
        excel = win32com.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        try:
            wb = excel.Workbooks.Add()
            ws = wb.Worksheets(1)
            # Input data in column A
            for i in range(1, 6):
                ws.Cells(i, 1).Value = i
            # LET+FILTER: keep values > 2
            ws.Cells(1, 2).Formula2 = (
                "=LET(src,A1:A5,filt,FILTER(src,src>2,\"\"),filt)"
            )
            wb.SaveAs(str(path), 51)  # 51 = xlOpenXMLWorkbook
            wb.Close(SaveChanges=False)
        finally:
            excel.Quit()
            excel = None

        # Verify aca="1" is present in the COM-saved file
        with zipfile.ZipFile(str(path), "r") as z:
            sheet_xml = z.read("xl/worksheets/sheet1.xml").decode("utf-8")
        assert 'aca="1"' in sheet_xml, "COM save should produce aca=\"1\" Formula2 marker"

        # Step 2: MCP operation — apply_number_format triggers openpyxl save
        os.environ["EXCEL_ALLOWLIST_ROOTS"] = str(tmp_path)
        os.environ["EXCEL_ENABLE_WRITE"] = "true"
        apply_number_format(str(path), "Sheet1", "C1:C10", "0.00", confirm=True)

        # Step 3: verify aca="1" is still present after MCP save
        with zipfile.ZipFile(str(path), "r") as z:
            sheet_xml_after = z.read("xl/worksheets/sheet1.xml").decode("utf-8")
        assert 'aca="1"' in sheet_xml_after, (
            "aca=\"1\" Formula2 marker must survive MCP openpyxl save — "
            "snapshot/restore broken?"
        )

        # Step 4: Excel can open the file without repair
        result = probe_excel_reopen(str(path))
        assert result["ok"] is True, (
            f"LET+FILTER Formula2 failed probe after MCP save: {result}"
        )
