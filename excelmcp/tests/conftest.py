import pytest
import openpyxl
from unittest.mock import MagicMock

from excelmcp._core import _wb_cache

# ---------------------------------------------------------------------------
# FastMCP 2.x compat: make FunctionTool callable so tests that import
# @mcp.tool()-decorated symbols and call them directly continue to work.
# In FastMCP 2.x @mcp.tool() wraps the function in a FunctionTool which is
# NOT callable by default; in 3.x the original function is returned.
# ---------------------------------------------------------------------------
try:
    from fastmcp.tools.tool import FunctionTool
    if not hasattr(FunctionTool, "_call_compat_patched"):
        FunctionTool.__call__ = lambda self, *args, **kwargs: self.fn(*args, **kwargs)
        FunctionTool._call_compat_patched = True
except ImportError:
    pass


def pytest_configure(config):
    config.addinivalue_line("markers", "unit: unit tests")
    config.addinivalue_line("markers", "smoke: smoke tests")
    config.addinivalue_line(
        "markers", "integration: integration tests requiring live Office app"
    )


@pytest.fixture(autouse=True)
def _clear_wb_cache():
    _wb_cache.clear()
    yield
    _wb_cache.clear()


@pytest.fixture
def sample_workbook(tmp_path, monkeypatch):
    """Create a minimal test.xlsx, set EXCEL_ALLOWLIST_ROOTS, and return the path as str."""
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    path = tmp_path / "test.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "hello"
    ws["A2"] = 42
    wb.save(str(path))
    return str(path)


@pytest.fixture
def valid_com_env(monkeypatch, tmp_path):
    """Set all required env vars for COM tools and return tmp_path."""
    monkeypatch.setenv("EXCEL_ENABLE_COM", "true")
    monkeypatch.setenv("EXCEL_ENABLE_WRITE", "true")
    monkeypatch.setenv("EXCEL_ALLOWLIST_ROOTS", str(tmp_path))
    monkeypatch.setenv("EXCEL_EXPORT_ROOTS", str(tmp_path))
    return tmp_path


@pytest.fixture
def mock_com_dispatch(monkeypatch):
    """Patch excelmcp._com._ensure_com_available so no win32com import occurs.

    Builds a MagicMock hierarchy:
        mock_pythoncom  — CoInitialize/CoUninitialize calls
        mock_excel      — Excel.Application proxy (returned by win32c.Dispatch)
        mock_wb         — Workbook proxy (returned by mock_excel.Workbooks.Open)
            .Sheets.Count           = 1
            .Sheets.side_effect     = lambda n: sheet1  (sheet1.Name = "Sheet1")
            .Application            = mock_excel
            .Names.Count            = 0  (empty named range list)

    DC-4: wb.Sheets is called as a callable (wb.Sheets(i+1)) not an attribute
    access. Use side_effect, NOT return_value, to handle different integer args.
    """
    mock_pythoncom = MagicMock()
    mock_excel = MagicMock()
    mock_wb = MagicMock()
    sheet1 = MagicMock()
    sheet1.Name = "Sheet1"
    mock_wb.Sheets.Count = 1
    mock_wb.Sheets.side_effect = lambda n: sheet1
    mock_wb.Application = mock_excel
    mock_wb.Names.Count = 0  # prevent TypeError in range(wb.Names.Count)
    mock_win32c = MagicMock()
    mock_win32c.Dispatch.return_value = mock_excel
    mock_excel.Workbooks.Open.return_value = mock_wb

    monkeypatch.setattr(
        "excelmcp._com._ensure_com_available",
        lambda: (mock_pythoncom, mock_win32c),
    )
    return mock_pythoncom, mock_excel, mock_wb


# ---------------------------------------------------------------------------
# Sprint H pivot fixtures
# ---------------------------------------------------------------------------

@pytest.fixture
def mock_pivot_table():
    """Single PivotTable COM proxy — SourceType=1 (xlDatabase), DataBodyRange = 2x2."""
    mock_pt = MagicMock()
    mock_pt.Name = "PT1"
    mock_cache = MagicMock()
    mock_cache.SourceType = 1
    mock_pt.PivotCache.return_value = mock_cache
    mock_pt.DataBodyRange = MagicMock()
    mock_pt.DataBodyRange.Count = 4
    mock_pt.DataBodyRange.Value = ((10, 20), (30, 40))
    field1 = MagicMock()
    field1.Name = "Region"
    field1.Orientation = 1
    field2 = MagicMock()
    field2.Name = "Sales"
    field2.Orientation = 4
    pf_proxy = MagicMock()
    pf_proxy.Count = 2
    def _pf_side_effect(*args):
        if not args:
            return pf_proxy
        return field1 if args[0] == 1 else field2
    mock_pt.PivotFields.side_effect = _pf_side_effect
    return mock_pt


@pytest.fixture
def mock_sheet_with_pivots(mock_pivot_table):
    """Mock worksheet containing one pivot table (mock_pivot_table)."""
    mock_ws = MagicMock()
    mock_ws.Name = "Sheet1"
    pt_proxy = MagicMock()
    pt_proxy.Count = 1
    def _pt_side_effect(*args):
        if not args:
            return pt_proxy
        return mock_pivot_table
    mock_ws.PivotTables.side_effect = _pt_side_effect
    return mock_ws


@pytest.fixture
def mock_wb_two_pivots():
    """Standalone mock workbook (no mock_com_dispatch dependency) with 2 xlDatabase pivots."""
    mock_wb = MagicMock()
    mock_wb.Names.Count = 0
    pt1 = MagicMock()
    pt1.Name = "PT1"
    pt1.PivotCache.return_value.SourceType = 1
    pt2 = MagicMock()
    pt2.Name = "PT2"
    pt2.PivotCache.return_value.SourceType = 1
    pt_proxy = MagicMock()
    pt_proxy.Count = 2
    def _pt_se(*args):
        return pt_proxy if not args else (pt1 if args[0] == 1 else pt2)
    mock_ws = MagicMock()
    mock_ws.Name = "Sheet1"
    mock_ws.PivotTables.side_effect = _pt_se
    mock_wb.Sheets.Count = 1
    mock_wb.Sheets.side_effect = lambda n: mock_ws
    mock_wb.Application = MagicMock()
    return mock_wb


@pytest.fixture
def mock_active_excel_session(monkeypatch):
    """Patch _active_wb._ensure_com_available so GetActiveObject returns a mock xl_app.

    Returns (mock_excel, mock_wb) tuple.

    Usage::

        def test_something(mock_active_excel_session):
            mock_excel, mock_wb = mock_active_excel_session
            # mock_excel.ActiveWorkbook == mock_wb
            # mock_excel.Workbooks == [mock_wb]
    """
    mock_pythoncom = MagicMock()
    mock_excel = MagicMock()
    mock_wb = MagicMock()
    mock_wb.FullName = r"C:\fake\test.xlsx"
    mock_wb.Name = "test.xlsx"
    mock_excel.ActiveWorkbook = mock_wb
    mock_excel.Workbooks = [mock_wb]
    mock_win32c = MagicMock()
    mock_win32c.GetActiveObject.return_value = mock_excel
    monkeypatch.setattr(
        "excelmcp._active_wb._ensure_com_available",
        lambda: (mock_pythoncom, mock_win32c),
    )
    return mock_excel, mock_wb
